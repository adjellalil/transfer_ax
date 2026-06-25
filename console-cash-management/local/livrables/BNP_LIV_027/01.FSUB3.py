"""
MONEXT COMPARAISON ANNUELLE [FSUB3]
===================================

DESCRIPTION :
-------------
Analyseur BNP comparant le PNB par client (ID RP) entre les periodes
janvier-fevrier 2025 et janvier-fevrier 2026, a partir du CSV detail
produit par MONEXT ANALYZER [K2P8N]. Calcule pour chaque client les 6
types de PNB sur N et N+1 (Corporate, Cotisations, Commissions, NDF,
Interets, Total), les ecarts (montant et pourcentage), la tendance et le
type de PNB a plus forte variation. Traitement 100% vectorise (zero
iterrows dans le calcul). Genere un fichier XLSX a 3 tableaux cote a
cote (Cumul Jan+Fev, Janvier, Fevrier). Version CLI autonome, sans GUI.

SOURCES REQUISES :
------------------
- CSV detail MONEXT ANALYZER [K2P8N] (--source-monext-detail)
    Colonnes positionnelles attendues (defauts) :
      MOIS (format 2025_JANVIER) | ID RP | Corporate Acronym |
      ENTREPRISE | BPE | SALES | PNB Corporate/Cotis/Commis/NDF/Interets/Total

OUTPUTS PRODUITS :
------------------
- XLSX une seule sheet "Analyse Comparative", 3 tableaux cote a cote
  separes par 2 colonnes vides :
    [Cumul Jan+Fev] | [Janvier] | [Fevrier]
  Colonnes par tableau :
    ID_RP | CORPORATE_ACRONYM | ENTREPRISE | BPE | SALES_RESEAU
    | PNB_*_N | PNB_*_N+1 (x6 types) | ECART_MONTANT | ECART_POURCENTAGE
    | TENDANCE | TYPE_PNB_MAX_VARIATION
  Tri : ECART_MONTANT croissant (baisses les plus fortes en premier).

ARGUMENTS CLI :
---------------
  --source-monext-detail PATH   (obligatoire) CSV detail MONEXT [K2P8N]
  --output-dir PATH             (obligatoire) Dossier de sortie XLSX
  --output-filename NAME        (obligatoire) Nom du fichier XLSX

DECOMPOSITION :
---------------
FSUB3
|-- load_csv_smart           lecture CSV multi-separateur/encodage
|-- to_float_vec             conversion numerique vectorisee
|-- MonextComparaison_FSUB3
|   |-- __init__             affecte les attributs lus depuis argparse
|   |-- _prog               progression -> print
|   |-- run                 worker 100% vectorise
|   |   |-- chargement + filtre MOIS
|   |   |-- nettoyage ID_RP + conversion PNB
|   |   |-- segmentation client (groupby/mode)
|   |   |-- pivot vectorise PNB par client/mois
|   |   |-- build_tableau   calcul N/N+1, ecarts, tendances, tri
|   |   |-- _export_xlsx    ecriture des 3 tableaux cote a cote
|   |-- _export_xlsx        export Excel mis en forme
|-- main                     argparse + try/except, sorties 0/1/2

BNP Paribas Cash Management - Direction Monetique
Mars 2026
"""

from __future__ import annotations

import argparse
import sys
import traceback
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

VERSION_ID = "FSUB3"

MOIS_J25 = "2025_JANVIER"
MOIS_F25 = "2025_FEVRIER"
MOIS_J26 = "2026_JANVIER"
MOIS_F26 = "2026_FEVRIER"
MOIS_UTILES = [MOIS_J25, MOIS_F25, MOIS_J26, MOIS_F26]

DEFAULT_POSITIONS = {
    'col_mois':          1,
    'col_id_rp':         9,
    'col_corporate':     2,
    'col_entreprise':   63,
    'col_bpe':          64,
    'col_sales':        71,
    'col_pnb_corporate':72,
    'col_pnb_cotis':    73,
    'col_pnb_commis':   74,
    'col_pnb_ndf':      75,
    'col_pnb_interets': 76,
    'col_pnb_total':    77,
}

PNB_KEYS = [
    ("Corporate",   "col_pnb_corporate"),
    ("Cotisations", "col_pnb_cotis"),
    ("Commissions", "col_pnb_commis"),
    ("NDF",         "col_pnb_ndf"),
    ("Interets",    "col_pnb_interets"),
    ("Total",       "col_pnb_total"),
]

SEG_COLS_LABEL = [
    ("col_mois",       "Colonne MOIS (format 2025_JANVIER)"),
    ("col_id_rp",      "ID RP (cle client)"),
    ("col_corporate",  "Corporate Acronym (col 2)"),
    ("col_entreprise", "ENTREPRISE (YES/NO)"),
    ("col_bpe",        "BPE (YES/NO)"),
    ("col_sales",      "SALES_YANNICK"),
]


def load_csv_smart(path, nrows=None):
    for sep in [';', ',', '\t']:
        for enc in ['utf-8-sig', 'utf-8', 'latin1', 'cp1252']:
            try:
                df = pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                 keep_default_na=False, na_values=[],
                                 on_bad_lines='skip', nrows=5)
                if df.shape[1] > 1:
                    return pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                       keep_default_na=False, na_values=[],
                                       on_bad_lines='skip', nrows=nrows)
            except Exception:
                continue
    return pd.read_csv(path, sep=None, engine='python', dtype=str,
                       keep_default_na=False, na_values=[], on_bad_lines='skip',
                       nrows=nrows)


def to_float_vec(series):
    s = series.astype(str).str.strip()
    s = s.str.replace(' ', '', regex=False).str.replace('\xa0', '', regex=False)
    s = s.str.replace(',', '.', regex=False)
    return pd.to_numeric(s, errors='coerce').fillna(0.0)


class MonextComparaison_FSUB3:
    def __init__(self, args: argparse.Namespace) -> None:
        self.fichier_source: str = str(args.source_monext_detail)
        self.output_dir: Path = Path(args.output_dir)
        self.output_filename: str = str(args.output_filename)

        # Mapping des colonnes par defaut (lecture positionnelle du CSV detail).
        # Les attributs ci-dessous remplacent les ComboBox de l'ancienne GUI :
        # le worker lit le nom de colonne via self.gcn(key).
        self.df_preview: Optional[pd.DataFrame] = None
        self.original_cols: list = []
        self.col_map: dict = {}

    def _prog(self, val: float, txt: str) -> None:
        print(f"[{int(val * 100):3d}%] {txt}")

    def gcn(self, key: str) -> str:
        return self.col_map[key]

    # =========================================================================
    # WORKER — 100% VECTORISE
    # =========================================================================
    def run(self) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # --- Resolution du mapping colonnes (positionnel par defaut) ---
        self._prog(0.02, "Lecture de l'entete du fichier source...")
        self.df_preview = load_csv_smart(self.fichier_source, nrows=5)
        self.original_cols = list(self.df_preview.columns)
        for key, _lbl in SEG_COLS_LABEL:
            self.col_map[key] = self.original_cols[DEFAULT_POSITIONS[key] - 1]
        for _label, key in PNB_KEYS:
            self.col_map[key] = self.original_cols[DEFAULT_POSITIONS[key] - 1]

        col_mois       = self.gcn("col_mois")
        col_id_rp      = self.gcn("col_id_rp")
        col_corporate  = self.gcn("col_corporate")
        col_entreprise = self.gcn("col_entreprise")
        col_bpe        = self.gcn("col_bpe")
        col_sales      = self.gcn("col_sales")
        pnb_cols       = {label: self.gcn(key) for label, key in PNB_KEYS}

        # --- Chargement ---
        self._prog(0.05, "Chargement du fichier source...")
        df = load_csv_smart(self.fichier_source)
        self._prog(0.10, f"Charge — {len(df):,} lignes")

        # --- Nettoyage MOIS + filtre ---
        df[col_mois] = df[col_mois].astype(str).str.strip().str.upper()
        df_f = df[df[col_mois].isin(MOIS_UTILES)].copy()

        if df_f.empty:
            vals_uniq = df[col_mois].unique()[:15]
            raise ValueError(
                "Aucune ligne pour les mois attendus :\n"
                f"{', '.join(MOIS_UTILES)}\n\n"
                "Valeurs trouvees :\n"
                f"{', '.join(str(v) for v in vals_uniq)}"
            )

        # --- Nettoyage ID_RP ---
        df_f[col_id_rp] = df_f[col_id_rp].astype(str).str.strip()
        df_f[col_id_rp] = df_f[col_id_rp].replace(
            ['', 'nan', 'NAN', 'NONE', 'None', 'N/A', 'NA'], 'INCONNU'
        )

        # --- Conversion PNB vectorisee ---
        self._prog(0.18, "Conversion PNB en numerique...")
        for col in pnb_cols.values():
            if col in df_f.columns:
                df_f[col] = to_float_vec(df_f[col])

        # --- SALES_RESEAU vectorise ---
        df_f['_SALES_RESEAU'] = np.where(
            df_f[col_sales].astype(str).str.strip().str.upper().isin(
                ['N/A', 'NA', '', 'NAN', 'NONE']
            ),
            'RESEAU', 'SALES'
        )

        # --- Segmentation par client ---
        self._prog(0.25, "Calcul segmentation clients...")

        def mode_vec(s):
            m = s.mode()
            return m.iloc[0] if not m.empty else ''

        df_seg = df_f.groupby(col_id_rp).agg(
            CORPORATE_ACRONYM=(col_corporate,  mode_vec),
            ENTREPRISE=(col_entreprise,        mode_vec),
            BPE=(col_bpe,                      mode_vec),
            SALES_RESEAU=('_SALES_RESEAU',     mode_vec),
        ).reset_index().rename(columns={col_id_rp: 'ID_RP'})

        # --- Pivot vectorise ---
        self._prog(0.35, "Agregation PNB par client et mois...")
        agg_dict = {col: 'sum' for col in pnb_cols.values() if col in df_f.columns}
        df_agg = df_f.groupby([col_id_rp, col_mois], as_index=False).agg(agg_dict)
        df_agg.rename(columns={col_id_rp: 'ID_RP'}, inplace=True)

        df_pivot = df_agg.pivot_table(
            index='ID_RP',
            columns=col_mois,
            values=list(pnb_cols.values()),
            aggfunc='sum',
            fill_value=0.0
        )
        df_pivot.columns = [f"{col}__{mois}" for col, mois in df_pivot.columns]
        df_pivot = df_pivot.reset_index().merge(df_seg, on='ID_RP', how='left')

        # =====================================================================
        # BUILD TABLEAU — fonction vectorisee
        # =====================================================================
        def build_tableau(mois_n_list, mois_n1_list, label_n, label_n1):
            """
            Construit un tableau comparatif — 100% vectorise.
            mois_n_list / mois_n1_list : liste de valeurs MOIS a sommer.
            """
            rows = pd.DataFrame()
            rows['ID_RP']             = df_pivot['ID_RP']
            rows['CORPORATE_ACRONYM'] = df_pivot.get('CORPORATE_ACRONYM',
                                        pd.Series('', index=df_pivot.index))
            rows['ENTREPRISE']        = df_pivot.get('ENTREPRISE',
                                        pd.Series('', index=df_pivot.index))
            rows['BPE']               = df_pivot.get('BPE',
                                        pd.Series('', index=df_pivot.index))
            rows['SALES_RESEAU']      = df_pivot.get('SALES_RESEAU',
                                        pd.Series('', index=df_pivot.index))

            pnb_totals_n  = pd.Series(0.0, index=df_pivot.index)
            pnb_totals_n1 = pd.Series(0.0, index=df_pivot.index)

            for lbl, col in pnb_cols.items():
                # Somme vectorisee sur la liste de mois
                vals_n = sum(
                    df_pivot.get(f"{col}__{m}", pd.Series(0.0, index=df_pivot.index))
                    for m in mois_n_list
                )
                vals_n1 = sum(
                    df_pivot.get(f"{col}__{m}", pd.Series(0.0, index=df_pivot.index))
                    for m in mois_n1_list
                )
                rows[f"PNB_{lbl}_{label_n}"]  = vals_n.values
                rows[f"PNB_{lbl}_{label_n1}"] = vals_n1.values

                if lbl == 'Total':
                    pnb_totals_n  = vals_n
                    pnb_totals_n1 = vals_n1

            ecart = (pnb_totals_n1 - pnb_totals_n).values
            base  = pnb_totals_n.abs().values

            rows['ECART_MONTANT']     = np.round(ecart, 2)
            rows['ECART_POURCENTAGE'] = np.where(
                base > 0.01,
                np.round(ecart / base * 100, 1),
                0.0
            )
            rows['TENDANCE'] = np.where(ecart > 0, 'HAUSSE',
                               np.where(ecart < 0, 'BAISSE', 'STABLE'))

            # TYPE_PNB_MAX_VARIATION — vectorise
            pnb_labels_sub = [lbl for lbl, _ in PNB_KEYS if lbl != 'Total']
            ecart_matrix = np.column_stack([
                (rows[f"PNB_{lbl}_{label_n1}"].values -
                 rows[f"PNB_{lbl}_{label_n}"].values)
                for lbl in pnb_labels_sub
            ])
            idx_max = np.argmax(np.abs(ecart_matrix), axis=1)
            rows['TYPE_PNB_MAX_VARIATION'] = np.array(pnb_labels_sub)[idx_max]

            return rows.sort_values('ECART_MONTANT', ascending=True).reset_index(drop=True)

        self._prog(0.55, "Construction tableau Cumul Jan+Fev...")
        tab_cum = build_tableau([MOIS_J25, MOIS_F25], [MOIS_J26, MOIS_F26],
                                "JanFev2025", "JanFev2026")

        self._prog(0.68, "Construction tableau Janvier...")
        tab_jan = build_tableau([MOIS_J25], [MOIS_J26],
                                "Jan2025", "Jan2026")

        self._prog(0.78, "Construction tableau Fevrier...")
        tab_fev = build_tableau([MOIS_F25], [MOIS_F26],
                                "Fev2025", "Fev2026")

        # =====================================================================
        # EXPORT XLSX — 3 tableaux cote a cote
        # =====================================================================
        self._prog(0.85, "Export XLSX...")

        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl non disponible.")

        self.output_dir.mkdir(parents=True, exist_ok=True)
        save_path = self.output_dir / self.output_filename

        self._export_xlsx(tab_cum, tab_jan, tab_fev, str(save_path))
        return save_path

    # =========================================================================
    # EXPORT XLSX
    # =========================================================================
    def _export_xlsx(self, tab_cum, tab_jan, tab_fev, save_path):
        self._prog(0.90, "Ecriture Excel...")

        wb = Workbook()
        ws = wb.active
        ws.title = "Analyse Comparative"

        # Styles
        fill_vert    = PatternFill("solid", fgColor="00915A")
        fill_cum_hdr = PatternFill("solid", fgColor="1B5E20")   # vert fonce
        fill_jan_hdr = PatternFill("solid", fgColor="0D47A1")   # bleu fonce
        fill_fev_hdr = PatternFill("solid", fgColor="880E4F")   # bordeaux
        fill_neg     = PatternFill("solid", fgColor="FFCDD2")   # rouge clair
        fill_pos     = PatternFill("solid", fgColor="C8E6C9")   # vert clair

        font_title   = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
        font_hdr     = Font(name="Segoe UI", size=9,  bold=True, color="FFFFFF")
        font_norm    = Font(name="Segoe UI", size=9)
        font_bold    = Font(name="Segoe UI", size=9,  bold=True)
        align_ctr    = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_right  = Alignment(horizontal="right")
        align_left   = Alignment(horizontal="left")

        SEPARATEUR = 2   # colonnes vides entre tableaux
        tableaux = [
            (tab_cum, "CUMUL JAN+FEV 2025 vs JAN+FEV 2026", fill_cum_hdr),
            (tab_jan, "JANVIER 2025 vs JANVIER 2026",        fill_jan_hdr),
            (tab_fev, "FEVRIER 2025 vs FEVRIER 2026",        fill_fev_hdr),
        ]

        col_start = 1   # colonne Excel de depart pour le premier tableau
        ROW_TITRE = 2
        ROW_HDR   = 4
        ROW_DATA  = 5

        for tab_df, titre, fill_hdr in tableaux:
            cols = list(tab_df.columns)
            nb_cols = len(cols)

            # --- Titre du tableau (ligne 2, fusionne) ---
            titre_cell = ws.cell(row=ROW_TITRE, column=col_start, value=titre)
            titre_cell.font      = font_title
            titre_cell.fill      = fill_hdr
            titre_cell.alignment = align_ctr
            end_col = col_start + nb_cols - 1
            ws.merge_cells(
                start_row=ROW_TITRE, start_column=col_start,
                end_row=ROW_TITRE,   end_column=end_col
            )

            # --- Headers (ligne 4) ---
            for ci, col_name in enumerate(cols):
                cell = ws.cell(row=ROW_HDR, column=col_start + ci, value=col_name)
                cell.font      = font_hdr
                cell.fill      = fill_hdr
                cell.alignment = align_ctr

            # --- Donnees ---
            # Identifier les colonnes numeriques et les colonnes ECART/TENDANCE
            ecart_col_idx    = cols.index("ECART_MONTANT")       if "ECART_MONTANT"    in cols else None
            tendance_col_idx = cols.index("TENDANCE")            if "TENDANCE"         in cols else None

            for ri, row_vals in enumerate(tab_df.itertuples(index=False, name=None)):
                excel_row = ROW_DATA + ri
                tendance_val = row_vals[tendance_col_idx] if tendance_col_idx is not None else None

                for ci, val in enumerate(row_vals):
                    excel_col = col_start + ci
                    cell = ws.cell(row=excel_row, column=excel_col)

                    if isinstance(val, float) or isinstance(val, np.floating):
                        cell.value         = round(float(val), 2)
                        cell.number_format = '#,##0.00'
                        cell.alignment     = align_right
                        cell.font          = font_norm
                    else:
                        cell.value     = str(val) if val is not None else ''
                        cell.alignment = align_left
                        cell.font      = font_norm

                    # Coloration ECART_MONTANT selon tendance
                    if ecart_col_idx is not None and ci == ecart_col_idx:
                        if tendance_val == 'BAISSE':
                            cell.fill = fill_neg
                        elif tendance_val == 'HAUSSE':
                            cell.fill = fill_pos

            # --- Largeur colonnes ---
            for ci, col_name in enumerate(cols):
                letter = get_column_letter(col_start + ci)
                if col_name in ('ID_RP', 'CORPORATE_ACRONYM'):
                    ws.column_dimensions[letter].width = 20
                elif col_name.startswith('PNB_'):
                    ws.column_dimensions[letter].width = 16
                elif col_name in ('ECART_MONTANT', 'ECART_POURCENTAGE'):
                    ws.column_dimensions[letter].width = 15
                elif col_name == 'TYPE_PNB_MAX_VARIATION':
                    ws.column_dimensions[letter].width = 20
                else:
                    ws.column_dimensions[letter].width = 14

            # Avancer la colonne de depart pour le prochain tableau
            col_start = end_col + SEPARATEUR + 1

        # Figer la 1ere colonne (ID_RP) de chaque tableau n'est pas possible
        # dans openpyxl pour multi-zones, on fige juste le header
        ws.freeze_panes = ws.cell(row=ROW_DATA, column=1)

        self._prog(0.97, "Sauvegarde...")
        wb.save(save_path)
        self._prog(1.0, "Termine !")

        nb = len(tab_cum)
        print(
            f"Fichier Excel cree : {Path(save_path).name}\n"
            f"  - {nb:,} clients analyses\n"
            f"  - 3 tableaux cote a cote :\n"
            f"      1. Cumul Jan+Fev 2025 vs 2026\n"
            f"      2. Janvier 2025 vs 2026\n"
            f"      3. Fevrier 2025 vs 2026\n"
            f"  Cellules rouges = baisses | Cellules vertes = hausses"
        )


def main(argv: Optional[list] = None) -> int:
    parser = argparse.ArgumentParser(
        prog="01.FSUB3",
        description="MONEXT COMPARAISON ANNUELLE [FSUB3] — comparaison PNB "
                    "par client (ID RP) entre jan-fev 2025 et jan-fev 2026.",
    )
    parser.add_argument(
        "--source-monext-detail", required=True, type=Path,
        help="CSV detail produit par MONEXT ANALYZER [K2P8N] (obligatoire)."
    )
    parser.add_argument(
        "--output-dir", required=True, type=Path,
        help="Dossier de sortie du fichier XLSX (obligatoire)."
    )
    parser.add_argument(
        "--output-filename", required=True, type=str,
        help="Nom du fichier XLSX de sortie (obligatoire)."
    )
    args = parser.parse_args(argv)

    if not args.source_monext_detail.is_file():
        print(f"[ERREUR] Fichier source introuvable : {args.source_monext_detail}",
              file=sys.stderr)
        return 2

    try:
        app = MonextComparaison_FSUB3(args)
        save_path = app.run()
    except Exception as e:
        print(f"[ERREUR] {e}", file=sys.stderr)
        traceback.print_exc()
        return 1

    print(f"[OK] Analyse comparative [FSUB3] terminee : {save_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
