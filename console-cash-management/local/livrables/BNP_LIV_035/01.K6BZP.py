"""
CIB COMMISSION ANALYZER v14 [K6BZP] — CLI
=========================================
BNP Paribas Cash Management — Direction Monétique
Livrable BNP_LIV_035

DESCRIPTION
-----------
Analyseur de commissions CIB (apporteur d'affaires) BNP Paribas. Recoupe les flux
WORLDLINE (CPC) et MONEXT (CCO) avec les référentiels d'identité/segment, convertit
les flux WORLDLINE en EUR ligne à ligne, nettoie les noms de pays en anglais, puis
recalcule COMPLÈTEMENT par année (plafond / périodicité / différé) les coûts CCO/CPC,
RWA, EAD, commissions HT/TTC, applicabilité et totaux en valeurs natives (sans formule
Excel). REBATE cloisonné par (année, RMPM). Produit 7 onglets XLSX + CSV.

Cette version est la transposition CLI (argparse, sans GUI) de l'application
customtkinter d'origine. La LOGIQUE MÉTIER est PRÉSERVÉE À L'IDENTIQUE : aucun calcul,
aucune condition, aucune colonne, aucun recalcul par année n'a été modifié. Seules les
couches d'interface (customtkinter, filedialog, messagebox, mainloop, threading) ont
été retirées et remplacées par des entrées en ligne de commande. Le mapping des
colonnes reproduit le comportement PAR DÉFAUT de l'UI (positions présélectionnées de
DEFAULT_POSITIONS) ; l'étape visuelle de vérification est déléguée à l'UI web.

SOURCES REQUISES (8 obligatoires + 3 optionnelles)
--------------------------------------------------
  Obligatoires :
    WORLDLINE   00. PRGM Worldline — Flux CPC
    MONEXT      01. MONEXT — Flux CCO
    ACCOUNT     02. IBAN_ACCOUNT — Pivot 1 (IBAN/RMPM -> identité)
    REF_CLIENT  03. REFERENTIEL_CLIENT — Pivot 2 (RC/RIB/RP -> identité + segment)
    IDSEG       04. IDENTIFIANT_SEGMENT — Pivot 3 (segment fallback)
    DEVISES     05. DEVISES — Conversion Worldline -> EUR
    REBATE      06. REBATE — Année | RMPM | Plateforme | Montant
    COUNTRY     10. COUNTRY — Mapping pays (Original -> Anglais)
  Optionnelles :
    OVERRIDE     07. OVERRIDE_PAYS — écrase le pays apporteur
    BEJO_CARTES  08. BEJO cartes — override MX par entité (entité | nb)
    BEJO_FLUX    09. BEJO flux — override MX par entité (entité | IBAN)

OUTPUTS PRODUITS
----------------
  <output-dir>/<output-filename>.csv   Données clients consolidées (sep ';', utf-8-sig)
  <output-dir>/<output-filename>.xlsx  Classeur 7 onglets :
       DICTIONNAIRE, COMMISSIONS PAR PAYS, GLOBAL/CCO/CPC par année, SYNTHÈSE.

ARGUMENTS CLI
-------------
  --worldline PATH        (oblig) Fichier WORLDLINE PRGM (CPC)
  --monext PATH           (oblig) Fichier MONEXT (CCO)
  --account PATH          (oblig) Fichier IBAN_ACCOUNT
  --ref-client PATH       (oblig) Fichier REFERENTIEL_CLIENT
  --idseg PATH            (oblig) Fichier IDENTIFIANT_SEGMENT
  --devises PATH          (oblig) Fichier DEVISES
  --rebate PATH           (oblig) Fichier REBATE
  --country PATH          (oblig) Fichier COUNTRY (mapping pays)
  --override-pays-file PATH        (opt) Fichier OVERRIDE_PAYS
  --override-pays VALUE            (opt) Pays de remplacement (défaut UI : LUXEMBOURG)
  --bejo-cartes PATH               (opt) Fichier BEJO_CARTES
  --bejo-flux PATH                 (opt) Fichier BEJO_FLUX
  --output-dir PATH       (oblig) Dossier de sortie
  --output-filename NAME  (oblig) Nom de base commun CSV + XLSX (sans extension)

DECOMPOSITION
-------------
  main()
    └─ argparse -> instancie CIBCommissionAnalyzer_K6BZP(args)
        └─ worker()                          recoupement + recalcul par année
            ├─ load_csv_smart()              lecture robuste CSV
            ├─ _load_country_map()           mapping pays Original -> Anglais
            ├─ _clean_pays_value()           normalisation pays
            ├─ _build_idseg_dicts()          dictionnaires segment (cascade)
            ├─ resolve_taux()                conversion devise -> EUR (ligne à ligne)
            └─ _build_dataframe_and_save()   consolidation par client
                ├─ _compute_financials()     coûts/RWA/EAD/commissions par année
                ├─ _apply_rebate_to_df_year()  REBATE cloisonné (année, RMPM)
                └─ generate_xlsx()           7 onglets natifs
                    ├─ _write_native_sheet()        GLOBAL / CCO / CPC par année
                    ├─ _build_pays_table_sheet()    COMMISSIONS PAR PAYS
                    ├─ _build_dictionnaire_sheet()  DICTIONNAIRE
                    └─ _build_synthese_sheet()      SYNTHÈSE 4 blocs
"""

import argparse
import sys
import pandas as pd
import numpy as np
import os
import unicodedata
import re as _re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

BNP_GREEN  = "#00915A"
VERSION_ID = "K6BZP"

MOIS_NOMS = {
    '01': 'JANVIER', '02': 'FEVRIER', '03': 'MARS', '04': 'AVRIL',
    '05': 'MAI', '06': 'JUIN', '07': 'JUILLET', '08': 'AOUT',
    '09': 'SEPTEMBRE', '10': 'OCTOBRE', '11': 'NOVEMBRE', '12': 'DECEMBRE'
}

# Pays exclus de la table « Commissions par Pays » (noms ANGLAIS normalisés)
PAYS_EXCLUS_COMMISSION = {'BELGIUM', 'FRANCE', 'LUXEMBOURG'}
# Mot-clé identifiant la « carte achat » (exclusion CPC) — recherché dans NOM_PROG_CPC.
CARTE_ACHAT_KEYWORD = 'ACHAT'

# ─────────────────────────────────────────────────────────────────────────────
# POSITIONS PAR DÉFAUT (présélection des colonnes dans l'UI)
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_POSITIONS = {
    # 00 — PRGM WORLDLINE (flux CPC)
    'wl_mois': 2, 'wl_nom_prog': 3, 'wl_rs': 8, 'wl_iban': 9,
    'wl_devise': 12, 'wl_plafond': 13, 'wl_periodicite': 14, 'wl_nb_cartes': 17,
    'wl_nb_tr_fourn': 21, 'wl_nb_tr_cb': 22, 'wl_nb_tr_visa': 23,
    'wl_dep_1': 28, 'wl_dep_2': 29, 'wl_pnb_cols': [30, 31, 32, 33, 34, 35, 36],
    'wl_rc': 40, 'wl_differe': 41, 'wl_conv_first': 24, 'wl_conv_last': 36,

    # 01 — MONEXT (flux CCO)
    'mx_mois': 1, 'mx_nom_prog': 2, 'mx_rs': 4, 'mx_rp': 9, 'mx_rc': 10,
    'mx_iban': 11, 'mx_nb_cartes': 12, 'mx_differe': 13,
    'mx_depenses': 15, 'mx_nb_transactions': 16, 'mx_retraits': 17,
    'mx_pnb_first': 19, 'mx_pnb_last': 55, 'mx_pnb_excl': 33, 'mx_interchange': 21,

    # 02 — IBAN_ACCOUNT (pivot 1)
    'acc_pays_ga': 1, 'acc_code_ga': 2, 'acc_nom_ga': 3,
    'acc_pays_le': 4, 'acc_rmpm': 5, 'acc_nom_le': 6, 'acc_iban': 7,

    # 03 — REFERENTIEL_CLIENT (pivot 2)
    'ref_rib': 2, 'ref_rmpm': 4, 'ref_rc': 7,
    'ref_segment': 9, 'ref_code_ga': 10, 'ref_nom_ga': 11,

    # 04 — IDENTIFIANT_SEGMENT (pivot 3)
    'idseg_type': 1, 'idseg_id': 2, 'idseg_segment': 3,

    # 05 — DEVISES
    'devise_date': 1, 'devise_code': 2, 'devise_taux': 3,

    # 06 — REBATE (Année | RMPM | Plateforme | Montant)
    'reb_annee': 1, 'reb_rmpm': 2, 'reb_plateforme': 3, 'reb_montant': 4,

    # 07 — OVERRIDE_PAYS (optionnel)
    'ovr_type': 1, 'ovr_valeur': 2, 'ovr_nom': 3,

    # 08-09 — BEJO (optionnels)
    'bejo_cartes_entite': 1, 'bejo_cartes_nb': 3,
    'bejo_flux_entite': 1, 'bejo_flux_iban': 6,

    # 10 — COUNTRY (mapping pays → anglais) [NOUVEAU]
    'country_original': 1, 'country_new': 2,
}

DEFAULT_CONSTANTS = {
    'taux_refinancement': 0.0099,         # 2025
    'taux_refinancement_2026': 0.0166,    # 2026 (NOUVEAU)
    'part_capital_banque': 0.12,
    'almt': 0.142, 'rw_defaut': 0.65,
    'cout_carte_cco': 40.14, 'cout_transaction_cpc': 0.59,
    'cout_rwa_plafond_cco': 13.32,
    'taux_ead_porteur': 0.10, 'taux_ead_entreprise': 0.40,
    'taux_commission': 0.20, 'taux_tva': 0.20,
}

# ─────────────────────────────────────────────────────────────────────────────
# PALETTE COULEURS (palette sobre BNP — reprise NO7WK)
# ─────────────────────────────────────────────────────────────────────────────
BLOC_COLORS = {
    'A_IDENTITY': {'name': 'Identité client', 'header': '1F4E79', 'banner': '4A78A8',
                   'cell': 'EAF1F8', 'font': 'FFFFFF'},
    'B_MATCHING': {'name': 'Méthode de matching', 'header': '00665E', 'banner': '4E8E88',
                   'cell': 'E6EFEE', 'font': 'FFFFFF'},
    'C_FLUX_MX': {'name': 'Flux MONEXT (CCO)', 'header': '5C5470',
                  'banner': '8A82A0', 'cell': 'EFEDF2', 'font': 'FFFFFF'},
    'D_FLUX_WL': {'name': 'Flux WORLDLINE (CPC)', 'header': '6B4423',
                  'banner': '9C7A55', 'cell': 'F2EDE6', 'font': 'FFFFFF'},
    'E_FINANCE': {'name': 'Calculs financiers (valeurs natives)', 'header': '0D3B66',
                  'banner': '4A6A8E', 'cell': 'E6ECF2', 'font': 'FFFFFF'},
    'F_DIAG': {'name': 'Diagnostic', 'header': '595959', 'banner': '8C8C8C',
               'cell': 'F7F7F7', 'font': 'FFFFFF'},
}

# Raccourcis couleurs
GRN = "00915A"; GRN2 = "E8F5E9"; DARK = "1C3A2D"; WHT = "FFFFFF"
BLU = "1F4E79"; BLU2 = "EAF1F8"
ORA = "E65100"; ORA2 = "FFF3E0"
PUR = "5C5470"; PUR2 = "EFEDF2"
BRN = "6B4423"; BRN2 = "F2EDE6"
COL_GAP = 2

# UI variants
GRN_UI = "#00915A"; GRN2_UI = "#E8F5E9"
BLU_UI = "#1F4E79"; BLU2_UI = "#EAF1F8"
ORA_UI = "#E65100"; ORA2_UI = "#FFF3E0"
PUR_UI = "#5C5470"; PUR2_UI = "#EFEDF2"
BRN_UI = "#6B4423"; BRN2_UI = "#F2EDE6"


# ═════════════════════════════════════════════════════════════════════════════
# APPLICATION
# ═════════════════════════════════════════════════════════════════════════════
class CIBCommissionAnalyzer_K6BZP:

    def __init__(self, args: argparse.Namespace) -> None:
        # Slots fichiers (8 obligatoires + 3 optionnels) — alimentés par argparse.
        self.files: Dict[str, str] = {
            "WORLDLINE": str(args.worldline),     # 00 PRGM
            "MONEXT": str(args.monext),           # 01
            "ACCOUNT": str(args.account),         # 02 IBAN_ACCOUNT
            "REF_CLIENT": str(args.ref_client),   # 03
            "IDSEG": str(args.idseg),             # 04
            "DEVISES": str(args.devises),         # 05
            "REBATE": str(args.rebate),           # 06
            "COUNTRY": str(args.country),         # 10 — nettoyage pays
            "OVERRIDE": str(args.override_pays_file) if args.override_pays_file else "",  # 07 — opt
            "BEJO_CARTES": str(args.bejo_cartes) if args.bejo_cartes else "",            # 08 — opt
            "BEJO_FLUX": str(args.bejo_flux) if args.bejo_flux else "",                  # 09 — opt
        }

        # Options optionnelles (équivalent des checkboxes UI)
        self.use_override: bool = bool(args.override_pays_file)
        self.use_bejo: bool = bool(args.bejo_cartes and args.bejo_flux)
        # Valeur de remplacement pays (défaut UI : LUXEMBOURG)
        self.override_pays_value: str = str(args.override_pays) if args.override_pays else ""

        # Sortie (remplace la boîte de dialogue _ask_save_path)
        self.output_dir: str = str(args.output_dir)
        self.output_filename: str = str(args.output_filename)

        # NOUVEAU — nettoyage des pays (mapping + valeurs non mappées)
        self.manual_map_pays: Dict[str, str] = {}
        self._pays_targets: List[str] = []

    # ═════════════════════════════════════════════════════════════════════════
    # PRÉPARATION CLI — mapping colonnes par DÉFAUT (positions présélectionnées UI)
    # ═════════════════════════════════════════════════════════════════════════
    def _default_col_name(self, fkey: str, pos: int) -> str:
        """Renvoie le nom de colonne à la position 1-based `pos` du fichier `fkey`.
        Reproduit la présélection des combobox de l'UI (DEFAULT_POSITIONS)."""
        cols = self.original_cols.get(fkey, [])
        if 1 <= pos <= len(cols):
            return cols[pos - 1]
        return ""

    def _build_default_mapping(self) -> Dict[str, str]:
        """Construit le dict `m` {clé_mapping: nom_colonne} reproduisant le comportement
        PAR DÉFAUT de l'UI. Vérification colonnes : déléguée à l'UI web."""
        # Affectation clé_mapping -> (fichier source, position par défaut 1-based)
        field_source = {
            'wl_mois': 'WORLDLINE', 'wl_nom_prog': 'WORLDLINE', 'wl_rs': 'WORLDLINE',
            'wl_iban': 'WORLDLINE', 'wl_devise': 'WORLDLINE', 'wl_plafond': 'WORLDLINE',
            'wl_periodicite': 'WORLDLINE', 'wl_nb_cartes': 'WORLDLINE',
            'wl_nb_tr_fourn': 'WORLDLINE', 'wl_nb_tr_cb': 'WORLDLINE',
            'wl_nb_tr_visa': 'WORLDLINE', 'wl_dep_1': 'WORLDLINE', 'wl_dep_2': 'WORLDLINE',
            'wl_rc': 'WORLDLINE', 'wl_differe': 'WORLDLINE',
            'wl_conv_first': 'WORLDLINE', 'wl_conv_last': 'WORLDLINE',
            'mx_mois': 'MONEXT', 'mx_nom_prog': 'MONEXT', 'mx_rs': 'MONEXT',
            'mx_rp': 'MONEXT', 'mx_rc': 'MONEXT', 'mx_iban': 'MONEXT',
            'mx_nb_cartes': 'MONEXT', 'mx_differe': 'MONEXT', 'mx_depenses': 'MONEXT',
            'mx_nb_transactions': 'MONEXT', 'mx_retraits': 'MONEXT',
            'acc_pays_ga': 'ACCOUNT', 'acc_code_ga': 'ACCOUNT', 'acc_nom_ga': 'ACCOUNT',
            'acc_pays_le': 'ACCOUNT', 'acc_rmpm': 'ACCOUNT', 'acc_nom_le': 'ACCOUNT',
            'acc_iban': 'ACCOUNT',
            'ref_rib': 'REF_CLIENT', 'ref_rmpm': 'REF_CLIENT', 'ref_rc': 'REF_CLIENT',
            'ref_segment': 'REF_CLIENT', 'ref_code_ga': 'REF_CLIENT',
            'ref_nom_ga': 'REF_CLIENT',
            'idseg_type': 'IDSEG', 'idseg_id': 'IDSEG', 'idseg_segment': 'IDSEG',
            'devise_date': 'DEVISES', 'devise_code': 'DEVISES', 'devise_taux': 'DEVISES',
            'reb_annee': 'REBATE', 'reb_rmpm': 'REBATE', 'reb_plateforme': 'REBATE',
            'reb_montant': 'REBATE',
            'country_original': 'COUNTRY', 'country_new': 'COUNTRY',
        }
        m: Dict[str, str] = {}
        for key, fkey in field_source.items():
            pos = DEFAULT_POSITIONS.get(key)
            if isinstance(pos, int):
                m[key] = self._default_col_name(fkey, pos)
        # OVERRIDE (optionnel)
        if self.use_override and "OVERRIDE" in self.original_cols:
            for key in ('ovr_type', 'ovr_valeur', 'ovr_nom'):
                m[key] = self._default_col_name('OVERRIDE', DEFAULT_POSITIONS[key])
        # BEJO (optionnel)
        if self.use_bejo:
            if "BEJO_CARTES" in self.original_cols:
                m['bejo_cartes_entite'] = self._default_col_name(
                    'BEJO_CARTES', DEFAULT_POSITIONS['bejo_cartes_entite'])
                m['bejo_cartes_nb'] = self._default_col_name(
                    'BEJO_CARTES', DEFAULT_POSITIONS['bejo_cartes_nb'])
            if "BEJO_FLUX" in self.original_cols:
                m['bejo_flux_entite'] = self._default_col_name(
                    'BEJO_FLUX', DEFAULT_POSITIONS['bejo_flux_entite'])
                m['bejo_flux_iban'] = self._default_col_name(
                    'BEJO_FLUX', DEFAULT_POSITIONS['bejo_flux_iban'])
        return m

    # ═════════════════════════════════════════════════════════════════════════
    # CHARGEMENT CSV
    # ═════════════════════════════════════════════════════════════════════════
    def load_csv_smart(self, path, nrows=None):
        for sep in [';', ',', '\t']:
            for enc in ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']:
                try:
                    df = pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                     keep_default_na=False, na_values=[],
                                     on_bad_lines='skip', nrows=5)
                    if df.shape[1] > 1:
                        return pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                           keep_default_na=False, na_values=[],
                                           on_bad_lines='skip', nrows=nrows)
                except Exception:
                    continue
        return pd.read_csv(path, sep=None, engine='python', dtype=str,
                           on_bad_lines='skip', nrows=nrows)

    # ═════════════════════════════════════════════════════════════════════════
    # NETTOYAGE DES PAYS — même logique que le programme REVENUS
    # ═════════════════════════════════════════════════════════════════════════
    @staticmethod
    def norm_map(v):
        """Normalise une valeur de pays pour servir de clé de mapping :
        majuscules, sans accents, espaces compressés."""
        if v is None:
            return ''
        s = unicodedata.normalize('NFD', str(v).strip().upper())
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        s = _re.sub(r'\s+', ' ', s).strip()
        return s

    def _load_country_map(self):
        """Charge le fichier COUNTRY → dict {norm(original): new_anglais} + set des cibles."""
        country_map = {}
        targets = set()
        try:
            df_c = self.load_csv_smart(self.files["COUNTRY"])
        except Exception:
            return country_map, sorted(targets)
        m = getattr(self, '_m', {})
        col_o = m['country_original'] \
            if m.get('country_original') else df_c.columns[0]
        col_n = m['country_new'] \
            if m.get('country_new') else (
                df_c.columns[1] if len(df_c.columns) > 1 else df_c.columns[0])
        for o, n in zip(df_c[col_o].astype(str).values, df_c[col_n].astype(str).values):
            ko = self.norm_map(o)
            new_clean = str(n).strip()
            if not ko or not new_clean:
                continue
            country_map[ko] = new_clean
            targets.add(new_clean.upper())
            # la cible se mappe aussi sur elle-même
            country_map.setdefault(self.norm_map(new_clean), new_clean)
        return country_map, sorted(targets)

    def _clean_pays_value(self, raw):
        """Renvoie le nom de pays nettoyé (anglais) à partir d'une valeur brute."""
        if raw is None:
            return ''
        s = str(raw).strip()
        if s == '':
            return ''
        k = self.norm_map(s)
        cmap = getattr(self, '_country_map', {})
        mmap = getattr(self, 'manual_map_pays', {})
        if k in cmap:
            return cmap[k]
        if k in mmap:
            return mmap[k]
        return s
    # ═════════════════════════════════════════════════════════════════════════
    # UTILITAIRES (statiques)
    # ═════════════════════════════════════════════════════════════════════════
    def _col_name(self, s):
        return s.split(". ", 1)[1] if ". " in s else s

    def _col_idx(self, s):
        try:
            return int(s.split(". ")[0])
        except Exception:
            return 0

    @staticmethod
    def clean_id_safe(series):
        s = series.astype(str).str.strip()
        s = s.replace(['', 'nan', 'NaN', 'None', 'NULL', 'NA', 'N/A', 'NAN', 'NONE'], '')
        mask = s.str.startswith('="') & s.str.endswith('"')
        s = s.where(~mask, s.str[2:-1])
        s = s.str.lstrip("'")
        mask2 = s.str.endswith('.0') & s.str[:-2].str.isdigit()
        return s.where(~mask2, s.str[:-2]).str.strip()

    @staticmethod
    def clean_id_strip0(series):
        s = CIBCommissionAnalyzer_K6BZP.clean_id_safe(series)
        stripped = s.str.lstrip('0')
        return stripped.where(stripped != '', s)

    @staticmethod
    def clean_iban_truncated(series):
        s = CIBCommissionAnalyzer_K6BZP.clean_id_safe(series).str.upper().str.replace(' ', '', regex=False)
        return s.str[4:].where(s.str.len() > 4, s)

    @staticmethod
    def clean_iban_full(series):
        return CIBCommissionAnalyzer_K6BZP.clean_id_safe(series).str.upper().str.replace(' ', '', regex=False)

    @staticmethod
    def clean_rib_ref(series):
        return CIBCommissionAnalyzer_K6BZP.clean_id_safe(series).str.upper().str.replace(' ', '', regex=False)

    @staticmethod
    def normalize_rs(series):
        def _n(v):
            if pd.isna(v) or str(v).strip() == '':
                return ''
            s = unicodedata.normalize('NFD', str(v).strip().upper())
            return ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return series.apply(_n)

    @staticmethod
    def to_float(series):
        s = series.astype(str)
        for rep in ['"', "'", ' ', '\xa0', '\u202f', '\u20ac']:
            s = s.str.replace(rep, '', regex=False)
        s = s.str.replace('EUR', '', regex=False)
        mask = s.str.endswith('-')
        s = s.where(~mask, '-' + s.str[:-1])
        s = s.str.replace(',', '.', regex=False)
        return pd.to_numeric(s, errors='coerce').fillna(0.0)

    @staticmethod
    def parse_mois(series):
        def _p(val):
            if pd.isna(val):
                return ''
            s = str(val).strip()
            if s.startswith('="') and s.endswith('"'):
                s = s[2:-1].strip()
            s = s.lstrip("'").strip()
            if s.endswith('.0') and s[:-2].isdigit():
                s = s[:-2]
            if _re.fullmatch(r'\d{6}', s):
                return s
            m = _re.match(r'^(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})$', s)
            if m:
                mo, an = int(m.group(2)), m.group(3)
                if 1 <= mo <= 12:
                    return f"{an}{str(mo).zfill(2)}"
            m = _re.match(r'^(\d{1,2})[/\-\.](\d{4})$', s)
            if m:
                mo, an = int(m.group(1)), m.group(2)
                if 1 <= mo <= 12:
                    return f"{an}{str(mo).zfill(2)}"
            m = _re.match(r'^(\d{4})[/\-\.](\d{1,2})$', s)
            if m:
                an, mo = m.group(1), int(m.group(2))
                if 1 <= mo <= 12:
                    return f"{an}{str(mo).zfill(2)}"
            try:
                dt = pd.to_datetime(s, dayfirst=True, errors='coerce')
                if pd.notna(dt):
                    return dt.strftime('%Y%m')
            except Exception:
                pass
            return ''
        return series.apply(_p)

    @staticmethod
    def mois_label(yyyymm):
        if len(yyyymm) == 6:
            return f"{yyyymm[:4]}_{MOIS_NOMS.get(yyyymm[4:], yyyymm[4:])}"
        return yyyymm

    @staticmethod
    def protect_id(val):
        if pd.isna(val) or str(val).strip() == '':
            return ''
        s = str(val).strip()
        if s.startswith('="') and s.endswith('"'):
            return s
        if s.isdigit():
            return f'="{s}"'
        if s.startswith('0') and len(s) > 1 and s[1:].isdigit():
            return f'="{s}"'
        return s

    @staticmethod
    def _differe_type(raw, source='CCO'):
        if raw is None:
            return ''
        s = str(raw).strip().upper()
        if s == '' or s in ('NAN', 'NONE', 'NULL'):
            return ''
        if 'IMM' in s:
            return 'IMM'
        s_num = s.replace(',', '.').replace(' ', '')
        try:
            v = float(s_num)
            if v == 0:
                return 'FIN_MOIS' if source == 'CCO' else 'IMM'
            return 'DIFFERE'
        except ValueError:
            return s[:16]

    @staticmethod
    def _differe_jours(raw):
        if raw is None:
            return 0
        s = str(raw).strip().upper()
        if s == '' or 'IMM' in s:
            return 0
        try:
            return int(float(s.replace(',', '.').replace(' ', '')))
        except ValueError:
            return 0

    @staticmethod
    def normalize_segment(raw):
        if pd.isna(raw):
            return ''
        s = str(raw).strip().upper()
        if 'BPE' in s:
            return 'BPE'
        if 'ENTREPRISE' in s or 'CORPORATE' in s or 'BCEF' in s:
            return 'ENTREPRISE'
        return ''

    def _prog(self, v: float, t: str) -> None:
        # Remplace la barre de progression UI : impression en console.
        print(f"[{v*100:5.1f}%] {t}")

    # ═════════════════════════════════════════════════════════════════════════
    # PRÉPARATION CLI — équivalent de load_previews + start_thread (sans GUI)
    # Lit les en-têtes de colonnes, charge le mapping pays, fixe les colonnes pays
    # (mêmes positions que l'UI : WL col 38, MX col 6), construit le mapping par
    # défaut et tous les arguments du worker (comportement présélectionné de l'UI).
    # Vérification colonnes : déléguée à l'UI web.
    # ═════════════════════════════════════════════════════════════════════════
    def _load_previews(self) -> None:
        required = ["WORLDLINE", "MONEXT", "ACCOUNT", "REF_CLIENT", "IDSEG",
                    "DEVISES", "REBATE", "COUNTRY"]
        for k in required:
            if not self.files[k]:
                raise ValueError(f"Fichier {k} obligatoire manquant.")
        keys = list(required)
        if self.use_override:
            keys.append("OVERRIDE")
        if self.use_bejo:
            keys.extend(["BEJO_CARTES", "BEJO_FLUX"])
        self.dfs_preview: Dict[str, Any] = {}
        self.original_cols: Dict[str, List[str]] = {}
        for k in keys:
            df = self.load_csv_smart(self.files[k], nrows=5)
            self.dfs_preview[k] = df
            self.original_cols[k] = list(df.columns)

    def _prepare_worker_args(self) -> Dict[str, Any]:
        """Construit tous les arguments du worker à partir du comportement PAR DÉFAUT
        de l'UI (positions présélectionnées, France pré-cochée, plage complète)."""
        self._load_previews()

        # Mapping colonnes par défaut (positions présélectionnées de l'UI)
        m = self._build_default_mapping()
        self._m = m

        # Colonnes pays apporteurs (mêmes positions que l'UI : WL col 38, MX col 6)
        df_wl = self.load_csv_smart(self.files["WORLDLINE"])
        df_mx = self.load_csv_smart(self.files["MONEXT"])
        wl_cols = list(df_wl.columns)
        mx_cols = list(df_mx.columns)
        self._wl_pays_col = wl_cols[38] if len(wl_cols) > 38 else wl_cols[-1]
        self._mx_pays_col = mx_cols[6] if len(mx_cols) > 6 else mx_cols[-1]

        # Nettoyage pays : mapping COUNTRY chargé. Les valeurs non mappées sont
        # gardées telles quelles (équivalent du parcours "skip" de l'UI : aucune
        # association manuelle). Vérification colonnes : déléguée à l'UI web.
        country_map, targets = self._load_country_map()
        self._country_map = country_map
        self._pays_targets = targets
        self.manual_map_pays = {}

        # Pays à exclure : reproduit la présélection de l'UI (France pré-cochée).
        pays_wl = sorted({self._clean_pays_value(v)
                          for v in df_wl[self._wl_pays_col].astype(str).tolist()
                          if self._clean_pays_value(v)})
        pays_mx = sorted({self._clean_pays_value(v)
                          for v in df_mx[self._mx_pays_col].astype(str).tolist()
                          if self._clean_pays_value(v)})
        pays_excl_wl = {p for p in pays_wl if p.strip().upper() in ['FRANCE', 'FR']}
        pays_excl_mx = {p for p in pays_mx if p.strip().upper() in ['FRANCE', 'FR']}

        # Colonnes PNB WORLDLINE (présélection DEFAULT_POSITIONS['wl_pnb_cols'])
        wl_pnb_cols = []
        for pos in DEFAULT_POSITIONS['wl_pnb_cols']:
            if pos <= len(wl_cols):
                wl_pnb_cols.append(wl_cols[pos - 1])

        # PNB MONEXT : plage + exclusion + interchange (indices 1-based présélectionnés)
        mx_pnb_cfg = {
            'first': DEFAULT_POSITIONS['mx_pnb_first'],
            'last': DEFAULT_POSITIONS['mx_pnb_last'],
            'excl': DEFAULT_POSITIONS['mx_pnb_excl'],
            'interchange': DEFAULT_POSITIONS['mx_interchange'],
        }
        # Conversion devises WORLDLINE : bornes (indices 1-based présélectionnés)
        wl_conv_cfg = {
            'first': DEFAULT_POSITIONS['wl_conv_first'],
            'last': DEFAULT_POSITIONS['wl_conv_last'],
        }

        # Constantes financières : valeurs par défaut (champs UI non édités).
        constants = {k: float(v) for k, v in DEFAULT_CONSTANTS.items()}

        # Plage de mois : vide => plage complète (présélection début..fin de l'UI).
        plage_cfg = {'debut': '', 'fin': ''}

        # Pays de remplacement (override) — nettoyé comme dans l'UI.
        override_pays = self.override_pays_value.strip() if self.use_override else ''
        if override_pays:
            override_pays = self._clean_pays_value(override_pays)

        return {
            'm': m, 'pays_excl_wl': pays_excl_wl, 'pays_excl_mx': pays_excl_mx,
            'wl_pnb_cols': wl_pnb_cols, 'mx_pnb_cfg': mx_pnb_cfg,
            'wl_conv_cfg': wl_conv_cfg, 'constants': constants,
            'plage_cfg': plage_cfg, 'override_pays': override_pays,
            'use_bejo': self.use_bejo,
        }

    def run(self) -> None:
        """Point d'entrée CLI : prépare les arguments puis exécute le worker."""
        a = self._prepare_worker_args()
        self.worker(a['m'], a['pays_excl_wl'], a['pays_excl_mx'], a['wl_pnb_cols'],
                    a['mx_pnb_cfg'], a['wl_conv_cfg'], a['constants'], a['plage_cfg'],
                    a['override_pays'], a['use_bejo'])

    # ═════════════════════════════════════════════════════════════════════════
    # WORKER PRINCIPAL
    # ═════════════════════════════════════════════════════════════════════════
    def worker(self, m, pays_excl_wl, pays_excl_mx, wl_pnb_cols, mx_pnb_cfg,
               wl_conv_cfg, constants, plage_cfg, override_pays, use_bejo):
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")

            self._prog(0.02, "Chargement WORLDLINE...")
            df_wl = self.load_csv_smart(self.files["WORLDLINE"])
            self._prog(0.04, "Chargement MONEXT...")
            df_mx = self.load_csv_smart(self.files["MONEXT"])
            self._prog(0.06, "Chargement IBAN_ACCOUNT...")
            df_account = self.load_csv_smart(self.files["ACCOUNT"])
            self._prog(0.07, "Chargement REFERENTIEL_CLIENT...")
            df_ref = self.load_csv_smart(self.files["REF_CLIENT"])
            self._prog(0.08, "Chargement IDENTIFIANT_SEGMENT...")
            df_idseg = self.load_csv_smart(self.files["IDSEG"])
            self._prog(0.09, "Chargement DEVISES...")
            df_devises = self.load_csv_smart(self.files["DEVISES"])
            self._prog(0.095, "Chargement REBATE...")
            df_rebate = self.load_csv_smart(self.files["REBATE"])

            ovr_rmpm_set = set()
            ovr_idprog_set = set()
            if override_pays and self.files.get("OVERRIDE"):
                self._prog(0.098, "Chargement OVERRIDE...")
                df_ovr = self.load_csv_smart(self.files["OVERRIDE"])
                col_t = m.get('ovr_type', df_ovr.columns[0])
                col_v = m.get('ovr_valeur', df_ovr.columns[1])
                col_n = m.get('ovr_nom', df_ovr.columns[2]) if len(df_ovr.columns) >= 3 else None
                ovr_types = df_ovr[col_t].astype(str).str.strip().str.upper()
                ovr_vals = self.clean_id_safe(df_ovr[col_v])
                ovr_noms = df_ovr[col_n].astype(str).str.strip() if col_n \
                    else pd.Series('', index=df_ovr.index)
                for t, v, nom in zip(ovr_types.values, ovr_vals.values, ovr_noms.values):
                    if not v:
                        continue
                    if 'RMPM' in t:
                        ovr_rmpm_set.add(v)
                        stripped = v.lstrip('0')
                        if stripped and stripped != v:
                            ovr_rmpm_set.add(stripped)
                    elif 'PROG' in t or 'ID_PROG' in t:
                        ovr_idprog_set.add(v)
                        if nom:
                            ovr_idprog_set.add(nom)
                self._prog(0.098, f"OVERRIDE: RMPM={len(ovr_rmpm_set)} "
                                  f"ID_PROG={len(ovr_idprog_set)} → '{override_pays}'")

            # REBATE — dico (annee, rmpm) → {WORLDLINE, MONEXT}
            self._prog(0.10, "Indexation REBATE...")
            rebate_by_year_rmpm = {}
            col_reb_annee = m.get('reb_annee', df_rebate.columns[0])
            col_reb_rmpm = m.get('reb_rmpm', df_rebate.columns[1])
            col_reb_plt = m.get('reb_plateforme', df_rebate.columns[2])
            col_reb_mnt = m.get('reb_montant', df_rebate.columns[3])
            reb_annees = df_rebate[col_reb_annee].astype(str).str.strip()
            reb_rmpms = self.clean_id_safe(df_rebate[col_reb_rmpm])
            reb_plts = df_rebate[col_reb_plt].astype(str).str.strip().str.upper()
            reb_mnts = self.to_float(df_rebate[col_reb_mnt])
            for annee, rmpm, plt, mnt in zip(reb_annees.values, reb_rmpms.values,
                                              reb_plts.values, reb_mnts.values):
                annee_clean = ''.join(ch for ch in str(annee) if ch.isdigit())
                if len(annee_clean) >= 4:
                    annee_clean = annee_clean[:4]
                if not annee_clean or not rmpm:
                    continue
                plt_norm = ''
                if 'WORLDLINE' in plt or 'WL' in plt or 'CPC' in plt:
                    plt_norm = 'WORLDLINE'
                elif 'MONEXT' in plt or 'MX' in plt or 'CCO' in plt:
                    plt_norm = 'MONEXT'
                else:
                    continue
                key_ar = (annee_clean, rmpm)
                if key_ar not in rebate_by_year_rmpm:
                    rebate_by_year_rmpm[key_ar] = {'WORLDLINE': 0.0, 'MONEXT': 0.0}
                rebate_by_year_rmpm[key_ar][plt_norm] += float(mnt)
                stripped = rmpm.lstrip('0')
                if stripped and stripped != rmpm:
                    key_strip = (annee_clean, stripped)
                    if key_strip not in rebate_by_year_rmpm:
                        rebate_by_year_rmpm[key_strip] = {'WORLDLINE': 0.0, 'MONEXT': 0.0}
                    rebate_by_year_rmpm[key_strip][plt_norm] = rebate_by_year_rmpm[key_ar][plt_norm]
            self._prog(0.105, f"REBATE: {len(rebate_by_year_rmpm)} couples (annee, RMPM)")

            # REFERENTIEL_CLIENT dicts
            self._prog(0.11, "Indexation REFERENTIEL_CLIENT...")
            ref_rc_c = self.clean_id_strip0(df_ref[m['ref_rc']])
            ref_rc_r = self.clean_id_safe(df_ref[m['ref_rc']])
            ref_rib_c = self.clean_rib_ref(df_ref[m['ref_rib']])
            ref_seg = df_ref[m['ref_segment']].astype(str).str.strip().str.upper()
            ref_ga = self.clean_id_strip0(df_ref[m['ref_code_ga']])
            ref_nga = df_ref[m['ref_nom_ga']].astype(str).str.strip()
            ref_rmpm = self.clean_id_safe(df_ref[m['ref_rmpm']])
            ref_sn = ref_seg.apply(self.normalize_segment)

            d_ref_rc = {}
            for rc, rr, seg, ga, nga, rmpm in zip(ref_rc_c.values, ref_rc_r.values,
                                                    ref_sn.values, ref_ga.values,
                                                    ref_nga.values, ref_rmpm.values):
                if rr and rr not in d_ref_rc:
                    d_ref_rc[rr] = (seg, ga, nga, rmpm)
                if rc and rc != rr and rc not in d_ref_rc:
                    d_ref_rc[rc] = (seg, ga, nga, rmpm)
            d_ref_rib = {}
            for rib, seg, ga, nga, rmpm in zip(ref_rib_c.values, ref_sn.values,
                                                ref_ga.values, ref_nga.values,
                                                ref_rmpm.values):
                if rib and rib not in d_ref_rib:
                    d_ref_rib[rib] = (seg, ga, nga, rmpm)
            self._prog(0.12, f"REF: RC={len(d_ref_rc):,} RIB={len(d_ref_rib):,}")

            # IDSEG dicts
            self._prog(0.125, "Indexation IDSEG...")
            seg_by_rc, seg_by_rmpm, seg_by_rp, seg_by_ca = self._build_idseg_dicts(
                df_idseg, m['idseg_type'], m['idseg_id'], m['idseg_segment'])

            # DEVISES
            self._prog(0.13, "Indexation DEVISES...")
            dev_dates = df_devises[m['devise_date']].astype(str).str.strip()
            dev_codes = df_devises[m['devise_code']].astype(str).str.strip().str.upper()
            dev_taux = self.to_float(df_devises[m['devise_taux']])
            taux_dict = {}
            mois_par_devise = {}
            for d, c, t in zip(dev_dates.values, dev_codes.values, dev_taux.values):
                if not d or not c or t <= 0:
                    continue
                d_clean = ''.join(ch for ch in d if ch.isdigit())
                if len(d_clean) != 6:
                    continue
                taux_dict[(d_clean, c)] = float(t)
                mois_par_devise.setdefault(c, set()).add(d_clean)
            mois_par_devise = {dev: sorted(list(s)) for dev, s in mois_par_devise.items()}
            taux_dict[('', 'EUR')] = 1.0

            def resolve_taux(mois_val, devise_val):
                if not devise_val or devise_val == 'EUR':
                    return 1.0
                if devise_val not in mois_par_devise:
                    return None
                m_clean = ''.join(ch for ch in str(mois_val) if ch.isdigit())
                if len(m_clean) != 6:
                    return None
                key = (m_clean, devise_val)
                if key in taux_dict:
                    return taux_dict[key]
                candidat = None
                for md in mois_par_devise[devise_val]:
                    if md <= m_clean:
                        candidat = md
                    else:
                        break
                return taux_dict[(candidat, devise_val)] if candidat else None

            # IBAN_ACCOUNT dicts
            self._prog(0.15, "Indexation IBAN_ACCOUNT...")
            df_account['_IBAN'] = self.clean_iban_full(df_account[m['acc_iban']])
            df_account['_GA'] = self.clean_id_strip0(df_account[m['acc_code_ga']])
            df_account['_NOM_GA'] = df_account[m['acc_nom_ga']].astype(str).str.strip()
            df_account['_PAYS_GA'] = df_account[m['acc_pays_ga']].astype(str).str.strip()
            df_account['_RMPM'] = self.clean_id_safe(df_account[m['acc_rmpm']])
            df_account['_NOM_LE'] = df_account[m['acc_nom_le']].astype(str).str.strip()
            df_account['_PAYS_LE'] = df_account[m['acc_pays_le']].astype(str).str.strip()

            acc_iban_dict = {}
            for ib, ga, nga, pga, rmpm, nle, ple in zip(
                df_account['_IBAN'].values, df_account['_GA'].values,
                df_account['_NOM_GA'].values, df_account['_PAYS_GA'].values,
                df_account['_RMPM'].values, df_account['_NOM_LE'].values,
                df_account['_PAYS_LE'].values
            ):
                if ib and ib not in acc_iban_dict:
                    acc_iban_dict[ib] = (ga, nga, pga, rmpm, nle, ple)

            acc_ga2n, acc_ga2p, acc_r2n, acc_r2p = {}, {}, {}, {}
            for ga, nga, pga in zip(df_account['_GA'].values,
                                     df_account['_NOM_GA'].values,
                                     df_account['_PAYS_GA'].values):
                if ga and ga not in acc_ga2n:
                    acc_ga2n[ga] = nga
                if ga and ga not in acc_ga2p:
                    acc_ga2p[ga] = pga
            for rmpm, nle, ple in zip(df_account['_RMPM'].values,
                                       df_account['_NOM_LE'].values,
                                       df_account['_PAYS_LE'].values):
                if rmpm and rmpm not in acc_r2n:
                    acc_r2n[rmpm] = nle
                if rmpm and rmpm not in acc_r2p:
                    acc_r2p[rmpm] = ple

            # BEJO (optionnel)
            bejo_cartes = {}
            bejo_flux = {}
            if use_bejo:
                self._prog(0.17, "Indexation BEJO...")
                df_bc = self.load_csv_smart(self.files["BEJO_CARTES"])
                df_bf = self.load_csv_smart(self.files["BEJO_FLUX"])
                col_bc_ent = m.get('bejo_cartes_entite', df_bc.columns[0])
                col_bc_nb = m.get('bejo_cartes_nb',
                                  df_bc.columns[2] if len(df_bc.columns) > 2 else df_bc.columns[-1])
                col_bf_ent = m.get('bejo_flux_entite', df_bf.columns[0])
                col_bf_iban = m.get('bejo_flux_iban',
                                    df_bf.columns[5] if len(df_bf.columns) > 5 else df_bf.columns[-1])
                df_bc['_K_ENT'] = self.normalize_rs(df_bc[col_bc_ent])
                for _, r in df_bc[df_bc['_K_ENT'] != ''].drop_duplicates('_K_ENT').iterrows():
                    bejo_cartes[r['_K_ENT']] = r[col_bc_nb]
                df_bf['_K_ENT'] = self.normalize_rs(df_bf[col_bf_ent])
                df_bf['_K_IBAN'] = self.clean_iban_full(df_bf[col_bf_iban])
                for _, r in df_bf[(df_bf['_K_ENT'] != '') & (df_bf['_K_IBAN'] != '')].iterrows():
                    bejo_flux.setdefault(r['_K_ENT'], [])
                    if r['_K_IBAN'] not in bejo_flux[r['_K_ENT']]:
                        bejo_flux[r['_K_ENT']].append(r['_K_IBAN'])

            def resolve_row(rp, rc, rs, iban_tr, iban_full):
                rmpm, ga, nom_ga, pays_ga, nom_le, pays_le = '', '', '', '', '', ''
                rp_res, rc_res = rp, rc
                source = "NON_TROUVE"
                segment_val = ''
                seg_source = 'FALLBACK'

                def _s(v):
                    if v is None:
                        return ''
                    s = str(v).strip()
                    if s in ('', 'nan', 'NaN', 'None', 'NULL', 'NA', 'N/A'):
                        return ''
                    return s

                rc_safe = rc
                for rc_candidate in [str(rc).strip(), rc_safe]:
                    if rc_candidate and rc_candidate in d_ref_rc:
                        seg, g, ng, rm = d_ref_rc[rc_candidate]
                        seg, g, ng, rm = _s(seg), _s(g), _s(ng), _s(rm)
                        if seg:
                            segment_val = seg; seg_source = "REF_RC"
                        if g:
                            ga = g
                        if ng:
                            nom_ga = ng
                        if rm:
                            rmpm = rm
                        source = "REF_RC"
                        break

                if source == "NON_TROUVE":
                    ref_rib_hit = None
                    if iban_tr and iban_tr in d_ref_rib:
                        ref_rib_hit = d_ref_rib[iban_tr]
                    elif iban_full and iban_full in d_ref_rib:
                        ref_rib_hit = d_ref_rib[iban_full]
                    if ref_rib_hit:
                        seg, g, ng, rm = ref_rib_hit
                        seg, g, ng, rm = _s(seg), _s(g), _s(ng), _s(rm)
                        if seg:
                            segment_val = seg; seg_source = "REF_IBAN"
                        if g:
                            ga = g
                        if ng:
                            nom_ga = ng
                        if rm:
                            rmpm = rm
                        source = "REF_IBAN"

                if not segment_val:
                    for rc_candidate in [str(rc).strip(), rc_safe]:
                        if rc_candidate and rc_candidate in seg_by_rc:
                            segment_val = seg_by_rc[rc_candidate]; seg_source = "IDSEG_RC"
                            break
                if not segment_val and rmpm:
                    for rm_candidate in [rmpm, rmpm.lstrip('0') or rmpm]:
                        if rm_candidate in seg_by_rmpm:
                            segment_val = seg_by_rmpm[rm_candidate]; seg_source = "IDSEG_RMPM"
                            break
                if not segment_val and rp:
                    rp_stripped = rp.lstrip('0') or rp
                    for rp_candidate in [rp, rp_stripped]:
                        if rp_candidate in seg_by_rp:
                            segment_val = seg_by_rp[rp_candidate]; seg_source = "IDSEG_RP"
                            break

                acc_hit = None
                if iban_full and iban_full in acc_iban_dict:
                    acc_hit = acc_iban_dict[iban_full]
                elif iban_tr and iban_tr in acc_iban_dict:
                    acc_hit = acc_iban_dict[iban_tr]
                if acc_hit:
                    a_ga, a_nga, a_pga, a_rm, a_nle, a_ple = [_s(x) for x in acc_hit]
                    if not ga and a_ga:
                        ga = a_ga
                    if not nom_ga and a_nga:
                        nom_ga = a_nga
                    if not pays_ga and a_pga:
                        pays_ga = a_pga
                    if not rmpm and a_rm:
                        rmpm = a_rm
                    if not nom_le and a_nle:
                        nom_le = a_nle
                    if not pays_le and a_ple:
                        pays_le = a_ple
                    if source == "NON_TROUVE":
                        source = "ACCOUNT_IBAN"

                if not nom_ga and ga and ga in acc_ga2n:
                    nom_ga = _s(acc_ga2n[ga])
                if not pays_ga and ga and ga in acc_ga2p:
                    pays_ga = _s(acc_ga2p[ga])
                if not nom_le and rmpm and rmpm in acc_r2n:
                    nom_le = _s(acc_r2n[rmpm])
                if not pays_le and rmpm and rmpm in acc_r2p:
                    pays_le = _s(acc_r2p[rmpm])

                return (rmpm, ga, nom_ga, pays_ga, nom_le, pays_le,
                        source, rp_res, rc_res, segment_val, seg_source)

            # PRÉPARATION WORLDLINE
            self._prog(0.20, "Préparation WORLDLINE...")
            wl_cols = list(df_wl.columns)
            df_wl['_MOIS'] = self.parse_mois(df_wl[m['wl_mois']])
            df_wl['_NOM_PROG'] = df_wl[m['wl_nom_prog']].astype(str).str.strip()
            df_wl['_RS'] = self.normalize_rs(df_wl[m['wl_rs']])
            df_wl['_IBAN_TR'] = self.clean_iban_truncated(df_wl[m['wl_iban']])
            df_wl['_IBAN_FULL'] = self.clean_iban_full(df_wl[m['wl_iban']])
            df_wl['_DEVISE'] = df_wl[m['wl_devise']].astype(str).str.strip().str.upper()
            df_wl['_PERIO'] = self.to_float(df_wl[m['wl_periodicite']])
            df_wl['_NB_CARTES'] = self.to_float(df_wl[m['wl_nb_cartes']])
            df_wl['_NB_TR_F'] = self.to_float(df_wl[m['wl_nb_tr_fourn']])
            df_wl['_NB_TR_CB'] = self.to_float(df_wl[m['wl_nb_tr_cb']])
            df_wl['_NB_TR_VISA'] = self.to_float(df_wl[m['wl_nb_tr_visa']])
            df_wl['_NB_TRANS'] = df_wl['_NB_TR_F'] + df_wl['_NB_TR_CB'] + df_wl['_NB_TR_VISA']
            df_wl['_RC'] = self.clean_id_strip0(df_wl[m['wl_rc']])
            df_wl['_RC_SAFE'] = self.clean_id_safe(df_wl[m['wl_rc']])
            df_wl['_DIFFERE'] = df_wl[m['wl_differe']].astype(str).str.strip()
            # NETTOYAGE PAYS (anglais)
            df_wl['_PAYS'] = df_wl[self._wl_pays_col].apply(self._clean_pays_value)

            self._prog(0.22, "Conversion devises WL → EUR...")
            arr_mois_wl = df_wl['_MOIS'].values
            arr_devise_wl = df_wl['_DEVISE'].values
            n_wl = len(df_wl)
            cache_t = {}
            arr_taux_wl = np.ones(n_wl, dtype=float)
            for i in range(n_wl):
                key = (arr_mois_wl[i], arr_devise_wl[i])
                if key not in cache_t:
                    cache_t[key] = resolve_taux(arr_mois_wl[i], arr_devise_wl[i]) or 1.0
                arr_taux_wl[i] = cache_t[key]
            df_wl['_TAUX'] = arr_taux_wl
            conv_first = wl_conv_cfg['first'] - 1
            conv_last = wl_conv_cfg['last'] - 1
            for c in [c for c in wl_cols[conv_first:conv_last + 1] if c in df_wl.columns]:
                df_wl[c] = np.round(self.to_float(df_wl[c]).values * arr_taux_wl, 2)
            df_wl['_PLAFOND'] = self.to_float(df_wl[m['wl_plafond']]) * arr_taux_wl
            df_wl['_DEP1'] = self.to_float(df_wl[m['wl_dep_1']])
            df_wl['_DEP2'] = self.to_float(df_wl[m['wl_dep_2']])
            df_wl['_FLUX'] = df_wl['_DEP1'] + df_wl['_DEP2']
            if wl_pnb_cols:
                for c in wl_pnb_cols:
                    df_wl[c] = self.to_float(df_wl[c])
                df_wl['_PNB'] = df_wl[wl_pnb_cols].sum(axis=1)
            else:
                df_wl['_PNB'] = 0.0
            if pays_excl_wl:
                df_wl = df_wl[~df_wl['_PAYS'].isin(pays_excl_wl)].copy()
            nb_lignes_non_eur = int((arr_devise_wl != 'EUR').sum())

            # PRÉPARATION MONEXT
            self._prog(0.25, "Préparation MONEXT...")
            mx_cols = list(df_mx.columns)
            df_mx['_MOIS'] = self.parse_mois(df_mx[m['mx_mois']])
            df_mx['_NOM_PROG'] = df_mx[m['mx_nom_prog']].astype(str).str.strip()
            df_mx['_RS'] = self.normalize_rs(df_mx[m['mx_rs']])
            df_mx['_RP_SAFE'] = self.clean_id_safe(df_mx[m['mx_rp']])
            df_mx['_RP'] = self.clean_id_strip0(df_mx[m['mx_rp']])
            df_mx['_RC_SAFE'] = self.clean_id_safe(df_mx[m['mx_rc']])
            df_mx['_RC'] = self.clean_id_strip0(df_mx[m['mx_rc']])
            df_mx['_IBAN_TR'] = self.clean_iban_truncated(df_mx[m['mx_iban']])
            df_mx['_IBAN_FULL'] = self.clean_iban_full(df_mx[m['mx_iban']])
            df_mx['_NB_CARTES'] = self.to_float(df_mx[m['mx_nb_cartes']])
            df_mx['_DIFFERE'] = df_mx[m['mx_differe']].astype(str).str.strip()
            df_mx['_DEP'] = self.to_float(df_mx[m['mx_depenses']])
            df_mx['_NB_TRANS'] = self.to_float(df_mx[m['mx_nb_transactions']])
            df_mx['_RET'] = self.to_float(df_mx[m['mx_retraits']])
            df_mx['_FLUX'] = df_mx['_DEP'] + df_mx['_RET']
            # NETTOYAGE PAYS (anglais)
            df_mx['_PAYS'] = df_mx[self._mx_pays_col].apply(self._clean_pays_value)

            f0 = mx_pnb_cfg['first'] - 1
            l0 = mx_pnb_cfg['last'] - 1
            e0 = mx_pnb_cfg['excl'] - 1 if mx_pnb_cfg['excl'] > 0 else -1
            ic0 = mx_pnb_cfg['interchange'] - 1 if mx_pnb_cfg['interchange'] > 0 else -1
            pnb_range = mx_cols[f0:l0 + 1]
            excl_name = mx_cols[e0] if 0 <= e0 < len(mx_cols) else None
            ic_name = mx_cols[ic0] if 0 <= ic0 < len(mx_cols) else None
            for c in pnb_range:
                if c in df_mx.columns:
                    df_mx[c] = self.to_float(df_mx[c])
            if ic_name and ic_name in df_mx.columns:
                df_mx[ic_name] = df_mx[ic_name] * -1
            pnb_filtered = [c for c in pnb_range if c != excl_name]
            df_mx['_PNB'] = df_mx[pnb_filtered].sum(axis=1) if pnb_filtered else 0.0
            if pays_excl_mx:
                df_mx = df_mx[~df_mx['_PAYS'].isin(pays_excl_mx)].copy()

            # MATCHING WL
            self._prog(0.35, "Matching WORLDLINE...")
            n_wl = len(df_wl)
            wl_rmpm = np.empty(n_wl, dtype=object); wl_ga = np.empty(n_wl, dtype=object)
            wl_nom_ga = np.empty(n_wl, dtype=object); wl_pays_ga = np.empty(n_wl, dtype=object)
            wl_nom_le = np.empty(n_wl, dtype=object); wl_pays_le = np.empty(n_wl, dtype=object)
            wl_source = np.empty(n_wl, dtype=object); wl_rp_res = np.empty(n_wl, dtype=object)
            wl_rc_res = np.empty(n_wl, dtype=object); wl_seg_raw = np.empty(n_wl, dtype=object)
            wl_seg_src = np.empty(n_wl, dtype=object)
            arr_wl_rs = df_wl['_RS'].values; arr_wl_rc = df_wl['_RC'].values
            arr_wl_iban_tr = df_wl['_IBAN_TR'].values; arr_wl_iban_full = df_wl['_IBAN_FULL'].values
            for i in range(n_wl):
                if i % 5000 == 0:
                    self._prog(0.35 + 0.08 * i / max(n_wl, 1), f"WL {i:,}/{n_wl:,}")
                res = resolve_row('', arr_wl_rc[i], arr_wl_rs[i],
                                  arr_wl_iban_tr[i], arr_wl_iban_full[i])
                wl_rmpm[i], wl_ga[i], wl_nom_ga[i], wl_pays_ga[i] = res[0], res[1], res[2], res[3]
                wl_nom_le[i], wl_pays_le[i] = res[4], res[5]
                wl_source[i], wl_rp_res[i], wl_rc_res[i] = res[6], res[7], res[8]
                wl_seg_raw[i], wl_seg_src[i] = res[9], res[10]
            df_wl['_RMPM_R'] = wl_rmpm; df_wl['_GA_R'] = wl_ga
            df_wl['_NOM_GA'] = wl_nom_ga; df_wl['_PAYS_GA'] = wl_pays_ga
            df_wl['_NOM_LE'] = wl_nom_le; df_wl['_PAYS_LE'] = wl_pays_le
            df_wl['_SOURCE'] = wl_source; df_wl['_RP_R'] = wl_rp_res
            df_wl['_RC_R'] = wl_rc_res; df_wl['_SEG_RAW'] = wl_seg_raw
            df_wl['_SEG_SRC'] = wl_seg_src

            if override_pays and (ovr_rmpm_set or ovr_idprog_set):
                wl_rmpm_vals = df_wl['_RMPM_R'].values
                wl_rc_vals = df_wl['_RC_R'].values
                wl_nom_prog_vals = df_wl['_NOM_PROG'].values
                n_ovr_wl = 0
                for i in range(len(df_wl)):
                    hit = False
                    rmpm_v = str(wl_rmpm_vals[i]).strip()
                    if ovr_rmpm_set and rmpm_v and rmpm_v in ovr_rmpm_set:
                        hit = True
                    if not hit:
                        rc_v = str(wl_rc_vals[i]).strip()
                        if ovr_rmpm_set and rc_v and rc_v in ovr_rmpm_set:
                            hit = True
                    if not hit:
                        nom = str(wl_nom_prog_vals[i]).strip()
                        if ovr_idprog_set and nom and nom in ovr_idprog_set:
                            hit = True
                    if hit:
                        df_wl.iat[i, df_wl.columns.get_loc('_PAYS')] = override_pays
                        n_ovr_wl += 1
                self._prog(0.435, f"Override WL: {n_ovr_wl:,} lignes")

            # MATCHING MX
            self._prog(0.43, "Matching MONEXT...")
            n_mx = len(df_mx)
            mx_rmpm = np.empty(n_mx, dtype=object); mx_ga = np.empty(n_mx, dtype=object)
            mx_nom_ga = np.empty(n_mx, dtype=object); mx_pays_ga = np.empty(n_mx, dtype=object)
            mx_nom_le = np.empty(n_mx, dtype=object); mx_pays_le = np.empty(n_mx, dtype=object)
            mx_source = np.empty(n_mx, dtype=object); mx_rp_res = np.empty(n_mx, dtype=object)
            mx_rc_res = np.empty(n_mx, dtype=object); mx_seg_raw = np.empty(n_mx, dtype=object)
            mx_seg_src = np.empty(n_mx, dtype=object)
            arr_mx_rp = df_mx['_RP'].values; arr_mx_rc = df_mx['_RC'].values
            arr_mx_rs = df_mx['_RS'].values
            arr_mx_iban_tr = df_mx['_IBAN_TR'].values; arr_mx_iban_full = df_mx['_IBAN_FULL'].values
            for i in range(n_mx):
                if i % 5000 == 0:
                    self._prog(0.43 + 0.08 * i / max(n_mx, 1), f"MX {i:,}/{n_mx:,}")
                res = resolve_row(arr_mx_rp[i], arr_mx_rc[i], arr_mx_rs[i],
                                  arr_mx_iban_tr[i], arr_mx_iban_full[i])
                mx_rmpm[i], mx_ga[i], mx_nom_ga[i], mx_pays_ga[i] = res[0], res[1], res[2], res[3]
                mx_nom_le[i], mx_pays_le[i] = res[4], res[5]
                mx_source[i], mx_rp_res[i], mx_rc_res[i] = res[6], res[7], res[8]
                mx_seg_raw[i], mx_seg_src[i] = res[9], res[10]
            df_mx['_RMPM_R'] = mx_rmpm; df_mx['_GA_R'] = mx_ga
            df_mx['_NOM_GA'] = mx_nom_ga; df_mx['_PAYS_GA'] = mx_pays_ga
            df_mx['_NOM_LE'] = mx_nom_le; df_mx['_PAYS_LE'] = mx_pays_le
            df_mx['_SOURCE'] = mx_source; df_mx['_RP_R'] = mx_rp_res
            df_mx['_RC_R'] = mx_rc_res; df_mx['_SEG_RAW'] = mx_seg_raw
            df_mx['_SEG_SRC'] = mx_seg_src

            if override_pays and ovr_rmpm_set:
                mx_rmpm_vals = df_mx['_RMPM_R'].values
                mx_rc_vals = df_mx['_RC_R'].values
                mx_rp_vals = df_mx['_RP_SAFE'].values
                n_ovr_mx = 0
                for i in range(len(df_mx)):
                    rmpm_v = str(mx_rmpm_vals[i]).strip()
                    rc_v = str(mx_rc_vals[i]).strip()
                    rp_v = str(mx_rp_vals[i]).strip()
                    if ((rmpm_v and rmpm_v in ovr_rmpm_set)
                            or (rc_v and rc_v in ovr_rmpm_set)
                            or (rp_v and rp_v in ovr_rmpm_set)):
                        df_mx.iat[i, df_mx.columns.get_loc('_PAYS')] = override_pays
                        n_ovr_mx += 1
                self._prog(0.515, f"Override MX: {n_ovr_mx:,} lignes")

            n_bejo_ok = 0
            if use_bejo and bejo_cartes and bejo_flux:
                self._prog(0.52, "BEJO override MX...")
                arr_ent_src = self.normalize_rs(df_mx['_RS']).values
                df_mx_iban_agg = df_mx[df_mx['_IBAN_FULL'] != ''].groupby('_IBAN_FULL').agg(
                    _FLUX=('_FLUX', 'sum'), _PNB=('_PNB', 'sum'),
                    _NB=('_NB_CARTES', 'sum')
                ).to_dict('index')
                for i in range(len(df_mx)):
                    ent_norm = arr_ent_src[i]
                    if not ent_norm or ent_norm not in bejo_cartes:
                        continue
                    ibans = bejo_flux.get(ent_norm, [])
                    if not ibans:
                        continue
                    ibans_ok = [ib for ib in ibans if ib in df_mx_iban_agg]
                    if not ibans_ok:
                        continue
                    n_bejo_ok += 1
                    nb_ref = self.to_float(pd.Series([str(bejo_cartes[ent_norm])])).iloc[0]
                    df_mx.iat[i, df_mx.columns.get_loc('_NB_CARTES')] = nb_ref
                if n_bejo_ok:
                    self._prog(0.525, f"BEJO: {n_bejo_ok} entités matchées")

            # AGRÉGATS + CLIENT_KEY
            self._prog(0.55, "Agrégats mensuels...")
            mois_wl = sorted([x for x in df_wl['_MOIS'].unique() if x and len(x) == 6])
            mois_mx = sorted([x for x in df_mx['_MOIS'].unique() if x and len(x) == 6])
            all_mois = sorted(set(mois_wl) | set(mois_mx))
            plage_debut = plage_cfg.get('debut', '')
            plage_fin = plage_cfg.get('fin', '')
            mois_plage = [mo for mo in all_mois if plage_debut <= mo <= plage_fin] \
                if plage_debut and plage_fin else list(all_mois)

            mx_global_by_mois = {}
            for mo in mois_mx:
                d = df_mx[df_mx['_MOIS'] == mo]
                mx_global_by_mois[mo] = {'flux': d['_FLUX'].sum(),
                                         'pnb': d['_PNB'].sum(), 'nb': d['_NB_CARTES'].sum()}
            wl_global_by_mois = {}
            for mo in mois_wl:
                d = df_wl[df_wl['_MOIS'] == mo]
                wl_global_by_mois[mo] = {'flux': d['_FLUX'].sum(),
                                         'pnb': d['_PNB'].sum(), 'nb': d['_NB_CARTES'].sum()}

            self._prog(0.57, "Construction clients...")

            def bck(rmpm, rc, rp, rs, tag, idx):
                if rmpm:
                    return f"RMPM|{rmpm}"
                if rc:
                    return f"RC|{rc}"
                if rp:
                    return f"RP|{rp}"
                if rs:
                    return f"RS|{rs}"
                return f"ORPH|{tag}|{idx}"

            df_mx['_CLIENT_KEY'] = [bck(rm, rc, rp, rs, 'MX', i)
                                    for i, (rm, rc, rp, rs) in enumerate(
                                        zip(df_mx['_RMPM_R'].values, df_mx['_RC_R'].values,
                                            df_mx['_RP_R'].values, df_mx['_RS'].values))]
            df_wl['_CLIENT_KEY'] = [bck(rm, rc, rp, rs, 'WL', i)
                                    for i, (rm, rc, rp, rs) in enumerate(
                                        zip(df_wl['_RMPM_R'].values, df_wl['_RC_R'].values,
                                            df_wl['_RP_R'].values, df_wl['_RS'].values))]

            self._prog(0.60, "Agrégats par client...")
            mx_mo_agg = {mo: {} for mo in all_mois}
            for mo in mois_mx:
                d = df_mx[df_mx['_MOIS'] == mo]
                mx_mo_agg[mo] = d.groupby('_CLIENT_KEY').agg({
                    '_FLUX': 'sum', '_PNB': 'sum',
                    '_NB_CARTES': 'sum', '_NB_TRANS': 'sum'}).T.to_dict('list')
            wl_mo_agg = {mo: {} for mo in all_mois}
            for mo in mois_wl:
                d = df_wl[df_wl['_MOIS'] == mo]
                wl_mo_agg[mo] = d.groupby('_CLIENT_KEY').agg({
                    '_FLUX': 'sum', '_PNB': 'sum',
                    '_NB_CARTES': 'sum', '_NB_TRANS': 'sum'}).T.to_dict('list')

            self._df_wl_processed = df_wl
            self._df_mx_processed = df_mx
            self._mx_mo_agg = mx_mo_agg
            self._wl_mo_agg = wl_mo_agg
            self._all_mois = all_mois
            self._mois_mx = mois_mx
            self._mois_wl = mois_wl
            self._mois_plage = mois_plage
            self._mx_global_by_mois = mx_global_by_mois
            self._wl_global_by_mois = wl_global_by_mois
            self._rebate_by_year_rmpm = rebate_by_year_rmpm
            self._n_bejo_ok = n_bejo_ok
            self._nb_lignes_non_eur = nb_lignes_non_eur

            self._build_dataframe_and_save(ts, constants)

        except Exception:
            self._prog(0, "Erreur")
            import traceback
            traceback.print_exc()
            raise

    @staticmethod
    def _build_idseg_dicts(df_seg, col_type, col_id, col_segment):
        seg_by_rc, seg_by_rmpm, seg_by_rp, seg_by_ca = {}, {}, {}, {}
        TYPE_MAP = {'RC': seg_by_rc, 'RMPM': seg_by_rmpm,
                    'RP': seg_by_rp, 'CODE_AGENCE': seg_by_ca, 'CODE GA': seg_by_ca}
        for t, raw_id, seg_val in zip(
            df_seg[col_type].astype(str).str.strip().str.upper().values,
            df_seg[col_id].astype(str).str.strip().values,
            df_seg[col_segment].astype(str).str.strip().str.upper().values
        ):
            if 'ENTREPRISE' in seg_val:
                segment = 'ENTREPRISE'
            elif 'BPE' in seg_val:
                segment = 'BPE'
            else:
                continue
            td = TYPE_MAP.get(t)
            if td is None:
                continue
            clean = raw_id
            if clean.startswith('="') and clean.endswith('"'):
                clean = clean[2:-1]
            clean = clean.lstrip("'").strip()
            if clean.endswith('.0') and clean[:-2].isdigit():
                clean = clean[:-2]
            if not clean:
                continue
            if clean not in td:
                td[clean] = segment
            stripped = clean.lstrip('0')
            if stripped and stripped != clean and stripped not in td:
                td[stripped] = segment
        return seg_by_rc, seg_by_rmpm, seg_by_rp, seg_by_ca

    # ═════════════════════════════════════════════════════════════════════════
    # CONSTRUCTION DU DATAFRAME FINAL + SAUVEGARDE (fenêtre unique)
    # ═════════════════════════════════════════════════════════════════════════
    def _build_dataframe_and_save(self, ts, constants):
        df_mx = self._df_mx_processed
        df_wl = self._df_wl_processed
        mx_mo_agg = self._mx_mo_agg
        wl_mo_agg = self._wl_mo_agg
        all_mois = self._all_mois
        mois_plage = self._mois_plage

        def fne(series):
            for v in series:
                if v and str(v).strip():
                    return v
            return ''

        self._prog(0.65, "Consolidation infos par client...")
        mx_info = {}
        for key, grp in df_mx.groupby('_CLIENT_KEY'):
            mx_info[key] = {
                'rmpm': fne(grp['_RMPM_R'].values), 'ga': fne(grp['_GA_R'].values),
                'nom_ga': fne(grp['_NOM_GA'].values), 'pays_ga': fne(grp['_PAYS_GA'].values),
                'nom_le': fne(grp['_NOM_LE'].values), 'pays_le': fne(grp['_PAYS_LE'].values),
                'source': fne(grp['_SOURCE'].values), 'rp': fne(grp['_RP_SAFE'].values),
                'rc': fne(grp['_RC_SAFE'].values), 'rs': fne(grp['_RS'].values),
                'nom_prog': fne(grp['_NOM_PROG'].values), 'pays': fne(grp['_PAYS'].values),
                'differe': fne(grp['_DIFFERE'].values), 'seg_raw': fne(grp['_SEG_RAW'].values),
                'seg_src': fne(grp['_SEG_SRC'].values),
            }
        wl_info = {}
        for key, grp in df_wl.groupby('_CLIENT_KEY'):
            wl_info[key] = {
                'rmpm': fne(grp['_RMPM_R'].values), 'ga': fne(grp['_GA_R'].values),
                'nom_ga': fne(grp['_NOM_GA'].values), 'pays_ga': fne(grp['_PAYS_GA'].values),
                'nom_le': fne(grp['_NOM_LE'].values), 'pays_le': fne(grp['_PAYS_LE'].values),
                'source': fne(grp['_SOURCE'].values), 'rc': fne(grp['_RC_SAFE'].values),
                'rs': fne(grp['_RS'].values), 'nom_prog': fne(grp['_NOM_PROG'].values),
                'pays': fne(grp['_PAYS'].values), 'differe': fne(grp['_DIFFERE'].values),
                'plafond': float(grp['_PLAFOND'].max()) if len(grp) > 0 else 0.0,
                'perio': float(grp['_PERIO'].iloc[0]) if len(grp) > 0 else 1.0,
                'seg_raw': fne(grp['_SEG_RAW'].values), 'seg_src': fne(grp['_SEG_SRC'].values),
            }

        all_client_keys = sorted(set(mx_info.keys()) | set(wl_info.keys()))
        n_clients = len(all_client_keys)
        self._prog(0.70, f"Construction DATA pour {n_clients:,} clients...")

        rows_data = []
        for i, key in enumerate(all_client_keys):
            if i % 2000 == 0:
                self._prog(0.70 + 0.06 * i / max(n_clients, 1), f"DATA {i:,}/{n_clients:,}")
            mx = mx_info.get(key, {})
            wl = wl_info.get(key, {})
            info_main = mx if mx.get('source', '').startswith(('REF_', 'ACCOUNT')) else (
                wl if wl.get('source', '').startswith(('REF_', 'ACCOUNT')) else (mx or wl))
            seg_raw = info_main.get('seg_raw', '') or mx.get('seg_raw', '') or wl.get('seg_raw', '')
            seg_src = info_main.get('seg_src', '') or mx.get('seg_src', '') or wl.get('seg_src', 'FALLBACK')
            is_bpe = (seg_raw == 'BPE')
            rmpm_client = info_main.get('rmpm', '') or mx.get('rmpm', '') or wl.get('rmpm', '')

            row = {
                'CLIENT_KEY': key, 'RMPM': rmpm_client,
                'ID_RP': self.protect_id(mx.get('rp', '')),
                'ID_RC_MX': self.protect_id(mx.get('rc', '')),
                'ID_RC_WL': self.protect_id(wl.get('rc', '')),
                'CODE_GA': info_main.get('ga', ''),
                'NOM_GA': info_main.get('nom_ga', '') or mx.get('nom_ga', '') or wl.get('nom_ga', ''),
                'PAYS_GA': info_main.get('pays_ga', '') or mx.get('pays_ga', '') or wl.get('pays_ga', ''),
                'NOM_ENTITE': info_main.get('nom_le', '') or mx.get('nom_le', '') or wl.get('nom_le', ''),
                'PAYS_ENTITE': info_main.get('pays_le', '') or mx.get('pays_le', '') or wl.get('pays_le', ''),
                'NOM_PROG_CCO': mx.get('nom_prog', ''), 'NOM_PROG_CPC': wl.get('nom_prog', ''),
                'RS_CCO': mx.get('rs', ''), 'RS_CPC': wl.get('rs', ''),
                'SOURCE_MATCHING': info_main.get('source', 'NON_TROUVE'),
                'PRESENT_CCO': 'YES' if key in mx_info else 'NO',
                'PRESENT_CPC': 'YES' if key in wl_info else 'NO',
                'IS_BPE': 'YES' if is_bpe else 'NO', 'IS_ENT': 'NO' if is_bpe else 'YES',
                'SEGMENT': 'BPE' if is_bpe else 'ENTREPRISE', 'SEGMENT_SOURCE': seg_src,
                'PAYS_APPORTEUR_CCO': mx.get('pays', ''), 'PAYS_APPORTEUR_CPC': wl.get('pays', ''),
                'PAYS_FINAL': mx.get('pays', '') or wl.get('pays', ''),
                'DIFFERE_CCO_TYPE': self._differe_type(mx.get('differe', ''), 'CCO'),
                'DIFFERE_CCO_JOURS': self._differe_jours(mx.get('differe', '')),
                'DIFFERE_CPC_TYPE': self._differe_type(wl.get('differe', ''), 'CPC'),
                'DIFFERE_CPC_JOURS': self._differe_jours(wl.get('differe', '')),
                'PLAFOND_CPC_EUR': wl.get('plafond', 0.0),
                'PERIODICITE_CPC': wl.get('perio', 1.0),
                'REBATE_CCO_MONTANT': 0.0, 'REBATE_CPC_MONTANT': 0.0,
            }

            mx_flux_p = sum(mx_mo_agg[mo].get(key, [0, 0, 0, 0])[0] for mo in mois_plage if mo in mx_mo_agg)
            mx_pnb_p = sum(mx_mo_agg[mo].get(key, [0, 0, 0, 0])[1] for mo in mois_plage if mo in mx_mo_agg)
            mx_nb_p = sum(mx_mo_agg[mo].get(key, [0, 0, 0, 0])[2] for mo in mois_plage if mo in mx_mo_agg)
            mx_tr_p = sum(mx_mo_agg[mo].get(key, [0, 0, 0, 0])[3] for mo in mois_plage if mo in mx_mo_agg)
            wl_flux_p = sum(wl_mo_agg[mo].get(key, [0, 0, 0, 0])[0] for mo in mois_plage if mo in wl_mo_agg)
            wl_pnb_p = sum(wl_mo_agg[mo].get(key, [0, 0, 0, 0])[1] for mo in mois_plage if mo in wl_mo_agg)
            wl_nb_p = sum(wl_mo_agg[mo].get(key, [0, 0, 0, 0])[2] for mo in mois_plage if mo in wl_mo_agg)
            wl_tr_p = sum(wl_mo_agg[mo].get(key, [0, 0, 0, 0])[3] for mo in mois_plage if mo in wl_mo_agg)
            row.update({
                'FLUX_CCO_PLAGE': mx_flux_p, 'PNB_CCO_PLAGE': mx_pnb_p,
                'NB_CARTES_CCO_PLAGE': mx_nb_p, 'NB_TRANS_CCO_PLAGE': mx_tr_p,
                'FLUX_CPC_PLAGE': wl_flux_p, 'PNB_CPC_PLAGE': wl_pnb_p,
                'NB_CARTES_CPC_PLAGE': wl_nb_p, 'NB_TRANS_CPC_PLAGE': wl_tr_p,
            })
            row['TAG_SANS_FLUX_MX'] = 'OUI' if (mx_flux_p == 0 and mx_pnb_p == 0 and mx_nb_p == 0) else 'NON'
            row['TAG_SANS_FLUX_WL'] = 'OUI' if (wl_flux_p == 0 and wl_pnb_p == 0 and wl_nb_p == 0) else 'NON'

            for mo in all_mois:
                lbl = self.mois_label(mo)
                mx_v = mx_mo_agg[mo].get(key, [0, 0, 0, 0]) if mo in mx_mo_agg else [0, 0, 0, 0]
                wl_v = wl_mo_agg[mo].get(key, [0, 0, 0, 0]) if mo in wl_mo_agg else [0, 0, 0, 0]
                row[f'NB_CARTES_CCO_{lbl}'] = mx_v[2]
                row[f'FLUX_CCO_{lbl}'] = mx_v[0]
                row[f'PNB_CCO_{lbl}'] = mx_v[1]
                row[f'NB_CARTES_CPC_{lbl}'] = wl_v[2]
                row[f'FLUX_CPC_{lbl}'] = wl_v[0]
                row[f'PNB_CPC_{lbl}'] = wl_v[1]

            last_3 = mois_plage[-3:] if len(mois_plage) >= 3 else mois_plage
            row['FLUX_CCO_TRIMESTRE'] = sum(mx_mo_agg[mo].get(key, [0, 0, 0, 0])[0]
                                             for mo in last_3 if mo in mx_mo_agg)
            row['FLUX_CPC_TRIMESTRE'] = sum(wl_mo_agg[mo].get(key, [0, 0, 0, 0])[0]
                                             for mo in last_3 if mo in wl_mo_agg)
            row['TRIMESTRE_LIBELLE'] = (f"{self.mois_label(last_3[0])} -> {self.mois_label(last_3[-1])}"
                                        if last_3 else "")
            rows_data.append(row)

        df_data = pd.DataFrame(rows_data)
        self._n_clients = n_clients

        # ── SAUVEGARDE — base commune CSV + XLSX (output-dir / output-filename) ──
        self._prog(0.80, "Préparation de l'emplacement de sauvegarde...")
        out_dir = Path(self.output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        base = str(out_dir / self.output_filename)
        # Nettoyage extension éventuelle
        for ext in ('.xlsx', '.csv', '.xls'):
            if base.lower().endswith(ext):
                base = base[:-len(ext)]
        csv_path = base + ".csv"
        xlsx_path = base + ".xlsx"

        self._prog(0.81, "Export CSV...")
        try:
            df_data.to_csv(csv_path, sep=';', index=False, encoding='utf-8-sig')
        except Exception as e:
            print(f"[CSV] Erreur écriture : {e}")

        if not OPENPYXL_OK:
            self._prog(1.0, "Terminé (CSV uniquement).")
            return

        self.generate_xlsx(df_data, ts, constants, xlsx_path)

    # ═════════════════════════════════════════════════════════════════════════
    # CALCUL FINANCIER NATIF — réimplémentation Python de toutes les formules
    # ─────────────────────────────────────────────────────────────────────────
    # Écrit dans df_year, en NOMBRES NATIFS, toutes les colonnes de coûts,
    # RWA, EAD, commissions, applicabilité et totaux. Aucune formule Excel.
    # Hypothèses figées : bonus impl/signing = 0, RW = rw_defaut, type = Entreprise.
    # Refinancement : 0,99 % (2025), 1,66 % (2026), sinon constante de repli.
    # ═════════════════════════════════════════════════════════════════════════
    def _compute_financials(self, df_year, year, constants):
        if df_year is None or len(df_year) == 0:
            return df_year

        n = len(df_year)
        # Constantes
        if str(year) == '2025':
            refin = constants.get('taux_refinancement', 0.0099)
        elif str(year) == '2026':
            refin = constants.get('taux_refinancement_2026', 0.0166)
        else:
            refin = constants.get('taux_refinancement', 0.0099)
        PART_CAP = constants.get('part_capital_banque', 0.12)
        ALMT = constants.get('almt', 0.142)
        RW = constants.get('rw_defaut', 0.65)
        COUT_CARTE = constants.get('cout_carte_cco', 40.14)
        COUT_TRANS = constants.get('cout_transaction_cpc', 0.59)
        COUT_RWA_CCO = constants.get('cout_rwa_plafond_cco', 13.32)
        TAUX_ENT = constants.get('taux_ead_entreprise', 0.40)
        TAUX_COMM = constants.get('taux_commission', 0.20)
        TVA = constants.get('taux_tva', 0.20)

        def col(name, default=0.0):
            if name in df_year.columns:
                return pd.to_numeric(df_year[name], errors='coerce').fillna(0.0).values.astype(float)
            return np.full(n, default, dtype=float)

        def scol(name):
            if name in df_year.columns:
                return df_year[name].astype(str).str.strip().str.upper().values
            return np.array([''] * n, dtype=object)

        # ── CCO ──────────────────────────────────────────────────────────────
        typ_cco = scol('DIFFERE_CCO_TYPE')
        jours_cco = col('DIFFERE_CCO_JOURS')
        flux_cco = col('FLUX_CCO_ANNEE')
        fluxtri_cco = col('FLUX_CCO_TRIMESTRE_FIN')
        pnb_cco = col('PNB_CCO_ANNEE')
        nbmoy = col('NB_CARTE_MOYEN')
        reb_cco = col('REBATE_CCO_MONTANT')

        imm_cco = (typ_cco == 'IMM')
        cout_diff_cco = np.where(imm_cco | (flux_cco == 0), 0.0,
                                 flux_cco * (jours_cco + 15) / 365.0 * refin)
        cout_run_cco = nbmoy * COUT_CARTE
        rwa_plaf_cco = np.where(imm_cco, 0.0, nbmoy * COUT_RWA_CCO)
        ead_bilan_cco = np.where(imm_cco | (fluxtri_cco == 0), 0.0,
                                 np.where(jours_cco > 27, fluxtri_cco, fluxtri_cco * 2.0 / 3.0))
        rwa_bilan_cco = ead_bilan_cco * RW * PART_CAP * ALMT
        total_couts_cco = cout_diff_cco + cout_run_cco + rwa_plaf_cco + rwa_bilan_cco
        comm_cco_ttc = np.maximum(0.0, (pnb_cco - reb_cco) * TAUX_COMM)
        comm_cco_ht = comm_cco_ttc / (1.0 + TVA)
        res_cco = pnb_cco - reb_cco - total_couts_cco - comm_cco_ttc
        applic_cco = np.where(res_cco > 0, 'OUI', 'NON')

        # ── CPC ──────────────────────────────────────────────────────────────
        typ_cpc = scol('DIFFERE_CPC_TYPE')
        jours_cpc = col('DIFFERE_CPC_JOURS')
        flux_cpc = col('FLUX_CPC_ANNEE')
        fluxtri_cpc = col('FLUX_CPC_TRIMESTRE_FIN')
        pnb_cpc = col('PNB_CPC_ANNEE')
        nbtrans_cpc = col('NB_TRANS_CPC_ANNEE')
        plaf_cpc = col('PLAFOND_CPC_EUR')
        perio_cpc = col('PERIODICITE_CPC')
        reb_cpc = col('REBATE_CPC_MONTANT')

        imm_cpc = (typ_cpc == 'IMM')
        cout_diff_cpc = np.where(imm_cpc | (jours_cpc == 0) | (flux_cpc == 0), 0.0,
                                 flux_cpc * (jours_cpc + 15) / 365.0 * refin)
        cout_run_cpc = nbtrans_cpc * COUT_TRANS
        mult_perio = np.where((perio_cpc == 1) | (perio_cpc == 12), 1.0,
                              np.where((perio_cpc == 3) | (perio_cpc == 6), 2.0, 1.0))
        ead_plaf_cpc = plaf_cpc * mult_perio
        rwa_plaf_cpc = ead_plaf_cpc * TAUX_ENT * RW * PART_CAP * ALMT
        ead_bilan_cpc = np.where(imm_cpc | (jours_cpc == 0) | (fluxtri_cpc == 0), 0.0,
                                 np.where(jours_cpc > 27, fluxtri_cpc, fluxtri_cpc * 2.0 / 3.0))
        rwa_bilan_cpc = ead_bilan_cpc * RW * PART_CAP * ALMT
        total_couts_cpc = cout_diff_cpc + cout_run_cpc + rwa_plaf_cpc + rwa_bilan_cpc
        comm_cpc_ttc = np.maximum(0.0, (pnb_cpc - reb_cpc) * TAUX_COMM)
        comm_cpc_ht = comm_cpc_ttc / (1.0 + TVA)
        res_cpc = pnb_cpc - reb_cpc - total_couts_cpc - comm_cpc_ttc
        applic_cpc = np.where(res_cpc > 0, 'OUI', 'NON')

        # ── Totaux ─────────────────────────────────────────────────────────
        comm_tot_ttc = (np.where(applic_cco == 'OUI', comm_cco_ttc, 0.0)
                        + np.where(applic_cpc == 'OUI', comm_cpc_ttc, 0.0))
        comm_tot_ht = comm_tot_ttc / (1.0 + TVA)
        couts_tot = total_couts_cco + total_couts_cpc
        pnb_tot = pnb_cco + pnb_cpc

        def put(name, arr, rnd=2):
            if isinstance(arr, np.ndarray) and arr.dtype.kind in 'fi':
                df_year[name] = np.round(arr.astype(float), rnd)
            else:
                df_year[name] = arr

        # Hypothèses tracées
        df_year['TYPE_DEBIT_CCO'] = 'Entreprise'
        put('RW_RETENU', np.full(n, RW), 4)
        put('TAUX_EAD_CCO', np.full(n, TAUX_ENT), 4)
        put('IMPL_BONUS_CCO', np.zeros(n))
        put('SIGNING_BONUS_CCO', np.zeros(n))
        put('IMPL_BONUS_CPC', np.zeros(n))
        put('SIGNING_BONUS_CPC', np.zeros(n))
        # CCO
        put('COUT_DIFFERE_CCO', cout_diff_cco)
        put('COUT_RUN_CCO', cout_run_cco)
        put('RWA_PLAFOND_CCO', rwa_plaf_cco)
        put('EAD_BILAN_CCO', ead_bilan_cco)
        put('RWA_BILAN_CCO', rwa_bilan_cco)
        put('TOTAL_COUTS_CCO', total_couts_cco)
        put('COMMISSION_CCO_TTC', comm_cco_ttc)
        put('COMMISSION_CCO_HT', comm_cco_ht)
        put('RESULTAT_COND_CCO', res_cco)
        df_year['APPLICABILITE_CCO'] = applic_cco
        # CPC
        put('COUT_DIFFERE_CPC', cout_diff_cpc)
        put('COUT_RUN_CPC', cout_run_cpc)
        put('MULT_PERIODICITE', mult_perio, 0)
        put('EAD_PLAFOND_CPC', ead_plaf_cpc)
        put('RWA_PLAFOND_CPC', rwa_plaf_cpc)
        put('EAD_BILAN_CPC', ead_bilan_cpc)
        put('RWA_BILAN_CPC', rwa_bilan_cpc)
        put('TOTAL_COUTS_CPC', total_couts_cpc)
        put('COMMISSION_CPC_TTC', comm_cpc_ttc)
        put('COMMISSION_CPC_HT', comm_cpc_ht)
        put('RESULTAT_COND_CPC', res_cpc)
        df_year['APPLICABILITE_CPC'] = applic_cpc
        # Totaux
        put('COMMISSION_TOTALE_TTC', comm_tot_ttc)
        put('COMMISSION_TOTAL_HT', comm_tot_ht)
        put('COUTS_TOTAUX_CLIENT', couts_tot)
        put('PNB_TOTAL_CLIENT', pnb_tot)
        return df_year

    # ═════════════════════════════════════════════════════════════════════════
    # GÉNÉRATION XLSX
    # ═════════════════════════════════════════════════════════════════════════
    COLOR_HDR = {'B': BLU, 'G': GRN, 'P': PUR, 'O': ORA, 'D': DARK}
    COLOR_CELL = {'B': BLU2, 'G': GRN2, 'P': PUR2, 'O': ORA2, 'D': 'E0E0E0'}

    def generate_xlsx(self, df_data, ts, constants, xlsx_path):
        self._prog(0.82, "Génération XLSX...")
        try:
            wb = Workbook()
            wb.remove(wb.active)

            all_mois = self._all_mois
            df_wl = self._df_wl_processed
            df_mx = self._df_mx_processed
            mx_mo_agg = self._mx_mo_agg
            wl_mo_agg = self._wl_mo_agg

            years = sorted(set(mo[:4] for mo in all_mois))
            years_final = [y for y in ['2025', '2026'] if y in years] + \
                          [y for y in years if y not in ('2025', '2026')]
            if not years_final:
                years_final = years

            df_year_by_year = {}
            for year in years_final:
                self._prog(0.84, f"Préparation année {year}...")
                yr_mois = [mo for mo in all_mois if mo.startswith(year)]
                if not yr_mois:
                    continue
                yr_flux_cols = []
                for mo in yr_mois:
                    lbl = self.mois_label(mo)
                    yr_flux_cols += [f'FLUX_CCO_{lbl}', f'FLUX_CPC_{lbl}',
                                     f'PNB_CCO_{lbl}', f'PNB_CPC_{lbl}',
                                     f'NB_CARTES_CCO_{lbl}', f'NB_CARTES_CPC_{lbl}']
                yr_flux_cols = [c for c in yr_flux_cols if c in df_data.columns]
                if yr_flux_cols:
                    mask = pd.Series(False, index=df_data.index)
                    for c in yr_flux_cols:
                        mask = mask | (df_data[c].astype(float).abs() > 0)
                    df_year = df_data[mask].copy()
                else:
                    df_year = df_data.copy()

                wl_yr = df_wl[df_wl['_MOIS'].str.startswith(year)].copy() if len(df_wl) > 0 else df_wl
                mx_yr = df_mx[df_mx['_MOIS'].str.startswith(year)].copy() if len(df_mx) > 0 else df_mx

                wl_yr_info = {}
                if len(wl_yr) > 0 and '_CLIENT_KEY' in wl_yr.columns:
                    for key_wl, grp_wl in wl_yr.groupby('_CLIENT_KEY'):
                        wl_yr_info[key_wl] = {
                            'plafond': float(grp_wl['_PLAFOND'].max()) if len(grp_wl) > 0 else 0.0,
                            'perio': float(grp_wl['_PERIO'].iloc[0]) if len(grp_wl) > 0 else 1.0,
                            'differe': str(grp_wl['_DIFFERE'].iloc[0]).strip() if len(grp_wl) > 0 else '',
                        }
                mx_yr_info = {}
                if len(mx_yr) > 0 and '_CLIENT_KEY' in mx_yr.columns:
                    for key_mx, grp_mx in mx_yr.groupby('_CLIENT_KEY'):
                        mx_yr_info[key_mx] = {
                            'differe': str(grp_mx['_DIFFERE'].iloc[0]).strip() if len(grp_mx) > 0 else '',
                        }

                yr_last3 = yr_mois[-3:] if len(yr_mois) >= 3 else yr_mois
                for idx in df_year.index:
                    key = df_year.at[idx, 'CLIENT_KEY']
                    flux_cco = sum(mx_mo_agg[mo].get(key, [0, 0, 0, 0])[0] for mo in yr_mois if mo in mx_mo_agg)
                    pnb_cco = sum(mx_mo_agg[mo].get(key, [0, 0, 0, 0])[1] for mo in yr_mois if mo in mx_mo_agg)
                    nb_cco = sum(mx_mo_agg[mo].get(key, [0, 0, 0, 0])[2] for mo in yr_mois if mo in mx_mo_agg)
                    tr_cco = sum(mx_mo_agg[mo].get(key, [0, 0, 0, 0])[3] for mo in yr_mois if mo in mx_mo_agg)
                    flux_cco_t = sum(mx_mo_agg[mo].get(key, [0, 0, 0, 0])[0] for mo in yr_last3 if mo in mx_mo_agg)
                    df_year.at[idx, 'FLUX_CCO_ANNEE'] = flux_cco
                    df_year.at[idx, 'PNB_CCO_ANNEE'] = pnb_cco
                    df_year.at[idx, 'NB_CARTES_CCO_ANNEE'] = nb_cco
                    df_year.at[idx, 'NB_TRANS_CCO_ANNEE'] = tr_cco
                    df_year.at[idx, 'FLUX_CCO_TRIMESTRE_FIN'] = flux_cco_t
                    df_year.at[idx, 'NB_CARTE_MOYEN'] = round(nb_cco / 12.0, 4)
                    flux_cpc = sum(wl_mo_agg[mo].get(key, [0, 0, 0, 0])[0] for mo in yr_mois if mo in wl_mo_agg)
                    pnb_cpc = sum(wl_mo_agg[mo].get(key, [0, 0, 0, 0])[1] for mo in yr_mois if mo in wl_mo_agg)
                    nb_cpc = sum(wl_mo_agg[mo].get(key, [0, 0, 0, 0])[2] for mo in yr_mois if mo in wl_mo_agg)
                    tr_cpc = sum(wl_mo_agg[mo].get(key, [0, 0, 0, 0])[3] for mo in yr_mois if mo in wl_mo_agg)
                    flux_cpc_t = sum(wl_mo_agg[mo].get(key, [0, 0, 0, 0])[0] for mo in yr_last3 if mo in wl_mo_agg)
                    df_year.at[idx, 'FLUX_CPC_ANNEE'] = flux_cpc
                    df_year.at[idx, 'PNB_CPC_ANNEE'] = pnb_cpc
                    df_year.at[idx, 'NB_CARTES_CPC_ANNEE'] = nb_cpc
                    df_year.at[idx, 'NB_TRANS_CPC_ANNEE'] = tr_cpc
                    df_year.at[idx, 'FLUX_CPC_TRIMESTRE_FIN'] = flux_cpc_t
                    if key in wl_yr_info:
                        df_year.at[idx, 'PLAFOND_CPC_EUR'] = wl_yr_info[key]['plafond']
                        df_year.at[idx, 'PERIODICITE_CPC'] = wl_yr_info[key]['perio']
                        diff_wl = wl_yr_info[key]['differe']
                        df_year.at[idx, 'DIFFERE_CPC_TYPE'] = self._differe_type(diff_wl, 'CPC')
                        df_year.at[idx, 'DIFFERE_CPC_JOURS'] = self._differe_jours(diff_wl)
                    else:
                        df_year.at[idx, 'PLAFOND_CPC_EUR'] = 0.0
                        df_year.at[idx, 'PERIODICITE_CPC'] = 1.0
                        df_year.at[idx, 'DIFFERE_CPC_TYPE'] = ''
                        df_year.at[idx, 'DIFFERE_CPC_JOURS'] = 0
                    if key in mx_yr_info:
                        diff_mx = mx_yr_info[key]['differe']
                        df_year.at[idx, 'DIFFERE_CCO_TYPE'] = self._differe_type(diff_mx, 'CCO')
                        df_year.at[idx, 'DIFFERE_CCO_JOURS'] = self._differe_jours(diff_mx)
                    else:
                        df_year.at[idx, 'DIFFERE_CCO_TYPE'] = ''
                        df_year.at[idx, 'DIFFERE_CCO_JOURS'] = 0
                    df_year.at[idx, 'TAG_SANS_FLUX_MX'] = 'OUI' if (flux_cco == 0 and pnb_cco == 0 and nb_cco == 0) else 'NON'
                    df_year.at[idx, 'TAG_SANS_FLUX_WL'] = 'OUI' if (flux_cpc == 0 and pnb_cpc == 0 and nb_cpc == 0) else 'NON'

                # Rebate + calcul natif
                self._apply_rebate_to_df_year(df_year, year)
                self._compute_financials(df_year, year, constants)
                df_year_by_year[year] = df_year

            self._year_refin = {y: (constants.get('taux_refinancement_2026', 0.0166)
                                    if y == '2026' else constants.get('taux_refinancement', 0.0099))
                                for y in years_final}

            # ── 1) DICTIONNAIRE (1re feuille) ───────────────────────────────
            self._prog(0.855, "Onglet DICTIONNAIRE...")
            self._build_dictionnaire_sheet(wb)

            # ── 2) COMMISSIONS PAR PAYS ─────────────────────────────────────
            self._prog(0.86, "Onglet COMMISSIONS PAR PAYS...")
            self._build_pays_table_sheet(wb, df_year_by_year, years_final)

            # ── 3) Onglets de calcul (Global/CCO/CPC × années) ──────────────
            for year in years_final:
                if year not in df_year_by_year:
                    continue
                self._prog(0.88, f"Onglet COMMISSION GLOBAL {year}...")
                self._write_native_sheet(
                    wb, f"COMMISSION GLOBAL {year}", df_year_by_year[year],
                    self._spec_global(), year, constants,
                    "GLOBAL — vue complète (valeurs natives, sans formule)")
            for year in years_final:
                if year not in df_year_by_year:
                    continue
                self._prog(0.90, f"Onglet COMMISSION CCO {year}...")
                self._write_native_sheet(
                    wb, f"COMMISSION CCO {year}",
                    df_year_by_year[year].sort_values(
                        by='NOM_ENTITE', key=lambda s: s.astype(str).str.upper()),
                    self._spec_cco(), year, constants,
                    "CCO (MONEXT) — tri alphabétique Nom Entité", accent='G')
            for year in years_final:
                if year not in df_year_by_year:
                    continue
                self._prog(0.92, f"Onglet COMMISSION CPC {year}...")
                self._write_native_sheet(
                    wb, f"COMMISSION CPC {year}",
                    df_year_by_year[year].sort_values(
                        by='NOM_ENTITE', key=lambda s: s.astype(str).str.upper()),
                    self._spec_cpc(), year, constants,
                    "CPC (WORLDLINE) — tri alphabétique Nom Entité", accent='P')

            # ── 4) SYNTHÈSE ─────────────────────────────────────────────────
            self._prog(0.96, "Onglet SYNTHÈSE...")
            self._build_synthese_sheet(wb, df_data, df_year_by_year, all_mois)

            self._prog(0.98, "Sauvegarde XLSX...")
            wb.save(xlsx_path)
            self._prog(1.0, "Terminé !")
            print(f"[INFO] Fichier généré ! [{VERSION_ID}]")
            print(f"[INFO]   - {self._n_clients:,} clients")
            print("[INFO]   - Valeurs natives (aucune formule)")
            print(f"[INFO]   - Refin 2025 = {self._year_refin.get('2025', 0.0099)*100:.2f}% / "
                  f"2026 = {self._year_refin.get('2026', 0.0166)*100:.2f}%")
            print(f"[INFO]   - REBATE indexés : {len(self._rebate_by_year_rmpm)} (annee, RMPM)")
            print("[INFO]   Onglets : DICTIONNAIRE, COMMISSIONS PAR PAYS, "
                  "GLOBAL/CCO/CPC par année, SYNTHÈSE.")
        except Exception:
            import traceback
            traceback.print_exc()
            raise

    # ── Rebate par année (inchangé sur le principe NO7WK) ───────────────────
    def _apply_rebate_to_df_year(self, df_year, year):
        if df_year is None or len(df_year) == 0:
            return
        rebate_dict = getattr(self, '_rebate_by_year_rmpm', {})
        df_year['REBATE_CCO_MONTANT'] = 0.0
        df_year['REBATE_CPC_MONTANT'] = 0.0
        if not rebate_dict:
            return
        year_str = str(year)
        for idx in df_year.index:
            rmpm = str(df_year.at[idx, 'RMPM']).strip()
            if not rmpm:
                continue
            hit = rebate_dict.get((year_str, rmpm))
            if hit is None:
                stripped = rmpm.lstrip('0')
                if stripped and stripped != rmpm:
                    hit = rebate_dict.get((year_str, stripped))
            if hit is None:
                for (yr_k, rmpm_k), val_dict in rebate_dict.items():
                    if yr_k != year_str:
                        continue
                    if rmpm.lstrip('0') == rmpm_k.lstrip('0'):
                        hit = val_dict
                        break
            if hit is None:
                continue
            cco = float(hit.get('MONEXT', 0.0))
            cpc = float(hit.get('WORLDLINE', 0.0))
            if cco > 0:
                df_year.at[idx, 'REBATE_CCO_MONTANT'] = cco
            if cpc > 0:
                df_year.at[idx, 'REBATE_CPC_MONTANT'] = cpc

    # ── Spécifications de colonnes (df_col, display, width, kind, color) ─────
    # kind : 't'=texte 'm'=montant 'i'=entier 'p'=pourcentage
    def _spec_global(self):
        return [
            ('RMPM', 'RMPM', 14, 't', 'B'), ('ID_RP', 'ID_RP', 16, 't', 'B'),
            ('ID_RC_MX', 'ID_RC_MX', 16, 't', 'B'), ('ID_RC_WL', 'ID_RC_WL', 16, 't', 'B'),
            ('CODE_GA', 'CODE_GA', 12, 't', 'B'), ('NOM_GA', 'NOM_GA', 28, 't', 'B'),
            ('PAYS_GA', 'PAYS_GA', 14, 't', 'B'), ('NOM_ENTITE', 'NOM_ENTITE', 30, 't', 'B'),
            ('PAYS_ENTITE', 'PAYS_ENTITE', 14, 't', 'B'),
            ('PAYS_APPORTEUR_CCO', 'PAYS_APPORTEUR_CCO', 16, 't', 'B'),
            ('PAYS_APPORTEUR_CPC', 'PAYS_APPORTEUR_CPC', 16, 't', 'B'),
            ('PAYS_FINAL', 'PAYS_FINAL', 14, 't', 'B'),
            ('NOM_PROG_CCO', 'NOM_PROG_CCO', 22, 't', 'G'),
            ('NOM_PROG_CPC', 'NOM_PROG_CPC', 22, 't', 'G'),
            ('SOURCE_MATCHING', 'SOURCE_MATCHING', 15, 't', 'G'),
            ('SEGMENT', 'SEGMENT', 12, 't', 'G'), ('SEGMENT_SOURCE', 'SEGMENT_SOURCE', 14, 't', 'G'),
            ('PRESENT_CCO', 'PRESENT_CCO', 11, 't', 'G'), ('PRESENT_CPC', 'PRESENT_CPC', 11, 't', 'G'),
            ('TAG_SANS_FLUX_MX', 'TAG_SANS_FLUX_MX', 14, 't', 'G'),
            ('TAG_SANS_FLUX_WL', 'TAG_SANS_FLUX_WL', 14, 't', 'G'),
            ('TYPE_DEBIT_CCO', 'TYPE_DEBIT_CCO', 14, 't', 'O'),
            ('RW_RETENU', 'RW_RETENU', 12, 'p', 'O'),
            ('DIFFERE_CCO_TYPE', 'DIFFERE_CCO_TYPE', 14, 't', 'G'),
            ('DIFFERE_CCO_JOURS', 'DIFFERE_CCO_JOURS', 12, 'i', 'G'),
            ('NB_CARTES_CCO_ANNEE', 'NB_CARTES_CCO_ANNEE', 14, 'm', 'G'),
            ('NB_CARTE_MOYEN', 'NB_CARTE_MOYEN', 14, 'm', 'G'),
            ('NB_TRANS_CCO_ANNEE', 'NB_TRANS_CCO_ANNEE', 14, 'm', 'G'),
            ('FLUX_CCO_ANNEE', 'FLUX_CCO_ANNEE', 16, 'm', 'G'),
            ('FLUX_CCO_TRIMESTRE_FIN', 'FLUX_CCO_TRIMESTRE_FIN', 16, 'm', 'G'),
            ('PNB_CCO_ANNEE', 'PNB_CCO_ANNEE', 16, 'm', 'G'),
            ('COUT_DIFFERE_CCO', 'COUT_DIFFERE_CCO', 14, 'm', 'G'),
            ('COUT_RUN_CCO', 'COUT_RUN_CCO', 14, 'm', 'G'),
            ('RWA_PLAFOND_CCO', 'RWA_PLAFOND_CCO', 14, 'm', 'G'),
            ('EAD_BILAN_CCO', 'EAD_BILAN_CCO', 14, 'm', 'G'),
            ('RWA_BILAN_CCO', 'RWA_BILAN_CCO', 14, 'm', 'G'),
            ('TOTAL_COUTS_CCO', 'TOTAL_COUTS_CCO', 14, 'm', 'G'),
            ('REBATE_CCO_MONTANT', 'REBATE_CCO_MONTANT', 14, 'm', 'O'),
            ('COMMISSION_CCO_TTC', 'COMMISSION_CCO_TTC', 16, 'm', 'G'),
            ('COMMISSION_CCO_HT', 'COMMISSION_CCO_HT', 16, 'm', 'G'),
            ('RESULTAT_COND_CCO', 'RESULTAT_COND_CCO', 16, 'm', 'G'),
            ('APPLICABILITE_CCO', 'APPLICABILITE_CCO', 14, 't', 'G'),
            ('DIFFERE_CPC_TYPE', 'DIFFERE_CPC_TYPE', 14, 't', 'P'),
            ('DIFFERE_CPC_JOURS', 'DIFFERE_CPC_JOURS', 12, 'i', 'P'),
            ('NB_CARTES_CPC_ANNEE', 'NB_CARTES_CPC_ANNEE', 14, 'm', 'P'),
            ('NB_TRANS_CPC_ANNEE', 'NB_TRANS_CPC_ANNEE', 14, 'm', 'P'),
            ('FLUX_CPC_ANNEE', 'FLUX_CPC_ANNEE', 16, 'm', 'P'),
            ('FLUX_CPC_TRIMESTRE_FIN', 'FLUX_CPC_TRIMESTRE_FIN', 16, 'm', 'P'),
            ('PNB_CPC_ANNEE', 'PNB_CPC_ANNEE', 16, 'm', 'P'),
            ('PLAFOND_CPC_EUR', 'PLAFOND_CPC_EUR', 14, 'm', 'P'),
            ('PERIODICITE_CPC', 'PERIODICITE_CPC', 12, 'i', 'P'),
            ('COUT_DIFFERE_CPC', 'COUT_DIFFERE_CPC', 14, 'm', 'P'),
            ('COUT_RUN_CPC', 'COUT_RUN_CPC', 14, 'm', 'P'),
            ('MULT_PERIODICITE', 'MULT_PERIODICITE', 12, 'i', 'P'),
            ('EAD_PLAFOND_CPC', 'EAD_PLAFOND_CPC', 14, 'm', 'P'),
            ('RWA_PLAFOND_CPC', 'RWA_PLAFOND_CPC', 14, 'm', 'P'),
            ('EAD_BILAN_CPC', 'EAD_BILAN_CPC', 14, 'm', 'P'),
            ('RWA_BILAN_CPC', 'RWA_BILAN_CPC', 14, 'm', 'P'),
            ('TOTAL_COUTS_CPC', 'TOTAL_COUTS_CPC', 14, 'm', 'P'),
            ('REBATE_CPC_MONTANT', 'REBATE_CPC_MONTANT', 14, 'm', 'O'),
            ('COMMISSION_CPC_TTC', 'COMMISSION_CPC_TTC', 16, 'm', 'P'),
            ('COMMISSION_CPC_HT', 'COMMISSION_CPC_HT', 16, 'm', 'P'),
            ('RESULTAT_COND_CPC', 'RESULTAT_COND_CPC', 16, 'm', 'P'),
            ('APPLICABILITE_CPC', 'APPLICABILITE_CPC', 14, 't', 'P'),
            ('COMMISSION_TOTALE_TTC', 'COMMISSION_TOTALE_TTC', 16, 'm', 'D'),
            ('COMMISSION_TOTAL_HT', 'COMMISSION_TOTAL_HT', 16, 'm', 'D'),
            ('COUTS_TOTAUX_CLIENT', 'COUTS_TOTAUX_CLIENT', 16, 'm', 'D'),
            ('PNB_TOTAL_CLIENT', 'PNB_TOTAL_CLIENT', 16, 'm', 'D'),
        ]

    def _spec_cco(self):
        return [
            ('RMPM', 'RMPM', 14, 't', 'B'), ('CODE_GA', 'CODE_GA', 12, 't', 'B'),
            ('NOM_GA', 'NOM_GA', 28, 't', 'B'), ('NOM_ENTITE', 'NOM_ENTITE', 30, 't', 'B'),
            ('PAYS_FINAL', 'PAYS_FINAL', 14, 't', 'B'),
            ('PAYS_APPORTEUR_CCO', 'PAYS_APPORTEUR_CCO', 16, 't', 'B'),
            ('SEGMENT', 'SEGMENT', 12, 't', 'B'), ('NOM_PROG_CCO', 'NOM_PROG_CCO', 22, 't', 'B'),
            ('TAG_SANS_FLUX_MX', 'TAG_SANS_FLUX_MX', 14, 't', 'B'),
            ('TYPE_DEBIT_CCO', 'TYPE_DEBIT_CCO', 14, 't', 'O'),
            ('RW_RETENU', 'RW_RETENU', 12, 'p', 'O'),
            ('DIFFERE_CCO_TYPE', 'DIFFERE_CCO_TYPE', 14, 't', 'G'),
            ('DIFFERE_CCO_JOURS', 'DIFFERE_CCO_JOURS', 12, 'i', 'G'),
            ('NB_CARTES_CCO_ANNEE', 'NB_CARTES_CCO_ANNEE', 14, 'm', 'G'),
            ('NB_CARTE_MOYEN', 'NB_CARTE_MOYEN', 14, 'm', 'G'),
            ('FLUX_CCO_ANNEE', 'FLUX_CCO_ANNEE', 16, 'm', 'G'),
            ('FLUX_CCO_TRIMESTRE_FIN', 'FLUX_CCO_TRIMESTRE_FIN', 16, 'm', 'G'),
            ('PNB_CCO_ANNEE', 'PNB_CCO_ANNEE', 16, 'm', 'G'),
            ('COUT_DIFFERE_CCO', 'COUT_DIFFERE_CCO', 14, 'm', 'G'),
            ('COUT_RUN_CCO', 'COUT_RUN_CCO', 14, 'm', 'G'),
            ('RWA_PLAFOND_CCO', 'RWA_PLAFOND_CCO', 14, 'm', 'G'),
            ('EAD_BILAN_CCO', 'EAD_BILAN_CCO', 14, 'm', 'G'),
            ('RWA_BILAN_CCO', 'RWA_BILAN_CCO', 14, 'm', 'G'),
            ('TOTAL_COUTS_CCO', 'TOTAL_COUTS_CCO', 14, 'm', 'G'),
            ('REBATE_CCO_MONTANT', 'REBATE_CCO_MONTANT', 14, 'm', 'O'),
            ('COMMISSION_CCO_TTC', 'COMMISSION_CCO_TTC', 16, 'm', 'G'),
            ('COMMISSION_CCO_HT', 'COMMISSION_CCO_HT', 16, 'm', 'D'),
            ('RESULTAT_COND_CCO', 'RESULTAT_COND_CCO', 16, 'm', 'G'),
            ('APPLICABILITE_CCO', 'APPLICABILITE_CCO', 14, 't', 'G'),
        ]

    def _spec_cpc(self):
        return [
            ('RMPM', 'RMPM', 14, 't', 'B'), ('CODE_GA', 'CODE_GA', 12, 't', 'B'),
            ('NOM_GA', 'NOM_GA', 28, 't', 'B'), ('NOM_ENTITE', 'NOM_ENTITE', 30, 't', 'B'),
            ('PAYS_FINAL', 'PAYS_FINAL', 14, 't', 'B'),
            ('PAYS_APPORTEUR_CPC', 'PAYS_APPORTEUR_CPC', 16, 't', 'B'),
            ('SEGMENT', 'SEGMENT', 12, 't', 'B'), ('NOM_PROG_CPC', 'NOM_PROG_CPC', 22, 't', 'B'),
            ('TAG_SANS_FLUX_WL', 'TAG_SANS_FLUX_WL', 14, 't', 'B'),
            ('RW_RETENU', 'RW_RETENU', 12, 'p', 'O'),
            ('DIFFERE_CPC_TYPE', 'DIFFERE_CPC_TYPE', 14, 't', 'P'),
            ('DIFFERE_CPC_JOURS', 'DIFFERE_CPC_JOURS', 12, 'i', 'P'),
            ('PLAFOND_CPC_EUR', 'PLAFOND_CPC_EUR', 14, 'm', 'P'),
            ('PERIODICITE_CPC', 'PERIODICITE_CPC', 12, 'i', 'P'),
            ('NB_CARTES_CPC_ANNEE', 'NB_CARTES_CPC_ANNEE', 14, 'm', 'P'),
            ('NB_TRANS_CPC_ANNEE', 'NB_TRANS_CPC_ANNEE', 14, 'm', 'P'),
            ('FLUX_CPC_ANNEE', 'FLUX_CPC_ANNEE', 16, 'm', 'P'),
            ('FLUX_CPC_TRIMESTRE_FIN', 'FLUX_CPC_TRIMESTRE_FIN', 16, 'm', 'P'),
            ('PNB_CPC_ANNEE', 'PNB_CPC_ANNEE', 16, 'm', 'P'),
            ('COUT_DIFFERE_CPC', 'COUT_DIFFERE_CPC', 14, 'm', 'P'),
            ('COUT_RUN_CPC', 'COUT_RUN_CPC', 14, 'm', 'P'),
            ('MULT_PERIODICITE', 'MULT_PERIODICITE', 12, 'i', 'P'),
            ('EAD_PLAFOND_CPC', 'EAD_PLAFOND_CPC', 14, 'm', 'P'),
            ('RWA_PLAFOND_CPC', 'RWA_PLAFOND_CPC', 14, 'm', 'P'),
            ('EAD_BILAN_CPC', 'EAD_BILAN_CPC', 14, 'm', 'P'),
            ('RWA_BILAN_CPC', 'RWA_BILAN_CPC', 14, 'm', 'P'),
            ('TOTAL_COUTS_CPC', 'TOTAL_COUTS_CPC', 14, 'm', 'P'),
            ('REBATE_CPC_MONTANT', 'REBATE_CPC_MONTANT', 14, 'm', 'O'),
            ('COMMISSION_CPC_TTC', 'COMMISSION_CPC_TTC', 16, 'm', 'P'),
            ('COMMISSION_CPC_HT', 'COMMISSION_CPC_HT', 16, 'm', 'D'),
            ('RESULTAT_COND_CPC', 'RESULTAT_COND_CPC', 16, 'm', 'P'),
            ('APPLICABILITE_CPC', 'APPLICABILITE_CPC', 14, 't', 'P'),
        ]

    # ── Écrivain de feuille natif générique ─────────────────────────────────
    def _write_native_sheet(self, wb, title, df, spec, year, constants,
                            subtitle, accent='G'):
        ws = wb.create_sheet(title=title)
        ws.column_dimensions["A"].width = 2.5

        def fill_(h):
            return PatternFill(start_color=h, end_color=h, fill_type="solid")

        def fnt_(c=WHT, bold=True, sz=10):
            return Font(name="Segoe UI", size=sz, bold=bold, color=c)

        thin = Side(style="thin", color="CCCCCC")
        brd = Border(left=thin, right=thin, top=thin, bottom=thin)
        acc2 = self.COLOR_CELL.get(accent, GRN2)

        refin = (constants.get('taux_refinancement_2026', 0.0166) if str(year) == '2026'
                 else constants.get('taux_refinancement', 0.0099))

        ws["B2"] = (f"CIB COMMISSIONNEMENT — {subtitle} — {year} — "
                    f"{datetime.now().strftime('%d/%m/%Y')} — [{VERSION_ID}]")
        ws["B2"].font = fnt_(DARK, sz=14)
        ws["B2"].fill = fill_(acc2)
        ws["B3"] = ("Valeurs natives (aucune formule). "
                    "Bonus impl/signing = 0, RW = RW défaut, type = Entreprise (hypothèses figées).")
        ws["B3"].font = fnt_(DARK, bold=False, sz=9)
        ws["B3"].fill = fill_(acc2)
        ws["B5"] = (f"Hypothèses {year} : refinancement = {refin*100:.2f}%  |  "
                    f"RW = {constants.get('rw_defaut', 0.65)*100:.0f}%  |  "
                    f"part capital = {constants.get('part_capital_banque', 0.12)*100:.0f}%  |  "
                    f"ALMT = {constants.get('almt', 0.142)*100:.1f}%  |  "
                    f"commission = {constants.get('taux_commission', 0.20)*100:.0f}%  |  "
                    f"TVA = {constants.get('taux_tva', 0.20)*100:.0f}%")
        ws["B5"].font = fnt_(ORA, bold=True, sz=9)
        ws["B5"].fill = fill_(ORA2)

        HDR_ROW = 7
        for ci, (src, disp, width, kind, color) in enumerate(spec):
            col_idx = 2 + ci
            c = ws.cell(row=HDR_ROW, column=col_idx, value=disp)
            c.font = fnt_(WHT, sz=9)
            c.fill = fill_(self.COLOR_HDR.get(color, DARK))
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = brd
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[HDR_ROW].height = 42

        DATA_START = HDR_ROW + 1
        df_reset = df.reset_index(drop=True)
        n_rows = len(df_reset)
        for i, row_obj in enumerate(df_reset.itertuples(index=False)):
            rn = DATA_START + i
            bg = WHT if i % 2 == 0 else acc2
            dd = row_obj._asdict()
            for ci, (src, disp, width, kind, color) in enumerate(spec):
                c = ws.cell(row=rn, column=2 + ci)
                c.font = Font(name="Segoe UI", size=9)
                c.border = brd
                c.fill = fill_(bg)
                val = dd.get(src, '')
                if kind in ('m', 'i', 'p'):
                    try:
                        num = float(val) if val not in ('', None) else 0.0
                    except (ValueError, TypeError):
                        num = 0.0
                    c.value = round(num, 2) if kind == 'm' else (round(num, 4) if kind == 'p' else int(round(num)))
                    c.number_format = {'m': '#,##0.00', 'i': '0', 'p': '0.00%'}[kind]
                    c.alignment = Alignment(horizontal="right")
                else:
                    c.value = str(val) if val not in (None, 'nan', 'NaN') else ''
                    c.alignment = Alignment(horizontal="left")
        ws.freeze_panes = f"B{DATA_START}"
        last_col = get_column_letter(1 + len(spec))
        last_row = DATA_START + max(n_rows - 1, 0)
        ws.auto_filter.ref = f"B{HDR_ROW}:{last_col}{last_row}"

    # ═════════════════════════════════════════════════════════════════════════
    # ONGLET COMMISSIONS PAR PAYS (demande Bénédicte) — une table par année
    # ─────────────────────────────────────────────────────────────────────────
    # CCO : somme COMMISSION_CCO_HT par PAYS_APPORTEUR_CCO,
    #       hors Belgium/France/Luxembourg, APPLICABILITE_CCO = OUI.
    # CPC : somme COMMISSION_CPC_HT par PAYS_APPORTEUR_CPC,
    #       hors Belgium/France/Luxembourg, hors « carte achat » (NOM_PROG_CPC
    #       contient ACHAT), APPLICABILITE_CPC = OUI.
    # Noms de pays en anglais (déjà nettoyés). Tri alpha + « Total général ».
    # ═════════════════════════════════════════════════════════════════════════
    def _build_pays_table_sheet(self, wb, df_year_by_year, years_final):
        ws = wb.create_sheet(title="COMMISSIONS PAR PAYS")
        ws.column_dimensions["A"].width = 2.5

        def fill_(h):
            return PatternFill(start_color=h, end_color=h, fill_type="solid")

        def fnt_(c=WHT, bold=True, sz=10):
            return Font(name="Segoe UI", size=sz, bold=bold, color=c)

        thin = Side(style="thin", color="CCCCCC")
        brd = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws["B2"] = (f"CIB COMMISSIONNEMENT — COMMISSIONS PAR PAYS — "
                    f"{datetime.now().strftime('%d/%m/%Y')} — [{VERSION_ID}]")
        ws["B2"].font = fnt_(DARK, sz=14)
        ws["B2"].fill = fill_(BLU2)
        ws["B3"] = ("CCO : hors Belgium/France/Luxembourg, applicabilité CCO = Oui.  "
                    "CPC : idem + hors « carte achat », applicabilité CPC = Oui.  "
                    "Montants HT, noms de pays en anglais.")
        ws["B3"].font = fnt_(DARK, bold=False, sz=9)
        ws["B3"].fill = fill_(BLU2)

        for k, w in enumerate([34, 18, 18, 18]):
            ws.column_dimensions[get_column_letter(2 + k)].width = w

        row = 5
        kw = CARTE_ACHAT_KEYWORD.upper()
        for year in years_final:
            df_year = df_year_by_year.get(year)
            if df_year is None or len(df_year) == 0:
                continue

            per_pays = {}

            def _acc(pays, key, montant):
                p = str(pays).strip()
                if not p:
                    return
                if p.upper() in PAYS_EXCLUS_COMMISSION:
                    return
                d = per_pays.setdefault(p, {'cco': 0.0, 'cpc': 0.0})
                d[key] += float(montant)

            # CCO
            if 'COMMISSION_CCO_HT' in df_year.columns:
                for pays, applic, montant in zip(
                    df_year.get('PAYS_APPORTEUR_CCO', pd.Series([''] * len(df_year))).values,
                    df_year.get('APPLICABILITE_CCO', pd.Series([''] * len(df_year))).astype(str).values,
                    pd.to_numeric(df_year['COMMISSION_CCO_HT'], errors='coerce').fillna(0.0).values
                ):
                    if str(applic).strip().upper() == 'OUI' and montant != 0:
                        _acc(pays, 'cco', montant)

            # CPC (hors carte achat)
            if 'COMMISSION_CPC_HT' in df_year.columns:
                nom_prog = df_year.get('NOM_PROG_CPC', pd.Series([''] * len(df_year))).astype(str)
                is_achat = nom_prog.apply(lambda s: kw in self.norm_map(s))
                for pays, applic, montant, achat in zip(
                    df_year.get('PAYS_APPORTEUR_CPC', pd.Series([''] * len(df_year))).values,
                    df_year.get('APPLICABILITE_CPC', pd.Series([''] * len(df_year))).astype(str).values,
                    pd.to_numeric(df_year['COMMISSION_CPC_HT'], errors='coerce').fillna(0.0).values,
                    is_achat.values
                ):
                    if str(applic).strip().upper() == 'OUI' and montant != 0 and not achat:
                        _acc(pays, 'cpc', montant)

            # Titre année
            tcell = ws.cell(row=row, column=2, value=f"Tableau commissions par pays — {year}")
            tcell.font = fnt_(WHT, sz=12)
            tcell.fill = fill_(DARK)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            tcell.alignment = Alignment(horizontal="left", vertical="center")
            row += 1

            # En-têtes
            headers = ["Pays Apporteur d'affaire", "Commission CCO HT",
                       "Commission CPC HT", "TOTAL"]
            for k, h in enumerate(headers):
                c = ws.cell(row=row, column=2 + k, value=h)
                c.font = fnt_(WHT, sz=10)
                c.fill = fill_(BLU)
                c.alignment = Alignment(horizontal="center", wrap_text=True)
                c.border = brd
            row += 1

            data_start = row
            tot_cco = tot_cpc = tot_tot = 0.0
            for i, pays in enumerate(sorted(per_pays.keys(), key=lambda s: s.upper())):
                cco = round(per_pays[pays]['cco'], 2)
                cpc = round(per_pays[pays]['cpc'], 2)
                total = round(cco + cpc, 2)
                tot_cco += cco; tot_cpc += cpc; tot_tot += total
                bg = WHT if i % 2 == 0 else BLU2
                vals = [pays,
                        cco if cco != 0 else None,
                        cpc if cpc != 0 else None,
                        total if total != 0 else None]
                for k, v in enumerate(vals):
                    c = ws.cell(row=row, column=2 + k, value=v)
                    c.font = Font(name="Segoe UI", size=9)
                    c.fill = fill_(bg)
                    c.border = brd
                    if k == 0:
                        c.alignment = Alignment(horizontal="left")
                    else:
                        c.number_format = '#,##0.00'
                        c.alignment = Alignment(horizontal="right")
                row += 1

            # Total général
            tot_vals = ["Total général", round(tot_cco, 2), round(tot_cpc, 2), round(tot_tot, 2)]
            for k, v in enumerate(tot_vals):
                c = ws.cell(row=row, column=2 + k, value=v)
                c.font = fnt_(WHT, sz=10)
                c.fill = fill_(GRN)
                c.border = brd
                if k == 0:
                    c.alignment = Alignment(horizontal="left")
                else:
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right")
            row += 1
            ws.row_dimensions[data_start - 1].height = 28
            row += 2  # espace entre années

    # ═════════════════════════════════════════════════════════════════════════
    # ONGLET DICTIONNAIRE (1re feuille) — toutes les colonnes créées
    # ═════════════════════════════════════════════════════════════════════════
    def _build_dictionnaire_sheet(self, wb):
        ws = wb.create_sheet(title="DICTIONNAIRE")
        ws.column_dimensions["A"].width = 2.5
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 64

        def fill_(h):
            return PatternFill(start_color=h, end_color=h, fill_type="solid")

        def fnt_(c=WHT, bold=True, sz=10):
            return Font(name="Segoe UI", size=sz, bold=bold, color=c)

        thin = Side(style="thin", color="CCCCCC")
        brd = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws["B2"] = f"CIB COMMISSIONNEMENT — DICTIONNAIRE DES COLONNES — [{VERSION_ID}]"
        ws["B2"].font = fnt_(DARK, sz=14)
        ws["B2"].fill = fill_(GRN2)
        ws.merge_cells("B2:D2")
        ws["B3"] = ("Toutes les colonnes créées par le programme. "
                    "Valeurs natives (aucune formule). Hypothèses figées : "
                    "bonus = 0, RW = RW défaut, type débit = Entreprise.")
        ws["B3"].font = fnt_(DARK, bold=False, sz=9)
        ws["B3"].fill = fill_(GRN2)
        ws.merge_cells("B3:D3")

        DEFS = [
            ("— IDENTITÉ", "", ""),
            ("CLIENT_KEY", "Clé d'unicité client", "RMPM, sinon RC, sinon RP, sinon RS ; ORPH si rien."),
            ("RMPM", "Identifiant Relation/Marché", "Issu de la cascade ACCOUNT > REF_CLIENT > IDSEG."),
            ("ID_RP / ID_RC_MX / ID_RC_WL", "Identifiants relation/contrat", "Protégés en texte (=\"...\") pour garder les zéros."),
            ("CODE_GA / NOM_GA / PAYS_GA", "Business group (agrégat)", "Code, nom et pays du groupe d'affaires."),
            ("NOM_ENTITE / PAYS_ENTITE", "Entité juridique", "Nom et pays de l'entité légale (RMPM)."),
            ("RS_CCO / RS_CPC", "Raison sociale source", "Raison sociale normalisée MONEXT / Worldline."),
            ("— MATCHING & SEGMENT", "", ""),
            ("SOURCE_MATCHING", "Origine de l'identité", "REF_RC, REF_IBAN, ACCOUNT_IBAN ou NON_TROUVE."),
            ("SEGMENT / SEGMENT_SOURCE", "Segment client + origine", "BPE ou ENTREPRISE ; source du segment retenu."),
            ("PRESENT_CCO / PRESENT_CPC", "Présence par flux", "YES/NO selon présence dans MONEXT / Worldline."),
            ("IS_BPE / IS_ENT", "Indicateurs de segment", "YES/NO dérivés du segment."),
            ("PAYS_APPORTEUR_CCO/CPC", "Pays apporteur (nettoyé)", "Pays MONEXT / Worldline uniformisé en anglais."),
            ("PAYS_FINAL", "Pays apporteur consolidé", "PAYS_APPORTEUR_CCO sinon CPC (anglais)."),
            ("TAG_SANS_FLUX_MX/WL", "Flux nuls", "OUI si flux, PNB et nb cartes nuls sur la période."),
            ("— FLUX ANNUELS", "", ""),
            ("NB_CARTES_CCO_ANNEE", "Nombre de cartes CCO/an", "Somme des cartes MONEXT sur l'année."),
            ("NB_CARTE_MOYEN", "Nb cartes moyen", "NB_CARTES_CCO_ANNEE / 12 ; base du coût de run CCO."),
            ("NB_TRANS_CCO/CPC_ANNEE", "Nb transactions/an", "Somme des transactions sur l'année."),
            ("FLUX_CCO/CPC_ANNEE", "Flux annuel (EUR)", "Somme des dépenses (WL converti en EUR)."),
            ("FLUX_CCO/CPC_TRIMESTRE_FIN", "Flux des 3 derniers mois", "Base du calcul EAD bilan."),
            ("PNB_CCO/CPC_ANNEE", "PNB annuel (EUR)", "Somme des colonnes PNB ; base de la commission."),
            ("PLAFOND_CPC_EUR", "Plafond CPC (EUR)", "Plafond Worldline max sur l'année (converti EUR)."),
            ("PERIODICITE_CPC", "Périodicité plafond", "1/12 → ×1 ; 3/6 → ×2 (multiplicateur EAD)."),
            ("DIFFERE_*_TYPE / _JOURS", "Type et jours de différé", "IMM / DIFFERE / FIN_MOIS et nb de jours."),
            ("— HYPOTHÈSES (figées)", "", ""),
            ("TYPE_DEBIT_CCO", "Type de débit", "Figé à « Entreprise » (plus de saisie Excel)."),
            ("RW_RETENU", "Risk weight retenu", "RW défaut (0,65) — figé."),
            ("TAUX_EAD_CCO", "Taux EAD CCO", "Taux EAD entreprise (informatif)."),
            ("IMPL_BONUS_* / SIGNING_BONUS_*", "Bonus", "Figés à 0 (plus de saisie Excel)."),
            ("REBATE_CCO/CPC_MONTANT", "Rebate appliqué", "Issu du fichier REBATE par (année, RMPM)."),
            ("— COÛTS CCO", "", ""),
            ("COUT_DIFFERE_CCO", "Coût du différé CCO", "0 si IMM/flux nul, sinon flux×(jours+15)/365×refin(année)."),
            ("COUT_RUN_CCO", "Coût de run CCO", "NB_CARTE_MOYEN × coût carte CCO."),
            ("RWA_PLAFOND_CCO", "RWA plafond CCO", "0 si IMM, sinon NB_CARTE_MOYEN × coût RWA CCO."),
            ("EAD_BILAN_CCO", "EAD bilan CCO", "Flux trimestre (×2/3 si différé ≤ 27 j)."),
            ("RWA_BILAN_CCO", "RWA bilan CCO", "EAD × RW × part capital × ALMT."),
            ("TOTAL_COUTS_CCO", "Total coûts CCO", "Différé + run + RWA plafond + RWA bilan."),
            ("— COÛTS CPC", "", ""),
            ("COUT_DIFFERE_CPC", "Coût du différé CPC", "0 si IMM/jours nul/flux nul, sinon flux×(jours+15)/365×refin(année)."),
            ("COUT_RUN_CPC", "Coût de run CPC", "NB_TRANS_CPC_ANNEE × coût transaction CPC."),
            ("MULT_PERIODICITE", "Multiplicateur périodicité", "1 (1/12 mois) ou 2 (3/6 mois)."),
            ("EAD_PLAFOND_CPC", "EAD plafond CPC", "Plafond CPC × multiplicateur périodicité."),
            ("RWA_PLAFOND_CPC", "RWA plafond CPC", "EAD plafond × taux EAD ent × RW × part cap × ALMT."),
            ("EAD_BILAN_CPC / RWA_BILAN_CPC", "EAD / RWA bilan CPC", "Mêmes règles que CCO côté Worldline."),
            ("TOTAL_COUTS_CPC", "Total coûts CPC", "Différé + run + RWA plafond + RWA bilan."),
            ("— COMMISSIONS & APPLICABILITÉ", "", ""),
            ("COMMISSION_CCO/CPC_TTC", "Commission TTC", "max(0 ; (PNB − rebate) × taux commission)."),
            ("COMMISSION_CCO/CPC_HT", "Commission HT", "Commission TTC / (1 + TVA) — base table par pays."),
            ("RESULTAT_COND_CCO/CPC", "Résultat conditionnel", "PNB − rebate − coûts − commission TTC."),
            ("APPLICABILITE_CCO/CPC", "Applicabilité", "OUI si résultat conditionnel > 0, sinon NON."),
            ("COMMISSION_TOTALE_TTC", "Commission totale TTC", "Somme des commissions applicables (OUI)."),
            ("COMMISSION_TOTAL_HT", "Commission totale HT", "Commission totale TTC / (1 + TVA)."),
            ("COUTS_TOTAUX_CLIENT", "Coûts totaux", "TOTAL_COUTS_CCO + TOTAL_COUTS_CPC."),
            ("PNB_TOTAL_CLIENT", "PNB total", "PNB_CCO_ANNEE + PNB_CPC_ANNEE."),
        ]

        HDR = 5
        for k, h in enumerate(["NOM_COLONNE", "DÉFINITION", "LOGIQUE DE CALCUL"]):
            c = ws.cell(row=HDR, column=2 + k, value=h)
            c.font = fnt_(WHT, sz=10)
            c.fill = fill_(BLU)
            c.alignment = Alignment(horizontal="left")
            c.border = brd
        r = HDR + 1
        for nom, defi, logique in DEFS:
            is_section = nom.startswith("—")
            for k, v in enumerate([nom, defi, logique]):
                c = ws.cell(row=r, column=2 + k, value=v)
                if is_section:
                    c.font = fnt_(WHT, bold=True, sz=10)
                    c.fill = fill_(DARK)
                else:
                    c.font = Font(name="Segoe UI", size=9,
                                  bold=(k == 0), color=("1F4E79" if k == 0 else "000000"))
                    c.fill = fill_(WHT if (r % 2 == 0) else BLU2)
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                c.border = brd
            if is_section:
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            r += 1
        ws.freeze_panes = "B6"

    # ═════════════════════════════════════════════════════════════════════════
    # ONGLET SYNTHÈSE — 4 blocs dashboard (repris NO7WK, valeurs natives)
    # ═════════════════════════════════════════════════════════════════════════
    def _build_synthese_sheet(self, wb, df_data, df_year_by_year, all_mois):
        ws = wb.create_sheet(title="SYNTHÈSE")
        ws.column_dimensions["A"].width = 2.5

        def fill_(h):
            return PatternFill(start_color=h, end_color=h, fill_type="solid")

        def fnt_(c=WHT, bold=True, sz=10):
            return Font(name="Segoe UI", size=sz, bold=bold, color=c)

        thin = Side(style="thin", color="CCCCCC")
        brd = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws["B2"] = (f"CIB COMMISSIONNEMENT — SYNTHÈSE DASHBOARD — "
                    f"{datetime.now().strftime('%d/%m/%Y')} — [{VERSION_ID}]")
        ws["B2"].font = fnt_(DARK, sz=14)
        ws["B2"].fill = fill_(GRN2)
        ws["B3"] = ("Quatre blocs d'analyse : classement des pays, "
                    "comparatif mensuel 2025 vs 2026, top 20 clients, répartition géographique.")
        ws["B3"].font = fnt_(DARK, bold=False, sz=9)
        ws["B3"].fill = fill_(GRN2)

        TABLE_GAP_COLS = 2
        BLOC_GAP_ROWS = 3
        years = sorted(df_year_by_year.keys())

        def years_label(year):
            yr_mois = [mo for mo in all_mois if mo.startswith(year)]
            if not yr_mois:
                return f"Année {year} (aucune donnée)"
            return (f"Période {self.mois_label(yr_mois[0])} → {self.mois_label(yr_mois[-1])} "
                    f"({len(yr_mois)} mois)")

        def put_table_title(row, col, title, subtitle, width, color_main, color_sub):
            c = ws.cell(row=row, column=col, value=title)
            c.font = fnt_(WHT, sz=11)
            c.fill = fill_(color_main)
            c.alignment = Alignment(horizontal="left", vertical="center")
            if width > 1:
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + width - 1)
            c2 = ws.cell(row=row + 1, column=col, value=subtitle)
            c2.font = fnt_(DARK, bold=False, sz=9)
            c2.fill = fill_(color_sub)
            c2.alignment = Alignment(horizontal="left", vertical="center")
            if width > 1:
                ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + width - 1)

        def put_headers(row, col, headers, color_main):
            for k, h in enumerate(headers):
                c = ws.cell(row=row, column=col + k, value=h)
                c.font = fnt_(WHT, sz=9)
                c.fill = fill_(color_main)
                c.alignment = Alignment(horizontal="center", wrap_text=True)
                c.border = brd

        def put_data_row(row, col, values, formats, bg):
            for k, (val, fmt) in enumerate(zip(values, formats)):
                c = ws.cell(row=row, column=col + k, value=val)
                c.font = Font(name="Segoe UI", size=9)
                c.fill = fill_(bg)
                c.border = brd
                if fmt:
                    c.number_format = fmt
                c.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")

        # BLOC 1 — Top 15 pays
        bloc1_title_row = 5
        c = ws.cell(row=bloc1_title_row, column=2,
                    value="BLOC 1 — Classement des pays par volume monétique")
        c.font = fnt_(WHT, sz=12); c.fill = fill_(DARK)
        ws.merge_cells(start_row=bloc1_title_row, start_column=2, end_row=bloc1_title_row, end_column=23)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=bloc1_title_row + 1, column=2,
                value="Top 15 des pays par volume de flux (CCO et CPC) — 2025 vs 2026.").font = \
            fnt_(DARK, bold=False, sz=9)
        table_start_row = bloc1_title_row + 3

        def build_top_pays(df_year, prefix):
            if df_year is None or len(df_year) == 0:
                return []
            flux_col = f'FLUX_{prefix}_ANNEE'; pnb_col = f'PNB_{prefix}_ANNEE'
            if flux_col not in df_year.columns:
                return []
            agg = df_year.groupby('PAYS_FINAL', dropna=False).agg(
                flux=(flux_col, 'sum'), pnb=(pnb_col, 'sum')).reset_index()
            agg = agg[agg['PAYS_FINAL'].astype(str).str.strip() != ''].copy()
            agg = agg[agg['flux'].astype(float) > 0]
            total_flux = float(agg['flux'].sum()) or 1.0
            agg['pct'] = agg['flux'].astype(float) / total_flux
            agg = agg.sort_values(by='flux', ascending=False).head(15)
            return list(agg.itertuples(index=False, name=None))

        configs_b1 = []
        for year in years:
            configs_b1.append(("CCO", year, GRN, GRN2))
            configs_b1.append(("CPC", year, PUR, PUR2))
        col_cursor = 2
        for prefix, year, col_main, col_sub in configs_b1:
            top = build_top_pays(df_year_by_year.get(year), prefix)
            put_table_title(table_start_row, col_cursor, f"Top 15 pays — {prefix} {year}",
                            years_label(year), 4, col_main, col_sub)
            put_headers(table_start_row + 2, col_cursor, ['PAYS', 'FLUX', 'PNB', '% TOTAL'], col_main)
            for k in range(4):
                ws.column_dimensions[get_column_letter(col_cursor + k)].width = [22, 16, 16, 10][k]
            for i, row_t in enumerate(top):
                pays, flux, pnb, pct = row_t
                bg = WHT if i % 2 == 0 else col_sub
                put_data_row(table_start_row + 3 + i, col_cursor,
                             [str(pays), float(flux), float(pnb), float(pct)],
                             [None, '#,##0.00', '#,##0.00', '0.00%'], bg)
            col_cursor += 4 + TABLE_GAP_COLS
        bloc1_end_row = table_start_row + 2 + 15

        # BLOC 2 — Comparatif mensuel
        bloc2_title_row = bloc1_end_row + BLOC_GAP_ROWS
        c = ws.cell(row=bloc2_title_row, column=2, value="BLOC 2 — Comparatif mensuel 2025 vs 2026")
        c.font = fnt_(WHT, sz=12); c.fill = fill_(DARK)
        ws.merge_cells(start_row=bloc2_title_row, start_column=2, end_row=bloc2_title_row, end_column=29)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=bloc2_title_row + 1, column=2,
                value="Suivi mensuel des flux et du PNB ; tendance et variation. Mois 2026 manquants laissés vides.").font = \
            fnt_(DARK, bold=False, sz=9)
        b2_table_start = bloc2_title_row + 3
        MOIS_ORDRE = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']

        def get_monthly_totals(prefix, metric):
            res = {}
            for mo in all_mois:
                lbl = self.mois_label(mo)
                col = f'{metric}_{prefix}_{lbl}'
                total = float(df_data[col].astype(float).sum()) if col in df_data.columns else 0.0
                res[(mo[:4], mo[4:])] = total
            return res

        def build_monthly_compare(prefix, metric):
            data = get_monthly_totals(prefix, metric)
            rows = []
            for mo in MOIS_ORDRE:
                v25 = data.get(('2025', mo), None)
                v26 = data.get(('2026', mo), None)
                lbl = MOIS_NOMS.get(mo, mo)
                if v25 is None:
                    v25 = 0.0
                if v26 is None or v26 == 0.0:
                    rows.append((lbl, float(v25), None, 'N/A', None))
                else:
                    if v25 == 0:
                        tend, pct = 'N/A', None
                    else:
                        delta = (v26 - v25) / v25
                        tend = 'HAUSSE' if delta > 0.001 else ('BAISSE' if delta < -0.001 else 'STABLE')
                        pct = float(delta)
                    rows.append((lbl, float(v25), float(v26), tend, pct))
            return rows

        configs_b2 = [
            ('FLUX', 'CCO', GRN, GRN2, "Flux CCO mois par mois"),
            ('FLUX', 'CPC', PUR, PUR2, "Flux CPC mois par mois"),
            ('PNB', 'CCO', GRN, GRN2, "PNB CCO mois par mois"),
            ('PNB', 'CPC', PUR, PUR2, "PNB CPC mois par mois"),
        ]
        col_cursor = 2
        for metric, prefix, col_main, col_sub, title in configs_b2:
            rows = build_monthly_compare(prefix, metric)
            put_table_title(b2_table_start, col_cursor, title,
                            "Comparatif 2025 vs 2026.", 5, col_main, col_sub)
            put_headers(b2_table_start + 2, col_cursor, ['MOIS', '2025', '2026', 'TENDANCE', '% VAR'], col_main)
            for k, w in enumerate([14, 16, 16, 12, 10]):
                ws.column_dimensions[get_column_letter(col_cursor + k)].width = w
            for i, (lbl, v25, v26, tend, pct) in enumerate(rows):
                bg = WHT if i % 2 == 0 else col_sub
                put_data_row(b2_table_start + 3 + i, col_cursor,
                             [lbl, v25, v26 if v26 is not None else '', tend, pct if pct is not None else ''],
                             [None, '#,##0.00', '#,##0.00', None, '0.00%'], bg)
            col_cursor += 5 + TABLE_GAP_COLS
        bloc2_end_row = b2_table_start + 2 + 12

        # BLOC 3 — Top 20 business groups
        bloc3_title_row = bloc2_end_row + BLOC_GAP_ROWS
        c = ws.cell(row=bloc3_title_row, column=2,
                    value="BLOC 3 — Top 20 clients par business group (NOM_GA)")
        c.font = fnt_(WHT, sz=12); c.fill = fill_(DARK)
        ws.merge_cells(start_row=bloc3_title_row, start_column=2, end_row=bloc3_title_row, end_column=23)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=bloc3_title_row + 1, column=2,
                value="Top 20 des business groups par flux et PNB (agrégation NOM_GA).").font = \
            fnt_(DARK, bold=False, sz=9)

        def build_top_clients(df_year, prefix, metric):
            if df_year is None or len(df_year) == 0:
                return []
            col = f'{metric}_{prefix}_ANNEE'
            if col not in df_year.columns:
                return []
            agg = df_year.groupby('NOM_GA', dropna=False).agg(val=(col, 'sum')).reset_index()
            agg = agg[agg['NOM_GA'].astype(str).str.strip() != ''].copy()
            agg = agg[agg['val'].astype(float) > 0]
            total = float(agg['val'].sum()) or 1.0
            agg['pct'] = agg['val'].astype(float) / total
            agg = agg.sort_values(by='val', ascending=False).head(20)
            return list(agg.itertuples(index=False, name=None))

        b3_start_row = bloc3_title_row + 3
        configs_b3_flux = []
        for year in years:
            configs_b3_flux.append(("CCO", year, "Flux", GRN, GRN2))
            configs_b3_flux.append(("CPC", year, "Flux", PUR, PUR2))
        col_cursor = 2
        for prefix, year, metric_label, col_main, col_sub in configs_b3_flux:
            top = build_top_clients(df_year_by_year.get(year), prefix, 'FLUX')
            put_table_title(b3_start_row, col_cursor, f"Top 20 — {metric_label} {prefix} {year}",
                            years_label(year), 4, col_main, col_sub)
            put_headers(b3_start_row + 2, col_cursor, ['RANG', 'BUSINESS GROUP', 'FLUX', '% TOTAL'], col_main)
            for k, w in enumerate([6, 28, 16, 10]):
                ws.column_dimensions[get_column_letter(col_cursor + k)].width = \
                    max(ws.column_dimensions[get_column_letter(col_cursor + k)].width or 0, w)
            for i, row_t in enumerate(top):
                ga, val, pct = row_t
                bg = WHT if i % 2 == 0 else col_sub
                put_data_row(b3_start_row + 3 + i, col_cursor,
                             [i + 1, str(ga), float(val), float(pct)],
                             ['0', None, '#,##0.00', '0.00%'], bg)
            col_cursor += 4 + TABLE_GAP_COLS
        b3_mid_row = b3_start_row + 2 + 20
        b3_row2_start = b3_mid_row + BLOC_GAP_ROWS
        col_cursor = 2
        configs_b3_pnb = []
        for year in years:
            configs_b3_pnb.append(("CCO", year, "PNB", GRN, GRN2))
            configs_b3_pnb.append(("CPC", year, "PNB", PUR, PUR2))
        for prefix, year, metric_label, col_main, col_sub in configs_b3_pnb:
            top = build_top_clients(df_year_by_year.get(year), prefix, 'PNB')
            put_table_title(b3_row2_start, col_cursor, f"Top 20 — {metric_label} {prefix} {year}",
                            years_label(year), 4, col_main, col_sub)
            put_headers(b3_row2_start + 2, col_cursor, ['RANG', 'BUSINESS GROUP', 'PNB', '% TOTAL'], col_main)
            for i, row_t in enumerate(top):
                ga, val, pct = row_t
                bg = WHT if i % 2 == 0 else col_sub
                put_data_row(b3_row2_start + 3 + i, col_cursor,
                             [i + 1, str(ga), float(val), float(pct)],
                             ['0', None, '#,##0.00', '0.00%'], bg)
            col_cursor += 4 + TABLE_GAP_COLS
        bloc3_end_row = b3_row2_start + 2 + 20

        # BLOC 4 — Répartition géographique
        bloc4_title_row = bloc3_end_row + BLOC_GAP_ROWS
        c = ws.cell(row=bloc4_title_row, column=2, value="BLOC 4 — Répartition géographique des flux")
        c.font = fnt_(WHT, sz=12); c.fill = fill_(DARK)
        ws.merge_cells(start_row=bloc4_title_row, start_column=2, end_row=bloc4_title_row, end_column=23)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=bloc4_title_row + 1, column=2,
                value="Répartition complète des flux par pays apporteur (PAYS_FINAL).").font = \
            fnt_(DARK, bold=False, sz=9)
        b4_start_row = bloc4_title_row + 3

        def build_repartition_pays(df_year, prefix):
            if df_year is None or len(df_year) == 0:
                return []
            flux_col = f'FLUX_{prefix}_ANNEE'
            if flux_col not in df_year.columns:
                return []
            agg = df_year.groupby('PAYS_FINAL', dropna=False).agg(flux=(flux_col, 'sum')).reset_index()
            agg = agg[agg['PAYS_FINAL'].astype(str).str.strip() != ''].copy()
            agg = agg[agg['flux'].astype(float) > 0]
            total = float(agg['flux'].sum()) or 1.0
            agg['pct'] = agg['flux'].astype(float) / total
            agg = agg.sort_values(by='flux', ascending=False)
            return list(agg.itertuples(index=False, name=None))

        configs_b4 = []
        for year in years:
            configs_b4.append(("CCO", year, GRN, GRN2))
            configs_b4.append(("CPC", year, PUR, PUR2))
        col_cursor = 2
        for prefix, year, col_main, col_sub in configs_b4:
            rep = build_repartition_pays(df_year_by_year.get(year), prefix)
            put_table_title(b4_start_row, col_cursor, f"Répartition géographique — {prefix} {year}",
                            years_label(year), 3, col_main, col_sub)
            put_headers(b4_start_row + 2, col_cursor, ['PAYS', 'FLUX', '% TOTAL'], col_main)
            for k, w in enumerate([24, 16, 12]):
                ws.column_dimensions[get_column_letter(col_cursor + k)].width = \
                    max(ws.column_dimensions[get_column_letter(col_cursor + k)].width or 0, w)
            for i, row_t in enumerate(rep):
                pays, flux, pct = row_t
                bg = WHT if i % 2 == 0 else col_sub
                put_data_row(b4_start_row + 3 + i, col_cursor,
                             [str(pays), float(flux), float(pct)],
                             [None, '#,##0.00', '0.00%'], bg)
            col_cursor += 3 + TABLE_GAP_COLS

        ws.sheet_view.zoomScale = 90


# ════════════════════════════════════════════════════════════════════════════
# Entry point — CLI argparse autonome
# ════════════════════════════════════════════════════════════════════════════
def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="01.K6BZP",
        description=f"CIB COMMISSION ANALYZER v14 [{VERSION_ID}] — CLI (sans GUI). "
                    "Recoupe WORLDLINE/MONEXT, convertit en EUR, recalcule par année "
                    "et produit un CSV + un classeur XLSX (7 onglets).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    # 8 sources obligatoires
    p.add_argument("--worldline", required=True, type=Path,
                   help="00. Fichier PRGM Worldline (flux CPC).")
    p.add_argument("--monext", required=True, type=Path,
                   help="01. Fichier MONEXT (flux CCO).")
    p.add_argument("--account", required=True, type=Path,
                   help="02. Fichier IBAN_ACCOUNT (pivot 1).")
    p.add_argument("--ref-client", required=True, type=Path,
                   help="03. Fichier REFERENTIEL_CLIENT (pivot 2).")
    p.add_argument("--idseg", required=True, type=Path,
                   help="04. Fichier IDENTIFIANT_SEGMENT (pivot 3).")
    p.add_argument("--devises", required=True, type=Path,
                   help="05. Fichier DEVISES (conversion -> EUR).")
    p.add_argument("--rebate", required=True, type=Path,
                   help="06. Fichier REBATE (Année | RMPM | Plateforme | Montant).")
    p.add_argument("--country", required=True, type=Path,
                   help="10. Fichier COUNTRY (mapping pays Original -> Anglais).")
    # 3 sources optionnelles
    p.add_argument("--override-pays-file", default=None, type=Path,
                   help="07. (opt) Fichier OVERRIDE_PAYS.")
    p.add_argument("--override-pays", default="LUXEMBOURG",
                   help="(opt) Pays de remplacement appliqué via OVERRIDE_PAYS.")
    p.add_argument("--bejo-cartes", default=None, type=Path,
                   help="08. (opt) Fichier BEJO_CARTES (entité | nb).")
    p.add_argument("--bejo-flux", default=None, type=Path,
                   help="09. (opt) Fichier BEJO_FLUX (entité | IBAN).")
    # Sorties obligatoires
    p.add_argument("--output-dir", required=True, type=Path,
                   help="Dossier de sortie (créé si absent).")
    p.add_argument("--output-filename", required=True,
                   help="Nom de base commun CSV + XLSX (sans extension).")
    return p


def main(argv: Optional[List[str]] = None) -> int:
    parser = _build_parser()
    args = parser.parse_args(argv)

    # Validation BEJO : les deux fichiers sont nécessaires si l'un est fourni.
    if bool(args.bejo_cartes) != bool(args.bejo_flux):
        parser.error("BEJO requiert les deux fichiers : --bejo-cartes ET --bejo-flux.")

    try:
        app = CIBCommissionAnalyzer_K6BZP(args)
        app.run()
    except ValueError as e:
        # Erreurs d'entrée (fichier obligatoire manquant, argument invalide).
        print(f"[ERREUR] {e}")
        return 2
    except Exception as e:
        print(f"[ECHEC] Traitement interrompu : {e}")
        return 1

    base = Path(args.output_dir) / args.output_filename
    print(f"[OK] Traitement terminé. Sorties : {base}.csv / {base}.xlsx")
    return 0


if __name__ == "__main__":
    sys.exit(main())