"""
WORLDLINE ANALYZER v16 [V2MRG]
================================
BNP Paribas Cash Management - Direction Monetique.

DESCRIPTION
-----------
Analyseur Worldline dual-source : fusionne REFERENTIEL CLIENT (RC/RIB ->
segment + GA + RMPM) et IDENTIFIANT-SEGMENT (RC/RMPM/RP/CODE_AGENCE -> segment)
en une hierarchie unique pour maximiser la couverture segment, puis cascade
ACCOUNT -> PARC -> MC1/MC2 pour le GA/RMPM, convertit les devises en EUR,
classe ENTREPRISE/BPE, derive la gestion directe/indirecte (YANNICK, shortlist,
CWT) et produit DETAIL CSV + TABLEAU CSV 4 zones geographiques + XLSX agrege.
Version CLI autonome (sans interface graphique). Logique metier identique a la
version GUI : aucun calcul / condition / colonne / plage / hierarchie modifie.

SOURCES REQUISES (CSV)
----------------------
PRGM        Worldline source (programmes, PNB, IBAN, RC, devise) - OBLIGATOIRE
REF_CLIENT  Referentiel client Olivier (RC/RIB -> segment+GA+RMPM)  - OBLIGATOIRE
SEGMENT     Identifiant-segment (TYPE | ID | SEGMENT)               - OBLIGATOIRE
PARC        Parc client (RC/RS/RMPM/GA)                             - OBLIGATOIRE
OPTI        Optiflux (RS / IBAN)                                    - OBLIGATOIRE
YANNICK     Gestion directe/indirecte (GA / Sales)                 - OBLIGATOIRE
DEVISES     Date (YYYYMM) | Devise | Taux                          - OBLIGATOIRE
ACCOUNT     BG-LE-RMPM Account (IBAN / GA / Entite / Pays)         - OBLIGATOIRE
USAGE       Matching usage (ID prog / produit / usage)            - OPTIONNEL
BPE_RETAIL  BPE Retail (code agence)                              - OPTIONNEL
MC1         Matching client 1 (ID prog / produit / GA)            - OPTIONNEL
MC2         Matching client 2 (ID prog / produit / GA)            - OPTIONNEL

OUTPUTS PRODUITS (dans --output-dir, prefixe par --output-filename)
------------------------------------------------------------------
<filename>_DETAIL.csv    Detail ligne a ligne enrichi (sep ';', utf-8-sig)
<filename>_TABLEAU.csv   Tableaux PNB 4 zones geographiques (sep ';')
<filename>_AGREGE.xlsx   Classeur Excel DETAIL + TABLEAU (si openpyxl dispo)

ARGUMENTS CLI
-------------
--prgm PATH (obligatoire) Fichier PRGM Worldline source
--ref-client PATH (obligatoire) Fichier REFERENTIEL CLIENT (Olivier)
--segment PATH (obligatoire) Fichier IDENTIFIANT-SEGMENT
--parc PATH (obligatoire) Fichier PARC_CLIENT
--opti PATH (obligatoire) Fichier OPTIFLUX
--yannick PATH (obligatoire) Fichier YANNICK gestion directe/indirecte
--devises PATH (obligatoire) Fichier DEVISES (Date|Devise|Taux)
--account PATH (obligatoire) Fichier BG-LE-RMPM ACCOUNT
--usage PATH (optionnel) Fichier MATCHING_USAGE
--bpe-retail PATH (optionnel) Fichier BPE RETAIL
--mc1 PATH (optionnel) Fichier MATCHING_CLIENT 1
--mc2 PATH (optionnel) Fichier MATCHING_CLIENT 2
--output-dir PATH (obligatoire) Repertoire de sortie des fichiers produits
--output-filename NAME (obligatoire) Prefixe de nom des fichiers produits (sans extension)

DECOMPOSITION
-------------
1. Chargement et configuration
   1.1 Lecture des arguments CLI et affectation aux attributs lus par le worker
   1.2 Resolution du mapping de colonnes par positions par defaut (presets UI)
2. Traitement (worker, logique inchangee)
   2.1 Conversion devises en EUR (taux par mois, fallback temporel)
   2.2 Construction des dictionnaires REF_CLIENT (RC, RIB)
   2.3 Construction des dictionnaires IDSEG (RC, RMPM, RP, CODE_AGENCE)
   2.4 Construction des dictionnaires ACCOUNT (IBAN, GA, RMPM, pays)
   2.5 Construction PARC / OPTI / YANNICK / USAGE / BPE / MC1 / MC2
   2.6 Preparation PRGM (nettoyage identifiants, PNB total)
   2.7 Etape 1 : resolution segment via REF (RC puis RIB)
   2.8 Etape 2 : resolution segment via IDSEG sur lignes non resolues
   2.9 Etape 3-4 : ACCOUNT via IBAN puis PARC ; OPTI ; MC1/MC2
   2.10 Cascade GA/RMPM (REF -> ACCOUNT -> PARC -> MC) et enrichissement pays
   2.11 YANNICK gestion directe/indirecte, shortlist, CWT
   2.12 Classification ENTREPRISE/BPE et colonnes de sortie + diagnostic
   2.13 Construction des tableaux PNB 4 zones geographiques
   2.14 Export CSV DETAIL + CSV TABLEAU
   2.15 Export XLSX agrege (DETAIL + TABLEAU)
3. Orchestration CLI
   3.1 main() : parse argparse -> instancie -> worker dans try/except
   3.2 Codes de sortie 0 (succes) / 1 (erreur traitement) / 2 (arguments)
"""

import argparse
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

VERSION_ID = "V2MRG"
CWT_PATTERN = "CWT "
GEO_COL_GAP = 2


def get_mois_mapping():
    mapping = {}
    for year in ['2024', '2025', '2026', '2027']:
        mois_names = ['JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN',
                      'JUILLET', 'AOUT', 'SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DECEMBRE']
        for i, mois in enumerate(mois_names, 1):
            mapping[f"{year}{str(i).zfill(2)}"] = f"{year}_{mois}"
    return mapping


MOIS_MAPPING = get_mois_mapping()

DEFAULT_POSITIONS = {
    'prgm_mois': 2, 'prgm_nom_programme': 3, 'prgm_id_prog': 4, 'prgm_produit': 5,
    'prgm_code_agence': 6, 'prgm_rs': 8, 'prgm_iban': 9, 'prgm_rc': 41,
    'parc_rp': 1, 'parc_code_agence': 4, 'parc_rmpm': 6,
    'parc_rs': 8, 'parc_code_ga': 11, 'parc_rc': 14,
    'opti_rs': 4, 'opti_iban': 65,
    'yannick_code_ga': 1, 'yannick_sales': 3, 'yannick_gestion_indirecte': 4,
    'usage_id_prog': 2, 'usage_produit': 3, 'usage_usage': 4,
    'bpe_retail_code_agence': 3,
    'mc1_id_prog': 1, 'mc1_produit': 3, 'mc1_code_ga': 5,
    'mc2_id_prog': 2, 'mc2_produit': 3, 'mc2_code_ga': 5,
    'prgm_devise': 12, 'devise_date': 1, 'devise_code': 2, 'devise_taux': 3,
    'conv_first': 24, 'conv_last': 36,
    'acc_pays_ga': 1, 'acc_code_ga': 2, 'acc_nom_ga': 3,
    'acc_pays_le': 4, 'acc_rmpm': 5, 'acc_nom_le': 6, 'acc_iban': 7,
    'ref_rib': 2, 'ref_rmpm': 4, 'ref_rc': 7,
    'ref_segment': 9, 'ref_code_ga': 10, 'ref_nom_ga': 11,
    'seg_type': 1, 'seg_id': 2, 'seg_segment': 3,
}


class WorldlineAnalyzer_V2MRG:
    def __init__(self, args: argparse.Namespace) -> None:
        self.files = {"PRGM": "", "PARC": "", "OPTI": "", "YANNICK": "",
                      "USAGE": "", "BPE_RETAIL": "", "MC1": "", "MC2": "",
                      "DEVISES": "", "ACCOUNT": "", "REF_CLIENT": "", "SEGMENT": ""}
        self.files["PRGM"] = str(args.prgm)
        self.files["REF_CLIENT"] = str(args.ref_client)
        self.files["SEGMENT"] = str(args.segment)
        self.files["PARC"] = str(args.parc)
        self.files["OPTI"] = str(args.opti)
        self.files["YANNICK"] = str(args.yannick)
        self.files["DEVISES"] = str(args.devises)
        self.files["ACCOUNT"] = str(args.account)
        self.files["USAGE"] = str(args.usage) if args.usage else ""
        self.files["BPE_RETAIL"] = str(args.bpe_retail) if args.bpe_retail else ""
        self.files["MC1"] = str(args.mc1) if args.mc1 else ""
        self.files["MC2"] = str(args.mc2) if args.mc2 else ""

        # Options optionnelles (equivalent des BooleanVar de l'UI) : actives si le
        # fichier correspondant a ete fourni.
        self.use_usage = bool(args.usage)
        self.use_bpe_retail = bool(args.bpe_retail)
        self.use_mc1 = bool(args.mc1)
        self.use_mc2 = bool(args.mc2)
        # Forcage CWT et shortlist produits : comportement par defaut de l'UI
        # (case decochee / aucune selection).
        self.use_cwt = False
        self.produits_shortlist = {}

        self.output_dir = Path(args.output_dir)
        self.output_filename = str(args.output_filename)

    # ── Compat options GUI (BooleanVar.get()) ──────────────────────────────
    class _Flag:
        def __init__(self, value: bool) -> None:
            self._value = value

        def get(self) -> bool:
            return self._value

    @property
    def use_usage_var(self):
        return WorldlineAnalyzer_V2MRG._Flag(self.use_usage)

    @property
    def use_bpe_retail_var(self):
        return WorldlineAnalyzer_V2MRG._Flag(self.use_bpe_retail)

    @property
    def use_mc1_var(self):
        return WorldlineAnalyzer_V2MRG._Flag(self.use_mc1)

    @property
    def use_mc2_var(self):
        return WorldlineAnalyzer_V2MRG._Flag(self.use_mc2)

    @property
    def use_cwt_var(self):
        return WorldlineAnalyzer_V2MRG._Flag(self.use_cwt)

    def load_csv_smart(self, path, nrows=None):
        for sep in [';', ',', '\t']:
            for enc in ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']:
                try:
                    df = pd.read_csv(path, sep=sep, encoding=enc, dtype=str, keep_default_na=False, na_values=[], on_bad_lines='skip', nrows=5)
                    if df.shape[1] > 1: return pd.read_csv(path, sep=sep, encoding=enc, dtype=str, keep_default_na=False, na_values=[], on_bad_lines='skip', nrows=nrows)
                except: continue
        return pd.read_csv(path, sep=None, engine='python', dtype=str, on_bad_lines='skip', nrows=nrows)

    @staticmethod
    def clean_id(series):
        s = series.astype(str).str.strip(); s = s.replace(['', 'nan', 'NAN', 'None', 'NULL', 'NA', 'N/A'], '')
        mask = s.str.startswith('="') & s.str.endswith('"'); s = s.where(~mask, s.str[2:-1]); s = s.str.lstrip("'")
        mask2 = s.str.endswith('.0') & s.str[:-2].str.isdigit(); return s.where(~mask2, s.str[:-2]).str.strip()
    @staticmethod
    def clean_ga(series):
        s = WorldlineAnalyzer_V2MRG.clean_id(series); stripped = s.str.lstrip('0'); return stripped.where(stripped != '', s)
    @staticmethod
    def clean_iban_wl(series):
        s = WorldlineAnalyzer_V2MRG.clean_id(series).str.upper().str.replace(' ', '', regex=False); return s.str[4:].where(s.str.len() > 4, s)
    @staticmethod
    def clean_iban_opti(series): return WorldlineAnalyzer_V2MRG.clean_id(series).str.upper().str.replace(' ', '', regex=False)
    @staticmethod
    def clean_iban_account(series): return WorldlineAnalyzer_V2MRG.clean_id(series).str.upper().str.replace(' ', '', regex=False)
    @staticmethod
    def clean_rib_ref(series): return WorldlineAnalyzer_V2MRG.clean_id(series).str.upper().str.replace(' ', '', regex=False)
    @staticmethod
    def norm_rs(series):
        def n(v):
            if pd.isna(v) or str(v).strip() == "": return ""
            s = unicodedata.normalize('NFD', str(v).strip().upper()); return ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return series.apply(n)
    @staticmethod
    def clean_gest(series):
        s = series.astype(str).str.strip().str.upper(); r = pd.Series('', index=series.index)
        r = r.where(~s.isin(['OUI', 'O', 'YES', 'Y', '1']), 'OUI'); return r.where(~s.isin(['NON', 'N', 'NO', '0']), 'NON')
    @staticmethod
    def to_float(series):
        s = series.astype(str).str.replace(',', '.', regex=False).str.replace(' ', '').str.replace('\xa0', ''); return pd.to_numeric(s, errors='coerce').fillna(0.0)
    @staticmethod
    def pays_to_geo(pays_series):
        NON_TROUVE = {'', 'nan', 'none', 'null', 'non trouvé', 'non trouve', 'pays non trouvé', 'pays non trouve'}
        def _geo(v):
            v = str(v).strip()
            if not v or v.lower() in NON_TROUVE: return 'Pays non trouvé'
            return 'France' if v.upper() in {'FR', 'FRANCE', 'FRA'} else 'Hors France'
        return pays_series.apply(_geo)
    @staticmethod
    def build_segment_dicts(df_seg, col_type, col_id, col_segment):
        seg_by_rc, seg_by_rmpm, seg_by_rp, seg_by_ca = {}, {}, {}, {}
        TYPE_MAP = {'RC': seg_by_rc, 'RMPM': seg_by_rmpm, 'RP': seg_by_rp, 'CODE_AGENCE': seg_by_ca}
        for t, raw_id, seg_val in zip(df_seg[col_type].astype(str).str.strip().str.upper().values, df_seg[col_id].astype(str).str.strip().values, df_seg[col_segment].astype(str).str.strip().str.upper().values):
            if 'ENTREPRISE' in seg_val: segment = 'ENTREPRISE'
            elif 'BPE' in seg_val: segment = 'BPE'
            else: continue
            td = TYPE_MAP.get(t)
            if td is None: continue
            clean = raw_id
            if clean.startswith('="') and clean.endswith('"'): clean = clean[2:-1]
            clean = clean.lstrip("'").strip()
            if clean.endswith('.0') and clean[:-2].isdigit(): clean = clean[:-2]
            if not clean: continue
            if clean not in td: td[clean] = segment
            stripped = clean.lstrip('0')
            if stripped and stripped != clean and stripped not in td: td[stripped] = segment
        return seg_by_rc, seg_by_rmpm, seg_by_rp, seg_by_ca

    def _build_config(self):
        """Reproduit le mapping de colonnes par defaut de l'UI (positions
        preselectionnees dans DEFAULT_POSITIONS / blocs UI), a partir des entetes
        reels des CSV charges. Verification colonnes : deleguee a l'UI web.

        Retourne (m, pnb_cols, devise_cfg) au format exact attendu par worker().
        """
        def numbered(cols):
            return [f"{i+1:02d}. {c}" for i, c in enumerate(cols)]

        def col_at(cols, pos):
            # pos 1-indexee ; comportement UI : selectionne cols[pos-1] si dispo.
            return cols[pos - 1] if pos and pos <= len(cols) else (cols[-1] if cols else "")

        # Entetes (apercu 5 lignes, comme load_previews dans l'UI).
        prgm_cols = list(self.load_csv_smart(self.files["PRGM"], nrows=5).columns)
        ref_cols = list(self.load_csv_smart(self.files["REF_CLIENT"], nrows=5).columns)
        seg_cols = list(self.load_csv_smart(self.files["SEGMENT"], nrows=5).columns)
        parc_cols = list(self.load_csv_smart(self.files["PARC"], nrows=5).columns)
        opti_cols = list(self.load_csv_smart(self.files["OPTI"], nrows=5).columns)
        yan_cols = list(self.load_csv_smart(self.files["YANNICK"], nrows=5).columns)
        dev_cols = list(self.load_csv_smart(self.files["DEVISES"], nrows=5).columns)
        acc_cols = list(self.load_csv_smart(self.files["ACCOUNT"], nrows=5).columns)

        m = {}
        # PRGM (create_map_block PRGM)
        for key in ['prgm_mois', 'prgm_nom_programme', 'prgm_id_prog', 'prgm_produit',
                    'prgm_code_agence', 'prgm_rs', 'prgm_iban', 'prgm_rc']:
            m[key] = col_at(prgm_cols, DEFAULT_POSITIONS[key])
        # REF_CLIENT (create_ref_block)
        for key in ['ref_rib', 'ref_rmpm', 'ref_rc', 'ref_segment', 'ref_code_ga', 'ref_nom_ga']:
            m[key] = col_at(ref_cols, DEFAULT_POSITIONS[key])
        # SEGMENT (create_segment_block)
        for key in ['seg_type', 'seg_id', 'seg_segment']:
            m[key] = col_at(seg_cols, DEFAULT_POSITIONS[key])
        # PARC (create_map_block PARC)
        for key in ['parc_rp', 'parc_code_agence', 'parc_rmpm', 'parc_rs', 'parc_code_ga', 'parc_rc']:
            m[key] = col_at(parc_cols, DEFAULT_POSITIONS[key])
        # OPTI
        for key in ['opti_rs', 'opti_iban']:
            m[key] = col_at(opti_cols, DEFAULT_POSITIONS[key])
        # YANNICK
        for key in ['yannick_code_ga', 'yannick_sales', 'yannick_gestion_indirecte']:
            m[key] = col_at(yan_cols, DEFAULT_POSITIONS[key])
        # ACCOUNT (create_account_block)
        for key in ['acc_pays_ga', 'acc_code_ga', 'acc_nom_ga', 'acc_pays_le',
                    'acc_rmpm', 'acc_nom_le', 'acc_iban']:
            m[key] = col_at(acc_cols, DEFAULT_POSITIONS[key])

        # Optionnels : presents seulement si fichier fourni (comme l'UI).
        if self.use_usage:
            usage_cols = list(self.load_csv_smart(self.files["USAGE"], nrows=5).columns)
            for key in ['usage_id_prog', 'usage_produit', 'usage_usage']:
                m[key] = col_at(usage_cols, DEFAULT_POSITIONS[key])
        if self.use_bpe_retail:
            bpe_cols = list(self.load_csv_smart(self.files["BPE_RETAIL"], nrows=5).columns)
            m['bpe_retail_code_agence'] = col_at(bpe_cols, DEFAULT_POSITIONS['bpe_retail_code_agence'])
        if self.use_mc1:
            mc1_cols = list(self.load_csv_smart(self.files["MC1"], nrows=5).columns)
            for key in ['mc1_id_prog', 'mc1_produit', 'mc1_code_ga']:
                m[key] = col_at(mc1_cols, DEFAULT_POSITIONS[key])
        if self.use_mc2:
            mc2_cols = list(self.load_csv_smart(self.files["MC2"], nrows=5).columns)
            for key in ['mc2_id_prog', 'mc2_produit', 'mc2_code_ga']:
                m[key] = col_at(mc2_cols, DEFAULT_POSITIONS[key])

        # PNB (create_pnb_block) : defaults positions (numbered inclut "(Aucune)"
        # en index 0, donc numbered[defaults[i]] == colonne en position defaults[i]).
        defaults = [36, 35, 34, 33, 32, 31, 30]
        pnb_cols = [prgm_cols[d - 1] for d in defaults if d <= len(prgm_cols)]

        # DEVISES / conversion (create_devise_block)
        devise_cfg = {
            'col_prgm_devise': col_at(prgm_cols, DEFAULT_POSITIONS['prgm_devise']),
            'col_devise_date': col_at(dev_cols, DEFAULT_POSITIONS['devise_date']),
            'col_devise_code': col_at(dev_cols, DEFAULT_POSITIONS['devise_code']),
            'col_devise_taux': col_at(dev_cols, DEFAULT_POSITIONS['devise_taux']),
            'conv_first': DEFAULT_POSITIONS['conv_first'],
            'conv_last': DEFAULT_POSITIONS['conv_last'],
        }
        return m, pnb_cols, devise_cfg

    def _prog(self, value: float, text: str) -> None:
        print(text)

    # ══════════════════════════════════════════════════════════════════════════
    # WORKER — v16 DUAL SOURCE VECTORISÉ
    # ══════════════════════════════════════════════════════════════════════════
    def worker(self, m, pnb_cols, devise_cfg):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self._prog(0.02, "Chargement..."); df = self.load_csv_smart(self.files["PRGM"])

        # ── DEVISES (identique GA14B) ────────────────────────────────────
        self._prog(0.03, "Devises..."); df_devises = self.load_csv_smart(self.files["DEVISES"]); dev_cols = list(df_devises.columns)
        col_dd = devise_cfg["col_devise_date"] or dev_cols[0]; col_dc = devise_cfg["col_devise_code"] or dev_cols[1]; col_dt = devise_cfg["col_devise_taux"] or dev_cols[2]
        dev_dates = df_devises[col_dd].astype(str).str.strip(); dev_codes = df_devises[col_dc].astype(str).str.strip().str.upper()
        dev_taux_num = pd.to_numeric(df_devises[col_dt].astype(str).str.strip().str.replace(",", ".", regex=False), errors="coerce")
        taux_dict = {}; mois_par_devise = {}
        for d, c, t in zip(dev_dates.values, dev_codes.values, dev_taux_num.values):
            if not d or not c or pd.isna(t) or t <= 0: continue
            d_clean = ''.join(ch for ch in d if ch.isdigit())
            if len(d_clean) != 6: continue
            taux_dict[(d_clean, c)] = float(t); mois_par_devise.setdefault(c, set()).add(d_clean)
        mois_par_devise = {dev: sorted(list(s)) for dev, s in mois_par_devise.items()}
        def resolve_taux(mois_val, devise_val):
            if not devise_val or devise_val not in mois_par_devise: return None
            m_clean = ''.join(ch for ch in str(mois_val) if ch.isdigit())
            if len(m_clean) != 6: return None
            key = (m_clean, devise_val)
            if key in taux_dict: return taux_dict[key]
            candidat = None
            for md in mois_par_devise[devise_val]:
                if md <= m_clean: candidat = md
                else: break
            return taux_dict[(candidat, devise_val)] if candidat else None
        prgm_cols = list(df.columns); col_devise_prgm = devise_cfg["col_prgm_devise"] or prgm_cols[DEFAULT_POSITIONS["prgm_devise"] - 1]
        col_mois_prgm = m.get("prgm_mois", "") or prgm_cols[DEFAULT_POSITIONS["prgm_mois"] - 1]
        self._prog(0.05, "Résolution taux..."); cache = {}; n = len(df); arr_taux = np.empty(n, dtype=object)
        arr_d = df[col_devise_prgm].astype(str).str.strip().str.upper().values; arr_m = df[col_mois_prgm].astype(str).str.strip().values
        for i in range(n):
            key = (arr_m[i], arr_d[i])
            if key not in cache: cache[key] = resolve_taux(arr_m[i], arr_d[i])
            arr_taux[i] = cache[key]
        nb_resolus = sum(1 for t in arr_taux if t is not None)
        idx_dev = list(df.columns).index(col_devise_prgm); df.insert(idx_dev + 1, "TAUX_CHANGE", [t if t is not None else "" for t in arr_taux])
        conv_first = devise_cfg["conv_first"] - 1; conv_last = devise_cfg["conv_last"] - 1
        cols_to_convert = [c for c in prgm_cols[conv_first:conv_last + 1] if c in df.columns]
        mask_resolu = np.array([t is not None for t in arr_taux]); taux_num = np.ones(n, dtype=float)
        for i in range(n):
            if arr_taux[i] is not None: taux_num[i] = float(arr_taux[i])
        rename_map = {}
        for col in cols_to_convert:
            vals = pd.to_numeric(df[col].astype(str).str.replace(",", ".", regex=False).str.replace(" ", "", regex=False), errors="coerce").fillna(0.0).values
            vc = vals.copy(); vc[mask_resolu] = np.round(vals[mask_resolu] * taux_num[mask_resolu], 2); df[col] = vc; rename_map[col] = f"{col} (EUR)"
        df.rename(columns=rename_map, inplace=True); pnb_cols = [rename_map.get(c, c) for c in pnb_cols]

        # ── CHARGEMENT ───────────────────────────────────────────────────
        self._prog(0.09, "Chargement..."); df_ref = self.load_csv_smart(self.files["REF_CLIENT"]); df_seg = self.load_csv_smart(self.files["SEGMENT"])
        df_parc = self.load_csv_smart(self.files["PARC"]); df_opti = self.load_csv_smart(self.files["OPTI"])
        df_yan = self.load_csv_smart(self.files["YANNICK"]); df_account = self.load_csv_smart(self.files["ACCOUNT"])
        df_usage = self.load_csv_smart(self.files["USAGE"]) if self.use_usage_var.get() else None
        df_bpe = self.load_csv_smart(self.files["BPE_RETAIL"]) if self.use_bpe_retail_var.get() else None
        df_mc1 = self.load_csv_smart(self.files["MC1"]) if self.use_mc1_var.get() else None
        df_mc2 = self.load_csv_smart(self.files["MC2"]) if self.use_mc2_var.get() else None

        # ── REF CLIENT dicts ─────────────────────────────────────────────
        self._prog(0.11, "REF CLIENT..."); ref_rc_c = self.clean_ga(df_ref[m['ref_rc']]); ref_rc_r = self.clean_id(df_ref[m['ref_rc']])
        ref_rib_c = self.clean_rib_ref(df_ref[m['ref_rib']]); ref_seg = df_ref[m['ref_segment']].astype(str).str.strip().str.upper()
        ref_ga = self.clean_ga(df_ref[m['ref_code_ga']]); ref_nga = df_ref[m['ref_nom_ga']].astype(str).str.strip(); ref_rmpm = self.clean_id(df_ref[m['ref_rmpm']])
        def ns(s):
            if 'ENTREPRISE' in s: return 'ENTREPRISE'
            elif 'BPE' in s: return 'BPE'
            return ''
        ref_sn = ref_seg.apply(ns)
        d_ref_rc = {}
        for rc, rr, seg, ga, nga, rmpm in zip(ref_rc_c.values, ref_rc_r.values, ref_sn.values, ref_ga.values, ref_nga.values, ref_rmpm.values):
            if rr and rr not in d_ref_rc: d_ref_rc[rr] = (seg, ga, nga, rmpm)
            if rc and rc != rr and rc not in d_ref_rc: d_ref_rc[rc] = (seg, ga, nga, rmpm)
        d_ref_rib = {}
        for rib, seg, ga, nga, rmpm in zip(ref_rib_c.values, ref_sn.values, ref_ga.values, ref_nga.values, ref_rmpm.values):
            if rib and rib not in d_ref_rib: d_ref_rib[rib] = (seg, ga, nga, rmpm)
        ref_rc_seg = {k: v[0] for k, v in d_ref_rc.items()}; ref_rc_ga = {k: v[1] for k, v in d_ref_rc.items()}
        ref_rc_nga = {k: v[2] for k, v in d_ref_rc.items()}; ref_rc_rmpm = {k: v[3] for k, v in d_ref_rc.items()}
        ref_rib_seg = {k: v[0] for k, v in d_ref_rib.items()}; ref_rib_ga = {k: v[1] for k, v in d_ref_rib.items()}
        ref_rib_nga = {k: v[2] for k, v in d_ref_rib.items()}; ref_rib_rmpm = {k: v[3] for k, v in d_ref_rib.items()}

        # ── IDSEG dicts ──────────────────────────────────────────────────
        self._prog(0.13, "IDSEG dicts..."); seg_by_rc, seg_by_rmpm, seg_by_rp, seg_by_ca = self.build_segment_dicts(df_seg, m['seg_type'], m['seg_id'], m['seg_segment'])

        # ── ACCOUNT dicts ────────────────────────────────────────────────
        self._prog(0.14, "ACCOUNT..."); df_account['_IBAN'] = self.clean_iban_account(df_account[m['acc_iban']]); df_account['_GA'] = self.clean_ga(df_account[m['acc_code_ga']])
        df_account['_NOM_GA'] = df_account[m['acc_nom_ga']].astype(str).str.strip(); df_account['_PAYS_GA'] = df_account[m['acc_pays_ga']].astype(str).str.strip()
        df_account['_RMPM'] = self.clean_id(df_account[m['acc_rmpm']]); df_account['_NOM_LE'] = df_account[m['acc_nom_le']].astype(str).str.strip(); df_account['_PAYS_LE'] = df_account[m['acc_pays_le']].astype(str).str.strip()
        acc_iban_dict = {}
        for ib, ga, nga, pga, rmpm, nle, ple in zip(df_account['_IBAN'].values, df_account['_GA'].values, df_account['_NOM_GA'].values, df_account['_PAYS_GA'].values, df_account['_RMPM'].values, df_account['_NOM_LE'].values, df_account['_PAYS_LE'].values):
            if ib and ib not in acc_iban_dict: acc_iban_dict[ib] = (ga, nga, pga, rmpm, nle, ple)
        acc_ga2n, acc_ga2p, acc_r2n, acc_r2p = {}, {}, {}, {}
        for ga, nga, pga in zip(df_account['_GA'].values, df_account['_NOM_GA'].values, df_account['_PAYS_GA'].values):
            if ga and ga not in acc_ga2n: acc_ga2n[ga] = nga
            if ga and ga not in acc_ga2p: acc_ga2p[ga] = pga
        for rmpm, nle, ple in zip(df_account['_RMPM'].values, df_account['_NOM_LE'].values, df_account['_PAYS_LE'].values):
            if rmpm and rmpm not in acc_r2n: acc_r2n[rmpm] = nle
            if rmpm and rmpm not in acc_r2p: acc_r2p[rmpm] = ple
        ai_ga = {k: v[0] for k, v in acc_iban_dict.items()}; ai_nga = {k: v[1] for k, v in acc_iban_dict.items()}; ai_pga = {k: v[2] for k, v in acc_iban_dict.items()}
        ai_rmpm = {k: v[3] for k, v in acc_iban_dict.items()}; ai_nle = {k: v[4] for k, v in acc_iban_dict.items()}; ai_ple = {k: v[5] for k, v in acc_iban_dict.items()}

        # ── PARC / OPTI / YANNICK / MC / BPE ────────────────────────────
        self._prog(0.16, "PARC..."); df_parc['_RC'] = self.clean_ga(df_parc[m['parc_rc']]); df_parc['_RS'] = self.norm_rs(df_parc[m['parc_rs']])
        df_parc['_RMPM'] = self.clean_id(df_parc[m['parc_rmpm']]); df_parc['_GA'] = self.clean_ga(df_parc[m['parc_code_ga']])
        drc_r = df_parc[df_parc['_RC']!=''].drop_duplicates('_RC').set_index('_RC')['_RMPM'].to_dict(); drc_g = df_parc[df_parc['_RC']!=''].drop_duplicates('_RC').set_index('_RC')['_GA'].to_dict()
        drs_r = df_parc[df_parc['_RS']!=''].drop_duplicates('_RS').set_index('_RS')['_RMPM'].to_dict(); drs_g = df_parc[df_parc['_RS']!=''].drop_duplicates('_RS').set_index('_RS')['_GA'].to_dict()
        incl = list(df_parc[(df_parc['_RS']!='') & (df_parc['_RS'].str.len()>=14)][['_RS','_RMPM','_GA']].itertuples(index=False, name=None))
        self._prog(0.17, "OPTI..."); df_opti['_IBAN'] = self.clean_iban_opti(df_opti[m['opti_iban']]); df_opti['_RS'] = self.norm_rs(df_opti[m['opti_rs']])
        o_iban = set(df_opti[df_opti['_IBAN']!='']['_IBAN'].unique()); o_rs = set(df_opti[df_opti['_RS']!='']['_RS'].unique())
        o_rs_incl = list(df_opti[df_opti['_RS'].str.len()>=14]['_RS'].unique())
        self._prog(0.18, "YANNICK..."); df_yan['_CODE'] = self.clean_ga(df_yan[m['yannick_code_ga']]); df_yan['_SALES'] = df_yan[m['yannick_sales']].astype(str).str.strip()
        df_yan['_DIR'] = self.clean_gest(df_yan[m['yannick_gestion_indirecte']]) == 'NON'
        y_sales = df_yan[df_yan['_CODE']!=''].drop_duplicates('_CODE').set_index('_CODE')['_SALES'].to_dict()
        y_dir = df_yan[df_yan['_CODE']!=''].drop_duplicates('_CODE').set_index('_CODE')['_DIR'].to_dict()
        u_dict = {}
        if df_usage is not None and m.get('usage_id_prog'):
            df_usage['_K'] = df_usage[m['usage_id_prog']].astype(str).str.strip() + '-' + df_usage[m['usage_produit']].astype(str).str.strip(); df_usage['_U'] = df_usage[m['usage_usage']].astype(str).str.strip()
            u_dict = df_usage[(df_usage['_K']!='-') & (df_usage['_U']!='')].drop_duplicates('_K').set_index('_K')['_U'].to_dict()
        bpe_set = set()
        if df_bpe is not None and m.get('bpe_retail_code_agence'): codes = self.clean_ga(df_bpe[m['bpe_retail_code_agence']]); bpe_set = set(codes[codes!=''].unique())
        mc1_d, mc2_d = {}, {}
        if df_mc1 is not None and m.get('mc1_id_prog'): df_mc1['_K'] = self.clean_id(df_mc1[m['mc1_id_prog']]) + '|' + df_mc1[m['mc1_produit']].astype(str).str.upper(); df_mc1['_G'] = self.clean_ga(df_mc1[m['mc1_code_ga']]); mc1_d = df_mc1[(df_mc1['_K']!='|') & (df_mc1['_G']!='')].drop_duplicates('_K').set_index('_K')['_G'].to_dict()
        if df_mc2 is not None and m.get('mc2_id_prog'): df_mc2['_K'] = self.clean_id(df_mc2[m['mc2_id_prog']]) + '|' + df_mc2[m['mc2_produit']].astype(str).str.upper(); df_mc2['_G'] = self.clean_ga(df_mc2[m['mc2_code_ga']]); mc2_d = df_mc2[(df_mc2['_K']!='|') & (df_mc2['_G']!='')].drop_duplicates('_K').set_index('_K')['_G'].to_dict()

        # ── PRGM PREP ────────────────────────────────────────────────────
        self._prog(0.20, "Prep PRGM...")
        for c in pnb_cols: df[c] = self.to_float(df[c])
        df['PNB_TOTAL'] = df[pnb_cols].sum(axis=1) if pnb_cols else 0.0
        df['_RC'] = self.clean_ga(df[m['prgm_rc']]); df['_RC_RAW'] = self.clean_id(df[m['prgm_rc']])
        df['_RS'] = self.norm_rs(df[m['prgm_rs']]); df['_IBAN'] = self.clean_iban_wl(df[m['prgm_iban']]); df['_IBAN_FULL'] = self.clean_iban_opti(df[m['prgm_iban']])
        df['_CA'] = self.clean_ga(df[m['prgm_code_agence']]); df['_CA_RAW'] = self.clean_id(df[m['prgm_code_agence']])
        df['_MOIS'] = df[m['prgm_mois']].astype(str).str.strip(); df['_ID'] = self.clean_id(df[m['prgm_id_prog']])
        df['_PROD'] = df[m['prgm_produit']].astype(str).str.strip(); df['_PROD_UP'] = df['_PROD'].str.upper()
        df['_NOM'] = df[m['prgm_nom_programme']].astype(str).str.strip(); df['_MC_KEY'] = df['_ID'] + '|' + df['_PROD_UP']

        # ══════════════════════════════════════════════════════════════════
        # ÉTAPE 1 : REF via RC puis RIB
        # ══════════════════════════════════════════════════════════════════
        self._prog(0.23, "REF vectorisé...")
        hrc_raw = df['_RC_RAW'].map(ref_rc_seg); hrc = df['_RC'].map(ref_rc_seg)
        hrc_seg = hrc_raw.where(hrc_raw.notna() & (hrc_raw != ''), hrc)
        f_ref_rc = hrc_seg.notna() & (hrc_seg != '')
        hrib_seg = df['_IBAN'].map(ref_rib_seg)
        f_ref_rib = ~f_ref_rc & hrib_seg.notna() & (hrib_seg != '')
        f_ref = f_ref_rc | f_ref_rib
        seg_from_ref = np.select([f_ref_rc, f_ref_rib], [hrc_seg, hrib_seg], '')
        ga_from_ref = np.select([f_ref_rc, f_ref_rib], [df['_RC_RAW'].map(ref_rc_ga).where(hrc_raw.notna(), df['_RC'].map(ref_rc_ga)), df['_IBAN'].map(ref_rib_ga)], '')
        nga_from_ref = np.select([f_ref_rc, f_ref_rib], [df['_RC_RAW'].map(ref_rc_nga).where(hrc_raw.notna(), df['_RC'].map(ref_rc_nga)), df['_IBAN'].map(ref_rib_nga)], '')
        rmpm_from_ref = np.select([f_ref_rc, f_ref_rib], [df['_RC_RAW'].map(ref_rc_rmpm).where(hrc_raw.notna(), df['_RC'].map(ref_rc_rmpm)), df['_IBAN'].map(ref_rib_rmpm)], '')
        ref_source = np.select([f_ref_rc, f_ref_rib], ['REF_RC', 'REF_IBAN'], 'NONE')
        self._prog(0.26, f"REF: RC={int(f_ref_rc.sum()):,} RIB={int(f_ref_rib.sum()):,}")

        # ══════════════════════════════════════════════════════════════════
        # ÉTAPE 2 : IDSEG sur lignes non résolues par REF
        # ══════════════════════════════════════════════════════════════════
        self._prog(0.28, "IDSEG vectorisé...")
        need_idseg = ~f_ref
        irc_raw = df['_RC_RAW'].map(seg_by_rc); irc = df['_RC'].map(seg_by_rc); irc_hit = irc_raw.fillna(irc)
        f_irc = need_idseg & irc_hit.notna()
        rmpm_s = pd.Series(rmpm_from_ref, index=df.index).fillna('')
        irmpm = rmpm_s.map(seg_by_rmpm); irmpm2 = rmpm_s.str.lstrip('0').where(rmpm_s.str.lstrip('0')!='', rmpm_s).map(seg_by_rmpm)
        irmpm_hit = irmpm.fillna(irmpm2); f_irmpm = need_idseg & ~f_irc & irmpm_hit.notna()
        ica_raw = df['_CA_RAW'].map(seg_by_ca); ica = df['_CA'].map(seg_by_ca); ica_hit = ica_raw.fillna(ica)
        f_ica = need_idseg & ~f_irc & ~f_irmpm & ica_hit.notna()
        f_idseg = f_irc | f_irmpm | f_ica
        seg_from_idseg = np.select([f_irc, f_irmpm, f_ica], [irc_hit, irmpm_hit, ica_hit], '')
        idseg_source = np.select([f_irc, f_irmpm, f_ica], ['IDSEG_RC', 'IDSEG_RMPM', 'IDSEG_CA'], 'NONE')
        self._prog(0.30, f"IDSEG: RC={int(f_irc.sum()):,} RMPM={int(f_irmpm.sum()):,} CA={int(f_ica.sum()):,}")

        # Segment combiné (REF > IDSEG)
        f_seg_any = f_ref | f_idseg
        segment_raw = pd.Series(seg_from_ref, index=df.index).fillna('')
        segment_raw = segment_raw.where(segment_raw != '', pd.Series(seg_from_idseg, index=df.index).fillna(''))
        segment_source = np.select([f_ref_rc, f_ref_rib, f_irc, f_irmpm, f_ica], ['REF_RC', 'REF_IBAN', 'IDSEG_RC', 'IDSEG_RMPM', 'IDSEG_CA'], 'FALLBACK')

        # ══════════════════════════════════════════════════════════════════
        # ÉTAPE 3 : ACCOUNT via IBAN
        # ══════════════════════════════════════════════════════════════════
        self._prog(0.32, "ACCOUNT...")
        hit_ib = df['_IBAN'].map(ai_ga); hit_ibf = df['_IBAN_FULL'].map(ai_ga)
        fa = hit_ib.notna() | hit_ibf.notna(); fa_t = hit_ib.notna()
        acc_ga = hit_ib.where(fa_t, hit_ibf); acc_rmpm = df['_IBAN'].map(ai_rmpm).where(fa_t, df['_IBAN_FULL'].map(ai_rmpm))
        acc_nga = df['_IBAN'].map(ai_nga).where(fa_t, df['_IBAN_FULL'].map(ai_nga))
        acc_pga = df['_IBAN'].map(ai_pga).where(fa_t, df['_IBAN_FULL'].map(ai_pga))
        acc_nle = df['_IBAN'].map(ai_nle).where(fa_t, df['_IBAN_FULL'].map(ai_nle))
        acc_ple = df['_IBAN'].map(ai_ple).where(fa_t, df['_IBAN_FULL'].map(ai_ple))
        df['FOUND_ACCOUNT'] = np.where(fa, 'YES', 'NO')

        # ══════════════════════════════════════════════════════════════════
        # ÉTAPE 4 : PARC
        # ══════════════════════════════════════════════════════════════════
        self._prog(0.34, "PARC...")
        need_parc = ~f_ref & ~fa
        prc_r = df['_RC'].map(drc_r); prc_g = df['_RC'].map(drc_g); fp_rc = need_parc & prc_r.notna()
        prs_r = df['_RS'].map(drs_r); prs_g = df['_RS'].map(drs_g); fp_rs = need_parc & ~fp_rc & prs_r.notna()
        fp_ri = pd.Series(False, index=df.index); pi_r = pd.Series('', index=df.index); pi_g = pd.Series('', index=df.index)
        if incl:
            mn = need_parc & ~fp_rc & ~fp_rs & (df['_RS']!='') & (df['_RS'].str.len()>=14); idxs = df.index[mn]
            if len(idxs) > 0:
                rs_v = df.loc[idxs, '_RS'].values
                for j, idx in enumerate(idxs):
                    for (prs, prm, pg) in incl:
                        if rs_v[j] in prs: fp_ri.iloc[idx] = True; pi_r.iloc[idx] = prm; pi_g.iloc[idx] = pg; break
        fp = fp_rc | fp_rs | fp_ri
        df['FOUND_PARC'] = np.where(fp, 'YES', 'NO')

        # OPTIFLUX
        self._prog(0.36, "OPTI...")
        fo_ib = df['_IBAN'].isin(o_iban) & (df['_IBAN']!=''); fo_rs = ~fo_ib & df['_RS'].isin(o_rs) & (df['_RS']!='')
        fo_ri = pd.Series(False, index=df.index)
        if o_rs_incl:
            mn = ~fo_ib & ~fo_rs & (df['_RS']!='') & (df['_RS'].str.len()>=14); idxs = df.index[mn]
            if len(idxs) > 0:
                rs_v = df.loc[idxs, '_RS'].values
                for j, idx in enumerate(idxs):
                    for o in o_rs_incl:
                        if rs_v[j] in o: fo_ri.iloc[idx] = True; break
        fo = fo_ib | fo_rs | fo_ri; df['FOUND_OPTIFLUX'] = np.where(fo, 'YES', 'NO')

        # MC1/MC2
        self._prog(0.38, "MC..."); need_mc = ~f_ref & ~fa & ~fp
        mc1_h = df['_MC_KEY'].map(mc1_d) if mc1_d else pd.Series(np.nan, index=df.index); fmc1 = need_mc & mc1_h.notna()
        mc2_h = df['_MC_KEY'].map(mc2_d) if mc2_d else pd.Series(np.nan, index=df.index); fmc2 = need_mc & mc2_h.notna()
        df['FOUND_MC1'] = np.where(fmc1, 'YES', 'NO'); df['FOUND_MC2'] = np.where(fmc2, 'YES', 'NO')
        df['FOUND_BPE_RETAIL'] = np.where(df['_CA'].isin(bpe_set) & (df['_CA']!=''), 'YES', 'NO') if bpe_set else 'NO'

        # ══════════════════════════════════════════════════════════════════
        # GA/RMPM CASCADÉ : REF → ACCOUNT → PARC → MC
        # ══════════════════════════════════════════════════════════════════
        self._prog(0.40, "GA cascadé...")
        ga = pd.Series(ga_from_ref, index=df.index).fillna(''); rmpm = pd.Series(rmpm_from_ref, index=df.index).fillna(''); nom_ga = pd.Series(nga_from_ref, index=df.index).fillna('')
        ga = ga.where(ga!='', acc_ga.fillna('')); rmpm = rmpm.where(rmpm!='', acc_rmpm.fillna('')); nom_ga = nom_ga.where(nom_ga!='', acc_nga.fillna(''))
        m_prc = (ga=='') & fp_rc; ga = ga.where(~m_prc, prc_g); rmpm = rmpm.where(~m_prc, prc_r)
        m_prs = (ga=='') & fp_rs; ga = ga.where(~m_prs, prs_g); rmpm = rmpm.where(~m_prs, prs_r)
        m_pri = (ga=='') & fp_ri; ga = ga.where(~m_pri, pi_g); rmpm = rmpm.where(~m_pri, pi_r)
        m_mc1 = (ga=='') & fmc1; ga = np.where(m_mc1, mc1_h.fillna(''), ga)
        m_mc2 = (pd.Series(ga, index=df.index)=='') & fmc2; ga = np.where(m_mc2, mc2_h.fillna(''), ga)
        ga = pd.Series(ga, index=df.index).fillna(''); rmpm = pd.Series(rmpm, index=df.index).fillna('')
        ga_source = pd.Series('NON_TROUVE', index=df.index)
        ga_source = np.where(f_ref_rc, 'REF_RC', ga_source); ga_source = np.where(f_ref_rib & (ga_source=='NON_TROUVE'), 'REF_IBAN', ga_source)
        ga_source = np.where(fa & (ga_source=='NON_TROUVE'), 'ACCOUNT_IBAN', ga_source)
        ga_source = np.where(fp_rc & (ga_source=='NON_TROUVE'), 'PARC_RC', ga_source); ga_source = np.where(fp_rs & (ga_source=='NON_TROUVE'), 'PARC_RS', ga_source)
        ga_source = np.where(fp_ri & (ga_source=='NON_TROUVE'), 'PARC_RS_INCL', ga_source)
        ga_source = np.where(m_mc1, 'MC1', ga_source); ga_source = np.where(m_mc2, 'MC2', ga_source)
        ga_source = pd.Series(ga_source, index=df.index)

        # ENRICHISSEMENT PAYS
        self._prog(0.43, "Pays...")
        pays_ga = acc_pga.fillna(''); pays_le = acc_ple.fillna(''); nom_le = acc_nle.fillna('')
        pays_ga = pays_ga.where(pays_ga!='', ga.map(acc_ga2p).fillna('')); pays_le = pays_le.where(pays_le!='', rmpm.map(acc_r2p).fillna(''))
        nom_le = nom_le.where(nom_le!='', rmpm.map(acc_r2n).fillna('')); nom_ga = nom_ga.where(nom_ga!='', ga.map(acc_ga2n).fillna(''))

        # YANNICK
        self._prog(0.45, "YANNICK..."); yd_ga = ga.map(y_dir); yd_rmpm = rmpm.map(y_dir); fyan = yd_ga.notna() | yd_rmpm.notna()
        df['FOUND_YANNICK'] = np.where(fyan, 'YES', 'NO')
        sales = ga.map(y_sales).fillna('').where(yd_ga.notna(), rmpm.map(y_sales).fillna('')).fillna('N/A')
        is_dir = yd_ga.where(yd_ga.notna(), yd_rmpm).fillna(False)

        # ══════════════════════════════════════════════════════════════════
        # CLASSIFICATION : seg_any → direct, sinon PARC/OPTI combo
        # ══════════════════════════════════════════════════════════════════
        self._prog(0.47, "Classification...")
        seg_is_ent = segment_raw == 'ENTREPRISE'; seg_is_bpe = segment_raw == 'BPE'
        fb = ~f_seg_any; tc = fa | fp
        fb_ent_c = fb & tc & ~fo; fb_bpe_o = fb & fo; fb_ent_mc1 = fb & ~tc & ~fo & fmc1; fb_ent_mc2 = fb & ~tc & ~fo & ~fmc1 & fmc2
        fb_bpe_r = fb & ~tc & ~fo & ~fmc1 & ~fmc2 & (df['FOUND_BPE_RETAIL']=='YES')
        fb_ent_d = fb & ~fb_bpe_o & ~fb_bpe_r & ~fb_ent_c & ~fb_ent_mc1 & ~fb_ent_mc2
        is_ent = seg_is_ent | fb_ent_c | fb_ent_mc1 | fb_ent_mc2 | fb_ent_d
        is_bpe = seg_is_bpe | fb_bpe_o | fb_bpe_r
        class_source = np.select([f_ref_rc & seg_is_ent, f_ref_rc & seg_is_bpe, f_ref_rib & seg_is_ent, f_ref_rib & seg_is_bpe, f_irc, f_irmpm, f_ica,
            fb_ent_c & fa & ~fp, fb_ent_c & fp, fb_bpe_o & tc, fb_bpe_o & ~tc, fb_ent_mc1, fb_ent_mc2, fb_bpe_r, fb_ent_d],
            ['REF_RC', 'REF_RC', 'REF_IBAN', 'REF_IBAN', 'IDSEG_RC', 'IDSEG_RMPM', 'IDSEG_CA',
             'ACCOUNT', 'PARC', 'CLIENT_OPTI', 'OPTI', 'MC1', 'MC2', 'BPE_RETAIL', 'ENT_DEF'], 'DEFAUT')
        df['ENTREPRISE'] = np.where(is_ent, 'YES', 'NO'); df['BPE'] = np.where(is_bpe, 'YES', 'NO')
        df['CLASS_SOURCE'] = class_source; df['REF_SOURCE'] = ref_source; df['SEGMENT_RAW'] = segment_raw; df['SEGMENT_SOURCE'] = segment_source

        # GESTION D/I
        self._prog(0.50, "Gestion...")
        gd = np.where(fyan, np.where(is_dir, 'YES', 'NO'), 'NO'); gs = np.where(fyan, 'YANNICK', 'DEFAUT')
        usage_key = df['_ID'] + '-' + df['_PROD']; usage_val = usage_key.map(u_dict).fillna('N/A'); pou = usage_val.where(usage_val != 'N/A', df['_PROD'])
        if self.produits_shortlist:
            for prod, gt in self.produits_shortlist.items():
                ms = pou == prod; gd = np.where(ms, 'YES' if gt == "GESTION_DIRECTE" else 'NO', gd); gs = np.where(ms, 'SHORTLIST', gs)
        cwt_cnt = 0
        if self.use_cwt_var.get():
            mc = df['_NOM'].str.contains(CWT_PATTERN, na=False); cwt_cnt = int(mc.sum()); gd = np.where(mc, 'YES', gd); gs = np.where(mc, 'CWT', gs)
        gi = np.where(pd.Series(gd) == 'YES', 'NO', 'YES')
        df['GESTION_DIRECTE'] = gd; df['GESTION_INDIRECTE'] = gi; df['GEST_SOURCE'] = gs; df['SALES'] = sales

        # COLONNES OUTPUT
        self._prog(0.53, "Output...")
        df['CODE_GA'] = ga; df['NOM_GA'] = nom_ga; df['RMPM'] = rmpm; df['NOM_ENTITE_JURIDIQUE'] = nom_le
        df['PAYS_GA'] = pays_ga.where(pays_ga!='', 'Pays non trouvé'); df['PAYS_ENTITE_JURIDIQUE'] = pays_le.where(pays_le!='', 'Pays non trouvé')
        df['GEO_GA'] = self.pays_to_geo(df['PAYS_GA']); df['GEO_ENTITE_JURIDIQUE'] = self.pays_to_geo(df['PAYS_ENTITE_JURIDIQUE'])
        df['PRODUIT'] = df['_PROD']; df['USAGE'] = usage_val; df['PRODUIT_OU_USAGE'] = pou; df['GA_SOURCE'] = ga_source

        # DIAGNOSTIC
        self._prog(0.55, "Diagnostic...")
        tr_rc = np.where(f_ref_rc, '✓', np.where(df['_RC']!='', '✗', '∅')); tr_rib = np.where(f_ref_rib, '✓', np.where(df['_IBAN']!='', '✗', '∅'))
        ti_rc = np.where(f_irc, '✓', '·'); ti_rmpm = np.where(f_irmpm, '✓', '·'); ti_ca = np.where(f_ica, '✓', '·')
        ta = np.where(fa, '✓', np.where((df['_IBAN']!='')|(df['_IBAN_FULL']!=''), '✗', '∅'))
        diag = ('REF[RC'+pd.Series(tr_rc, index=df.index)+',RIB'+pd.Series(tr_rib, index=df.index)+']'
            +' | IDSEG[RC'+pd.Series(ti_rc, index=df.index)+',RMPM'+pd.Series(ti_rmpm, index=df.index)+',CA'+pd.Series(ti_ca, index=df.index)+']'
            +' | ACC['+pd.Series(ta, index=df.index)+']'
            +' | Class='+pd.Series(class_source, index=df.index)
            +' | GA='+ga_source.astype(str)
            +' | SEG='+pd.Series(segment_source, index=df.index))
        df['DIAGNOSTIC'] = diag
        df.drop(columns=['_RC','_RC_RAW','_RS','_IBAN','_IBAN_FULL','_CA','_CA_RAW','_ID','_PROD','_PROD_UP','_NOM','_MC_KEY','_MOIS'], inplace=True, errors='ignore')

        # ══════════════════════════════════════════════════════════════════
        # TABLEAUX PNB 4 ZONES (identique GA14B)
        # ══════════════════════════════════════════════════════════════════
        self._prog(0.60, "Tableaux PNB...")
        df['_MT'] = df[m['prgm_mois']].astype(str).str.strip() if m['prgm_mois'] in df.columns else ''
        ml = sorted([x for x in df['_MT'].unique() if x])
        da = df.groupby(['_MT','PRODUIT_OU_USAGE','ENTREPRISE','BPE','GESTION_DIRECTE','GESTION_INDIRECTE','GEO_ENTITE_JURIDIQUE'])['PNB_TOTAL'].sum().reset_index()
        ape = sorted([p for p in da[da['ENTREPRISE']=='YES']['PRODUIT_OU_USAGE'].unique() if p and p!='N/A'])
        apb = sorted([p for p in da[da['BPE']=='YES']['PRODUIT_OU_USAGE'].unique() if p and p!='N/A'])
        def mkp(ds, ap):
            rows = []; tm = {MOIS_MAPPING.get(mc, mc): 0.0 for mc in ml}
            for prod in ap:
                dp = ds[ds['PRODUIT_OU_USAGE']==prod]; row = {'': prod}
                for mc in ml: val = dp[dp['_MT']==mc]['PNB_TOTAL'].sum(); cl = MOIS_MAPPING.get(mc, mc); row[cl] = val; tm[cl] += val
                rows.append(row)
                rgd = {'': '  |__ GESTION_DIRECTE'}
                for mc in ml: rgd[MOIS_MAPPING.get(mc, mc)] = dp[dp['GESTION_DIRECTE']=='YES'][dp['_MT']==mc]['PNB_TOTAL'].sum()
                rows.append(rgd)
                rgi = {'': '  |__ GESTION_INDIRECTE'}
                for mc in ml: rgi[MOIS_MAPPING.get(mc, mc)] = dp[dp['GESTION_INDIRECTE']=='YES'][dp['_MT']==mc]['PNB_TOTAL'].sum()
                rows.append(rgi)
            rtm = {'': 'TOTAL_MOIS'}; rtm.update(tm); rows.append(rtm)
            gt = sum(tm.values()); rt = {'': 'TOTAL'}
            for mc in ml: rt[MOIS_MAPPING.get(mc, mc)] = ''
            if ml: rt[MOIS_MAPPING.get(ml[0], ml[0])] = gt
            rows.append(rt); return pd.DataFrame(rows)
        mf = da['GEO_ENTITE_JURIDIQUE']=='France'; mhf = da['GEO_ENTITE_JURIDIQUE']=='Hors France'; mnt = da['GEO_ENTITE_JURIDIQUE']=='Pays non trouvé'
        ae = da[da['ENTREPRISE']=='YES']; ab = da[da['BPE']=='YES']
        peg = mkp(ae, ape); pef = mkp(ae[mf], ape); peh = mkp(ae[mhf], ape); pen = mkp(ae[mnt], ape)
        pbg = mkp(ab, apb); pbf = mkp(ab[mf], apb); pbh = mkp(ab[mhf], apb); pbn = mkp(ab[mnt], apb)
        apg = sorted([p for p in df['PRODUIT_OU_USAGE'].unique() if p and p!='N/A'])
        def mkpg(ds):
            t = ds.groupby('PRODUIT_OU_USAGE')['PNB_TOTAL'].sum().reindex(apg, fill_value=0.0).reset_index(); t.columns = ['PRODUIT', 'PNB_TOTAL']
            t = t.sort_values('PNB_TOTAL', ascending=False); return pd.concat([t, pd.DataFrame([{'PRODUIT': 'TOTAL', 'PNB_TOTAL': t['PNB_TOTAL'].sum()}])], ignore_index=True)
        def mkgg(ds):
            gv = ds[ds['GESTION_DIRECTE']=='YES']['PNB_TOTAL'].sum(); iv = ds[ds['GESTION_INDIRECTE']=='YES']['PNB_TOTAL'].sum()
            return pd.DataFrame({'GESTION': ['DIRECTE', 'INDIRECTE', 'TOTAL'], 'PNB_TOTAL': [gv, iv, gv+iv]})
        dpf = df[df['GEO_ENTITE_JURIDIQUE']=='France']; dph = df[df['GEO_ENTITE_JURIDIQUE']=='Hors France']; dpn = df[df['GEO_ENTITE_JURIDIQUE']=='Pays non trouvé']
        ag = mkpg(df); af = mkpg(dpf); ah = mkpg(dph); an = mkpg(dpn)
        gg = mkgg(df); gf = mkgg(dpf); gh = mkgg(dph); gn = mkgg(dpn)
        df.drop(columns=['_MT'], inplace=True, errors='ignore')

        # EXPORTS CSV
        self._prog(0.75, "Export DETAIL...")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        save_det = str(self.output_dir / f"{self.output_filename}_DETAIL.csv")
        df.to_csv(save_det, sep=';', index=False, encoding='utf-8-sig', decimal=',')
        self._prog(0.80, "Export TABLEAU...")
        save_tab = str(self.output_dir / f"{self.output_filename}_TABLEAU.csv")
        def ff(d):
            d2 = d.copy()
            for c in d2.columns:
                if c!='': d2[c] = d2[c].apply(lambda x: f"{x:.2f}".replace('.', ',') if isinstance(x, (int, float)) else x)
            return d2
        with open(save_tab, 'w', encoding='utf-8-sig') as f:
            for gl, pe, pb, pp, pg in [("GLOBAL",peg,pbg,ag,gg),("FRANCE",pef,pbf,af,gf),("HORS FRANCE",peh,pbh,ah,gh),("PAYS NON TROUVE",pen,pbn,an,gn)]:
                f.write(f"=== {gl} ===\n\n--- ENTREPRISE ---\n"); ff(pe).to_csv(f, sep=';', index=False, lineterminator='\n')
                f.write("\n--- BPE ---\n"); ff(pb).to_csv(f, sep=';', index=False, lineterminator='\n')
                f.write("\n--- SOMME PAR PRODUIT ---\n"); tp=pp.copy(); tp['PNB_TOTAL']=tp['PNB_TOTAL'].apply(lambda x: f"{x:.2f}".replace('.',',')); tp.to_csv(f, sep=';', index=False, lineterminator='\n')
                f.write("\n--- SOMME PAR GESTION ---\n"); tg=pg.copy(); tg['PNB_TOTAL']=tg['PNB_TOTAL'].apply(lambda x: f"{x:.2f}".replace('.',',')); tg.to_csv(f, sep=';', index=False, lineterminator='\n')
                f.write("\n\n")
        self._prog(0.85, "Terminé !")
        nrc = int(f_ref_rc.sum()); nrib = int(f_ref_rib.sum()); nirc = int(f_irc.sum()); nirmpm = int(f_irmpm.sum()); nica = int(f_ica.sum()); nfb = int((~f_seg_any).sum())
        print(f"""Fichiers CSV générés !

• Total : {n:,} | Taux : {nb_resolus:,}
• ACCOUNT : {int(fa.sum()):,} | PARC : {int(fp.sum()):,} | OPTI : {int(fo.sum()):,}
• ENT : {int(is_ent.sum()):,} | BPE : {int(is_bpe.sum()):,}
• GD : {int((pd.Series(gd)=='YES').sum()):,} | GI : {int((pd.Series(gi)=='YES').sum()):,}
• PNB : {df['PNB_TOTAL'].sum():,.2f} EUR
{f"• CWT : {cwt_cnt:,}" if cwt_cnt else ""}

── Sources segment (v16 dual) ──
• REF_RC   : {nrc:,}
• REF_IBAN : {nrib:,}
• IDSEG_RC : {nirc:,}
• IDSEG_RMPM: {nirmpm:,}
• IDSEG_CA : {nica:,}
• FALLBACK : {nfb:,}""")
        # Generation XLSX (toujours, equivalent du "Oui" par defaut de l'UI).
        self.create_xlsx(save_det, ts, peg, pef, peh, pen, pbg, pbf, pbh, pbn, ag, af, ah, an, gg, gf, gh, gn, ml)

    def create_xlsx(self, det_path, ts, peg, pef, peh, pen, pbg, pbf, pbh, pbn, ag, af, ah, an, gg, gf, gh, gn, ml):
        if not OPENPYXL_AVAILABLE: print("[ERREUR] openpyxl manquant"); return
        self._prog(0.5, "XLSX...")
        save_x = str(self.output_dir / f"{self.output_filename}_AGREGE.xlsx")
        wb = Workbook(); ws1 = wb.active; ws1.title = "DETAIL"; ws1.column_dimensions['A'].width = 2.5
        ws1['B4'] = f"WORLDLINE - Analyse détaillée - {datetime.now().strftime('%d/%m/%Y')}"; ws1['B4'].font = Font(bold=True, size=14)
        df_d = pd.read_csv(det_path, sep=';', encoding='utf-8-sig', dtype=str, keep_default_na=False)
        nc = {c for c in df_d.columns if c=='PNB_TOTAL' or c=='TAUX_CHANGE' or c.endswith('(EUR)')}
        hf = PatternFill("solid", fgColor="00915A"); ht = Font(bold=True, color="FFFFFF", size=10)
        for ci, cn in enumerate(df_d.columns, start=2): c = ws1.cell(row=6, column=ci, value=cn); c.font = ht; c.fill = hf; c.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(df_d.itertuples(index=False), start=7):
            for ci, (cn, val) in enumerate(zip(df_d.columns, row), start=2):
                if cn in nc:
                    try:
                        sv = str(val).strip()
                        if sv=='' or sv.lower() in ('nan','none'): ws1.cell(row=ri, column=ci, value="")
                        else: ws1.cell(row=ri, column=ci, value=float(sv.replace(',','.')))
                    except: ws1.cell(row=ri, column=ci, value=val)
                else: ws1.cell(row=ri, column=ci, value=val)
        for ci in range(2, len(df_d.columns)+2): ws1.column_dimensions[get_column_letter(ci)].width = 15
        ws2 = wb.create_sheet(title="TABLEAU"); ws2.column_dimensions['A'].width = 2.5
        CG,CF,CH,CN = "006B43","1565C0","B71C1C","6A1B9A"; SG,SF,SH,SN = "E8F5E9","E3F2FD","FFEBEE","F3E5F5"
        nm = len(ml); bdw = max(1+nm, 2)
        def wsh(ws, r, c, t, bg):
            cell = ws.cell(row=r, column=c, value=t); cell.font = Font(bold=True, color="FFFFFF", size=11); cell.fill = PatternFill("solid", fgColor=bg)
            for cc in range(c+1, c+bdw): ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=bg)
        def wdb(ws, sr, sc, dff, hc, sbc):
            if dff is None or len(dff)==0: ws.cell(row=sr, column=sc, value="(vide)"); return 2
            r = sr
            for ci, cn in enumerate(dff.columns): c = ws.cell(row=r, column=sc+ci, value=cn); c.font = Font(bold=True, color="FFFFFF", size=9); c.fill = PatternFill("solid", fgColor=hc)
            r += 1
            for _, rd in dff.iterrows():
                for ci, (cn, val) in enumerate(rd.items()):
                    cell = ws.cell(row=r, column=sc+ci, value=val)
                    if not str(rd.iloc[0]).startswith("  |__"): cell.fill = PatternFill("solid", fgColor=sbc)
                    if isinstance(val, (int, float)) and cn!='': cell.number_format = '#,##0.00'
                r += 1
            return r - sr
        ws2['B4'] = f"WORLDLINE - Tableaux PNB - {datetime.now().strftime('%d/%m/%Y')}"; ws2['B4'].font = Font(bold=True, size=14)
        SRH = 6; cg = 2; cf = cg+bdw+GEO_COL_GAP; ch = cf+bdw+GEO_COL_GAP; cn_ = ch+bdw+GEO_COL_GAP
        for cs, t, co in [(cg,"GLOBAL",CG),(cf,"FRANCE",CF),(ch,"HORS FRANCE",CH),(cn_,"PAYS NON TROUVÉ",CN)]:
            cell = ws2.cell(row=SRH, column=cs, value=t); cell.font = Font(bold=True, color="FFFFFF", size=13); cell.fill = PatternFill("solid", fgColor=co)
            for cc in range(cs+1, cs+bdw): ws2.cell(row=SRH, column=cc).fill = PatternFill("solid", fgColor=co)
        SC = {"ENTREPRISE": ("004D33",SG,"0D47A1",SF,"7F0000",SH,"4A148C",SN), "BPE": ("005A3C",SG,"1565C0",SF,"8B0000",SH,"6A1B9A",SN), "PRODUITS": ("006B43",SG,"1976D2",SF,"B71C1C",SH,"8E24AA",SN), "GESTION": ("00513A",SG,"1A237E",SF,"880E4F",SH,"5E35B1",SN)}
        cr = SRH + 2
        for sn, dq in [("ENTREPRISE",[peg,pef,peh,pen]),("BPE",[pbg,pbf,pbh,pbn]),("PRODUITS",[ag,af,ah,an]),("GESTION",[gg,gf,gh,gn])]:
            sc = SC[sn]; hcg,sg_,hcf,sf_,hch,sh_,hcn,sn_ = sc
            wsh(ws2,cr,cg,f"  {sn}",CG); wsh(ws2,cr,cf,f"  {sn}",CF); wsh(ws2,cr,ch,f"  {sn}",CH); wsh(ws2,cr,cn_,f"  {sn}",CN); cr += 1
            ru = [wdb(ws2,cr,cg,dq[0],hcg,sg_),wdb(ws2,cr,cf,dq[1],hcf,sf_),wdb(ws2,cr,ch,dq[2],hch,sh_),wdb(ws2,cr,cn_,dq[3],hcn,sn_)]; cr += max(ru) + 2
        for sc in [cg,cf,ch,cn_]:
            ws2.column_dimensions[get_column_letter(sc)].width = 32
            for ci in range(1, bdw): ws2.column_dimensions[get_column_letter(sc+ci)].width = 14
        wb.save(save_x); self._prog(1.0, "Terminé !")
        print(f"Fichier XLSX créé : {Path(save_x).name}")


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="01.V2MRG",
        description="Worldline Analyzer v16 [V2MRG] - CLI autonome (dual source REF + IDSEG).",
    )
    p.add_argument("--prgm", type=Path, required=True, help="Fichier PRGM Worldline source (obligatoire)")
    p.add_argument("--ref-client", type=Path, required=True, help="Fichier REFERENTIEL CLIENT / Olivier (obligatoire)")
    p.add_argument("--segment", type=Path, required=True, help="Fichier IDENTIFIANT-SEGMENT (obligatoire)")
    p.add_argument("--parc", type=Path, required=True, help="Fichier PARC_CLIENT (obligatoire)")
    p.add_argument("--opti", type=Path, required=True, help="Fichier OPTIFLUX (obligatoire)")
    p.add_argument("--yannick", type=Path, required=True, help="Fichier YANNICK gestion directe/indirecte (obligatoire)")
    p.add_argument("--devises", type=Path, required=True, help="Fichier DEVISES Date|Devise|Taux (obligatoire)")
    p.add_argument("--account", type=Path, required=True, help="Fichier BG-LE-RMPM ACCOUNT (obligatoire)")
    p.add_argument("--usage", type=Path, default=None, help="Fichier MATCHING_USAGE (optionnel)")
    p.add_argument("--bpe-retail", type=Path, default=None, help="Fichier BPE RETAIL (optionnel)")
    p.add_argument("--mc1", type=Path, default=None, help="Fichier MATCHING_CLIENT 1 (optionnel)")
    p.add_argument("--mc2", type=Path, default=None, help="Fichier MATCHING_CLIENT 2 (optionnel)")
    p.add_argument("--output-dir", type=Path, required=True, help="Repertoire de sortie (obligatoire)")
    p.add_argument("--output-filename", type=str, required=True, help="Prefixe de nom des fichiers produits, sans extension (obligatoire)")
    return p


def main(argv=None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        app = WorldlineAnalyzer_V2MRG(args)
        m, pnb_cols, devise_cfg = app._build_config()
        app.worker(m, pnb_cols, devise_cfg)
    except Exception as e:
        print(f"[ERREUR] {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1
    print("[OK] Traitement terminé.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
