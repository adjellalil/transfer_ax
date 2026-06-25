"""
================================================================================
CPC ANIMATION COMMERCIALE v16 [TXVLV]
================================================================================

DESCRIPTION
    Reporting d'animation commerciale CPC (BNP Paribas Cash Management,
    Direction Monetique). Suivi de la baisse de flux clients Worldline :
    croisement PRGM Worldline (41 colonnes, flux/PNB mensuels) et ACHETEUR
    Casper/Noyan (72 colonnes, metadonnees), enrichissement optionnel
    PARC_CLIENT / ACCOUNT / SALES, calcul des flux/PNB par trimestre, des
    cumuls periode N et N-1, des evolutions, du plafond proxy (annualise)
    et des alertes de baisse. Production d'un classeur XLSX (feuilles DATA,
    SYNTHESE, DASHBOARD). Version CLI autonome, sans interface graphique :
    la logique metier est preservee a l'identique.

SOURCES REQUISES
    PRGM      (obligatoire) : export Worldline 41 colonnes (flux/PNB mensuels).
    ACHETEUR  (obligatoire) : export Casper/Noyan 72 colonnes (metadonnees).
    PARC      (optionnel)   : PARC_CLIENT (col6=RMPM, col14=FDC).
    ACCOUNT   (optionnel)   : BG-LE-RMPM-ACCOUNT (IBAN -> GA/RMPM).
    SALES     (optionnel)   : SALES/YANNICK (GA -> Sales -> GI).
    Matching : IBAN -> ACCOUNT -> RMPM -> PARC -> FDC.
    Jointure : ID Programme (PRGM col 4 = ACHETEUR col 1).

OUTPUTS PRODUITS
    Un fichier XLSX (<output-dir>/<output-filename>) contenant :
      - feuille DATA      : dataset complet (identification, flux/PNB par
                            trimestre, cumuls, evolutions, alertes, mensuels).
      - feuille SYNTHESE  : parametres, KPI et tableau filtre dynamique.
      - feuille DASHBOARD  : tables dataviz (trimestres, tops, indicateurs).

ARGUMENTS CLI
    --source-prgm       (obligatoire) Chemin du fichier PRGM Worldline.
    --source-acheteur   (obligatoire) Chemin du fichier ACHETEUR Casper/Noyan.
    --source-parc       (optionnel)   Chemin du fichier PARC_CLIENT.
    --source-account    (optionnel)   Chemin du fichier ACCOUNT.
    --source-sales      (optionnel)   Chemin du fichier SALES.
    --output-dir        (obligatoire) Repertoire de sortie du XLSX.
    --output-filename   (obligatoire) Nom du fichier XLSX produit.
    --seuil-pct         (optionnel)   Seuil d'alerte en % (defaut 10).
    --seuil-eur         (optionnel)   Seuil d'alerte en EUR (defaut 100000).
    --annee-ref         (optionnel)   Annee de reference N (defaut: derniere detectee).
    --annee-prec        (optionnel)   Annee precedente N-1 (defaut: avant-derniere).
    --mois-fin          (optionnel)   Mois de fin AAAAMM ou AAAA-MM (defaut: dernier).

DECOMPOSITION
    01.TXVLV.py
    |-- Constantes & mappings (PRGM, ACHET, PARC, ACC, SAL, MOIS, TRIM, couleurs)
    |-- Helpers VERBATIM (lcsv, cid, cid0, ciban, tof, pmois, pid, fnz, _f, _ft, _col)
    |-- App
    |   |-- __init__(args)        : affecte self.files / options depuis argparse
    |   |-- _resolve_config()     : reproduit la derivation GUI (m, pnb, dep, ar/ap, am, mf)
    |   |-- _p(v, t)              : remplace la progressbar GUI par un print
    |   |-- _work(cfg)           : worker VERBATIM (calcul flux/PNB/activite,
    |   |                          plafond proxy, indicateurs, XLSX)
    |   |-- _patch(path)         : patch Dynamic Arrays (VERBATIM)
    |   `-- run()                : orchestration (config -> worker)
    `-- main()                   : argparse -> App.run() -> sorties 0/1/2
================================================================================
"""

import argparse
import os
import re as _re
import shutil
import sys
import tempfile
import unicodedata
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter as C
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL = True
except ImportError:
    OPENPYXL = False

VER = "TXVLV"
BG = "#00915A"

# Couleurs XLSX (sans #)
GRN="00915A"; GRN2="E8F5E9"; DK="1C3A2D"; WH="FFFFFF"
BL="1565C0"; OR="E65100"; OR2="FFF3E0"; RD="B71C1C"; RD2="FFEBEE"; GY="F5F5F5"

MOIS = {'01':'JAN','02':'FEV','03':'MAR','04':'AVR','05':'MAI','06':'JUN',
        '07':'JUL','08':'AOU','09':'SEP','10':'OCT','11':'NOV','12':'DEC'}
TRIM = {'01':'T1','02':'T1','03':'T1','04':'T2','05':'T2','06':'T2',
        '07':'T3','08':'T3','09':'T3','10':'T4','11':'T4','12':'T4'}

# -- POSITIONS CONFIRMEES ----------------------------------------------------
PRGM = {
    'mois':2, 'nom':3, 'id':4, 'produit':5,
    'code_agence':6, 'agence':7, 'rs':8, 'iban':9,
    'groupe':10, 'sous_groupe':11, 'devise':12,
    'plafond':13, 'periodicite':14,
    'nb_cartes':15,
    'dep_cols':[24,25,26,27,28,29],  # Mt proxi+retraits+VAD+fourn+CB+VISA
    'pnb_cols':[30,31,32,33,34,35,36],  # cotis+abon+comm+ichgCB+ichgVI+ichgVIi+autres
    'rc':40, 'differe':41,
}
ACHET = {
    'id':1, 'libelle':2, 'rs':3,
    'nom_groupe':8, 'nom_sous_groupe':9,
    'periode':13, 'type':15, 'code_agence':16, 'iban':17,
    'etat_actif':19, 'pays':24,
    'adh_proc':36, 'plaf_proc':37, 'adh_virt':39, 'plaf_virt':40,
    'adh_voyage':41, 'plaf_voyage':42, 'adh_achat':43, 'plaf_achat':44,
    'adh_cvv':46, 'plaf_cvv':47,
    'differe_yn':63, 'differe_j':64,
    'rc':71, 'date_creation':72,
}
PARC = {'rmpm':6, 'code_ga':11, 'rc':13, 'fdc':14}
ACC = {'code_ga':2, 'nom_ga':3, 'rmpm':5, 'iban':7}
SAL = {'code_ga':1, 'sales':3, 'gi':4}

NUM_PFX = ('FLUX_','PNB_','CUMUL_','EUR_','PCT_','TAUX_','PLAFOND_PERIODIQUE','PLAFOND_ANNUALISE')
PCT_PFX = ('PCT_','TAUX_')

# -- UTILS -------------------------------------------------------------------
def lcsv(p, nrows=None):
    for s in [';',',','\t']:
        for e in ['utf-8-sig','utf-8','latin1','cp1252']:
            try:
                t=pd.read_csv(p,sep=s,encoding=e,dtype=str,keep_default_na=False,na_values=[],on_bad_lines='skip',nrows=5)
                if t.shape[1]>1: return pd.read_csv(p,sep=s,encoding=e,dtype=str,keep_default_na=False,na_values=[],on_bad_lines='skip',nrows=nrows)
            except: continue
    return pd.read_csv(p,sep=None,engine='python',dtype=str,keep_default_na=False,na_values=[],on_bad_lines='skip',nrows=nrows)

def cid(s):
    s=s.astype(str).str.strip().replace(['','nan','NaN','None','NULL','NA','N/A'],'')
    m=s.str.startswith('="')&s.str.endswith('"'); s=s.where(~m,s.str[2:-1])
    s=s.str.lstrip("'"); m2=s.str.endswith('.0')&s.str[:-2].str.isdigit(); s=s.where(~m2,s.str[:-2])
    return s.str.strip()

def cid0(s): s=cid(s); r=s.str.lstrip('0'); return r.where(r!='',s)
def ciban(s): return cid(s).str.upper().str.replace(' ','',regex=False)

def tof(s):
    s=s.astype(str)
    for r in ['"',"'",' ','\xa0',' ','€']: s=s.str.replace(r,'',regex=False)
    s=s.str.replace('EUR','',regex=False); m=s.str.endswith('-'); s=s.where(~m,'-'+s.str[:-1])
    s=s.str.replace(',','.',regex=False); return pd.to_numeric(s,errors='coerce').fillna(0.0)

def pmois(series):
    def _p(v):
        if pd.isna(v): return ''
        s=str(v).strip()
        if s.startswith('="') and s.endswith('"'): s=s[2:-1].strip()
        s=s.lstrip("'").strip()
        if s.endswith('.0') and s[:-2].isdigit(): s=s[:-2]
        if _re.fullmatch(r'\d{6}',s): return s
        for pat,fn in [
            (r'^(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})$',lambda m:f"{m.group(3)}{int(m.group(2)):02d}" if 1<=int(m.group(2))<=12 else ''),
            (r'^(\d{1,2})[/\-\.](\d{4})$',lambda m:f"{m.group(2)}{int(m.group(1)):02d}" if 1<=int(m.group(1))<=12 else ''),
            (r'^(\d{4})[/\-\.](\d{1,2})$',lambda m:f"{m.group(1)}{int(m.group(2)):02d}" if 1<=int(m.group(2))<=12 else ''),
        ]:
            match=_re.match(pat,s)
            if match:
                r=fn(match)
                if r: return r
        try:
            dt=pd.to_datetime(s,dayfirst=True,errors='coerce')
            if pd.notna(dt): return dt.strftime('%Y%m')
        except: pass
        return ''
    return series.apply(_p)

def pid(v):
    if pd.isna(v) or str(v).strip()=='': return ''
    s=str(v).strip()
    if s.startswith('="'): return s
    if s.isdigit() or (s.startswith('0') and len(s)>1): return f'="{s}"'
    return s

def fnz(arr):
    for v in arr:
        if v and str(v).strip() and str(v).strip() not in ('0','0.0','nan'): return v
    return ''

def _f(h): return PatternFill(start_color=h,end_color=h,fill_type='solid')
def _ft(c="FFFFFF",b=True,sz=9): return Font(name='Arial',size=sz,bold=b,color=c)
_tn=Side(style='thin',color='CCCCCC'); _bd=Border(left=_tn,right=_tn,top=_tn,bottom=_tn)

def _col(df, pos):
    """Get column name by 1-indexed position, or None."""
    cols = list(df.columns)
    return cols[pos-1] if pos and 1 <= pos <= len(cols) else None


# ======================================================================
class App:
    def __init__(self, args: argparse.Namespace) -> None:
        # Fichiers (memes cles que la version GUI)
        self.files: Dict[str, str] = {
            "PRGM": str(args.source_prgm),
            "ACHETEUR": str(args.source_acheteur),
            "PARC": str(args.source_parc) if args.source_parc else "",
            "ACCOUNT": str(args.source_account) if args.source_account else "",
            "SALES": str(args.source_sales) if args.source_sales else "",
        }
        # Options (remplacent les BooleanVar/checkboxes GUI)
        self._use_parc = bool(args.source_parc)
        self._use_acc = bool(args.source_account)
        self._use_sal = bool(args.source_sales)
        # Sortie (remplace asksaveasfilename)
        self.output_dir = Path(args.output_dir)
        self.output_filename = str(args.output_filename)
        # Parametres (remplacent les entries GUI)
        self._seuil_pct = float(str(args.seuil_pct).replace(',', '.')) / 100
        self._seuil_eur = float(str(args.seuil_eur).replace(',', '.'))
        self._annee_ref = str(args.annee_ref) if args.annee_ref else ''
        self._annee_prec = str(args.annee_prec) if args.annee_prec else ''
        self._mois_fin = str(args.mois_fin) if args.mois_fin else ''
        # Etats internes
        self.previews: Dict[str, Any] = {}
        self.ocols: Dict[str, List[str]] = {}
        self._am: List[str] = []

    # Equivalents des accesseurs de checkboxes GUI -----------------------
    def use_parc_get(self) -> bool: return self._use_parc
    def use_acc_get(self) -> bool: return self._use_acc
    def use_sal_get(self) -> bool: return self._use_sal

    # -- Reproduction de la derivation GUI (mapping par defaut + periodes) --
    def _resolve_config(self) -> Dict[str, Any]:
        """Reproduit, sans GUI, la construction de cfg realisee par _go().

        Le mapping (m, pnb, dep) reprend les positions par defaut que la GUI
        pre-selectionnait dans les combobox (les memes colonnes positionnelles).
        Les annees/mois sont detectes depuis la colonne mois du PRGM.
        """
        keys = ["PRGM", "ACHETEUR"]
        if self.use_parc_get(): keys.append("PARC")
        if self.use_acc_get(): keys.append("ACCOUNT")
        if self.use_sal_get(): keys.append("SALES")
        for k in keys:
            self.previews[k] = lcsv(self.files[k], nrows=5)
            self.ocols[k] = list(self.previews[k].columns)

        # Mapping par defaut PRGM (positions confirmees) -> noms de colonnes
        def _pcol(pos: int) -> Optional[str]:
            return _col(self.previews["PRGM"], pos) if "PRGM" in self.previews else None
        m: Dict[str, Optional[str]] = {
            'p_mois': _pcol(PRGM['mois']),
            'p_id': _pcol(PRGM['id']),
            'p_prod': _pcol(PRGM['produit']),
            'p_cag': _pcol(PRGM['code_agence']),
            'p_rs': _pcol(PRGM['rs']),
            'p_iban': _pcol(PRGM['iban']),
            'p_plaf': _pcol(PRGM['plafond']),
            'p_per': _pcol(PRGM['periodicite']),
            'p_rc': _pcol(PRGM['rc']),
        }
        # ACHETEUR (verif) -> defauts positionnels
        def _acol(pos: int) -> Optional[str]:
            return _col(self.previews["ACHETEUR"], pos) if "ACHETEUR" in self.previews else None
        m['a_id'] = _acol(ACHET['id']); m['a_rs'] = _acol(ACHET['rs'])
        m['a_iban'] = _acol(ACHET['iban']); m['a_rc'] = _acol(ACHET['rc'])
        # PARC / ACCOUNT / SALES (si actives)
        if self.use_parc_get() and "PARC" in self.previews:
            dfpc = self.previews["PARC"]
            m['pc_rmpm'] = _col(dfpc, PARC['rmpm']); m['pc_ga'] = _col(dfpc, PARC['code_ga'])
            m['pc_rc'] = _col(dfpc, PARC['rc']); m['pc_fdc'] = _col(dfpc, PARC['fdc'])
        if self.use_acc_get() and "ACCOUNT" in self.previews:
            dfac = self.previews["ACCOUNT"]
            m['ac_ga'] = _col(dfac, ACC['code_ga']); m['ac_nga'] = _col(dfac, ACC['nom_ga'])
            m['ac_rmpm'] = _col(dfac, ACC['rmpm']); m['ac_iban'] = _col(dfac, ACC['iban'])
        if self.use_sal_get() and "SALES" in self.previews:
            dfsl = self.previews["SALES"]
            m['sl_ga'] = _col(dfsl, SAL['code_ga']); m['sl_s'] = _col(dfsl, SAL['sales'])
            m['sl_gi'] = _col(dfsl, SAL['gi'])

        # PNB / DEP : colonnes positionnelles par defaut (comme combobox GUI)
        prgm_cols = self.ocols.get("PRGM", [])
        pnb = [prgm_cols[i-1] for i in PRGM['pnb_cols'] if 1 <= i <= len(prgm_cols)]
        dep = [prgm_cols[i-1] for i in PRGM['dep_cols'] if 1 <= i <= len(prgm_cols)]

        # Detection annees / mois depuis la colonne mois du PRGM (comme _cfg_ui)
        try:
            df = lcsv(self.files["PRGM"])
            if m.get('p_mois'):
                am_s = pmois(df[m['p_mois']]); al = sorted([x for x in am_s.unique() if x and len(x) == 6])
                ann = sorted(set(mo[:4] for mo in al))
            else:
                ann = []; al = []
        except:
            ann = []; al = []
        if not ann: ann = [str(datetime.now().year-1), str(datetime.now().year)]
        self._am = al

        # Annee P / P-1 (defaut : derniere / avant-derniere detectee)
        ar = self._annee_ref or ann[-1]
        ap = self._annee_prec or (ann[-2] if len(ann) >= 2 else ann[0])
        if ar not in ann:
            raise ValueError(f"Annee ref '{ar}' absente des annees detectees {ann}")
        if ap not in ann:
            raise ValueError(f"Annee prec '{ap}' absente des annees detectees {ann}")

        # Mois de fin (defaut : dernier mois detecte)
        if self._mois_fin:
            mf = self._mois_fin.replace('-', '')
        else:
            mf = al[-1] if al else ''

        sp = self._seuil_pct
        se = self._seuil_eur
        op = str(self.output_dir / self.output_filename)

        return {'m': m, 'pnb': pnb, 'dep': dep, 'sp': sp, 'se': se,
                'ar': ar, 'ap': ap, 'am': self._am, 'mf': mf, 'op': op}

    def _p(self, v, t):
        """Remplace la progressbar GUI : journalisation sur stdout."""
        print(f"[{v*100:5.1f}%] {t}")

    # =================================================================
    # WORKER
    # =================================================================
    def _work(self, cfg):
        m=cfg['m']; pnb=cfg['pnb']; dep=cfg['dep']; sp=cfg['sp']; se=cfg['se']
        ar=cfg['ar']; ap=cfg['ap']; am=cfg['am']; mf=cfg['mf']; op=cfg['op']
        mf_mm=mf[4:] if len(mf)==6 else '12'
        m_ref=[mo for mo in am if mo[:4]==ar and mo[4:]<=mf_mm]
        m_prec=[mo for mo in am if mo[:4]==ap and mo[4:]<=mf_mm]

        # LOAD
        self._p(0.02,"Chargement PRGM..."); dfP=lcsv(self.files["PRGM"])
        self._p(0.06,"Chargement ACHETEUR..."); dfA=lcsv(self.files["ACHETEUR"])
        dfPC=lcsv(self.files["PARC"]) if self.use_parc_get() and self.files.get("PARC") else None
        dfAC=lcsv(self.files["ACCOUNT"]) if self.use_acc_get() and self.files.get("ACCOUNT") else None
        dfSL=lcsv(self.files["SALES"]) if self.use_sal_get() and self.files.get("SALES") else None

        # PRGM prep
        self._p(0.12,"Préparation PRGM...")
        if m.get('p_mois'): dfP['_MO']=pmois(dfP[m['p_mois']])
        if m.get('p_id'): dfP['_ID']=cid(dfP[m['p_id']])
        if m.get('p_prod'): dfP['_PROD']=dfP[m['p_prod']].astype(str).str.strip()
        if m.get('p_cag'): dfP['_CAG']=cid(dfP[m['p_cag']])
        prgm_nom_col = _col(dfP, PRGM['nom'])
        if prgm_nom_col: dfP['_NOM']=dfP[prgm_nom_col].astype(str).str.strip()
        prgm_agence_col = _col(dfP, PRGM['agence'])
        if prgm_agence_col: dfP['_AGENCE']=dfP[prgm_agence_col].astype(str).str.strip()
        if m.get('p_rs'): dfP['_RS']=dfP[m['p_rs']].astype(str).str.strip()
        if m.get('p_iban'): dfP['_IBAN']=ciban(dfP[m['p_iban']])
        if m.get('p_plaf'): dfP['_PLAF']=tof(dfP[m['p_plaf']])
        if m.get('p_per'): dfP['_PER']=tof(dfP[m['p_per']])
        prgm_grp_col = _col(dfP, PRGM['groupe'])
        if prgm_grp_col: dfP['_GRP']=dfP[prgm_grp_col].astype(str).str.strip()
        prgm_sgrp_col = _col(dfP, PRGM['sous_groupe'])
        if prgm_sgrp_col: dfP['_SGRP']=dfP[prgm_sgrp_col].astype(str).str.strip()
        if m.get('p_rc'): dfP['_RC']=cid0(dfP[m['p_rc']])
        # Flux = sum of dep cols
        if dep:
            for c in dep:
                if c in dfP.columns: dfP[c]=tof(dfP[c])
            dfP['_FLUX']=dfP[[c for c in dep if c in dfP.columns]].sum(axis=1)
        else: dfP['_FLUX']=0.0
        # PNB = sum of pnb cols
        if pnb:
            for c in pnb:
                if c in dfP.columns: dfP[c]=tof(dfP[c])
            dfP['_PNB']=dfP[[c for c in pnb if c in dfP.columns]].sum(axis=1)
        else: dfP['_PNB']=0.0
        self._p(0.20,f"PRGM: {len(dfP):,} lignes")

        # ACHETEUR prep
        self._p(0.22,"Préparation ACHETEUR...")
        ac=list(dfA.columns)
        def _a(k):
            p=ACHET.get(k); return ac[p-1] if p and 1<=p<=len(ac) else None
        a_id=m.get('a_id') or _a('id'); a_rs=m.get('a_rs') or _a('rs')
        a_iban=m.get('a_iban') or _a('iban'); a_rc=m.get('a_rc') or _a('rc')
        if a_id: dfA['_ID']=cid(dfA[a_id])
        if _a('libelle'): dfA['_LIB']=dfA[_a('libelle')].astype(str).str.strip()
        if a_rs: dfA['_RS']=dfA[a_rs].astype(str).str.strip()
        if _a('code_agence'): dfA['_CAG']=cid(dfA[_a('code_agence')])
        if a_iban: dfA['_IBAN']=ciban(dfA[a_iban])
        if a_rc: dfA['_RC']=cid(dfA[a_rc])
        if _a('date_creation'): dfA['_DCREA']=dfA[_a('date_creation')].astype(str).str.strip()
        for k,d in [('nom_groupe','_NGRP'),('nom_sous_groupe','_NSGRP'),('type','_TYPE'),
                     ('pays','_PAYS'),('etat_actif','_ETAT')]:
            c=_a(k)
            if c: dfA[d]=dfA[c].astype(str).str.strip()
        if _a('periode'): dfA['_PER']=tof(dfA[_a('periode')])
        if _a('differe_yn'): dfA['_DIFYN']=dfA[_a('differe_yn')].astype(str).str.strip().str.upper()
        if _a('differe_j'): dfA['_DIFJ']=tof(dfA[_a('differe_j')])
        # Adhésions → produit
        for ak,d in [('adh_proc','_AP'),('adh_virt','_AV'),('adh_voyage','_AVO'),
                      ('adh_achat','_AA'),('adh_cvv','_AC')]:
            c=_a(ak)
            if c: dfA[d]=dfA[c].astype(str).str.strip().str.upper()
        for pk,d in [('plaf_proc','_PP'),('plaf_virt','_PV'),('plaf_voyage','_PVO'),
                      ('plaf_achat','_PA'),('plaf_cvv','_PC')]:
            c=_a(pk)
            if c: dfA[d]=tof(dfA[c])
        def _dprod(r):
            for a,l in [('_AV','Virtuelle'),('_AP','Procurement'),('_AVO','Voyage'),('_AA','Carte Achat'),('_AC','CVV')]:
                if r.get(a,'') in ('OUI','O','YES','Y','TRUE'): return l
            return ''
        if any(c in dfA.columns for c in ['_AP','_AV','_AVO','_AA','_AC']):
            dfA['_PROD']=dfA.apply(_dprod,axis=1)
        else: dfA['_PROD']=''
        pcs=[c for c in ['_PP','_PV','_PVO','_PA','_PC'] if c in dfA.columns]
        dfA['_PLMAX']=dfA[pcs].max(axis=1) if pcs else 0.0
        # Dict acheteur
        ad={}
        if '_ID' in dfA.columns:
            aw=[c for c in dfA.columns if c.startswith('_')]
            for _,r in dfA.iterrows():
                idp=r.get('_ID','')
                if idp: ad[idp]={c:r.get(c,'') for c in aw}
        self._p(0.30,f"ACHETEUR: {len(ad):,} prog")

        # DICTS enrichissement
        acc_iban={}
        if dfAC is not None and m.get('ac_iban'):
            self._p(0.32,"Dict ACCOUNT...")
            dfAC['_IBAN']=ciban(dfAC[m['ac_iban']])
            dfAC['_GA']=cid0(dfAC[m['ac_ga']]) if m.get('ac_ga') else ''
            dfAC['_NGA']=dfAC[m['ac_nga']].astype(str).str.strip() if m.get('ac_nga') else ''
            dfAC['_RMPM']=cid(dfAC[m['ac_rmpm']]) if m.get('ac_rmpm') else ''
            for ib,ga,ng,rm in zip(dfAC['_IBAN'].values,dfAC['_GA'].values,dfAC['_NGA'].values,dfAC['_RMPM'].values):
                if ib and ib not in acc_iban: acc_iban[ib]=(ga,ng,rm)

        parc_rmpm={}  # RMPM → (GA, FDC)
        parc_rc={}    # RC → (RMPM, GA, FDC, RS)
        if dfPC is not None and m.get('pc_rmpm'):
            self._p(0.34,"Dict PARC...")
            dfPC['_RMPM']=cid(dfPC[m['pc_rmpm']])
            dfPC['_GA']=cid0(dfPC[m['pc_ga']]) if m.get('pc_ga') else ''
            dfPC['_FDC']=cid(dfPC[m['pc_fdc']]) if m.get('pc_fdc') else ''
            dfPC['_RC']=cid0(dfPC[m['pc_rc']]) if m.get('pc_rc') else ''
            # Nom GA depuis PARC col 12 si dispo
            parc_cols=list(dfPC.columns)
            dfPC['_NGA_P']=dfPC[parc_cols[11]].astype(str).str.strip() if len(parc_cols)>11 else ''
            dfPC['_RS_P']=dfPC[parc_cols[7]].astype(str).str.strip() if len(parc_cols)>7 else ''
            for rm,ga,fdc,rc,nga,rs in zip(dfPC['_RMPM'].values,dfPC['_GA'].values,
                dfPC['_FDC'].values,dfPC['_RC'].values,dfPC['_NGA_P'].values,dfPC['_RS_P'].values):
                if rm and rm not in parc_rmpm: parc_rmpm[rm]=(ga,fdc,nga)
                if rc and rc not in parc_rc: parc_rc[rc]=(rm,ga,fdc,rs)

        # Reverse dict ACCOUNT: RMPM → (GA, NGA) pour fallback
        acc_rmpm={}
        if dfAC is not None and m.get('ac_rmpm'):
            for rm,ga,ng in zip(dfAC['_RMPM'].values,dfAC['_GA'].values,dfAC['_NGA'].values):
                if rm and rm not in acc_rmpm: acc_rmpm[rm]=(ga,ng)

        sal_ga={}
        if dfSL is not None and m.get('sl_ga'):
            self._p(0.36,"Dict SALES...")
            dfSL['_GA']=cid0(dfSL[m['sl_ga']])
            dfSL['_SL']=dfSL[m['sl_s']].astype(str).str.strip() if m.get('sl_s') else ''
            dfSL['_GI']=dfSL[m['sl_gi']].astype(str).str.strip().str.upper() if m.get('sl_gi') else ''
            for ga,sl,gi in zip(dfSL['_GA'].values,dfSL['_SL'].values,dfSL['_GI'].values):
                if ga and ga not in sal_ga: sal_ga[ga]=(sl,gi)

        # AGRÉGATS PRGM par ID_PROG + MOIS
        self._p(0.40,"Agrégats...")
        pagg=dfP.groupby(['_ID','_MO']).agg({'_FLUX':'sum','_PNB':'sum'}).reset_index()
        # Info statique par programme (premier non-vide)
        pi={}
        for idp,g in dfP.groupby('_ID'):
            pi[idp]={
                'nom':fnz(g['_NOM'].values) if '_NOM' in g else '',
                'prod':fnz(g['_PROD'].values) if '_PROD' in g else '',
                'rs':fnz(g['_RS'].values) if '_RS' in g else '',
                'iban':fnz(g['_IBAN'].values) if '_IBAN' in g else '',
                'rc':fnz(g['_RC'].values) if '_RC' in g else '',
                'cag':fnz(g['_CAG'].values) if '_CAG' in g else '',
                'agence':fnz(g['_AGENCE'].values) if '_AGENCE' in g else '',
                'grp':fnz(g['_GRP'].values) if '_GRP' in g else '',
                'sgrp':fnz(g['_SGRP'].values) if '_SGRP' in g else '',
                'plaf':float(g['_PLAF'].max()) if '_PLAF' in g else 0,
                'per':float(g['_PER'].iloc[0]) if '_PER' in g else 1,
            }
        pmf={}; pmp={}
        for _,r in pagg.iterrows():
            pmf.setdefault(r['_ID'],{})[r['_MO']]=r['_FLUX']
            pmp.setdefault(r['_ID'],{})[r['_MO']]=r['_PNB']

        # FUSION
        self._p(0.50,"Fusion...")
        aids=sorted(set(pi.keys())|set(ad.keys()))
        self._p(0.52,f"{len(aids):,} programmes")

        # Colonnes trimestres
        all_ann=sorted(set(mo[:4] for mo in am))
        tcols=[]
        for yr in all_ann:
            for tq in ['T1','T2','T3','T4']:
                tmos=[mo for mo in am if mo[:4]==yr and TRIM.get(mo[4:])==tq]
                if tmos: tcols.append((f'{tq}_{yr}_FLUX',f'{tq}_{yr}_PNB',tmos))

        # Noms cumuls
        md=MOIS.get('01','JAN'); mfl=MOIS.get(mf_mm,'??')
        cfp=f'CUMUL_FLUX_{md}_{mfl}_{ap}'; cfr=f'CUMUL_FLUX_{md}_{mfl}_{ar}'
        cpp=f'CUMUL_PNB_{md}_{mfl}_{ap}'; cpr=f'CUMUL_PNB_{md}_{mfl}_{ar}'

        # DATASET
        self._p(0.55,"Dataset...")
        rows=[]
        for i,idp in enumerate(aids):
            if i%1000==0: self._p(0.55+0.15*i/max(len(aids),1),f"DATA: {i:,}/{len(aids):,}")
            p=pi.get(idp,{}); a=ad.get(idp,{})
            inp=idp in pi; ina=idp in ad
            # Matching — CASCADE
            iban=p.get('iban','') or a.get('_IBAN','')
            rc=p.get('rc','') or str(a.get('_RC','')).strip()
            ga=''; nga=''; rmpm=''; fdc=''; src_match=''

            # 1) IBAN → ACCOUNT
            if iban and iban in acc_iban:
                ga,nga,rmpm=acc_iban[iban]
                src_match='IBAN→ACCOUNT'

            # 2) Fallback: RC → PARC (si pas de RMPM trouvé)
            if not rmpm and rc and rc in parc_rc:
                rmpm_p,ga_p,fdc_p,rs_p=parc_rc[rc]
                rmpm=rmpm_p
                if not ga: ga=ga_p
                fdc=fdc_p
                src_match='RC→PARC' if not src_match else src_match

            # 3) RMPM → PARC (FDC + GA fallback)
            if rmpm and rmpm in parc_rmpm:
                ga_p2,fdc_p2,nga_p2=parc_rmpm[rmpm]
                if not fdc: fdc=fdc_p2
                if not ga: ga=ga_p2
                if not nga and nga_p2: nga=nga_p2
                if not src_match: src_match='RMPM→PARC'

            # 4) RMPM → ACCOUNT reverse (si GA/NGA manquants)
            if rmpm and (not ga or not nga) and rmpm in acc_rmpm:
                ga_r,nga_r=acc_rmpm[rmpm]
                if not ga: ga=ga_r
                if not nga: nga=nga_r

            # SALES via GA
            sales=''; seg='RESEAU'
            if ga and ga in sal_ga:
                sl,gi=sal_ga[ga]; sales=sl
                seg='RESEAU' if gi in ('OUI','O','YES','Y','TRUE') else 'SALES'
            # Code agence : priorité PRGM (col 6), fallback ACHETEUR (col 16) — pas de zéros ajoutés
            cag=p.get('cag','') or a.get('_CAG','')
            # Strip leading zeros for code agence (828, not 00828)
            cag_clean=cag.lstrip('0') if cag else ''
            if not cag_clean: cag_clean=cag

            prod=a.get('_PROD','') or p.get('prod','')
            rs=p.get('rs','') or a.get('_RS','')
            # ID RC : PRGM col 40 ou ACHETEUR col 71
            id_rc=p.get('rc','') or str(a.get('_RC','')).strip()
            dcrea=a.get('_DCREA','')
            # Date : format JJ/MM/AAAA sans heure
            if dcrea:
                try:
                    dt=pd.to_datetime(dcrea,dayfirst=True,errors='coerce')
                    if pd.notna(dt): dcrea=dt.strftime('%d/%m/%Y')
                except: pass
            plaf_p=p.get('plaf',0); plaf_a=float(a.get('_PLMAX',0) or 0)
            plaf=plaf_p if plaf_p>0 else plaf_a
            per=p.get('per',0) or float(a.get('_PER',0) or 0)
            plaf_ann=plaf*(12/per) if per>0 else 0
            fx=pmf.get(idp,{}); px=pmp.get(idp,{})

            # Ordre colonnes : identification
            row={
                'CODE_AGENCE':cag_clean,
                'FDC':pid(fdc),
                'ID_RC':pid(id_rc),
                'RMPM':pid(rmpm),
                'RAISON_SOCIALE':rs,
                'CODE_GA':pid(ga),
                'NOM_GA':nga,
                'SALES':sales,
                'PRODUIT':prod,
                'ID_PROGRAMME':pid(idp),
                'LIBELLE_PROGRAMME':a.get('_LIB','') or p.get('nom',''),
                'DATE_CREATION':dcrea,
                'NOM_GROUPE':a.get('_NGRP','') or p.get('grp',''),
                'NOM_SOUS_GROUPE':a.get('_NSGRP','') or p.get('sgrp',''),
                'TYPE':a.get('_TYPE',''),
                'IBAN':iban,
                'PERIODE_MOIS':int(per) if per else '',
                'PLAFOND_PERIODIQUE':plaf,
            }
            # Trimestres : FLUX groupés, puis PNB groupés
            for fc,pc,tmos in tcols:
                row[fc]=sum(fx.get(mo,0) for mo in tmos)
            for fc,pc,tmos in tcols:
                row[pc]=sum(px.get(mo,0) for mo in tmos)
            # Cumuls
            fr=sum(fx.get(mo,0) for mo in m_ref); fp=sum(fx.get(mo,0) for mo in m_prec)
            pr_=sum(px.get(mo,0) for mo in m_ref); pp=sum(px.get(mo,0) for mo in m_prec)
            row[cfp]=fp; row[cfr]=fr; row[cpp]=pp; row[cpr]=pr_
            pct_f=(fr-fp)/abs(fp) if fp!=0 else (0 if fr==0 else 1)
            row['PCT_EVOL_FLUX_PP1']=pct_f; row['EUR_EVOL_FLUX_PP1']=fr-fp
            pct_pn=(pr_-pp)/abs(pp) if pp!=0 else (0 if pr_==0 else 1)
            row['PCT_EVOL_PNB_PP1']=pct_pn; row['EUR_EVOL_PNB_PP1']=pr_-pp
            # Alerte
            has_flux=fr>0
            alerte=(pct_f<-sp and abs(fr-fp)>se) and has_flux
            is_new=False
            if dcrea:
                try:
                    dc=pd.to_datetime(dcrea,dayfirst=True,errors='coerce')
                    if pd.notna(dc) and str(dc.year)==ar: is_new=True
                except: pass
            if not inp and not has_flux: st='INACTIF'
            elif is_new and not has_flux: st='NOUVEAU_SANS_FLUX'
            elif has_flux: st='ACTIF'
            else: st='SANS_FLUX'
            if st!='ACTIF': alerte=False
            row['ALERTE_BAISSE']='OUI' if alerte else 'NON'
            row['NB_PROG_SANS_FLUX']='OUI' if (is_new and not has_flux) else 'NON'
            row['COMMENTAIRES']=''
            row['SEGMENT']=seg
            # Mois détaillés à la fin (TOUS)
            for yr in all_ann:
                for mo in [m2 for m2 in am if m2[:4]==yr]:
                    row[f'FLUX_{mo[:4]}_{MOIS.get(mo[4:],"??")}']=fx.get(mo,0)
            for yr in all_ann:
                for mo in [m2 for m2 in am if m2[:4]==yr]:
                    row[f'PNB_{mo[:4]}_{MOIS.get(mo[4:],"??")}']=px.get(mo,0)
            rows.append(row)

        df=pd.DataFrame(rows)
        df=df.sort_values('CODE_AGENCE',key=lambda x: x.astype(str)).reset_index(drop=True)
        self._p(0.72,f"DATA: {len(df):,} programmes")

        # XLSX
        self._p(0.75,"XLSX...")
        if not OPENPYXL: raise RuntimeError("openpyxl manquant")
        wb=Workbook()
        headers=list(df.columns)
        num_c={h for h in headers if any(h.startswith(p) or h==p for p in NUM_PFX)}
        pct_c={h for h in headers if any(h.startswith(p) for p in PCT_PFX)}

        # SHEET DATA — with color-coded header blocks
        self._p(0.78,"Sheet DATA...")
        ws=wb.active; ws.title="DATA"

        # Identify column sections for coloring
        id_cols={'CODE_AGENCE','FDC','ID_RC','RMPM','RAISON_SOCIALE','CODE_GA','NOM_GA','SALES',
                 'PRODUIT','ID_PROGRAMME','LIBELLE_PROGRAMME','DATE_CREATION','NOM_GROUPE',
                 'NOM_SOUS_GROUPE','TYPE','IBAN','PERIODE_MOIS','PLAFOND_PERIODIQUE'}
        flux_t_cols={h for h in headers if h.endswith('_FLUX') and not h.startswith(('CUMUL','PCT','EUR'))}
        pnb_t_cols={h for h in headers if h.endswith('_PNB') and not h.startswith(('CUMUL','PCT','EUR'))}
        cumul_cols={h for h in headers if h.startswith('CUMUL_')}
        evol_cols={h for h in headers if h.startswith(('PCT_','EUR_'))}
        alert_cols={'ALERTE_BAISSE','NB_PROG_SANS_FLUX','COMMENTAIRES','SEGMENT','TAUX_UTILISATION_PLAFOND'}
        mois_flux={h for h in headers if h.startswith('FLUX_') and h not in flux_t_cols}
        mois_pnb={h for h in headers if h.startswith('PNB_') and h not in pnb_t_cols}

        def _hdr_color(h):
            if h in id_cols: return GRN
            if h in flux_t_cols: return BL
            if h in pnb_t_cols: return "7B1FA2"  # purple
            if h in cumul_cols: return "00695C"   # teal
            if h in evol_cols: return OR
            if h in alert_cols: return RD
            if h in mois_flux: return "37474F"
            if h in mois_pnb: return "4E342E"
            return GRN

        # Row 1 : bloc labels (merged where possible)
        # Row 2 : column headers
        HDR_ROW = 2
        # Write bloc labels in row 1
        bloc_names = [
            (id_cols, "IDENTIFICATION", GRN),
            (flux_t_cols, "FLUX PAR TRIMESTRE", BL),
            (pnb_t_cols, "PNB PAR TRIMESTRE", "7B1FA2"),
            (cumul_cols, "CUMULS PÉRIODE", "00695C"),
            (evol_cols, "ÉVOLUTIONS", OR),
            (alert_cols, "ALERTES", RD),
            (mois_flux, "FLUX MENSUELS", "37474F"),
            (mois_pnb, "PNB MENSUELS", "4E342E"),
        ]
        for bloc_set, bloc_name, bloc_color in bloc_names:
            positions = [ci for ci,h in enumerate(headers,1) if h in bloc_set]
            if positions:
                first, last = min(positions), max(positions)
                c = ws.cell(row=1, column=first, value=bloc_name)
                c.font = _ft(sz=8)
                c.fill = _f(bloc_color)
                c.alignment = Alignment(horizontal='center')
                for ci in range(first+1, last+1):
                    ws.cell(row=1, column=ci).fill = _f(bloc_color)

        # Row 2 : column names
        for ci,h in enumerate(headers,1):
            c=ws.cell(row=HDR_ROW,column=ci,value=h)
            c.font=_ft(sz=8); c.fill=_f(_hdr_color(h))
            c.alignment=Alignment(horizontal='center',wrap_text=True); c.border=_bd
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 30

        # Data rows (start at row 3)
        DATA_START = HDR_ROW + 1
        for ri,(_,row) in enumerate(df.iterrows(), DATA_START):
            rf=_f(GY) if ri%2==0 else _f(WH)
            for ci,h in enumerate(headers,1):
                v=row[h]; c=ws.cell(row=ri,column=ci)
                if h in pct_c and isinstance(v,(int,float)): c.value=v; c.number_format='0.0%'
                elif h in num_c and isinstance(v,(int,float)): c.value=round(v,2); c.number_format='#,##0 €'
                else:
                    if pd.isna(v) or str(v) in ('nan','None','NaN'): c.value=''
                    else: c.value=v
                c.font=Font(name='Arial',size=9); c.fill=rf; c.border=_bd
                if h=='ALERTE_BAISSE' and v=='OUI': c.fill=_f(RD2)
        for ci,h in enumerate(headers,1): ws.column_dimensions[C(ci)].width=max(11,min(len(h)+3,26))
        ws.freeze_panes=f'A{DATA_START}'
        ws.auto_filter.ref=f"A{HDR_ROW}:{C(len(headers))}{HDR_ROW}"

        # SHEET SYNTHESE
        self._p(0.87,"Sheet SYNTHESE...")
        wss=wb.create_sheet("SYNTHESE")
        t1=wss.cell(row=1,column=1,value="PARAMÈTRES"); t1.font=_ft(sz=9); t1.fill=_f(GRN)
        wss.cell(row=1,column=2).fill=_f(GRN)
        for r,lbl,val,fmt in [(2,"Indicateur","FLUX",None),(3,"Seuil %",sp,'0.0%'),(4,"Seuil €",se,'#,##0 €')]:
            wss.cell(row=r,column=1,value=lbl).font=Font(name='Arial',size=9,color=DK)
            wss.cell(row=r,column=1).fill=_f("F0F4F0"); wss.cell(row=r,column=1).border=_bd
            cb=wss.cell(row=r,column=2,value=val); cb.font=Font(name='Arial',size=9,bold=True,color="1565C0"); cb.border=_bd
            if fmt: cb.number_format=fmt
        wss.cell(row=5,column=1,value="Modifiez B2:B4 → résultats mis à jour").font=Font(name='Arial',size=8,italic=True,color="666666")
        dv=DataValidation(type="list",formula1='"FLUX,PNB"'); wss.add_data_validation(dv); dv.add(wss['B2'])
        wss.column_dimensions['A'].width=20; wss.column_dimensions['B'].width=14
        # KPIs
        wss.cell(row=1,column=4,value="RÉSUMÉ").font=_ft(sz=9); wss.cell(row=1,column=4).fill=_f(GRN); wss.cell(row=1,column=5).fill=_f(GRN)
        # Data source masquée
        dfa=df[df[cfr]!=0].sort_values('EUR_EVOL_FLUX_PP1').copy()
        n_act=len(dfa)
        SC=16; SH=7; SD=SH+1
        sk=['RMPM','ID_RC','RAISON_SOCIALE','PRODUIT','CODE_AGENCE',cfp,cfr,'EUR_EVOL_FLUX_PP1','PCT_EVOL_FLUX_PP1',
            cpp,cpr,'EUR_EVOL_PNB_PP1','PCT_EVOL_PNB_PP1','SEGMENT']
        ns=len(sk)
        for ci,h in enumerate(sk): wss.cell(row=SH,column=SC+ci,value=h).font=Font(name='Arial',size=8,color="999999")
        for ri,(_,row) in enumerate(dfa.iterrows()):
            for ci,h in enumerate(sk):
                v=row.get(h,'')
                c=wss.cell(row=SD+ri,column=SC+ci)
                if isinstance(v,(int,float)): c.value=round(v,2) if isinstance(v,float) else v
                else: c.value='' if (pd.isna(v) or str(v) in ('nan','None','NaN')) else v
        SE=SD+n_act-1
        for ci in range(SC,SC+ns): wss.column_dimensions[C(ci)].hidden=True
        # Visible table
        TH=7
        vn=['RMPM','ID RC','RAISON SOCIALE','PRODUIT','CODE AGENCE',f'FLUX {ap}',f'FLUX {ar}','ECART FLUX','EVOL FLUX %',
            f'PNB {ap}',f'PNB {ar}','ECART PNB','EVOL PNB %','SEGMENT']
        for ci,h in enumerate(vn,1):
            c=wss.cell(row=TH,column=ci,value=h); c.font=_ft(); c.fill=_f(DK); c.alignment=Alignment(horizontal='center',wrap_text=True); c.border=_bd
        # FILTER formula
        sr=f"{C(SC)}{SD}:{C(SC+ns-1)}{SE}"
        ef=f"{C(SC+8)}{SD}:{C(SC+8)}{SE}"; ec=f"{C(SC+7)}{SD}:{C(SC+7)}{SE}"
        ep=f"{C(SC+12)}{SD}:{C(SC+12)}{SE}"; epc=f"{C(SC+11)}{SD}:{C(SC+11)}{SE}"
        filt=f'=_xlfn._xlws.SORT(_xlfn._xlws.FILTER({sr},IF($B$2="FLUX",({ef}<-$B$3)*(-{ec}>$B$4),({ep}<-$B$3)*(-{epc}>$B$4)),"Aucun résultat"),IF($B$2="FLUX",8,12),1)'
        wss.cell(row=TH+1,column=1,value=filt)
        wss.freeze_panes=f'A{TH+1}'
        for ci,w in enumerate([14,14,28,16,14,15,15,15,12,15,15,15,12,12],1): wss.column_dimensions[C(ci)].width=w
        wss.column_dimensions['D'].width=22; wss.column_dimensions['E'].width=16
        # KPI formulas
        for r,lbl,val,fmt in [
            (2,"Programmes avec flux",str(n_act),'0'),
            (3,"Nb alertes",f'=SUMPRODUCT(IF($B$2="FLUX",({ef}<-$B$3)*(-{ec}>$B$4),({ep}<-$B$3)*(-{epc}>$B$4)))','0'),
            (4,"Écart total",f'=SUMPRODUCT(IF($B$2="FLUX",({ef}<-$B$3)*(-{ec}>$B$4),({ep}<-$B$3)*(-{epc}>$B$4))*IF($B$2="FLUX",{ec},{epc}))','#,##0 €'),
        ]:
            wss.cell(row=r,column=4,value=lbl).font=Font(name='Arial',size=9,bold=True,color=DK)
            c=wss.cell(row=r,column=5,value=val); c.font=Font(name='Arial',size=9,bold=True,color=GRN); c.number_format=fmt

        n_al=len(df[df['ALERTE_BAISSE']=='OUI'])

        # SHEET DASHBOARD (tables dataviz — hidden)
        self._p(0.92,"Sheet DASHBOARD...")
        wsd=wb.create_sheet("DASHBOARD")
        wsd.sheet_state = 'hidden'
        wsd.cell(row=1,column=1,value=f"DASHBOARD [{VER}]").font=Font(name='Arial',size=11,bold=True,color=GRN)
        # Table 1: Trimestres
        dr=3
        wsd.cell(row=dr,column=1,value="TRIMESTRE").font=_ft(DK); wsd.cell(row=dr,column=2,value="FLUX").font=_ft(DK); wsd.cell(row=dr,column=3,value="PNB").font=_ft(DK)
        for ci in range(1,4): wsd.cell(row=dr,column=ci).fill=_f(GRN2)
        for fc,pc,tmos in tcols:
            dr+=1; tlbl=fc.replace('_FLUX','')
            wsd.cell(row=dr,column=1,value=tlbl).font=Font(name='Arial',size=8)
            wsd.cell(row=dr,column=2,value=round(df[fc].sum(),0)).number_format='#,##0 €'
            wsd.cell(row=dr,column=3,value=round(df[pc].sum(),0)).number_format='#,##0 €'
        # Table 2: Top 10 baisses
        dr+=2
        wsd.cell(row=dr,column=1,value="TOP 10 BAISSES FLUX").font=_ft(RD,sz=9); dr+=1
        wsd.cell(row=dr,column=1,value="CLIENT").font=_ft(DK); wsd.cell(row=dr,column=2,value="ECART €").font=_ft(DK); wsd.cell(row=dr,column=3,value="EVOL %").font=_ft(DK)
        for ci in range(1,4): wsd.cell(row=dr,column=ci).fill=_f(RD2)
        top_b=df[df['ALERTE_BAISSE']=='OUI'].sort_values('EUR_EVOL_FLUX_PP1').head(10)
        for _,r in top_b.iterrows():
            dr+=1
            wsd.cell(row=dr,column=1,value=r['RAISON_SOCIALE']).font=Font(name='Arial',size=8)
            wsd.cell(row=dr,column=2,value=round(r['EUR_EVOL_FLUX_PP1'],0)).number_format='#,##0 €'
            wsd.cell(row=dr,column=3,value=r['PCT_EVOL_FLUX_PP1']).number_format='0.0%'
        # Table 3: Top 10 PNB
        dr+=2
        wsd.cell(row=dr,column=1,value=f"TOP 10 PNB {ar}").font=_ft(GRN,sz=9); dr+=1
        wsd.cell(row=dr,column=1,value="CLIENT").font=_ft(DK); wsd.cell(row=dr,column=2,value="PNB").font=_ft(DK)
        for ci in range(1,3): wsd.cell(row=dr,column=ci).fill=_f(GRN2)
        top_p=df.sort_values(cpr,ascending=False).head(10)
        for _,r in top_p.iterrows():
            dr+=1
            wsd.cell(row=dr,column=1,value=r['RAISON_SOCIALE']).font=Font(name='Arial',size=8)
            wsd.cell(row=dr,column=2,value=round(r[cpr],0)).number_format='#,##0 €'
        # Table 4: Top 10 FLUX
        dr+=2
        wsd.cell(row=dr,column=1,value=f"TOP 10 FLUX {ar}").font=_ft(GRN,sz=9); dr+=1
        wsd.cell(row=dr,column=1,value="CLIENT").font=_ft(DK); wsd.cell(row=dr,column=2,value="FLUX").font=_ft(DK)
        for ci in range(1,3): wsd.cell(row=dr,column=ci).fill=_f(GRN2)
        top_f=df.sort_values(cfr,ascending=False).head(10)
        for _,r in top_f.iterrows():
            dr+=1
            wsd.cell(row=dr,column=1,value=r['RAISON_SOCIALE']).font=Font(name='Arial',size=8)
            wsd.cell(row=dr,column=2,value=round(r[cfr],0)).number_format='#,##0 €'
        # KPIs
        dr+=2
        wsd.cell(row=dr,column=1,value="INDICATEURS").font=_ft(DK); wsd.cell(row=dr,column=1).fill=_f(GRN2)
        for ki,(kl,kv) in enumerate([
            ("Actifs avec flux",n_act),("Alertes",n_al),
            (f"Flux total {ar}",df[cfr].sum()),(f"PNB total {ar}",df[cpr].sum()),
        ]):
            wsd.cell(row=dr+1+ki,column=1,value=kl).font=Font(name='Arial',size=8,bold=True)
            c=wsd.cell(row=dr+1+ki,column=2,value=kv); c.font=Font(name='Arial',size=8)
            if isinstance(kv,float): c.number_format='#,##0 €'
        wsd.column_dimensions['A'].width=28; wsd.column_dimensions['B'].width=16; wsd.column_dimensions['C'].width=12

        # SAVE + PATCH
        self._p(0.96,"Sauvegarde..."); wb.save(op)
        self._p(0.98,"Patch Dynamic Arrays..."); self._patch(op)
        self._p(1.0,f"Terminé ! {os.path.basename(op)}")
        return {'aids': len(aids), 'n_act': n_act, 'n_al': n_al, 'op': op}

    @staticmethod
    def _patch(path):
        tmp=tempfile.mkdtemp()
        try:
            with zipfile.ZipFile(path,'r') as z: z.extractall(tmp)
            meta=('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                  '<metadata xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                  'xmlns:xda="http://schemas.microsoft.com/office/spreadsheetml/2017/dynamicarray">'
                  '<metadataTypes count="1"><metadataType name="XLDAPR" minSupportedVersion="120000" '
                  'copy="1" pasteAll="1" pasteValues="1" merge="1" splitFirst="1" rowColShift="1" '
                  'clearFormats="1" clearComments="1" assign="1" coerce="1" cellMeta="1"/></metadataTypes>'
                  '<futureMetadata name="XLDAPR" count="1"><bk><extLst><ext uri="{bdbb8cdc-fa1e-496e-a857-3c3f30c029c3}">'
                  '<xda:dynamicArrayProperties fDynamic="1" fCollapsed="0"/></ext></extLst></bk></futureMetadata>'
                  '<cellMetadata count="1"><bk><rc t="1" v="0"/></bk></cellMetadata></metadata>')
            with open(os.path.join(tmp,'xl','metadata.xml'),'w',encoding='utf-8') as f: f.write(meta)
            rp=os.path.join(tmp,'xl','_rels','workbook.xml.rels')
            with open(rp,'r',encoding='utf-8') as f: c=f.read()
            if 'sheetMetadata' not in c:
                c=c.replace('</Relationships>','<Relationship Id="rIdMeta" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sheetMetadata" Target="metadata.xml"/></Relationships>')
                with open(rp,'w',encoding='utf-8') as f: f.write(c)
            ct=os.path.join(tmp,'[Content_Types].xml')
            with open(ct,'r',encoding='utf-8') as f: c=f.read()
            if 'metadata.xml' not in c:
                c=c.replace('</Types>','<Override PartName="/xl/metadata.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheetMetadata+xml"/></Types>')
                with open(ct,'w',encoding='utf-8') as f: f.write(c)
            import re
            sp=os.path.join(tmp,'xl','worksheets','sheet2.xml')
            if os.path.exists(sp):
                with open(sp,'r',encoding='utf-8') as f: c=f.read()
                c=re.sub(r'(<c r="A8")',r'\1 cm="1"',c)
                with open(sp,'w',encoding='utf-8') as f: f.write(c)
            os.remove(path)
            with zipfile.ZipFile(path,'w',zipfile.ZIP_DEFLATED) as z:
                for root,dirs,files in os.walk(tmp):
                    for fn in files: z.write(os.path.join(root,fn),os.path.relpath(os.path.join(root,fn),tmp))
        finally: shutil.rmtree(tmp,ignore_errors=True)

    def run(self) -> Dict[str, Any]:
        """Orchestration CLI : verifie les entrees, derive la config, lance le worker."""
        for k in ["PRGM", "ACHETEUR"]:
            if not self.files[k] or not Path(self.files[k]).is_file():
                raise FileNotFoundError(f"Source {k} introuvable : {self.files[k]}")
        for use, k in [(self.use_parc_get(), "PARC"), (self.use_acc_get(), "ACCOUNT"), (self.use_sal_get(), "SALES")]:
            if use and (not self.files[k] or not Path(self.files[k]).is_file()):
                raise FileNotFoundError(f"Source {k} introuvable : {self.files[k]}")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        cfg = self._resolve_config()
        return self._work(cfg)


def main() -> int:
    parser = argparse.ArgumentParser(
        prog="01.TXVLV.py",
        description="CPC ANIMATION COMMERCIALE v16 [TXVLV] — reporting animation commerciale (CLI, sans GUI).",
    )
    parser.add_argument("--source-prgm", required=True, help="Fichier PRGM Worldline (41 colonnes) [OBLIGATOIRE]")
    parser.add_argument("--source-acheteur", required=True, help="Fichier ACHETEUR Casper/Noyan (72 colonnes) [OBLIGATOIRE]")
    parser.add_argument("--source-parc", default=None, help="Fichier PARC_CLIENT [OPTIONNEL]")
    parser.add_argument("--source-account", default=None, help="Fichier ACCOUNT (IBAN->GA/RMPM) [OPTIONNEL]")
    parser.add_argument("--source-sales", default=None, help="Fichier SALES (GA->Sales->GI) [OPTIONNEL]")
    parser.add_argument("--output-dir", required=True, help="Répertoire de sortie du XLSX [OBLIGATOIRE]")
    parser.add_argument("--output-filename", required=True, help="Nom du fichier XLSX produit [OBLIGATOIRE]")
    parser.add_argument("--seuil-pct", default="10", help="Seuil d'alerte en %% (défaut 10)")
    parser.add_argument("--seuil-eur", default="100000", help="Seuil d'alerte en EUR (défaut 100000)")
    parser.add_argument("--annee-ref", default=None, help="Année de référence N (défaut: dernière détectée)")
    parser.add_argument("--annee-prec", default=None, help="Année précédente N-1 (défaut: avant-dernière détectée)")
    parser.add_argument("--mois-fin", default=None, help="Mois de fin AAAAMM ou AAAA-MM (défaut: dernier détecté)")

    try:
        args = parser.parse_args()
    except SystemExit as exc:
        return int(exc.code) if exc.code is not None else 2

    try:
        if not OPENPYXL:
            raise RuntimeError("openpyxl manquant : installez 'openpyxl' pour produire le XLSX.")
        app = App(args)
        res = app.run()
        print(f"[OK] Programmes: {res['aids']:,} | Actifs: {res['n_act']:,} | "
              f"Alertes: {res['n_al']:,} -> {res['op']}")
        return 0
    except Exception as exc:
        import traceback
        traceback.print_exc()
        print(f"[ERREUR] {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
