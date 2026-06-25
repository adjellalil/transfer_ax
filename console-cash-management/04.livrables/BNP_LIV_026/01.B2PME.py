
"""
TITRE [B2PME]
=============
CCO PME DASHBOARD v14 [B2PME] — version CLI autonome (sans interface graphique).

DESCRIPTION
-----------
Outil en ligne de commande qui consolide le parc client PME BNP Paribas (Direction
Monetique) avec les flux et stocks CCO, les programmes Worldline (PRGM) et les notes
de frais MONEXT (NDF), puis produit un DATASET CSV sacralise et un classeur XLSX a
4 feuilles (SOURCE / DATA pivot / ANALYSE RECHERCHEX / DASHBOARD).

La logique metier est strictement identique a l'application graphique d'origine
(customtkinter) : aucun calcul, pivot, formule RECHERCHEX ni mise en forme n'a ete
modifie. Seules les couches GUI (customtkinter, tkinter, filedialog, messagebox,
threading, mainloop) ont ete retirees et remplacees par argparse + print + chemins
de sortie explicites.

ARCHITECTURE 4 FEUILLES (inchangee) :
  Sheet 1 SOURCE    : en-tetes uniquement, donnees a coller manuellement depuis le CSV
  Sheet 2 DATA      : tableau pivote calcule par Python (RS x PRODUIT, colonnes = mois)
  Sheet 3 ANALYSE   : encadre client + liste deroulante + formules RECHERCHEX
  Sheet 4 DASHBOARD : valeurs en dur + graphiques lineaires

SOURCES REQUISES
----------------
  PARC_CLIENT            (--parc)              [obligatoire]  Parc client (segment PME)
  FLUX AMOUNT  [N3096]   (--flux-amount)       [obligatoire]  Flux CCO agrege - montants
  FLUX NUMBER  [N3096]   (--flux-number)       [obligatoire]  Flux CCO agrege - nombres
  FLUX INTERCHANGE       (--flux-interchange)  [obligatoire]  Flux CCO agrege - interchange
  STOCK        [66Z3Q]   (--stock)             [obligatoire]  Stock cartes agrege
  WORLDLINE PRGM [6AICC] (--worldline)         [obligatoire]  Programmes Worldline consolides
  MONEXT       [K2P8N]   (--monext)            [obligatoire]  Notes de frais MONEXT consolidees
  MATCHING_USAGE         (--matching-usage)    [optionnel]    Mapping usage Worldline
  FEF Eric Tardy         (--fef)               [optionnel]    Liste FEF (adossement)
  SCOPE Olivier Attali   (--scope)             [optionnel]    Codes APE (holding)

OUTPUTS PRODUITS
----------------
  <output-dir>/<output-filename>.csv   : DATASET sacralise (RP/RC/RMPM/SIREN = "VALEUR")
  <output-dir>/<output-filename>.xlsx  : Classeur 4 feuilles (SOURCE/DATA/ANALYSE/DASHBOARD)

ARGUMENTS CLI
-------------
  --parc PATH                (obligatoire)
  --flux-amount PATH         (obligatoire)
  --flux-number PATH         (obligatoire)
  --flux-interchange PATH    (obligatoire)
  --stock PATH               (obligatoire)
  --worldline PATH           (obligatoire)
  --monext PATH              (obligatoire)
  --matching-usage PATH      (optionnel)
  --fef PATH                 (optionnel)
  --scope PATH               (optionnel)
  --pme-codes STR            (optionnel, defaut "ER,ME")
  --stock-active STR         (optionnel, defaut "Active")
  --ape-holding STR          (optionnel, codes APE holding supplementaires)
  --no-default-ape           (optionnel, desactive les 4 codes APE holding par defaut)
  --output-dir PATH          (obligatoire)
  --output-filename STR      (obligatoire, sans extension)

DECOMPOSITION (arborescent)
---------------------------
  main()
   |-- parse argparse
   |-- App(args)
   |    `-- __init__        : affecte files{} + parametres depuis args
   `-- App.worker()
        |-- 1. PARC -> PME      (filtre segment, cles RP/RC/RS)
        |-- 2. FEF              (adossement, optionnel)
        |-- 3. FLUX CCO         (amount / number / interchange)
        |-- 4. STOCK            (cartes actives)
        |-- 5. WORLDLINE        (PRGM + matching usage)
        |-- 6. MONEXT           (PNB, NDF, flux)
        |-- 7. Periodes         (union YYYYMM)
        |-- 8. Pre-agregation   (dictionnaires O(1))
        |-- 9. Construction DATASET (CSV)
        |-- 10. Export CSV      (QUOTE_NONE, sacralisation)
        |-- 11. Export XLSX     -> _export_xlsx()
        `-- _export_xlsx()      : SOURCE / DATA pivot / ANALYSE RECHERCHEX / DASHBOARD

SACRALISATION - 2 niveaux (regles agent_skills generation_python) :
  NIVEAU 1 sacralise_id()  : preserve les zeros initiaux -> ="VALEUR"
  NIVEAU 2 protect_id()    : sacralise tout ID numerique (evite notation scientifique)
  Export CSV : quoting=csv.QUOTE_NONE, escapechar='\\'

BNP Paribas Cash Management - Direction Monetique
Avril 2026
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ─── Résolution auto des sources + lecture DuckDB (inliné depuis 06.fonctions) ──
def _find_sources_root():
    here = Path(__file__).resolve()
    for d in [here] + list(here.parents):
        c = d / "03.sources"
        if c.is_dir():
            return c
    return None


def resolve_source(name, required=False):
    root = _find_sources_root()
    if not root:
        if required:
            raise FileNotFoundError(f"Dossier 03.sources introuvable depuis {__file__}")
        return None
    folder = next((m for m in root.glob("*/" + name) if m.is_dir()), None)
    if not folder:
        if required:
            raise FileNotFoundError(f"Source introuvable : 03.sources/*/{name}")
        return None
    exts = (".csv", ".xlsx", ".xls", ".txt")
    files = [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() in exts]
    if not files:
        if required:
            raise FileNotFoundError(f"Aucun fichier source dans {folder}")
        return None

    def _np(f):
        m = re.match(r"\s*(\d+)", f.name)
        return (int(m.group(1)) if m else -1, f.name)
    return max(files, key=_np)


def resolve_source_kw(name, keyword, required=False):
    """Comme resolve_source mais ne garde que les fichiers dont le NOM contient `keyword`
    (sert aux 3 datasets CCO_FLUX : AMOUNT / NUMBER / INTERCHANGE dans un même dossier)."""
    root = _find_sources_root()
    folder = next((m for m in (root.glob("*/" + name) if root else []) if m.is_dir()), None)
    if not folder:
        if required:
            raise FileNotFoundError(f"Source introuvable : 03.sources/*/{name}")
        return None
    exts = (".csv", ".xlsx", ".xls", ".txt")
    files = [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() in exts
             and keyword.lower() in f.name.lower()]
    if not files:
        if required:
            raise FileNotFoundError(f"Fichier '{keyword}' introuvable dans 03.sources/*/{name}")
        return None

    def _np(f):
        m = re.match(r"\s*(\d+)", f.name)
        return (int(m.group(1)) if m else -1, f.name)
    return max(files, key=_np)


def _read_duck(path, nrows=None):
    p = Path(path)
    if nrows is not None or p.suffix.lower() not in (".csv", ".txt", ""):
        return None
    try:
        import duckdb
        return duckdb.read_csv(str(p), all_varchar=True).df()
    except Exception:
        return None


VERSION_ID = "B2PME"

DEFAULT_PME_CODES    = "ER,ME"
DEFAULT_STOCK_ACTIVE = "Active"
RS_MIN_MATCH         = 14
SEP_KEY              = "§"   # séparateur clé RECHERCHEX

DEFAULT_APE_HOLDING = {
    "6420Z": "Holding pure / passive",
    "7010Z": "Holding animatrice / active",
    "7022Z": "Holding conseil",
    "6630Z": "Holding gestion de fonds",
}

MOIS_FR_ORDER = [
    "JANVIER","FEVRIER","MARS","AVRIL","MAI","JUIN",
    "JUILLET","AOUT","SEPTEMBRE","OCTOBRE","NOVEMBRE","DECEMBRE"
]
MOIS_FR_NUM = {m: i+1 for i, m in enumerate(MOIS_FR_ORDER)}

# Positions par défaut PARC (1-indexed) — identiques aux autres programmes
PARC_POS = {
    "rp": 1, "rc": 14, "rmpm": 6, "siren": 7,
    "rs": 8, "segment": 9, "code_ga": 11, "nom_ga": 12,
}
SCOPE_POS  = {"id_rp": 7, "rs": 10, "code_ga": 17, "nom_ga": 18, "siren": 9, "code_ape": 15}
MONEXT_POS = {"mois": 1, "rs": 4, "id_rp": 9, "id_rc": 10,
              "depenses": 15, "retraits": 17,
              "pnb_first": 19, "pnb_last": 55, "pnb_excl": 33,
              "pnb_ndf_1": 22, "pnb_ndf_2": 55}
WL_POS     = {"mois": 2, "id_prog": 4, "produit": 5, "rs": 8, "id_rc": 40}  # RC en col 40 (col 41 = DIFFERE ; aligne sur le consensus, etait 41)
WL_PNB_DEF = [30, 31, 32, 33, 34, 35, 36]
USAGE_POS  = {"id_prog": 2, "produit": 3, "usage": 4}


# =============================================================================
# UTILITAIRES — patterns validés (K2P8N, 6AICC, C2SME)
# =============================================================================

def load_csv_smart(path, nrows=None):
    _d = _read_duck(path, nrows)
    if _d is not None:
        return _d
    for sep in [";", ",", "\t"]:
        for enc in ["utf-8-sig", "utf-8", "latin1", "cp1252"]:
            try:
                df = pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                 keep_default_na=False, na_values=[],
                                 on_bad_lines="skip", nrows=5)
                if df.shape[1] > 1:
                    return pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                       keep_default_na=False, na_values=[],
                                       on_bad_lines="skip", nrows=nrows)
            except Exception:
                continue
    return pd.read_csv(path, sep=None, engine="python", dtype=str,
                       keep_default_na=False, na_values=[],
                       on_bad_lines="skip", nrows=nrows)


def clean_id_safe(series):
    s = series.astype(str).str.strip()
    s = s.replace(["", "nan", "NaN", "None", "NULL", "NA", "N/A"], "")
    mask = s.str.startswith('="') & s.str.endswith('"')
    s = s.where(~mask, s.str[2:-1])
    s = s.str.lstrip("'")
    mask2 = s.str.endswith(".0") & s.str[:-2].str.isdigit()
    s = s.where(~mask2, s.str[:-2])
    return s.str.strip()


def clean_id_strip0(series):
    s = clean_id_safe(series)
    stripped = s.str.lstrip("0")
    return stripped.where(stripped != "", s)


# NIVEAU 1 — préserve zéros initiaux
def sacralise_id(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s == "" or s.lower() in ("nan","none","null","na","n/a"):
        return ""
    if s.startswith('="') and s.endswith('"'):
        return s
    if s.endswith(".0") and s[:-2].replace("-","").isdigit():
        s = s[:-2]
    s = s.lstrip("'")
    return f'="{s}"'


# NIVEAU 2 — sacralise tout ID numérique (évite notation scientifique)
def protect_id(val):
    if pd.isna(val) or str(val).strip() == "":
        return ""
    s = str(val).strip()
    if s.startswith('="') and s.endswith('"'):
        return s
    if s.isdigit():
        return f'="{s}"'
    if s.startswith("0") and len(s) > 1 and s[1:].isdigit():
        return f'="{s}"'
    return s


def sacralise_vec(series):
    """Vectorisé — niveau 1."""
    return series.apply(sacralise_id)


def protect_vec(series):
    """Vectorisé — niveau 2."""
    return series.apply(protect_id)


def norm_rs(series):
    def _n(val):
        if pd.isna(val) or str(val).strip() == "":
            return ""
        s = str(val).strip().upper()
        s = unicodedata.normalize("NFD", s)
        return "".join(c for c in s if unicodedata.category(c) != "Mn")
    return series.apply(_n)


def to_float(series):
    s = series.astype(str)
    for ch in ['"', "'", " ", "\xa0", " ", "€", "EUR"]:
        s = s.str.replace(ch, "", regex=False)
    mask = s.str.endswith("-")
    s = s.where(~mask, "-" + s.str[:-1])
    s = s.str.replace(",", ".", regex=False)
    return pd.to_numeric(s, errors="coerce").fillna(0.0)


def parse_date_cco(date_str):
    """JANVIER_2025 → (2025, 1, 'JANVIER')"""
    if not date_str or str(date_str).strip() == "":
        return None, None, None
    s = str(date_str).strip().upper()
    for mois_fr, num in MOIS_FR_NUM.items():
        if mois_fr in s:
            m = re.search(r"(\d{4})", s)
            if m:
                return int(m.group(1)), num, mois_fr
    return None, None, None


def parse_mois_yyyymm(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.startswith('="') and s.endswith('"'):
        s = s[2:-1].strip()
    s = s.lstrip("'").strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    if re.fullmatch(r"\d{6}", s):
        return s
    m = re.match(r"^(\d{1,2})[/\-\.](\d{4})$", s)
    if m:
        mo, yr = int(m.group(1)), m.group(2)
        if 1 <= mo <= 12:
            return f"{yr}{str(mo).zfill(2)}"
    m = re.match(r"^(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})$", s)
    if m:
        mo, yr = int(m.group(2)), m.group(3)
        if 1 <= mo <= 12:
            return f"{yr}{str(mo).zfill(2)}"
    for mois_fr, num in MOIS_FR_NUM.items():
        if mois_fr in s.upper():
            m2 = re.search(r"(\d{4})", s)
            if m2:
                return f"{m2.group(1)}{str(num).zfill(2)}"
    return ""


def yyyymm_to_label(yyyymm):
    """202501 → JANVIER_2025"""
    s = str(yyyymm)
    if len(s) != 6 or not s.isdigit():
        return s
    yr, mo = s[:4], int(s[4:6])
    if 1 <= mo <= 12:
        return f"{MOIS_FR_ORDER[mo-1]}_{yr}"
    return s


def detect_col(df, keywords, exact=False):
    for col in df.columns:
        nh = col.strip().lower().replace(" ", "").replace("_", "")
        for kw in keywords:
            kw_n = kw.lower().replace(" ", "").replace("_", "")
            if exact and nh == kw_n:
                return col
            if not exact and kw_n in nh:
                return col
    return None


def build_fef_lookup(df_fef, col_rp, col_rs=None):
    keys_rp = clean_id_strip0(df_fef[col_rp])
    set_rp  = set(keys_rp[keys_rp != ""].unique())
    dict_rs = {}
    if col_rs and col_rs in df_fef.columns:
        for k, r in zip(keys_rp, norm_rs(df_fef[col_rs])):
            if r and r not in dict_rs:
                dict_rs[r] = k
    return set_rp, dict_rs


def match_fef(k_rp, k_rs, set_rp, dict_rs):
    if k_rp and k_rp in set_rp:
        return True
    if k_rs:
        if k_rs in dict_rs:
            return True
        if len(k_rs) >= RS_MIN_MATCH:
            for rs_fef in dict_rs:
                if len(rs_fef) >= RS_MIN_MATCH and (k_rs in rs_fef or rs_fef in k_rs):
                    return True
    return False


def build_agg_dict(df, key_cols, val_col):
    valid = [c for c in key_cols if c in df.columns]
    if not valid or val_col not in df.columns:
        return {}
    df2 = df.dropna(subset=[valid[0]]).copy()
    df2 = df2[df2[valid[0]] != ""]
    if df2.empty:
        return {}
    return df2.groupby(key_cols)[val_col].sum().to_dict()


# =============================================================================
# SHIM CLI — remplace les widgets GUI (.get()) par des valeurs simples
# Le corps du worker reste identique : il continue d'appeler .get() / _gcol().
# =============================================================================

class _Val:
    """Substitut minimal d'un widget Entry/ComboBox : expose .get()."""
    def __init__(self, value: str = ""):
        self._value = value

    def get(self) -> str:
        return self._value


class _Bool:
    """Substitut minimal d'une BooleanVar : expose .get()."""
    def __init__(self, value: bool):
        self._value = value

    def get(self) -> bool:
        return self._value


# =============================================================================
# APPLICATION (CLI) — logique métier identique à la version GUI
# =============================================================================

class CCOPMEDashboard_B2PME:

    def __init__(self, args: argparse.Namespace) -> None:
        # ── Fichiers sources (mêmes clés que la version GUI) ──────────────
        self.files: dict[str, str] = {
            "parc":             args.parc or "",
            "flux_amount":      args.flux_amount or "",
            "flux_number":      args.flux_number or "",
            "flux_interchange": args.flux_interchange or "",
            "stock":            args.stock or "",
            "worldline":        args.worldline or "",
            "matching_usage":   args.matching_usage or "",
            "monext":           args.monext or "",
            "fef":              args.fef or "",
            "scope":            args.scope or "",
        }

        # ── Sorties (remplacent filedialog.asksaveasfilename) ─────────────
        self.output_dir: Path = Path(args.output_dir)
        self.output_filename: str = args.output_filename

        # ── Paramètres métier (remplacent les Entry GUI, via shim .get()) ─
        self.pme_codes = _Val(args.pme_codes)
        self.stk_active = _Val(args.stock_active)

        # Codes APE holding : cases à cocher GUI -> --no-default-ape,
        # champ "codes supplémentaires" GUI -> --ape-holding
        self.ape_vars: dict[str, _Bool] = {
            code: _Bool(not args.no_default_ape)
            for code in DEFAULT_APE_HOLDING
        }
        self.ape_custom = _Val(args.ape_holding)

        # ── Mapping/sélection de colonnes ─────────────────────────────────
        # La version GUI proposait un mapping manuel des colonnes (ComboBox).
        # En CLI cette étape visuelle est supprimée : _gcol() renvoie None,
        # ce qui force l'utilisation des positions par défaut de l'UI
        # (PARC_POS / MONEXT_POS / WL_POS), strictement identiques.
        # Colonnes MONEXT PNB (première/dernière/exclusion/interchange) ->
        # valeurs par défaut UI ; Interchange MONEXT = "(Aucune)" comme l'UI.
        self.mx_first = _Val("")   # vide -> MONEXT_POS["pnb_first"] (19)
        self.mx_last  = _Val("")   # vide -> MONEXT_POS["pnb_last"]  (55)
        self.mx_excl  = _Val("")   # vide -> MONEXT_POS["pnb_excl"]  (33)
        self.mx_ic    = _Val("(Aucune)")  # défaut UI : aucune colonne interchange

        # PNB Worldline : défaut UI = colonnes 30->36 (WL_PNB_DEF).
        # Représentées par des shims renvoyant le nom de colonne attendu ;
        # résolu dynamiquement dans worker() (df_wl non chargé ici).
        self.wl_pnb_vars: list[_Val] = []  # rempli dans worker() après chargement

    # ──────────────────────────────────────────────────────────────────────────
    # Remplacement des hooks GUI
    # ──────────────────────────────────────────────────────────────────────────

    def _prog(self, val: float, txt: str) -> None:
        # Remplace progress.set + lbl_status.configure : simple affichage console
        print(f"[{int(val * 100):3d}%] {txt}")

    def _gcol(self, mk: str) -> None:
        # Mapping manuel de colonnes supprimé (étape visuelle) -> défaut UI.
        return None

    def _cn(self, s):
        return s.split(". ", 1)[1] if ". " in s else s

    def _ci(self, s):
        try:
            return int(s.split(". ")[0])
        except Exception:
            return 0

    # ──────────────────────────────────────────────────────────────────────────
    # WORKER — logique métier VERBATIM (aucun calcul modifié)
    # ──────────────────────────────────────────────────────────────────────────

    def worker(self) -> None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        pme_codes  = [c.strip().upper()
                      for c in self.pme_codes.get().split(",") if c.strip()]
        stk_active = self.stk_active.get().strip().lower()

        ape_holding = set()
        for code, var in self.ape_vars.items():
            if var.get():
                ape_holding.add(code.strip().upper())
        if self.ape_custom.get().strip():
            for c in self.ape_custom.get().split(","):
                ape_holding.add(c.strip().upper())

        # ── 1. PARC → PME ────────────────────────────────────────────────
        self._prog(0.02, "Chargement PARC_CLIENT...")
        df_parc = load_csv_smart(self.files["parc"])
        pc_cols = list(df_parc.columns)
        def pc(mk, pos):
            col = self._gcol(mk)
            if col and col in df_parc.columns:
                return col
            return pc_cols[pos-1] if 0 < pos <= len(pc_cols) else pc_cols[0]

        col_rp    = pc("parc_rp",    PARC_POS["rp"])
        col_rc    = pc("parc_rc",    PARC_POS["rc"])
        col_rmpm  = pc("parc_rmpm",  PARC_POS["rmpm"])
        col_siren = pc("parc_siren", PARC_POS["siren"])
        col_rs    = pc("parc_rs",    PARC_POS["rs"])
        col_seg   = pc("parc_seg",   PARC_POS["segment"])
        col_cga   = pc("parc_cga",   PARC_POS["code_ga"])
        col_nga   = pc("parc_nga",   PARC_POS["nom_ga"])

        df_parc["_K_RP"]   = clean_id_strip0(df_parc[col_rp])
        df_parc["_K_RC"]   = clean_id_strip0(df_parc[col_rc])
        df_parc["_K_RS"]   = norm_rs(df_parc[col_rs])
        df_parc["_SEG"]    = df_parc[col_seg].astype(str).str.strip().str.upper()
        df_parc["_RMPM"]   = clean_id_safe(df_parc[col_rmpm])
        df_parc["_SIREN"]  = clean_id_safe(df_parc[col_siren])

        df_pme = df_parc[df_parc["_SEG"].isin(pme_codes)].copy().reset_index(drop=True)
        n_pme  = len(df_pme)
        self._prog(0.04, f"{n_pme:,} clients PME...")
        if n_pme == 0:
            raise ValueError(
                f"Aucun client PME avec : {', '.join(pme_codes)}")

        set_rp_pme = set(df_pme["_K_RP"][df_pme["_K_RP"] != ""].unique())
        set_rc_pme = set(df_pme["_K_RC"][df_pme["_K_RC"] != ""].unique())

        df_dedup   = (df_pme[df_pme["_K_RP"] != ""]
                      .drop_duplicates("_K_RP").copy())
        rc_to_rp   = dict(zip(df_pme["_K_RC"], df_pme["_K_RP"]))

        parc_info  = df_dedup.set_index("_K_RP")[
            [col_rs, col_rc, col_rmpm, col_siren,
             col_seg, col_cga, col_nga, "_K_RS", "_RMPM", "_SIREN"]
        ].to_dict("index")

        # APE depuis SCOPE
        scope_ape = {}
        if self.files.get("scope"):
            self._prog(0.05, "SCOPE...")
            df_sc = load_csv_smart(self.files["scope"])
            sc_c  = list(df_sc.columns)
            def scc(p): return sc_c[p-1] if 0 < p <= len(sc_c) else None
            c_sc_rp  = scc(SCOPE_POS["id_rp"])
            c_sc_ape_cfg = self._gcol("scope_ape")
            c_sc_ape = (c_sc_ape_cfg if c_sc_ape_cfg and c_sc_ape_cfg in df_sc.columns
                        else scc(SCOPE_POS["code_ape"]))
            if c_sc_rp and c_sc_ape:
                df_sc["_K_RP"] = clean_id_strip0(df_sc[c_sc_rp])
                scope_ape = (df_sc[df_sc["_K_RP"] != ""]
                             .drop_duplicates("_K_RP")
                             .set_index("_K_RP")[c_sc_ape]
                             .astype(str).str.strip().str.upper()
                             .to_dict())

        # ── 2. FEF ───────────────────────────────────────────────────────
        set_rp_fef, dict_rs_fef = set(), {}
        if self.files.get("fef"):
            self._prog(0.06, "FEF...")
            df_fef = load_csv_smart(self.files["fef"])
            fc = list(df_fef.columns)
            if fc:
                set_rp_fef, dict_rs_fef = build_fef_lookup(
                    df_fef, fc[0], fc[2] if len(fc) >= 3 else None)

        fef_flag = {rp: match_fef(rp,
                        parc_info.get(rp, {}).get("_K_RS", ""),
                        set_rp_fef, dict_rs_fef)
                    for rp in set_rp_pme}

        # ── 3. FLUX CCO ──────────────────────────────────────────────────
        def load_flux(file_key, prog_val):
            if not self.files.get(file_key):
                return pd.DataFrame(columns=["_K_RP","_PRODUCT","_YYYYMM","_VAL"])
            self._prog(prog_val, f"Chargement {file_key}...")
            df = load_csv_smart(self.files[file_key])
            c_rp  = detect_col(df, ["id_rp","idrp"])
            c_pr  = detect_col(df, ["product"])
            c_dt  = detect_col(df, ["date"])
            c_val = detect_col(df, ["valeur","value"])
            if not c_rp or not c_val:
                return pd.DataFrame(columns=["_K_RP","_PRODUCT","_YYYYMM","_VAL"])
            df["_K_RP"]    = clean_id_strip0(df[c_rp])
            df["_PRODUCT"] = df[c_pr].astype(str).str.strip() if c_pr else ""
            df["_VAL"]     = to_float(df[c_val])
            # Interchange stocké en négatif dans la source → abs() systématique
            if file_key == "flux_interchange":
                df["_VAL"] = df["_VAL"].abs()
            if c_dt:
                parsed = df[c_dt].apply(parse_date_cco)
                df["_AN"] = [p[0] for p in parsed]
                df["_MO"] = [p[1] for p in parsed]
                df["_YYYYMM"] = df.apply(
                    lambda r: f"{int(r['_AN'])}{str(int(r['_MO'])).zfill(2)}"
                    if r["_AN"] and r["_MO"] else "", axis=1)
            else:
                df["_YYYYMM"] = ""
            return df[df["_K_RP"].isin(set_rp_pme)][
                ["_K_RP","_PRODUCT","_YYYYMM","_VAL"]].copy()

        df_amt = load_flux("flux_amount",      0.08)
        df_num = load_flux("flux_number",      0.12)
        df_int = load_flux("flux_interchange", 0.16)

        produits_cco = sorted(
            df_amt["_PRODUCT"][df_amt["_PRODUCT"] != ""].unique().tolist())

        # ── 4. STOCK ─────────────────────────────────────────────────────
        self._prog(0.20, "STOCK...")
        stk_dict = {}
        df_stk = load_csv_smart(self.files["stock"])
        c_s_rp  = detect_col(df_stk, ["id_rp","idrp"])
        c_s_pr  = detect_col(df_stk, ["product"])
        c_s_dt  = detect_col(df_stk, ["data"], exact=True)
        c_s_stk = detect_col(df_stk, ["stock"])
        c_s_d   = detect_col(df_stk, ["date"])
        if c_s_rp and c_s_stk:
            df_stk["_K_RP"]    = clean_id_strip0(df_stk[c_s_rp])
            df_stk["_PRODUCT"] = df_stk[c_s_pr].astype(str).str.strip() if c_s_pr else ""
            df_stk["_DATA"]    = df_stk[c_s_dt].astype(str).str.strip().str.lower() if c_s_dt else ""
            df_stk["_STOCK"]   = to_float(df_stk[c_s_stk])
            if c_s_d:
                parsed = df_stk[c_s_d].apply(parse_date_cco)
                df_stk["_AN"] = [p[0] for p in parsed]
                df_stk["_MO"] = [p[1] for p in parsed]
            else:
                df_stk["_AN"] = None; df_stk["_MO"] = None
            df_stk = df_stk[df_stk["_K_RP"].isin(set_rp_pme)].copy()
            df_act = df_stk[df_stk["_DATA"] == stk_active].dropna(
                subset=["_AN","_MO"]).copy()
            if not df_act.empty:
                df_act["_AN"] = df_act["_AN"].astype(int)
                df_act["_MO"] = df_act["_MO"].astype(int)
                # Index par YYYYMM pour avoir 1 valeur par mois (pas une seule par an)
                df_act["_YYYYMM"] = (df_act["_AN"].astype(str)
                                      + df_act["_MO"].astype(str).str.zfill(2))
                stk_dict = (df_act.groupby(["_K_RP","_PRODUCT","_YYYYMM"])["_STOCK"]
                            .sum().to_dict())

        # ── 5. WORLDLINE ─────────────────────────────────────────────────
        self._prog(0.24, "WORLDLINE...")
        df_wl = load_csv_smart(self.files["worldline"])
        wl_c  = list(df_wl.columns)
        def wlc(p): return wl_c[p-1] if 0 < p <= len(wl_c) else None

        # Colonnes Worldline — depuis UI si configurées, sinon défaut
        c_wl_rc   = self._gcol("wl_rc")      or wlc(WL_POS["id_rc"])
        c_wl_mo   = self._gcol("wl_mois")    or wlc(WL_POS["mois"])
        c_wl_pr   = self._gcol("wl_produit") or wlc(WL_POS["produit"])
        c_wl_rs   = wlc(WL_POS["rs"])
        c_wl_idp  = wlc(WL_POS["id_prog"])
        c_wl_dep1 = wlc(28)
        c_wl_dep2 = wlc(29)

        # PNB Worldline : sélection manuelle GUI supprimée -> défaut UI 30->36
        # (WL_PNB_DEF). On reconstruit les shims à partir des colonnes réelles.
        numbered_wl = [f"{i+1:02d}. {c}" for i, c in enumerate(wl_c)]
        self.wl_pnb_vars = []
        for i in range(7):
            pos = WL_PNB_DEF[i] if i < len(WL_PNB_DEF) else 0
            self.wl_pnb_vars.append(
                _Val(numbered_wl[pos-1] if 0 < pos <= len(numbered_wl) else "(Aucune)"))

        wl_pnb_names = [self._cn(cb.get()) for cb in self.wl_pnb_vars
                        if cb.get() and cb.get() != "(Aucune)"]

        df_wl["_K_RC"]    = clean_id_strip0(df_wl[c_wl_rc]) if c_wl_rc else ""
        df_wl["_K_RS"]    = norm_rs(df_wl[c_wl_rs]) if c_wl_rs else ""
        df_wl["_PRODUCT"] = df_wl[c_wl_pr].astype(str).str.strip() if c_wl_pr else ""
        df_wl["_IDPROG"]  = df_wl[c_wl_idp].astype(str).str.strip() if c_wl_idp else ""
        df_wl["_DEP1"]    = to_float(df_wl[c_wl_dep1]) if c_wl_dep1 else 0.0
        df_wl["_DEP2"]    = to_float(df_wl[c_wl_dep2]) if c_wl_dep2 else 0.0
        df_wl["_FLUX"]    = df_wl["_DEP1"] + df_wl["_DEP2"]
        if wl_pnb_names:
            for c in wl_pnb_names:
                if c in df_wl.columns:
                    df_wl[c] = to_float(df_wl[c])
            df_wl["_PNB"] = df_wl[
                [c for c in wl_pnb_names if c in df_wl.columns]
            ].sum(axis=1)
        else:
            df_wl["_PNB"] = 0.0
        df_wl["_YYYYMM"] = df_wl[c_wl_mo].apply(parse_mois_yyyymm) if c_wl_mo else ""

        # MATCHING_USAGE
        usage_dict = {}
        if self.files.get("matching_usage"):
            self._prog(0.27, "MATCHING_USAGE...")
            df_mu = load_csv_smart(self.files["matching_usage"])
            mu_c  = list(df_mu.columns)
            def muc(p): return mu_c[p-1] if 0 < p <= len(mu_c) else None
            c_mu_id = muc(USAGE_POS["id_prog"])
            c_mu_pr = muc(USAGE_POS["produit"])
            c_mu_us = muc(USAGE_POS["usage"])
            if c_mu_id and c_mu_us:
                df_mu["_K"] = (df_mu[c_mu_id].astype(str).str.strip()
                               + "|" + (df_mu[c_mu_pr].astype(str).str.strip().str.upper()
                                        if c_mu_pr else ""))
                df_mu["_U"] = df_mu[c_mu_us].astype(str).str.strip()
                usage_dict  = (df_mu[df_mu["_K"] != "|"]
                               .drop_duplicates("_K")
                               .set_index("_K")["_U"].to_dict())
        df_wl["_PROD_MAP"] = df_wl.apply(
            lambda r: usage_dict.get(
                r["_IDPROG"] + "|" + r["_PRODUCT"].upper(), r["_PRODUCT"]),
            axis=1) if usage_dict else df_wl["_PRODUCT"]

        df_wl_pme = df_wl[df_wl["_K_RC"].isin(set_rc_pme)].copy()
        if df_wl_pme.empty:
            rs_pme = set(df_pme["_K_RS"][df_pme["_K_RS"] != ""].unique())
            df_wl_pme = df_wl[df_wl["_K_RS"].isin(rs_pme)].copy()
        df_wl_pme["_K_RP"] = df_wl_pme["_K_RC"].map(rc_to_rp).fillna("")
        df_wl_pme["_YYYYMM"] = df_wl_pme["_YYYYMM"].astype(str)

        produits_wl = sorted(
            df_wl_pme["_PROD_MAP"][df_wl_pme["_PROD_MAP"] != ""].unique().tolist())

        # ── 6. MONEXT ────────────────────────────────────────────────────
        self._prog(0.33, "MONEXT...")
        df_mx = load_csv_smart(self.files["monext"])
        mx_c  = list(df_mx.columns)
        def mxc(p): return mx_c[p-1] if 0 < p <= len(mx_c) else None

        c_mx_mo  = self._gcol("mx_mois") or mxc(MONEXT_POS["mois"])
        c_mx_rp  = self._gcol("mx_rp")   or mxc(MONEXT_POS["id_rp"])
        c_mx_rc  = self._gcol("mx_rc")   or mxc(MONEXT_POS["id_rc"])
        c_mx_dep = self._gcol("mx_dep")  or mxc(MONEXT_POS["depenses"])
        c_mx_ret = self._gcol("mx_ret")  or mxc(MONEXT_POS["retraits"])
        c_mx_ndf1= self._gcol("mx_ndf1") or mxc(MONEXT_POS["pnb_ndf_1"])
        c_mx_ndf2= self._gcol("mx_ndf2") or mxc(MONEXT_POS["pnb_ndf_2"])

        pnb_f = max(0, self._ci(self.mx_first.get()) - 1) if self.mx_first.get() else MONEXT_POS["pnb_first"]-1
        pnb_l = max(0, self._ci(self.mx_last.get())  - 1) if self.mx_last.get()  else MONEXT_POS["pnb_last"]-1
        pnb_range = mx_c[pnb_f:pnb_l+1]

        excl_ch = self.mx_excl.get()
        ic_ch   = self.mx_ic.get()
        excl_col = mx_c[self._ci(excl_ch)-1] \
                   if excl_ch != "(Aucune)" and self._ci(excl_ch) > 0 else None
        ic_col   = mx_c[self._ci(ic_ch)-1] \
                   if ic_ch != "(Aucune)" and self._ci(ic_ch) > 0 else None

        # NB : la version GUI laissait par défaut "Exclure doublon" sur la
        # colonne 33 (MONEXT_POS["pnb_excl"]). En CLI, l'étape de sélection
        # visuelle est supprimée : on conserve le comportement par défaut UI
        # via excl_col = colonne 33.
        if excl_col is None and 0 < MONEXT_POS["pnb_excl"] <= len(mx_c):
            excl_col = mx_c[MONEXT_POS["pnb_excl"] - 1]

        for c in pnb_range:
            if c in df_mx.columns:
                df_mx[c] = to_float(df_mx[c])
        if ic_col and ic_col in df_mx.columns:
            df_mx[ic_col] = to_float(df_mx[ic_col]) * -1

        pnb_filtered = [c for c in pnb_range if c != excl_col and c in df_mx.columns]
        df_mx["_PNB_TOT"] = df_mx[pnb_filtered].sum(axis=1) if pnb_filtered else 0.0
        # Le PNB MONEXT est méthodologiquement positif.
        # Si l'interchange n'a pas été sélectionné comme colonne × -1,
        # il peut tirer _PNB_TOT en négatif → on prend la valeur absolue.
        if not ic_col:
            df_mx["_PNB_TOT"] = df_mx["_PNB_TOT"].abs()

        ndf_cols = []
        for c in [c_mx_ndf1, c_mx_ndf2]:
            if c and c in df_mx.columns and c not in ndf_cols:
                df_mx[c] = to_float(df_mx[c])
                ndf_cols.append(c)
        df_mx["_PNB_NDF"] = df_mx[ndf_cols].sum(axis=1) if ndf_cols else 0.0
        df_mx["_FLUX"]    = ((to_float(df_mx[c_mx_dep]) if c_mx_dep else 0.0)
                             + (to_float(df_mx[c_mx_ret]) if c_mx_ret else 0.0))
        df_mx["_K_RP"]    = clean_id_strip0(df_mx[c_mx_rp]) if c_mx_rp else ""
        df_mx["_K_RC"]    = clean_id_strip0(df_mx[c_mx_rc]) if c_mx_rc else ""
        df_mx["_YYYYMM"]  = df_mx[c_mx_mo].apply(parse_mois_yyyymm) if c_mx_mo else ""
        df_mx["_IS_NDF"]  = df_mx["_PNB_NDF"].abs() > 0.001

        mask_mx = df_mx["_K_RP"].isin(set_rp_pme) | df_mx["_K_RC"].isin(set_rc_pme)
        df_mx_pme = df_mx[mask_mx].copy()
        no_rp = df_mx_pme["_K_RP"] == ""
        df_mx_pme.loc[no_rp, "_K_RP"] = (df_mx_pme.loc[no_rp, "_K_RC"]
                                           .map(rc_to_rp).fillna(""))
        df_mx_pme["_YYYYMM"] = df_mx_pme["_YYYYMM"].astype(str)

        # ── 7. Périodes ──────────────────────────────────────────────────
        self._prog(0.40, "Périodes...")
        all_yyyymm = set()
        for df_f in [df_amt, df_num, df_int, df_wl_pme, df_mx_pme]:
            if "_YYYYMM" in df_f.columns:
                all_yyyymm.update(
                    v for v in df_f["_YYYYMM"].unique()
                    if v and len(v) == 6 and v.isdigit())
        periods = sorted(all_yyyymm)
        annees  = sorted({p[:4] for p in periods})
        self._prog(0.42, f"{len(periods)} mois | années {annees}")

        # ── 8. Pré-agrégation O(1) ───────────────────────────────────────
        self._prog(0.44, "Pré-agrégation FLUX CCO...")
        for df_f in [df_amt, df_num, df_int]:
            df_f["_YYYYMM"] = df_f["_YYYYMM"].astype(str)

        agg_amt = build_agg_dict(df_amt, ["_K_RP","_PRODUCT","_YYYYMM"], "_VAL")
        agg_num = build_agg_dict(df_num, ["_K_RP","_PRODUCT","_YYYYMM"], "_VAL")
        agg_int = build_agg_dict(df_int, ["_K_RP","_PRODUCT","_YYYYMM"], "_VAL")

        self._prog(0.48, "Pré-agrégation WORLDLINE...")
        agg_wl_fl = build_agg_dict(df_wl_pme, ["_K_RP","_PROD_MAP","_YYYYMM"], "_FLUX")
        agg_wl_pn = build_agg_dict(df_wl_pme, ["_K_RP","_PROD_MAP","_YYYYMM"], "_PNB")

        self._prog(0.52, "Pré-agrégation MONEXT...")
        agg_mx_fl  = build_agg_dict(df_mx_pme, ["_K_RP","_YYYYMM"], "_FLUX")
        agg_mx_pn  = build_agg_dict(df_mx_pme, ["_K_RP","_YYYYMM"], "_PNB_TOT")
        df_ndf     = df_mx_pme[df_mx_pme["_IS_NDF"]].copy()
        agg_ndf_pn = build_agg_dict(df_ndf, ["_K_RP","_YYYYMM"], "_PNB_NDF")
        agg_ndf_fl = build_agg_dict(df_ndf, ["_K_RP","_YYYYMM"], "_FLUX")
        agg_ndf_nb = {}
        if not df_ndf.empty and "_K_RP" in df_ndf.columns:
            agg_ndf_nb = (df_ndf[df_ndf["_K_RP"] != ""]
                          .groupby(["_K_RP","_YYYYMM"])["_K_RP"]
                          .count().to_dict())

        # ── 9. Construction DATASET (CSV) ─────────────────────────────────
        self._prog(0.56, f"Dataset {n_pme:,} EJ...")

        SOURCES = (
            [("CCO", p) for p in produits_cco]
            + [("WL",  p) for p in produits_wl]
            + [("NDF", "NDF")]
        )

        rows = []
        arr_rp = df_dedup["_K_RP"].values

        for i_ej, rp in enumerate(arr_rp):
            if i_ej % 200 == 0:
                self._prog(0.56 + 0.24 * i_ej / max(n_pme, 1),
                           f"Dataset {i_ej:,}/{n_pme:,}...")
            if not rp:
                continue
            info   = parc_info.get(rp, {})
            rs_raw = str(info.get(col_rs, "")).strip()
            rc_raw = str(info.get(col_rc, "")).strip()
            rmpm   = str(info.get("_RMPM", "")).strip()
            siren  = str(info.get("_SIREN", "")).strip()
            cga    = str(info.get(col_cga, "")).strip()
            nga    = str(info.get(col_nga, "")).strip()
            k_rs   = str(info.get("_K_RS", "")).strip()
            ape     = scope_ape.get(rp, "")
            holding = "OUI" if ape in ape_holding else "NON"
            adosse  = "OUI" if fef_flag.get(rp, False) else "NON"
            segment = str(info.get(col_seg, "")).strip()

            for (src, produit) in SOURCES:
                for yyyymm in periods:
                    annee  = int(yyyymm[:4])
                    mo_num = int(yyyymm[4:6])
                    label  = yyyymm_to_label(yyyymm)
                    if src == "CCO":
                        k   = (rp, produit, yyyymm)
                        nbc = stk_dict.get((rp, produit, yyyymm), 0.0)
                        nbt = agg_num.get(k, 0.0)
                        flx = agg_amt.get(k, 0.0)
                        pnb = agg_int.get(k, 0.0)
                        if nbc == 0 and nbt == 0 and flx == 0 and pnb == 0:
                            continue
                        rows.append({"ID_RP": sacralise_id(rp),
                                     "ID_RC": sacralise_id(rc_raw),
                                     "RMPM": protect_id(rmpm),
                                     "SIREN": protect_id(siren),
                                     "RAISON_SOCIALE": rs_raw,
                                     "SEGMENT": segment,
                                     "CODE_GA": cga, "NOM_GA": nga,
                                     "CODE_APE": ape, "HOLDING": holding,
                                     "ADOSSE_FEF": adosse,
                                     "SOURCE": "CCO", "PRODUIT": produit,
                                     "YYYYMM": yyyymm, "PERIODE": label,
                                     "ANNEE": annee, "MOIS": MOIS_FR_ORDER[mo_num-1],
                                     "NB_CARTES": nbc, "NB_TRX": nbt,
                                     "FLUX": flx, "PNB": pnb})
                    elif src == "WL":
                        k   = (rp, produit, yyyymm)
                        flx = agg_wl_fl.get(k, 0.0)
                        pnb = agg_wl_pn.get(k, 0.0)
                        if flx == 0 and pnb == 0:
                            continue
                        rows.append({"ID_RP": sacralise_id(rp),
                                     "ID_RC": sacralise_id(rc_raw),
                                     "RMPM": protect_id(rmpm),
                                     "SIREN": protect_id(siren),
                                     "RAISON_SOCIALE": rs_raw,
                                     "SEGMENT": segment,
                                     "CODE_GA": cga, "NOM_GA": nga,
                                     "CODE_APE": ape, "HOLDING": holding,
                                     "ADOSSE_FEF": adosse,
                                     "SOURCE": "WORLDLINE", "PRODUIT": produit,
                                     "YYYYMM": yyyymm, "PERIODE": label,
                                     "ANNEE": annee, "MOIS": MOIS_FR_ORDER[mo_num-1],
                                     "NB_CARTES": 0.0, "NB_TRX": 0.0,
                                     "FLUX": flx, "PNB": pnb})
                    else:  # NDF
                        k   = (rp, yyyymm)
                        nb  = agg_ndf_nb.get(k, 0.0)
                        flx = agg_ndf_fl.get(k, 0.0)
                        pnb = agg_ndf_pn.get(k, 0.0)
                        if nb == 0 and flx == 0 and pnb == 0:
                            continue
                        rows.append({"ID_RP": sacralise_id(rp),
                                     "ID_RC": sacralise_id(rc_raw),
                                     "RMPM": protect_id(rmpm),
                                     "SIREN": protect_id(siren),
                                     "RAISON_SOCIALE": rs_raw,
                                     "SEGMENT": segment,
                                     "CODE_GA": cga, "NOM_GA": nga,
                                     "CODE_APE": ape, "HOLDING": holding,
                                     "ADOSSE_FEF": adosse,
                                     "SOURCE": "MONEXT_NDF", "PRODUIT": "NDF",
                                     "YYYYMM": yyyymm, "PERIODE": label,
                                     "ANNEE": annee, "MOIS": MOIS_FR_ORDER[mo_num-1],
                                     "NB_CARTES": nb, "NB_TRX": 0.0,
                                     "FLUX": flx, "PNB": pnb})

        DS_COLS = ["ID_RP","ID_RC","RMPM","SIREN","RAISON_SOCIALE","SEGMENT",
                   "CODE_GA","NOM_GA","CODE_APE","HOLDING","ADOSSE_FEF",
                   "SOURCE","PRODUIT","YYYYMM","PERIODE","ANNEE","MOIS",
                   "NB_CARTES","NB_TRX","FLUX","PNB"]
        df_ds = pd.DataFrame(rows) if rows else pd.DataFrame(columns=DS_COLS)
        for c in DS_COLS:
            if c not in df_ds.columns:
                df_ds[c] = ""
        df_ds = df_ds[DS_COLS]
        self._prog(0.82, f"Dataset : {len(df_ds):,} lignes")

        # ── 10. Export CSV — QUOTE_NONE pour sacralisation ────────────────
        self._prog(0.84, "Export CSV...")
        # asksaveasfilename remplacé par self.output_dir / self.output_filename
        self.output_dir.mkdir(parents=True, exist_ok=True)
        save_csv = str(self.output_dir / f"{self.output_filename}.csv")

        df_ds.to_csv(save_csv, sep=";", index=False, encoding="utf-8-sig",
                     quoting=csv.QUOTE_NONE, escapechar="\\")
        self._prog(0.87, "CSV enregistré")

        print(f"[INFO] DATASET CSV créé : {save_csv}")
        print(f"[INFO] {len(df_ds):,} lignes | {n_pme:,} EJ | {len(periods)} mois")
        print('[INFO] Identifiants RP/RC/RMPM/SIREN sacralisés (="VALEUR").')

        # ── 11. Export XLSX ───────────────────────────────────────────────
        self._prog(0.89, "Sélection XLSX...")
        # asksaveasfilename remplacé par self.output_dir / self.output_filename
        save_xlsx = str(self.output_dir / f"{self.output_filename}.xlsx")

        self._prog(0.91, "Génération XLSX...")
        # Liste RS pour liste déroulante
        rs_list = sorted(
            df_ds["RAISON_SOCIALE"].astype(str).str.strip()
            .replace("", pd.NA).dropna().unique().tolist())

        self._export_xlsx(df_ds, DS_COLS, rs_list,
                          produits_cco, produits_wl,
                          periods, annees, save_xlsx, n_pme, pme_codes, ts)
        self._prog(1.0, "Terminé !")
        print(f"[INFO] XLSX généré : {save_xlsx}")
        print("[INFO] Sheet SOURCE  : en-têtes + espace pour collage du CSV")
        print("[INFO] Sheet DATA    : tableau pivoté par Python (RS×PRODUIT × mois)")
        print("[INFO] Sheet ANALYSE : liste déroulante client + formules RECHERCHEX")
        print("[INFO] Sheet DASHBOARD: valeurs en dur + graphiques")

        self._csv_path = save_csv
        self._xlsx_path = save_xlsx

    # ──────────────────────────────────────────────────────────────────────────
    # EXPORT XLSX — 4 sheets (VERBATIM)
    # ──────────────────────────────────────────────────────────────────────────

    def _export_xlsx(self, df_ds, ds_cols, rs_list,
                     produits_cco, produits_wl,
                     periods, annees, save_xlsx, n_pme, pme_codes, ts):
        """
        4 sheets : SOURCE / DATA (pivot) / ANALYSE (RECHERCHEX) / DASHBOARD (valeurs en dur)
        Pattern couleur C2SME : hex 6 chars, SANS FF, SANS #, JAMAIS tabColor.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.chart import LineChart, Reference

        wb = Workbook()
        C = get_column_letter

        # ── Palette C2SME — SEUL PATTERN VALIDÉ SANS ERREUR ─────────────────
        GRN  = "00915A"; GRN2 = "E8F5E9"; DARK = "1C3A2D"
        GREY = "F5F5F5"; WHT  = "FFFFFF"; BLU  = "1565C0"
        BLU2 = "E3F2FD"; ORG  = "E65100"; ORG2 = "FFF3E0"
        WINE = "880E4F"; GRY2 = "607D8B"

        def fill(h):
            return PatternFill(start_color=h, end_color=h, fill_type="solid")
        def fnt(c="FFFFFF", bold=True, sz=10):
            return Font(name="Segoe UI", size=sz, bold=bold, color=c)
        thin = Side(style="thin", color="CCCCCC")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

        SEP = "§"

        COL_COLORS = {
            "ID_RP": DARK, "ID_RC": DARK, "RMPM": DARK, "SIREN": DARK,
            "RAISON_SOCIALE": DARK, "SEGMENT": GRY2, "CODE_GA": DARK, "NOM_GA": DARK,
            "CODE_APE": DARK, "HOLDING": ORG, "ADOSSE_FEF": ORG,
            "SOURCE": GRN, "PRODUIT": GRN,
            "YYYYMM": GRY2, "PERIODE": GRY2, "ANNEE": GRY2, "MOIS": GRY2,
            "NB_CARTES": "2E7D32", "NB_TRX": BLU, "FLUX": BLU, "PNB": WINE,
        }
        COL_W = {
            "ID_RP": 18, "ID_RC": 18, "RMPM": 14, "SIREN": 13,
            "RAISON_SOCIALE": 32, "SEGMENT": 8, "CODE_GA": 10, "NOM_GA": 26,
            "CODE_APE": 8, "HOLDING": 10, "ADOSSE_FEF": 12,
            "SOURCE": 14, "PRODUIT": 24, "YYYYMM": 8,
            "PERIODE": 16, "ANNEE": 7, "MOIS": 6,
            "NB_CARTES": 12, "NB_TRX": 10, "FLUX": 14, "PNB": 14,
        }

        # ═══════════════════════════════════════════════════════════════
        # SHEET 1 — SOURCE (en-têtes + espace pour collage CSV)
        # ═══════════════════════════════════════════════════════════════
        ws_src = wb.active
        ws_src.title = "SOURCE"
        ws_src["A1"].value = (f"SOURCE [{VERSION_ID}] — "
                               "Collez le contenu du CSV à partir de A3 (sans les en-têtes)")
        ws_src["A1"].font = Font(name="Segoe UI", size=9, italic=True, color="888888")
        for ci, col in enumerate(ds_cols, start=1):
            cell = ws_src.cell(row=2, column=ci, value=col)
            cell.font      = fnt(sz=9)
            cell.fill      = fill(COL_COLORS.get(col, DARK))
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = brd
            ws_src.column_dimensions[C(ci)].width = COL_W.get(col, 14)
        ws_src.row_dimensions[2].height = 38
        ws_src.freeze_panes = "A3"

        # ═══════════════════════════════════════════════════════════════
        # SHEET 2 — DATA (pivot calculé par Python)
        # ═══════════════════════════════════════════════════════════════
        ws_data = wb.create_sheet("DATA")

        FIXED = ["ID_RP","ID_RC","RMPM","SIREN","RAISON_SOCIALE","SEGMENT",
                 "CODE_GA","NOM_GA","CODE_APE","HOLDING","ADOSSE_FEF",
                 "SOURCE","PRODUIT","CLE"]
        N_FIX = len(FIXED)
        INDS  = ["NB_CARTES","NB_TRX","FLUX","PNB"]

        th_cols = []
        for p in periods:
            lbl = yyyymm_to_label(p)
            for ind in INDS:
                th_cols.append((f"{lbl}_{ind}", p, ind, "mois"))
        for a in annees:
            for ind in INDS:
                th_cols.append((f"ANNEE_{a}_{ind}", a, ind, "annee"))

        all_dc = FIXED + [t[0] for t in th_cols]

        for ci, col in enumerate(all_dc, start=1):
            if col in FIXED:
                hc = COL_COLORS.get(col, DARK)
            elif "NB_CARTES" in col: hc = "2E7D32"
            elif "NB_TRX"   in col: hc = BLU
            elif "FLUX"     in col: hc = BLU
            elif "PNB"      in col: hc = WINE
            else:                   hc = GRY2
            cell = ws_data.cell(row=1, column=ci, value=col)
            cell.font = fnt(sz=8); cell.fill = fill(hc)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = brd
            ws_data.column_dimensions[C(ci)].width = (
                18 if col in ["ID_RP","ID_RC"] else
                14 if col in ["RMPM","SIREN","CLE"] else
                28 if col == "RAISON_SOCIALE" else
                10 if col in ["CODE_GA","CODE_APE","HOLDING","ADOSSE_FEF"] else
                22 if col == "NOM_GA" else 14)
        ws_data.row_dimensions[1].height = 42
        ws_data.freeze_panes = f"{C(N_FIX+1)}2"

        # Pivot vectorisé
        col_data_map = {}
        for ci_off, (col_lbl, pv, ind, tt) in enumerate(th_cols):
            col_data_map[col_lbl] = C(N_FIX + 1 + ci_off)

        if not df_ds.empty:
            grp_c = ["ID_RP","ID_RC","RMPM","SIREN","RAISON_SOCIALE","SEGMENT",
                     "CODE_GA","NOM_GA","CODE_APE","HOLDING","ADOSSE_FEF",
                     "SOURCE","PRODUIT"]
            df_uid = (df_ds[grp_c].drop_duplicates(["RAISON_SOCIALE","SEGMENT","SOURCE","PRODUIT"])
                      .copy().reset_index(drop=True))
            df_uid["CLE"] = df_uid["RAISON_SOCIALE"] + SEP + df_uid["SOURCE"] + SEP + df_uid["PRODUIT"]

            df_n = df_ds.copy()
            for ind in INDS:
                if ind in df_n.columns:
                    df_n[ind] = pd.to_numeric(
                        df_n[ind].astype(str).str.replace(",", ".", regex=False),
                        errors="coerce").fillna(0.0)
                else:
                    df_n[ind] = 0.0
            df_n["_ANN"] = df_n["ANNEE"].astype(str)

            piv_mo, piv_an = {}, {}
            for ind in INDS:
                gm = (df_n[df_n[ind] != 0.0]
                      .groupby(["RAISON_SOCIALE","SOURCE","PRODUIT","YYYYMM"])[ind].sum())
                for (rs,src,pr,ym), v in gm.items():
                    piv_mo[(rs,src,pr,ym,ind)] = v
                ga = (df_n[df_n[ind] != 0.0]
                      .groupby(["RAISON_SOCIALE","SOURCE","PRODUIT","_ANN"])[ind].sum())
                for (rs,src,pr,a), v in ga.items():
                    piv_an[(rs,src,pr,a,ind)] = v

            for ri, row_id in enumerate(df_uid.itertuples(index=False), start=2):
                rs = row_id.RAISON_SOCIALE; src = row_id.SOURCE; pr = row_id.PRODUIT
                bg = GREY if ri % 2 == 0 else WHT
                fv = [row_id.ID_RP, row_id.ID_RC, row_id.RMPM, row_id.SIREN,
                      rs, row_id.SEGMENT, row_id.CODE_GA, row_id.NOM_GA,
                      row_id.CODE_APE, row_id.HOLDING, row_id.ADOSSE_FEF,
                      src, pr, row_id.CLE]
                for ci, v in enumerate(fv, start=1):
                    c = ws_data.cell(row=ri, column=ci, value=v)
                    c.font = Font(name="Segoe UI", size=9)
                    c.fill = fill(bg); c.border = brd
                    c.alignment = Alignment(horizontal="center")
                for ci_off, (col_lbl, pv, ind, tt) in enumerate(th_cols):
                    ci = N_FIX + 1 + ci_off
                    v = (piv_mo.get((rs,src,pr,pv,ind))
                         if tt == "mois"
                         else piv_an.get((rs,src,pr,str(pv),ind)))
                    c = ws_data.cell(row=ri, column=ci, value=v)
                    c.font = Font(name="Segoe UI", size=9)
                    c.fill = fill(bg); c.border = brd
                    c.alignment = Alignment(horizontal="center")
                    if v:
                        c.number_format = "#,##0.00"

        # ═══════════════════════════════════════════════════════════════
        # SHEET 3 — _CLIENTS masquée
        # ═══════════════════════════════════════════════════════════════
        ws_cl = wb.create_sheet("_CLIENTS")
        ws_cl.sheet_state = "hidden"
        for i, rs in enumerate(rs_list, start=1):
            ws_cl.cell(row=i, column=1, value=rs)

        # ═══════════════════════════════════════════════════════════════
        # SHEET 4 — ANALYSE (liste déroulante + RECHERCHEX)
        # ═══════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("ANALYSE")
        CLE_C = C(N_FIX); RS_REF = "$B$1"
        c_rs = "E"
        c_nga = C(all_dc.index("NOM_GA")     + 1)
        c_hld = C(all_dc.index("HOLDING")    + 1)
        c_fef = C(all_dc.index("ADOSSE_FEF") + 1)

        def rx(sv, pv, dc):
            cle = f'({RS_REF}&"{SEP}{sv}{SEP}{pv}")'
            return f"=SIERREUR(RECHERCHEX({cle},DATA!${CLE_C}:${CLE_C},DATA!${dc}:${dc},0),0)"

        def rx_an(sv, pv, an, ind):
            lbl = f"ANNEE_{an}_{ind}"; dc = col_data_map.get(lbl)
            if not dc: return "0"
            cle = f'({RS_REF}&"{SEP}{sv}{SEP}{pv}")'
            return f"=SIERREUR(RECHERCHEX({cle},DATA!${CLE_C}:${CLE_C},DATA!${dc}:${dc},0),0)"

        tca = []
        for p in periods: tca.append((yyyymm_to_label(p), p, "mois"))
        for a in annees:  tca.append((f"ANNEE_{a}", a, "annee"))

        SC=1; PC=2; TC=3; HR=6; DR=7

        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 36

        for i, lbl in enumerate(["Nom du client","Nom du GA","Holding ?"], start=1):
            ca = ws2.cell(row=i, column=1, value=lbl)
            ca.font = fnt(sz=10); ca.fill = fill(DARK)
            ca.border = brd; ca.alignment = Alignment(horizontal="left", vertical="center")
            ws2.row_dimensions[i].height = 22

        b1 = ws2.cell(row=1, column=2, value=rs_list[0] if rs_list else "")
        b1.font = Font(name="Segoe UI", size=11, bold=True, color="000000")
        b1.fill = fill("FFFDE7"); b1.border = brd
        b1.alignment = Alignment(horizontal="left", vertical="center")

        dv = DataValidation(type="list",
                            formula1=f"_CLIENTS!$A$1:$A${len(rs_list)}",
                            allow_blank=True, showDropDown=False, showErrorMessage=False)
        ws2.add_data_validation(dv); dv.sqref = "B1"

        ws2.cell(row=2, column=2).value = (
            f'=SIERREUR(RECHERCHEX({RS_REF},DATA!${c_rs}:${c_rs},DATA!${c_nga}:${c_nga},"N/A"),"N/A")')
        ws2.cell(row=2, column=2).font = Font(name="Segoe UI", size=10)
        ws2.cell(row=2, column=2).fill = fill(GRN2); ws2.cell(row=2, column=2).border = brd

        ws2.cell(row=3, column=2).value = (
            f'=SIERREUR(RECHERCHEX({RS_REF},DATA!${c_rs}:${c_rs},DATA!${c_hld}:${c_hld},"N/A"),"N/A")')
        ws2.cell(row=3, column=2).font = Font(name="Segoe UI", size=10)
        ws2.cell(row=3, column=2).fill = fill(GRN2); ws2.cell(row=3, column=2).border = brd

        ws2.row_dimensions[HR].height = 42
        for ci_s, ls in enumerate(["Section","Produit / Indicateur"], start=1):
            c = ws2.cell(row=HR, column=ci_s, value=ls)
            c.font = fnt(sz=9); c.fill = fill(DARK); c.border = brd
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for ci, (lbl, val, tt) in enumerate(tca, start=TC):
            c = ws2.cell(row=HR, column=ci, value=lbl)
            c.font = fnt(sz=8)
            c.fill = fill(BLU if tt == "mois" else GRY2)
            c.border = brd
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            ws2.column_dimensions[C(ci)].width = 14
        ws2.column_dimensions[C(SC)].width = 14
        ws2.column_dimensions[C(PC)].width = 34

        SECTS = [("PAIEMENT","NB_CARTES",DARK,GRN2),
                 ("FLUX","FLUX",BLU,BLU2),
                 ("PNB","PNB",WINE,"FCE4EC")]
        lignes = []
        for sect, ind_def, sc_c, bg_c in SECTS:
            lignes.append((sect, f"Total {sect} CCO", "CCO",        None,  ind_def, sc_c, bg_c))
            for p in produits_cco:
                lignes.append((sect, f"CORPORATE  {p}", "CCO",      p,     ind_def, sc_c, bg_c))
            for p in produits_wl:
                lignes.append((sect, f"PROC/VIRT  {p}", "WORLDLINE",p,     ind_def, sc_c, bg_c))
            if sect == "PAIEMENT":
                lignes.append((sect, "NDF — Nb lignes","MONEXT_NDF","NDF","NB_CARTES",sc_c,bg_c))
            lignes.append((sect, f"NDF — {sect}","MONEXT_NDF","NDF",ind_def,sc_c,bg_c))
            if sect == "PAIEMENT":
                lignes.append((sect, "FEF — Adossé FEF","_FEF_",None,None,sc_c,bg_c))

        prev_sect = None
        for l_idx, (sect, plbl, sv, pv, ind, sc_c, bg_c) in enumerate(lignes):
            rn = DR + l_idx
            ws2.row_dimensions[rn].height = 18
            cs = ws2.cell(row=rn, column=SC)
            cs.value = sect if sect != prev_sect else ""
            cs.font = fnt(sz=9); cs.fill = fill(sc_c); cs.border = brd
            cs.alignment = Alignment(horizontal="center", vertical="center")
            if sect != prev_sect: prev_sect = sect

            is_tot = pv is None and sv not in ("_FEF_",)
            cp = ws2.cell(row=rn, column=PC, value=plbl)
            cp.font = Font(name="Segoe UI", size=9, bold=is_tot)
            cp.fill = fill(GRN2 if "CORPORATE" in plbl.upper()
                          else (BLU2 if "PROC/VIRT" in plbl.upper() else GREY))
            cp.border = brd; cp.alignment = Alignment(horizontal="left", vertical="center")

            for ci_off, (lbl_t, val_t, tt) in enumerate(tca):
                ci = TC + ci_off
                cd = ws2.cell(row=rn, column=ci)
                cd.font = Font(name="Segoe UI", size=9)
                cd.fill = fill(bg_c); cd.border = brd
                cd.alignment = Alignment(horizontal="center")
                cd.number_format = "#,##0.00"
                if sv == "_FEF_":
                    cd.value = (f'=SIERREUR(RECHERCHEX({RS_REF},DATA!${c_rs}:${c_rs},'
                                f'DATA!${c_fef}:${c_fef},"N/A"),"N/A")')
                    cd.number_format = "General"
                elif is_tot:
                    cd.value = None
                elif tt == "mois":
                    dc = col_data_map.get(f"{lbl_t}_{ind}")
                    cd.value = rx(sv, pv, dc) if dc else 0
                else:
                    cd.value = rx_an(sv, pv, val_t, ind)

        ws2.freeze_panes = f"C{DR}"

        # ═══════════════════════════════════════════════════════════════
        # SHEET 5 — DASHBOARD (toutes valeurs en dur — zéro formule)
        # Python agrège depuis df_ds et écrit les valeurs numériques
        # Structure : tableau mensuel gauche | graphiques droite
        # ═══════════════════════════════════════════════════════════════
        ws_db = wb.create_sheet("DASHBOARD")

        # ── Agrégation Python depuis df_ds ──────────────────────────────
        # Toutes sources et produits confondus par mois
        mois_only = sorted([p for p in periods if len(p) == 6 and p.isdigit()])

        agg_mois = {}   # yyyymm → {NB_CARTES, FLUX, PNB}
        agg_moy  = {}   # yyyymm → {NB_CARTES_MOY, FLUX_MOY, PNB_MOY} (par EJ)

        if not df_ds.empty:
            df_n2 = df_ds.copy()
            for ind in ["NB_CARTES","NB_TRX","FLUX","PNB"]:
                if ind in df_n2.columns:
                    df_n2[ind] = pd.to_numeric(
                        df_n2[ind].astype(str).str.replace(",",".",regex=False),
                        errors="coerce").fillna(0.0)
                else:
                    df_n2[ind] = 0.0

            for mo in mois_only:
                df_mo = df_n2[df_n2["YYYYMM"] == mo]
                nb_c = float(df_mo["NB_CARTES"].sum())
                flx  = float(df_mo["FLUX"].sum())
                pnb  = float(df_mo["PNB"].sum())
                n_ej = max(df_mo["RAISON_SOCIALE"].nunique(), 1)
                agg_mois[mo] = {"NB_CARTES": nb_c, "FLUX": flx, "PNB": pnb}
                agg_moy[mo]  = {"NB_CARTES_MOY": nb_c/n_ej,
                                 "FLUX_MOY": flx/n_ej,
                                 "PNB_MOY": pnb/n_ej}

        # ── Titre dashboard ─────────────────────────────────────────────
        ws_db["A1"].value = (f"CCO PME DASHBOARD [{VERSION_ID}] — "
                             f"{n_pme:,} EJ | {', '.join(pme_codes)} | "
                             f"{len(mois_only)} mois | "
                             f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        ws_db["A1"].font = fnt(GRN, sz=12)
        ws_db.row_dimensions[1].height = 26
        ws_db.merge_cells("A1:N1")

        # ── Tableau 1 : évolution mensuelle (totaux) ────────────────────
        ws_db["A3"].value = "ÉVOLUTION MENSUELLE — TOTAUX"
        ws_db["A3"].font  = fnt(GRN, sz=11)
        ws_db.row_dimensions[3].height = 22

        HDRS1 = ["MOIS","NB_CARTES","FLUX (EUR)","PNB (EUR)"]
        for ci, h in enumerate(HDRS1, start=1):
            c = ws_db.cell(row=4, column=ci, value=h)
            c.font = fnt(sz=9)
            c.fill = fill(DARK if ci==1 else
                         "2E7D32" if ci==2 else BLU if ci==3 else WINE)
            c.border = brd; c.alignment = Alignment(horizontal="center")
        ws_db.row_dimensions[4].height = 32

        DB_DATA_START = 5
        for mi, mo in enumerate(mois_only):
            r = DB_DATA_START + mi
            lbl = yyyymm_to_label(mo)
            d   = agg_mois.get(mo, {"NB_CARTES": 0, "FLUX": 0, "PNB": 0})
            bg  = GREY if mi % 2 == 0 else WHT
            vals = [lbl, d["NB_CARTES"], d["FLUX"], d["PNB"]]
            for ci, v in enumerate(vals, start=1):
                c = ws_db.cell(row=r, column=ci, value=v)
                c.font = Font(name="Segoe UI", size=9)
                c.fill = fill(bg); c.border = brd
                c.alignment = Alignment(horizontal="center")
                if ci > 1: c.number_format = "#,##0.00"
            ws_db.row_dimensions[r].height = 16

        DB_DATA_END = DB_DATA_START + len(mois_only) - 1

        # ── Tableau 2 : moyennes par EJ ─────────────────────────────────
        T2_START = DB_DATA_END + 3
        ws_db.cell(row=T2_START, column=1,
                   value="ÉVOLUTION MENSUELLE — MOYENNES PAR EJ").font = fnt(GRN, sz=11)
        ws_db.row_dimensions[T2_START].height = 22

        HDRS2 = ["MOIS","NB_CARTES MOY","FLUX MOY (EUR)","PNB MOY (EUR)"]
        for ci, h in enumerate(HDRS2, start=1):
            c = ws_db.cell(row=T2_START+1, column=ci, value=h)
            c.font = fnt(sz=9)
            c.fill = fill(DARK if ci==1 else
                         "2E7D32" if ci==2 else BLU if ci==3 else WINE)
            c.border = brd; c.alignment = Alignment(horizontal="center")
        ws_db.row_dimensions[T2_START+1].height = 32

        MOY_DATA_START = T2_START + 2
        for mi, mo in enumerate(mois_only):
            r = MOY_DATA_START + mi
            lbl = yyyymm_to_label(mo)
            d   = agg_moy.get(mo, {"NB_CARTES_MOY":0, "FLUX_MOY":0, "PNB_MOY":0})
            bg  = GREY if mi % 2 == 0 else WHT
            vals = [lbl, d["NB_CARTES_MOY"], d["FLUX_MOY"], d["PNB_MOY"]]
            for ci, v in enumerate(vals, start=1):
                c = ws_db.cell(row=r, column=ci, value=v)
                c.font = Font(name="Segoe UI", size=9)
                c.fill = fill(bg); c.border = brd
                c.alignment = Alignment(horizontal="center")
                if ci > 1: c.number_format = "#,##0.00"
            ws_db.row_dimensions[r].height = 16

        MOY_DATA_END = MOY_DATA_START + len(mois_only) - 1

        # ── Largeurs colonnes dashboard ──────────────────────────────────
        for ci, w in enumerate([18, 16, 18, 16], start=1):
            ws_db.column_dimensions[C(ci)].width = w
        ws_db.freeze_panes = "A5"

        # ── 3 graphiques linéaires (si données disponibles) ──────────────
        if len(mois_only) > 0:

            def make_line(title, col_idx, color_hex, anchor, data_start, data_end):
                ch = LineChart()
                ch.title  = title
                ch.width  = 16; ch.height = 9
                ch.legend.position = "b"; ch.legend.overlay = False
                try:
                    ch.y_axis.majorGridlines = None
                    ch.y_axis.minorGridlines = None
                except: pass
                # Données en dur dans ws_db
                ref = Reference(ws_db, min_col=col_idx, max_col=col_idx,
                                 min_row=data_start - 1, max_row=data_end)
                ch.add_data(ref, titles_from_data=True)
                ch.series[0].graphicalProperties.line.solidFill = color_hex
                ch.series[0].graphicalProperties.line.width     = 25000
                cats = Reference(ws_db, min_col=1,
                                  min_row=data_start, max_row=data_end)
                ch.set_categories(cats)
                ws_db.add_chart(ch, anchor)
                return ch

            # Graphique 1 — NB CARTES (col 2, ligne DB_DATA_START→END)
            make_line("Évolution Paiements — Nb cartes", 2, GRN,
                      "F3", DB_DATA_START, DB_DATA_END)
            # Graphique 2 — FLUX (col 3)
            make_line("Évolution Flux (EUR)", 3, BLU,
                      "F20", DB_DATA_START, DB_DATA_END)
            # Graphique 3 — PNB (col 4)
            make_line("Évolution PNB (EUR)", 4, WINE,
                      "F37", DB_DATA_START, DB_DATA_END)

        wb.save(save_xlsx)


# =============================================================================
# MAIN — argparse + gestion des codes de sortie (0 / 1 / 2)
# =============================================================================

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="01.B2PME.py",
        description="CCO PME DASHBOARD v14 [B2PME] — CLI autonome "
                    "(DATASET CSV + XLSX 4 feuilles).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    # ── Sources obligatoires ──────────────────────────────────────────────
    p.add_argument("--parc", required=False, default=None, help="PARC_CLIENT (obligatoire)")
    p.add_argument("--flux-amount", required=False, default=None,
                   help="FLUX AMOUNT agrégé [N3096] (obligatoire)")
    p.add_argument("--flux-number", required=False, default=None,
                   help="FLUX NUMBER agrégé [N3096] (obligatoire)")
    p.add_argument("--flux-interchange", required=False, default=None,
                   help="FLUX INTERCHANGE agrégé [N3096] (obligatoire)")
    p.add_argument("--stock", required=False, default=None, help="STOCK agrégé [66Z3Q] (obligatoire)")
    p.add_argument("--worldline", required=False, default=None,
                   help="WORLDLINE PRGM consolidé [6AICC] (obligatoire)")
    p.add_argument("--monext", required=False, default=None,
                   help="MONEXT consolidé [K2P8N] (obligatoire)")
    # ── Sources optionnelles ──────────────────────────────────────────────
    p.add_argument("--matching-usage", default=None,
                   help="MATCHING_USAGE Worldline (optionnel)")
    p.add_argument("--fef", default=None, help="FEF Éric Tardy (optionnel)")
    p.add_argument("--scope", default=None,
                   help="SCOPE CLIENTS Olivier Attali (optionnel)")
    # ── Paramètres métier (optionnels, défauts UI) ────────────────────────
    p.add_argument("--pme-codes", default=DEFAULT_PME_CODES,
                   help="Codes segment PME (séparés par virgule)")
    p.add_argument("--stock-active", default=DEFAULT_STOCK_ACTIVE,
                   help="Valeur Data=actif pour le STOCK")
    p.add_argument("--ape-holding", default="",
                   help="Codes APE holding supplémentaires (séparés par virgule)")
    p.add_argument("--no-default-ape", action="store_true",
                   help="Désactive les 4 codes APE holding par défaut")
    # ── Sorties obligatoires ──────────────────────────────────────────────
    p.add_argument("--output-dir", required=True,
                   help="Répertoire de sortie (obligatoire)")
    p.add_argument("--output-filename", required=True,
                   help="Nom de fichier de sortie sans extension (obligatoire)")
    return p


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    # Auto-résolution des sources obligatoires non fournies depuis 03.sources/
    try:
        if not args.parc:
            args.parc = resolve_source("PARC_CLIENT", required=True)
        if not args.stock:
            args.stock = resolve_source("CCO_STOCK", required=True)
        if not args.worldline:
            args.worldline = resolve_source("PRGM_AGREGE", required=True)
        if not args.monext:
            args.monext = resolve_source("MONEXT_AGREGE", required=True)
        if not args.flux_amount:
            args.flux_amount = resolve_source_kw("CCO_FLUX", "AMOUNT", required=True)
        if not args.flux_number:
            args.flux_number = resolve_source_kw("CCO_FLUX", "NUMBER", required=True)
        if not args.flux_interchange:
            args.flux_interchange = resolve_source_kw("CCO_FLUX", "INTERCHANGE", required=True)
    except FileNotFoundError as exc:
        print(f"[ERREUR] {exc}", file=sys.stderr)
        return 1

    # Vérification existence des sources fournies (oblig + opt si renseignées)
    file_args = {
        "--parc": args.parc,
        "--flux-amount": args.flux_amount,
        "--flux-number": args.flux_number,
        "--flux-interchange": args.flux_interchange,
        "--stock": args.stock,
        "--worldline": args.worldline,
        "--monext": args.monext,
        "--matching-usage": args.matching_usage,
        "--fef": args.fef,
        "--scope": args.scope,
    }
    for flag, path in file_args.items():
        if path and not Path(path).is_file():
            print(f"[ERREUR] Fichier introuvable pour {flag} : {path}",
                  file=sys.stderr)
            return 2

    try:
        app = CCOPMEDashboard_B2PME(args)
        app.worker()
    except Exception as exc:  # noqa: BLE001
        import traceback
        traceback.print_exc()
        print(f"[ERREUR] {exc}", file=sys.stderr)
        return 1

    print(f"[OK] DATASET CSV + XLSX générés dans : {args.output_dir}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
