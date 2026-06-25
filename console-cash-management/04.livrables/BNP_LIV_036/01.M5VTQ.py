
"""
MONEXT REVENUS PERIODE v1 [M5VTQ]
=================================
BNP Paribas Cash Management - Direction Monetique
Livrable BNP_LIV_036 - Analyse des revenus MONEXT par client (ID RP) sur une
periode selectionnable [premier mois -> dernier mois]. Version CLI autonome
(argparse, sans GUI), logique metier preservee a l'identique du script original.

DESCRIPTION
-----------
Agregation des flux MONEXT par client (ID RP) sur une periode bornee. Une seule
entree : le fichier MONEXT consolide multi-mois. Produit un classeur Excel a deux
feuilles : DATA (une ligne par ID RP) et SYNTHESE (vue d'ensemble + detail mensuel).
Le client est identifie directement par son ID RP (pas d'IBAN_ACCOUNT / SINGLETON /
PARC pour ce livrable : agregation MONEXT pure).

SOURCES REQUISES
----------------
- MONEXT CONSOLIDE (--source-monext, obligatoire) : CSV multi-mois portant le mois,
  l'ID RP, le nom client (RS), le nombre de cartes (mensuel), les depenses, les
  retraits et les colonnes PNB. Colonne differe optionnelle.

OUTPUTS PRODUITS
----------------
- XLSX 2 feuilles : MONEXT_REVENUS_PERIODE_<debut>_<fin>_<timestamp>_M5VTQ.xlsx
    * Feuille DATA     : une ligne par ID RP
        1. ID_RP
        2. NOM_CLIENT (RS MONEXT, valeur la plus frequente)
        3. FLUX TOTAL [premier -> dernier]   (somme des flux sur la periode)
        4. NB_CARTES_MOYEN                   (somme des cartes / nb de mois selectionnes)
        5. DIFFERES                          (liste triee croissante des differes distincts)
        6. PNB TOTAL [premier -> dernier]
        puis decomposition mensuelle : bloc FLUX, bloc CARTES, bloc PNB par mois.
    * Feuille SYNTHESE : KPIs periode + detail mensuel (FLUX / NB_CARTES / PNB / CLIENTS_ACTIFS).

ARGUMENTS CLI
-------------
- --source-monext   (obligatoire) : chemin du CSV MONEXT consolide.
- --mois-debut      (obligatoire) : premier mois inclus, format YYYYMM (ex 202501).
- --mois-fin        (obligatoire) : dernier mois inclus, format YYYYMM (ex 202512).
- --differe-col     (optionnel)   : position (1-based) de la colonne differe dans MONEXT.
- --output-dir      (obligatoire) : dossier de sortie.
- --output-filename (obligatoire) : nom du fichier XLSX produit.

DECOMPOSITION
-------------
- main()
  |- argparse : lecture des flags CLI
  |- MonextRevenusPeriode(args)
  |   |- __init__        : affecte fichier, mois debut/fin, mapping colonnes (defauts), differe
  |   |- run()           : pipeline metier
  |       |- load_csv_smart      : detection separateur / encodage
  |       |- clean_id / to_float / parse_mois / differe_to_num : helpers (verbatim)
  |       |- calcul FLUX (depenses + retraits)
  |       |- calcul CARTES (mensuel)
  |       |- calcul PNB (somme plage hors col exclue, interchange x -1)
  |       |- filtre periode + agregats par RP
  |       |- pivots mensuels FLUX / CARTES / PNB
  |       |- assemblage DATA + SYNTHESE + KPIs
  |       |- write_xlsx          : export 2 feuilles
  |- codes de sortie : 0 OK / 1 erreur metier / 2 erreur arguments

CALCULS (repris du LIV_034 W3RKN)
---------------------------------
  - FLUX   = Depenses (col 15) + Retraits (col 17)
  - CARTES = col 12 (mensuel)
  - PNB    = somme colonnes [first..last] (19..55) HORS colonne exclue (33),
             colonne interchange (21) x -1
  - NB_CARTES_MOYEN = somme cartes periode / nb de mois selectionnes
  - DIFFERES        = valeurs distinctes de la colonne differe, triees croissant, entre crochets

STANDARDS : METHOD_CODING_BNP (vectorisation pandas, openpyxl hex 6 sans FF,
conversion numerique ciblee, preservation des zeros initiaux des identifiants).
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ─── Résolution auto des sources + lecture DuckDB (inliné depuis 06.fonctions) ──
# Le script remonte jusqu'au dossier contenant "03.sources" puis prend, dans
# 03.sources/*/<NOM_SOURCE>/, le fichier au préfixe numérique le plus élevé.
def _find_sources_root():
    here = Path(__file__).resolve()
    for d in [here] + list(here.parents):
        c = d / "03.sources"
        if c.is_dir():
            return c
    return None


def resolve_source(name, required=False):
    root = _find_sources_root()
    if not root:
        if required:
            raise FileNotFoundError(f"Dossier 03.sources introuvable depuis {__file__}")
        return None
    folder = next((m for m in root.glob("*/" + name) if m.is_dir()), None)
    if not folder:
        if required:
            raise FileNotFoundError(f"Source introuvable : 03.sources/*/{name}")
        return None
    exts = (".csv", ".xlsx", ".xls", ".txt")
    files = [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() in exts]
    if not files:
        if required:
            raise FileNotFoundError(f"Aucun fichier source dans {folder}")
        return None

    def _np(f):
        m = re.match(r"\s*(\d+)", f.name)
        return (int(m.group(1)) if m else -1, f.name)
    return max(files, key=_np)


def _read_duck(path, nrows=None):
    """Lecture CSV via DuckDB (tout en texte). Renvoie None si indisponible ou
    non-CSV (ou nrows demandé) pour laisser le loader d'origine prendre le relais."""
    p = Path(path)
    if nrows is not None or p.suffix.lower() not in (".csv", ".txt", ""):
        return None
    try:
        import duckdb
        return duckdb.read_csv(str(p), all_varchar=True).df()
    except Exception:
        return None


# ═════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═════════════════════════════════════════════════════════════════════════════
VERSION_ID = "M5VTQ"

# Couleurs openpyxl — hex 6 chars, SANS # ni préfixe FF (cf. METHOD_CODING_BNP)
GRN = "00915A"
DARK = "1C3A2D"
BLU = "1565C0"
WHT = "FFFFFF"
GREY = "F5F5F5"
GREY_HDR = "4A5560"

MOIS_NOMS = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
             'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
MOIS_LABELS = {
    '01': 'JANVIER', '02': 'FEVRIER', '03': 'MARS', '04': 'AVRIL',
    '05': 'MAI', '06': 'JUIN', '07': 'JUILLET', '08': 'AOUT',
    '09': 'SEPTEMBRE', '10': 'OCTOBRE', '11': 'NOVEMBRE', '12': 'DECEMBRE'
}

# Positions par défaut (présélection UI — confirmées à l'écran de mapping)
DEFAULT_POSITIONS = {
    'monext_mois':         1,
    'monext_rs':           4,    # nom client
    'monext_id_rp':        9,
    'monext_nb_cartes':   12,
    'monext_depenses':    15,
    'monext_retraits':    17,
    'monext_pnb_first':   19,
    'monext_pnb_last':    55,
    'monext_pnb_exclude': 33,
    'monext_interchange': 21,
}


def mois_label(yyyymm):
    """'202501' -> 'JANVIER 2025'"""
    if not yyyymm or len(yyyymm) != 6:
        return str(yyyymm)
    return f"{MOIS_LABELS.get(yyyymm[4:6], yyyymm[4:6])} {yyyymm[:4]}"


def months_range(start_yyyymm, end_yyyymm):
    """Liste ordonnée des YYYYMM entre deux bornes incluses."""
    out = []
    y, mo = int(start_yyyymm[:4]), int(start_yyyymm[4:6])
    ey, em = int(end_yyyymm[:4]), int(end_yyyymm[4:6])
    while (y, mo) <= (ey, em):
        out.append(f"{y}{mo:02d}")
        mo += 1
        if mo > 12:
            mo = 1
            y += 1
    return out


class MonextRevenusPeriode:
    def __init__(self, args: argparse.Namespace):
        # --- Entree unique : MONEXT consolide ---
        self.files = {"MONEXT": str(args.source_monext)}

        # --- Periode [premier mois -> dernier mois] (YYYYMM) ---
        self.prem = str(args.mois_debut)
        self.dern = str(args.mois_fin)

        # --- Sortie ---
        self.output_dir = Path(args.output_dir)
        self.output_filename = str(args.output_filename)

        # --- Mapping colonnes MONEXT (positions par defaut, 1-based) ---
        # Reprise des DEFAULT_POSITIONS (pres-selection de l'ancienne UI de mapping),
        # converties en noms de colonnes a la lecture du fichier (cf. run()).
        self.positions = dict(DEFAULT_POSITIONS)

        # --- Differe (optionnel) : mapping colonne differe par position 1-based ---
        self.differe_col = int(args.differe_col) if args.differe_col else 0

    def load_csv_smart(self, path, nrows=None):
        _d = _read_duck(path, nrows)
        if _d is not None:
            return _d
        for sep in [';', ',', '\t']:
            for enc in ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']:
                try:
                    test = pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                       keep_default_na=False, na_values=[],
                                       on_bad_lines='skip', nrows=5)
                    if test.shape[1] > 1:
                        return pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                           keep_default_na=False, na_values=[],
                                           on_bad_lines='skip', nrows=nrows)
                except Exception:
                    continue
        return pd.read_csv(path, sep=None, engine='python', dtype=str,
                           on_bad_lines='skip', nrows=nrows)

    # ══════════════════════════════════════════════════════════════════════
    # HELPERS STATIQUES
    # ══════════════════════════════════════════════════════════════════════
    @staticmethod
    def clean_id(series):
        """Nettoie un identifiant en PRÉSERVANT les zéros initiaux."""
        s = series.astype(str).str.strip()
        s = s.replace(['nan', 'NaN', 'None', 'NULL', 'NA', 'N/A', 'NAN', 'NONE'], '')
        mask = s.str.startswith('="') & s.str.endswith('"')
        s = s.where(~mask, s.str[2:-1])
        s = s.str.lstrip("'")
        mask2 = s.str.endswith('.0') & s.str[:-2].str.isdigit()
        return s.where(~mask2, s.str[:-2]).str.strip()

    @staticmethod
    def to_float(series):
        s = series.astype(str)
        s = s.str.replace('"', '', regex=False).str.replace("'", '', regex=False)
        s = s.str.replace(' ', '', regex=False).str.replace('\xa0', '', regex=False).str.replace('\u202f', '', regex=False)
        mask = s.str.endswith('-')
        s = s.where(~mask, '-' + s.str[:-1])
        s = s.str.replace(',', '.', regex=False)
        return pd.to_numeric(s, errors='coerce').fillna(0.0)

    @staticmethod
    def parse_mois(series):
        def _p(val):
            if pd.isna(val):
                return ''
            s = str(val)
            sc = ''.join(c for c in s if c.isdigit() or c in '/-.')
            if not sc:
                return ''
            for pat, fn in [
                (r'^(\d{1,2})[/\-.](\d{4})$', lambda m: f"{m.group(2)}{int(m.group(1)):02d}"),
                (r'^(\d{4})[/\-.](\d{1,2})$', lambda m: f"{m.group(1)}{int(m.group(2)):02d}" if 1 <= int(m.group(2)) <= 12 else None),
                (r'^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})$', lambda m: f"{m.group(3)}{int(m.group(2)):02d}" if 1 <= int(m.group(2)) <= 12 else None),
                (r'^(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})$', lambda m: f"{m.group(1)}{int(m.group(2)):02d}" if 1 <= int(m.group(2)) <= 12 else None),
            ]:
                mt = re.match(pat, sc)
                if mt:
                    r = fn(mt)
                    if r:
                        return r
            if len(sc) == 6 and sc.isdigit() and 1 <= int(sc[4:6]) <= 12:
                return sc
            if len(sc) == 8 and sc.isdigit() and 1 <= int(sc[4:6]) <= 12:
                return sc[:6]
            try:
                dt = pd.to_datetime(s, dayfirst=True, errors='coerce')
                if pd.notna(dt):
                    return dt.strftime('%Y%m')
            except Exception:
                pass
            return ''
        return series.apply(_p)

    @staticmethod
    def differe_to_num(val):
        """Extrait le premier entier d'une valeur différé. None si rien."""
        s = str(val).strip()
        if not s or s.lower() in ('nan', 'none', 'null', ''):
            return None
        m = re.search(r'\d+', s)
        return int(m.group(0)) if m else None

    def upd(self, v, t):
        # Remplace la progress bar / status GUI par un print console.
        print(f"[{int(v * 100):3d}%] {t}")

    # ══════════════════════════════════════════════════════════════════════
    # LANCEMENT (CLI)
    # ══════════════════════════════════════════════════════════════════════
    def run(self):
        # Resolution du mapping a partir des positions par defaut, converties en
        # noms de colonnes apres lecture de l'entete MONEXT. Equivalent CLI de
        # l'ancien ecran de mapping (memes positions DEFAULT_POSITIONS).
        if not self.files["MONEXT"]:
            raise ValueError("Fichier MONEXT obligatoire manquant.")
        src = Path(self.files["MONEXT"])
        if not src.is_file():
            raise FileNotFoundError(f"Fichier MONEXT introuvable : {src}")

        header = self.load_csv_smart(self.files["MONEXT"], nrows=5)
        cols = list(header.columns)
        ncols = len(cols)

        def name_at(pos: int) -> str:
            if pos and 1 <= pos <= ncols:
                return cols[pos - 1]
            raise ValueError(
                f"Position de colonne {pos} hors limites (fichier MONEXT : {ncols} colonnes)."
            )

        # Mapping colonnes obligatoires (memes positions que l'ancienne UI).
        m = {
            'monext_mois':      name_at(self.positions['monext_mois']),
            'monext_rs':        name_at(self.positions['monext_rs']),
            'monext_id_rp':     name_at(self.positions['monext_id_rp']),
            'monext_nb_cartes': name_at(self.positions['monext_nb_cartes']),
            'monext_depenses':  name_at(self.positions['monext_depenses']),
            'monext_retraits':  name_at(self.positions['monext_retraits']),
        }
        # Differe optionnel : mapping colonne differe (position 1-based) ou vide.
        if self.differe_col:
            m['monext_differe'] = name_at(self.differe_col)
        else:
            m['monext_differe'] = ""

        pnb_cfg = {
            'first': self.positions['monext_pnb_first'],
            'last': self.positions['monext_pnb_last'],
            'excl': self.positions['monext_pnb_exclude'],
            'inter': self.positions['monext_interchange'],
        }
        if pnb_cfg['first'] == 0 or pnb_cfg['last'] == 0:
            raise ValueError("Plage PNB non configurée.")
        if pnb_cfg['first'] > pnb_cfg['last']:
            raise ValueError("Première colonne PNB > dernière.")

        prem = self.prem
        dern = self.dern
        if prem > dern:
            raise ValueError("Le premier mois est postérieur au dernier mois.")

        return self.worker(m, pnb_cfg, prem, dern)

    # ══════════════════════════════════════════════════════════════════════
    # WORKER
    # ══════════════════════════════════════════════════════════════════════
    def worker(self, m, pnb_cfg, prem, dern):
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            period_months = months_range(prem, dern)
            nb_months = len(period_months)

            self.upd(0.05, "Chargement MONEXT...")
            df = self.load_csv_smart(self.files["MONEXT"])
            cols = list(df.columns)

            # --- Clés & mesures ---
            self.upd(0.20, "Préparation des colonnes...")
            df['_RP'] = self.clean_id(df[m['monext_id_rp']])
            df['_RS'] = df[m['monext_rs']].astype(str).str.strip()
            df['_MOIS'] = self.parse_mois(df[m['monext_mois']])
            df['_NB_CARTES'] = self.to_float(df[m['monext_nb_cartes']])
            df['_FLUX'] = self.to_float(df[m['monext_depenses']]) + self.to_float(df[m['monext_retraits']])

            # --- PNB ---
            self.upd(0.30, "Calcul PNB...")
            f0 = pnb_cfg['first'] - 1
            l0 = pnb_cfg['last'] - 1
            e0 = pnb_cfg['excl'] - 1 if pnb_cfg['excl'] > 0 else -1
            ic0 = pnb_cfg['inter'] - 1 if pnb_cfg['inter'] > 0 else -1
            pnb_range = cols[f0:l0 + 1]
            excl_name = cols[e0] if 0 <= e0 < len(cols) else None
            ic_name = cols[ic0] if 0 <= ic0 < len(cols) else None
            for c in pnb_range:
                if c in df.columns:
                    df[c] = self.to_float(df[c])
            if ic_name and ic_name in df.columns:
                df[ic_name] = self.to_float(df[ic_name]) * -1
            pnb_filtered = [c for c in pnb_range if c != excl_name]
            df['_PNB'] = df[pnb_filtered].sum(axis=1) if pnb_filtered else 0.0

            # --- Différé (optionnel) ---
            if m['monext_differe'] and m['monext_differe'] in df.columns:
                df['_DIFF_NUM'] = df[m['monext_differe']].apply(self.differe_to_num)
            else:
                df['_DIFF_NUM'] = np.nan

            # --- Filtre période ---
            self.upd(0.40, f"Filtre période ({nb_months} mois)...")
            df_p = df[df['_MOIS'].isin(period_months)].copy()
            if df_p.empty:
                self.upd(0, "Aucune donnée")
                raise ValueError(
                    f"Aucune ligne MONEXT sur la période {mois_label(prem)} -> {mois_label(dern)}.")
            df_p = df_p[df_p['_RP'] != '']

            # --- Agrégats période par RP ---
            self.upd(0.55, "Agrégation par client...")
            base = df_p.groupby('_RP', sort=False).agg(
                FLUX_TOTAL=('_FLUX', 'sum'),
                PNB_TOTAL=('_PNB', 'sum'),
                CARTES_SUM=('_NB_CARTES', 'sum'),
            )
            base['NB_CARTES_MOYEN'] = (base['CARTES_SUM'] / nb_months).round(2)
            base['FLUX_TOTAL'] = base['FLUX_TOTAL'].round(2)
            base['PNB_TOTAL'] = base['PNB_TOTAL'].round(2)

            # nom client = RS la plus fréquente (non vide)
            dfn = df_p[df_p['_RS'] != '']
            if not dfn.empty:
                nom_map = dfn.groupby('_RP')['_RS'].agg(
                    lambda s: s.value_counts().index[0]).to_dict()
            else:
                nom_map = {}

            # différés distincts triés croissant entre crochets
            dfd = df_p[df_p['_DIFF_NUM'].notna()]
            if not dfd.empty:
                diff_map = dfd.groupby('_RP')['_DIFF_NUM'].agg(
                    lambda s: '[' + ', '.join(str(int(x)) for x in sorted(set(s))) + ']').to_dict()
            else:
                diff_map = {}

            # --- Pivots mensuels ---
            self.upd(0.65, "Décomposition mensuelle...")
            piv_flux = df_p.pivot_table(index='_RP', columns='_MOIS', values='_FLUX',
                                        aggfunc='sum', fill_value=0.0).reindex(columns=period_months, fill_value=0.0)
            piv_cartes = df_p.pivot_table(index='_RP', columns='_MOIS', values='_NB_CARTES',
                                          aggfunc='sum', fill_value=0.0).reindex(columns=period_months, fill_value=0.0)
            piv_pnb = df_p.pivot_table(index='_RP', columns='_MOIS', values='_PNB',
                                       aggfunc='sum', fill_value=0.0).reindex(columns=period_months, fill_value=0.0)

            piv_flux.columns = [f"FLUX {mois_label(mc)}" for mc in period_months]
            piv_cartes.columns = [f"CARTES {mois_label(mc)}" for mc in period_months]
            piv_pnb.columns = [f"PNB {mois_label(mc)}" for mc in period_months]

            piv_flux = piv_flux.round(2)
            piv_cartes = piv_cartes.round(2)
            piv_pnb = piv_pnb.round(2)

            # --- Assemblage DATA ---
            self.upd(0.75, "Assemblage feuille DATA...")
            prem_lbl, dern_lbl = mois_label(prem), mois_label(dern)
            col_flux_tot = f"FLUX TOTAL [{prem_lbl} -> {dern_lbl}]"
            col_pnb_tot = f"PNB TOTAL [{prem_lbl} -> {dern_lbl}]"

            ident = pd.DataFrame(index=base.index)
            ident.insert(0, "ID_RP", base.index)
            ident["NOM_CLIENT"] = ident["ID_RP"].map(nom_map).fillna('')
            ident[col_flux_tot] = base['FLUX_TOTAL']
            ident["NB_CARTES_MOYEN"] = base['NB_CARTES_MOYEN']
            ident["DIFFERES"] = ident["ID_RP"].map(diff_map).fillna('')
            ident[col_pnb_tot] = base['PNB_TOTAL']

            data = pd.concat([ident, piv_flux, piv_cartes, piv_pnb], axis=1)
            data = data.sort_values(col_flux_tot, ascending=False).reset_index(drop=True)

            numeric_cols = ([col_flux_tot, "NB_CARTES_MOYEN", col_pnb_tot]
                            + list(piv_flux.columns) + list(piv_cartes.columns) + list(piv_pnb.columns))

            # --- SYNTHESE ---
            self.upd(0.82, "Construction synthèse...")
            gm = df_p.groupby('_MOIS').agg(
                FLUX=('_FLUX', 'sum'), CARTES=('_NB_CARTES', 'sum'), PNB=('_PNB', 'sum')
            ).reindex(period_months, fill_value=0.0)
            clients_actifs = df_p.groupby('_MOIS')['_RP'].nunique().reindex(period_months, fill_value=0)

            synth_rows = []
            for mc in period_months:
                synth_rows.append({
                    'MOIS': mois_label(mc),
                    'FLUX': round(float(gm.loc[mc, 'FLUX']), 2),
                    'NB_CARTES': round(float(gm.loc[mc, 'CARTES']), 2),
                    'PNB': round(float(gm.loc[mc, 'PNB']), 2),
                    'CLIENTS_ACTIFS': int(clients_actifs.loc[mc]),
                })
            df_synth = pd.DataFrame(synth_rows)

            kpis = {
                'periode': f"{prem_lbl} -> {dern_lbl}",
                'nb_mois': nb_months,
                'nb_clients': int(data.shape[0]),
                'flux_total': round(float(base['FLUX_TOTAL'].sum()), 2),
                'pnb_total': round(float(base['PNB_TOTAL'].sum()), 2),
                'cartes_total': round(float(base['CARTES_SUM'].sum()), 2),
                'cartes_moyen_global': round(float(base['CARTES_SUM'].sum()) / nb_months, 2),
            }

            # --- Export XLSX ---
            if not OPENPYXL_OK:
                raise RuntimeError("openpyxl non installé : pip install openpyxl")

            self.upd(0.88, "Préparation du dossier de sortie...")
            out_dir = self.output_dir
            out_dir.mkdir(parents=True, exist_ok=True)

            out_path = str(out_dir / self.output_filename)
            self.upd(0.92, "Écriture XLSX...")
            self.write_xlsx(out_path, data, numeric_cols, df_synth, kpis, prem_lbl, dern_lbl)

            self.upd(1.0, "Terminé !")
            return out_path, kpis, prem_lbl, dern_lbl, nb_months

        except Exception:
            self.upd(0, "Erreur")
            import traceback
            traceback.print_exc()
            raise

    # ══════════════════════════════════════════════════════════════════════
    # ÉCRITURE XLSX — DATA + SYNTHESE
    # ══════════════════════════════════════════════════════════════════════
    def write_xlsx(self, out_path, data, numeric_cols, df_synth, kpis, prem_lbl, dern_lbl):
        wb = Workbook()
        numeric_set = set(numeric_cols)

        hfill = PatternFill("solid", fgColor=GRN)
        hfont = Font(bold=True, color=WHT, size=10)
        center = Alignment(horizontal="center", vertical="center")
        right = Alignment(horizontal="right")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Feuille DATA ──
        ws = wb.active
        ws.title = "DATA"
        ws.column_dimensions['A'].width = 2.5
        ws['B2'] = f"MONEXT — Revenus par client  [{prem_lbl} -> {dern_lbl}]"
        ws['B2'].font = Font(bold=True, size=14)
        ws['B3'] = f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}  —  v{VERSION_ID}"
        ws['B3'].font = Font(italic=True, size=9, color="888888")

        hdr_row = 5
        for ci, cn in enumerate(data.columns, start=2):
            c = ws.cell(row=hdr_row, column=ci, value=cn)
            c.font = hfont
            c.fill = hfill
            c.alignment = center
            c.border = border
            ws.column_dimensions[get_column_letter(ci)].width = 18 if ci <= 3 else 15

        for ri, row_data in enumerate(data.itertuples(index=False), start=hdr_row + 1):
            for ci, (cn, val) in enumerate(zip(data.columns, row_data), start=2):
                cell = ws.cell(row=ri, column=ci)
                if cn in numeric_set:
                    try:
                        cell.value = float(val)
                        cell.number_format = '#,##0.00'
                        cell.alignment = right
                    except (ValueError, TypeError):
                        cell.value = ''
                else:
                    s = '' if val is None else str(val)
                    cell.value = '' if s in ('nan', 'None', 'NaT') else s
        ws.freeze_panes = ws.cell(row=hdr_row + 1, column=4)

        # ── Feuille SYNTHESE ──
        ws2 = wb.create_sheet("SYNTHESE")
        ws2.column_dimensions['A'].width = 2.5
        ws2.column_dimensions['B'].width = 24
        for col in ('C', 'D', 'E', 'F'):
            ws2.column_dimensions[col].width = 18

        ws2['B2'] = "SYNTHÈSE — MONEXT Revenus par période"
        ws2['B2'].font = Font(bold=True, size=14)

        kpi_lines = [
            ("Période", kpis['periode']),
            ("Nombre de mois", kpis['nb_mois']),
            ("Nombre de clients (ID RP)", kpis['nb_clients']),
            ("Flux total", kpis['flux_total']),
            ("PNB total", kpis['pnb_total']),
            ("Cartes — total cumulé", kpis['cartes_total']),
            ("Cartes — moyenne / mois", kpis['cartes_moyen_global']),
        ]
        r = 4
        for label, val in kpi_lines:
            cl = ws2.cell(row=r, column=2, value=label)
            cl.font = Font(bold=True, color=WHT, size=10)
            cl.fill = PatternFill("solid", fgColor=GREY_HDR)
            cl.border = border
            cv = ws2.cell(row=r, column=3, value=val)
            cv.border = border
            if isinstance(val, (int, float)) and label not in ("Nombre de mois", "Nombre de clients (ID RP)"):
                cv.number_format = '#,##0.00'
                cv.alignment = right
            r += 1

        # tableau mensuel
        r += 2
        ws2.cell(row=r, column=2, value="DÉTAIL MENSUEL").font = Font(bold=True, size=12, color="7B1FA2")
        r += 1
        headers = ['MOIS', 'FLUX', 'NB_CARTES', 'PNB', 'CLIENTS_ACTIFS']
        for ci, h in enumerate(headers, start=2):
            c = ws2.cell(row=r, column=ci, value=h)
            c.font = hfont
            c.fill = hfill
            c.alignment = center
            c.border = border
        r += 1
        start_tbl = r
        for _, rd in df_synth.iterrows():
            ws2.cell(row=r, column=2, value=rd['MOIS']).border = border
            for ci, key in enumerate(['FLUX', 'NB_CARTES', 'PNB'], start=3):
                c = ws2.cell(row=r, column=ci, value=float(rd[key]))
                c.number_format = '#,##0.00'
                c.alignment = right
                c.border = border
            c = ws2.cell(row=r, column=6, value=int(rd['CLIENTS_ACTIFS']))
            c.alignment = right
            c.border = border
            r += 1
        # ligne total
        ct = ws2.cell(row=r, column=2, value="TOTAL")
        ct.font = Font(bold=True)
        ct.fill = PatternFill("solid", fgColor=GREY)
        ct.border = border
        for ci, col_letter in zip([3, 4, 5], ['C', 'D', 'E']):
            c = ws2.cell(row=r, column=ci, value=f"=SUM({col_letter}{start_tbl}:{col_letter}{r-1})")
            c.number_format = '#,##0.00'
            c.alignment = right
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=GREY)
            c.border = border
        ws2.cell(row=r, column=6).fill = PatternFill("solid", fgColor=GREY)
        ws2.cell(row=r, column=6).border = border

        wb.save(out_path)


# ═════════════════════════════════════════════════════════════════════════════
# CLI
# ═════════════════════════════════════════════════════════════════════════════
def _valid_yyyymm(value: str) -> str:
    s = str(value).strip()
    if len(s) != 6 or not s.isdigit() or not (1 <= int(s[4:6]) <= 12):
        raise argparse.ArgumentTypeError(
            f"Mois invalide '{value}' : format attendu YYYYMM (ex 202501).")
    return s


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="01.M5VTQ",
        description=(
            f"MONEXT REVENUS PERIODE v1 [{VERSION_ID}] - Agregation des revenus MONEXT "
            "par client (ID RP) sur une periode [premier mois -> dernier mois]. "
            "Sortie XLSX 2 feuilles (DATA + SYNTHESE)."
        ),
    )
    parser.add_argument("--source-monext", required=False, default=None, type=Path,
                        help="CSV MONEXT consolide multi-mois. Defaut: auto-resolu depuis "
                             "03.sources/*/MONEXT_AGREGE/ (dernier fichier numerote).")
    parser.add_argument("--mois-debut", required=True, type=_valid_yyyymm,
                        help="Premier mois inclus, format YYYYMM (ex 202501).")
    parser.add_argument("--mois-fin", required=True, type=_valid_yyyymm,
                        help="Dernier mois inclus, format YYYYMM (ex 202512).")
    parser.add_argument("--differe-col", type=int, default=0,
                        help="Position 1-based de la colonne differe dans MONEXT (optionnel).")
    parser.add_argument("--output-dir", required=True, type=Path,
                        help="Dossier de sortie (obligatoire).")
    parser.add_argument("--output-filename", required=True, type=str,
                        help="Nom du fichier XLSX produit (obligatoire).")
    return parser


def main(argv=None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        # Auto-résolution des sources non fournies (arborescence 03.sources/)
        if not args.source_monext:
            args.source_monext = resolve_source("MONEXT_AGREGE", required=True)
        app = MonextRevenusPeriode(args)
        result = app.run()
        out_path, kpis, prem_lbl, dern_lbl, nb_months = result
        print(
            f"[OK] {Path(out_path).name} genere dans {Path(out_path).parent} | "
            f"Periode {prem_lbl} -> {dern_lbl} ({nb_months} mois) | "
            f"Clients (ID RP) {kpis['nb_clients']} | Flux total {kpis['flux_total']:,.2f} | "
            f"PNB total {kpis['pnb_total']:,.2f} | Cartes (moy/mois) {kpis['cartes_moyen_global']:,.2f}"
        )
        return 0
    except Exception as exc:  # erreur metier / execution
        print(f"[ERREUR] {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())