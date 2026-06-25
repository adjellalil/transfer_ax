# -*- coding: utf-8 -*-
"""
SME REVENUE ANALYZER v2 [C7MWZ]
===============================

DESCRIPTION
    Analyse des revenus SME (cartes / flux / PNB) par marche et segment sur
    2025/2026, successeur de C2SME. Logique inversee : on part des fichiers
    financiers (MONEXT Corporate + WORLDLINE/PRGM Centralise), pas du PARC.
      1. Consolidation MONEXT + WORLDLINE -> 1 ligne par programme unique
      2. Matching vers PARC CLIENT (RP -> RC -> RS exacte -> RS inclusion)
      3. Classification : tout programme -> BPE ou ENTREPRISE (ENT_DEF si non trouve)
         via OPTIFLUX (IBAN / RS) et BPE_RETAIL (codes agences) optionnel.
    Version CLI autonome (sans GUI) : la logique metier est PRESERVEE A L'IDENTIQUE
    par rapport a l'application customtkinter d'origine.

    CODES DE CALCUL (valides C2SME) :
      FLUX      MX = col 15 + col 17
      PNB       MX = col 19 -> 55, sauf col 33, interchange x -1
      NB_CARTES MX = col 12
      FLUX      WL = col 28 + col 29
      PNB       WL = col 30 -> 36
      NB_CARTES WL = col 17

SOURCES REQUISES
    --monext        (obligatoire)  MONEXT consolide CORPORATE (CSV multi-mois)
    --worldline     (obligatoire)  WORLDLINE PRGM CENTRALISE (CSV multi-mois)
    --parc          (obligatoire)  PARC_CLIENT (enrichissement segment/GA/RMPM)
    --optiflux      (obligatoire)  OPTIFLUX (identification BPE via IBAN/RS)
    --bpe-retail    (optionnel)    BPE RETAIL (codes agences, col 3)

OUTPUTS PRODUITS
    Un classeur Excel (.xlsx) ecrit dans --output-dir avec 2 feuilles :
      - DETAIL   : 1 ligne par programme (MONEXT + WORLDLINE), categorise.
      - SYNTHESE : 2 tableaux 2025 / 2026, ventiles BPE / ENTREPRISE x
                   Corporate (MONEXT) / Centralise (WORLDLINE) x segments.

ARGUMENTS CLI
    --monext PATH            (obligatoire) chemin CSV MONEXT
    --worldline PATH         (obligatoire) chemin CSV WORLDLINE
    --parc PATH              (obligatoire) chemin CSV PARC_CLIENT
    --optiflux PATH          (obligatoire) chemin CSV OPTIFLUX
    --bpe-retail PATH        (optionnel)   chemin CSV BPE RETAIL ; active
                                           l'identification BPE par codes agences
    --segments CODES         (optionnel)   codes segment SME, separes par virgule
                                           (defaut : ER,ME)
    --output-dir DIR         (optionnel)   dossier de sortie (defaut : dossier courant)
    --output-filename NOM    (optionnel)   nom du fichier .xlsx (defaut horodate)

DECOMPOSITION
    main()
    |-- parse argparse -> SMERevenueAnalyzer_C7MWZ(args)
    `-- app.run()
        `-- worker(m, mx_pnb_cfg, wl_pnb_cols, sme_codes)
            |-- load_csv_smart()        chargement CSV robuste (sep/encodage)
            |-- _resolve_mapping()      mapping colonnes depuis DEFAULT_POSITIONS
            |-- Preparation MONEXT      clean_id / norm_rs / to_float / parse_mois,
            |                           PNB (plage + exclusion + interchange),
            |                           dedoublonnage anti-zero
            |-- Preparation WORLDLINE   flux / PNB / plafond actif
            |-- _build_dict()           dictionnaires PARC (RP / RC / RS)
            |-- lookup_parc()           hierarchie de matching RP->RC->RS
            |-- Consolidation MONEXT    1 ligne par programme (ID_RP/ID_RC)
            |-- Consolidation WORLDLINE 1 ligne par ID_RC
            |-- classify()              BPE / ENTREPRISE via OPTIFLUX + BPE_RETAIL
            `-- _generate_xlsx()        feuilles DETAIL + SYNTHESE
"""

import argparse
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np  # conserve a l'identique (dependance C2SME)
import pandas as pd

# ─── Résolution auto des sources + lecture DuckDB (inliné depuis 06.fonctions) ──
def _find_sources_root():
    here = Path(__file__).resolve()
    for d in [here] + list(here.parents):
        c = d / "03.sources"
        if c.is_dir():
            return c
    return None


def resolve_source(name, required=False):
    root = _find_sources_root()
    if not root:
        if required:
            raise FileNotFoundError(f"Dossier 03.sources introuvable depuis {__file__}")
        return None
    folder = next((m for m in root.glob("*/" + name) if m.is_dir()), None)
    if not folder:
        if required:
            raise FileNotFoundError(f"Source introuvable : 03.sources/*/{name}")
        return None
    exts = (".csv", ".xlsx", ".xls", ".txt")
    files = [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() in exts]
    if not files:
        if required:
            raise FileNotFoundError(f"Aucun fichier source dans {folder}")
        return None

    def _np(f):
        m = re.match(r"\s*(\d+)", f.name)
        return (int(m.group(1)) if m else -1, f.name)
    return max(files, key=_np)


def _read_duck(path, nrows=None):
    p = Path(path)
    if nrows is not None or p.suffix.lower() not in (".csv", ".txt", ""):
        return None
    try:
        import duckdb
        return duckdb.read_csv(str(p), all_varchar=True).df()
    except Exception:
        return None


# --- CONFIGURATION ------------------------------------------------------------
VERSION_ID = "C7MWZ"

DEFAULT_SME_CODES = "ER,ME"

DEFAULT_POSITIONS: Dict[str, Any] = {
    'mx_mois':       1,  'mx_rs':         4,  'mx_id_rp':      9,
    'mx_id_rc':      10, 'mx_iban':       11, 'mx_nb_cartes':  12, 'mx_depenses':   15,
    'mx_retraits':   17, 'mx_pnb_first':  19, 'mx_pnb_last':   55,
    'mx_pnb_excl':   33,
    'wl_mois':       2,  'wl_rs':         8,  'wl_id_rc':      40,
    'wl_plafond':    13,
    'wl_nb_cartes':  17, 'wl_dep_1':      28, 'wl_dep_2':      29,
    'wl_pnb_cols':   [30, 31, 32, 33, 34, 35, 36],
    'parc_rp':       1,  'parc_code_agence': 4, 'parc_rmpm':   6,
    'parc_rs':       8,  'parc_segment':  9,  'parc_code_ga':  11,
    'parc_nom_ga':   12, 'parc_rc':       13,
    'bpe_code_agence': 3,
}


# ==============================================================================
class SMERevenueAnalyzer_C7MWZ:

    def __init__(self, args: argparse.Namespace) -> None:
        # Affectation des memes attributs que la version GUI, depuis argparse.
        self.files: Dict[str, str] = {
            "MONEXT":     str(Path(args.monext)) if args.monext else "",
            "WORLDLINE":  str(Path(args.worldline)) if args.worldline else "",
            "PARC":       str(Path(args.parc)) if args.parc else "",
            "OPTIFLUX":   str(Path(args.optiflux)) if args.optiflux else "",
            "BPE_RETAIL": str(Path(args.bpe_retail)) if args.bpe_retail else "",
        }
        self.dfs_preview: Dict[str, pd.DataFrame] = {}
        self.original_cols: Dict[str, List[str]] = {}

        # use_bpe_var (BooleanVar GUI) -> active si --bpe-retail fourni
        self.use_bpe: bool = bool(args.bpe_retail)

        # Codes segment SME (sme_codes_var GUI -> --segments)
        self.sme_codes_str: str = args.segments

        # Sorties (asksaveasfilename GUI -> --output-dir / --output-filename)
        self.output_dir: Path = Path(args.output_dir) if args.output_dir else Path.cwd()
        self.output_filename: Optional[str] = args.output_filename

    # --------------------------------------------------------------------------
    # CHARGEMENT
    # --------------------------------------------------------------------------
    def load_csv_smart(self, path: str, nrows: Optional[int] = None) -> pd.DataFrame:
        _d = _read_duck(path, nrows)
        if _d is not None:
            return _d
        for sep in [';', ',', '\t']:
            for enc in ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']:
                try:
                    df_t = pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                       keep_default_na=False, na_values=[],
                                       on_bad_lines='skip', nrows=5)
                    if df_t.shape[1] > 1:
                        return pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                           keep_default_na=False, na_values=[],
                                           on_bad_lines='skip', nrows=nrows)
                except Exception:
                    continue
        return pd.read_csv(path, sep=None, engine='python', dtype=str,
                           keep_default_na=False, na_values=[], on_bad_lines='skip', nrows=nrows)

    def load_previews(self) -> None:
        required = ["MONEXT", "WORLDLINE", "PARC", "OPTIFLUX"]
        for k in required:
            if not self.files[k]:
                raise ValueError(f"Fichier {k} non selectionne.")
        if self.use_bpe and not self.files["BPE_RETAIL"]:
            raise ValueError("Fichier BPE RETAIL non selectionne.")
        self._prog(0.0, "Lecture des colonnes...")
        keys = required[:]
        if self.use_bpe:
            keys.append("BPE_RETAIL")
        for k in keys:
            self.dfs_preview[k]   = self.load_csv_smart(self.files[k], nrows=5)
            self.original_cols[k] = list(self.dfs_preview[k].columns)

    # --------------------------------------------------------------------------
    # MAPPING DES COLONNES (etape visuelle supprimee)
    # --------------------------------------------------------------------------
    # NOTE : dans la version GUI, l'utilisateur validait/ajustait manuellement le
    # mapping des colonnes (ETAPE 2). Ici, le mapping est resolu automatiquement a
    # partir des positions par defaut DEFAULT_POSITIONS (memes valeurs que les
    # selections par defaut de l'UI). La selection manuelle est donc supprimee, mais
    # les structures m / mx_pnb_cfg / wl_pnb_cols produites sont identiques a celles
    # passees au worker par start_thread() dans la version GUI.
    def _resolve_mapping(self) -> Tuple[Dict[str, str], Dict[str, int], List[str]]:
        # Champs de mapping (libelle UI -> cle de mapping -> position par defaut),
        # repris a l'identique de show_mapping_ui().
        map_fields: List[Tuple[str, str, int]] = [
            # MONEXT
            ("MONEXT", "mx_mois",      DEFAULT_POSITIONS['mx_mois']),
            ("MONEXT", "mx_rs",        DEFAULT_POSITIONS['mx_rs']),
            ("MONEXT", "mx_id_rp",     DEFAULT_POSITIONS['mx_id_rp']),
            ("MONEXT", "mx_id_rc",     DEFAULT_POSITIONS['mx_id_rc']),
            ("MONEXT", "mx_iban",      DEFAULT_POSITIONS['mx_iban']),
            ("MONEXT", "mx_nb_cartes", DEFAULT_POSITIONS['mx_nb_cartes']),
            ("MONEXT", "mx_depenses",  DEFAULT_POSITIONS['mx_depenses']),
            ("MONEXT", "mx_retraits",  DEFAULT_POSITIONS['mx_retraits']),
            # WORLDLINE
            ("WORLDLINE", "wl_mois",      DEFAULT_POSITIONS['wl_mois']),
            ("WORLDLINE", "wl_rs",        DEFAULT_POSITIONS['wl_rs']),
            ("WORLDLINE", "wl_id_rc",     DEFAULT_POSITIONS['wl_id_rc']),
            ("WORLDLINE", "wl_nb_cartes", DEFAULT_POSITIONS['wl_nb_cartes']),
            ("WORLDLINE", "wl_dep_1",     DEFAULT_POSITIONS['wl_dep_1']),
            ("WORLDLINE", "wl_dep_2",     DEFAULT_POSITIONS['wl_dep_2']),
            ("WORLDLINE", "wl_plafond",   DEFAULT_POSITIONS['wl_plafond']),
            # PARC
            ("PARC", "parc_rp",          DEFAULT_POSITIONS['parc_rp']),
            ("PARC", "parc_rmpm",        DEFAULT_POSITIONS['parc_rmpm']),
            ("PARC", "parc_rs",          DEFAULT_POSITIONS['parc_rs']),
            ("PARC", "parc_segment",     DEFAULT_POSITIONS['parc_segment']),
            ("PARC", "parc_code_ga",     DEFAULT_POSITIONS['parc_code_ga']),
            ("PARC", "parc_nom_ga",      DEFAULT_POSITIONS['parc_nom_ga']),
            ("PARC", "parc_rc",          DEFAULT_POSITIONS['parc_rc']),
            ("PARC", "parc_code_agence", DEFAULT_POSITIONS['parc_code_agence']),
            # OPTIFLUX (positions par defaut UI)
            ("OPTIFLUX", "opti_rs",   4),
            ("OPTIFLUX", "opti_iban", 65),
        ]

        def _col_name_at(fkey: str, pos: int) -> str:
            cols = self.original_cols.get(fkey, [])
            if pos and pos <= len(cols):
                return cols[pos - 1]
            # Repli : la version GUI imposait une selection ; en mode par defaut on
            # conserve la derniere colonne si la position depasse (rare).
            return cols[-1] if cols else ""

        m: Dict[str, str] = {}
        for fkey, mk, default_pos in map_fields:
            m[mk] = _col_name_at(fkey, default_pos)

        if self.use_bpe and "BPE_RETAIL" in self.original_cols:
            m["bpe_code_agence"] = _col_name_at(
                "BPE_RETAIL", DEFAULT_POSITIONS['bpe_code_agence'])

        # PNB MONEXT : configuration par INDICES (1-based), comme start_thread().
        mx_pnb_cfg: Dict[str, int] = {
            'first':       DEFAULT_POSITIONS['mx_pnb_first'],
            'last':        DEFAULT_POSITIONS['mx_pnb_last'],
            'excl':        DEFAULT_POSITIONS['mx_pnb_excl'],
            # Interchange : "(Aucune)" par defaut dans l'UI -> 0
            'interchange': 0,
        }

        # PNB WORLDLINE : noms de colonnes (col 30 -> 36).
        wl_cols = self.original_cols.get("WORLDLINE", [])
        wl_pnb_cols: List[str] = [
            wl_cols[pos - 1]
            for pos in DEFAULT_POSITIONS['wl_pnb_cols']
            if pos <= len(wl_cols)
        ]

        return m, mx_pnb_cfg, wl_pnb_cols

    # --------------------------------------------------------------------------
    # NETTOYAGE (valide C2SME / R7SME)
    # --------------------------------------------------------------------------
    @staticmethod
    def clean_id(series: pd.Series) -> pd.Series:
        s = series.astype(str).str.strip()
        s = s.replace(['', 'nan', 'NaN', 'None', 'NULL', 'NA', 'N/A', 'NAN', 'NONE'], '')
        mask_excel = s.str.startswith('="') & s.str.endswith('"')
        s = s.where(~mask_excel, s.str[2:-1])
        s = s.str.lstrip("'")
        mask_dec = s.str.endswith('.0') & s.str[:-2].str.isdigit()
        s = s.where(~mask_dec, s.str[:-2])
        return s.str.strip()

    @staticmethod
    def clean_id_strip0(series: pd.Series) -> pd.Series:
        s = SMERevenueAnalyzer_C7MWZ.clean_id(series)
        stripped = s.str.lstrip('0')
        return stripped.where(stripped != '', s)

    @staticmethod
    def norm_rs(series: pd.Series) -> pd.Series:
        def _n(val: Any) -> str:
            if pd.isna(val) or str(val).strip() == '':
                return ''
            s = str(val).strip().upper()
            s = unicodedata.normalize('NFD', s)
            return ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return series.apply(_n)

    @staticmethod
    def to_float(series: pd.Series) -> pd.Series:
        s = series.astype(str)
        s = s.str.replace('"', '', regex=False).str.replace("'", '', regex=False)
        s = s.str.replace(' ', '', regex=False).str.replace('\xa0', '', regex=False)
        s = s.str.replace(' ', '', regex=False)
        mask_minus = s.str.endswith('-')
        s = s.where(~mask_minus, '-' + s.str[:-1])
        s = s.str.replace(',', '.', regex=False)
        return pd.to_numeric(s, errors='coerce').fillna(0.0)

    @staticmethod
    def parse_mois(series: pd.Series) -> pd.Series:
        def _p(val: Any) -> str:
            if pd.isna(val):
                return ''
            s = str(val).strip()
            if s.startswith('="') and s.endswith('"'):
                s = s[2:-1].strip()
            s = s.lstrip("'").strip()
            if s.endswith('.0') and s[:-2].isdigit():
                s = s[:-2]
            if re.fullmatch(r'\d{6}', s):
                return s
            m = re.match(r'^(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})$', s)
            if m:
                mo = int(m.group(2)); an = m.group(3)
                if 1 <= mo <= 12:
                    return f"{an}{str(mo).zfill(2)}"
            m = re.match(r'^(\d{1,2})[/\-\.](\d{4})$', s)
            if m:
                mo = int(m.group(1)); an = m.group(2)
                if 1 <= mo <= 12:
                    return f"{an}{str(mo).zfill(2)}"
            try:
                dt = pd.to_datetime(s, dayfirst=True, errors='coerce')
                if pd.notna(dt):
                    return dt.strftime('%Y%m')
            except Exception:
                pass
            return ''
        return series.apply(_p)

    # --------------------------------------------------------------------------
    # LANCEMENT
    # --------------------------------------------------------------------------
    def run(self) -> str:
        # Equivalent CLI de start_thread() : charge previews, resout le mapping
        # (anciennement saisi via l'UI), valide, puis appelle worker() en direct
        # (sans threading GUI).
        self.load_previews()
        m, mx_pnb_cfg, wl_pnb_cols = self._resolve_mapping()

        for k, v in m.items():
            if not v:
                raise ValueError(f"Colonne non selectionnee : {k}")

        if not mx_pnb_cfg['first'] or not mx_pnb_cfg['last']:
            raise ValueError("Plage PNB MONEXT non configuree")

        sme_codes = [c.strip().upper() for c in self.sme_codes_str.split(',') if c.strip()]
        if not sme_codes:
            raise ValueError("Codes segment SME vides")

        return self.worker(m, mx_pnb_cfg, wl_pnb_cols, sme_codes)

    def _prog(self, val: float, txt: str) -> None:
        # self._prog (GUI : progressbar + label) -> print console.
        print(f"[{val * 100:5.1f}%] {txt}")

    # --------------------------------------------------------------------------
    # WORKER
    # --------------------------------------------------------------------------
    def worker(self, m: Dict[str, str], mx_pnb_cfg: Dict[str, int],
               wl_pnb_cols: List[str], sme_codes: List[str]) -> str:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        # -- Chargement ------------------------------------------------------
        self._prog(0.02, "Chargement MONEXT...")
        df_mx   = self.load_csv_smart(self.files["MONEXT"])
        self._prog(0.05, "Chargement WORLDLINE...")
        df_wl   = self.load_csv_smart(self.files["WORLDLINE"])
        self._prog(0.08, "Chargement PARC CLIENT...")
        df_parc = self.load_csv_smart(self.files["PARC"])

        self._prog(0.09, "Chargement OPTIFLUX...")
        df_opti = self.load_csv_smart(self.files["OPTIFLUX"])

        set_bpe: set = set()
        if self.use_bpe and self.files["BPE_RETAIL"]:
            self._prog(0.095, "Chargement BPE RETAIL...")
            df_bpe  = self.load_csv_smart(self.files["BPE_RETAIL"])
            bpe_col = m.get('bpe_code_agence', df_bpe.columns[2])
            set_bpe = set(
                self.clean_id_strip0(df_bpe[bpe_col])
                .replace('', pd.NA).dropna().unique()
            )

        # -- Preparation MONEXT ----------------------------------------------
        self._prog(0.10, "Preparation MONEXT...")

        # CRITIQUE : capturer les colonnes ORIGINALES avant tout ajout
        # Les indices f0/l0/e0/ic0 referencent les positions du fichier source
        mx_cols_orig = list(df_mx.columns)

        df_mx['_K_RP']      = self.clean_id_strip0(df_mx[m['mx_id_rp']])
        df_mx['_K_RP_SAFE'] = self.clean_id(df_mx[m['mx_id_rp']])
        df_mx['_K_RC']      = self.clean_id_strip0(df_mx[m['mx_id_rc']])
        df_mx['_K_RS']      = self.norm_rs(df_mx[m['mx_rs']])
        # IBAN MONEXT col 11 : meme nettoyage que WL (retrait 4 premiers chars code pays)
        _mx_iban_raw        = self.clean_id(df_mx[m['mx_iban']]).str.upper().str.replace(' ', '', regex=False)
        df_mx['_K_IBAN']    = _mx_iban_raw.str[4:].where(_mx_iban_raw.str.len() > 4, _mx_iban_raw)
        df_mx['_MOIS']      = self.parse_mois(df_mx[m['mx_mois']])
        df_mx['_NB_CARTES'] = self.to_float(df_mx[m['mx_nb_cartes']])
        df_mx['_DEPENSES']  = self.to_float(df_mx[m['mx_depenses']])
        df_mx['_RETRAITS']  = self.to_float(df_mx[m['mx_retraits']])
        df_mx['_FLUX']      = df_mx['_DEPENSES'] + df_mx['_RETRAITS']

        # PNB MONEXT - indices bases sur colonnes originales uniquement
        mx_cols     = mx_cols_orig
        f0          = mx_pnb_cfg['first'] - 1
        l0          = mx_pnb_cfg['last']  - 1
        e0          = mx_pnb_cfg['excl']  - 1 if mx_pnb_cfg['excl'] > 0 else -1
        ic0         = mx_pnb_cfg['interchange'] - 1 if mx_pnb_cfg['interchange'] > 0 else -1
        pnb_range   = mx_cols[f0:l0+1]
        excl_name   = mx_cols[e0]  if 0 <= e0  < len(mx_cols) else None
        ic_col_name = mx_cols[ic0] if 0 <= ic0 < len(mx_cols) else None
        for c in pnb_range:
            if c in df_mx.columns:
                df_mx[c] = self.to_float(df_mx[c])
        if ic_col_name and ic_col_name in df_mx.columns:
            df_mx[ic_col_name] = df_mx[ic_col_name] * -1
        pnb_filtered  = [c for c in pnb_range if c != excl_name]
        df_mx['_PNB'] = df_mx[pnb_filtered].sum(axis=1) if pnb_filtered else 0.0

        # Dedoublonnage anti-zero MONEXT (regle 9 agent_skills)
        # Garder la ligne _K_RP_SAFE la plus longue par (RP, MOIS)
        df_mx['_K_RP_LEN'] = df_mx['_K_RP_SAFE'].str.len()
        mask_rp    = df_mx['_K_RP'] != ''
        df_mx_rp   = (df_mx[mask_rp]
                      .sort_values('_K_RP_LEN', ascending=False)
                      .drop_duplicates(subset=['_K_RP', '_MOIS']))
        df_mx_norp = df_mx[~mask_rp]
        df_mx      = pd.concat([df_mx_rp, df_mx_norp], ignore_index=True)
        df_mx.drop(columns=['_K_RP_LEN'], inplace=True)

        # -- Preparation WORLDLINE -------------------------------------------
        self._prog(0.14, "Preparation WORLDLINE...")
        df_wl['_K_RC']      = self.clean_id_strip0(df_wl[m['wl_id_rc']])
        df_wl['_K_RS']      = self.norm_rs(df_wl[m['wl_rs']])
        df_wl['_MOIS']      = self.parse_mois(df_wl[m['wl_mois']])
        df_wl['_NB_CARTES'] = self.to_float(df_wl[m['wl_nb_cartes']])
        df_wl['_DEP1']      = self.to_float(df_wl[m['wl_dep_1']])
        df_wl['_DEP2']      = self.to_float(df_wl[m['wl_dep_2']])
        df_wl['_FLUX']      = df_wl['_DEP1'] + df_wl['_DEP2']
        # Plafond col 13 : un programme WL est "actif" ssi plafond > 0
        df_wl['_PLAFOND']   = self.to_float(df_wl[m['wl_plafond']])
        if wl_pnb_cols:
            for c in wl_pnb_cols:
                df_wl[c] = self.to_float(df_wl[c])
            df_wl['_PNB'] = df_wl[wl_pnb_cols].sum(axis=1)
        else:
            df_wl['_PNB'] = 0.0

        # -- Dictionnaires PARC (regle 8 agent_skills) -----------------------
        # CRITIQUE : la colonne cle du set_index NE FIGURE PAS dans data_cols.
        # On definit des listes separees pour chaque cle possible.
        self._prog(0.18, "Construction dictionnaires PARC...")

        df_parc['_K_RP']      = self.clean_id_strip0(df_parc[m['parc_rp']])
        df_parc['_K_RC']      = self.clean_id_strip0(df_parc[m['parc_rc']])
        df_parc['_K_RS']      = self.norm_rs(df_parc[m['parc_rs']])
        df_parc['_K_RMPM']    = self.clean_id(df_parc[m['parc_rmpm']])
        df_parc['_K_SEGMENT'] = df_parc[m['parc_segment']].astype(str).str.strip().str.upper()
        df_parc['_K_CODE_GA'] = self.clean_id_strip0(df_parc[m['parc_code_ga']])
        df_parc['_K_CODE_AG'] = self.clean_id_strip0(df_parc[m['parc_code_agence']])
        df_parc['_NOM_GA']    = df_parc[m['parc_nom_ga']].astype(str).str.strip()

        # Cle RP -> donnees sans _K_RP (la cle est retiree par set_index)
        DATA_SANS_RP = ['_K_RC', '_K_RMPM', '_K_SEGMENT', '_K_CODE_GA', '_K_CODE_AG', '_NOM_GA']
        # Cle RC -> donnees sans _K_RC
        DATA_SANS_RC = ['_K_RP', '_K_RMPM', '_K_SEGMENT', '_K_CODE_GA', '_K_CODE_AG', '_NOM_GA']
        # Cle RS -> donnees sans _K_RS
        DATA_SANS_RS = ['_K_RP', '_K_RC', '_K_RMPM', '_K_SEGMENT', '_K_CODE_GA', '_K_CODE_AG', '_NOM_GA']

        def _build_dict(df_sub: pd.DataFrame, key_col: str, data_cols: List[str]) -> Dict[str, list]:
            """data_cols NE DOIT PAS contenir key_col (regle 8)."""
            tmp    = df_sub[df_sub[key_col] != ''].drop_duplicates(key_col).copy()
            needed = [key_col] + data_cols
            return tmp[needed].set_index(key_col)[data_cols].T.to_dict('list')

        d_rp = _build_dict(df_parc, '_K_RP', DATA_SANS_RP)
        d_rc = _build_dict(df_parc, '_K_RC', DATA_SANS_RC)
        d_rs = _build_dict(df_parc, '_K_RS', DATA_SANS_RS)

        # -- Dictionnaires OPTIFLUX ------------------------------------------
        # OPTIFLUX identifie les BPE : si IBAN ou RS present dans OPTIFLUX -> BPE
        df_opti['_K_IBAN'] = self.clean_id(df_opti[m.get('opti_iban', df_opti.columns[64] if len(df_opti.columns) > 64 else df_opti.columns[-1])])
        df_opti['_K_IBAN'] = df_opti['_K_IBAN'].str.upper().str.replace(' ', '', regex=False)
        # Retirer 4 premiers chars IBAN WL (code pays)
        df_opti['_K_RS']   = self.norm_rs(df_opti[m.get('opti_rs', df_opti.columns[3] if len(df_opti.columns) > 3 else df_opti.columns[0])])

        set_opti_iban = set(df_opti[df_opti['_K_IBAN'] != '']['_K_IBAN'].unique())
        set_opti_rs   = set(df_opti[df_opti['_K_RS'] != '']['_K_RS'].unique())
        set_opti_rs_long = set(r for r in set_opti_rs if len(r) >= 14)

        RS_MIN  = 14
        rs_incl = [(k, v) for k, v in d_rs.items() if len(k) >= RS_MIN]

        def lookup_parc(rp: str, rc: str, rs: str) -> Tuple[str, Optional[Dict[str, Any]]]:
            """
            Hierarchie : RP -> RC -> RS exacte -> RS inclusion.
            Retourne (method, dict_enrichissement) ou ('NO', None).
            Le dict retourne a toujours les memes cles.
            """
            if rp and rp in d_rp:
                r = d_rp[rp]
                # DATA_SANS_RP = [RC, RMPM, SEGMENT, CODE_GA, CODE_AG, NOM_GA]
                return 'RP', {'ID_RP': rp,    'ID_RC': r[0], 'RMPM': r[1],
                              'SEGMENT': r[2], 'CODE_GA': r[3],
                              'CODE_AGENCE': r[4], 'NOM_GA': r[5]}
            if rc and rc in d_rc:
                r = d_rc[rc]
                # DATA_SANS_RC = [RP, RMPM, SEGMENT, CODE_GA, CODE_AG, NOM_GA]
                return 'RC', {'ID_RP': r[0], 'ID_RC': rc,    'RMPM': r[1],
                              'SEGMENT': r[2], 'CODE_GA': r[3],
                              'CODE_AGENCE': r[4], 'NOM_GA': r[5]}
            if rs and rs in d_rs:
                r = d_rs[rs]
                # DATA_SANS_RS = [RP, RC, RMPM, SEGMENT, CODE_GA, CODE_AG, NOM_GA]
                return 'RS', {'ID_RP': r[0], 'ID_RC': r[1], 'RMPM': r[2],
                              'SEGMENT': r[3], 'CODE_GA': r[4],
                              'CODE_AGENCE': r[5], 'NOM_GA': r[6]}
            if rs and len(rs) >= RS_MIN:
                for (k, r) in rs_incl:
                    if rs in k or k in rs:
                        return 'RS_INCLUSION', {
                            'ID_RP': r[0], 'ID_RC': r[1], 'RMPM': r[2],
                            'SEGMENT': r[3], 'CODE_GA': r[4],
                            'CODE_AGENCE': r[5], 'NOM_GA': r[6]}
            return 'NO', None

        # -- Consolidation MONEXT : 1 ligne par programme (ID_RP unique) -----
        self._prog(0.22, "Consolidation MONEXT...")

        # Cle de groupe = ID_RP si disponible, sinon ID_RC en fallback
        df_mx['_GRP'] = df_mx['_K_RP'].where(df_mx['_K_RP'] != '', df_mx['_K_RC'])
        df_mx = df_mx[df_mx['_GRP'] != ''].copy()

        grp_mx      = df_mx.groupby('_GRP', sort=False)
        grp_mx_keys = list(grp_mx.groups.keys())
        n_mx        = len(grp_mx_keys)
        mx_records  = []

        for i, grp_key in enumerate(grp_mx_keys):
            if i % 2000 == 0:
                self._prog(0.22 + 0.12 * i / max(n_mx, 1),
                           f"MONEXT : {i:,}/{n_mx:,}...")
            rows = grp_mx.get_group(grp_key)

            rp_vals      = rows[rows['_K_RP'] != '']['_K_RP'].unique()
            rp_safe_vals = rows[rows['_K_RP_SAFE'] != '']['_K_RP_SAFE'].unique()
            rc_vals      = rows[rows['_K_RC'] != '']['_K_RC'].unique()
            rs_vals      = rows[rows['_K_RS'] != '']['_K_RS'].unique()
            iban_vals    = rows[rows['_K_IBAN'] != '']['_K_IBAN'].unique()

            id_rp      = rp_vals[0]      if len(rp_vals)      else ''
            id_rp_safe = rp_safe_vals[0] if len(rp_safe_vals) else id_rp
            id_rc      = rc_vals[0]      if len(rc_vals)      else ''
            rs         = rs_vals[0]      if len(rs_vals)      else ''
            iban_mx    = iban_vals[0]    if len(iban_vals)    else ''

            def _agg(df_r: pd.DataFrame, annee: str) -> Tuple[float, float, float]:
                r = df_r[df_r['_MOIS'].str.startswith(annee)]
                if not len(r):
                    return 0.0, 0.0, 0.0
                return (round(float(r['_FLUX'].sum()),      2),
                        round(float(r['_PNB'].sum()),       2),
                        round(float(r['_NB_CARTES'].sum()), 2))

            fx25, pnb25, nc25 = _agg(rows, '2025')
            fx26, pnb26, nc26 = _agg(rows, '2026')

            method, parc = lookup_parc(id_rp, id_rc, rs)

            mx_records.append({
                'SOURCE':         'MONEXT',
                'ID_RP':          id_rp_safe,
                'ID_RC':          parc['ID_RC']      if parc else id_rc,
                'RAISON_SOCIALE': rs,
                'RMPM':           parc['RMPM']       if parc else '',
                'SEGMENT':        parc['SEGMENT']    if parc else '',
                'CODE_GA':        parc['CODE_GA']    if parc else '',
                'NOM_GA':         parc['NOM_GA']     if parc else '',
                'CODE_AGENCE':    parc['CODE_AGENCE'] if parc else '',
                'MATCH_PARC':     method,
                'IBAN':           iban_mx,
                # MONEXT : actif dans l'annee si flux ou PNB > 0
                'NB_PROG_ACTIF_2025': 1 if (fx25 != 0.0 or pnb25 != 0.0 or nc25 != 0.0) else 0,
                'NB_PROG_ACTIF_2026': 1 if (fx26 != 0.0 or pnb26 != 0.0 or nc26 != 0.0) else 0,
                'FLUX_2025':  fx25, 'PNB_2025':  pnb25, 'NB_CARTES_2025': nc25,
                'FLUX_2026':  fx26, 'PNB_2026':  pnb26, 'NB_CARTES_2026': nc26,
            })

        # -- Consolidation WORLDLINE : 1 ligne par ID_RC unique --------------
        self._prog(0.36, "Consolidation WORLDLINE...")

        df_wl = df_wl[df_wl['_K_RC'] != ''].copy()

        # Programmes actifs WL par annee : ID_RC avec au moins 1 ligne plafond > 0
        # et mois dans l'annee concernee
        rc_actifs_2025 = set(
            df_wl[(df_wl['_PLAFOND'] > 0) & (df_wl['_MOIS'].str.startswith('2025'))]['_K_RC'].unique()
        )
        rc_actifs_2026 = set(
            df_wl[(df_wl['_PLAFOND'] > 0) & (df_wl['_MOIS'].str.startswith('2026'))]['_K_RC'].unique()
        )

        grp_wl      = df_wl.groupby('_K_RC', sort=False)
        grp_wl_keys = list(grp_wl.groups.keys())
        n_wl        = len(grp_wl_keys)
        wl_records  = []

        for i, rc_key in enumerate(grp_wl_keys):
            if i % 2000 == 0:
                self._prog(0.36 + 0.12 * i / max(n_wl, 1),
                           f"WORLDLINE : {i:,}/{n_wl:,}...")
            rows    = grp_wl.get_group(rc_key)
            rs_vals = rows[rows['_K_RS'] != '']['_K_RS'].unique()
            rs      = rs_vals[0] if len(rs_vals) else ''

            def _agg_wl(df_r: pd.DataFrame, annee: str) -> Tuple[float, float, float]:
                r = df_r[df_r['_MOIS'].str.startswith(annee)]
                if not len(r):
                    return 0.0, 0.0, 0.0
                return (round(float(r['_FLUX'].sum()),      2),
                        round(float(r['_PNB'].sum()),       2),
                        round(float(r['_NB_CARTES'].sum()), 2))

            fx25, pnb25, nc25 = _agg_wl(rows, '2025')
            fx26, pnb26, nc26 = _agg_wl(rows, '2026')

            method, parc = lookup_parc('', rc_key, rs)

            # NB_PROG actif par annee : plafond > 0 dans l'annee concernee
            prog_wl_2025 = 1 if rc_key in rc_actifs_2025 else 0
            prog_wl_2026 = 1 if rc_key in rc_actifs_2026 else 0

            wl_records.append({
                'SOURCE':                  'WORLDLINE',
                'ID_RP':                   parc['ID_RP']       if parc else '',
                'ID_RC':                   rc_key,
                'RAISON_SOCIALE':          rs,
                'RMPM':                    parc['RMPM']        if parc else '',
                'SEGMENT':                 parc['SEGMENT']     if parc else '',
                'CODE_GA':                 parc['CODE_GA']     if parc else '',
                'NOM_GA':                  parc['NOM_GA']      if parc else '',
                'CODE_AGENCE':             parc['CODE_AGENCE'] if parc else '',
                'MATCH_PARC':              method,
                'IBAN':                    '',   # WL : matching via RC, pas IBAN
                'NB_PROG_ACTIF_2025':      prog_wl_2025,
                'NB_PROG_ACTIF_2026':      prog_wl_2026,
                'FLUX_2025':  fx25, 'PNB_2025':  pnb25, 'NB_CARTES_2025': nc25,
                'FLUX_2026':  fx26, 'PNB_2026':  pnb26, 'NB_CARTES_2026': nc26,
            })

        # -- Assemblage + classification -------------------------------------
        self._prog(0.50, "Assemblage dataset...")
        df_all = pd.DataFrame(mx_records + wl_records)

        # -- Classification BPE / ENTREPRISE ---------------------------------
        # Logique alignee sur R7GEO / K2P8N (validee mission BNP) :
        #   PARC + OPTIFLUX  -> BPE          (source : PARC_OPTI)
        #   PARC seul        -> ENTREPRISE   (source : PARC)
        #   OPTIFLUX seul    -> BPE          (source : OPTI)
        #   BPE_RETAIL (fbt) -> BPE          (source : BPE_RETAIL)
        #   Rien trouve      -> ENTREPRISE   (source : ENT_DEF)
        # Pas de NON IDENTIFIE - tout finit en BPE ou ENTREPRISE.

        def in_optiflux(rs: str, iban: str = '') -> bool:
            """
            True si present dans OPTIFLUX.
            Ordre : IBAN exact -> RS exacte -> RS inclusion >=14 chars.
            Aligne sur K2P8N / R7GEO.
            """
            # 1. IBAN exact (MONEXT col 11, deja nettoye 4 chars)
            if iban and iban in set_opti_iban:
                return True
            if not rs:
                return False
            # 2. RS exacte
            if rs in set_opti_rs:
                return True
            # 3. RS inclusion
            if len(rs) >= 14:
                for o in set_opti_rs_long:
                    if rs in o or o in rs:
                        return True
            return False

        def classify(row: pd.Series) -> str:
            rs       = str(row.get('RAISON_SOCIALE', '')).strip()
            iban     = str(row.get('IBAN', '')).strip()
            match    = row['MATCH_PARC']    # RP / RC / RS / RS_INCLUSION / NO
            code_ag  = str(row.get('CODE_AGENCE', '')).strip()
            found_op = in_optiflux(rs, iban)

            if match != 'NO':
                # Trouve dans PARC
                if found_op:
                    return 'BPE'        # PARC + OPTIFLUX
                if set_bpe and code_ag in set_bpe:
                    return 'BPE'        # PARC + BPE_RETAIL
                return 'ENTREPRISE'     # PARC seul
            else:
                # Non trouve dans PARC
                if found_op:
                    return 'BPE'        # OPTIFLUX seul
                if set_bpe and code_ag in set_bpe:
                    return 'BPE'        # BPE_RETAIL seul
                return 'ENTREPRISE'     # ENT_DEF (fallback)

        df_all['CATEGORIE'] = df_all.apply(classify, axis=1)

        # -- Export XLSX -----------------------------------------------------
        self._prog(0.55, "Generation XLSX...")
        # asksaveasfilename (GUI) -> chemin construit depuis output_dir/output_filename.
        filename = self.output_filename or f"SME_REVENUE_{ts}_{VERSION_ID}.xlsx"
        if not filename.lower().endswith(".xlsx"):
            filename += ".xlsx"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        save_xlsx = str(self.output_dir / filename)

        self._generate_xlsx(df_all, sme_codes, save_xlsx)
        return save_xlsx

    # --------------------------------------------------------------------------
    # GENERATION XLSX
    # --------------------------------------------------------------------------
    def _generate_xlsx(self, df: pd.DataFrame, sme_codes: List[str], save_xlsx: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # Palette validee C2SME - 6 chars, sans # ni FF (regle 15 agent_skills)
        GRN  = "00915A"; GRN2 = "E8F5E9"; DARK = "1C3A2D"
        GREY = "F5F5F5"; WHT  = "FFFFFF"; BLU  = "1565C0"
        BLU2 = "E3F2FD"; ORG  = "E65100"; ORG2 = "FFF3E0"
        RED2 = "FFEBEE"

        def fill(h):
            return PatternFill(start_color=h, end_color=h, fill_type="solid")
        def fnt(c="FFFFFF", bold=True, sz=10):
            return Font(name="Segoe UI", size=sz, bold=bold, color=c)
        thin = Side(style="thin", color="CCCCCC")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        def C(n):
            return get_column_letter(n)

        n_total = len(df)

        # ======================================================================
        # FEUILLE 1 - DETAIL
        # ======================================================================
        ws1 = wb.active
        ws1.title = "DETAIL"

        ws1.merge_cells("A1:P1")
        ws1["A1"].value = (f"SME REVENUE ANALYZER v2 [{VERSION_ID}] | "
                           f"Segments SME : {', '.join(sme_codes)} | "
                           f"Genere le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        ws1["A1"].font      = fnt(GRN, sz=13)
        ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws1.row_dimensions[1].height = 28

        ws1["A2"].value = (f"{n_total:,} programmes | "
                           f"MONEXT : {len(df[df['SOURCE']=='MONEXT']):,} | "
                           f"WORLDLINE : {len(df[df['SOURCE']=='WORLDLINE']):,}")
        ws1["A2"].font  = fnt("666666", bold=False, sz=9)

        cols_detail = [
            ('SOURCE',         'SOURCE',         12, DARK),
            ('ID_RP',          'ID RP',          14, DARK),
            ('ID_RC',          'ID RC',          14, DARK),
            ('RAISON_SOCIALE', 'RAISON SOCIALE', 35, DARK),
            ('RMPM',           'RMPM',           14, DARK),
            ('SEGMENT',        'SEGMENT',        10, DARK),
            ('CODE_GA',        'CODE GA',        12, DARK),
            ('NOM_GA',         'NOM GA',         28, DARK),
            ('CATEGORIE',      'CATEGORIE',      18, GRN),
            ('MATCH_PARC',     'MATCH PARC',     14, GRN),
            ('NB_PROG_ACTIF_2025', 'NB PROG 2025',  16, BLU),
            ('NB_PROG_ACTIF_2026', 'NB PROG 2026',  16, ORG),
            ('FLUX_2025',      'FLUX 2025',      18, BLU),
            ('PNB_2025',       'PNB 2025',       18, BLU),
            ('NB_CARTES_2025', 'NB CARTES 2025', 16, BLU),
            ('FLUX_2026',      'FLUX 2026',      18, ORG),
            ('PNB_2026',       'PNB 2026',       18, ORG),
            ('NB_CARTES_2026', 'NB CARTES 2026', 16, ORG),
        ]
        num_fields = {'NB_PROG_ACTIF_2025','NB_PROG_ACTIF_2026','FLUX_2025','PNB_2025','NB_CARTES_2025',
                      'FLUX_2026','PNB_2026','NB_CARTES_2026'}

        HDR = 4
        for ci, (field, header, width, color) in enumerate(cols_detail, start=1):
            cell = ws1.cell(row=HDR, column=ci, value=header)
            cell.font      = fnt()
            cell.fill      = fill(color)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = brd
            ws1.column_dimensions[C(ci)].width = width
        ws1.row_dimensions[HDR].height = 40

        cat_bg = {'BPE': GRN2, 'ENTREPRISE': BLU2}

        for ri, row_data in enumerate(df.itertuples(index=False), start=HDR+1):
            cat = str(getattr(row_data, 'CATEGORIE', ''))
            bg  = fill(cat_bg.get(cat, GREY))
            for ci, (field, _, _, _) in enumerate(cols_detail, start=1):
                val = getattr(row_data, field, '')
                if field in num_fields:
                    try:
                        val = float(val)
                    except Exception:
                        pass
                cell = ws1.cell(row=ri, column=ci, value=val)
                cell.font      = Font(name="Segoe UI", size=9)
                cell.fill      = bg
                cell.border    = brd
                cell.alignment = Alignment(
                    horizontal="right" if field in num_fields else "left")
                if field in num_fields and isinstance(val, float):
                    cell.number_format = '#,##0.00'

        ws1.freeze_panes = f"C{HDR+1}"

        # ======================================================================
        # FEUILLE 2 - SYNTHESE
        # Structure :
        #   BPE
        #     Corporate (MONEXT)
        #       BPE / ER
        #       BPE / ME
        #       BPE / Non identifie   (segment vide = ENT_DEF)
        #       BPE / Autres segments (segments presents hors ER/ME)
        #     Centralise (WORLDLINE)
        #       (idem)
        #   ENTREPRISE
        #     Corporate (MONEXT) + sous-lignes
        #     Centralise (WORLDLINE) + sous-lignes
        #   TOTAL GENERAL
        # ======================================================================
        ws2 = wb.create_sheet("SYNTHESE")

        for k in ['NB_PROG_ACTIF_2025','NB_PROG_ACTIF_2026',
                  'FLUX_2025','PNB_2025','NB_CARTES_2025',
                  'FLUX_2026','PNB_2026','NB_CARTES_2026']:
            df[k] = pd.to_numeric(df[k], errors='coerce').fillna(0.0)

        # Segments "nommes" presents dans les donnees (hors vide)
        segs_nommes = sorted([
            s for s in df['SEGMENT'].unique()
            if s and str(s).strip() not in ('', 'nan', 'None')
        ])
        segs_standard = [s for s in sme_codes if s in segs_nommes]   # ER, ME en premier
        segs_autres   = [s for s in segs_nommes if s not in sme_codes]  # autres segments

        def sub_rows_source(sub_cat, cat_lbl, source_lbl):
            """
            Retourne les sous-lignes d'une source (MONEXT ou WORLDLINE)
            pour une categorie donnee (BPE ou ENTREPRISE).
            Lignes : ER, ME, ..., Non identifie, Autres segments
            """
            rows = []
            sub = sub_cat[sub_cat['SOURCE'] == source_lbl]
            if not len(sub):
                return rows

            # Sous-lignes segments standard (ER, ME)
            for seg in segs_standard:
                s = sub[sub['SEGMENT'] == seg]
                if len(s):
                    rows.append((f"    {cat_lbl} / {seg}", 2, s))

            # Sous-ligne Non identifie (segment vide = ENT_DEF)
            s_ni = sub[sub['SEGMENT'].isin(['', 'nan', 'None']) | sub['SEGMENT'].isna()]
            if len(s_ni):
                rows.append((f"    {cat_lbl} / Non identifie", 2, s_ni))

            # Sous-lignes autres segments (GE, TPE, etc.)
            for seg in segs_autres:
                s = sub[sub['SEGMENT'] == seg]
                if len(s):
                    rows.append((f"    {cat_lbl} / {seg}", 2, s))

            return rows

        def build_all_rows():
            """
            Construit toutes les lignes du tableau.
            Chaque element : (libelle, niveau_indent, subset, bg_color, is_total)
            niveau_indent : 0=categorie, 1=source, 2=segment
            """
            rows = []
            for cat, bg_cat, bg_src in [
                ('BPE',        GRN2, "E0F2E9"),
                ('ENTREPRISE', BLU2, "D6EAF8"),
            ]:
                sub_cat = df[df['CATEGORIE'] == cat]
                if not len(sub_cat):
                    continue

                # Ligne categorie
                rows.append((cat, 0, sub_cat, bg_cat, False))

                for src_lbl, src_display in [
                    ('MONEXT',    'Corporate (MONEXT)'),
                    ('WORLDLINE', 'Centralise (WORLDLINE)'),
                ]:
                    sub_src = sub_cat[sub_cat['SOURCE'] == src_lbl]
                    if not len(sub_src):
                        continue

                    # Ligne source
                    rows.append((f"  {src_display}", 1, sub_src, bg_src, False))

                    # Sous-lignes segments
                    for lbl, niveau, sub_seg in sub_rows_source(sub_cat, cat, src_lbl):
                        rows.append((lbl, niveau, sub_seg, GREY, False))

            # Total general
            rows.append(('TOTAL GENERAL', 0, df, DARK, True))
            return rows

        tableau_rows = build_all_rows()

        # -- Mise en page : 2 tableaux cote a cote 2025 / 2026 ----------------
        NB_COLS = 5
        GAP     = 2
        C_2025  = 1
        C_2026  = C_2025 + NB_COLS + GAP
        total_w = NB_COLS * 2 + GAP

        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_w)
        ws2.cell(row=1, column=1,
                 value=(f"SME - Synthese | {', '.join(sme_codes)} | "
                        f"{datetime.now().strftime('%d/%m/%Y %H:%M')}")
                 ).font = fnt(GRN, sz=13)
        ws2.row_dimensions[1].height = 28

        ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_w)
        ws2.cell(row=2, column=1,
                 value=f"{n_total:,} programmes | MONEXT (Corporate) + WORLDLINE (Centralise)"
                 ).font = fnt("666666", bold=False, sz=9)

        ROW_START = 4

        # Couleurs de fond par niveau d'indentation
        INDENT_FILLS = {0: None, 1: None, 2: GREY}  # calcule dynamiquement

        def write_tableau(ws, annee, col_start, rows_data):
            color_titre = BLU if annee == '2025' else ORG

            ws.merge_cells(start_row=ROW_START - 1, start_column=col_start,
                           end_row=ROW_START - 1, end_column=col_start + NB_COLS - 1)
            ct = ws.cell(row=ROW_START - 1, column=col_start, value=f"TABLEAU {annee}")
            ct.font      = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
            ct.fill      = fill(color_titre)
            ct.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[ROW_START - 1].height = 24

            headers    = ["PERIMETRE", "NB PROGRAMMES",
                          f"FLUX {annee} (EUR)", f"NB CARTES {annee}", f"PNB {annee} (EUR)"]
            col_widths = [36, 16, 20, 16, 20]
            for ci, (h, w) in enumerate(zip(headers, col_widths)):
                cell = ws.cell(row=ROW_START, column=col_start + ci, value=h)
                cell.font      = fnt()
                cell.fill      = fill(DARK)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border    = brd
                ws.column_dimensions[C(col_start + ci)].width = w
            ws.row_dimensions[ROW_START].height = 36

            nb_prog_col = f'NB_PROG_ACTIF_{annee}'

            for ri, (lbl, niveau, subset, bg_color, is_total) in enumerate(rows_data):
                row_xlsx  = ROW_START + 1 + ri
                bg        = fill(bg_color)

                if is_total:
                    color_txt = WHT; bold_txt = True
                elif niveau == 0:
                    color_txt = "000000"; bold_txt = True
                elif niveau == 1:
                    color_txt = "000000"; bold_txt = True
                else:
                    color_txt = "444444"; bold_txt = False

                # NB_PROGRAMMES : somme NB_PROG_ACTIF_2025 ou 2026
                nb_prog = int(subset[nb_prog_col].sum()) if nb_prog_col in subset.columns else len(subset)
                flux    = round(float(subset[f'FLUX_{annee}'].sum()),      2)
                nb_c    = round(float(subset[f'NB_CARTES_{annee}'].sum()), 2)
                pnb     = round(float(subset[f'PNB_{annee}'].sum()),       2)

                for ci, val in enumerate([lbl, nb_prog, flux, nb_c, pnb]):
                    cell = ws.cell(row=row_xlsx, column=col_start + ci, value=val)
                    cell.font   = Font(name="Segoe UI", size=9, bold=bold_txt, color=color_txt)
                    cell.fill   = bg
                    cell.border = brd
                    if ci == 0:
                        indent = niveau * 2
                        cell.alignment = Alignment(horizontal="left", indent=indent)
                    else:
                        cell.alignment = Alignment(horizontal="right")
                        if isinstance(val, float):
                            cell.number_format = '#,##0.00'
                ws.row_dimensions[row_xlsx].height = 18

        write_tableau(ws2, '2025', C_2025, tableau_rows)
        write_tableau(ws2, '2026', C_2026, tableau_rows)

        for g in range(GAP):
            ws2.column_dimensions[C(C_2025 + NB_COLS + g)].width = 3

        # -- Sauvegarde -------------------------------------------------------
        self._prog(0.97, "Sauvegarde XLSX...")
        wb.save(save_xlsx)
        self._prog(1.0, "Termine !")

        n_bpe = len(df[df['CATEGORIE'] == 'BPE'])
        n_ent = len(df[df['CATEGORIE'] == 'ENTREPRISE'])

        # messagebox.showinfo (GUI) -> print console.
        print(
            f"Fichier XLSX genere :\n{Path(save_xlsx).name}\n\n"
            f"TOTAL PROGRAMMES : {n_total:,}\n"
            f"  BPE        : {n_bpe:,}\n"
            f"  ENTREPRISE : {n_ent:,}\n\n"
            f"TOTAUX 2025 :\n"
            f"  Flux : {df['FLUX_2025'].sum():,.0f} EUR\n"
            f"  PNB  : {df['PNB_2025'].sum():,.0f} EUR\n\n"
            f"2 feuilles : DETAIL + SYNTHESE"
        )


# --- MAIN ---------------------------------------------------------------------
def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        prog="01.C7MWZ",
        description=f"SME REVENUE ANALYZER v2 [{VERSION_ID}] - "
                    "analyse des revenus SME (CLI autonome, sans GUI).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--monext",    required=False, default=None,
                        help="Chemin du CSV MONEXT consolide CORPORATE (obligatoire).")
    parser.add_argument("--worldline", required=False, default=None,
                        help="Chemin du CSV WORLDLINE PRGM CENTRALISE (obligatoire).")
    parser.add_argument("--parc",      required=False, default=None,
                        help="Chemin du CSV PARC_CLIENT (obligatoire).")
    parser.add_argument("--optiflux",  required=False, default=None,
                        help="Chemin du CSV OPTIFLUX (obligatoire).")
    parser.add_argument("--bpe-retail", dest="bpe_retail", default=None,
                        help="Chemin du CSV BPE RETAIL (optionnel) ; active "
                             "l'identification BPE par codes agences.")
    parser.add_argument("--segments", default=DEFAULT_SME_CODES,
                        help="Codes segment SME, separes par virgule.")
    parser.add_argument("--output-dir", dest="output_dir", default=None,
                        help="Dossier de sortie du fichier .xlsx "
                             "(defaut : dossier courant).")
    parser.add_argument("--output-filename", dest="output_filename", default=None,
                        help="Nom du fichier .xlsx (defaut : SME_REVENUE_<horodate>_"
                             f"{VERSION_ID}.xlsx).")

    try:
        args = parser.parse_args(argv)
    except SystemExit as exc:
        # argparse renvoie 2 en cas d'erreur d'arguments.
        return int(exc.code) if exc.code is not None else 2

    try:
        if not args.monext:
            args.monext = resolve_source("MONEXT_AGREGE", required=True)
        if not args.worldline:
            args.worldline = resolve_source("PRGM_AGREGE", required=True)
        if not args.parc:
            args.parc = resolve_source("PARC_CLIENT", required=True)
        if not args.optiflux:
            args.optiflux = resolve_source("OPTIFLUX", required=True)
        app = SMERevenueAnalyzer_C7MWZ(args)
        out_path = app.run()
    except FileNotFoundError as exc:
        print(f"[ERREUR] Fichier introuvable : {exc}", file=sys.stderr)
        return 1
    except ValueError as exc:
        print(f"[ERREUR] {exc}", file=sys.stderr)
        return 1
    except Exception as exc:  # pragma: no cover
        import traceback
        traceback.print_exc()
        print(f"[ERREUR] {exc}", file=sys.stderr)
        return 1

    print(f"[OK] Analyse terminee. Fichier genere : {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
