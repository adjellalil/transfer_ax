"""
RENAMER EXCEL [HRNAM]
======================

DESCRIPTION
-----------
Outil CLI autonome de renommage en masse de fichiers et dossiers via un
fichier Excel intermediaire servant de table de correspondance :
  - Colonne A : nom actuel (NE PAS MODIFIER)
  - Colonne B : nouveau nom (a modifier)

Le workflow original (application GUI customtkinter) est preserve a
l'identique, mais decoupe en deux sous-commandes (--action) :

  1. generate : scanne --folder, ecrit un XLSX de correspondance
                (--mapping-file) avec la colonne A pre-remplie (nom
                actuel) et la colonne B pre-remplie (= nom actuel, a
                modifier). L'utilisateur edite ensuite la colonne B dans
                Excel puis sauvegarde.

  2. apply    : relit le XLSX (--mapping-file), construit le plan de
                renommage et l'applique sur --folder. Avec --dry-run,
                seul l'apercu est affiche (aucune action disque) ; cela
                reproduit le bouton "Calculer l'apercu" de la GUI.

SECURITES (logique inchangee) :
  - Apercu avant execution (dry-run)
  - Skip si col B vide ou identique au nom actuel
  - Gestion des collisions : suggestion de suffixe _1, _2, ... automatique
  - Distingue fichiers / dossiers
  - Renommage en deux passes (noms temporaires) pour gerer A->B, B->C
  - Validation des caracteres interdits

SOURCES REQUISES
----------------
  - generate : un dossier existant (--folder) contenant les elements a
               renommer.
  - apply    : le dossier (--folder) + le XLSX de correspondance
               (--mapping-file) genere prealablement et edite.
  - Module Python : openpyxl (pip install openpyxl).

OUTPUTS PRODUITS
----------------
  - generate : un fichier XLSX (--mapping-file) contenant la table de
               correspondance (col A nom actuel, col B nouveau nom,
               col C type, col D statut).
  - apply    : les fichiers / dossiers de --folder sont renommes selon la
               colonne B du XLSX. En mode --dry-run : aucune ecriture,
               seulement l'apercu imprime.

ARGUMENTS CLI
-------------
  --action {generate,apply}   (requis) operation a executer.
  --folder PATH               (requis) dossier source a traiter.
  --mapping-file PATH         (requis) fichier XLSX de correspondance
                              (sortie de generate / entree de apply).
  --include-files             inclure les fichiers (defaut : actif).
  --no-include-files          exclure les fichiers.
  --include-dirs              inclure les dossiers (defaut : actif).
  --no-include-dirs           exclure les dossiers.
  --dry-run                   (apply) apercu uniquement, aucune ecriture.

DECOMPOSITION
-------------
  1. parse_args()            : definition argparse + validations.
  2. list_items()            : scan filtre du dossier (fichiers/dossiers).
  3. generate_excel()        : construction et sauvegarde du XLSX.
  4. read_excel_plan()       : relecture XLSX -> plan de renommage.
  5. print_preview()         : affichage de l'apercu (dry-run).
  6. execute_rename()        : application en deux passes.
  7. main()                  : orchestration argparse + codes de sortie.

CODES DE SORTIE
---------------
  0 : succes.
  1 : erreur d'execution (IO, XLSX, renommage, etc.).
  2 : erreur d'arguments / openpyxl manquant.

BNP Paribas Cash Management - Direction Monetique
Mai 2026
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

VERSION_ID = "HRNAM"


# --- ETAPE 1 : listing -----------------------------------------------------
def list_items(source_dir: str, include_files: bool, include_dirs: bool) -> List[Tuple[str, str]]:
    """Retourne la liste des elements du dossier source, filtres."""
    if not source_dir:
        return []
    items: List[Tuple[str, str]] = []
    for name in sorted(os.listdir(source_dir)):
        full = os.path.join(source_dir, name)
        if os.path.isfile(full) and include_files:
            items.append((name, "FICHIER"))
        elif os.path.isdir(full) and include_dirs:
            items.append((name, "DOSSIER"))
    return items


# --- ETAPE 2 : generer Excel ----------------------------------------------
def generate_excel(source_dir: str, save_path: str, include_files: bool, include_dirs: bool) -> int:
    """Genere le XLSX de correspondance. Retourne le nombre de lignes ecrites."""
    items = list_items(source_dir, include_files, include_dirs)
    if not items:
        raise ValueError(
            "Aucun element a renommer (verifier les options Fichiers / Dossiers)"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "RENAMING"

    # En-tetes
    green_fill = PatternFill(start_color="00915A", end_color="00915A", fill_type="solid")
    light_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    white_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Notice en lignes 1-3
    ws["A1"] = f"RENAMER [{VERSION_ID}] - Generer le {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = Font(name="Segoe UI", size=13, bold=True, color="003D2E")
    ws["A1"].fill = light_fill
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Dossier source : {source_dir}"
    ws["A2"].font = Font(name="Segoe UI", size=10, italic=True)
    ws["A2"].fill = light_fill
    ws.merge_cells("A2:D2")

    ws["A3"] = ("NE PAS MODIFIER la colonne A (nom actuel). "
                "Modifier UNIQUEMENT la colonne B (nouveau nom). "
                "Laisser vide ou identique = pas de renommage.")
    ws["A3"].font = Font(name="Segoe UI", size=10, bold=True, color="BF360C")
    ws["A3"].fill = warn_fill
    ws.merge_cells("A3:D3")
    ws.row_dimensions[3].height = 30

    # En-tetes colonnes (ligne 5)
    headers = ["NOM ACTUEL (NE PAS MODIFIER)", "NOUVEAU NOM (MODIFIER ICI)", "TYPE", "STATUT"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=ci, value=h)
        c.font = white_font
        c.fill = green_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd
    ws.row_dimensions[5].height = 35

    # Donnees
    for ri, (name, kind) in enumerate(items, start=6):
        ws.cell(row=ri, column=1, value=name).font = Font(name="Consolas", size=10)
        ws.cell(row=ri, column=2, value=name).font = Font(name="Consolas", size=10, bold=True, color="00915A")
        ws.cell(row=ri, column=3, value=kind).font = Font(name="Segoe UI", size=9, color="666666")
        ws.cell(row=ri, column=4, value="").font = Font(name="Segoe UI", size=9, italic=True, color="999999")
        # Bordures + zebrage
        bg = "F5F5F5" if ri % 2 == 0 else "FFFFFF"
        for ci in range(1, 5):
            cell = ws.cell(row=ri, column=ci)
            cell.border = brd
            if ci != 2:  # col B garde le vert
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

    # Largeurs
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20

    # Freeze panes apres l'en-tete
    ws.freeze_panes = "A6"

    # Sauvegarde
    wb.save(save_path)
    return len(items)


# --- ETAPE 3 : apercu ------------------------------------------------------
def read_excel_plan(source_dir: str, excel_path: str) -> Tuple[List[Dict], List[str]]:
    """Relit le XLSX et construit le plan de renommage.

    Retourne (plan, errors) avec plan = liste de dicts :
    {old, new, kind, status} ou status appartient a
    {RENAME, SKIP_SAME, SKIP_EMPTY, COLLISION, NOT_FOUND}.
    """
    if not excel_path or not os.path.exists(excel_path):
        return [], ["Aucun XLSX genere"]

    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
    except Exception as e:
        return [], [f"Lecture XLSX : {e}"]

    plan: List[Dict] = []
    errors: List[str] = []
    # Donnees a partir de la ligne 6
    new_names_used = set()
    existing_in_dir = set(os.listdir(source_dir)) if source_dir else set()

    for ri in range(6, ws.max_row + 1):
        old = ws.cell(row=ri, column=1).value
        new = ws.cell(row=ri, column=2).value
        kind = ws.cell(row=ri, column=3).value or "?"
        if old is None or str(old).strip() == "":
            continue
        old = str(old).strip()
        new = str(new).strip() if new is not None else ""

        full_old = os.path.join(source_dir, old)
        if not os.path.exists(full_old):
            plan.append({"old": old, "new": new, "kind": kind, "status": "NOT_FOUND",
                         "detail": "n'existe plus dans le dossier source"})
            continue

        if new == "" or new == old:
            plan.append({"old": old, "new": new or old, "kind": kind, "status": "SKIP_SAME",
                         "detail": "nom identique ou vide"})
            continue

        # Validation du nouveau nom
        forbidden = set('/\\:*?"<>|')
        if any(c in forbidden for c in new):
            plan.append({"old": old, "new": new, "kind": kind, "status": "INVALID",
                         "detail": "caracteres interdits"})
            continue

        # Collision : le nouveau nom existe deja (sauf si c'est nous-meme mais la casse change)
        collision = False
        full_new_target = os.path.join(source_dir, new)
        if new in existing_in_dir and new != old:
            collision = True
        if new in new_names_used:
            collision = True

        if collision:
            # Suggerer un nom avec suffixe
            base, ext = os.path.splitext(new)
            suggest = None
            for k in range(1, 100):
                cand = f"{base}_{k}{ext}"
                if cand not in existing_in_dir and cand not in new_names_used:
                    suggest = cand
                    break
            plan.append({"old": old, "new": new, "kind": kind, "status": "COLLISION",
                         "detail": f"existe deja -> suggere : {suggest or '???'}",
                         "suggested": suggest})
            continue

        plan.append({"old": old, "new": new, "kind": kind, "status": "RENAME", "detail": ""})
        new_names_used.add(new)

    return plan, errors


def print_preview(plan: List[Dict], errors: List[str]) -> Tuple[int, int, int, int, int]:
    """Affiche l'apercu (dry-run) et retourne les compteurs."""
    # Compteurs
    n_rename = sum(1 for p in plan if p["status"] == "RENAME")
    n_skip = sum(1 for p in plan if p["status"] == "SKIP_SAME")
    n_coll = sum(1 for p in plan if p["status"] == "COLLISION")
    n_inv = sum(1 for p in plan if p["status"] == "INVALID")
    n_404 = sum(1 for p in plan if p["status"] == "NOT_FOUND")

    print(
        f"A renommer : {n_rename} - Skip : {n_skip} - Collisions : {n_coll} - "
        f"Invalides : {n_inv} - Disparus : {n_404}"
    )

    # Affichage
    if errors:
        print("ERREURS :")
        for e in errors:
            print(f"  - {e}")
        print("")

    print(f"{'STATUT':<12} {'TYPE':<8} {'ANCIEN':<45} -> NOUVEAU")
    print("-" * 130)
    for p in plan:
        sym = {
            "RENAME": "[OK]    ",
            "SKIP_SAME": "[skip]  ",
            "COLLISION": "[COLL!] ",
            "INVALID": "[INV!]  ",
            "NOT_FOUND": "[404]   ",
        }.get(p["status"], "[?]    ")
        old = p["old"][:43] + "..." if len(p["old"]) > 44 else p["old"]
        new = p["new"][:50]
        line = f"{sym:<12} {p['kind']:<8} {old:<45} -> {new}"
        if p.get("detail"):
            line += f"   ({p['detail']})"
        print(line)

    print(f"Apercu : {n_rename} renommage(s) pret(s)")
    return n_rename, n_skip, n_coll, n_inv, n_404


# --- ETAPE 4 : execution ---------------------------------------------------
def execute_rename(source_dir: str, rename_plan: List[Dict]) -> Tuple[int, List[Tuple[str, str]]]:
    """Applique le plan de renommage en deux passes. Retourne (final_ok, ko)."""
    n_rename = sum(1 for p in rename_plan if p["status"] == "RENAME")

    if n_rename == 0:
        raise ValueError("Aucun renommage planifie.")

    ok = 0
    ko: List[Tuple[str, str]] = []

    # Renommage en deux passes pour gerer les cas A->B, B->C : on passe par
    # un nom temporaire pour eviter qu'un renommage en ecrase un autre.
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Phase 1 : tous -> noms temporaires
    temp_map = []
    for p in rename_plan:
        if p["status"] != "RENAME":
            continue
        old_path = os.path.join(source_dir, p["old"])
        tmp_name = f"__RENAMING_{ts}_{ok:05d}__{p['old']}"
        tmp_path = os.path.join(source_dir, tmp_name)
        try:
            os.rename(old_path, tmp_path)
            temp_map.append((tmp_path, os.path.join(source_dir, p["new"]), p))
            ok += 1
        except Exception as e:
            ko.append((p["old"], f"phase1: {e}"))

    # Phase 2 : noms temporaires -> noms finaux
    final_ok = 0
    for tmp_path, final_path, p in temp_map:
        try:
            os.rename(tmp_path, final_path)
            final_ok += 1
        except Exception as e:
            ko.append((p["old"], f"phase2: {e}"))
            # tenter rollback
            try:
                os.rename(tmp_path, os.path.join(source_dir, p["old"]))
            except Exception:
                pass

    return final_ok, ko


# --- ORCHESTRATION ---------------------------------------------------------
def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="07.HRNAM.py",
        description=f"RENAMER EXCEL [{VERSION_ID}] - renommage en masse via XLSX.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--action", required=True, choices=["generate", "apply"],
        help="generate : creer le XLSX de correspondance ; "
             "apply : appliquer le renommage depuis le XLSX.",
    )
    parser.add_argument(
        "--folder", required=True,
        help="Dossier source a traiter.",
    )
    parser.add_argument(
        "--mapping-file", required=True,
        help="Fichier XLSX de correspondance (sortie de generate / entree de apply).",
    )
    parser.add_argument(
        "--include-files", dest="include_files", action="store_true", default=True,
        help="Inclure les fichiers (defaut : actif).",
    )
    parser.add_argument(
        "--no-include-files", dest="include_files", action="store_false",
        help="Exclure les fichiers.",
    )
    parser.add_argument(
        "--include-dirs", dest="include_dirs", action="store_true", default=True,
        help="Inclure les dossiers (defaut : actif).",
    )
    parser.add_argument(
        "--no-include-dirs", dest="include_dirs", action="store_false",
        help="Exclure les dossiers.",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="(apply) Apercu uniquement, aucune ecriture sur le disque.",
    )
    return parser.parse_args(argv)


def main(argv: List[str]) -> int:
    args = parse_args(argv)

    if not OPENPYXL_OK:
        print("ERREUR : le module openpyxl est requis. Installer avec : pip install openpyxl")
        return 2

    source_dir = str(Path(args.folder))
    mapping_file = str(Path(args.mapping_file))

    if not os.path.isdir(source_dir):
        print(f"ERREUR : dossier introuvable : {source_dir}")
        return 2

    if args.action == "generate":
        try:
            n_rows = generate_excel(
                source_dir, mapping_file, args.include_files, args.include_dirs
            )
        except ValueError as e:
            print(f"ERREUR : {e}")
            return 2
        except Exception as e:
            print(f"ERREUR : Generation XLSX : {e}")
            return 1
        print(f"XLSX genere : {mapping_file} ({n_rows} ligne(s))")
        print("Editer la colonne B dans Excel, sauvegarder, puis relancer avec --action apply.")
        return 0

    # action == apply
    if not os.path.exists(mapping_file):
        print(f"ERREUR : fichier de correspondance introuvable : {mapping_file}")
        return 2

    try:
        plan, errors = read_excel_plan(source_dir, mapping_file)
    except Exception as e:
        print(f"ERREUR : Lecture du plan : {e}")
        return 1

    n_rename, n_skip, n_coll, n_inv, n_404 = print_preview(plan, errors)

    if args.dry_run:
        print("=== DRY-RUN : aucune action effectuee sur le disque ===")
        return 0

    if n_rename == 0:
        print("Rien a faire : aucun renommage planifie.")
        return 0

    print("")
    print(
        f"Confirmation : {n_rename} element(s) seront renommes ; "
        f"{n_coll} collision(s) ignoree(s) ; {n_inv} nom(s) invalide(s) ignore(s)."
    )
    print("Cette action est IRREVERSIBLE.")

    try:
        final_ok, ko = execute_rename(source_dir, plan)
    except ValueError as e:
        print(f"Rien a faire : {e}")
        return 0
    except Exception as e:
        print(f"ERREUR : Renommage : {e}")
        return 1

    # Resultat
    print("")
    print("RENOMMAGE TERMINE")
    print(f"  OK : {final_ok} element(s) renomme(s) avec succes")
    if ko:
        print(f"  KO : {len(ko)} erreur(s)")
        for old, err in ko[:5]:
            print(f"  - {old}: {err[:80]}")
    print(f"=== RENOMMAGE EFFECTUE : {final_ok} OK, {len(ko)} KO ===")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main(sys.argv[1:]))
    except KeyboardInterrupt:
        print("Interrompu par l'utilisateur.")
        sys.exit(1)
