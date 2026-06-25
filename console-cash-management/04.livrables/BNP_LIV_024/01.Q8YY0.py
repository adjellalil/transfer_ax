
"""
CIB COMMISSIONNEMENT ANALYZER v12 [Q8YY0]
==========================================
BNP Paribas Cash Management — Direction Monétique

Successeur de 9RQD9. Inclut tous les fixes précédents + recalcul complet par année.

CHANGEMENTS Q8YY0 vs 9RQD9 :
------------------------------
HÉRITÉ :
  - 3ALDI : fix IBAN tronqué vs complet.
  - 2U6U3 : fix OVERRIDE post-matching + PAYS_FINAL priorité override.
  - SBECI : resolve_row enrichissement indépendant par champ.
  - 1VTD2 : fix FINANCIAL par année (recalcul PLAGE/TRIMESTRE).
  - 9RQD9 : _s() sécurisation + diagnostic console.

NOUVEAU — RECALCUL COMPLET PAR ANNÉE (bugs #2 et #3) :
  Bug #2 : PLAFOND_CPC_EUR et PERIODICITE_CPC étaient calculés sur la
  plage complète (max/first de toutes les lignes WL, tous mois confondus).
  Si un client a changé de plafond entre 2025 et 2026, les deux sheets
  FINANCIAL utilisaient le même plafond → coûts RWA faux → APPLICABILITE
  fausse → perte/gain de commission indu.

  Bug #3 : DIFFERE_CCO_TYPE, DIFFERE_CCO_JOURS, DIFFERE_CPC_TYPE,
  DIFFERE_CPC_JOURS étaient figés via fne() sur toutes les lignes du
  client, tous mois confondus. Si un client change de différé en cours
  d'année, les coûts de différé étaient faux.

  FIX : on passe df_wl et df_mx à generate_xlsx_mn8k3_compat. Pour chaque
  année, on filtre les lignes brutes WL/MX par mois, et on recalcule :
  - PLAFOND_CPC_EUR = max des plafonds WL de l'année
  - PERIODICITE_CPC = première périodicité WL de l'année
  - DIFFERE_CCO_TYPE/JOURS = depuis les lignes MX de l'année
  - DIFFERE_CPC_TYPE/JOURS = depuis les lignes WL de l'année

TOUT LE RESTE EST 100% IDENTIQUE à 9RQD9.

=============================================================================
REFACTORING CLI [Q8YY0]
=============================================================================
CIB COMMISSIONNEMENT ANALYZER v12 [Q8YY0]

DESCRIPTION
  Analyseur de commissionnement CIB BNP Paribas (Direction Monetique).
  Recalcule par annee les couts (plafond, periodicite, differe, RWA) et
  produit un CSV + un classeur Excel (DATA + ANALYSE + FINANCIAL avec
  formules Excel vivantes). Version CLI autonome, sans interface graphique.
  La logique metier (worker) est strictement identique a la version GUI ;
  seules les couches d'interface (saisie/sortie) ont ete remplacees.

SOURCES REQUISES
  - WORLDLINE PRGM : flux monetiques Worldline CPC (CSV)
  - MONEXT         : flux monetiques Monext CCO (CSV)
  - REF CLIENT     : referentiel client (RC/RIB -> segment + GA + RMPM) (CSV)
  - IDSEG          : identifiant-segment 3 colonnes, fallback ENT/BPE (CSV)
  - PARC           : parc client, fallback resolution GA/RMPM (CSV)
  - ACCOUNT        : referentiel IBAN -> GA, RMPM, pays (CSV)
  - DEVISES        : taux de change par mois, conversion WL en EUR (CSV)
  - MONITORING     : base monitoring/onboarding CIB, pays apporteur (CSV)
  - (optionnels)   : OPTIFLUX, BPE_RETAIL, SEG_AGENCE, USAGE, MC1, MC2, OVERRIDE

OUTPUTS PRODUITS
  - <output-dir>/<output-filename>.csv   : sheet DATA exportee en CSV
  - <output-dir>/<output-filename>.xlsx  : DATA + ANALYSE_<annee> +
                                           FINANCIAL_<annee> (formules vivantes)

ARGUMENTS CLI
  --wl-prgm PATH (obligatoire) Fichier WORLDLINE PRGM (source CPC)
  --mx-monext PATH (obligatoire) Fichier MONEXT (source CCO)
  --ref-client PATH (obligatoire) Fichier REFERENTIEL CLIENT
  --idseg PATH (obligatoire) Fichier IDENTIFIANT-SEGMENT
  --parc PATH (obligatoire) Fichier PARC_CLIENT
  --account PATH (obligatoire) Fichier IBAN_ACCOUNT
  --devises PATH (obligatoire) Fichier DEVISES
  --monitoring PATH (obligatoire) Fichier MONITORING CIB
  --optiflux PATH (optionnel) Fichier OPTIFLUX (flag BPE par IBAN/RS)
  --bpe-retail PATH (optionnel) Fichier CODE_AGENCE_RETAIL (codes agences BPE)
  --seg-agence PATH (optionnel) Fichier SEG_AGENCE (mapping code agence)
  --usage PATH (optionnel) Fichier MATCHING_USAGE (shortlist produits)
  --mc1 PATH (optionnel) Fichier MATCHING_CLIENT_1 (fallback code GA)
  --mc2 PATH (optionnel) Fichier MATCHING_CLIENT_2 (fallback code GA)
  --override PATH (optionnel) Fichier OVERRIDE_PAYS (ecraser pays clients)
  --override-pays VALEUR (optionnel) Pays de remplacement (defaut: LUXEMBOURG)
  --output-dir PATH (obligatoire) Repertoire de sortie
  --output-filename NOM (obligatoire) Nom de base des fichiers de sortie (sans extension)

DECOMPOSITION
  1. Chargement des sources (CSV multi-encodage/separateur)
     1.1 Sources obligatoires (8) + optionnelles (7)
  2. Construction des dictionnaires de resolution
     2.1 REF CLIENT (RC/RIB), IDSEG, MONITORING CIB, ACCOUNT, PARC, DEVISES
  3. Preparation WORLDLINE / MONEXT
     3.1 Parsing mois, conversion devises, PNB, exclusions pays, override
  4. Matching et resolution (segment, GA, RMPM, pays)
  5. Agregats mensuels et par client
  6. Construction DATA + exports
     6.1 CSV (sheet DATA)
     6.2 XLSX : DATA + ANALYSE_<annee> + FINANCIAL_<annee> (recalcul par annee)
"""

import argparse
import sys
import unicodedata
import re
import re as _re
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ─── Résolution auto des sources + lecture DuckDB (inliné depuis 06.fonctions) ──
def _find_sources_root():
    here = Path(__file__).resolve()
    for d in [here] + list(here.parents):
        c = d / "03.sources"
        if c.is_dir():
            return c
    return None


def resolve_source(name, required=False):
    root = _find_sources_root()
    if not root:
        if required:
            raise FileNotFoundError(f"Dossier 03.sources introuvable depuis {__file__}")
        return None
    folder = next((m for m in root.glob("*/" + name) if m.is_dir()), None)
    if not folder:
        if required:
            raise FileNotFoundError(f"Source introuvable : 03.sources/*/{name}")
        return None
    exts = (".csv", ".xlsx", ".xls", ".txt")
    files = [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() in exts]
    if not files:
        if required:
            raise FileNotFoundError(f"Aucun fichier source dans {folder}")
        return None

    def _np(f):
        m = re.match(r"\s*(\d+)", f.name)
        return (int(m.group(1)) if m else -1, f.name)
    return max(files, key=_np)


def _read_duck(path, nrows=None):
    p = Path(path)
    if nrows is not None or p.suffix.lower() not in (".csv", ".txt", ""):
        return None
    try:
        import duckdb
        return duckdb.read_csv(str(p), all_varchar=True).df()
    except Exception:
        return None


BNP_GREEN  = "#00915A"
VERSION_ID = "Q8YY0"


class _Flag:
    """Substitut CLI d'une variable booleenne Tk : expose .get()."""
    def __init__(self, value: bool = False):
        self._v = bool(value)
    def get(self) -> bool:
        return self._v

MOIS_NOMS = {'01':'JANVIER','02':'FEVRIER','03':'MARS','04':'AVRIL','05':'MAI','06':'JUIN','07':'JUILLET','08':'AOUT','09':'SEPTEMBRE','10':'OCTOBRE','11':'NOVEMBRE','12':'DECEMBRE'}

DEFAULT_POSITIONS = {
    'wl_mois':2,'wl_nom_prog':3,'wl_rs':8,'wl_iban':9,'wl_devise':12,'wl_plafond':13,'wl_periodicite':14,'wl_nb_cartes':17,
    'wl_nb_tr_fourn':21,'wl_nb_tr_cb':22,'wl_nb_tr_visa':23,'wl_dep_1':28,'wl_dep_2':29,'wl_pnb_cols':[30,31,32,33,34,35,36],
    'wl_rc':40,'wl_differe':41,'wl_conv_first':24,'wl_conv_last':36,
    'mx_mois':1,'mx_nom_prog':2,'mx_rs':4,'mx_rp':9,'mx_rc':10,'mx_iban':11,'mx_nb_cartes':12,'mx_differe':13,
    'mx_depenses':15,'mx_nb_transactions':16,'mx_retraits':17,'mx_pnb_first':19,'mx_pnb_last':55,'mx_pnb_excl':33,'mx_interchange':21,
    'parc_rp':1,'parc_rmpm':6,'parc_rs':8,'parc_code_ga':11,'parc_rc':14,
    'acc_pays_ga':1,'acc_code_ga':2,'acc_nom_ga':3,'acc_pays_le':4,'acc_rmpm':5,'acc_nom_le':6,'acc_iban':7,
    'devise_date':1,'devise_code':2,'devise_taux':3,
    'opti_rs':4,'opti_iban':65,
    'bpe_retail_code_agence':3,'seg_code_agence':1,'seg_source':7,
    'usage_id_prog':2,'usage_produit':3,'usage_usage':4,
    'mc1_id_prog':1,'mc1_produit':3,'mc1_code_ga':5,'mc2_id_prog':2,'mc2_produit':3,'mc2_code_ga':5,
    # RÉFÉRENTIEL CLIENT (v12)
    'ref_rib':2,'ref_rmpm':4,'ref_rc':7,'ref_segment':9,'ref_code_ga':10,'ref_nom_ga':11,
    # IDENTIFIANT - SEGMENT (v12)
    'idseg_type':1,'idseg_id':2,'idseg_segment':3,
    # MONITORING CIB (v12b)
    'mon_source':1,'mon_pays_monitoring':3,'mon_pays_onboarding':4,'mon_rc':11,'mon_rp':12,'mon_rmpm':15,
    # OVERRIDE PAYS (J5COM)
    'ovr_type':1,'ovr_valeur':2,'ovr_nom':3,
}

DEFAULT_CONSTANTS = {
    'taux_refinancement':0.0099,'part_capital_banque':0.12,'almt':0.142,'rw_defaut':0.65,
    'cout_carte_cco':40.14,'cout_transaction_cpc':0.59,'cout_rwa_plafond_cco':13.32,
    'taux_ead_porteur':0.10,'taux_ead_entreprise':0.40,'taux_commission':0.20,'taux_tva':0.20,
}

GRN="00915A";GRN2="E8F5E9";DARK="1C3A2D";WHT="FFFFFF";BLU="1565C0";BLU2="E3F2FD";ORA="E65100";ORA2="FFF3E0";PUR="7B1FA2";PUR2="F3E5F5"
COL_GAP=2
GRN_UI="#00915A";GRN2_UI="#E8F5E9";BLU_UI="#1565C0";BLU2_UI="#E3F2FD";ORA_UI="#E65100";ORA2_UI="#FFF3E0";PUR_UI="#7B1FA2";PUR2_UI="#F3E5F5"


class CIBCommissionAnalyzer_Q8YY0:
    def __init__(self, args: argparse.Namespace):
        # ── Fichiers (8 obligatoires + 7 optionnels), memes clefs que la GUI ──
        self.files = {
            "WORLDLINE":  str(args.wl_prgm),
            "MONEXT":     str(args.mx_monext),
            "PARC":       str(args.parc),
            "ACCOUNT":    str(args.account),
            "DEVISES":    str(args.devises),
            "REF_CLIENT": str(args.ref_client),
            "IDSEG":      str(args.idseg),
            "MONITORING": str(args.monitoring),
            "OVERRIDE":   str(args.override) if args.override else "",
            "OPTIFLUX":   str(args.optiflux) if args.optiflux else "",
            "BPE_RETAIL": str(args.bpe_retail) if args.bpe_retail else "",
            "SEG_AGENCE": str(args.seg_agence) if args.seg_agence else "",
            "USAGE":      str(args.usage) if args.usage else "",
            "MC1":        str(args.mc1) if args.mc1 else "",
            "MC2":        str(args.mc2) if args.mc2 else "",
        }
        # ── Flags optionnels (substituts CLI des BooleanVar Tk) ───────────────
        self.use_opti_var        = _Flag(bool(args.optiflux))
        self.use_bpe_retail_var  = _Flag(bool(args.bpe_retail))
        self.use_seg_agence_var  = _Flag(bool(args.seg_agence))
        self.use_usage_var       = _Flag(bool(args.usage))
        self.use_mc1_var         = _Flag(bool(args.mc1))
        self.use_mc2_var         = _Flag(bool(args.mc2))
        self.use_override_var    = _Flag(bool(args.override))

        # ── Pays de remplacement OVERRIDE (defaut UI : LUXEMBOURG) ────────────
        self._override_pays = str(args.override_pays).strip()

        # ── Sorties ───────────────────────────────────────────────────────────
        self.output_dir = Path(args.output_dir)
        self.output_filename = str(args.output_filename)

        # ── Mapping colonnes : comportement par DEFAUT de l'UI (positions
        #    preselectionnees DEFAULT_POSITIONS). Verification colonnes :
        #    deleguee a l'UI web. ──────────────────────────────────────────────
        self.dfs_preview = {}
        self.original_cols = {}
        self._prepare_default_mapping()

    # ──────────────────────────────────────────────────────────────────────────
    # Construction du mapping/parametrage par DEFAUT (remplace les etapes UI)
    # ──────────────────────────────────────────────────────────────────────────
    def _prepare_default_mapping(self):
        DP = DEFAULT_POSITIONS

        def cols_of(key):
            return list(self.load_csv_smart(self.files[key], nrows=5).columns)

        # Colonnes des sources obligatoires
        wl_cols = cols_of("WORLDLINE"); mx_cols = cols_of("MONEXT")
        parc_cols = cols_of("PARC"); acc_cols = cols_of("ACCOUNT")
        dev_cols = cols_of("DEVISES"); ref_cols = cols_of("REF_CLIENT")
        idseg_cols = cols_of("IDSEG"); mon_cols = cols_of("MONITORING")

        def name_at(cols, pos):
            # pos : position 1-based preselectionnee dans l'UI
            idx = pos - 1
            return cols[idx] if 0 <= idx < len(cols) else (cols[-1] if cols else "")

        m = {}
        # WORLDLINE
        for mk in ('wl_mois','wl_nom_prog','wl_rs','wl_iban','wl_devise','wl_plafond',
                   'wl_periodicite','wl_nb_cartes','wl_nb_tr_fourn','wl_nb_tr_cb',
                   'wl_nb_tr_visa','wl_dep_1','wl_dep_2','wl_rc','wl_differe',
                   'wl_conv_first','wl_conv_last'):
            m[mk] = name_at(wl_cols, DP[mk])
        # MONEXT
        for mk in ('mx_mois','mx_nom_prog','mx_rs','mx_rp','mx_rc','mx_iban','mx_nb_cartes',
                   'mx_differe','mx_depenses','mx_nb_transactions','mx_retraits'):
            m[mk] = name_at(mx_cols, DP[mk])
        # PARC
        for mk in ('parc_rp','parc_rmpm','parc_rs','parc_code_ga','parc_rc'):
            m[mk] = name_at(parc_cols, DP[mk])
        # ACCOUNT
        for mk in ('acc_pays_ga','acc_code_ga','acc_nom_ga','acc_pays_le','acc_rmpm',
                   'acc_nom_le','acc_iban'):
            m[mk] = name_at(acc_cols, DP[mk])
        # DEVISES
        for mk in ('devise_date','devise_code','devise_taux'):
            m[mk] = name_at(dev_cols, DP[mk])
        # REF CLIENT
        for mk in ('ref_rib','ref_rmpm','ref_rc','ref_segment','ref_code_ga','ref_nom_ga'):
            m[mk] = name_at(ref_cols, DP[mk])
        # IDSEG
        for mk in ('idseg_type','idseg_id','idseg_segment'):
            m[mk] = name_at(idseg_cols, DP[mk])
        # MONITORING
        for mk in ('mon_source','mon_pays_monitoring','mon_pays_onboarding','mon_rc',
                   'mon_rp','mon_rmpm'):
            m[mk] = name_at(mon_cols, DP[mk])

        # Sources optionnelles (mapping uniquement si activees)
        if self.use_opti_var.get():
            opti_cols = cols_of("OPTIFLUX")
            for mk in ('opti_rs','opti_iban'):
                m[mk] = name_at(opti_cols, DP[mk])
        if self.use_bpe_retail_var.get():
            bpe_cols = cols_of("BPE_RETAIL")
            m['bpe_retail_code_agence'] = name_at(bpe_cols, DP['bpe_retail_code_agence'])
        if self.use_seg_agence_var.get():
            seg_cols = cols_of("SEG_AGENCE")
            for mk in ('seg_code_agence','seg_source'):
                m[mk] = name_at(seg_cols, DP[mk])
        if self.use_usage_var.get():
            usage_cols = cols_of("USAGE")
            for mk in ('usage_id_prog','usage_produit','usage_usage'):
                m[mk] = name_at(usage_cols, DP[mk])
        if self.use_mc1_var.get():
            mc1_cols = cols_of("MC1")
            for mk in ('mc1_id_prog','mc1_produit','mc1_code_ga'):
                m[mk] = name_at(mc1_cols, DP[mk])
        if self.use_mc2_var.get():
            mc2_cols = cols_of("MC2")
            for mk in ('mc2_id_prog','mc2_produit','mc2_code_ga'):
                m[mk] = name_at(mc2_cols, DP[mk])
        if self.use_override_var.get():
            ovr_cols = cols_of("OVERRIDE")
            for mk in ('ovr_type','ovr_valeur','ovr_nom'):
                m[mk] = name_at(ovr_cols, DP[mk])
        self.m = m

        # ── Colonnes pays WORLDLINE / MONEXT (logique identique a _load_pays :
        #    col index 38 pour WL, 6 pour MX, sinon derniere colonne) ──────────
        df_wl_full = self.load_csv_smart(self.files["WORLDLINE"])
        df_mx_full = self.load_csv_smart(self.files["MONEXT"])
        wl_cols_full = list(df_wl_full.columns); mx_cols_full = list(df_mx_full.columns)
        self._wl_pays_col = wl_cols_full[38] if len(wl_cols_full) > 38 else wl_cols_full[-1]
        self._mx_pays_col = mx_cols_full[6] if len(mx_cols_full) > 6 else mx_cols_full[-1]

        # ── Pays a EXCLURE : reproduit le pre-cochage UI (FRANCE / FR) ────────
        pays_wl = sorted(df_wl_full[self._wl_pays_col].astype(str).str.strip()
                         .replace('', np.nan).dropna().unique().tolist())
        pays_mx = sorted(df_mx_full[self._mx_pays_col].astype(str).str.strip()
                         .replace('', np.nan).dropna().unique().tolist())
        self.pays_excl_wl = {p for p in pays_wl if p.strip().upper() in ('FRANCE', 'FR')}
        self.pays_excl_mx = {p for p in pays_mx if p.strip().upper() in ('FRANCE', 'FR')}

        # ── PNB WORLDLINE (positions DEFAULT_POSITIONS['wl_pnb_cols']) ────────
        self.wl_pnb_cols = [wl_cols_full[pos - 1] for pos in DP['wl_pnb_cols']
                            if 0 <= pos - 1 < len(wl_cols_full)]
        # ── PNB MONEXT : plage + exclusion + interchange (indices 1-based) ────
        self.mx_pnb_cfg = {
            'first': DP['mx_pnb_first'], 'last': DP['mx_pnb_last'],
            'excl': DP['mx_pnb_excl'], 'interchange': DP['mx_interchange'],
        }
        # ── Conversion devises WORLDLINE : bornes (indices 1-based) ───────────
        self.wl_conv_cfg = {'first': DP['wl_conv_first'], 'last': DP['wl_conv_last']}

        # ── Constantes financieres (valeurs par defaut) ──────────────────────
        self.constants = dict(DEFAULT_CONSTANTS)

        # ── Plage de mois : pleine plage disponible (defaut UI) ──────────────
        mois_wl = self.parse_mois(df_wl_full[m['wl_mois']]).unique().tolist()
        mois_mx = self.parse_mois(df_mx_full[m['mx_mois']]).unique().tolist()
        all_mois_avail = sorted(set(mo for mo in mois_wl + mois_mx if mo and len(mo) == 6))
        self.plage_cfg = {
            'debut': all_mois_avail[0] if all_mois_avail else '',
            'fin': all_mois_avail[-1] if all_mois_avail else '',
        }

    def load_csv_smart(self, path, nrows=None):
        _d = _read_duck(path, nrows)
        if _d is not None:
            return _d
        for sep in [';', ',', '\t']:
            for enc in ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']:
                try:
                    df = pd.read_csv(path, sep=sep, encoding=enc, dtype=str, keep_default_na=False, na_values=[], on_bad_lines='skip', nrows=5)
                    if df.shape[1] > 1: return pd.read_csv(path, sep=sep, encoding=enc, dtype=str, keep_default_na=False, na_values=[], on_bad_lines='skip', nrows=nrows)
                except: continue
        return pd.read_csv(path, sep=None, engine='python', dtype=str, on_bad_lines='skip', nrows=nrows)

    # ══════════════════════════════════════════════════════════════════════════
    # UTILITAIRES (identique J5COM)
    # ══════════════════════════════════════════════════════════════════════════
    def _col_name(self, s): return s.split(". ", 1)[1] if ". " in s else s
    def _col_idx(self, s):
        try: return int(s.split(". ")[0])
        except: return 0

    @staticmethod
    def clean_id_safe(series):
        s = series.astype(str).str.strip(); s = s.replace(['','nan','NaN','None','NULL','NA','N/A','NAN','NONE'], '')
        mask = s.str.startswith('="') & s.str.endswith('"'); s = s.where(~mask, s.str[2:-1]); s = s.str.lstrip("'")
        mask2 = s.str.endswith('.0') & s.str[:-2].str.isdigit(); return s.where(~mask2, s.str[:-2]).str.strip()
    @staticmethod
    def clean_id_strip0(series):
        s = CIBCommissionAnalyzer_Q8YY0.clean_id_safe(series); stripped = s.str.lstrip('0'); return stripped.where(stripped != '', s)
    @staticmethod
    def clean_iban_truncated(series):
        s = CIBCommissionAnalyzer_Q8YY0.clean_id_safe(series).str.upper().str.replace(' ', '', regex=False); return s.str[4:].where(s.str.len() > 4, s)
    @staticmethod
    def clean_iban_full(series): return CIBCommissionAnalyzer_Q8YY0.clean_id_safe(series).str.upper().str.replace(' ', '', regex=False)
    @staticmethod
    def clean_rib_ref(series): return CIBCommissionAnalyzer_Q8YY0.clean_id_safe(series).str.upper().str.replace(' ', '', regex=False)
    @staticmethod
    def normalize_rs(series):
        def _n(v):
            if pd.isna(v) or str(v).strip() == '': return ''
            s = unicodedata.normalize('NFD', str(v).strip().upper()); return ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return series.apply(_n)
    @staticmethod
    def to_float(series):
        s = series.astype(str)
        for rep in ['"', "'", ' ', '\xa0', '\u202f', '\u20ac']: s = s.str.replace(rep, '', regex=False)
        s = s.str.replace('EUR', '', regex=False); mask = s.str.endswith('-'); s = s.where(~mask, '-' + s.str[:-1])
        s = s.str.replace(',', '.', regex=False); return pd.to_numeric(s, errors='coerce').fillna(0.0)
    @staticmethod
    def parse_mois(series):
        def _p(val):
            if pd.isna(val): return ''
            s = str(val).strip()
            if s.startswith('="') and s.endswith('"'): s = s[2:-1].strip()
            s = s.lstrip("'").strip()
            if s.endswith('.0') and s[:-2].isdigit(): s = s[:-2]
            if _re.fullmatch(r'\d{6}', s): return s
            m = _re.match(r'^(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})$', s)
            if m:
                mo, an = int(m.group(2)), m.group(3)
                if 1 <= mo <= 12: return f"{an}{str(mo).zfill(2)}"
            m = _re.match(r'^(\d{1,2})[/\-\.](\d{4})$', s)
            if m:
                mo, an = int(m.group(1)), m.group(2)
                if 1 <= mo <= 12: return f"{an}{str(mo).zfill(2)}"
            m = _re.match(r'^(\d{4})[/\-\.](\d{1,2})$', s)
            if m:
                an, mo = m.group(1), int(m.group(2))
                if 1 <= mo <= 12: return f"{an}{str(mo).zfill(2)}"
            try:
                dt = pd.to_datetime(s, dayfirst=True, errors='coerce')
                if pd.notna(dt): return dt.strftime('%Y%m')
            except: pass
            return ''
        return series.apply(_p)
    @staticmethod
    def mois_label(yyyymm):
        if len(yyyymm)==6: return f"{yyyymm[:4]}_{MOIS_NOMS.get(yyyymm[4:], yyyymm[4:])}"
        return yyyymm
    @staticmethod
    def protect_id(val):
        if pd.isna(val) or str(val).strip() == '': return ''
        s = str(val).strip()
        if s.startswith('="') and s.endswith('"'): return s
        if s.isdigit(): return f'="{s}"'
        if s.startswith('0') and len(s) > 1 and s[1:].isdigit(): return f'="{s}"'
        return s
    @staticmethod
    def _differe_type(raw, source='CCO'):
        if raw is None: return ''
        s = str(raw).strip().upper()
        if s == '' or s in ('NAN','NONE','NULL'): return ''
        if 'IMM' in s: return 'IMM'
        s_num = s.replace(',','.').replace(' ','')
        try:
            v = float(s_num)
            if v == 0: return 'FIN_MOIS' if source == 'CCO' else 'IMM'
            return 'DIFFERE'
        except ValueError: return s[:16]
    @staticmethod
    def _differe_jours(raw):
        if raw is None: return 0
        s = str(raw).strip().upper()
        if s == '' or 'IMM' in s: return 0
        try: return int(float(s.replace(',','.').replace(' ','')))
        except ValueError: return 0
    @staticmethod
    def build_segment_dicts(df_seg, col_type, col_id, col_segment):
        seg_by_rc, seg_by_rmpm, seg_by_rp, seg_by_ca = {}, {}, {}, {}
        TYPE_MAP = {'RC': seg_by_rc, 'RMPM': seg_by_rmpm, 'RP': seg_by_rp, 'CODE_AGENCE': seg_by_ca}
        for t, raw_id, seg_val in zip(df_seg[col_type].astype(str).str.strip().str.upper().values, df_seg[col_id].astype(str).str.strip().values, df_seg[col_segment].astype(str).str.strip().str.upper().values):
            if 'ENTREPRISE' in seg_val: segment = 'ENTREPRISE'
            elif 'BPE' in seg_val: segment = 'BPE'
            else: continue
            td = TYPE_MAP.get(t)
            if td is None: continue
            clean = raw_id
            if clean.startswith('="') and clean.endswith('"'): clean = clean[2:-1]
            clean = clean.lstrip("'").strip()
            if clean.endswith('.0') and clean[:-2].isdigit(): clean = clean[:-2]
            if not clean: continue
            if clean not in td: td[clean] = segment
            stripped = clean.lstrip('0')
            if stripped and stripped != clean and stripped not in td: td[stripped] = segment
        return seg_by_rc, seg_by_rmpm, seg_by_rp, seg_by_ca

    def _prog(self, v, t): print(f"[{v*100:5.1f}%] {t}")

    # ══════════════════════════════════════════════════════════════════════════
    # LANCEMENT (CLI) — reprend la logique de start_thread, sans thread ni UI
    # ══════════════════════════════════════════════════════════════════════════
    def run(self):
        m = self.m
        pays_excl_wl = self.pays_excl_wl
        pays_excl_mx = self.pays_excl_mx
        wl_pnb_cols = self.wl_pnb_cols
        mx_pnb_cfg = self.mx_pnb_cfg
        wl_conv_cfg = self.wl_conv_cfg
        constants = self.constants
        plage_cfg = self.plage_cfg
        override_pays = self._override_pays if self.use_override_var.get() else ''
        self.worker(m, pays_excl_wl, pays_excl_mx, wl_pnb_cols, mx_pnb_cfg, wl_conv_cfg, constants, plage_cfg, override_pays)

    # ══════════════════════════════════════════════════════════════════════════
    # WORKER — Q8YY0 : resolve_row corrigée (IBAN tronqué vs complet)
    # ══════════════════════════════════════════════════════════════════════════
    def worker(self, m, pays_excl_wl, pays_excl_mx, wl_pnb_cols, mx_pnb_cfg, wl_conv_cfg, constants, plage_cfg, override_pays):
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            self._prog(0.02, "Chargement WORLDLINE..."); df_wl = self.load_csv_smart(self.files["WORLDLINE"])
            self._prog(0.04, "Chargement MONEXT..."); df_mx = self.load_csv_smart(self.files["MONEXT"])
            self._prog(0.06, "Chargement PARC..."); df_parc = self.load_csv_smart(self.files["PARC"])
            self._prog(0.07, "Chargement ACCOUNT..."); df_account = self.load_csv_smart(self.files["ACCOUNT"])
            self._prog(0.08, "Chargement DEVISES..."); df_devises = self.load_csv_smart(self.files["DEVISES"])
            self._prog(0.09, "Chargement REF CLIENT..."); df_ref = self.load_csv_smart(self.files["REF_CLIENT"])
            self._prog(0.095, "Chargement IDSEG..."); df_idseg = self.load_csv_smart(self.files["IDSEG"])
            self._prog(0.098, "Chargement MONITORING CIB..."); df_monitoring = self.load_csv_smart(self.files["MONITORING"])

            df_opti = self.load_csv_smart(self.files["OPTIFLUX"]) if self.use_opti_var.get() else None

            # ── J5COM : OVERRIDE PAYS ────────────────────────────────────────
            ovr_rmpm_set = set(); ovr_idprog_set = set()
            if self.use_override_var.get() and self.files.get("OVERRIDE"):
                self._prog(0.099, "Chargement OVERRIDE...")
                df_ovr = self.load_csv_smart(self.files["OVERRIDE"])
                ovr_types = df_ovr[m.get('ovr_type', df_ovr.columns[0])].astype(str).str.strip().str.upper()
                ovr_vals = self.clean_id_safe(df_ovr[m.get('ovr_valeur', df_ovr.columns[1])])
                ovr_noms = df_ovr[m.get('ovr_nom', df_ovr.columns[2])].astype(str).str.strip() if len(df_ovr.columns) >= 3 else pd.Series('', index=df_ovr.index)
                for t, v, nom in zip(ovr_types.values, ovr_vals.values, ovr_noms.values):
                    if not v: continue
                    if 'RMPM' in t:
                        ovr_rmpm_set.add(v)
                        stripped = v.lstrip('0')
                        if stripped and stripped != v: ovr_rmpm_set.add(stripped)
                    elif 'PROG' in t or 'ID_PROG' in t:
                        ovr_idprog_set.add(v)
                        ovr_idprog_set.add(nom)
                self._prog(0.099, f"OVERRIDE: RMPM={len(ovr_rmpm_set)} ID_PROG={len(ovr_idprog_set)} → pays='{override_pays}'")

            # ── v12 : RÉFÉRENTIEL CLIENT dicts ───────────────────────────────
            self._prog(0.10, "REF CLIENT dicts...")
            ref_rc_c = self.clean_id_strip0(df_ref[m['ref_rc']]); ref_rc_r = self.clean_id_safe(df_ref[m['ref_rc']])
            ref_rib_c = self.clean_rib_ref(df_ref[m['ref_rib']])
            ref_seg = df_ref[m['ref_segment']].astype(str).str.strip().str.upper()
            ref_ga = self.clean_id_strip0(df_ref[m['ref_code_ga']]); ref_nga = df_ref[m['ref_nom_ga']].astype(str).str.strip()
            ref_rmpm = self.clean_id_safe(df_ref[m['ref_rmpm']])
            def ns(s):
                if 'ENTREPRISE' in s: return 'ENTREPRISE'
                elif 'BPE' in s: return 'BPE'
                return ''
            ref_sn = ref_seg.apply(ns)
            d_ref_rc = {}
            for rc, rr, seg, ga, nga, rmpm in zip(ref_rc_c.values, ref_rc_r.values, ref_sn.values, ref_ga.values, ref_nga.values, ref_rmpm.values):
                if rr and rr not in d_ref_rc: d_ref_rc[rr] = (seg, ga, nga, rmpm)
                if rc and rc != rr and rc not in d_ref_rc: d_ref_rc[rc] = (seg, ga, nga, rmpm)
            d_ref_rib = {}
            for rib, seg, ga, nga, rmpm in zip(ref_rib_c.values, ref_sn.values, ref_ga.values, ref_nga.values, ref_rmpm.values):
                if rib and rib not in d_ref_rib: d_ref_rib[rib] = (seg, ga, nga, rmpm)
            self._prog(0.11, f"REF: RC={len(d_ref_rc):,} RIB={len(d_ref_rib):,}")

            # ── v12 : IDSEG dicts ────────────────────────────────────────────
            self._prog(0.12, "IDSEG dicts...")
            seg_by_rc, seg_by_rmpm, seg_by_rp, seg_by_ca = self.build_segment_dicts(df_idseg, m['idseg_type'], m['idseg_id'], m['idseg_segment'])

            # ── v12b : MONITORING CIB dicts ──────────────────────────────────
            self._prog(0.125, "MONITORING CIB dicts...")
            mon_rmpm = self.clean_id_safe(df_monitoring[m['mon_rmpm']])
            mon_rc = self.clean_id_strip0(df_monitoring[m['mon_rc']]); mon_rc_safe = self.clean_id_safe(df_monitoring[m['mon_rc']])
            mon_rp = self.clean_id_strip0(df_monitoring[m['mon_rp']]); mon_rp_safe = self.clean_id_safe(df_monitoring[m['mon_rp']])
            mon_source_col = df_monitoring[m['mon_source']].astype(str).str.strip().str.upper()
            mon_pays_m = df_monitoring[m['mon_pays_monitoring']].astype(str).str.strip()
            mon_pays_o = df_monitoring[m['mon_pays_onboarding']].astype(str).str.strip()
            NA_VALS = {'', 'N/A', 'NA', 'NON APPLICABLE', 'NON-APPLICABLE', 'NAN', 'NONE', 'NULL'}
            def resolve_mon_pays(src, pm, po):
                pm_clean = pm if pm.upper() not in NA_VALS else ''
                po_clean = po if po.upper() not in NA_VALS else ''
                if 'MONITORING' in src.upper(): return pm_clean or po_clean
                if 'ONBOARDING' in src.upper(): return po_clean or pm_clean
                return pm_clean or po_clean
            mon_rmpm_to_pays, mon_rc_to_pays, mon_rp_to_pays = {}, {}, {}
            mon_rmpm_set, mon_rc_set, mon_rp_set = set(), set(), set()
            for rmpm, rc, rcs, rp, rps, src, pm, po in zip(mon_rmpm.values, mon_rc.values, mon_rc_safe.values, mon_rp.values, mon_rp_safe.values, mon_source_col.values, mon_pays_m.values, mon_pays_o.values):
                pays = resolve_mon_pays(src, pm, po)
                if rmpm:
                    mon_rmpm_set.add(rmpm)
                    if rmpm not in mon_rmpm_to_pays and pays: mon_rmpm_to_pays[rmpm] = pays
                for rc_v in [rc, rcs]:
                    if rc_v:
                        mon_rc_set.add(rc_v)
                        if rc_v not in mon_rc_to_pays and pays: mon_rc_to_pays[rc_v] = pays
                for rp_v in [rp, rps]:
                    if rp_v:
                        mon_rp_set.add(rp_v)
                        if rp_v not in mon_rp_to_pays and pays: mon_rp_to_pays[rp_v] = pays
            self._prog(0.13, f"MONITORING: RMPM={len(mon_rmpm_set):,} RC={len(mon_rc_set):,} RP={len(mon_rp_set):,}")

            # ── DEVISES (identique J5COM) ────────────────────────────────────
            self._prog(0.13, "Table devises...")
            dev_dates = df_devises[m['devise_date']].astype(str).str.strip()
            dev_codes = df_devises[m['devise_code']].astype(str).str.strip().str.upper()
            dev_taux = self.to_float(df_devises[m['devise_taux']])
            taux_dict = {}; mois_par_devise = {}
            for d, c, t in zip(dev_dates.values, dev_codes.values, dev_taux.values):
                if not d or not c or t <= 0: continue
                d_clean = ''.join(ch for ch in d if ch.isdigit())
                if len(d_clean) != 6: continue
                taux_dict[(d_clean, c)] = float(t); mois_par_devise.setdefault(c, set()).add(d_clean)
            mois_par_devise = {dev: sorted(list(s)) for dev, s in mois_par_devise.items()}
            taux_dict[('', 'EUR')] = 1.0
            def resolve_taux(mois_val, devise_val):
                if not devise_val or devise_val == 'EUR': return 1.0
                if devise_val not in mois_par_devise: return None
                m_clean = ''.join(ch for ch in str(mois_val) if ch.isdigit())
                if len(m_clean) != 6: return None
                key = (m_clean, devise_val)
                if key in taux_dict: return taux_dict[key]
                candidat = None
                for md in mois_par_devise[devise_val]:
                    if md <= m_clean: candidat = md
                    else: break
                return taux_dict[(candidat, devise_val)] if candidat else None

            # ── ACCOUNT dicts (identique J5COM) ─────────────────────────────
            self._prog(0.15, "ACCOUNT dicts...")
            df_account['_IBAN'] = self.clean_iban_full(df_account[m['acc_iban']]); df_account['_GA'] = self.clean_id_strip0(df_account[m['acc_code_ga']])
            df_account['_NOM_GA'] = df_account[m['acc_nom_ga']].astype(str).str.strip(); df_account['_PAYS_GA'] = df_account[m['acc_pays_ga']].astype(str).str.strip()
            df_account['_RMPM'] = self.clean_id_safe(df_account[m['acc_rmpm']]); df_account['_NOM_LE'] = df_account[m['acc_nom_le']].astype(str).str.strip()
            df_account['_PAYS_LE'] = df_account[m['acc_pays_le']].astype(str).str.strip()
            acc_iban_dict = {}
            for ib, ga, nga, pga, rmpm, nle, ple in zip(df_account['_IBAN'].values, df_account['_GA'].values, df_account['_NOM_GA'].values, df_account['_PAYS_GA'].values, df_account['_RMPM'].values, df_account['_NOM_LE'].values, df_account['_PAYS_LE'].values):
                if ib and ib not in acc_iban_dict: acc_iban_dict[ib] = (ga, nga, pga, rmpm, nle, ple)
            acc_ga2n, acc_ga2p, acc_r2n, acc_r2p = {}, {}, {}, {}
            for ga, nga, pga in zip(df_account['_GA'].values, df_account['_NOM_GA'].values, df_account['_PAYS_GA'].values):
                if ga and ga not in acc_ga2n: acc_ga2n[ga] = nga
                if ga and ga not in acc_ga2p: acc_ga2p[ga] = pga
            for rmpm, nle, ple in zip(df_account['_RMPM'].values, df_account['_NOM_LE'].values, df_account['_PAYS_LE'].values):
                if rmpm and rmpm not in acc_r2n: acc_r2n[rmpm] = nle
                if rmpm and rmpm not in acc_r2p: acc_r2p[rmpm] = ple

            # ── PARC dicts (identique J5COM) ─────────────────────────────────
            self._prog(0.17, "PARC dicts...")
            df_parc['_RP'] = self.clean_id_strip0(df_parc[m['parc_rp']]); df_parc['_RP_SAFE'] = self.clean_id_safe(df_parc[m['parc_rp']])
            df_parc['_RC'] = self.clean_id_strip0(df_parc[m['parc_rc']]); df_parc['_RC_SAFE'] = self.clean_id_safe(df_parc[m['parc_rc']])
            df_parc['_RMPM'] = self.clean_id_safe(df_parc[m['parc_rmpm']]); df_parc['_RS'] = self.normalize_rs(df_parc[m['parc_rs']])
            df_parc['_GA'] = self.clean_id_strip0(df_parc[m['parc_code_ga']])
            PF = ['_RMPM', '_GA', '_RP_SAFE', '_RC_SAFE']
            parc_rp_dict = df_parc[df_parc['_RP']!=''].drop_duplicates('_RP').set_index('_RP')[PF].T.to_dict('list')
            parc_rc_dict = df_parc[df_parc['_RC']!=''].drop_duplicates('_RC').set_index('_RC')[PF].T.to_dict('list')
            parc_rs_dict = df_parc[df_parc['_RS']!=''].drop_duplicates('_RS').set_index('_RS')[PF].T.to_dict('list')

            opti_iban_set, opti_rs_set = set(), set()
            if df_opti is not None:
                df_opti['_IBAN'] = self.clean_iban_full(df_opti[m['opti_iban']]); df_opti['_RS'] = self.normalize_rs(df_opti[m['opti_rs']])
                opti_iban_set = set(df_opti[df_opti['_IBAN']!='']['_IBAN'].unique()); opti_rs_set = set(df_opti[df_opti['_RS']!='']['_RS'].unique())

            # ══════════════════════════════════════════════════════════════════
            # Q8YY0 FIX : resolve_row CORRIGÉE
            # ──────────────────────────────────────────────────────────────────
            # Bug J5COM : iban_tr (tronqué, sans FR76) était utilisé contre
            # d_ref_rib (clés complètes) et acc_iban_dict (clés complètes).
            # Fix : tester AUSSI iban_full contre ces dicts.
            # ══════════════════════════════════════════════════════════════════
            def resolve_row(rp, rc, rs, iban_tr, iban_full):
                """Résolution GA/RMPM/pays avec REF en priorité."""
                rmpm, ga, nom_ga, pays_ga, nom_le, pays_le = '', '', '', '', '', ''
                rp_res, rc_res = rp, rc
                source = "NON_TROUVE"
                segment_val = ''
                seg_source = 'FALLBACK'

                # ══════════════════════════════════════════════════════════════
                # Q8YY0 : ENRICHISSEMENT INDÉPENDANT PAR CHAMP
                # Sécurisé avec .strip() sur toutes les valeurs REF/ACCOUNT/PARC
                # pour éviter les espaces/caractères invisibles qui bloquent
                # les tests falsiness (not ga, not rmpm, etc.)
                # ══════════════════════════════════════════════════════════════

                def _s(v):
                    """Strip sécurisé : retourne '' si vide, espace, None, nan."""
                    if v is None: return ''
                    s = str(v).strip()
                    if s in ('', 'nan', 'NaN', 'None', 'NULL', 'NA', 'N/A'): return ''
                    return s

                # ── PASSE 1 : SEGMENT + GA/NOM_GA/RMPM via REF ──────────────
                # 1a. REF via RC
                rc_safe = rc
                for rc_candidate in [str(rc).strip(), rc_safe]:
                    if rc_candidate and rc_candidate in d_ref_rc:
                        seg, g, ng, rm = d_ref_rc[rc_candidate]
                        seg, g, ng, rm = _s(seg), _s(g), _s(ng), _s(rm)
                        if seg: segment_val = seg; seg_source = "REF_RC"
                        if g: ga = g
                        if ng: nom_ga = ng
                        if rm: rmpm = rm
                        source = "REF_RC"
                        break

                # 1b. REF via RIB (si RC n'a pas matché)
                if source == "NON_TROUVE":
                    ref_rib_hit = None
                    if iban_tr and iban_tr in d_ref_rib:
                        ref_rib_hit = d_ref_rib[iban_tr]
                    elif iban_full and iban_full in d_ref_rib:
                        ref_rib_hit = d_ref_rib[iban_full]
                    if ref_rib_hit:
                        seg, g, ng, rm = ref_rib_hit
                        seg, g, ng, rm = _s(seg), _s(g), _s(ng), _s(rm)
                        if seg: segment_val = seg; seg_source = "REF_IBAN"
                        if g: ga = g
                        if ng: nom_ga = ng
                        if rm: rmpm = rm
                        source = "REF_IBAN"

                # ── PASSE 2 : SEGMENT fallback via IDSEG ─────────────────────
                if not segment_val:
                    for rc_candidate in [str(rc).strip(), rc_safe]:
                        if rc_candidate and rc_candidate in seg_by_rc:
                            segment_val = seg_by_rc[rc_candidate]; seg_source = "IDSEG_RC"; break
                if not segment_val and rmpm:
                    for rm_candidate in [rmpm, rmpm.lstrip('0') or rmpm]:
                        if rm_candidate in seg_by_rmpm:
                            segment_val = seg_by_rmpm[rm_candidate]; seg_source = "IDSEG_RMPM"; break
                if not segment_val and rp:
                    rp_stripped = rp.lstrip('0') or rp
                    for rp_candidate in [rp, rp_stripped]:
                        if rp_candidate in seg_by_rp:
                            segment_val = seg_by_rp[rp_candidate]; seg_source = "IDSEG_RP"; break

                # ── PASSE 3 : ACCOUNT via IBAN — TOUJOURS consulté ───────────
                acc_hit = None
                if iban_full and iban_full in acc_iban_dict:
                    acc_hit = acc_iban_dict[iban_full]
                elif iban_tr and iban_tr in acc_iban_dict:
                    acc_hit = acc_iban_dict[iban_tr]
                if acc_hit:
                    a_ga, a_nga, a_pga, a_rm, a_nle, a_ple = [_s(x) for x in acc_hit]
                    if not ga and a_ga: ga = a_ga
                    if not nom_ga and a_nga: nom_ga = a_nga
                    if not pays_ga and a_pga: pays_ga = a_pga
                    if not rmpm and a_rm: rmpm = a_rm
                    if not nom_le and a_nle: nom_le = a_nle
                    if not pays_le and a_ple: pays_le = a_ple
                    if source == "NON_TROUVE": source = "ACCOUNT_IBAN"

                # ── PASSE 4 : PARC — comble les champs encore vides ──────────
                parc_hit = None
                parc_via = ''
                if rp and rp in parc_rp_dict:
                    parc_hit = parc_rp_dict[rp]; parc_via = "PARC_RP"
                elif rc and rc in parc_rc_dict:
                    parc_hit = parc_rc_dict[rc]; parc_via = "PARC_RC"
                elif rs and rs in parc_rs_dict:
                    parc_hit = parc_rs_dict[rs]; parc_via = "PARC_RS"
                if parc_hit:
                    p_rmpm, p_ga, p_rp, p_rc = [_s(x) for x in parc_hit]
                    if not rmpm and p_rmpm: rmpm = p_rmpm
                    if not ga and p_ga: ga = p_ga
                    if not rp_res and p_rp: rp_res = p_rp
                    if not rc_res and p_rc: rc_res = p_rc
                    if source == "NON_TROUVE": source = parc_via

                # ── PASSE 5 : Enrichissement pays via dicts dérivés ──────────
                if not nom_ga and ga and ga in acc_ga2n: nom_ga = _s(acc_ga2n[ga])
                if not pays_ga and ga and ga in acc_ga2p: pays_ga = _s(acc_ga2p[ga])
                if not nom_le and rmpm and rmpm in acc_r2n: nom_le = _s(acc_r2n[rmpm])
                if not pays_le and rmpm and rmpm in acc_r2p: pays_le = _s(acc_r2p[rmpm])

                return (rmpm, ga, nom_ga, pays_ga, nom_le, pays_le, source, rp_res, rc_res, segment_val, seg_source)

            def flag_bpe(iban_tr, rs, segment_val):
                """BPE si segment direct dit BPE, ou OPTIFLUX."""
                if segment_val == 'BPE': return True
                if segment_val == 'ENTREPRISE': return False
                if not opti_iban_set and not opti_rs_set: return False
                if iban_tr and iban_tr in opti_iban_set: return True
                if rs and rs in opti_rs_set: return True
                return False

            # ── PRÉPARATION WORLDLINE (identique J5COM) ─────────────────────
            self._prog(0.20, "Préparation WORLDLINE...")
            wl_cols = list(df_wl.columns)
            df_wl['_MOIS'] = self.parse_mois(df_wl[m['wl_mois']]); df_wl['_NOM_PROG'] = df_wl[m['wl_nom_prog']].astype(str).str.strip()
            df_wl['_RS'] = self.normalize_rs(df_wl[m['wl_rs']]); df_wl['_IBAN_TR'] = self.clean_iban_truncated(df_wl[m['wl_iban']])
            df_wl['_IBAN_FULL'] = self.clean_iban_full(df_wl[m['wl_iban']]); df_wl['_DEVISE'] = df_wl[m['wl_devise']].astype(str).str.strip().str.upper()
            df_wl['_PERIO'] = self.to_float(df_wl[m['wl_periodicite']]); df_wl['_NB_CARTES'] = self.to_float(df_wl[m['wl_nb_cartes']])
            df_wl['_NB_TR_F'] = self.to_float(df_wl[m['wl_nb_tr_fourn']]); df_wl['_NB_TR_CB'] = self.to_float(df_wl[m['wl_nb_tr_cb']])
            df_wl['_NB_TR_VISA'] = self.to_float(df_wl[m['wl_nb_tr_visa']]); df_wl['_NB_TRANS'] = df_wl['_NB_TR_F'] + df_wl['_NB_TR_CB'] + df_wl['_NB_TR_VISA']
            df_wl['_RC'] = self.clean_id_strip0(df_wl[m['wl_rc']]); df_wl['_RC_SAFE'] = self.clean_id_safe(df_wl[m['wl_rc']])
            df_wl['_DIFFERE'] = df_wl[m['wl_differe']].astype(str).str.strip(); df_wl['_PAYS'] = df_wl[self._wl_pays_col].astype(str).str.strip()

            # Conversion devises WL
            self._prog(0.22, "Taux WL...")
            arr_mois_wl = df_wl['_MOIS'].values; arr_devise_wl = df_wl['_DEVISE'].values; n_wl = len(df_wl)
            cache_t = {}; arr_taux_wl = np.ones(n_wl, dtype=float)
            for i in range(n_wl):
                key = (arr_mois_wl[i], arr_devise_wl[i])
                if key not in cache_t: cache_t[key] = resolve_taux(arr_mois_wl[i], arr_devise_wl[i]) or 1.0
                arr_taux_wl[i] = cache_t[key]
            df_wl['_TAUX'] = arr_taux_wl
            conv_first = wl_conv_cfg['first'] - 1; conv_last = wl_conv_cfg['last'] - 1
            for c in [c for c in wl_cols[conv_first:conv_last + 1] if c in df_wl.columns]:
                df_wl[c] = np.round(self.to_float(df_wl[c]).values * arr_taux_wl, 2)
            df_wl['_PLAFOND'] = self.to_float(df_wl[m['wl_plafond']]) * arr_taux_wl
            df_wl['_DEP1'] = self.to_float(df_wl[m['wl_dep_1']]); df_wl['_DEP2'] = self.to_float(df_wl[m['wl_dep_2']])
            df_wl['_FLUX'] = df_wl['_DEP1'] + df_wl['_DEP2']
            if wl_pnb_cols:
                for c in wl_pnb_cols: df_wl[c] = self.to_float(df_wl[c])
                df_wl['_PNB'] = df_wl[wl_pnb_cols].sum(axis=1)
            else: df_wl['_PNB'] = 0.0
            if pays_excl_wl: df_wl = df_wl[~df_wl['_PAYS'].isin(pays_excl_wl)].copy()
            # Q8YY0 : override WL déplacé APRÈS matching (voir après MATCHING WL)

            # MONEXT prep
            self._prog(0.25, "Préparation MONEXT...")
            mx_cols = list(df_mx.columns)
            df_mx['_MOIS'] = self.parse_mois(df_mx[m['mx_mois']]); df_mx['_NOM_PROG'] = df_mx[m['mx_nom_prog']].astype(str).str.strip()
            df_mx['_RS'] = self.normalize_rs(df_mx[m['mx_rs']]); df_mx['_RP_SAFE'] = self.clean_id_safe(df_mx[m['mx_rp']])
            df_mx['_RP'] = self.clean_id_strip0(df_mx[m['mx_rp']]); df_mx['_RC_SAFE'] = self.clean_id_safe(df_mx[m['mx_rc']])
            df_mx['_RC'] = self.clean_id_strip0(df_mx[m['mx_rc']]); df_mx['_IBAN_TR'] = self.clean_iban_truncated(df_mx[m['mx_iban']])
            df_mx['_IBAN_FULL'] = self.clean_iban_full(df_mx[m['mx_iban']]); df_mx['_NB_CARTES'] = self.to_float(df_mx[m['mx_nb_cartes']])
            df_mx['_DIFFERE'] = df_mx[m['mx_differe']].astype(str).str.strip()
            df_mx['_DEP'] = self.to_float(df_mx[m['mx_depenses']]); df_mx['_NB_TRANS'] = self.to_float(df_mx[m['mx_nb_transactions']])
            df_mx['_RET'] = self.to_float(df_mx[m['mx_retraits']]); df_mx['_FLUX'] = df_mx['_DEP'] + df_mx['_RET']
            df_mx['_PAYS'] = df_mx[self._mx_pays_col].astype(str).str.strip()
            f0 = mx_pnb_cfg['first'] - 1; l0 = mx_pnb_cfg['last'] - 1; e0 = mx_pnb_cfg['excl'] - 1 if mx_pnb_cfg['excl'] > 0 else -1
            ic0 = mx_pnb_cfg['interchange'] - 1 if mx_pnb_cfg['interchange'] > 0 else -1
            pnb_range = mx_cols[f0:l0+1]; excl_name = mx_cols[e0] if 0 <= e0 < len(mx_cols) else None; ic_name = mx_cols[ic0] if 0 <= ic0 < len(mx_cols) else None
            for c in pnb_range:
                if c in df_mx.columns: df_mx[c] = self.to_float(df_mx[c])
            if ic_name and ic_name in df_mx.columns: df_mx[ic_name] = df_mx[ic_name] * -1
            pnb_filtered = [c for c in pnb_range if c != excl_name]
            df_mx['_PNB'] = df_mx[pnb_filtered].sum(axis=1) if pnb_filtered else 0.0
            if pays_excl_mx: df_mx = df_mx[~df_mx['_PAYS'].isin(pays_excl_mx)].copy()
            # Q8YY0 : override MX déplacé APRÈS matching (voir après MATCHING MX)

            # ── MATCHING WL ──────────────────────────────────────────────────
            self._prog(0.35, "Matching WORLDLINE...")
            n_wl = len(df_wl)
            wl_rmpm=np.empty(n_wl,dtype=object); wl_ga=np.empty(n_wl,dtype=object); wl_nom_ga=np.empty(n_wl,dtype=object)
            wl_pays_ga=np.empty(n_wl,dtype=object); wl_nom_le=np.empty(n_wl,dtype=object); wl_pays_le=np.empty(n_wl,dtype=object)
            wl_source=np.empty(n_wl,dtype=object); wl_rp_res=np.empty(n_wl,dtype=object); wl_rc_res=np.empty(n_wl,dtype=object)
            wl_bpe=np.zeros(n_wl,dtype=bool); wl_seg_raw=np.empty(n_wl,dtype=object); wl_seg_src=np.empty(n_wl,dtype=object)
            arr_wl_rs=df_wl['_RS'].values; arr_wl_rc=df_wl['_RC'].values; arr_wl_iban_tr=df_wl['_IBAN_TR'].values; arr_wl_iban_full=df_wl['_IBAN_FULL'].values
            for i in range(n_wl):
                if i % 5000 == 0: self._prog(0.35 + 0.08*i/max(n_wl,1), f"WL {i:,}/{n_wl:,}")
                res = resolve_row('', arr_wl_rc[i], arr_wl_rs[i], arr_wl_iban_tr[i], arr_wl_iban_full[i])
                wl_rmpm[i],wl_ga[i],wl_nom_ga[i],wl_pays_ga[i]=res[0],res[1],res[2],res[3]
                wl_nom_le[i],wl_pays_le[i]=res[4],res[5]; wl_source[i],wl_rp_res[i],wl_rc_res[i]=res[6],res[7],res[8]
                wl_seg_raw[i]=res[9]; wl_seg_src[i]=res[10]
                wl_bpe[i]=flag_bpe(arr_wl_iban_tr[i], arr_wl_rs[i], res[9])
            df_wl['_RMPM_R']=wl_rmpm; df_wl['_GA_R']=wl_ga; df_wl['_NOM_GA']=wl_nom_ga; df_wl['_PAYS_GA']=wl_pays_ga
            df_wl['_NOM_LE']=wl_nom_le; df_wl['_PAYS_LE']=wl_pays_le; df_wl['_SOURCE']=wl_source
            df_wl['_RP_R']=wl_rp_res; df_wl['_RC_R']=wl_rc_res; df_wl['_IS_BPE']=wl_bpe
            df_wl['_SEG_RAW']=wl_seg_raw; df_wl['_SEG_SRC']=wl_seg_src

            # ── Q8YY0 : OVERRIDE WL POST-MATCHING ────────────────────────────
            # Appliqué APRÈS resolve_row pour avoir _RMPM_R et _RC_R disponibles.
            # Teste : RMPM résolu vs ovr_rmpm_set, RC vs ovr_rmpm_set,
            #         NOM_PROG vs ovr_idprog_set.
            if override_pays and (ovr_rmpm_set or ovr_idprog_set):
                wl_rmpm_vals = df_wl['_RMPM_R'].values
                wl_rc_vals = df_wl['_RC_R'].values
                wl_nom_prog_vals = df_wl['_NOM_PROG'].values
                n_ovr_wl = 0
                for i in range(len(df_wl)):
                    hit = False
                    # Test RMPM résolu
                    rmpm_v = str(wl_rmpm_vals[i]).strip()
                    if ovr_rmpm_set and rmpm_v and rmpm_v in ovr_rmpm_set: hit = True
                    # Test RC
                    if not hit:
                        rc_v = str(wl_rc_vals[i]).strip()
                        if ovr_rmpm_set and rc_v and rc_v in ovr_rmpm_set: hit = True
                    # Test NOM_PROG
                    if not hit:
                        nom = str(wl_nom_prog_vals[i]).strip()
                        if ovr_idprog_set and nom and nom in ovr_idprog_set: hit = True
                    if hit:
                        df_wl.iat[i, df_wl.columns.get_loc('_PAYS')] = override_pays; n_ovr_wl += 1
                self._prog(0.435, f"Override WL post-matching: {n_ovr_wl:,} lignes → '{override_pays}'")

            # ── MATCHING MX ──────────────────────────────────────────────────
            self._prog(0.43, "Matching MONEXT...")
            n_mx = len(df_mx)
            mx_rmpm=np.empty(n_mx,dtype=object); mx_ga=np.empty(n_mx,dtype=object); mx_nom_ga=np.empty(n_mx,dtype=object)
            mx_pays_ga=np.empty(n_mx,dtype=object); mx_nom_le=np.empty(n_mx,dtype=object); mx_pays_le=np.empty(n_mx,dtype=object)
            mx_source=np.empty(n_mx,dtype=object); mx_rp_res=np.empty(n_mx,dtype=object); mx_rc_res=np.empty(n_mx,dtype=object)
            mx_bpe=np.zeros(n_mx,dtype=bool); mx_seg_raw=np.empty(n_mx,dtype=object); mx_seg_src=np.empty(n_mx,dtype=object)
            arr_mx_rp=df_mx['_RP'].values; arr_mx_rc=df_mx['_RC'].values; arr_mx_rs=df_mx['_RS'].values
            arr_mx_iban_tr=df_mx['_IBAN_TR'].values; arr_mx_iban_full=df_mx['_IBAN_FULL'].values
            for i in range(n_mx):
                if i % 5000 == 0: self._prog(0.43 + 0.08*i/max(n_mx,1), f"MX {i:,}/{n_mx:,}")
                res = resolve_row(arr_mx_rp[i], arr_mx_rc[i], arr_mx_rs[i], arr_mx_iban_tr[i], arr_mx_iban_full[i])
                mx_rmpm[i],mx_ga[i],mx_nom_ga[i],mx_pays_ga[i]=res[0],res[1],res[2],res[3]
                mx_nom_le[i],mx_pays_le[i]=res[4],res[5]; mx_source[i],mx_rp_res[i],mx_rc_res[i]=res[6],res[7],res[8]
                mx_seg_raw[i]=res[9]; mx_seg_src[i]=res[10]
                mx_bpe[i]=flag_bpe(arr_mx_iban_tr[i], arr_mx_rs[i], res[9])
            df_mx['_RMPM_R']=mx_rmpm; df_mx['_GA_R']=mx_ga; df_mx['_NOM_GA']=mx_nom_ga; df_mx['_PAYS_GA']=mx_pays_ga
            df_mx['_NOM_LE']=mx_nom_le; df_mx['_PAYS_LE']=mx_pays_le; df_mx['_SOURCE']=mx_source
            df_mx['_RP_R']=mx_rp_res; df_mx['_RC_R']=mx_rc_res; df_mx['_IS_BPE']=mx_bpe
            df_mx['_SEG_RAW']=mx_seg_raw; df_mx['_SEG_SRC']=mx_seg_src

            # ── Q8YY0 : OVERRIDE MX POST-MATCHING ────────────────────────────
            # Appliqué APRÈS resolve_row pour avoir _RMPM_R disponible.
            # Teste : RMPM résolu vs ovr_rmpm_set, RC résolu vs ovr_rmpm_set,
            #         RP_SAFE vs ovr_rmpm_set (fallback).
            if override_pays and ovr_rmpm_set:
                mx_rmpm_vals = df_mx['_RMPM_R'].values
                mx_rc_vals = df_mx['_RC_R'].values
                mx_rp_vals = df_mx['_RP_SAFE'].values
                n_ovr_mx = 0
                for i in range(len(df_mx)):
                    rmpm_v = str(mx_rmpm_vals[i]).strip()
                    rc_v = str(mx_rc_vals[i]).strip()
                    rp_v = str(mx_rp_vals[i]).strip()
                    if (rmpm_v and rmpm_v in ovr_rmpm_set) or (rc_v and rc_v in ovr_rmpm_set) or (rp_v and rp_v in ovr_rmpm_set):
                        df_mx.iat[i, df_mx.columns.get_loc('_PAYS')] = override_pays; n_ovr_mx += 1
                self._prog(0.515, f"Override MX post-matching: {n_ovr_mx:,} lignes → '{override_pays}'")

            # ── Q8YY0 DIAGNOSTIC : lignes MX avec RP mais sans GA ────────
            diag_count = 0
            print("\n" + "="*80)
            print(f"[Q8YY0] DIAGNOSTIC — Lignes MONEXT avec RP non vide mais GA vide")
            print("="*80)
            for i in range(len(df_mx)):
                rp_v = str(df_mx['_RP'].values[i]).strip()
                ga_v = str(df_mx['_GA_R'].values[i]).strip()
                if rp_v and not ga_v and diag_count < 20:
                    rmpm_v = str(df_mx['_RMPM_R'].values[i]).strip()
                    src_v = str(df_mx['_SOURCE'].values[i]).strip()
                    iban_tr_v = str(df_mx['_IBAN_TR'].values[i]).strip()
                    iban_full_v = str(df_mx['_IBAN_FULL'].values[i]).strip()
                    seg_src_v = str(df_mx['_SEG_SRC'].values[i]).strip()
                    # Check if RP is in parc
                    in_parc = rp_v in parc_rp_dict
                    parc_vals = parc_rp_dict.get(rp_v, ['?','?','?','?'])
                    # Check IBAN in acc
                    in_acc_full = iban_full_v in acc_iban_dict if iban_full_v else False
                    in_acc_tr = iban_tr_v in acc_iban_dict if iban_tr_v else False
                    # Check IBAN in ref
                    in_ref_tr = iban_tr_v in d_ref_rib if iban_tr_v else False
                    in_ref_full = iban_full_v in d_ref_rib if iban_full_v else False
                    print(f"\n  [{diag_count+1}] RP={rp_v!r} | GA={ga_v!r} | RMPM={rmpm_v!r} | SOURCE={src_v}")
                    print(f"      IBAN_TR={iban_tr_v[:20]!r} | IBAN_FULL={iban_full_v[:20]!r}")
                    print(f"      in_ref_rib(tr)={in_ref_tr} | in_ref_rib(full)={in_ref_full}")
                    print(f"      in_acc(full)={in_acc_full} | in_acc(tr)={in_acc_tr}")
                    print(f"      in_parc_rp={in_parc} | parc_vals={[str(x)[:20] for x in parc_vals]}")
                    print(f"      SEG_SRC={seg_src_v}")
                    diag_count += 1
            if diag_count == 0:
                print("  → AUCUNE ligne MX avec RP non vide et GA vide. Tout est résolu.")
            else:
                print(f"\n  → {diag_count} lignes diagnostiquées (max 20 affichées)")
            print("="*80 + "\n")

            # ── AGRÉGATS + CLIENT_KEY + DATA (identique 2U6U3) ───────────────
            self._prog(0.52, "Agrégats mensuels...")
            mois_wl = sorted([x for x in df_wl['_MOIS'].unique() if x and len(x)==6])
            mois_mx = sorted([x for x in df_mx['_MOIS'].unique() if x and len(x)==6])
            all_mois = sorted(set(mois_wl) | set(mois_mx))
            plage_debut = plage_cfg.get('debut',''); plage_fin = plage_cfg.get('fin','')
            mois_plage = [mo for mo in all_mois if plage_debut <= mo <= plage_fin] if plage_debut and plage_fin else list(all_mois)

            mx_global_by_mois = {}
            for mo in mois_mx:
                d = df_mx[df_mx['_MOIS']==mo]; mx_global_by_mois[mo] = {'flux':d['_FLUX'].sum(),'pnb':d['_PNB'].sum(),'nb':d['_NB_CARTES'].sum()}
            wl_global_by_mois = {}
            for mo in mois_wl:
                d = df_wl[df_wl['_MOIS']==mo]; wl_global_by_mois[mo] = {'flux':d['_FLUX'].sum(),'pnb':d['_PNB'].sum(),'nb':d['_NB_CARTES'].sum()}

            self._prog(0.56, "Construction clients...")
            def bck(rmpm, rc, rp, rs, tag, idx):
                if rmpm: return f"RMPM|{rmpm}"
                if rc: return f"RC|{rc}"
                if rp: return f"RP|{rp}"
                if rs: return f"RS|{rs}"
                return f"ORPH|{tag}|{idx}"
            df_mx['_CLIENT_KEY'] = [bck(rm,rc,rp,rs,'MX',i) for i,(rm,rc,rp,rs) in enumerate(zip(df_mx['_RMPM_R'].values,df_mx['_RC_R'].values,df_mx['_RP_R'].values,df_mx['_RS'].values))]
            df_wl['_CLIENT_KEY'] = [bck(rm,rc,rp,rs,'WL',i) for i,(rm,rc,rp,rs) in enumerate(zip(df_wl['_RMPM_R'].values,df_wl['_RC_R'].values,df_wl['_RP_R'].values,df_wl['_RS'].values))]

            self._prog(0.60, "Agrégats par client...")
            mx_mo_agg = {mo: {} for mo in all_mois}
            for mo in mois_mx:
                d = df_mx[df_mx['_MOIS']==mo]
                mx_mo_agg[mo] = d.groupby('_CLIENT_KEY').agg({'_FLUX':'sum','_PNB':'sum','_NB_CARTES':'sum','_NB_TRANS':'sum'}).T.to_dict('list')
            wl_mo_agg = {mo: {} for mo in all_mois}
            for mo in mois_wl:
                d = df_wl[df_wl['_MOIS']==mo]
                wl_mo_agg[mo] = d.groupby('_CLIENT_KEY').agg({'_FLUX':'sum','_PNB':'sum','_NB_CARTES':'sum','_NB_TRANS':'sum'}).T.to_dict('list')

            def fne(series):
                for v in series:
                    if v and str(v).strip(): return v
                return ''

            mx_info = {}
            for key, grp in df_mx.groupby('_CLIENT_KEY'):
                mx_info[key] = {'rmpm':fne(grp['_RMPM_R'].values),'ga':fne(grp['_GA_R'].values),'nom_ga':fne(grp['_NOM_GA'].values),
                    'pays_ga':fne(grp['_PAYS_GA'].values),'nom_le':fne(grp['_NOM_LE'].values),'pays_le':fne(grp['_PAYS_LE'].values),
                    'source':fne(grp['_SOURCE'].values),'rp':fne(grp['_RP_SAFE'].values),'rc':fne(grp['_RC_SAFE'].values),
                    'rs':fne(grp['_RS'].values),'nom_prog':fne(grp['_NOM_PROG'].values),'pays':fne(grp['_PAYS'].values),
                    'is_bpe':bool(grp['_IS_BPE'].any()),'differe':fne(grp['_DIFFERE'].values),
                    'seg_raw':fne(grp['_SEG_RAW'].values),'seg_src':fne(grp['_SEG_SRC'].values)}
            wl_info = {}
            for key, grp in df_wl.groupby('_CLIENT_KEY'):
                wl_info[key] = {'rmpm':fne(grp['_RMPM_R'].values),'ga':fne(grp['_GA_R'].values),'nom_ga':fne(grp['_NOM_GA'].values),
                    'pays_ga':fne(grp['_PAYS_GA'].values),'nom_le':fne(grp['_NOM_LE'].values),'pays_le':fne(grp['_PAYS_LE'].values),
                    'source':fne(grp['_SOURCE'].values),'rc':fne(grp['_RC_SAFE'].values),'rs':fne(grp['_RS'].values),
                    'nom_prog':fne(grp['_NOM_PROG'].values),'pays':fne(grp['_PAYS'].values),
                    'is_bpe':bool(grp['_IS_BPE'].any()),'differe':fne(grp['_DIFFERE'].values),
                    'plafond':float(grp['_PLAFOND'].max()) if len(grp)>0 else 0.0,
                    'perio':float(grp['_PERIO'].iloc[0]) if len(grp)>0 else 1.0,
                    'seg_raw':fne(grp['_SEG_RAW'].values),'seg_src':fne(grp['_SEG_SRC'].values)}

            all_client_keys = sorted(set(mx_info.keys()) | set(wl_info.keys()))
            n_clients = len(all_client_keys)

            self._prog(0.72, "Construction DATA...")
            rows_data = []
            for i, key in enumerate(all_client_keys):
                if i % 2000 == 0: self._prog(0.72 + 0.06*i/max(n_clients,1), f"DATA {i:,}/{n_clients:,}")
                mx = mx_info.get(key, {}); wl = wl_info.get(key, {})
                info_main = mx if mx.get('source','').startswith('ACCOUNT') or mx.get('source','').startswith('REF_') else (wl if wl.get('source','').startswith('ACCOUNT') or wl.get('source','').startswith('REF_') else (mx or wl))
                seg_raw = info_main.get('seg_raw', '') or mx.get('seg_raw', '') or wl.get('seg_raw', '')
                seg_src = info_main.get('seg_src', '') or mx.get('seg_src', '') or wl.get('seg_src', 'FALLBACK')
                is_bpe = seg_raw == 'BPE' if seg_raw else (mx.get('is_bpe') or wl.get('is_bpe'))

                row = {
                    'CLIENT_KEY':key, 'RMPM':info_main.get('rmpm',''),
                    'ID_RP':self.protect_id(mx.get('rp','')), 'ID_RC_MX':self.protect_id(mx.get('rc','')), 'ID_RC_WL':self.protect_id(wl.get('rc','')),
                    'CODE_GA':info_main.get('ga',''), 'NOM_GA':info_main.get('nom_ga',''),
                    'PAYS_GA':info_main.get('pays_ga',''), 'NOM_ENTITE':info_main.get('nom_le',''), 'PAYS_ENTITE':info_main.get('pays_le',''),
                    'NOM_PROG_CCO':mx.get('nom_prog',''), 'NOM_PROG_CPC':wl.get('nom_prog',''),
                    'RS_CCO':mx.get('rs',''), 'RS_CPC':wl.get('rs',''),
                    'SOURCE_MATCHING':info_main.get('source','NON_TROUVE'),
                    'PRESENT_CCO':'YES' if key in mx_info else 'NO', 'PRESENT_CPC':'YES' if key in wl_info else 'NO',
                    'IS_BPE':'YES' if is_bpe else 'NO', 'IS_ENT':'NO' if is_bpe else 'YES',
                    'SEGMENT':'BPE' if is_bpe else 'ENTREPRISE',
                    'SEGMENT_RAW': seg_raw, 'SEGMENT_SOURCE': seg_src,
                    'PAYS_APPORTEUR_CCO':mx.get('pays',''), 'PAYS_APPORTEUR_CPC':wl.get('pays',''),
                }
                rmpm_client = row['RMPM']; rp_client = mx.get('rp',''); rc_client = mx.get('rc','') or wl.get('rc','')
                # ── Q8YY0 : détection OVERRIDE au niveau client ──────────
                # Si RMPM, RC, RP ou NOM_PROG est dans les sets override,
                # le pays est écrasé. Priorité absolue sur CIB et plateforme.
                client_overridden = False
                if override_pays and (ovr_rmpm_set or ovr_idprog_set):
                    for id_v in [rmpm_client, rc_client, rp_client]:
                        if ovr_rmpm_set and id_v and id_v in ovr_rmpm_set:
                            client_overridden = True; break
                    if not client_overridden and ovr_idprog_set:
                        for prog in [mx.get('nom_prog',''), wl.get('nom_prog','')]:
                            if prog and prog in ovr_idprog_set:
                                client_overridden = True; break
                if client_overridden:
                    # Override priorité absolue : PAYS_APPORTEUR + PAYS_FINAL
                    row['PAYS_APPORTEUR_CCO'] = override_pays if mx else row['PAYS_APPORTEUR_CCO']
                    row['PAYS_APPORTEUR_CPC'] = override_pays if wl else row['PAYS_APPORTEUR_CPC']
                    # MONITORING résolution normale (pour info) mais PAYS_FINAL = override
                    pays_cib = ''
                    flag_cib = 'NO'
                    source_pays = ''
                    if rmpm_client and rmpm_client in mon_rmpm_to_pays:
                        pays_cib = mon_rmpm_to_pays[rmpm_client]; flag_cib = 'YES'
                    elif rmpm_client and rmpm_client in mon_rmpm_set:
                        flag_cib = 'YES'
                    if not pays_cib:
                        for rc_v in [rc_client, rc_client.lstrip('0') if rc_client else '']:
                            if rc_v and rc_v in mon_rc_to_pays:
                                pays_cib = mon_rc_to_pays[rc_v]; flag_cib = 'YES'; break
                            elif rc_v and rc_v in mon_rc_set and flag_cib != 'YES':
                                flag_cib = 'YES'
                    if not pays_cib:
                        for rp_v in [rp_client, rp_client.lstrip('0') if rp_client else '']:
                            if rp_v and rp_v in mon_rp_to_pays:
                                pays_cib = mon_rp_to_pays[rp_v]; flag_cib = 'YES'; break
                            elif rp_v and rp_v in mon_rp_set and flag_cib != 'YES':
                                flag_cib = 'YES'
                    pays_final = override_pays
                    source_pays = 'OVERRIDE'
                else:
                    # ── Résolution PAYS_FINAL standard (MONITORING > plateforme) ──
                    pays_cib = ''
                    flag_cib = 'NO'
                    source_pays = ''
                    if rmpm_client and rmpm_client in mon_rmpm_to_pays:
                        pays_cib = mon_rmpm_to_pays[rmpm_client]; flag_cib = 'YES'; source_pays = 'CIB_RMPM'
                    elif rmpm_client and rmpm_client in mon_rmpm_set:
                        flag_cib = 'YES'; source_pays = 'CIB_RMPM_NO_PAYS'
                    if not pays_cib:
                        for rc_v in [rc_client, rc_client.lstrip('0') if rc_client else '']:
                            if rc_v and rc_v in mon_rc_to_pays:
                                pays_cib = mon_rc_to_pays[rc_v]; flag_cib = 'YES'; source_pays = 'CIB_RC'; break
                            elif rc_v and rc_v in mon_rc_set and flag_cib != 'YES':
                                flag_cib = 'YES'; source_pays = 'CIB_RC_NO_PAYS'
                    if not pays_cib:
                        for rp_v in [rp_client, rp_client.lstrip('0') if rp_client else '']:
                            if rp_v and rp_v in mon_rp_to_pays:
                                pays_cib = mon_rp_to_pays[rp_v]; flag_cib = 'YES'; source_pays = 'CIB_RP'; break
                            elif rp_v and rp_v in mon_rp_set and flag_cib != 'YES':
                                flag_cib = 'YES'; source_pays = 'CIB_RP_NO_PAYS'
                    pays_cco = mx.get('pays',''); pays_cpc = wl.get('pays','')
                    if pays_cib:
                        pays_final = pays_cib
                        if not source_pays: source_pays = 'CIB'
                    else:
                        pays_final = pays_cco or pays_cpc
                        source_pays = 'PLATEFORME_CCO' if pays_cco else ('PLATEFORME_CPC' if pays_cpc else 'NON_TROUVE')
                row.update({'PAYS_CIB': pays_cib, 'PAYS_FINAL': pays_final, 'FLAG_CIB': flag_cib, 'SOURCE_PAYS': source_pays})
                row.update({
                    'DIFFERE_CCO_TYPE':self._differe_type(mx.get('differe',''),'CCO'), 'DIFFERE_CCO_JOURS':self._differe_jours(mx.get('differe','')),
                    'DIFFERE_CPC_TYPE':self._differe_type(wl.get('differe',''),'CPC'), 'DIFFERE_CPC_JOURS':self._differe_jours(wl.get('differe','')),
                    'PLAFOND_CPC_EUR':wl.get('plafond',0.0), 'PERIODICITE_CPC':wl.get('perio',1.0),
                })
                mx_flux_p = sum(mx_mo_agg[mo].get(key,[0,0,0,0])[0] for mo in mois_plage if mo in mx_mo_agg)
                mx_pnb_p = sum(mx_mo_agg[mo].get(key,[0,0,0,0])[1] for mo in mois_plage if mo in mx_mo_agg)
                mx_nb_p = sum(mx_mo_agg[mo].get(key,[0,0,0,0])[2] for mo in mois_plage if mo in mx_mo_agg)
                mx_tr_p = sum(mx_mo_agg[mo].get(key,[0,0,0,0])[3] for mo in mois_plage if mo in mx_mo_agg)
                wl_flux_p = sum(wl_mo_agg[mo].get(key,[0,0,0,0])[0] for mo in mois_plage if mo in wl_mo_agg)
                wl_pnb_p = sum(wl_mo_agg[mo].get(key,[0,0,0,0])[1] for mo in mois_plage if mo in wl_mo_agg)
                wl_nb_p = sum(wl_mo_agg[mo].get(key,[0,0,0,0])[2] for mo in mois_plage if mo in wl_mo_agg)
                wl_tr_p = sum(wl_mo_agg[mo].get(key,[0,0,0,0])[3] for mo in mois_plage if mo in wl_mo_agg)
                row.update({'FLUX_CCO_PLAGE':mx_flux_p,'PNB_CCO_PLAGE':mx_pnb_p,'NB_CARTES_CCO_PLAGE':mx_nb_p,'NB_TRANS_CCO_PLAGE':mx_tr_p,
                    'FLUX_CPC_PLAGE':wl_flux_p,'PNB_CPC_PLAGE':wl_pnb_p,'NB_CARTES_CPC_PLAGE':wl_nb_p,'NB_TRANS_CPC_PLAGE':wl_tr_p})
                for mo in all_mois:
                    lbl = self.mois_label(mo)
                    row[f'FLUX_CCO_{lbl}'] = mx_mo_agg[mo].get(key,[0,0,0,0])[0]
                    row[f'FLUX_CPC_{lbl}'] = wl_mo_agg[mo].get(key,[0,0,0,0])[0]
                last_3 = mois_plage[-3:] if len(mois_plage)>=3 else mois_plage
                row['FLUX_CCO_TRIMESTRE'] = sum(mx_mo_agg[mo].get(key,[0,0,0,0])[0] for mo in last_3 if mo in mx_mo_agg)
                row['FLUX_CPC_TRIMESTRE'] = sum(wl_mo_agg[mo].get(key,[0,0,0,0])[0] for mo in last_3 if mo in wl_mo_agg)
                row['TRIMESTRE_LIBELLE'] = f"{self.mois_label(last_3[0])} -> {self.mois_label(last_3[-1])}" if last_3 else ""
                rows_data.append(row)
            df_data = pd.DataFrame(rows_data)

            # ── EXPORTS ──────────────────────────────────────────────────────
            self._prog(0.82, "CSV...")
            self.output_dir.mkdir(parents=True, exist_ok=True)
            save_csv = str(self.output_dir / f"{self.output_filename}.csv")
            df_data.to_csv(save_csv, sep=';', index=False, encoding='utf-8-sig')
            print(f"[OK] CSV : {save_csv}")
            if not OPENPYXL_OK: self._prog(1.0, "Terminé (CSV)."); return

            nb_ref_rc = int(np.sum(np.array([s.startswith('REF_RC') for s in np.concatenate([wl_seg_src, mx_seg_src])])))
            nb_ref_rib = int(np.sum(np.array([s=='REF_IBAN' for s in np.concatenate([wl_seg_src, mx_seg_src])])))
            nb_idseg = int(np.sum(np.array([s.startswith('IDSEG') for s in np.concatenate([wl_seg_src, mx_seg_src])])))
            nb_fb = int(np.sum(np.array([s=='FALLBACK' for s in np.concatenate([wl_seg_src, mx_seg_src])])))

            self.generate_xlsx_mn8k3_compat(df_data, ts, all_mois, mois_mx, mois_wl, mois_plage,
                mx_global_by_mois, wl_global_by_mois, pays_excl_wl, pays_excl_mx, n_clients, constants,
                mx_mo_agg, wl_mo_agg, df_wl, df_mx)

            self._prog(1.0, "Terminé !")
            print(f"[OK] Analyse terminee [{VERSION_ID}] : {n_clients:,} clients "
                  f"(REF_RC={nb_ref_rc:,} REF_IBAN={nb_ref_rib:,} IDSEG={nb_idseg:,} FALLBACK={nb_fb:,})")

        except Exception:
            self._prog(0, "Erreur"); raise

    # ══════════════════════════════════════════════════════════════════════════
    # XLSX — structure identique J5COM (DATA + ANALYSE + FINANCIAL)
    # ══════════════════════════════════════════════════════════════════════════
    def generate_xlsx_mn8k3_compat(self, df_data, ts, all_mois, mois_mx, mois_wl, mois_plage,
                                    mx_global_by_mois, wl_global_by_mois,
                                    pays_excl_wl, pays_excl_mx, n_clients, constants,
                                    mx_mo_agg, wl_mo_agg, df_wl, df_mx):

        self.output_dir.mkdir(parents=True, exist_ok=True)
        save_xlsx = str(self.output_dir / f"{self.output_filename}.xlsx")

        self._prog(0.85, "XLSX...")
        try:
            def fill(h): return PatternFill(start_color=h, end_color=h, fill_type="solid")
            def fnt(c=WHT, bold=True, sz=10): return Font(name="Segoe UI", size=sz, bold=bold, color=c)
            thin = Side(style="thin", color="CCCCCC"); brd = Border(left=thin, right=thin, top=thin, bottom=thin)
            wb = Workbook()

            # SHEET DATA
            ws_data = wb.active; ws_data.title = "DATA"; ws_data.column_dimensions["A"].width = 2.5
            plage_lbl = f"{mois_plage[0] if mois_plage else '?'} -> {mois_plage[-1] if mois_plage else '?'}"
            ws_data["B2"] = f"CIB Commissionnement — DATA — {datetime.now().strftime('%d/%m/%Y')} — [{VERSION_ID}] — Plage : {plage_lbl}"
            ws_data["B2"].font = fnt(DARK, sz=13); ws_data["B2"].fill = fill(GRN2)
            HDR_ROW = 5
            NUMERIC_PREFIXES = ('FLUX_','PNB_','NB_CARTES_','NB_TRANS_','PLAFOND_','PERIODICITE_','DIFFERE_CCO_JOURS','DIFFERE_CPC_JOURS')
            def is_num(cn): return any(cn.startswith(p) for p in NUMERIC_PREFIXES)
            for ci, col in enumerate(df_data.columns, start=2):
                cell = ws_data.cell(row=HDR_ROW, column=ci, value=col); cell.font = fnt(); cell.fill = fill(GRN); cell.alignment = Alignment(horizontal="center", wrap_text=True); cell.border = brd
            num_idx = {ci for ci, col in enumerate(df_data.columns) if is_num(col)}
            for ri, row_data in enumerate(df_data.itertuples(index=False), start=HDR_ROW + 1):
                bg = GRN2 if ri % 2 == 0 else WHT
                for ci_0, val in enumerate(row_data):
                    cell = ws_data.cell(row=ri, column=ci_0 + 2); cell.fill = fill(bg); cell.border = brd; cell.font = Font(name="Segoe UI", size=9)
                    if ci_0 in num_idx and val not in ('', None):
                        try: cell.value = float(val); cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal="right")
                        except: cell.value = str(val) if val not in (None,'nan') else ''
                    else: cell.value = str(val) if val not in (None,'nan') else ''
            for ci, col in enumerate(df_data.columns, start=2):
                ws_data.column_dimensions[get_column_letter(ci)].width = 16
            ws_data.freeze_panes = f"B{HDR_ROW + 1}"

            # SHEETS ANALYSE par année
            self._prog(0.89, "ANALYSE par année...")
            global_by_mois = {}
            for mo in all_mois:
                mx_d = mx_global_by_mois.get(mo,{'flux':0,'pnb':0,'nb':0}); wl_d = wl_global_by_mois.get(mo,{'flux':0,'pnb':0,'nb':0})
                global_by_mois[mo] = {'flux':mx_d['flux']+wl_d['flux'],'pnb':mx_d['pnb']+wl_d['pnb'],'nb':mx_d['nb']+wl_d['nb']}
            years = sorted(set(mo[:4] for mo in all_mois))
            for year in years:
                ws_ana = wb.create_sheet(title=f"ANALYSE_{year}"); ws_ana.column_dimensions["A"].width = 2.5
                ws_ana["B2"] = f"CIB — ANALYSE {year} — {datetime.now().strftime('%d/%m/%Y')}"; ws_ana["B2"].font = fnt(DARK, sz=13); ws_ana["B2"].fill = fill(GRN2)
                BLOC_COLS = 4; col_g = 2; col_m = col_g + BLOC_COLS + COL_GAP; col_w = col_m + BLOC_COLS + COL_GAP
                BLOCS = [(col_g,"GLOBAL",GRN,GRN2),(col_m,"MONEXT",BLU,BLU2),(col_w,"WORLDLINE",PUR,PUR2)]
                START_HDR = 5
                for col_s, label, hdr_col, bg_col in BLOCS:
                    c = ws_ana.cell(row=START_HDR, column=col_s, value=label); c.font = fnt(sz=13); c.fill = fill(hdr_col)
                    for dc in range(1, BLOC_COLS): ws_ana.cell(row=START_HDR, column=col_s+dc).fill = fill(hdr_col)
                SUB = ["MOIS","NB CARTES","FLUX","PNB"]
                for col_s, _, hdr_col, _ in BLOCS:
                    for dc, sub in enumerate(SUB):
                        c = ws_ana.cell(row=START_HDR+1, column=col_s+dc, value=sub); c.font = fnt(sz=9); c.fill = fill(hdr_col); c.border = brd
                DATA_START = START_HDR + 2
                yr_all = [mo for mo in all_mois if mo.startswith(year)]; yr_mx = [mo for mo in mois_mx if mo.startswith(year)]; yr_wl = [mo for mo in mois_wl if mo.startswith(year)]
                sources_data = [(col_g,yr_all,global_by_mois,GRN2),(col_m,yr_mx,mx_global_by_mois,BLU2),(col_w,yr_wl,wl_global_by_mois,PUR2)]
                for col_s, ml, dd, bg_col in sources_data:
                    for ri_o, mo in enumerate(ml):
                        ri = DATA_START + ri_o; d = dd.get(mo,{'flux':0,'pnb':0,'nb':0}); bg = bg_col if ri_o%2==0 else WHT
                        c = ws_ana.cell(row=ri, column=col_s, value=self.mois_label(mo)); c.font = Font(name="Segoe UI", size=9, bold=True); c.fill = fill(bg); c.border = brd
                        for dc, kf in enumerate(['nb','flux','pnb'], start=1):
                            c = ws_ana.cell(row=ri, column=col_s+dc, value=round(d[kf],2)); c.font = Font(name="Segoe UI", size=9); c.fill = fill(bg); c.border = brd; c.number_format = '#,##0.00'
                    ri_tot = DATA_START + len(ml)
                    c = ws_ana.cell(row=ri_tot, column=col_s, value="TOTAL"); c.font = fnt(sz=9); c.fill = fill(DARK); c.border = brd
                    for dc, kf in enumerate(['nb','flux','pnb'], start=1):
                        tot = sum(dd.get(mo,{}).get(kf,0) for mo in ml)
                        c = ws_ana.cell(row=ri_tot, column=col_s+dc, value=round(tot,2)); c.font = fnt(sz=9); c.fill = fill(DARK); c.border = brd; c.number_format = '#,##0.00'
                for col_s,_,_,_ in BLOCS:
                    ws_ana.column_dimensions[get_column_letter(col_s)].width = 18
                    for dc in range(1, BLOC_COLS): ws_ana.column_dimensions[get_column_letter(col_s+dc)].width = 16

            # SHEETS FINANCIAL par année
            for year in years:
                self._prog(0.93, f"FINANCIAL_{year}...")
                yr_plage = [mo for mo in mois_plage if mo.startswith(year)]
                if not yr_plage: continue
                yr_flux_cols_cco = [c for c in df_data.columns if c.startswith(f'FLUX_CCO_{year}')]
                yr_flux_cols_cpc = [c for c in df_data.columns if c.startswith(f'FLUX_CPC_{year}')]
                if yr_flux_cols_cco or yr_flux_cols_cpc:
                    mask = pd.Series(False, index=df_data.index)
                    for c in yr_flux_cols_cco + yr_flux_cols_cpc:
                        mask = mask | (df_data[c].astype(float).abs() > 0)
                    df_year = df_data[mask].copy()
                else:
                    df_year = df_data.copy()

                # ── Q8YY0 : RECALCUL COMPLET PAR ANNÉE ──────────────────────
                # Filtrer les lignes brutes WL/MX pour cette année
                wl_yr = df_wl[df_wl['_MOIS'].str.startswith(year)].copy() if len(df_wl) > 0 else df_wl
                mx_yr = df_mx[df_mx['_MOIS'].str.startswith(year)].copy() if len(df_mx) > 0 else df_mx

                # Pré-calculer plafond/perio/differe par CLIENT_KEY pour l'année
                # WL : plafond max + periodicite première valeur + differe
                wl_yr_info = {}
                if len(wl_yr) > 0 and '_CLIENT_KEY' in wl_yr.columns:
                    for key_wl, grp_wl in wl_yr.groupby('_CLIENT_KEY'):
                        wl_yr_info[key_wl] = {
                            'plafond': float(grp_wl['_PLAFOND'].max()) if len(grp_wl) > 0 else 0.0,
                            'perio': float(grp_wl['_PERIO'].iloc[0]) if len(grp_wl) > 0 else 1.0,
                            'differe': str(grp_wl['_DIFFERE'].iloc[0]).strip() if len(grp_wl) > 0 else '',
                        }
                # MX : differe
                mx_yr_info = {}
                if len(mx_yr) > 0 and '_CLIENT_KEY' in mx_yr.columns:
                    for key_mx, grp_mx in mx_yr.groupby('_CLIENT_KEY'):
                        mx_yr_info[key_mx] = {
                            'differe': str(grp_mx['_DIFFERE'].iloc[0]).strip() if len(grp_mx) > 0 else '',
                        }

                yr_last3 = yr_plage[-3:] if len(yr_plage) >= 3 else yr_plage
                for idx in df_year.index:
                    key = df_year.at[idx, 'CLIENT_KEY']
                    # CCO agrégats
                    df_year.at[idx, 'FLUX_CCO_PLAGE'] = sum(mx_mo_agg[mo].get(key,[0,0,0,0])[0] for mo in yr_plage if mo in mx_mo_agg)
                    df_year.at[idx, 'PNB_CCO_PLAGE'] = sum(mx_mo_agg[mo].get(key,[0,0,0,0])[1] for mo in yr_plage if mo in mx_mo_agg)
                    df_year.at[idx, 'NB_CARTES_CCO_PLAGE'] = sum(mx_mo_agg[mo].get(key,[0,0,0,0])[2] for mo in yr_plage if mo in mx_mo_agg)
                    df_year.at[idx, 'NB_TRANS_CCO_PLAGE'] = sum(mx_mo_agg[mo].get(key,[0,0,0,0])[3] for mo in yr_plage if mo in mx_mo_agg)
                    df_year.at[idx, 'FLUX_CCO_TRIMESTRE'] = sum(mx_mo_agg[mo].get(key,[0,0,0,0])[0] for mo in yr_last3 if mo in mx_mo_agg)
                    # CPC agrégats
                    df_year.at[idx, 'FLUX_CPC_PLAGE'] = sum(wl_mo_agg[mo].get(key,[0,0,0,0])[0] for mo in yr_plage if mo in wl_mo_agg)
                    df_year.at[idx, 'PNB_CPC_PLAGE'] = sum(wl_mo_agg[mo].get(key,[0,0,0,0])[1] for mo in yr_plage if mo in wl_mo_agg)
                    df_year.at[idx, 'NB_CARTES_CPC_PLAGE'] = sum(wl_mo_agg[mo].get(key,[0,0,0,0])[2] for mo in yr_plage if mo in wl_mo_agg)
                    df_year.at[idx, 'NB_TRANS_CPC_PLAGE'] = sum(wl_mo_agg[mo].get(key,[0,0,0,0])[3] for mo in yr_plage if mo in wl_mo_agg)
                    df_year.at[idx, 'FLUX_CPC_TRIMESTRE'] = sum(wl_mo_agg[mo].get(key,[0,0,0,0])[0] for mo in yr_last3 if mo in wl_mo_agg)

                    # ── Q8YY0 FIX #2 : PLAFOND + PERIODICITE recalculés par année ──
                    if key in wl_yr_info:
                        df_year.at[idx, 'PLAFOND_CPC_EUR'] = wl_yr_info[key]['plafond']
                        df_year.at[idx, 'PERIODICITE_CPC'] = wl_yr_info[key]['perio']
                    else:
                        # Client pas présent côté CPC cette année → plafond/perio à 0/1
                        df_year.at[idx, 'PLAFOND_CPC_EUR'] = 0.0
                        df_year.at[idx, 'PERIODICITE_CPC'] = 1.0

                    # ── Q8YY0 FIX #3 : DIFFERE recalculé par année ──
                    if key in mx_yr_info:
                        diff_mx = mx_yr_info[key]['differe']
                        df_year.at[idx, 'DIFFERE_CCO_TYPE'] = self._differe_type(diff_mx, 'CCO')
                        df_year.at[idx, 'DIFFERE_CCO_JOURS'] = self._differe_jours(diff_mx)
                    else:
                        df_year.at[idx, 'DIFFERE_CCO_TYPE'] = ''
                        df_year.at[idx, 'DIFFERE_CCO_JOURS'] = 0

                    if key in wl_yr_info:
                        diff_wl = wl_yr_info[key]['differe']
                        df_year.at[idx, 'DIFFERE_CPC_TYPE'] = self._differe_type(diff_wl, 'CPC')
                        df_year.at[idx, 'DIFFERE_CPC_JOURS'] = self._differe_jours(diff_wl)
                    else:
                        df_year.at[idx, 'DIFFERE_CPC_TYPE'] = ''
                        df_year.at[idx, 'DIFFERE_CPC_JOURS'] = 0

                self._build_financial_sheet(wb, df_year, constants, brd, yr_plage, year_suffix=year)

            self._prog(0.98, "Sauvegarde...")
            wb.save(save_xlsx)
            print(f"[OK] XLSX : {save_xlsx} "
                  f"(DATA={n_clients:,} clients, ANALYSE={len(all_mois)} mois, "
                  f"FINANCIAL formules vivantes [{VERSION_ID}])")
        except Exception:
            raise

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET FINANCIAL — identique J5COM
    # ══════════════════════════════════════════════════════════════════════════
    def _build_financial_sheet(self, wb, df_data, constants, brd, mois_plage, year_suffix=''):
        sheet_name = f"FINANCIAL_{year_suffix}" if year_suffix else "FINANCIAL"
        ws = wb.create_sheet(title=sheet_name); ws.column_dimensions["A"].width = 2.5
        def fill(h): return PatternFill(start_color=h, end_color=h, fill_type="solid")
        def fnt(c=WHT, bold=True, sz=10): return Font(name="Segoe UI", size=sz, bold=bold, color=c)
        plage_lbl = f"{mois_plage[0] if mois_plage else '?'} -> {mois_plage[-1] if mois_plage else '?'}"
        ws["B2"] = f"CIB Commissionnement — FINANCIAL — {datetime.now().strftime('%d/%m/%Y')} — [{VERSION_ID}] — Plage : {plage_lbl}"
        ws["B2"].font = fnt(DARK, sz=13); ws["B2"].fill = fill(GRN2)
        ws["B3"] = "Cellules EDITABLES en orange. Tout le reste se recalcule automatiquement."
        ws["B3"].font = fnt(DARK, bold=False, sz=9); ws["B3"].fill = fill(GRN2)

        CONST_ROW_LBL = 5; CONST_ROW_VAL = 6
        const_specs = [('TAUX_REFIN',constants['taux_refinancement']),('PART_CAP',constants['part_capital_banque']),('ALMT',constants['almt']),('RW_DEFAUT',constants['rw_defaut']),
            ('COUT_CARTE',constants['cout_carte_cco']),('COUT_TRANS',constants['cout_transaction_cpc']),('COUT_RWA_CCO',constants['cout_rwa_plafond_cco']),
            ('TAUX_PORTEUR',constants['taux_ead_porteur']),('TAUX_ENT',constants['taux_ead_entreprise']),('TAUX_COMM',constants['taux_commission']),('TAUX_TVA',constants['taux_tva'])]
        EUR_CONST = ('COUT_CARTE','COUT_TRANS','COUT_RWA_CCO')
        const_cols = {}
        for i, (lbl, val) in enumerate(const_specs):
            ci = 2 + i; cl = get_column_letter(ci); const_cols[lbl] = f"${cl}${CONST_ROW_VAL}"
            c = ws.cell(row=CONST_ROW_LBL, column=ci, value=lbl); c.font = fnt(DARK, bold=True, sz=9); c.fill = fill(ORA2); c.alignment = Alignment(horizontal="center")
            c = ws.cell(row=CONST_ROW_VAL, column=ci, value=val); c.font = fnt(DARK, bold=False, sz=10); c.fill = fill(ORA2); c.border = brd
            c.number_format = '0.00' if lbl in EUR_CONST else '0.00%'

        rw_def = constants['rw_defaut']
        COLS = []
        COLS += [('ID_RP','DATA','ID_RP',16,BLU2),('ID_RC','DATA','ID_RC_CONSOL',16,BLU2),('CODE_GA','DATA','CODE_GA',14,BLU2),('NOM_GA','DATA','NOM_GA',25,BLU2),
            ('PAYS_GA','DATA','PAYS_GA',14,BLU2),('RMPM','DATA','RMPM',14,BLU2),('NOM_ENTITE','DATA','NOM_ENTITE',25,BLU2),('PAYS_ENTITE','DATA','PAYS_ENTITE',14,BLU2),
            ('PAYS_APPORTEUR_CCO','DATA','PAYS_APPORTEUR_CCO',16,BLU2),('PAYS_APPORTEUR_CPC','DATA','PAYS_APPORTEUR_CPC',16,BLU2),
            ('PAYS_CIB','DATA','PAYS_CIB',16,BLU2),('PAYS_FINAL','DATA','PAYS_FINAL',16,BLU2),('FLAG_CIB','DATA','FLAG_CIB',10,BLU2)]
        COLS += [('NOM_PROG_CCO','DATA','NOM_PROG_CCO',22,GRN2),('NOM_PROG_CPC','DATA','NOM_PROG_CPC',22,GRN2),('SOURCE_MATCHING','DATA','SOURCE_MATCHING',16,GRN2),
            ('PRESENT_CCO','DATA','PRESENT_CCO',12,GRN2),('PRESENT_CPC','DATA','PRESENT_CPC',12,GRN2),('IS_ENTREPRISE','DATA','IS_ENT',12,GRN2),('IS_BPE','DATA','IS_BPE',10,GRN2),('SEGMENT','DATA','SEGMENT',12,GRN2)]
        COLS += [('TYPE_DEBIT_CCO','SAISIE','Entreprise',16,ORA2),('RW_AJUSTE','SAISIE',rw_def,12,ORA2)]
        COLS += [('DIFFERE_CCO_TYPE','DATA','DIFFERE_CCO_TYPE',14,GRN2),('DIFFERE_CCO_JOURS','DATA','DIFFERE_CCO_JOURS',14,GRN2),
            ('NB_CARTES_CCO_PLAGE','DATA','NB_CARTES_CCO_PLAGE',14,GRN2),('NB_TRANS_CCO_PLAGE','DATA','NB_TRANS_CCO_PLAGE',14,GRN2),
            ('FLUX_CCO_PLAGE','DATA','FLUX_CCO_PLAGE',16,GRN2),('FLUX_CCO_TRIMESTRE','DATA','FLUX_CCO_TRIMESTRE',16,GRN2),('PNB_CCO_PLAGE','DATA','PNB_CCO_PLAGE',16,GRN2),
            ('COUT_DIFFERE_CCO','CALC',lambda r: f'=IF(OR(DIFFERE_CCO_TYPE_{r}="IMM",FLUX_CCO_PLAGE_{r}=0),0,FLUX_CCO_PLAGE_{r}*(DIFFERE_CCO_JOURS_{r}+15)/365*{const_cols["TAUX_REFIN"]})',16,GRN2),
            ('COUT_RUN_CCO','CALC',lambda r: f'=NB_CARTES_CCO_PLAGE_{r}*{const_cols["COUT_CARTE"]}',14,GRN2),
            ('TAUX_EAD_CCO','CALC',lambda r: f'=IF(TYPE_DEBIT_CCO_{r}="Porteur",{const_cols["TAUX_PORTEUR"]},{const_cols["TAUX_ENT"]})',12,GRN2),
            ('RW_RETENU','CALC',lambda r: f'=IF(RW_AJUSTE_{r}="",{const_cols["RW_DEFAUT"]},RW_AJUSTE_{r})',12,GRN2),
            ('RWA_PLAFOND_CCO','CALC',lambda r: f'=IF(DIFFERE_CCO_TYPE_{r}="IMM",0,NB_CARTES_CCO_PLAGE_{r}/12*{const_cols["COUT_RWA_CCO"]})',14,GRN2),
            ('EAD_BILAN_CCO','CALC',lambda r: f'=IF(OR(DIFFERE_CCO_TYPE_{r}="IMM",FLUX_CCO_TRIMESTRE_{r}=0),0,IF(DIFFERE_CCO_JOURS_{r}>27,FLUX_CCO_TRIMESTRE_{r},FLUX_CCO_TRIMESTRE_{r}*2/3))',16,GRN2),
            ('RWA_BILAN_CCO','CALC',lambda r: f'=EAD_BILAN_CCO_{r}*RW_RETENU_{r}*{const_cols["PART_CAP"]}*{const_cols["ALMT"]}',14,GRN2),
            ('TOTAL_COUTS_CCO','CALC',lambda r: f'=COUT_DIFFERE_CCO_{r}+COUT_RUN_CCO_{r}+RWA_PLAFOND_CCO_{r}+RWA_BILAN_CCO_{r}',14,GRN2)]
        COLS += [('DIFFERE_CPC_TYPE','DATA','DIFFERE_CPC_TYPE',14,PUR2),('DIFFERE_CPC_JOURS','DATA','DIFFERE_CPC_JOURS',14,PUR2),
            ('NB_CARTES_CPC_PLAGE','DATA','NB_CARTES_CPC_PLAGE',14,PUR2),('NB_TRANS_CPC_PLAGE','DATA','NB_TRANS_CPC_PLAGE',14,PUR2),
            ('FLUX_CPC_PLAGE','DATA','FLUX_CPC_PLAGE',16,PUR2),('FLUX_CPC_TRIMESTRE','DATA','FLUX_CPC_TRIMESTRE',16,PUR2),('PNB_CPC_PLAGE','DATA','PNB_CPC_PLAGE',16,PUR2),
            ('PLAFOND_CPC_EUR','DATA','PLAFOND_CPC_EUR',14,PUR2),('PERIODICITE_CPC','DATA','PERIODICITE_CPC',12,PUR2),
            ('COUT_DIFFERE_CPC','CALC',lambda r: f'=IF(OR(DIFFERE_CPC_TYPE_{r}="IMM",DIFFERE_CPC_JOURS_{r}=0,FLUX_CPC_PLAGE_{r}=0),0,FLUX_CPC_PLAGE_{r}*(DIFFERE_CPC_JOURS_{r}+15)/365*{const_cols["TAUX_REFIN"]})',16,PUR2),
            ('COUT_RUN_CPC','CALC',lambda r: f'=NB_TRANS_CPC_PLAGE_{r}*{const_cols["COUT_TRANS"]}',14,PUR2),
            ('MULT_PERIODICITE','CALC',lambda r: f'=IF(OR(PERIODICITE_CPC_{r}=1,PERIODICITE_CPC_{r}=12),1,IF(OR(PERIODICITE_CPC_{r}=3,PERIODICITE_CPC_{r}=6),2,1))',12,PUR2),
            ('EAD_PLAFOND_CPC','CALC',lambda r: f'=PLAFOND_CPC_EUR_{r}*MULT_PERIODICITE_{r}',14,PUR2),
            ('RWA_PLAFOND_CPC','CALC',lambda r: f'=EAD_PLAFOND_CPC_{r}*{const_cols["TAUX_ENT"]}*RW_RETENU_{r}*{const_cols["PART_CAP"]}*{const_cols["ALMT"]}',14,PUR2),
            ('EAD_BILAN_CPC','CALC',lambda r: f'=IF(OR(DIFFERE_CPC_TYPE_{r}="IMM",DIFFERE_CPC_JOURS_{r}=0,FLUX_CPC_TRIMESTRE_{r}=0),0,IF(DIFFERE_CPC_JOURS_{r}>27,FLUX_CPC_TRIMESTRE_{r},FLUX_CPC_TRIMESTRE_{r}*2/3))',16,PUR2),
            ('RWA_BILAN_CPC','CALC',lambda r: f'=EAD_BILAN_CPC_{r}*RW_RETENU_{r}*{const_cols["PART_CAP"]}*{const_cols["ALMT"]}',14,PUR2),
            ('TOTAL_COUTS_CPC','CALC',lambda r: f'=COUT_DIFFERE_CPC_{r}+COUT_RUN_CPC_{r}+RWA_PLAFOND_CPC_{r}+RWA_BILAN_CPC_{r}',14,PUR2)]
        COLS += [('IMPL_BONUS_CCO','SAISIE','',14,ORA2),('SIGNING_BONUS_CCO','SAISIE','',14,ORA2),('REBATE_CCO','SAISIE','',14,ORA2),
            ('COMMISSION_CCO_TTC','CALC',lambda r: f'=MAX(0,(PNB_CCO_PLAGE_{r}-IF(IMPL_BONUS_CCO_{r}="",0,IMPL_BONUS_CCO_{r})-IF(SIGNING_BONUS_CCO_{r}="",0,SIGNING_BONUS_CCO_{r})-IF(REBATE_CCO_{r}="",0,REBATE_CCO_{r}))*{const_cols["TAUX_COMM"]})',16,GRN2),
            ('RESULTAT_COND_CCO','CALC',lambda r: f'=PNB_CCO_PLAGE_{r}-IF(IMPL_BONUS_CCO_{r}="",0,IMPL_BONUS_CCO_{r})-IF(SIGNING_BONUS_CCO_{r}="",0,SIGNING_BONUS_CCO_{r})-IF(REBATE_CCO_{r}="",0,REBATE_CCO_{r})-TOTAL_COUTS_CCO_{r}-COMMISSION_CCO_TTC_{r}',16,GRN2),
            ('APPLICABILITE_CCO','CALC',lambda r: f'=IF(RESULTAT_COND_CCO_{r}>0,"OUI","NON")',14,GRN2)]
        COLS += [('IMPL_BONUS_CPC','SAISIE','',14,ORA2),('SIGNING_BONUS_CPC','SAISIE','',14,ORA2),('REBATE_CPC','SAISIE','',14,ORA2),
            ('COMMISSION_CPC_TTC','CALC',lambda r: f'=MAX(0,(PNB_CPC_PLAGE_{r}-IF(IMPL_BONUS_CPC_{r}="",0,IMPL_BONUS_CPC_{r})-IF(SIGNING_BONUS_CPC_{r}="",0,SIGNING_BONUS_CPC_{r})-IF(REBATE_CPC_{r}="",0,REBATE_CPC_{r}))*{const_cols["TAUX_COMM"]})',16,PUR2),
            ('RESULTAT_COND_CPC','CALC',lambda r: f'=PNB_CPC_PLAGE_{r}-IF(IMPL_BONUS_CPC_{r}="",0,IMPL_BONUS_CPC_{r})-IF(SIGNING_BONUS_CPC_{r}="",0,SIGNING_BONUS_CPC_{r})-IF(REBATE_CPC_{r}="",0,REBATE_CPC_{r})-TOTAL_COUTS_CPC_{r}-COMMISSION_CPC_TTC_{r}',16,PUR2),
            ('APPLICABILITE_CPC','CALC',lambda r: f'=IF(RESULTAT_COND_CPC_{r}>0,"OUI","NON")',14,PUR2)]
        COLS += [('COMMISSION_TOTALE_TTC','CALC',lambda r: f'=IF(APPLICABILITE_CCO_{r}="OUI",COMMISSION_CCO_TTC_{r},0)+IF(APPLICABILITE_CPC_{r}="OUI",COMMISSION_CPC_TTC_{r},0)',16,DARK),
            ('COMMISSION_TOTALE_HT','CALC',lambda r: f'=COMMISSION_TOTALE_TTC_{r}/(1+{const_cols["TAUX_TVA"]})',16,DARK),
            ('COUTS_TOTAUX_CLIENT','CALC',lambda r: f'=TOTAL_COUTS_CCO_{r}+TOTAL_COUTS_CPC_{r}',16,DARK),
            ('PNB_TOTAL_CLIENT','CALC',lambda r: f'=PNB_CCO_PLAGE_{r}+PNB_CPC_PLAGE_{r}',16,DARK)]

        FIN_HDR_ROW = 9; col_to_idx = {}
        for ci, (name, tp, src, width, color) in enumerate(COLS):
            col_idx = 2 + ci; col_to_idx[name] = col_idx
            c = ws.cell(row=FIN_HDR_ROW, column=col_idx, value=name)
            hdr_color = {BLU2:BLU, GRN2:GRN, PUR2:PUR, ORA2:ORA, DARK:DARK}.get(color, DARK)
            c.font = fnt(WHT, sz=9); c.fill = fill(hdr_color); c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = brd
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[FIN_HDR_ROW].height = 40

        name_to_letter = {name: get_column_letter(idx) for name, idx in col_to_idx.items()}
        def resolve_formula(ft, rn):
            if not callable(ft): return ft
            raw = ft(rn)
            for name in sorted(name_to_letter.keys(), key=len, reverse=True):
                raw = raw.replace(f'{name}_{rn}', f'{name_to_letter[name]}{rn}')
            return raw

        DATA_START_ROW = FIN_HDR_ROW + 1; n_rows = len(df_data)
        self._prog(0.94, f"FINANCIAL : {n_rows:,} lignes...")
        for i, data_row in enumerate(df_data.itertuples(index=False)):
            if i % 1000 == 0: self._prog(0.94 + 0.04*i/max(n_rows,1), f"FINANCIAL {i:,}/{n_rows:,}")
            rn = DATA_START_ROW + i; bg = WHT if i%2==0 else GRN2
            dd = data_row._asdict(); dd['ID_RC_CONSOL'] = dd.get('ID_RC_MX','') or dd.get('ID_RC_WL','')
            for ci, (name, tp, src, width, color) in enumerate(COLS):
                col_idx = 2 + ci; c = ws.cell(row=rn, column=col_idx); c.font = Font(name="Segoe UI", size=9); c.border = brd
                if tp == 'SAISIE': c.fill = fill(ORA2); c.font = Font(name="Segoe UI", size=9, bold=True, color=ORA)
                else: c.fill = fill(bg)
                if tp == 'DATA':
                    val = dd.get(src, '')
                    if isinstance(val, (int, float)) and not isinstance(val, bool):
                        c.value = float(val) if val != '' else 0.0
                        if any(name.startswith(p) for p in ('FLUX_','NB_CARTES_','NB_TRANS_','PLAFOND_','PNB_')): c.number_format = '#,##0.00'
                        elif name.endswith('_JOURS'): c.number_format = '0'
                        c.alignment = Alignment(horizontal="right")
                    else: c.value = str(val) if val not in (None,'nan','NaN') else ''; c.alignment = Alignment(horizontal="left")
                elif tp == 'SAISIE':
                    if src == '': c.value = None
                    elif isinstance(src, (int, float)): c.value = src; c.number_format = '0.00%' if src < 1 else '0.00'
                    else: c.value = str(src)
                    c.alignment = Alignment(horizontal="center")
                elif tp == 'CALC':
                    formula = resolve_formula(src, rn); c.value = formula
                    if any(k in name for k in ('COMMISSION','COUT','RWA','EAD','FLUX','PNB','RESULTAT','TOTAL')): c.number_format = '#,##0.00'
                    elif name in ('TAUX_EAD_CCO','RW_RETENU'): c.number_format = '0.00%'
                    elif name == 'MULT_PERIODICITE': c.number_format = '0'
                    c.alignment = Alignment(horizontal="right")
        self._prog(0.98, f"FINANCIAL [{VERSION_ID}] : {n_rows:,} lignes x {len(COLS)} colonnes")


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="01.Q8YY0",
        description=f"CIB COMMISSIONNEMENT ANALYZER v12 [{VERSION_ID}] - CLI autonome (sans GUI).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    # ── 8 sources obligatoires ────────────────────────────────────────────────
    p.add_argument("--wl-prgm", required=False, default=None, metavar="PATH",
                   help="Fichier WORLDLINE PRGM (source CPC) [obligatoire]")
    p.add_argument("--mx-monext", required=False, default=None, metavar="PATH",
                   help="Fichier MONEXT (source CCO) [obligatoire]")
    p.add_argument("--ref-client", required=False, default=None, metavar="PATH",
                   help="Fichier REFERENTIEL CLIENT [obligatoire]")
    p.add_argument("--idseg", required=False, default=None, metavar="PATH",
                   help="Fichier IDENTIFIANT-SEGMENT [obligatoire]")
    p.add_argument("--parc", required=False, default=None, metavar="PATH",
                   help="Fichier PARC_CLIENT [obligatoire]")
    p.add_argument("--account", required=False, default=None, metavar="PATH",
                   help="Fichier IBAN_ACCOUNT [obligatoire]")
    p.add_argument("--devises", required=False, default=None, metavar="PATH",
                   help="Fichier DEVISES [obligatoire]")
    p.add_argument("--monitoring", required=False, default=None, metavar="PATH",
                   help="Fichier MONITORING CIB [obligatoire]")
    # ── 7 sources optionnelles ────────────────────────────────────────────────
    p.add_argument("--optiflux", metavar="PATH", default=None,
                   help="Fichier OPTIFLUX (flag BPE par IBAN/RS) [optionnel]")
    p.add_argument("--bpe-retail", metavar="PATH", default=None,
                   help="Fichier CODE_AGENCE_RETAIL [optionnel]")
    p.add_argument("--seg-agence", metavar="PATH", default=None,
                   help="Fichier SEG_AGENCE [optionnel]")
    p.add_argument("--usage", metavar="PATH", default=None,
                   help="Fichier MATCHING_USAGE [optionnel]")
    p.add_argument("--mc1", metavar="PATH", default=None,
                   help="Fichier MATCHING_CLIENT_1 [optionnel]")
    p.add_argument("--mc2", metavar="PATH", default=None,
                   help="Fichier MATCHING_CLIENT_2 [optionnel]")
    p.add_argument("--override", metavar="PATH", default=None,
                   help="Fichier OVERRIDE_PAYS [optionnel]")
    p.add_argument("--override-pays", metavar="VALEUR", default="LUXEMBOURG",
                   help="Pays de remplacement OVERRIDE [optionnel]")
    # ── Sorties (obligatoires) ────────────────────────────────────────────────
    p.add_argument("--output-dir", required=True, metavar="PATH",
                   help="Repertoire de sortie [obligatoire]")
    p.add_argument("--output-filename", required=True, metavar="NOM",
                   help="Nom de base des fichiers de sortie, sans extension [obligatoire]")
    return p


def main(argv: Optional[list] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        # ── Résolution auto des sources obligatoires non fournies (03.sources) ──
        if not args.wl_prgm:
            args.wl_prgm = resolve_source("PRGM_AGREGE", required=True)
        if not args.mx_monext:
            args.mx_monext = resolve_source("MONEXT_AGREGE", required=True)
        if not args.ref_client:
            args.ref_client = resolve_source("REFERENTIEL_CLIENT", required=True)
        if not args.idseg:
            args.idseg = resolve_source("IDENTIFIANT_SEGMENT", required=True)
        if not args.parc:
            args.parc = resolve_source("PARC_CLIENT", required=True)
        if not args.account:
            args.account = resolve_source("IBAN_ACCOUNT", required=True)
        if not args.devises:
            args.devises = resolve_source("DEVISES", required=True)
        if not args.monitoring:
            args.monitoring = resolve_source("MONITORING", required=True)
        app = CIBCommissionAnalyzer_Q8YY0(args)
        app.run()
    except FileNotFoundError as exc:
        print(f"[ERREUR] Fichier introuvable : {exc}", file=sys.stderr)
        return 2
    except Exception as exc:  # noqa: BLE001
        print(f"[ERREUR] {type(exc).__name__} : {exc}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1
    print("[OK] Traitement termine avec succes.")
    return 0


if __name__ == "__main__":
    sys.exit(main())