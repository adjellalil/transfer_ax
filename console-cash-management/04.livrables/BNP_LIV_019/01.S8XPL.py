
"""
MONEXT ANALYZER v13 [S8XPL] - CLI
==================================
BNP Paribas Cash Management - Direction Monetique

DESCRIPTION
-----------
Analyseur MONEXT consolide (refonte CLI, sans GUI). Logique metier preservee
a l'identique depuis la version GUI customtkinter v13 [S8XPL] (successeur de
v12 [R7SEG]). REFERENTIEL CLIENT prioritaire (REF_RP > REF_RC > REF_IBAN) avec
fallback PARC / OPTIFLUX / CODE_AGENCE_PARC / BPE_RETAIL / SEG_AGENCE / DEFAUT.
Calcul vectorise des 5 types de PNB (Corporate, Cotisations, Commissions,
Interets, Note de frais), classification ENTREPRISE/BPE, gestion directe/
indirecte (YANNICK + critere CWT optionnel), geographie sur 3 colonnes
(Pays GA via BG, Pays Entite via RMPM) et restitution sur 3 zones geo
(GLOBAL / FRANCE / HORS FRANCE + NON TROUVE).

NOUVEAUTE v13 vs v12 :
- Correction bug XLSX : split automatique de la sheet DETAIL au-dela de
  1 000 000 lignes (limite Excel 1 048 576). Sheet TABLEAU inchangee.

Tout le reste est strictement identique a v12 [R7SEG].

SOURCES REQUISES
----------------
  MONEXT       (oblig)  CSV MONEXT consolide (source principale)
  REF_CLIENT   (oblig)  Referentiel client (fichier Olivier) - source de verite
  PARC         (oblig)  PARC_CLIENT (fallback)
  OPTI         (oblig)  OPTIFLUX (fallback)
  YANNICK      (oblig)  Gestion Directe/Indirecte
  GEO          (oblig)  BG_LE_RMPM_COUNTRY (geographie GA & Entite)
  BPE_RETAIL   (opt)    Codes agences BPE (fallback apres PARC/OPTI)
  SEG_AGENCE   (opt)    Segment agence (BCEF->ENT / BPE->BPE)

OUTPUTS PRODUITS
----------------
  MONEXT_DETAIL_<ts>_S8XPL.csv    Detail ligne a ligne (CSV ; utf-8-sig)
  MONEXT_TABLEAU_<ts>_S8XPL.csv   Tableaux PNB 3 zones geo (CSV ; utf-8-sig)
  MONEXT_AGREGE_<ts>_S8XPL.xlsx   Classeur agrege (si --xlsx et openpyxl dispo)
Les noms ci-dessus derivent de --output-filename (base) ; voir ARGUMENTS CLI.

ARGUMENTS CLI
-------------
  --monext PATH              (oblig) Fichier MONEXT consolide
  --ref-client PATH          (oblig) Referentiel client
  --parc PATH                (oblig) PARC_CLIENT
  --opti PATH                (oblig) OPTIFLUX
  --yannick PATH             (oblig) Gestion Directe/Indirecte
  --geo PATH                 (oblig) BG_LE_RMPM_COUNTRY
  --bpe-retail PATH          (opt)   BPE RETAIL ; active le fallback BPE RETAIL
  --seg-agence PATH          (opt)   SEGMENT AGENCE ; active le fallback SEG
  --cwt-force                (opt)   Force "CWT" -> Gestion Directe (en dernier)
  --xlsx                     (opt)   Genere aussi le classeur XLSX agrege
  --output-dir PATH          (oblig) Repertoire de sortie
  --output-filename NAME     (oblig) Base de nom des fichiers de sortie

DECOMPOSITION
-------------
main()
 |__ argparse (validation des chemins / options)
 |__ MonextAnalyzer_S8XPL(args)
 |     |__ __init__            : affecte fichiers + options + mapping par defaut
 |     |__ run()
 |           |__ worker()      : pipeline metier (identique GUI)
 |           |     |__ chargement CSV (load_csv_smart)
 |           |     |__ calcul PNB (plages cols 19-55, exclusion col 33, NDF/INT/COT/COM/IC)
 |           |     |__ cles MONEXT (clean_id / clean_ga / clean_iban / norm_rs)
 |           |     |__ REFERENTIEL CLIENT (dicts RP/RC/RIB, resolution vectorisee)
 |           |     |__ fallback PARC (RP/RC/RS/CA) vectorise
 |           |     |__ OPTIFLUX (IBAN/RS) vectorise
 |           |     |__ YANNICK (gestion D/I) vectorise
 |           |     |__ BPE_RETAIL / SEG_AGENCE (optionnels)
 |           |     |__ classification ENT/BPE (REF prioritaire + fallback)
 |           |     |__ gestion D/I + CWT_FORCE
 |           |     |__ identite / geo (Pays GA, Pays Entite, GEO France/HF)
 |           |     |__ tableaux PNB (3 zones : GLOBAL/FRANCE/HF+NT)
 |           |     |__ exports CSV DETAIL + TABLEAU
 |           |__ create_xlsx() : classeur agrege, split DETAIL > 1M lignes
"""

import argparse
import os
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ─── Résolution auto des sources + lecture DuckDB (inliné depuis 06.fonctions) ──
def _find_sources_root():
    here = Path(__file__).resolve()
    for d in [here] + list(here.parents):
        c = d / "03.sources"
        if c.is_dir():
            return c
    return None


def resolve_source(name, required=False):
    root = _find_sources_root()
    if not root:
        if required:
            raise FileNotFoundError(f"Dossier 03.sources introuvable depuis {__file__}")
        return None
    folder = next((m for m in root.glob("*/" + name) if m.is_dir()), None)
    if not folder:
        if required:
            raise FileNotFoundError(f"Source introuvable : 03.sources/*/{name}")
        return None
    exts = (".csv", ".xlsx", ".xls", ".txt")
    files = [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() in exts]
    if not files:
        if required:
            raise FileNotFoundError(f"Aucun fichier source dans {folder}")
        return None

    def _np(f):
        m = re.match(r"\s*(\d+)", f.name)
        return (int(m.group(1)) if m else -1, f.name)
    return max(files, key=_np)


def _read_duck(path, nrows=None):
    p = Path(path)
    if nrows is not None or p.suffix.lower() not in (".csv", ".txt", ""):
        return None
    try:
        import duckdb
        return duckdb.read_csv(str(p), all_varchar=True).df()
    except Exception:
        return None


VERSION_ID = "S8XPL"
GEO_COL_GAP = 2

# v13 [S8XPL] - Split sheet DETAIL au-dela de cette limite
# (Excel limite a 1 048 576 ; on garde une marge de securite)
XLSX_MAX_ROWS_PER_SHEET = 1_000_000

MOIS_NOMS = {
    '01': 'JANVIER', '02': 'FEVRIER', '03': 'MARS',
    '04': 'AVRIL', '05': 'MAI', '06': 'JUIN',
    '07': 'JUILLET', '08': 'AOUT', '09': 'SEPTEMBRE',
    '10': 'OCTOBRE', '11': 'NOVEMBRE', '12': 'DECEMBRE'
}

DEFAULT_POSITIONS = {
    'monext_mois': 1, 'monext_groupe_affaires': 2, 'monext_rs': 4,
    'monext_code_agence': 8, 'monext_id_rp': 9, 'monext_id_rc': 10, 'monext_iban': 11,
    'pnb_first': 19, 'pnb_last': 55,
    'pnb_cotisations': 19, 'pnb_commissions': 20,
    'pnb_notefrais_1': 22, 'pnb_notefrais_2': 55, 'pnb_interets': 28,
    'parc_rp': 1, 'parc_code_agence': 4, 'parc_rmpm': 6, 'parc_rs': 8,
    'parc_code_ga': 11, 'parc_rc': 14,
    'opti_rs': 4, 'opti_iban': 65,
    'yannick_code_ga': 1, 'yannick_sales': 3, 'yannick_gestion_indirecte': 4,
    'bpe_retail_code_agence': 3,
    'seg_code_agence': 1, 'seg_source': 7,
    'geo_pays_bg': 1, 'geo_code_ga': 2, 'geo_nom_bg': 3,
    'geo_pays_entite': 4, 'geo_rmpm_entite': 5,
    # REFERENTIEL CLIENT (v12)
    'ref_id_rp': 1, 'ref_rib': 2, 'ref_devise': 3, 'ref_rmpm': 4,
    'ref_siren': 5, 'ref_rs': 6, 'ref_rc': 7, 'ref_categorie': 8,
    'ref_segment': 9, 'ref_code_ga': 10, 'ref_nom_ga': 11,
}

# Mapping logique -> cle source/fichier, pour resolution par defaut des colonnes.
# Reproduit le comportement par defaut de l'UI : chaque champ est pris a sa
# position DEFAULT_POSITIONS dans le fichier source correspondant.
# (L'etape visuelle de mapping/selection manuelle des colonnes est supprimee.)
MAPPING_SOURCE = {
    'monext_mois': 'MONEXT', 'monext_groupe_affaires': 'MONEXT', 'monext_rs': 'MONEXT',
    'monext_code_agence': 'MONEXT', 'monext_id_rp': 'MONEXT', 'monext_id_rc': 'MONEXT',
    'monext_iban': 'MONEXT',
    'parc_rp': 'PARC', 'parc_code_agence': 'PARC', 'parc_rmpm': 'PARC', 'parc_rs': 'PARC',
    'parc_code_ga': 'PARC', 'parc_rc': 'PARC',
    'opti_rs': 'OPTI', 'opti_iban': 'OPTI',
    'yannick_code_ga': 'YANNICK', 'yannick_sales': 'YANNICK',
    'yannick_gestion_indirecte': 'YANNICK',
    'bpe_retail_code_agence': 'BPE_RETAIL',
    'seg_code_agence': 'SEG_AGENCE', 'seg_source': 'SEG_AGENCE',
    'geo_pays_bg': 'GEO', 'geo_code_ga': 'GEO', 'geo_nom_bg': 'GEO',
    'geo_pays_entite': 'GEO', 'geo_rmpm_entite': 'GEO',
    'ref_id_rp': 'REF_CLIENT', 'ref_rib': 'REF_CLIENT', 'ref_rmpm': 'REF_CLIENT',
    'ref_rc': 'REF_CLIENT', 'ref_segment': 'REF_CLIENT', 'ref_code_ga': 'REF_CLIENT',
    'ref_nom_ga': 'REF_CLIENT',
}

PNB_TYPES = [
    ('Corporate',    '_CORP_F'),
    ('Cotisations',  '_COT_F'),
    ('Commissions',  '_COM_F'),
    ('Interets',     '_INT_F'),
    ('Note de frais', '_NDF_F'),
]


def yyyymm_to_label(code: str) -> str:
    if not code or len(code) != 6:
        return code
    return f"{code[:4]}_{MOIS_NOMS.get(code[4:6], code[4:6])}"


class MonextAnalyzer_S8XPL:
    def __init__(self, args: argparse.Namespace) -> None:
        # ── Fichiers (memes cles que la version GUI) ─────────────────────────
        self.files = {
            "MONEXT": str(args.monext),
            "REF_CLIENT": str(args.ref_client),
            "PARC": str(args.parc),
            "OPTI": str(args.opti),
            "YANNICK": str(args.yannick),
            "GEO": str(args.geo),
            "BPE_RETAIL": str(args.bpe_retail) if args.bpe_retail else "",
            "SEG_AGENCE": str(args.seg_agence) if args.seg_agence else "",
        }

        # ── Options de traitement (memes attributs que la version GUI) ───────
        self.use_bpe_retail = bool(args.bpe_retail)
        self.use_seg_agence = bool(args.seg_agence)
        self.use_cwt_force = bool(args.cwt_force)
        self.gen_xlsx = bool(args.xlsx)

        # ── Sortie (remplace asksaveasfilename) ──────────────────────────────
        self.output_dir = Path(args.output_dir)
        self.output_filename = str(args.output_filename)

        # ── Colonnes a exclure (UI : champ pre-rempli "33") ──────────────────
        self.pnb_exclude_cols_text = "33"

        # ── Caches colonnes originales (resolution mapping par defaut) ───────
        self.original_cols: dict = {}

    # ══════════════════════════════════════════════════════════════════════════
    # SORTIE (remplace self.upd / progressbar / status)
    # ══════════════════════════════════════════════════════════════════════════
    def upd(self, v: float, t: str) -> None:
        print(f"  [{int(v * 100):3d}%] {t}")

    # ══════════════════════════════════════════════════════════════════════════
    # CHARGEMENT CSV — identique GUI
    # ══════════════════════════════════════════════════════════════════════════
    def load_csv_smart(self, path: str, nrows=None) -> pd.DataFrame:
        _d = _read_duck(path, nrows)
        if _d is not None:
            return _d
        for sep in [';', ',', '\t']:
            for enc in ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']:
                try:
                    df = pd.read_csv(path, sep=sep, encoding=enc, dtype=str, keep_default_na=False, na_values=[], on_bad_lines='skip', nrows=5)
                    if df.shape[1] > 1:
                        return pd.read_csv(path, sep=sep, encoding=enc, dtype=str, keep_default_na=False, na_values=[], on_bad_lines='skip', nrows=nrows)
                except Exception:
                    continue
        return pd.read_csv(path, sep=None, engine='python', dtype=str, on_bad_lines='skip', nrows=nrows)

    # ══════════════════════════════════════════════════════════════════════════
    # STATIC METHODS — identiques R7SEG
    # ══════════════════════════════════════════════════════════════════════════
    @staticmethod
    def clean_id(series):
        s = series.astype(str).str.strip()
        s = s.replace(['', 'nan', 'NaN', 'None', 'NULL', 'NA', 'N/A', 'NAN', 'NONE'], '')
        mask = s.str.startswith('="') & s.str.endswith('"')
        s = s.where(~mask, s.str[2:-1])
        s = s.str.lstrip("'")
        mask2 = s.str.endswith('.0') & s.str[:-2].str.isdigit()
        return s.where(~mask2, s.str[:-2]).str.strip()

    @staticmethod
    def clean_ga(series):
        s = MonextAnalyzer_S8XPL.clean_id(series)
        stripped = s.str.lstrip('0')
        return stripped.where(stripped != '', s)

    @staticmethod
    def clean_iban_monext(series):
        s = MonextAnalyzer_S8XPL.clean_id(series).str.upper().str.replace(' ', '', regex=False)
        return s.str[4:].where(s.str.len() > 4, s)

    @staticmethod
    def clean_iban_opti(series):
        return MonextAnalyzer_S8XPL.clean_id(series).str.upper().str.replace(' ', '', regex=False)

    @staticmethod
    def clean_rib_ref(series):
        """RIB du referentiel Olivier : 23 chars, sans FR76. Nettoyage identique."""
        return MonextAnalyzer_S8XPL.clean_id(series).str.upper().str.replace(' ', '', regex=False)

    @staticmethod
    def norm_rs(series):
        def n(v):
            if pd.isna(v) or str(v).strip() == '':
                return ''
            s = unicodedata.normalize('NFD', str(v).strip().upper())
            return ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return series.apply(n)

    @staticmethod
    def clean_gest(series):
        s = series.astype(str).str.strip().str.upper()
        r = pd.Series('', index=series.index)
        r = r.where(~s.isin(['OUI', 'O', 'YES', 'Y', '1', 'TRUE', 'VRAI']), 'OUI')
        return r.where(~s.isin(['NON', 'N', 'NO', '0', 'FALSE', 'FAUX']), 'NON')

    @staticmethod
    def to_float(series):
        s = series.astype(str)
        s = s.str.replace('"', '', regex=False).str.replace("'", '', regex=False)
        s = s.str.replace(' ', '', regex=False).str.replace('\xa0', '', regex=False).str.replace(' ', '', regex=False)
        mask = s.str.endswith('-'); s = s.where(~mask, '-' + s.str[:-1])
        s = s.str.replace(',', '.', regex=False)
        return pd.to_numeric(s, errors='coerce').fillna(0.0)

    @staticmethod
    def pays_to_geo(pays_series):
        FRANCE_VALUES = {'FR', 'FRANCE', 'FRA'}
        def _geo(v):
            v = str(v).strip()
            if not v or v.lower() in ('non trouve', 'non trouve', 'nan', '', 'none', 'null', 'pays non trouve', 'pays non trouve'):
                return 'Pays non trouve'
            return 'France' if v.upper() in FRANCE_VALUES else 'Hors France'
        return pays_series.apply(_geo)

    @staticmethod
    def parse_date(val):
        if pd.isna(val):
            return ''
        s = str(val); sc = ''.join(c for c in s if c.isdigit() or c in '/-.')
        if not sc:
            return ''
        for pat, fn in [
            (r'^(\d{1,2})[/\-\.](\d{4})$', lambda m: f"{m.group(2)}{int(m.group(1)):02d}"),
            (r'^(\d{4})[/\-\.](\d{1,2})$', lambda m: f"{m.group(1)}{int(m.group(2)):02d}" if 1 <= int(m.group(2)) <= 12 else None),
            (r'^(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})$', lambda m: f"{m.group(3)}{int(m.group(2)):02d}" if 1 <= int(m.group(2)) <= 12 else None),
            (r'^(\d{4})[/\-\.](\d{1,2})[/\-\.](\d{1,2})$', lambda m: f"{m.group(1)}{int(m.group(2)):02d}" if 1 <= int(m.group(2)) <= 12 else None),
        ]:
            match = re.match(pat, sc)
            if match:
                result = fn(match)
                if result:
                    return result
        if len(sc) == 6 and sc.isdigit() and 1 <= int(sc[4:6]) <= 12:
            return sc
        if len(sc) == 8 and sc.isdigit() and 1 <= int(sc[4:6]) <= 12:
            return sc[:6]
        try:
            dt = pd.to_datetime(s, dayfirst=True, errors='coerce')
            if pd.notna(dt):
                return dt.strftime('%Y%m')
        except Exception:
            pass
        return ''

    def parse_excluded(self) -> set:
        text = self.pnb_exclude_cols_text.strip()
        if not text:
            return set()
        return {int(p.strip()) for p in text.replace(';', ',').split(',') if p.strip().isdigit()}

    # ══════════════════════════════════════════════════════════════════════════
    # PREPARATION — remplace load_previews + show_mapping_ui + start_thread
    # Reproduit le comportement par defaut de l'UI (mapping = DEFAULT_POSITIONS).
    # ══════════════════════════════════════════════════════════════════════════
    def build_mapping_and_config(self):
        # Lecture des en-tetes (equivalent du preview nrows=5 de l'UI) pour
        # resoudre les positions DEFAULT_POSITIONS en noms de colonnes reels.
        preview_keys = ["MONEXT", "REF_CLIENT", "PARC", "OPTI", "YANNICK", "GEO"]
        if self.use_bpe_retail and self.files["BPE_RETAIL"]:
            preview_keys.append("BPE_RETAIL")
        if self.use_seg_agence and self.files["SEG_AGENCE"]:
            preview_keys.append("SEG_AGENCE")
        for key in preview_keys:
            df_prev = self.load_csv_smart(self.files[key], nrows=5)
            self.original_cols[key] = list(df_prev.columns)

        # ── Mapping colonnes : selection par defaut (UI pre-remplie) ─────────
        m: dict = {}
        for logical, src in MAPPING_SOURCE.items():
            if src not in self.original_cols:
                # Source optionnelle non fournie : colonne laissee vide
                # (equivalent UI : champ non renseigne / opt_keys).
                m[logical] = ""
                continue
            cols = self.original_cols[src]
            pos = DEFAULT_POSITIONS[logical]
            m[logical] = cols[pos - 1] if 1 <= pos <= len(cols) else ""

        # ── Config PNB : positions par defaut (UI pre-remplie) ───────────────
        # 'ic' (interchange) : "(aucune)" par defaut dans l'UI -> 0.
        pnb_cfg = {
            'first': DEFAULT_POSITIONS['pnb_first'],
            'last': DEFAULT_POSITIONS['pnb_last'],
            'nf1': DEFAULT_POSITIONS['pnb_notefrais_1'],
            'nf2': DEFAULT_POSITIONS['pnb_notefrais_2'],
            'inter': DEFAULT_POSITIONS['pnb_interets'],
            'cot': DEFAULT_POSITIONS['pnb_cotisations'],
            'com': DEFAULT_POSITIONS['pnb_commissions'],
            'ic': 0,
            'excl': self.parse_excluded(),
        }
        if pnb_cfg['first'] == 0 or pnb_cfg['last'] == 0:
            raise ValueError("Plage PNB non configuree")
        if pnb_cfg['first'] > pnb_cfg['last']:
            raise ValueError("Premiere > derniere colonne PNB")
        return m, pnb_cfg

    def run(self) -> None:
        m, pnb_cfg = self.build_mapping_and_config()
        self.worker(m, pnb_cfg, self.use_cwt_force)

    # ══════════════════════════════════════════════════════════════════════════
    # WORKER — identique R7SEG (REFERENTIEL CLIENT prioritaire, vectorise)
    # ══════════════════════════════════════════════════════════════════════════
    def worker(self, m, pnb_cfg, use_cwt):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        # ── CHARGEMENT ───────────────────────────────────────────────────
        self.upd(0.02, "Chargement MONEXT...")
        df = self.load_csv_smart(self.files["MONEXT"])
        n = len(df); cols = list(df.columns)

        self.upd(0.05, "Chargement fichiers...")
        df_ref = self.load_csv_smart(self.files["REF_CLIENT"])
        df_parc = self.load_csv_smart(self.files["PARC"])
        df_opti = self.load_csv_smart(self.files["OPTI"])
        df_yan = self.load_csv_smart(self.files["YANNICK"])
        df_geo = self.load_csv_smart(self.files["GEO"])
        df_bpe = self.load_csv_smart(self.files["BPE_RETAIL"]) if self.use_bpe_retail else None
        df_seg = self.load_csv_smart(self.files["SEG_AGENCE"]) if self.use_seg_agence else None

        # ── PNB (identique X5DET) ────────────────────────────────────────
        self.upd(0.08, "Calcul PNB...")
        fi = pnb_cfg['first'] - 1; la = pnb_cfg['last'] - 1
        excl_0 = {i - 1 for i in pnb_cfg['excl']}
        pnb_cols = cols[fi:la+1]; excl_names = [cols[i] for i in sorted(excl_0) if 0 <= i < len(cols)]
        pnb_filt = [c for c in pnb_cols if c not in excl_names]
        nf_cols = [cols[pnb_cfg['nf1']-1], cols[pnb_cfg['nf2']-1]] if pnb_cfg['nf1'] and pnb_cfg['nf2'] else []
        int_cols = [cols[pnb_cfg['inter']-1]] if pnb_cfg['inter'] else []
        cot_cols = [cols[pnb_cfg['cot']-1]] if pnb_cfg['cot'] else []
        com_cols = [cols[pnb_cfg['com']-1]] if pnb_cfg['com'] else []
        nf_cols = [c for c in nf_cols if 0 <= cols.index(c) < len(cols)]
        excl_from_corp = nf_cols + int_cols + cot_cols + com_cols
        corp_cols = [c for c in pnb_filt if c not in excl_from_corp]
        for c in list(set(pnb_cols + excl_names)):
            if c in df.columns:
                df[c] = self.to_float(df[c])
        if pnb_cfg['ic'] and 0 <= pnb_cfg['ic']-1 < len(cols):
            df[cols[pnb_cfg['ic']-1]] = df[cols[pnb_cfg['ic']-1]] * -1
        df['_CORP_F'] = df[corp_cols].sum(axis=1) if corp_cols else 0.0
        df['_NDF_F'] = df[nf_cols].sum(axis=1) if nf_cols else 0.0
        df['_INT_F'] = df[int_cols].sum(axis=1) if int_cols else 0.0
        df['_COT_F'] = df[cot_cols].sum(axis=1) if cot_cols else 0.0
        df['_COM_F'] = df[com_cols].sum(axis=1) if com_cols else 0.0
        df['_TOTAL_F'] = df['_CORP_F'] + df['_NDF_F'] + df['_INT_F'] + df['_COT_F'] + df['_COM_F']

        # ── MOIS ─────────────────────────────────────────────────────────
        self.upd(0.10, "Parsing mois..."); df['_MOIS'] = df[m['monext_mois']].apply(self.parse_date)

        # ── CLES MONEXT ──────────────────────────────────────────────────
        self.upd(0.12, "Cles MONEXT...")
        df['_RP'] = self.clean_ga(df[m['monext_id_rp']])
        df['_RC'] = self.clean_ga(df[m['monext_id_rc']])
        df['_RC_RAW'] = self.clean_id(df[m['monext_id_rc']])
        df['_RS'] = self.norm_rs(df[m['monext_rs']])
        df['_IBAN'] = self.clean_iban_monext(df[m['monext_iban']])
        df['_CA'] = self.clean_ga(df[m['monext_code_agence']])
        df['_GA_STR'] = df[m['monext_groupe_affaires']].astype(str).str.strip().str.upper()

        # ══════════════════════════════════════════════════════════════════
        # REFERENTIEL CLIENT — construction des 3 dicts de lookup
        # ══════════════════════════════════════════════════════════════════
        self.upd(0.14, "REFERENTIEL CLIENT dicts...")
        ref_rp_clean = self.clean_ga(df_ref[m['ref_id_rp']])
        ref_rc_clean = self.clean_ga(df_ref[m['ref_rc']])
        ref_rc_raw = self.clean_id(df_ref[m['ref_rc']])
        ref_rib_clean = self.clean_rib_ref(df_ref[m['ref_rib']])
        ref_segment = df_ref[m['ref_segment']].astype(str).str.strip().str.upper()
        ref_code_ga = self.clean_ga(df_ref[m['ref_code_ga']])
        ref_nom_ga = df_ref[m['ref_nom_ga']].astype(str).str.strip()
        ref_rmpm = self.clean_id(df_ref[m['ref_rmpm']])

        def norm_seg(s):
            if 'ENTREPRISE' in s:
                return 'ENTREPRISE'
            elif 'BPE' in s:
                return 'BPE'
            return ''
        ref_seg_norm = ref_segment.apply(norm_seg)

        d_ref_rp = {}
        for rp, seg, ga, nga, rmpm in zip(ref_rp_clean.values, ref_seg_norm.values, ref_code_ga.values, ref_nom_ga.values, ref_rmpm.values):
            if rp and rp not in d_ref_rp:
                d_ref_rp[rp] = (seg, ga, nga, rmpm)

        d_ref_rc = {}
        for rc, rc_r, seg, ga, nga, rmpm in zip(ref_rc_clean.values, ref_rc_raw.values, ref_seg_norm.values, ref_code_ga.values, ref_nom_ga.values, ref_rmpm.values):
            if rc_r and rc_r not in d_ref_rc:
                d_ref_rc[rc_r] = (seg, ga, nga, rmpm)
            if rc and rc != rc_r and rc not in d_ref_rc:
                d_ref_rc[rc] = (seg, ga, nga, rmpm)

        d_ref_rib = {}
        for rib, seg, ga, nga, rmpm in zip(ref_rib_clean.values, ref_seg_norm.values, ref_code_ga.values, ref_nom_ga.values, ref_rmpm.values):
            if rib and rib not in d_ref_rib:
                d_ref_rib[rib] = (seg, ga, nga, rmpm)

        self.upd(0.16, f"REF : RP={len(d_ref_rp):,} | RC={len(d_ref_rc):,} | RIB={len(d_ref_rib):,}")

        ref_rp_seg = {k: v[0] for k, v in d_ref_rp.items()}; ref_rp_ga = {k: v[1] for k, v in d_ref_rp.items()}
        ref_rp_nga = {k: v[2] for k, v in d_ref_rp.items()}; ref_rp_rmpm = {k: v[3] for k, v in d_ref_rp.items()}
        ref_rc_seg = {k: v[0] for k, v in d_ref_rc.items()}; ref_rc_ga = {k: v[1] for k, v in d_ref_rc.items()}
        ref_rc_nga = {k: v[2] for k, v in d_ref_rc.items()}; ref_rc_rmpm = {k: v[3] for k, v in d_ref_rc.items()}
        ref_rib_seg = {k: v[0] for k, v in d_ref_rib.items()}; ref_rib_ga = {k: v[1] for k, v in d_ref_rib.items()}
        ref_rib_nga = {k: v[2] for k, v in d_ref_rib.items()}; ref_rib_rmpm = {k: v[3] for k, v in d_ref_rib.items()}

        # ══════════════════════════════════════════════════════════════════
        # REFERENTIEL — RESOLUTION VECTORISEE (RP -> RC -> RIB)
        # ══════════════════════════════════════════════════════════════════
        self.upd(0.18, "REFERENTIEL vectorise...")
        hit_rp_seg = df['_RP'].map(ref_rp_seg); f_rp = hit_rp_seg.notna() & (hit_rp_seg != '')
        hit_rc_raw = df['_RC_RAW'].map(ref_rc_seg); hit_rc = df['_RC'].map(ref_rc_seg)
        hit_rc_seg = hit_rc_raw.where(hit_rc_raw.notna() & (hit_rc_raw != ''), hit_rc)
        f_rc = ~f_rp & hit_rc_seg.notna() & (hit_rc_seg != '')
        hit_rib_seg = df['_IBAN'].map(ref_rib_seg)
        f_rib = ~f_rp & ~f_rc & hit_rib_seg.notna() & (hit_rib_seg != '')
        f_ref = f_rp | f_rc | f_rib

        seg_from_ref = np.select([f_rp, f_rc, f_rib],
            [df['_RP'].map(ref_rp_seg), hit_rc_seg, df['_IBAN'].map(ref_rib_seg)], '')
        ga_from_ref = np.select([f_rp, f_rc, f_rib],
            [df['_RP'].map(ref_rp_ga), df['_RC_RAW'].map(ref_rc_ga).where(hit_rc_raw.notna(), df['_RC'].map(ref_rc_ga)), df['_IBAN'].map(ref_rib_ga)], '')
        nga_from_ref = np.select([f_rp, f_rc, f_rib],
            [df['_RP'].map(ref_rp_nga), df['_RC_RAW'].map(ref_rc_nga).where(hit_rc_raw.notna(), df['_RC'].map(ref_rc_nga)), df['_IBAN'].map(ref_rib_nga)], '')
        rmpm_from_ref = np.select([f_rp, f_rc, f_rib],
            [df['_RP'].map(ref_rp_rmpm), df['_RC_RAW'].map(ref_rc_rmpm).where(hit_rc_raw.notna(), df['_RC'].map(ref_rc_rmpm)), df['_IBAN'].map(ref_rib_rmpm)], '')
        ref_source = np.select([f_rp, f_rc, f_rib], ['REF_RP', 'REF_RC', 'REF_IBAN'], 'FALLBACK')

        self.upd(0.22, f"REF resolu : RP={int(f_rp.sum()):,} | RC={int(f_rc.sum()):,} | RIB={int(f_rib.sum()):,} | Non={int((~f_ref).sum()):,}")

        # ══════════════════════════════════════════════════════════════════
        # PARC (fallback — identique X5DET) — VECTORISE
        # ══════════════════════════════════════════════════════════════════
        self.upd(0.24, "PARC...")
        df_parc['_RP'] = self.clean_ga(df_parc[m['parc_rp']]); df_parc['_RC'] = self.clean_ga(df_parc[m['parc_rc']])
        df_parc['_RS'] = self.norm_rs(df_parc[m['parc_rs']]); df_parc['_CA'] = self.clean_ga(df_parc[m['parc_code_agence']])
        df_parc['_RMPM'] = self.clean_id(df_parc[m['parc_rmpm']]); df_parc['_CODE_GA'] = self.clean_ga(df_parc[m['parc_code_ga']])
        OUT = ['_RMPM', '_CODE_GA']
        d_rp = df_parc[df_parc['_RP'] != ''].drop_duplicates('_RP').set_index('_RP')[OUT].T.to_dict('list')
        d_rc = df_parc[df_parc['_RC'] != ''].drop_duplicates('_RC').set_index('_RC')[OUT].T.to_dict('list')
        d_rs = df_parc[df_parc['_RS'] != ''].drop_duplicates('_RS').set_index('_RS')[OUT].T.to_dict('list')
        d_ca = df_parc[df_parc['_CA'] != ''].drop_duplicates('_CA').set_index('_CA')[OUT].T.to_dict('list')

        parc_rp_rmpm = df['_RP'].map({k: v[0] for k, v in d_rp.items()})
        parc_rp_ga = df['_RP'].map({k: v[1] for k, v in d_rp.items()})
        fp_rp = ~f_ref & parc_rp_rmpm.notna()

        parc_rc_rmpm = df['_RC'].map({k: v[0] for k, v in d_rc.items()})
        parc_rc_ga = df['_RC'].map({k: v[1] for k, v in d_rc.items()})
        fp_rc = ~f_ref & ~fp_rp & parc_rc_rmpm.notna()

        parc_rs_rmpm = df['_RS'].map({k: v[0] for k, v in d_rs.items()})
        parc_rs_ga = df['_RS'].map({k: v[1] for k, v in d_rs.items()})
        fp_rs = ~f_ref & ~fp_rp & ~fp_rc & parc_rs_rmpm.notna()

        fp = fp_rp | fp_rc | fp_rs
        df['FOUND_PARC'] = np.where(fp, 'YES', 'NO')
        df['METHOD_PARC'] = np.select([fp_rp, fp_rc, fp_rs], ['ID_RP', 'ID_RC', 'RS_EXACT'], 'N/A')

        ga = pd.Series(ga_from_ref, index=df.index).fillna('')
        rmpm = pd.Series(rmpm_from_ref, index=df.index).fillna('')
        nom_ga = pd.Series(nga_from_ref, index=df.index).fillna('')

        ga = ga.where(ga != '', np.select([fp_rp, fp_rc, fp_rs], [parc_rp_ga, parc_rc_ga, parc_rs_ga], ''))
        rmpm = rmpm.where(rmpm != '', np.select([fp_rp, fp_rc, fp_rs], [parc_rp_rmpm, parc_rc_rmpm, parc_rs_rmpm], ''))

        parc_ca_rmpm = df['_CA'].map({k: v[0] for k, v in d_ca.items()})
        parc_ca_ga = df['_CA'].map({k: v[1] for k, v in d_ca.items()})
        f_ca_parc = ~f_ref & ~fp & parc_ca_rmpm.notna()
        ga = pd.Series(ga, index=df.index).where(pd.Series(ga, index=df.index) != '', parc_ca_ga.where(f_ca_parc, ''))
        rmpm = pd.Series(rmpm, index=df.index).where(pd.Series(rmpm, index=df.index) != '', parc_ca_rmpm.where(f_ca_parc, ''))

        # ══════════════════════════════════════════════════════════════════
        # OPTIFLUX — VECTORISE
        # ══════════════════════════════════════════════════════════════════
        self.upd(0.28, "OPTIFLUX...")
        df_opti['_IBAN'] = self.clean_iban_opti(df_opti[m['opti_iban']]); df_opti['_RS'] = self.norm_rs(df_opti[m['opti_rs']])
        set_opti_iban = set(df_opti[df_opti['_IBAN'] != '']['_IBAN'].unique())
        set_opti_rs = set(df_opti[df_opti['_RS'] != '']['_RS'].unique())
        fo_iban = df['_IBAN'].isin(set_opti_iban) & (df['_IBAN'] != '')
        fo_rs = ~fo_iban & df['_RS'].isin(set_opti_rs) & (df['_RS'] != '')
        fo = fo_iban | fo_rs
        df['FOUND_OPTIFLUX'] = np.where(fo, 'YES', 'NO')

        # ══════════════════════════════════════════════════════════════════
        # YANNICK — VECTORISE
        # ══════════════════════════════════════════════════════════════════
        self.upd(0.30, "YANNICK...")
        df_yan['_CODE'] = self.clean_ga(df_yan[m['yannick_code_ga']])
        df_yan['_SALES'] = df_yan[m['yannick_sales']].astype(str).str.strip()
        df_yan['_IS_DIR'] = (self.clean_gest(df_yan[m['yannick_gestion_indirecte']]) == 'NON')
        y_sales = df_yan[df_yan['_CODE'] != ''].drop_duplicates('_CODE').set_index('_CODE')['_SALES'].to_dict()
        y_dir = df_yan[df_yan['_CODE'] != ''].drop_duplicates('_CODE').set_index('_CODE')['_IS_DIR'].to_dict()

        ga_s = pd.Series(ga, index=df.index).fillna('')
        rmpm_s = pd.Series(rmpm, index=df.index).fillna('')
        yan_sales_ga = ga_s.map(y_sales).fillna(''); yan_dir_ga = ga_s.map(y_dir)
        yan_sales_rmpm = rmpm_s.map(y_sales).fillna(''); yan_dir_rmpm = rmpm_s.map(y_dir)
        fyan = yan_dir_ga.notna() | yan_dir_rmpm.notna()
        df['FOUND_YANNICK'] = np.where(fyan, 'YES', 'NO')
        sales = yan_sales_ga.where(yan_dir_ga.notna(), yan_sales_rmpm).fillna('N/A')
        is_dir_yan = yan_dir_ga.where(yan_dir_ga.notna(), yan_dir_rmpm).fillna(False)

        # ══════════════════════════════════════════════════════════════════
        # BPE RETAIL / SEG AGENCE (optionnels) ────────────────────────────
        # ══════════════════════════════════════════════════════════════════
        set_bpe_retail = set()
        if df_bpe is not None and m.get('bpe_retail_code_agence'):
            codes = self.clean_ga(df_bpe[m['bpe_retail_code_agence']]); set_bpe_retail = set(codes[codes != ''].unique())
        seg_dict = {}
        if df_seg is not None and m.get('seg_code_agence') and m.get('seg_source'):
            tmp_ca = self.clean_ga(df_seg[m['seg_code_agence']]); tmp_src = df_seg[m['seg_source']].astype(str).str.strip().str.upper()
            for ca, src in zip(tmp_ca, tmp_src):
                if ca and ca not in seg_dict:
                    if src == 'BPE':
                        seg_dict[ca] = 'BPE'
                    elif src == 'BCEF':
                        seg_dict[ca] = 'ENTREPRISE'

        # ══════════════════════════════════════════════════════════════════
        # CLASSIFICATION ENT/BPE — REF PRIORITAIRE + FALLBACK X5DET
        # ══════════════════════════════════════════════════════════════════
        self.upd(0.35, "Classification...")
        seg_raw = pd.Series(seg_from_ref, index=df.index).fillna('')
        ref_is_ent = f_ref & (seg_raw == 'ENTREPRISE')
        ref_is_bpe = f_ref & (seg_raw == 'BPE')
        ref_no_seg = f_ref & (seg_raw == '')

        fb = ~f_ref | ref_no_seg
        fb_parc_only = fb & fp & ~fo
        fb_parc_opti = fb & fp & fo
        fb_opti_only = fb & ~fp & fo
        fb_ca_parc = fb & ~fp & ~fo & f_ca_parc
        fb_bpe_retail = fb & ~fp & ~fo & ~f_ca_parc & df['_CA'].isin(set_bpe_retail) if set_bpe_retail else pd.Series(False, index=df.index)
        fb_seg_agence_bpe = pd.Series(False, index=df.index)
        fb_seg_agence_ent = pd.Series(False, index=df.index)
        if seg_dict:
            ca_seg = df['_CA'].map(seg_dict)
            fb_seg_bpe_mask = fb & ~fp & ~fo & ~f_ca_parc & ~fb_bpe_retail & (ca_seg == 'BPE')
            fb_seg_ent_mask = fb & ~fp & ~fo & ~f_ca_parc & ~fb_bpe_retail & (ca_seg == 'ENTREPRISE')
            fb_seg_agence_bpe = fb_seg_bpe_mask
            fb_seg_agence_ent = fb_seg_ent_mask
        fb_ent_def = fb & ~fb_parc_only & ~fb_parc_opti & ~fb_opti_only & ~f_ca_parc & ~fb_bpe_retail & ~fb_seg_agence_bpe & ~fb_seg_agence_ent

        is_ent = ref_is_ent | fb_parc_only | fb_ca_parc | fb_seg_agence_ent | fb_ent_def
        is_bpe = ref_is_bpe | fb_parc_opti | fb_opti_only | fb_bpe_retail | fb_seg_agence_bpe

        class_source = np.select([
            ref_is_ent, ref_is_bpe,
            fb_parc_only, fb_parc_opti, fb_opti_only, fb_ca_parc,
            fb_bpe_retail, fb_seg_agence_bpe, fb_seg_agence_ent, fb_ent_def
        ], [
            'REF_' + pd.Series(ref_source, index=df.index), 'REF_' + pd.Series(ref_source, index=df.index),
            'PARC_ONLY', 'PARC_ET_OPTIFLUX', 'OPTIFLUX_ONLY', 'CODE_AGENCE_PARC',
            'BPE_RETAIL', 'SEG_AGENCE_BPE', 'SEG_AGENCE_ENT', 'ENTREPRISE_DEFAUT'
        ], 'DEFAUT')

        df['ENTREPRISE'] = np.where(is_ent, 'YES', 'NO')
        df['BPE'] = np.where(is_bpe, 'YES', 'NO')
        df['CLASS_SOURCE'] = class_source
        df['REF_SOURCE'] = ref_source
        df['SEGMENT_RAW'] = seg_raw

        # ══════════════════════════════════════════════════════════════════
        # GESTION DIRECTE/INDIRECTE
        # ══════════════════════════════════════════════════════════════════
        self.upd(0.40, "Gestion D/I...")
        gd = np.where(fyan, np.where(is_dir_yan, 'YES', 'NO'), 'NO')
        g_src = np.where(fyan, 'YANNICK', 'DEFAUT')

        cwt_cnt = 0
        if use_cwt:
            mask_cwt = df['_GA_STR'].str.contains('CWT', na=False) & (pd.Series(gd) != 'YES')
            cwt_cnt = int(mask_cwt.sum())
            gd = np.where(mask_cwt, 'YES', gd)
            g_src = np.where(mask_cwt, 'CWT_FORCE', g_src)

        gi = np.where(pd.Series(gd) == 'YES', 'NO', 'YES')
        df['GESTION_DIRECTE'] = gd; df['GESTION_INDIRECTE'] = gi; df['GEST_SOURCE'] = g_src; df['SALES_YANNICK'] = sales

        # ══════════════════════════════════════════════════════════════════
        # IDENTITE / GEO
        # ══════════════════════════════════════════════════════════════════
        self.upd(0.45, "Identite/Geo...")
        ga_s = pd.Series(ga, index=df.index).fillna('')
        rmpm_s = pd.Series(rmpm, index=df.index).fillna('')
        df['CODE_GA'] = ga_s.where(ga_s != '', 'N/A')
        df['NOM_GA'] = pd.Series(nom_ga, index=df.index).fillna('')
        df['RMPM'] = rmpm_s.where(rmpm_s != '', 'N/A')
        df['MOIS'] = df['_MOIS'].apply(yyyymm_to_label)

        col_geo_code_ga = m.get('geo_code_ga', ''); col_geo_pays_bg = m.get('geo_pays_bg', '')
        col_geo_rmpm_entite = m.get('geo_rmpm_entite', ''); col_geo_pays_entite = m.get('geo_pays_entite', '')
        geo_ga_dict = {}
        if col_geo_code_ga and col_geo_pays_bg:
            tmp_ga = self.clean_ga(df_geo[col_geo_code_ga]); tmp_pays = df_geo[col_geo_pays_bg].astype(str).str.strip()
            for gv, pv in zip(tmp_ga, tmp_pays):
                if gv and gv not in geo_ga_dict:
                    geo_ga_dict[gv] = pv
        geo_rmpm_dict = {}
        if col_geo_rmpm_entite and col_geo_pays_entite:
            tmp_rmpm = self.clean_id(df_geo[col_geo_rmpm_entite]); tmp_payse = df_geo[col_geo_pays_entite].astype(str).str.strip()
            for rv, pv in zip(tmp_rmpm, tmp_payse):
                if rv and rv not in geo_rmpm_dict:
                    geo_rmpm_dict[rv] = pv

        df['PAYS_GA'] = ga_s.map(geo_ga_dict).fillna('Pays non trouve')
        df['PAYS_GA'] = df['PAYS_GA'].where(df['CODE_GA'] != 'N/A', 'Pays non trouve')
        df['PAYS_ENTITE_JURIDIQUE'] = rmpm_s.map(geo_rmpm_dict).fillna('Pays non trouve')
        df['PAYS_ENTITE_JURIDIQUE'] = df['PAYS_ENTITE_JURIDIQUE'].where(df['RMPM'] != 'N/A', 'Pays non trouve')
        df['GEO_GA'] = self.pays_to_geo(df['PAYS_GA'])
        df['GEO_ENTITE_JURIDIQUE'] = self.pays_to_geo(df['PAYS_ENTITE_JURIDIQUE'])

        for c in pnb_cols:
            if c in df.columns:
                df[c] = df[c].round(2).astype(str).str.replace('.', ',', regex=False)
        for c in excl_names:
            if c in df.columns and c not in pnb_cols:
                df[c] = df[c].round(2).astype(str).str.replace('.', ',', regex=False)
        df['PNB_CORPORATE'] = df['_CORP_F'].round(2).astype(str).str.replace('.', ',', regex=False)
        df['PNB_COTISATIONS'] = df['_COT_F'].round(2).astype(str).str.replace('.', ',', regex=False)
        df['PNB_COMMISSIONS'] = df['_COM_F'].round(2).astype(str).str.replace('.', ',', regex=False)
        df['PNB_NOTEFRAIS'] = df['_NDF_F'].round(2).astype(str).str.replace('.', ',', regex=False)
        df['PNB_INTERETS'] = df['_INT_F'].round(2).astype(str).str.replace('.', ',', regex=False)
        df['PNB_TOTAL'] = df['_TOTAL_F'].round(2).astype(str).str.replace('.', ',', regex=False)

        # ══════════════════════════════════════════════════════════════════
        # TABLEAUX PNB — identique X5DET
        # ══════════════════════════════════════════════════════════════════
        self.upd(0.55, "Tableaux PNB...")
        mois_list = sorted([x for x in df['_MOIS'].unique() if x and len(x) == 6])

        def mk_pnb_pivot(df_sub):
            rows = []; total_mois = {yyyymm_to_label(mc): 0.0 for mc in mois_list}
            for tl, tc in PNB_TYPES:
                row = {'': tl}
                for mc in mois_list:
                    val = df_sub[df_sub['_MOIS'] == mc][tc].sum(); cl = yyyymm_to_label(mc); row[cl] = val; total_mois[cl] += val
                rows.append(row)
                row_gd = {'': '  |__ DIRECTE'}; df_gd = df_sub[df_sub['GESTION_DIRECTE'] == 'YES']
                for mc in mois_list:
                    row_gd[yyyymm_to_label(mc)] = df_gd[df_gd['_MOIS'] == mc][tc].sum()
                rows.append(row_gd)
                row_gi = {'': '  |__ INDIRECTE'}; df_gi = df_sub[df_sub['GESTION_INDIRECTE'] == 'YES']
                for mc in mois_list:
                    row_gi[yyyymm_to_label(mc)] = df_gi[df_gi['_MOIS'] == mc][tc].sum()
                rows.append(row_gi)
            row_tm = {'': 'TOTAL_MOIS'}; row_tm.update(total_mois); rows.append(row_tm)
            gt = sum(total_mois.values()); row_tot = {'': 'TOTAL'}
            for mc in mois_list:
                row_tot[yyyymm_to_label(mc)] = ''
            if mois_list:
                row_tot[yyyymm_to_label(mois_list[0])] = gt
            rows.append(row_tot); return pd.DataFrame(rows)

        def mk_simple_pivot(df_sub, label, pnb_col):
            rows = []; total_mois = {yyyymm_to_label(mc): 0.0 for mc in mois_list}
            row = {'': label}
            for mc in mois_list:
                val = df_sub[df_sub['_MOIS'] == mc][pnb_col].sum(); cl = yyyymm_to_label(mc); row[cl] = val; total_mois[cl] += val
            rows.append(row)
            row_gd = {'': '  |__ DIRECTE'}; df_gd = df_sub[df_sub['GESTION_DIRECTE'] == 'YES']
            for mc in mois_list:
                row_gd[yyyymm_to_label(mc)] = df_gd[df_gd['_MOIS'] == mc][pnb_col].sum()
            rows.append(row_gd)
            row_gi = {'': '  |__ INDIRECTE'}; df_gi = df_sub[df_sub['GESTION_INDIRECTE'] == 'YES']
            for mc in mois_list:
                row_gi[yyyymm_to_label(mc)] = df_gi[df_gi['_MOIS'] == mc][pnb_col].sum()
            rows.append(row_gi)
            row_tm = {'': 'TOTAL_MOIS'}; row_tm.update(total_mois); rows.append(row_tm)
            gt = sum(total_mois.values()); row_tot = {'': 'TOTAL'}
            for mc in mois_list:
                row_tot[yyyymm_to_label(mc)] = ''
            if mois_list:
                row_tot[yyyymm_to_label(mois_list[0])] = gt
            rows.append(row_tot); return pd.DataFrame(rows)

        def mk_gest(df_sub):
            gd_v = df_sub[df_sub['GESTION_DIRECTE'] == 'YES']['_TOTAL_F'].sum()
            gi_v = df_sub[df_sub['GESTION_INDIRECTE'] == 'YES']['_TOTAL_F'].sum()
            return pd.DataFrame({'GESTION': ['DIRECTE', 'INDIRECTE', 'TOTAL'], 'PNB_TOTAL': [gd_v, gi_v, gd_v + gi_v]})

        df['_CCM_F'] = df['_CORP_F'] + df['_COT_F'] + df['_COM_F']
        mask_france = df['GEO_ENTITE_JURIDIQUE'] == 'France'
        mask_hf = df['GEO_ENTITE_JURIDIQUE'].isin(['Hors France', 'Pays non trouve'])
        df_ent = df[df['ENTREPRISE'] == 'YES']; df_bpe_f = df[df['BPE'] == 'YES']
        df_fra = df[mask_france]; df_hf = df[mask_hf]

        peg = mk_pnb_pivot(df_ent); pef = mk_pnb_pivot(df_ent[mask_france]); peh = mk_pnb_pivot(df_ent[mask_hf])
        pbg = mk_pnb_pivot(df_bpe_f); pbf = mk_pnb_pivot(df_bpe_f[mask_france]); pbh = mk_pnb_pivot(df_bpe_f[mask_hf])
        pcg = mk_simple_pivot(df, 'Corp+Cot+Com', '_CCM_F'); pcf = mk_simple_pivot(df_fra, 'Corp+Cot+Com', '_CCM_F'); pch = mk_simple_pivot(df_hf, 'Corp+Cot+Com', '_CCM_F')
        png = mk_simple_pivot(df, 'Note de frais', '_NDF_F'); pnf = mk_simple_pivot(df_fra, 'Note de frais', '_NDF_F'); pnh = mk_simple_pivot(df_hf, 'Note de frais', '_NDF_F')
        pgg = mk_gest(df); pgf = mk_gest(df_fra); pgh = mk_gest(df_hf)

        df.drop(columns=[c for c in df.columns if c.startswith('_')], inplace=True, errors='ignore')

        # ══════════════════════════════════════════════════════════════════
        # EXPORTS CSV
        # ══════════════════════════════════════════════════════════════════
        self.output_dir.mkdir(parents=True, exist_ok=True)
        base = self.output_filename

        self.upd(0.70, "Export DETAIL...")
        save_det = str(self.output_dir / f"{base}_DETAIL_{ts}_{VERSION_ID}.csv")
        df.to_csv(save_det, sep=';', index=False, encoding='utf-8-sig')

        self.upd(0.80, "Export TABLEAU...")
        save_tab = str(self.output_dir / f"{base}_TABLEAU_{ts}_{VERSION_ID}.csv")

        def fmt(d):
            d2 = d.copy()
            for c in d2.columns:
                if c not in ('', 'GESTION'):
                    d2[c] = d2[c].apply(lambda x: f"{x:.2f}".replace('.', ',') if isinstance(x, (int, float)) else x)
            return d2
        with open(save_tab, 'w', encoding='utf-8-sig') as f:
            for gl, pe, pb, pc, pn, pg in [("GLOBAL", peg, pbg, pcg, png, pgg), ("FRANCE", pef, pbf, pcf, pnf, pgf), ("HORS FRANCE + NON TROUVE", peh, pbh, pch, pnh, pgh)]:
                f.write(f"=== {gl} ===\n\n"); f.write("--- ENTREPRISE ---\n"); fmt(pe).to_csv(f, sep=';', index=False, lineterminator='\n')
                f.write("\n--- BPE ---\n"); fmt(pb).to_csv(f, sep=';', index=False, lineterminator='\n')
                f.write("\n--- CORP + COT + COM ---\n"); fmt(pc).to_csv(f, sep=';', index=False, lineterminator='\n')
                f.write("\n--- NOTE DE FRAIS ---\n"); fmt(pn).to_csv(f, sep=';', index=False, lineterminator='\n')
                f.write("\n--- GESTION ---\n"); tg = pg.copy(); tg['PNB_TOTAL'] = tg['PNB_TOTAL'].apply(lambda x: f"{x:.2f}".replace('.', ',')); tg.to_csv(f, sep=';', index=False, lineterminator='\n')
                f.write("\n\n")

        self.upd(0.90, "Termine !")
        nb_ref_rp = int(f_rp.sum()); nb_ref_rc = int(f_rc.sum()); nb_ref_rib = int(f_rib.sum()); nb_fallback = int((~f_ref).sum())

        print()
        print("Fichiers CSV generes !")
        print()
        print(f"DETAIL  : {os.path.basename(save_det)}")
        print(f"TABLEAU : {os.path.basename(save_tab)}")
        print()
        print(f"- Total : {n:,}")
        print(f"- PARC (fallback) : {int(fp.sum()):,} | OPTI : {int(fo.sum()):,}")
        print(f"- ENTREPRISE : {int(is_ent.sum()):,} | BPE : {int(is_bpe.sum()):,}")
        print(f"- Gestion Directe : {int((pd.Series(gd) == 'YES').sum()):,} | Indirecte : {int((pd.Series(gi) == 'YES').sum()):,}")
        if cwt_cnt:
            print(f"- CWT : {cwt_cnt:,}")
        print()
        print("-- REFERENTIEL CLIENT (v12) --")
        print(f"- REF via RP   : {nb_ref_rp:,}")
        print(f"- REF via RC   : {nb_ref_rc:,}")
        print(f"- REF via RIB  : {nb_ref_rib:,}")
        print(f"- FALLBACK     : {nb_fallback:,}")
        print()

        if self.gen_xlsx:
            self.create_xlsx(save_det, ts, peg, pef, peh, pbg, pbf, pbh, pcg, pcf, pch, png, pnf, pnh, pgg, pgf, pgh, mois_list)

    # ══════════════════════════════════════════════════════════════════════════
    # XLSX — v13 [S8XPL] : split DETAIL > 1M lignes en plusieurs sheets
    # TABLEAU reste inchangee
    # ══════════════════════════════════════════════════════════════════════════
    def create_xlsx(self, det_path, ts, peg, pef, peh, pbg, pbf, pbh, pcg, pcf, pch, png, pnf, pnh, pgg, pgf, pgh, mois_list):
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl non installe.")
        self.upd(0.5, "XLSX...")
        save_x = str(self.output_dir / f"{self.output_filename}_AGREGE_{ts}_{VERSION_ID}.xlsx")
        wb = Workbook()

        # ── Lecture du CSV DETAIL ──────────────────────────────────────
        df_d = pd.read_csv(det_path, sep=';', encoding='utf-8-sig', dtype=str, keep_default_na=False)
        total_rows = len(df_d)
        hfill = PatternFill("solid", fgColor="00915A"); hfont = Font(bold=True, color="FFFFFF", size=10)
        pnb_col_names = {c for c in df_d.columns if c.startswith('PNB_')}

        # ── Determination du split ─────────────────────────────────────
        if total_rows <= XLSX_MAX_ROWS_PER_SHEET:
            chunks = [(0, total_rows, "DETAIL")]
        else:
            chunks = []
            nb_chunks = (total_rows + XLSX_MAX_ROWS_PER_SHEET - 1) // XLSX_MAX_ROWS_PER_SHEET
            for i in range(nb_chunks):
                start = i * XLSX_MAX_ROWS_PER_SHEET
                end = min(start + XLSX_MAX_ROWS_PER_SHEET, total_rows)
                chunks.append((start, end, f"DETAIL_{i+1}"))

        # ── Ecriture des sheets DETAIL (1 ou N) ────────────────────────
        first_sheet = True
        for chunk_idx, (start, end, sheet_name) in enumerate(chunks, start=1):
            self.upd(0.5 + 0.4 * (chunk_idx / max(len(chunks), 1)),
                     f"DETAIL {chunk_idx}/{len(chunks)} ({end-start:,} lignes)...")
            if first_sheet:
                ws = wb.active
                ws.title = sheet_name
                first_sheet = False
            else:
                ws = wb.create_sheet(title=sheet_name)
            ws.column_dimensions['A'].width = 2.5
            title_suffix = "" if len(chunks) == 1 else f" - Partie {chunk_idx}/{len(chunks)} (lignes {start+1:,} a {end:,})"
            ws['B4'] = f"MONEXT - Analyse detaillee - {datetime.now().strftime('%d/%m/%Y')}{title_suffix}"
            ws['B4'].font = Font(bold=True, size=14)

            # Header (ligne 6)
            for ci, cn in enumerate(df_d.columns, start=2):
                c = ws.cell(row=6, column=ci, value=cn); c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(ci)].width = 15

            # Donnees — slice du DataFrame
            df_chunk = df_d.iloc[start:end]
            for ri, row_data in enumerate(df_chunk.itertuples(index=False), start=7):
                for ci, (cn, val) in enumerate(zip(df_d.columns, row_data), start=2):
                    s = str(val) if val is not None else ''
                    if cn in pnb_col_names and s not in ('', 'nan'):
                        try:
                            ws.cell(row=ri, column=ci, value=float(s.replace(',', '.')))
                        except ValueError:
                            ws.cell(row=ri, column=ci, value=s)
                    else:
                        ws.cell(row=ri, column=ci, value=s if s not in ('nan', 'None') else '')

        # ── Sheet TABLEAU — inchangee ─────────────────────────────────
        self.upd(0.92, "TABLEAU...")
        ws2 = wb.create_sheet(title="TABLEAU"); ws2.column_dimensions['A'].width = 2.5
        ws2['B4'] = f"MONEXT - Tableaux PNB - {datetime.now().strftime('%d/%m/%Y')}"; ws2['B4'].font = Font(bold=True, size=14)
        CG, CF, CH = "006B43", "1565C0", "B71C1C"; SG, SF, SH = "E8F5E9", "E3F2FD", "FFEBEE"
        nb_mois = len(mois_list); bdw = max(1 + nb_mois, 2)

        def wdb(ws, sr, sc, dff, hc, sbc):
            if dff is None or len(dff) == 0:
                ws.cell(row=sr, column=sc, value="(Aucune donnee)").font = Font(italic=True, color="999999"); return 2
            r = sr
            for ci, cn in enumerate(dff.columns):
                c = ws.cell(row=r, column=sc + ci, value=cn); c.font = Font(bold=True, color="FFFFFF", size=9); c.fill = PatternFill("solid", fgColor=hc); c.alignment = Alignment(horizontal="center" if ci > 0 else "left")
            r += 1
            for _, rd in dff.iterrows():
                for ci, (cn, val) in enumerate(rd.items()):
                    cell = ws.cell(row=r, column=sc + ci, value=val); lb = str(rd.iloc[0]) if ci == 0 else ""
                    if not lb.startswith("  |__"):
                        cell.fill = PatternFill("solid", fgColor=sbc)
                    if isinstance(val, (int, float)) and cn not in ('', 'GESTION'):
                        cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal="right")
                r += 1
            return r - sr

        def wsh(ws, r, c, t, bg):
            cell = ws.cell(row=r, column=c, value=t); cell.font = Font(bold=True, color="FFFFFF", size=11); cell.fill = PatternFill("solid", fgColor=bg); cell.alignment = Alignment(horizontal="left")
            for cc in range(c + 1, c + bdw):
                ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=bg)
        SRH = 6; cg = 2; cf = cg + bdw + GEO_COL_GAP; ch = cf + bdw + GEO_COL_GAP
        for cs, t, co in [(cg, "GLOBAL", CG), (cf, "FRANCE", CF), (ch, "HORS FRANCE + NON TROUVE", CH)]:
            cell = ws2.cell(row=SRH, column=cs, value=t); cell.font = Font(bold=True, color="FFFFFF", size=13); cell.fill = PatternFill("solid", fgColor=co); cell.alignment = Alignment(horizontal="left")
            for cc in range(cs + 1, cs + bdw):
                ws2.cell(row=SRH, column=cc).fill = PatternFill("solid", fgColor=co)
        SC = {"ENTREPRISE": ("004D33", SG, "0D47A1", SF, "7F0000", SH), "BPE": ("005A3C", SG, "1565C0", SF, "8B0000", SH), "CORP+COT+COM": ("006B43", SG, "1976D2", SF, "B71C1C", SH), "NOTE DE FRAIS": ("005A3C", SG, "1A237E", SF, "880E4F", SH), "GESTION": ("00513A", SG, "0A3D8F", SF, "6A0000", SH)}
        cr = SRH + 2
        for sn, dq in [("ENTREPRISE", [peg, pef, peh]), ("BPE", [pbg, pbf, pbh]), ("CORP+COT+COM", [pcg, pcf, pch]), ("NOTE DE FRAIS", [png, pnf, pnh]), ("GESTION", [pgg, pgf, pgh])]:
            sc = SC[sn]; hcg, sg_, hcf, sf_, hch, sh_ = sc
            wsh(ws2, cr, cg, f"  {sn}", CG); wsh(ws2, cr, cf, f"  {sn}", CF); wsh(ws2, cr, ch, f"  {sn}", CH)
            cr += 1
            ru = [wdb(ws2, cr, cg, dq[0], hcg, sg_), wdb(ws2, cr, cf, dq[1], hcf, sf_), wdb(ws2, cr, ch, dq[2], hch, sh_)]
            cr += max(ru) + 2
        for sc in [cg, cf, ch]:
            ws2.column_dimensions[get_column_letter(sc)].width = 28
            for ci in range(1, bdw):
                ws2.column_dimensions[get_column_letter(sc + ci)].width = 14
            for ci in range(bdw, bdw + GEO_COL_GAP):
                ws2.column_dimensions[get_column_letter(sc + ci)].width = 3
        wb.save(save_x)
        self.upd(1.0, "Termine !")

        if len(chunks) == 1:
            detail_info = f"- Sheet DETAIL : {total_rows:,} lignes"
        else:
            detail_info = f"- {len(chunks)} sheets DETAIL_1..DETAIL_{len(chunks)} ({total_rows:,} lignes au total)"

        print("Fichier Excel cree !")
        print(f"  {os.path.basename(save_x)}")
        print(f"  {detail_info}")
        print("  - Sheet TABLEAU (3 zones geo)")
        print("  - REF_SOURCE + SEGMENT_RAW + NOM_GA inclus (v12)")
        print("  - Split DETAIL >1M lignes (v13 [S8XPL])")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="01.S8XPL.py",
        description="MONEXT ANALYZER v13 [S8XPL] - CLI (BNP Paribas Cash Management).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    # Sources requises
    parser.add_argument("--monext", required=False, default=None, type=Path, help="(oblig) Fichier MONEXT consolide (CSV)")
    parser.add_argument("--ref-client", required=False, default=None, type=Path, help="(oblig) Referentiel client (fichier Olivier)")
    parser.add_argument("--parc", required=False, default=None, type=Path, help="(oblig) PARC_CLIENT")
    parser.add_argument("--opti", required=False, default=None, type=Path, help="(oblig) OPTIFLUX")
    parser.add_argument("--yannick", required=False, default=None, type=Path, help="(oblig) Gestion Directe/Indirecte")
    parser.add_argument("--geo", required=False, default=None, type=Path, help="(oblig) BG_LE_RMPM_COUNTRY")
    # Sources optionnelles
    parser.add_argument("--bpe-retail", type=Path, default=None, help="(opt) BPE RETAIL ; active le fallback BPE RETAIL")
    parser.add_argument("--seg-agence", type=Path, default=None, help="(opt) SEGMENT AGENCE ; active le fallback SEG AGENCE")
    # Options de traitement
    parser.add_argument("--cwt-force", action="store_true", help="(opt) Force les lignes 'CWT' en Gestion Directe (en dernier)")
    parser.add_argument("--xlsx", action="store_true", help="(opt) Genere aussi le classeur XLSX agrege")
    # Sortie
    parser.add_argument("--output-dir", required=True, type=Path, help="(oblig) Repertoire de sortie")
    parser.add_argument("--output-filename", required=True, type=str, help="(oblig) Base de nom des fichiers de sortie")
    return parser


def main(argv=None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    # ── Résolution auto des sources obligatoires non fournies (03.sources/) ──
    try:
        if not args.monext:
            args.monext = resolve_source("MONEXT_AGREGE", required=True)
        if not args.ref_client:
            args.ref_client = resolve_source("REFERENTIEL_CLIENT", required=True)
        if not args.parc:
            args.parc = resolve_source("PARC_CLIENT", required=True)
        if not args.opti:
            args.opti = resolve_source("OPTIFLUX", required=True)
        if not args.yannick:
            args.yannick = resolve_source("GA_GESTION_DIRECTE", required=True)
        if not args.geo:
            args.geo = resolve_source("IBAN_ACCOUNT", required=True)
    except FileNotFoundError as exc:
        print(f"[ERREUR] {exc}", file=sys.stderr)
        return 1

    # Validation des chemins d'entree
    required_inputs = {
        "--monext": args.monext, "--ref-client": args.ref_client, "--parc": args.parc,
        "--opti": args.opti, "--yannick": args.yannick, "--geo": args.geo,
    }
    for flag, path in required_inputs.items():
        if not Path(path).is_file():
            print(f"[ERREUR] Fichier introuvable pour {flag} : {path}", file=sys.stderr)
            return 2
    if args.bpe_retail is not None and not Path(args.bpe_retail).is_file():
        print(f"[ERREUR] Fichier introuvable pour --bpe-retail : {args.bpe_retail}", file=sys.stderr)
        return 2
    if args.seg_agence is not None and not Path(args.seg_agence).is_file():
        print(f"[ERREUR] Fichier introuvable pour --seg-agence : {args.seg_agence}", file=sys.stderr)
        return 2

    try:
        analyzer = MonextAnalyzer_S8XPL(args)
        analyzer.run()
    except Exception as exc:  # noqa: BLE001
        print(f"[ERREUR] {exc}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1

    print(f"[OK] MONEXT ANALYZER v13 [{VERSION_ID}] termine. Sorties dans : {args.output_dir}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
