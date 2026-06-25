# -*- coding: utf-8 -*-
"""
SME DATASET ANALYZER [C2SME]
=================================

DESCRIPTION :
  Extraction et analyse du segment SME (Small & Medium Enterprises) a partir du
  fichier Parc Clients, enrichi avec MONEXT (Corporate) et WORLDLINE (Centralise).
  Filtre le segment SME (codes ER, ME par defaut) du Parc Clients, enrichit via
  MONEXT (Corporate) et WORLDLINE (Centralise), realise le matching RP/RC/RS,
  calcule les agregats mensuels et annuels 2025/2026, puis produit un dataset CSV
  et (optionnel) un classeur XLSX (feuille Synthese + feuille Simulation full
  formule avec 4 graphiques openpyxl).

  Version CLI autonome (refactorisation de l'application GUI customtkinter).
  La LOGIQUE METIER est PRESERVEE A L'IDENTIQUE : filtres segment, nettoyage
  vectorise, matching RP/RC/RS, agregats mensuels/annuels et formules XLSX
  sont strictement inchanges (valides par Ali).

FORMULES DE CALCUL INCHANGEES (validees par Ali) :
  FLUX      MX = col 15 + col 17
  PNB       MX = col 19->55, sauf col 33, interchange x -1
  NB_CARTES MX = col 12
  FLUX      WL = col 28 + col 29
  PNB       WL = col 30+31+32+33+34+35+36
  NB_CARTES WL = col 17

SOURCES REQUISES :
  --parc           PARC_CLIENT (source SME)                         [OBLIGATOIRE]
  --monext         MONEXT consolide - CORPORATE (multi-mois)        [OBLIGATOIRE]
  --worldline      WORLDLINE PRGM - CENTRALISE (multi-mois)         [OBLIGATOIRE]
  --bpe-retail     BPE RETAIL (codes agences)                       [OPTIONNEL]

OUTPUTS PRODUITS :
  <output-dir>/<output-filename>.csv    Dataset SME par client (sep ';', utf-8-sig)
  <output-dir>/<output-filename>.xlsx   Classeur (Synthese + Simulation 4 graphiques)
                                        Genere sauf si --no-xlsx.

ARGUMENTS CLI :
  --parc PATH                 (obligatoire) Fichier PARC_CLIENT CSV.
  --monext PATH               (obligatoire) Fichier MONEXT Corporate CSV.
  --worldline PATH            (obligatoire) Fichier WORLDLINE Centralise CSV.
  --bpe-retail PATH           (optionnel)   Fichier BPE RETAIL CSV. Active
                                            l'identification BPE si fourni.
  --segments "ER,ME"          (optionnel)   Codes segment SME (defaut "ER,ME").
  --output-dir PATH           (optionnel)   Dossier de sortie (defaut : .).
  --output-filename NAME      (optionnel)   Base du nom des fichiers de sortie
                                            (sans extension). Defaut :
                                            SME_DATASET_<timestamp>_C2SME.
  --no-xlsx                   (optionnel)   Ne genere pas le classeur XLSX.

DECOMPOSITION :
  main()
   |- argparse : parse des arguments CLI
   |- AnalyseurC2SME(args)
   |   |- worker()
   |   |   |- load_csv_smart()      chargement CSV multi-sep / multi-encodage
   |   |   |- filtrage segment SME  (clean_id, norm_rs, isin segments)
   |   |   |- preparation MONEXT     (flux/pnb/nb_cartes - formules inchangees)
   |   |   |- preparation WORLDLINE  (flux/pnb/nb_cartes - formules inchangees)
   |   |   |- dictionnaires lookup   (RP/RC/RS, mensuels et globaux)
   |   |   |- matching vectorise     (DIRECT_RP/RC, RS_EXACTE, RS_INCLUSION)
   |   |   |- agregats mensuels      (par client, par mois)
   |   |   |- agregats annuels       (2025 / 2026, indicateurs)
   |   |   |- construction dataset   (export CSV)
   |   |   `- generate_xlsx()        (optionnel)
   |   `- generate_xlsx()
   |       |- Feuille Synthese       (donnees annuelles par client)
   |       `- Feuille Simulation     (full formule + 4 graphiques openpyxl)
   `- codes de sortie : 0 (OK), 1 (erreur metier), 2 (erreur arguments)

BNP Paribas Cash Management - Direction Monetique
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd


# ─── Résolution auto des sources + lecture DuckDB (inliné depuis 06.fonctions) ──
def _find_sources_root():
    here = Path(__file__).resolve()
    for d in [here] + list(here.parents):
        c = d / "03.sources"
        if c.is_dir():
            return c
    return None


def resolve_source(name, required=False):
    root = _find_sources_root()
    if not root:
        if required:
            raise FileNotFoundError(f"Dossier 03.sources introuvable depuis {__file__}")
        return None
    folder = next((m for m in root.glob("*/" + name) if m.is_dir()), None)
    if not folder:
        if required:
            raise FileNotFoundError(f"Source introuvable : 03.sources/*/{name}")
        return None
    exts = (".csv", ".xlsx", ".xls", ".txt")
    files = [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() in exts]
    if not files:
        if required:
            raise FileNotFoundError(f"Aucun fichier source dans {folder}")
        return None

    def _np(f):
        m = re.match(r"\s*(\d+)", f.name)
        return (int(m.group(1)) if m else -1, f.name)
    return max(files, key=_np)


def _read_duck(path, nrows=None):
    p = Path(path)
    if nrows is not None or p.suffix.lower() not in (".csv", ".txt", ""):
        return None
    try:
        import duckdb
        return duckdb.read_csv(str(p), all_varchar=True).df()
    except Exception:
        return None


# --- CONFIGURATION ------------------------------------------------------------
VERSION_ID = "C2SME"

MOIS_NOMS = {
    '01': 'JANVIER', '02': 'FEVRIER', '03': 'MARS',
    '04': 'AVRIL',   '05': 'MAI',     '06': 'JUIN',
    '07': 'JUILLET', '08': 'AOUT',    '09': 'SEPTEMBRE',
    '10': 'OCTOBRE', '11': 'NOVEMBRE','12': 'DECEMBRE'
}

DEFAULT_SME_CODES = "ER,ME"

DEFAULT_POSITIONS = {
    'parc_rp': 1, 'parc_rmpm': 6, 'parc_rs': 8, 'parc_segment': 9,
    'parc_code_ga': 11, 'parc_nom_ga': 12, 'parc_rc': 13, 'parc_code_agence': 4,
    'mx_mois': 1, 'mx_rs': 4, 'mx_id_rp': 9, 'mx_id_rc': 10,
    'mx_nb_cartes': 12, 'mx_depenses': 15, 'mx_retraits': 17,
    'mx_pnb_first': 19, 'mx_pnb_last': 55, 'mx_pnb_excl': 33,
    'wl_mois': 2, 'wl_rs': 8, 'wl_id_rc': 40, 'wl_nb_cartes': 17,
    'wl_dep_1': 28, 'wl_dep_2': 29,
    'wl_pnb_cols': [30, 31, 32, 33, 34, 35, 36],
    'bpe_code_agence': 3,
}


def yyyymm_to_label(code: str) -> str:
    if not code or len(code) != 6:
        return code
    return f"{code[:4]}_{MOIS_NOMS.get(code[4:6], code[4:6])}"


# ==============================================================================
class AnalyseurC2SME:

    def __init__(self, args: argparse.Namespace) -> None:
        # Les fichiers d'entree (anciens chemins selectionnes via filedialog).
        self.files: dict[str, str] = {
            "PARC": str(args.parc),
            "MONEXT": str(args.monext),
            "WORLDLINE": str(args.worldline),
            "BPE_RETAIL": str(args.bpe_retail) if args.bpe_retail else "",
        }
        # Activation BPE : equivalent de l'ancienne case a cocher use_bpe_var.
        self.use_bpe: bool = bool(args.bpe_retail)

        # Codes segment SME (anciens self.sme_codes_var de l'UI).
        self.sme_codes: list[str] = [
            c.strip().upper() for c in str(args.segments).split(',') if c.strip()
        ]

        # Sorties (remplacent asksaveasfilename).
        self.output_dir: Path = Path(args.output_dir)
        self.output_filename: str = args.output_filename or ""
        self.generate_xlsx_flag: bool = not args.no_xlsx

        # Caches de previsualisation (utilises pour resoudre le mapping colonnes).
        self.dfs_preview: dict[str, pd.DataFrame] = {}
        self.original_cols: dict[str, list[str]] = {}

    # --------------------------------------------------------------------------
    # CHARGEMENT (inchange)
    # --------------------------------------------------------------------------
    def load_csv_smart(self, path: str, nrows: int | None = None) -> pd.DataFrame:
        _d = _read_duck(path, nrows)
        if _d is not None:
            return _d
        for sep in [';', ',', '\t']:
            for enc in ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']:
                try:
                    df_t = pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                       keep_default_na=False, na_values=[],
                                       on_bad_lines='skip', nrows=5)
                    if df_t.shape[1] > 1:
                        return pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                           keep_default_na=False, na_values=[],
                                           on_bad_lines='skip', nrows=nrows)
                except Exception:
                    continue
        return pd.read_csv(path, sep=None, engine='python', dtype=str,
                           keep_default_na=False, na_values=[], on_bad_lines='skip', nrows=nrows)

    # --------------------------------------------------------------------------
    # RESOLUTION DU MAPPING DE COLONNES
    # Remplacement de l'etape visuelle "ETAPE 2 : Mapping des colonnes".
    # Le mapping et la selection manuelle de colonnes (combobox de l'UI) sont
    # supprimes : on applique directement les positions par defaut de l'UI
    # (DEFAULT_POSITIONS, base 1) resolues contre les colonnes reelles de chaque
    # fichier. Aucun calcul metier n'est modifie.
    # --------------------------------------------------------------------------
    def _resolve_mapping(self) -> tuple[dict[str, str], dict[str, int], list[str]]:
        # Charge les en-tetes reels (equivalent de load_previews).
        keys = ["PARC", "MONEXT", "WORLDLINE"]
        if self.use_bpe:
            keys.append("BPE_RETAIL")
        for k in keys:
            self.dfs_preview[k] = self.load_csv_smart(self.files[k], nrows=5)
            self.original_cols[k] = list(self.dfs_preview[k].columns)

        def col_at(fkey: str, pos1: int) -> str:
            cols = self.original_cols[fkey]
            if pos1 < 1 or pos1 > len(cols):
                raise ValueError(
                    f"Position {pos1} hors plage pour le fichier {fkey} "
                    f"({len(cols)} colonnes)."
                )
            return cols[pos1 - 1]

        # m : noms de colonnes reels selectionnes (defauts UI), comme dans l'UI.
        parc_fields = [
            'parc_rp', 'parc_rmpm', 'parc_rs', 'parc_segment',
            'parc_code_ga', 'parc_nom_ga', 'parc_rc', 'parc_code_agence',
        ]
        mx_fields = [
            'mx_mois', 'mx_rs', 'mx_id_rp', 'mx_id_rc',
            'mx_nb_cartes', 'mx_depenses', 'mx_retraits',
        ]
        wl_fields = [
            'wl_mois', 'wl_rs', 'wl_id_rc', 'wl_nb_cartes',
            'wl_dep_1', 'wl_dep_2',
        ]

        m: dict[str, str] = {}
        for f in parc_fields:
            m[f] = col_at("PARC", DEFAULT_POSITIONS[f])
        for f in mx_fields:
            m[f] = col_at("MONEXT", DEFAULT_POSITIONS[f])
        for f in wl_fields:
            m[f] = col_at("WORLDLINE", DEFAULT_POSITIONS[f])
        if self.use_bpe:
            m['bpe_code_agence'] = col_at("BPE_RETAIL", DEFAULT_POSITIONS['bpe_code_agence'])

        # Configuration PNB MONEXT (defauts UI : plage 19->55, exclusion 33,
        # interchange "(Aucune)" -> 0).
        mx_pnb_cfg = {
            'first': DEFAULT_POSITIONS['mx_pnb_first'],
            'last': DEFAULT_POSITIONS['mx_pnb_last'],
            'excl': DEFAULT_POSITIONS['mx_pnb_excl'],
            'interchange': 0,
        }

        # Colonnes PNB WORLDLINE (defauts UI : cols 30..36, par nom).
        wl_cols = self.original_cols["WORLDLINE"]
        wl_pnb_cols = [wl_cols[p - 1] for p in DEFAULT_POSITIONS['wl_pnb_cols']
                       if 1 <= p <= len(wl_cols)]

        return m, mx_pnb_cfg, wl_pnb_cols

    # --------------------------------------------------------------------------
    # NETTOYAGE VECTORISE (inchange - valide)
    # --------------------------------------------------------------------------
    @staticmethod
    def clean_id(series: pd.Series) -> pd.Series:
        s = series.astype(str).str.strip()
        s = s.replace(['', 'nan', 'NaN', 'None', 'NULL', 'NA', 'N/A', 'NAN', 'NONE'], '')
        mask_excel = s.str.startswith('="') & s.str.endswith('"')
        s = s.where(~mask_excel, s.str[2:-1])
        s = s.str.lstrip("'")
        mask_dec = s.str.endswith('.0') & s.str[:-2].str.isdigit()
        s = s.where(~mask_dec, s.str[:-2])
        return s.str.strip()

    @staticmethod
    def clean_id_strip0(series: pd.Series) -> pd.Series:
        s = AnalyseurC2SME.clean_id(series)
        stripped = s.str.lstrip('0')
        return stripped.where(stripped != '', s)

    @staticmethod
    def norm_rs(series: pd.Series) -> pd.Series:
        def _n(val: Any) -> str:
            if pd.isna(val) or str(val).strip() == '': return ''
            s = str(val).strip().upper()
            s = unicodedata.normalize('NFD', s)
            return ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return series.apply(_n)

    @staticmethod
    def to_float(series: pd.Series) -> pd.Series:
        s = series.astype(str)
        s = s.str.replace('"', '', regex=False).str.replace("'", '', regex=False)
        s = s.str.replace(' ', '', regex=False).str.replace('\xa0', '', regex=False)
        s = s.str.replace(' ', '', regex=False).str.replace('€', '', regex=False)
        mask_minus = s.str.endswith('-')
        s = s.where(~mask_minus, '-' + s.str[:-1])
        s = s.str.replace(',', '.', regex=False)
        return pd.to_numeric(s, errors='coerce').fillna(0.0)

    @staticmethod
    def parse_mois(series: pd.Series) -> pd.Series:
        def _p(val: Any) -> str:
            if pd.isna(val): return ''
            s = str(val).strip()
            if s.startswith('="') and s.endswith('"'): s = s[2:-1].strip()
            s = s.lstrip("'").strip()
            if s.endswith('.0') and s[:-2].isdigit(): s = s[:-2]
            if re.fullmatch(r'\d{6}', s): return s
            m = re.match(r'^(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})$', s)
            if m:
                mo = int(m.group(2)); an = m.group(3)
                if 1 <= mo <= 12: return f"{an}{str(mo).zfill(2)}"
            m = re.match(r'^(\d{1,2})[/\-\.](\d{4})$', s)
            if m:
                mo = int(m.group(1)); an = m.group(2)
                if 1 <= mo <= 12: return f"{an}{str(mo).zfill(2)}"
            try:
                dt = pd.to_datetime(s, dayfirst=True, errors='coerce')
                if pd.notna(dt): return dt.strftime('%Y%m')
            except Exception: pass
            return ''
        return series.apply(_p)

    @staticmethod
    def fmt(val: float) -> str:
        return str(round(val, 2)).replace('.', ',')

    # --------------------------------------------------------------------------
    # PROGRESSION (remplace l'ancien _prog GUI : barre + label -> print)
    # --------------------------------------------------------------------------
    def _prog(self, val: float, txt: str) -> None:
        print(f"[{val * 100:5.1f}%] {txt}")

    # --------------------------------------------------------------------------
    # WORKER (formules inchangees - validees)
    # --------------------------------------------------------------------------
    def worker(self, m: dict[str, str], mx_pnb_cfg: dict[str, int],
               wl_pnb_cols: list[str], sme_codes: list[str]) -> None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        self._prog(0.02, "Chargement PARC_CLIENT...")
        df_parc = self.load_csv_smart(self.files["PARC"])
        self._prog(0.04, "Chargement MONEXT Corporate...")
        df_mx = self.load_csv_smart(self.files["MONEXT"])
        self._prog(0.06, "Chargement WORLDLINE Centralise...")
        df_wl = self.load_csv_smart(self.files["WORLDLINE"])

        set_bpe: set = set()
        if self.use_bpe and self.files["BPE_RETAIL"]:
            self._prog(0.07, "Chargement BPE RETAIL...")
            df_bpe  = self.load_csv_smart(self.files["BPE_RETAIL"])
            bpe_col = m.get('bpe_code_agence', df_bpe.columns[2])
            set_bpe = set(self.clean_id_strip0(df_bpe[bpe_col]).replace('', pd.NA).dropna().unique())

        self._prog(0.10, "Filtrage clients SME...")
        df_parc['_K_RP']      = self.clean_id_strip0(df_parc[m['parc_rp']])
        df_parc['_K_RC']      = self.clean_id_strip0(df_parc[m['parc_rc']])
        df_parc['_K_RMPM']    = self.clean_id(df_parc[m['parc_rmpm']])
        df_parc['_K_RS']      = self.norm_rs(df_parc[m['parc_rs']])
        df_parc['_K_SEGMENT'] = df_parc[m['parc_segment']].astype(str).str.strip().str.upper()
        df_parc['_K_CODE_AG'] = self.clean_id_strip0(df_parc[m['parc_code_agence']])

        mask_sme = df_parc['_K_SEGMENT'].isin(sme_codes)
        df_sme   = df_parc[mask_sme].copy().reset_index(drop=True)
        n_sme    = len(df_sme)

        if n_sme == 0:
            raise RuntimeError(
                f"Aucun client trouve avec les codes : {', '.join(sme_codes)}"
            )

        df_sme['CATEGORIE'] = np.where(
            df_sme['_K_CODE_AG'].isin(set_bpe), 'BPE', 'ENTREPRISE'
        ) if set_bpe else 'ENTREPRISE'

        # -- MONEXT - formules inchangees ------------------------------------
        self._prog(0.15, "Preparation MONEXT Corporate...")
        df_mx['_K_RP']      = self.clean_id_strip0(df_mx[m['mx_id_rp']])
        df_mx['_K_RC']      = self.clean_id_strip0(df_mx[m['mx_id_rc']])
        df_mx['_K_RS']      = self.norm_rs(df_mx[m['mx_rs']])
        df_mx['_MOIS']      = self.parse_mois(df_mx[m['mx_mois']])
        df_mx['_NB_CARTES'] = self.to_float(df_mx[m['mx_nb_cartes']])
        df_mx['_DEPENSES']  = self.to_float(df_mx[m['mx_depenses']])
        df_mx['_RETRAITS']  = self.to_float(df_mx[m['mx_retraits']])
        df_mx['_FLUX']      = df_mx['_DEPENSES'] + df_mx['_RETRAITS']

        mx_cols     = list(df_mx.columns)
        f0          = mx_pnb_cfg['first'] - 1
        l0          = mx_pnb_cfg['last']  - 1
        e0          = mx_pnb_cfg['excl']  - 1 if mx_pnb_cfg['excl'] > 0 else -1
        ic0         = mx_pnb_cfg['interchange'] - 1 if mx_pnb_cfg['interchange'] > 0 else -1
        pnb_range   = mx_cols[f0:l0+1]
        excl_name   = mx_cols[e0]  if 0 <= e0  < len(mx_cols) else None
        ic_col_name = mx_cols[ic0] if 0 <= ic0 < len(mx_cols) else None
        for c in pnb_range:
            if c in df_mx.columns: df_mx[c] = self.to_float(df_mx[c])
        if ic_col_name and ic_col_name in df_mx.columns:
            df_mx[ic_col_name] = df_mx[ic_col_name] * -1
        pnb_filtered = [c for c in pnb_range if c != excl_name]
        df_mx['_PNB'] = df_mx[pnb_filtered].sum(axis=1) if pnb_filtered else 0.0

        # -- WORLDLINE - formules inchangees ---------------------------------
        self._prog(0.20, "Preparation WORLDLINE Centralise...")
        df_wl['_K_RC']      = self.clean_id_strip0(df_wl[m['wl_id_rc']])
        df_wl['_K_RS']      = self.norm_rs(df_wl[m['wl_rs']])
        df_wl['_MOIS']      = self.parse_mois(df_wl[m['wl_mois']])
        df_wl['_NB_CARTES'] = self.to_float(df_wl[m['wl_nb_cartes']])
        df_wl['_DEP1']      = self.to_float(df_wl[m['wl_dep_1']])
        df_wl['_DEP2']      = self.to_float(df_wl[m['wl_dep_2']])
        df_wl['_FLUX']      = df_wl['_DEP1'] + df_wl['_DEP2']
        if wl_pnb_cols:
            for c in wl_pnb_cols: df_wl[c] = self.to_float(df_wl[c])
            df_wl['_PNB'] = df_wl[wl_pnb_cols].sum(axis=1)
        else:
            df_wl['_PNB'] = 0.0

        all_mois  = sorted(set(df_mx['_MOIS'].unique()) | set(df_wl['_MOIS'].unique()) - {''})
        all_mois  = [mo for mo in all_mois if mo and len(mo) == 6]
        mois_2025 = [mo for mo in all_mois if mo.startswith('2025')]
        mois_2026 = [mo for mo in all_mois if mo.startswith('2026')]

        self._prog(0.25, "Construction dictionnaires de lookup...")
        agg = {'_FLUX': 'sum', '_PNB': 'sum', '_NB_CARTES': 'sum'}
        mx_rp_g = df_mx[df_mx['_K_RP'] != ''].groupby('_K_RP').agg(agg).T.to_dict('list')
        mx_rc_g = df_mx[df_mx['_K_RC'] != ''].groupby('_K_RC').agg(agg).T.to_dict('list')
        mx_rs_g = df_mx[df_mx['_K_RS'] != ''].groupby('_K_RS').agg(agg).T.to_dict('list')
        wl_rc_g = df_wl[df_wl['_K_RC'] != ''].groupby('_K_RC').agg(agg).T.to_dict('list')
        wl_rs_g = df_wl[df_wl['_K_RS'] != ''].groupby('_K_RS').agg(agg).T.to_dict('list')

        self._prog(0.30, f"Calcul agregats mensuels ({len(all_mois)} mois)...")
        mx_mo_rp: dict = {}; mx_mo_rc: dict = {}; mx_mo_rs: dict = {}
        wl_mo_rc: dict = {}; wl_mo_rs: dict = {}
        for i_mo, mo in enumerate(all_mois):
            if i_mo % 2000 == 0:
                self._prog(0.30 + 0.10 * i_mo / max(len(all_mois), 1),
                           f"Agregats mensuels {i_mo}/{len(all_mois)}...")
            dmx = df_mx[df_mx['_MOIS'] == mo]
            dwl = df_wl[df_wl['_MOIS'] == mo]
            mx_mo_rp[mo] = dmx[dmx['_K_RP'] != ''].groupby('_K_RP').agg(agg).T.to_dict('list')
            mx_mo_rc[mo] = dmx[dmx['_K_RC'] != ''].groupby('_K_RC').agg(agg).T.to_dict('list')
            mx_mo_rs[mo] = dmx[dmx['_K_RS'] != ''].groupby('_K_RS').agg(agg).T.to_dict('list')
            wl_mo_rc[mo] = dwl[dwl['_K_RC'] != ''].groupby('_K_RC').agg(agg).T.to_dict('list')
            wl_mo_rs[mo] = dwl[dwl['_K_RS'] != ''].groupby('_K_RS').agg(agg).T.to_dict('list')

        RS_MIN = 14

        self._prog(0.42, f"Matching vectorise {n_sme:,} clients SME...")
        arr_rp = df_sme['_K_RP'].values
        arr_rc = df_sme['_K_RC'].values
        arr_rs = df_sme['_K_RS'].values

        found_mx  = np.full(n_sme, "NO",  dtype=object)
        method_mx = np.full(n_sme, "N/A", dtype=object)
        found_wl  = np.full(n_sme, "NO",  dtype=object)
        method_wl = np.full(n_sme, "N/A", dtype=object)
        key_rp_mx = np.full(n_sme, "", dtype=object)
        key_rc_mx = np.full(n_sme, "", dtype=object)
        key_rc_wl = np.full(n_sme, "", dtype=object)
        key_rs_mx = np.full(n_sme, "", dtype=object)
        key_rs_wl = np.full(n_sme, "", dtype=object)

        s_rp = pd.Series(arr_rp); s_rc = pd.Series(arr_rc)
        mask = (s_rp != '') & s_rp.isin(mx_rp_g)
        if mask.any():
            idx = mask.values; found_mx[idx] = "YES"; method_mx[idx] = "DIRECT_RP"
            key_rp_mx[idx] = s_rp[mask].values
        mask = (found_mx == "NO") & (s_rc != '') & s_rc.isin(mx_rc_g)
        if mask.any():
            idx = mask.values; found_mx[idx] = "YES"; method_mx[idx] = "DIRECT_RC"
            key_rc_mx[idx] = s_rc[mask].values
        mask = (s_rc != '') & s_rc.isin(wl_rc_g)
        if mask.any():
            idx = mask.values; found_wl[idx] = "YES"; method_wl[idx] = "DIRECT_RC"
            key_rc_wl[idx] = s_rc[mask].values

        residual   = np.where((found_mx == "NO") | (found_wl == "NO"))[0]
        n_res      = len(residual)
        rs_incl_mx = [(k, v) for k, v in mx_rs_g.items() if len(k) >= RS_MIN]
        rs_incl_wl = [(k, v) for k, v in wl_rs_g.items() if len(k) >= RS_MIN]

        for loop_i, i in enumerate(residual):
            if loop_i % 2000 == 0:
                self._prog(0.42 + 0.20 * loop_i / max(n_res, 1),
                           f"Matching residuel {loop_i:,}/{n_res:,}...")
            rs = arr_rs[i]
            if not rs: continue
            if found_mx[i] == "NO":
                if rs in mx_rs_g:
                    found_mx[i] = "YES"; method_mx[i] = "RS_EXACTE"; key_rs_mx[i] = rs
                elif len(rs) >= RS_MIN:
                    for (k, _) in rs_incl_mx:
                        if rs in k or k in rs:
                            found_mx[i] = "YES"; method_mx[i] = "RS_INCLUSION"
                            key_rs_mx[i] = k; break
            if found_wl[i] == "NO":
                if rs in wl_rs_g:
                    found_wl[i] = "YES"; method_wl[i] = "RS_EXACTE"; key_rs_wl[i] = rs
                elif len(rs) >= RS_MIN:
                    for (k, _) in rs_incl_wl:
                        if rs in k or k in rs:
                            found_wl[i] = "YES"; method_wl[i] = "RS_INCLUSION"
                            key_rs_wl[i] = k; break

        n_mois = len(all_mois)
        mx_flux_mo: dict = {}; mx_pnb_mo: dict = {}; mx_nb_mo: dict = {}
        wl_flux_mo: dict = {}; wl_pnb_mo: dict = {}; wl_nb_mo: dict = {}

        for i_mo, mo in enumerate(all_mois):
            if i_mo % 2000 == 0:
                self._prog(0.63 + 0.15 * i_mo / max(n_mois, 1),
                           f"Donnees mensuelles {i_mo}/{n_mois}...")
            f_mx = np.zeros(n_sme); p_mx = np.zeros(n_sme); n_mx = np.zeros(n_sme)
            f_wl = np.zeros(n_sme); p_wl = np.zeros(n_sme); n_wl = np.zeros(n_sme)
            d_rp = mx_mo_rp.get(mo, {}); d_rc_mx = mx_mo_rc.get(mo, {}); d_rs_mx = mx_mo_rs.get(mo, {})
            d_rc_wl = wl_mo_rc.get(mo, {}); d_rs_wl = wl_mo_rs.get(mo, {})

            for i in range(n_sme):
                k = key_rp_mx[i]
                if k and k in d_rp:   f_mx[i], p_mx[i], n_mx[i] = d_rp[k]
                elif (k2 := key_rc_mx[i]) and k2 in d_rc_mx: f_mx[i], p_mx[i], n_mx[i] = d_rc_mx[k2]
                elif (k3 := key_rs_mx[i]) and k3 in d_rs_mx: f_mx[i], p_mx[i], n_mx[i] = d_rs_mx[k3]
                k4 = key_rc_wl[i]
                if k4 and k4 in d_rc_wl:   f_wl[i], p_wl[i], n_wl[i] = d_rc_wl[k4]
                elif (k5 := key_rs_wl[i]) and k5 in d_rs_wl: f_wl[i], p_wl[i], n_wl[i] = d_rs_wl[k5]

            mx_flux_mo[mo] = f_mx; mx_pnb_mo[mo] = p_mx; mx_nb_mo[mo] = n_mx
            wl_flux_mo[mo] = f_wl; wl_pnb_mo[mo] = p_wl; wl_nb_mo[mo] = n_wl

        self._prog(0.80, "Calcul agregats annuels et indicateurs...")

        def sum_mo(d: dict, liste: list[str]) -> np.ndarray:
            return sum(d.get(mo, np.zeros(n_sme)) for mo in liste)

        flux_corp_25 = sum_mo(mx_flux_mo, mois_2025); pnb_corp_25 = sum_mo(mx_pnb_mo, mois_2025); nb_corp_25 = sum_mo(mx_nb_mo, mois_2025)
        flux_cent_25 = sum_mo(wl_flux_mo, mois_2025); pnb_cent_25 = sum_mo(wl_pnb_mo, mois_2025); nb_cent_25 = sum_mo(wl_nb_mo, mois_2025)
        flux_corp_26 = sum_mo(mx_flux_mo, mois_2026); pnb_corp_26 = sum_mo(mx_pnb_mo, mois_2026); nb_corp_26 = sum_mo(mx_nb_mo, mois_2026)
        flux_cent_26 = sum_mo(wl_flux_mo, mois_2026); pnb_cent_26 = sum_mo(wl_pnb_mo, mois_2026); nb_cent_26 = sum_mo(wl_nb_mo, mois_2026)

        flux_tot_25 = flux_corp_25 + flux_cent_25
        pnb_tot_25  = pnb_corp_25  + pnb_cent_25
        nb_tot_25   = nb_corp_25   + nb_cent_25

        n_mois_25           = max(len(mois_2025), 1)
        flux_moy_par_carte  = np.where(nb_tot_25   > 0, flux_tot_25 / nb_tot_25, 0.0)
        taux_pnb_25         = np.where(flux_tot_25 > 0, pnb_tot_25  / flux_tot_25, 0.0)
        pnb_moy_par_carte   = np.where(nb_tot_25   > 0, pnb_tot_25  / nb_tot_25, 0.0)
        flux_mensuel_moy_25 = flux_tot_25 / n_mois_25
        nb_cartes_moy_25    = nb_tot_25   / n_mois_25

        self._prog(0.84, "Construction dataset final...")
        result = df_sme[[
            m['parc_rmpm'], m['parc_rp'], m['parc_rc'],
            m['parc_rs'], m['parc_segment'],
            m['parc_code_ga'], m['parc_nom_ga']
        ]].copy()
        result.columns = ['RMPM','ID_RP','ID_RC','RAISON_SOCIALE','SEGMENT','CODE_GA','NOM_GA']
        result['CATEGORIE']        = df_sme['CATEGORIE'].values
        result['FOUND_CORPORATE']  = found_mx
        result['FOUND_CENTRALISE'] = found_wl

        for mo in all_mois:
            lbl = yyyymm_to_label(mo)
            result[f'FLUX_CORPORATE_{lbl}']      = [self.fmt(v) for v in mx_flux_mo[mo]]
            result[f'PNB_CORPORATE_{lbl}']       = [self.fmt(v) for v in mx_pnb_mo[mo]]
            result[f'NB_CARTES_CORPORATE_{lbl}'] = [self.fmt(v) for v in mx_nb_mo[mo]]
            result[f'FLUX_CENTRALISE_{lbl}']      = [self.fmt(v) for v in wl_flux_mo[mo]]
            result[f'PNB_CENTRALISE_{lbl}']       = [self.fmt(v) for v in wl_pnb_mo[mo]]
            result[f'NB_CARTES_CENTRALISE_{lbl}'] = [self.fmt(v) for v in wl_nb_mo[mo]]

        for col, vals in [
            ('FLUX_CORPORATE_2026', flux_corp_26), ('PNB_CORPORATE_2026', pnb_corp_26), ('NB_CARTES_CORPORATE_2026', nb_corp_26),
            ('FLUX_CENTRALISE_2026', flux_cent_26), ('PNB_CENTRALISE_2026', pnb_cent_26), ('NB_CARTES_CENTRALISE_2026', nb_cent_26),
            ('FLUX_CORPORATE_2025', flux_corp_25), ('PNB_CORPORATE_2025', pnb_corp_25), ('NB_CARTES_CORPORATE_2025', nb_corp_25),
            ('FLUX_CENTRALISE_2025', flux_cent_25), ('PNB_CENTRALISE_2025', pnb_cent_25), ('NB_CARTES_CENTRALISE_2025', nb_cent_25),
            ('FLUX_MENSUEL_MOYEN_2025', flux_mensuel_moy_25), ('NB_CARTES_MENSUEL_MOYEN_2025', nb_cartes_moy_25),
            ('FLUX_MOY_PAR_CARTE_2025', flux_moy_par_carte), ('TAUX_PNB_2025', taux_pnb_25),
            ('PNB_MOY_PAR_CARTE_2025', pnb_moy_par_carte),
        ]:
            if 'TAUX_PNB' in col:
                result[col] = [self.fmt(v * 100) + '%' for v in vals]
            else:
                result[col] = [self.fmt(v) for v in vals]

        self._prog(0.88, "Sauvegarde CSV dataset...")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        base = self.output_filename or f"SME_DATASET_{ts}_{VERSION_ID}"
        save_csv = self.output_dir / f"{base}.csv"
        result.to_csv(save_csv, sep=';', index=False, encoding='utf-8-sig')
        self._prog(0.92, "CSV sauvegarde.")

        n_mx_ok   = int((found_mx == 'YES').sum())
        n_wl_ok   = int((found_wl == 'YES').sum())
        n_both_no = int(((found_mx == 'NO') & (found_wl == 'NO')).sum())

        print("")
        print(f"Dataset SME genere : {save_csv.name}")
        print(f"PERIMETRE : {n_sme:,} clients SME ({', '.join(sme_codes)})")
        print(f"MOIS : {len(all_mois)} | 2025 : {len(mois_2025)} | 2026 : {len(mois_2026)}")
        print("MATCHING :")
        print(f"  Corporate  (MX) : {n_mx_ok:,} / {n_sme:,} ({100*n_mx_ok/n_sme:.1f}%)")
        print(f"  Centralise (WL) : {n_wl_ok:,} / {n_sme:,} ({100*n_wl_ok/n_sme:.1f}%)")
        print(f"  Non trouves (aucun) : {n_both_no:,}")
        print("TOTAUX 2025 :")
        print(f"  Flux total : {float(flux_tot_25.sum()):,.0f} EUR")
        print(f"  PNB total  : {float(pnb_tot_25.sum()):,.0f} EUR")
        print(f"  Nb cartes  : {float(nb_tot_25.sum()):,.0f}")
        print("")

        if self.generate_xlsx_flag:
            self.generate_xlsx(
                result, n_sme, sme_codes, all_mois, mois_2025, mois_2026,
                float(flux_corp_25.sum()), float(pnb_corp_25.sum()), float(nb_corp_25.sum()),
                float(flux_cent_25.sum()), float(pnb_cent_25.sum()), float(nb_cent_25.sum()),
                float(flux_corp_26.sum()), float(pnb_corp_26.sum()), float(nb_corp_26.sum()),
                float(flux_cent_26.sum()), float(pnb_cent_26.sum()), float(nb_cent_26.sum()),
                float(flux_tot_25.sum()),  float(pnb_tot_25.sum()),  float(nb_tot_25.sum()),
                ts
            )
        else:
            self._prog(1.0, "Termine (CSV uniquement).")

    # --------------------------------------------------------------------------
    # GENERATION XLSX
    # --------------------------------------------------------------------------
    def generate_xlsx(self, result: pd.DataFrame, n_sme: int, sme_codes: list[str],
                      all_mois: list[str], mois_2025: list[str], mois_2026: list[str],
                      REF_CORP_FLUX: float, REF_CORP_PNB: float, REF_CORP_NB: float,
                      REF_CENT_FLUX: float, REF_CENT_PNB: float, REF_CENT_NB: float,
                      REF_CORP_26: float, REF_PNB_CORP_26: float, REF_NB_CORP_26: float,
                      REF_CENT_26: float, REF_PNB_CENT_26: float, REF_NB_CENT_26: float,
                      REF_TOT_FLUX: float, REF_TOT_PNB: float, REF_TOT_NB: float,
                      ts: str) -> None:

        base = self.output_filename or f"SME_DATASET_{ts}_{VERSION_ID}"
        save_xlsx = self.output_dir / f"{base}.xlsx"

        self._prog(0.94, "Generation XLSX...")

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference, PieChart
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.chart.label import DataLabel

        wb = Workbook()

        # Palette couleurs BNP
        GRN  = "00915A"; GRN2 = "E8F5E9"; DARK = "1C3A2D"
        GREY = "F5F5F5"; WHT  = "FFFFFF"; BLU  = "1565C0"
        BLU2 = "E3F2FD"; ORG  = "E65100"; ORG2 = "FFF3E0"
        YLW  = "FFFF00"; GRY2 = "607D8B"

        def fill(h):
            return PatternFill(start_color=h, end_color=h, fill_type='solid')
        def fnt(c="FFFFFF", bold=True, sz=10):
            return Font(name='Segoe UI', size=sz, bold=bold, color=c)
        thin = Side(style='thin', color='CCCCCC')
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

        def C(n): return get_column_letter(n)

        # -----------------------------------------------------------------
        # UTILITAIRES GRAPHIQUES
        # -----------------------------------------------------------------
        def style_chart(chart, title, w=14, h=9):
            """Applique un style commun a tous les graphiques."""
            chart.title    = title
            chart.width    = w
            chart.height   = h
            # Pas de grillage
            chart.plot_area.graphicalProperties = None
            # Legende en bas
            chart.legend.position = 'b'
            chart.legend.overlay  = False
            # Pas de titre d'axe superpose - on garde des titres courts
            return chart

        def no_grid(chart):
            """Supprime le grillage major/minor sur les deux axes."""
            try:
                chart.y_axis.majorGridlines = None
                chart.y_axis.minorGridlines = None
                chart.x_axis.majorGridlines = None
                chart.x_axis.minorGridlines = None
            except Exception: pass
            return chart

        # =================================================================
        # FEUILLE 1 : SYNTHESE (inchangee)
        # =================================================================
        ws1 = wb.active
        ws1.title = "Synthèse"

        ws1.merge_cells("A1:Q1")
        t = ws1["A1"]
        t.value = (f"SME — Synthèse par client | Segments : {', '.join(sme_codes)} | "
                   f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        t.font = fnt(GRN, sz=13)
        t.alignment = Alignment(horizontal='left', vertical='center')
        ws1.row_dimensions[1].height = 28

        ws1["A2"].value = (f"{n_sme:,} clients | {len(mois_2025)} mois 2025 | "
                           f"{len(mois_2026)} mois 2026 | Indicateurs calculés sur 2025")
        ws1["A2"].font = fnt("666666", bold=False, sz=9)

        synth_cols = [
            'RMPM','ID_RP','ID_RC','RAISON_SOCIALE','SEGMENT',
            'CODE_GA','NOM_GA','CATEGORIE',
            'FOUND_CORPORATE','FOUND_CENTRALISE',
            'FLUX_CORPORATE_2026','PNB_CORPORATE_2026','NB_CARTES_CORPORATE_2026',
            'FLUX_CENTRALISE_2026','PNB_CENTRALISE_2026','NB_CARTES_CENTRALISE_2026',
            'FLUX_CORPORATE_2025','PNB_CORPORATE_2025','NB_CARTES_CORPORATE_2025',
            'FLUX_CENTRALISE_2025','PNB_CENTRALISE_2025','NB_CARTES_CENTRALISE_2025',
            'FLUX_MENSUEL_MOYEN_2025','NB_CARTES_MENSUEL_MOYEN_2025',
            'FLUX_MOY_PAR_CARTE_2025','TAUX_PNB_2025','PNB_MOY_PAR_CARTE_2025',
        ]
        synth_cols = [c for c in synth_cols if c in result.columns]
        df_synth   = result[synth_cols].copy()

        def hdr_fill(c):
            if c in ['RMPM','ID_RP','ID_RC','RAISON_SOCIALE','SEGMENT',
                      'CODE_GA','NOM_GA','CATEGORIE','FOUND_CORPORATE','FOUND_CENTRALISE']:
                return DARK
            if '2026' in c: return BLU
            if c in ['FLUX_MENSUEL_MOYEN_2025','NB_CARTES_MENSUEL_MOYEN_2025',
                      'FLUX_MOY_PAR_CARTE_2025','TAUX_PNB_2025','PNB_MOY_PAR_CARTE_2025']:
                return ORG
            return GRN

        HDR = 4
        for ci, c in enumerate(synth_cols, start=1):
            cell = ws1.cell(row=HDR, column=ci, value=c)
            cell.font = fnt(); cell.fill = fill(hdr_fill(c))
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = brd
        ws1.row_dimensions[HDR].height = 40

        for ri, row_data in enumerate(df_synth.itertuples(index=False), start=HDR+1):
            for ci, val in enumerate(row_data, start=1):
                cell = ws1.cell(row=ri, column=ci, value=val)
                cell.font   = Font(name='Segoe UI', size=9)
                cell.fill   = fill(GREY) if ri % 2 == 0 else fill(WHT)
                cell.border = brd
                cell.alignment = Alignment(horizontal='center')

        col_w = {'RMPM':14,'ID_RP':14,'ID_RC':14,'RAISON_SOCIALE':35,
                 'SEGMENT':10,'CODE_GA':12,'NOM_GA':28,'CATEGORIE':14,
                 'FOUND_CORPORATE':16,'FOUND_CENTRALISE':18}
        for ci, c in enumerate(synth_cols, start=1):
            ws1.column_dimensions[C(ci)].width = col_w.get(c, 20)
        ws1.freeze_panes = f"D{HDR+1}"

        agg_r = HDR + len(df_synth) + 3
        ws1.cell(row=agg_r, column=1, value="TOTAUX GLOBAUX SME 2025").font = fnt(GRN, sz=11)
        for dr, (lbl, val) in enumerate([
            ("Flux Corporate 2025",  REF_CORP_FLUX),  ("Flux Centralisé 2025", REF_CENT_FLUX),
            ("Flux Total 2025",      REF_TOT_FLUX),   ("PNB Corporate 2025",   REF_CORP_PNB),
            ("PNB Centralisé 2025",  REF_CENT_PNB),   ("PNB Total 2025",       REF_TOT_PNB),
            ("Nb cartes Corp. 2025", REF_CORP_NB),    ("Nb cartes Cent. 2025", REF_CENT_NB),
            ("Nb cartes Total 2025", REF_TOT_NB),
            ("Taux PNB moyen 2025 (%)", round(REF_TOT_PNB/REF_TOT_FLUX*100 if REF_TOT_FLUX else 0, 3)),
            ("Flux / carte 2025",    round(REF_TOT_FLUX/REF_TOT_NB if REF_TOT_NB else 0, 2)),
            ("PNB / carte 2025",     round(REF_TOT_PNB/REF_TOT_NB  if REF_TOT_NB else 0, 2)),
        ], start=agg_r+1):
            c1 = ws1.cell(row=dr, column=1, value=lbl)
            c2 = ws1.cell(row=dr, column=2, value=round(val, 4))
            c1.font = Font(name='Segoe UI', size=9, bold=True)
            c2.font = Font(name='Segoe UI', size=9)
            c1.fill = fill(GRN2); c2.fill = fill(GRN2)
            c1.border = brd; c2.border = brd

        # =================================================================
        # FEUILLE 2 : SIMULATION - FULL FORMULE + 4 GRAPHIQUES
        # =================================================================
        ws2 = wb.create_sheet("Simulation")

        for ci in range(1, 12):
            ws2.column_dimensions[C(ci)].width = 24

        # Titre
        ws2.merge_cells("A1:J1")
        t2 = ws2["A1"]
        t2.value = "SME — Dashboard de Simulation Croissance"
        t2.font  = fnt(GRN, sz=14)
        t2.alignment = Alignment(horizontal='left', vertical='center')
        ws2.row_dimensions[1].height = 32

        ws2["A2"].value = ("Cellules jaunes = paramètres. "
                           "Toutes les autres valeurs sont calculées automatiquement. "
                           "Valeurs négatives autorisées.")
        ws2["A2"].font = Font(name='Segoe UI', sz=9, color="666666", italic=True)

        # -- Tableau de reference 2025 -------------------------------------
        r = 4
        ws2.cell(row=r, column=1, value="DONNÉES DE RÉFÉRENCE 2025").font = fnt(GRN, sz=11)
        r += 1

        for ci, h in enumerate(["", "CORPORATE (MX)", "CENTRALISÉ (WL)", "TOTAL"], start=1):
            c = ws2.cell(row=r, column=ci, value=h)
            c.font = fnt()
            c.fill = fill(DARK if ci==1 else GRN if ci==2 else BLU if ci==3 else DARK)
            c.border = brd; c.alignment = Alignment(horizontal='center')
        r += 1

        # Lignes de reference - on memorise les adresses
        addr: dict[str, str] = {}
        row_nb   = r; row_flux = r+1; row_pnb  = r+2
        row_tx   = r+3; row_fxct = r+4

        addr.update({
            'NB_CORP': f"B{row_nb}",   'NB_CENT': f"C{row_nb}",   'NB_TOT': f"D{row_nb}",
            'FLUX_CORP': f"B{row_flux}", 'FLUX_CENT': f"C{row_flux}", 'FLUX_TOT': f"D{row_flux}",
            'PNB_CORP': f"B{row_pnb}",  'PNB_CENT': f"C{row_pnb}",  'PNB_TOT': f"D{row_pnb}",
            'TX_CORP': f"B{row_tx}",   'TX_CENT': f"C{row_tx}",   'TX_TOT': f"D{row_tx}",
            'FXCT_CORP': f"B{row_fxct}", 'FXCT_CENT': f"C{row_fxct}", 'FXCT_TOT': f"D{row_fxct}",
        })

        ref_rows_data = [
            ("NB_CARTES_ANNEE_2025", REF_CORP_NB,   REF_CENT_NB),
            ("FLUX_2025 (EUR)",      REF_CORP_FLUX,  REF_CENT_FLUX),
            ("PNB_2025 (EUR)",       REF_CORP_PNB,   REF_CENT_PNB),
        ]
        tot_formulas = [
            f"={addr['NB_CORP']}+{addr['NB_CENT']}",
            f"={addr['FLUX_CORP']}+{addr['FLUX_CENT']}",
            f"={addr['PNB_CORP']}+{addr['PNB_CENT']}",
        ]
        for i, (lbl, vc, vw) in enumerate(ref_rows_data):
            c1 = ws2.cell(row=r, column=1, value=lbl)
            c2 = ws2.cell(row=r, column=2, value=round(vc, 2))
            c3 = ws2.cell(row=r, column=3, value=round(vw, 2))
            c4 = ws2.cell(row=r, column=4, value=tot_formulas[i])
            for cx in [c1,c2,c3,c4]:
                cx.font = Font(name='Segoe UI', sz=9, bold=(cx==c1))
                cx.fill = fill(GRN2); cx.border = brd
                if cx != c1: cx.alignment = Alignment(horizontal='right')
            r += 1

        # Taux PNB (formule)
        ws2.cell(row=r, column=1, value="Taux PNB_2025").font = Font(name='Segoe UI', sz=9, bold=True)
        for ci, (n, d) in enumerate([(addr['PNB_CORP'], addr['FLUX_CORP']),
                                      (addr['PNB_CENT'], addr['FLUX_CENT']),
                                      (addr['PNB_TOT'],  addr['FLUX_TOT'])], start=2):
            c = ws2.cell(row=r, column=ci, value=f"=IF({d}=0,0,{n}/{d})")
            c.font = Font(name='Segoe UI', sz=9)
            c.fill = fill(GRN2); c.border = brd; c.alignment = Alignment(horizontal='right')
        ws2.cell(row=r, column=1).fill = fill(GRN2); ws2.cell(row=r, column=1).border = brd
        r += 1

        # Flux / carte (formule)
        ws2.cell(row=r, column=1, value="Flux/carte_2025 (EUR)").font = Font(name='Segoe UI', sz=9, bold=True)
        for ci, (n, d) in enumerate([(addr['FLUX_CORP'], addr['NB_CORP']),
                                      (addr['FLUX_CENT'], addr['NB_CENT']),
                                      (addr['FLUX_TOT'],  addr['NB_TOT'])], start=2):
            c = ws2.cell(row=r, column=ci, value=f"=IF({d}=0,0,{n}/{d})")
            c.font = Font(name='Segoe UI', sz=9)
            c.fill = fill(GRN2); c.border = brd; c.alignment = Alignment(horizontal='right')
        ws2.cell(row=r, column=1).fill = fill(GRN2); ws2.cell(row=r, column=1).border = brd
        r += 2

        # -- Parametres - 4 cellules jaunes -------------------------------
        ws2.cell(row=r, column=1, value="PARAMÈTRES DE SIMULATION").font = fnt(ORG, sz=11)
        r += 1
        ws2.cell(row=r, column=2, value="CORPORATE (MX)").font = fnt(GRN, sz=10)
        ws2.cell(row=r, column=3, value="CENTRALISÉ (WL)").font = fnt(BLU, sz=10)
        r += 1

        for lbl_p, key_c, key_w in [
            ("X% — Variation flux / carte", "X_CORP", "X_CENT"),
            ("Y% — Variation nb cartes",    "Y_CORP", "Y_CENT"),
        ]:
            ws2.cell(row=r, column=1, value=lbl_p).font = Font(name='Segoe UI', sz=9, bold=True)
            ws2.cell(row=r, column=1).fill = fill(ORG2); ws2.cell(row=r, column=1).border = brd
            for ci in [2, 3]:
                c = ws2.cell(row=r, column=ci, value=0.0)
                c.font = Font(name='Segoe UI', sz=9, bold=True, color=ORG)
                c.fill = fill(YLW); c.border = brd; c.alignment = Alignment(horizontal='center')
            addr[key_c] = f"B{r}"; addr[key_w] = f"C{r}"
            r += 1
        r += 1

        # -- Valeurs projetees (full formule) ------------------------------
        ws2.cell(row=r, column=1, value="VALEURS PROJETÉES").font = fnt(BLU, sz=11)
        r += 1
        for ci, h in enumerate(["", "CORPORATE (MX)", "CENTRALISÉ (WL)", "TOTAL"], start=1):
            c = ws2.cell(row=r, column=ci, value=h)
            c.font = fnt()
            c.fill = fill(DARK if ci==1 else GRN if ci==2 else BLU if ci==3 else DARK)
            c.border = brd; c.alignment = Alignment(horizontal='center')
        r += 1

        row_proj_nb   = r
        row_proj_flux = r+1
        row_proj_pnb  = r+2
        addr.update({
            'PROJ_NB_CORP': f"B{row_proj_nb}",   'PROJ_NB_CENT': f"C{row_proj_nb}",   'PROJ_NB_TOT': f"D{row_proj_nb}",
            'PROJ_FLUX_CORP': f"B{row_proj_flux}", 'PROJ_FLUX_CENT': f"C{row_proj_flux}", 'PROJ_FLUX_TOT': f"D{row_proj_flux}",
            'PROJ_PNB_CORP': f"B{row_proj_pnb}",  'PROJ_PNB_CENT': f"C{row_proj_pnb}",  'PROJ_PNB_TOT': f"D{row_proj_pnb}",
        })

        proj_defs = [
            ("NB_CARTES projeté",
             f"={addr['NB_CORP']}*(1+{addr['Y_CORP']}/100)",
             f"={addr['NB_CENT']}*(1+{addr['Y_CENT']}/100)"),
            ("FLUX projeté (EUR)",
             f"={addr['NB_CORP']}*(1+{addr['Y_CORP']}/100)*IF({addr['NB_CORP']}=0,0,{addr['FLUX_CORP']}/{addr['NB_CORP']})*(1+{addr['X_CORP']}/100)",
             f"={addr['NB_CENT']}*(1+{addr['Y_CENT']}/100)*IF({addr['NB_CENT']}=0,0,{addr['FLUX_CENT']}/{addr['NB_CENT']})*(1+{addr['X_CENT']}/100)"),
            ("PNB projeté (EUR)",
             f"={addr['PROJ_FLUX_CORP']}*IF({addr['FLUX_CORP']}=0,0,{addr['PNB_CORP']}/{addr['FLUX_CORP']})",
             f"={addr['PROJ_FLUX_CENT']}*IF({addr['FLUX_CENT']}=0,0,{addr['PNB_CENT']}/{addr['FLUX_CENT']})"),
        ]
        proj_tot_keys = [
            (addr['PROJ_NB_CORP'],   addr['PROJ_NB_CENT']),
            (addr['PROJ_FLUX_CORP'], addr['PROJ_FLUX_CENT']),
            (addr['PROJ_PNB_CORP'],  addr['PROJ_PNB_CENT']),
        ]
        proj_fills = [GRN2, BLU2, ORG2]
        for pi, (lbl, fc, fw) in enumerate(proj_defs):
            bg = proj_fills[pi]
            c1 = ws2.cell(row=r, column=1, value=lbl)
            c2 = ws2.cell(row=r, column=2, value=fc)
            c3 = ws2.cell(row=r, column=3, value=fw)
            kc, kw = proj_tot_keys[pi]
            c4 = ws2.cell(row=r, column=4, value=f"={kc}+{kw}")
            for cx in [c1,c2,c3,c4]:
                cx.font = Font(name='Segoe UI', sz=9, bold=(cx==c1))
                cx.fill = fill(bg); cx.border = brd
                if cx != c1: cx.alignment = Alignment(horizontal='right')
            r += 1

        # Gain PNB
        row_gain = r
        addr['GAIN_CORP'] = f"B{row_gain}"
        addr['GAIN_CENT'] = f"C{row_gain}"
        addr['GAIN_TOT']  = f"D{row_gain}"
        ws2.cell(row=r, column=1, value="GAIN PNB (EUR)").font = fnt(BLU, sz=10)
        ws2.cell(row=r, column=1).fill = fill(BLU2); ws2.cell(row=r, column=1).border = brd
        for ci, (p, ref) in enumerate([
            (addr['PROJ_PNB_CORP'], addr['PNB_CORP']),
            (addr['PROJ_PNB_CENT'], addr['PNB_CENT']),
            (addr['PROJ_PNB_TOT'],  addr['PNB_TOT']),
        ], start=2):
            c = ws2.cell(row=r, column=ci, value=f"={p}-{ref}")
            c.font = Font(name='Segoe UI', sz=9, bold=True, color=BLU)
            c.fill = fill(BLU2); c.border = brd; c.alignment = Alignment(horizontal='right')
        r += 1

        row_gain_pct = r
        ws2.cell(row=r, column=1, value="GAIN PNB (%)").font = fnt(BLU, sz=10)
        ws2.cell(row=r, column=1).fill = fill(BLU2); ws2.cell(row=r, column=1).border = brd
        for ci, (p, ref) in enumerate([
            (addr['PROJ_PNB_CORP'], addr['PNB_CORP']),
            (addr['PROJ_PNB_CENT'], addr['PNB_CENT']),
            (addr['PROJ_PNB_TOT'],  addr['PNB_TOT']),
        ], start=2):
            c = ws2.cell(row=r, column=ci,
                          value=f"=IF({ref}=0,0,({p}-{ref})/ABS({ref})*100)")
            c.font = Font(name='Segoe UI', sz=9, bold=True, color=BLU)
            c.fill = fill(BLU2); c.border = brd; c.alignment = Alignment(horizontal='right')
        r += 3

        # -----------------------------------------------------------------
        # TABLE DE DONNEES POUR LES GRAPHIQUES
        # On construit une mini-table avec 3 lignes (Corp / Cent / Total)
        # et 2 periodes (2025 / 2026 projete) pour chaque indicateur.
        # Toutes les cellules de donnees 2026 sont des formules.
        # -----------------------------------------------------------------
        graph_anchor_row = r
        ws2.cell(row=r, column=1, value="TABLE DE DONNÉES — GRAPHIQUES").font = \
            fnt("888888", bold=False, sz=8)
        r += 1

        # Structure :
        # Col A   = Libelle (Corporate / Centralise / Total)
        # Col B   = NB_CARTES 2025
        # Col C   = NB_CARTES 2026 projete (formule)
        # Col D   = FLUX 2025
        # Col E   = FLUX 2026 projete (formule)
        # Col F   = PNB 2025
        # Col G   = PNB 2026 projete (formule)
        # Col H   = GAIN PNB (formule) - pour le graphique gain
        # Col I   = [vide - separateur]

        table_hdr_row = r
        for ci, h in enumerate([
            "Segment",
            "NB_CARTES 2025", "NB_CARTES 2026 projeté",
            "FLUX 2025 (EUR)", "FLUX 2026 projeté (EUR)",
            "PNB 2025 (EUR)", "PNB 2026 projeté (EUR)",
            "GAIN PNB (EUR)"
        ], start=1):
            c = ws2.cell(row=r, column=ci, value=h)
            c.font = fnt(sz=8); c.fill = fill(DARK); c.border = brd
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        ws2.row_dimensions[r].height = 28
        r += 1

        table_data_start = r
        # Donnees des 3 lignes
        table_rows = [
            # (libelle, nb25_addr, nb26_formula, flux25_addr, flux26_formula, pnb25_addr, pnb26_formula, gain_formula)
            ("Corporate (MX)",
             addr['NB_CORP'],   f"={addr['PROJ_NB_CORP']}",
             addr['FLUX_CORP'], f"={addr['PROJ_FLUX_CORP']}",
             addr['PNB_CORP'],  f"={addr['PROJ_PNB_CORP']}",
             f"={addr['GAIN_CORP']}"),
            ("Centralisé (WL)",
             addr['NB_CENT'],   f"={addr['PROJ_NB_CENT']}",
             addr['FLUX_CENT'], f"={addr['PROJ_FLUX_CENT']}",
             addr['PNB_CENT'],  f"={addr['PROJ_PNB_CENT']}",
             f"={addr['GAIN_CENT']}"),
            ("Total",
             addr['NB_TOT'],    f"={addr['PROJ_NB_TOT']}",
             addr['FLUX_TOT'],  f"={addr['PROJ_FLUX_TOT']}",
             addr['PNB_TOT'],   f"={addr['PROJ_PNB_TOT']}",
             f"={addr['GAIN_TOT']}"),
        ]
        fills_rows = [GRN2, BLU2, ORG2]
        for ri_t, (lbl, nb25, nb26, fx25, fx26, pnb25, pnb26, gain) in enumerate(table_rows):
            bg = fills_rows[ri_t]
            vals = [lbl, f"={nb25}", nb26, f"={fx25}", fx26, f"={pnb25}", pnb26, gain]
            for ci, val in enumerate(vals, start=1):
                c = ws2.cell(row=r, column=ci, value=val)
                c.font = Font(name='Segoe UI', sz=9, bold=(ci==1))
                c.fill = fill(bg); c.border = brd
                if ci > 1: c.alignment = Alignment(horizontal='right')
            r += 1
        table_data_end = r - 1

        # -----------------------------------------------------------------
        # GRAPHIQUE 1 - NB CARTES : 2025 vs 2026 projete
        # -----------------------------------------------------------------
        chart1 = BarChart()
        chart1.type     = "col"
        chart1.grouping = "clustered"
        style_chart(chart1, "Nombre de cartes — 2025 vs 2026 projeté", w=13, h=9)
        no_grid(chart1)
        chart1.y_axis.title = "Nb Cartes"

        # Serie 2025 (col B)
        ref_nb25 = Reference(ws2, min_col=2, max_col=2,
                              min_row=table_data_start, max_row=table_data_end)
        chart1.add_data(ref_nb25, titles_from_data=False)
        chart1.series[0].title = SeriesLabel(v="2025")
        chart1.series[0].graphicalProperties.solidFill = GRN

        # Serie 2026 projete (col C)
        ref_nb26 = Reference(ws2, min_col=3, max_col=3,
                              min_row=table_data_start, max_row=table_data_end)
        chart1.add_data(ref_nb26, titles_from_data=False)
        chart1.series[1].title = SeriesLabel(v="2026 projeté")
        chart1.series[1].graphicalProperties.solidFill = BLU

        cats1 = Reference(ws2, min_col=1,
                           min_row=table_data_start, max_row=table_data_end)
        chart1.set_categories(cats1)

        # -----------------------------------------------------------------
        # GRAPHIQUE 2 - FLUX : 2025 vs 2026 projete
        # -----------------------------------------------------------------
        chart2 = BarChart()
        chart2.type     = "col"
        chart2.grouping = "clustered"
        style_chart(chart2, "Flux — 2025 vs 2026 projeté", w=13, h=9)
        no_grid(chart2)
        chart2.y_axis.title = "EUR"

        ref_fx25 = Reference(ws2, min_col=4, max_col=4,
                              min_row=table_data_start, max_row=table_data_end)
        chart2.add_data(ref_fx25, titles_from_data=False)
        chart2.series[0].title = SeriesLabel(v="Flux 2025")
        chart2.series[0].graphicalProperties.solidFill = GRN

        ref_fx26 = Reference(ws2, min_col=5, max_col=5,
                              min_row=table_data_start, max_row=table_data_end)
        chart2.add_data(ref_fx26, titles_from_data=False)
        chart2.series[1].title = SeriesLabel(v="Flux 2026 projeté")
        chart2.series[1].graphicalProperties.solidFill = BLU

        cats2 = Reference(ws2, min_col=1,
                           min_row=table_data_start, max_row=table_data_end)
        chart2.set_categories(cats2)

        # -----------------------------------------------------------------
        # GRAPHIQUE 3 - PNB : 2025 vs 2026 projete
        # -----------------------------------------------------------------
        chart3 = BarChart()
        chart3.type     = "col"
        chart3.grouping = "clustered"
        style_chart(chart3, "PNB — 2025 vs 2026 projeté", w=13, h=9)
        no_grid(chart3)
        chart3.y_axis.title = "EUR"

        ref_pnb25 = Reference(ws2, min_col=6, max_col=6,
                               min_row=table_data_start, max_row=table_data_end)
        chart3.add_data(ref_pnb25, titles_from_data=False)
        chart3.series[0].title = SeriesLabel(v="PNB 2025")
        chart3.series[0].graphicalProperties.solidFill = GRN

        ref_pnb26 = Reference(ws2, min_col=7, max_col=7,
                               min_row=table_data_start, max_row=table_data_end)
        chart3.add_data(ref_pnb26, titles_from_data=False)
        chart3.series[1].title = SeriesLabel(v="PNB 2026 projeté")
        chart3.series[1].graphicalProperties.solidFill = BLU

        cats3 = Reference(ws2, min_col=1,
                           min_row=table_data_start, max_row=table_data_end)
        chart3.set_categories(cats3)

        # -----------------------------------------------------------------
        # GRAPHIQUE 4a - GAIN PNB total (barre unique)
        # -----------------------------------------------------------------
        chart4 = BarChart()
        chart4.type     = "col"
        chart4.grouping = "clustered"
        style_chart(chart4, "Gain PNB (EUR)", w=10, h=9)
        no_grid(chart4)
        chart4.y_axis.title = "EUR"

        ref_gain = Reference(ws2, min_col=8, max_col=8,
                              min_row=table_data_start, max_row=table_data_end)
        chart4.add_data(ref_gain, titles_from_data=False)
        chart4.series[0].title = SeriesLabel(v="Gain PNB")
        chart4.series[0].graphicalProperties.solidFill = ORG

        cats4 = Reference(ws2, min_col=1,
                           min_row=table_data_start, max_row=table_data_end)
        chart4.set_categories(cats4)

        # -----------------------------------------------------------------
        # GRAPHIQUE 4b - PIE CHART repartition gain Corp / Cent
        # (uniquement les 2 premieres lignes : Corp + Cent)
        # -----------------------------------------------------------------
        chart5 = PieChart()
        style_chart(chart5, "Répartition du gain PNB", w=8, h=9)

        ref_pie = Reference(ws2, min_col=8, max_col=8,
                             min_row=table_data_start, max_row=table_data_end - 1)  # Corp + Cent seulement
        chart5.add_data(ref_pie, titles_from_data=False)
        cats_pie = Reference(ws2, min_col=1,
                              min_row=table_data_start, max_row=table_data_end - 1)
        chart5.set_categories(cats_pie)
        chart5.series[0].graphicalProperties.solidFill = None
        # Couleurs des tranches
        try:
            from openpyxl.chart.data_source import NumDataSource
            chart5.series[0].dPt = []
            from openpyxl.chart.data_source import DataPoint
            dp1 = DataPoint(idx=0)
            dp1.spPr = None
            dp2 = DataPoint(idx=1)
            dp2.spPr = None
        except Exception: pass

        # -- Positionnement des 4 graphiques sur la feuille ----------------
        # Graphiques disposes en 2 colonnes x 2 lignes, a droite du tableau
        # G1 (NB CARTES)   | G2 (FLUX)
        # G3 (PNB)         | G4a (GAIN barre) + G4b (PIE)
        chart_col_left  = "F"   # colonne de depart gauche
        chart_col_right = "R"   # colonne de depart droite

        ws2.add_chart(chart1, f"{chart_col_left}4")
        ws2.add_chart(chart2, f"{chart_col_right}4")
        ws2.add_chart(chart3, f"{chart_col_left}24")
        ws2.add_chart(chart4, f"{chart_col_right}24")
        ws2.add_chart(chart5, "Z24")  # pie a cote de la barre gain

        # -- Sauvegarde ----------------------------------------------------
        self._prog(0.99, "Sauvegarde XLSX...")
        wb.save(save_xlsx)
        self._prog(1.0, "Termine !")

        print("")
        print(f"Fichier XLSX cree : {save_xlsx.name}")
        print(f"  Feuille 'Synthese'   : {n_sme:,} clients")
        print("  Feuille 'Simulation' :")
        print("    -> Tableau reference 2025 (full formule)")
        print("    -> 4 cellules jaunes : X_CORP%, Y_CORP%, X_CENT%, Y_CENT%")
        print("    -> Tableau projections 2026 (full formule)")
        print("    -> 4 graphiques : NB CARTES | FLUX | PNB | GAIN PNB + Pie")
        print("    -> Titre en haut, legende en bas, pas de grillage")
        print("    -> A 0%/0%/0%/0% = valeurs 2025 exactes")


# --- MAIN ---------------------------------------------------------------------
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="01.C2SME.py",
        description="SME DATASET ANALYZER [C2SME] - CLI. Extraction et analyse du "
                    "segment SME (Parc Clients) enrichi MONEXT + WORLDLINE.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--parc", required=False, default=None, type=Path,
                        help="(obligatoire) Fichier PARC_CLIENT (source SME), CSV.")
    parser.add_argument("--monext", required=False, default=None, type=Path,
                        help="(obligatoire) Fichier MONEXT consolide - CORPORATE, CSV.")
    parser.add_argument("--worldline", required=False, default=None, type=Path,
                        help="(obligatoire) Fichier WORLDLINE PRGM - CENTRALISE, CSV.")
    parser.add_argument("--bpe-retail", required=False, type=Path, default=None,
                        help="(optionnel) Fichier BPE RETAIL (codes agences). "
                             "Active l'identification BPE si fourni.")
    parser.add_argument("--segments", default=DEFAULT_SME_CODES,
                        help=f'(optionnel) Codes segment SME separes par virgule '
                             f'(defaut "{DEFAULT_SME_CODES}").')
    parser.add_argument("--output-dir", type=Path, default=Path("."),
                        help="(optionnel) Dossier de sortie (defaut : repertoire courant).")
    parser.add_argument("--output-filename", default="",
                        help="(optionnel) Base du nom des fichiers de sortie (sans "
                             "extension). Defaut : SME_DATASET_<timestamp>_C2SME.")
    parser.add_argument("--no-xlsx", action="store_true",
                        help="(optionnel) Ne genere pas le classeur XLSX (CSV uniquement).")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    # Validation des entrees obligatoires (uniquement si fournies en CLI ;
    # sinon resolution auto via 03.sources dans le bloc try ci-dessous).
    for label, p in [("PARC", args.parc), ("MONEXT", args.monext),
                     ("WORLDLINE", args.worldline)]:
        if p is not None and not Path(p).is_file():
            print(f"[ERREUR] Fichier {label} introuvable : {p}", file=sys.stderr)
            return 2
    if args.bpe_retail is not None and not Path(args.bpe_retail).is_file():
        print(f"[ERREUR] Fichier BPE_RETAIL introuvable : {args.bpe_retail}",
              file=sys.stderr)
        return 2

    if not [c.strip() for c in str(args.segments).split(',') if c.strip()]:
        print("[ERREUR] Codes segment SME vides (--segments).", file=sys.stderr)
        return 2

    try:
        # Resolution auto des sources obligatoires via 03.sources si non fournies.
        if not args.parc:
            args.parc = resolve_source("PARC_CLIENT", required=True)
        if not args.monext:
            args.monext = resolve_source("MONEXT_AGREGE", required=True)
        if not args.worldline:
            args.worldline = resolve_source("PRGM_AGREGE", required=True)

        app = AnalyseurC2SME(args)
        # Resolution du mapping de colonnes (defauts UI, remplace l'etape visuelle).
        m, mx_pnb_cfg, wl_pnb_cols = app._resolve_mapping()
        if not mx_pnb_cfg['first'] or not mx_pnb_cfg['last']:
            print("[ERREUR] Plage PNB MONEXT non configuree.", file=sys.stderr)
            return 1
        app.worker(m, mx_pnb_cfg, wl_pnb_cols, app.sme_codes)
    except Exception as e:
        print(f"[ERREUR] {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1

    print("[OK] Traitement C2SME termine.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
