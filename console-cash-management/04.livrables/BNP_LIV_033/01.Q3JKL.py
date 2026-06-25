
"""
RC IDENTIFIER ANALYZER [Q3JKL]
==============================

DESCRIPTION
-----------
Analyseur d'identifiants RC Worldline (BNP Paribas Cash Management,
Direction Monétique). Outil CLI autonome (refactorisé depuis une appli
GUI customtkinter ; logique métier strictement préservée). Resout les
identifiants RC Worldline via un pipeline de 16 méthodes ordonnées
(0→15 + 99) et une synthèse de resolution :
  - validation du format RC (longueur 17, zero-padding),
  - matching RC -> RMPM via le fichier PARC,
  - matching IBAN -> RMPM via les fichiers FORTIS (IBAN_GROUPE /
    IBAN_SINGLETON),
  - matching par raison sociale (tokenisation T1..T6, similarite,
    Jaccard, fuzzy Levenshtein, substrings),
  - absorption des RMPM (table ABSORBES : RMPM source -> RMPM nouveau).
Deux modes : MODE_1 (production PRGM + ACHETEUR -> table complete),
MODE_2 (lookup ciblé : fichier 4 colonnes -> fichier complété).

SOURCES REQUISES
----------------
  PRGM            CSV Worldline consolidé (programmes).
  ACHETEUR        CSV acheteurs.
  PARC            CSV PARC CLIENT (mapping RC -> RMPM).
  IBAN_GROUPE     CSV FORTIS groupe (IBAN -> RMPM, multi-RS).
  IBAN_SINGLETON  CSV FORTIS singleton (IBAN -> RMPM).
  ABSORBES        CSV RMPM absorbés (RMPM source -> RMPM nouveau).
  LOOKUP_INPUT    CSV 4 colonnes (positions absolues) — MODE_2 uniquement.

OUTPUTS PRODUITS
----------------
  Classeur XLSX unique ecrit dans --output-dir / --output-filename :
    - feuille DATA       : table brute (toutes lignes analysees),
    - feuille SYNTHESE   : table dedupliquee par ID_PROGRAMME,
    - feuille ANALYSE    : stats par phase et par status,
    - feuille LIVRABLE   : livrable Worldline final,
    - feuille MODE2      : presente uniquement en MODE_2.

ARGUMENTS CLI
-------------
  --mode {MODE_1,MODE_2}   Mode de fonctionnement (defaut MODE_1).
  --prgm PATH              (oblig) CSV PRGM.
  --acheteur PATH          (oblig) CSV ACHETEUR.
  --parc PATH              (oblig) CSV PARC.
  --iban-groupe PATH       (oblig) CSV IBAN_GROUPE (FORTIS groupe).
  --iban-singleton PATH    (oblig) CSV IBAN_SINGLETON (FORTIS singleton).
  --absorbes PATH          (oblig) CSV RMPM ABSORBES.
  --lookup-input PATH      (oblig en MODE_2) CSV 4 colonnes.
  --no-corruption          Desactive la methode 5 (RC_PARC_CORRUPTION).
  --output-dir PATH        Repertoire de sortie (defaut : repertoire courant).
  --output-filename NAME   Nom du fichier XLSX (defaut : auto-genere).

DECOMPOSITION
-------------
  main()
    └── RCIdentifierAnalyzer_Q3JKL(args).run()
          └── _worker()
                ├── BLOC 1 — CHARGEMENT
                │     PRGM, ACHETEUR, PARC, IBAN_GROUPE, IBAN_SINGLETON,
                │     ABSORBES (+ LOOKUP_INPUT en MODE_2).
                ├── BLOC 2 — PREPARATION SOURCES
                │     prep PRGM (dedup RC/PROD, derniere apparition),
                │     prep ACHETEUR (dedup, etat actif, dictionnaires),
                │     construction df_union (MODE_1 fusion / MODE_2 lookup),
                │     dictionnaires IBAN (FORTIS groupe + singleton),
                │     dictionnaire PARC (RC -> RMPM), dictionnaire absorbes.
                ├── BLOC 3 — INDEXES RS
                │     build_rs_indexes : T1->T6, tokens, fuzzy buckets, ngrams.
                ├── BLOC 4 — PIPELINE CASCADE (16 methodes 0->15 + 99)
                │     0 EXCLU / 1-2 IBAN / 3-5 RC PARC / 6-11 RS clean /
                │     12 Jaccard / 13 fuzzy / 14-15 substrings / 99 non resolu.
                ├── BLOC 5 — POST-TRAITEMENT
                │     absorption RMPM, enrichissement, construction df_out.
                └── SYNTHESE DE RESOLUTION
                      dedup intelligent par ID_PROGRAMME, stats par phase
                      et par status, export XLSX (DATA / SYNTHESE / ANALYSE /
                      LIVRABLE [/ MODE2]).

BNP Paribas Cash Management - Direction Monetique
Juin 2026

(Historique des patches conserve depuis la version GUI v10.3 — logique
metier identique.)
Architecture simplifiée à 16 méthodes
ordonnées (0→15 + 99), pipeline de nettoyage atomique T1→T6,
audit narratif détaillé, livrable Worldline final.

PATCH v10.3 vs v10.2 :
----------------------
- BUG FIX méthode 0 (EXCLU) : le dédoublonnage masquait les EXCLU.
  Quand un ID_PROGRAMME avait une ligne PRGM résolue (phase 3) + une ligne
  ACHETEUR avec libellé "NE_PAS_..." (phase 0), le dedup gardait la résolution
  PRGM. Fix : **EXCLU sticky** dans smart_dedup_by_id_programme — si au moins
  une ligne du groupe est EXCLU, le programme entier est EXCLU + le libellé
  utilisé est celui de la ligne EXCLU.

- BUG FIX cross-country matching :
  "BOPRO SAS - EUR" ne doit PAS matcher "BOPRO NV" (FR vs NL).
  "ALTO - EUR" ne doit PAS matcher "ALTO SRL" (juridique inconnue côté source).
  "FACILITY CORP - EUR" PEUT matcher "FACILITY" (target abbrégé).
  Implémenté via extract_legal_suffix + is_juridique_compatible, appliqué
  aux phases 9, 11, 12, 13, 14, 15. Règles :
    · source ∅ / target ∅   → ACCEPT
    · source ∅ / target plein → REJECT (source ambiguë)
    · source plein / target ∅ → ACCEPT (target abbrégé)
    · source = target       → ACCEPT
    · source ≠ target       → REJECT (jurisdictions différentes)

- Méthode 12 Jaccard : seuil 0.80 → **1.00** (exact token match strict)
- PHASES_FAIBLES étendu : {12, 13, 14, 15} (avant : {14, 15})
  → Méthodes 12 et 13 désormais classées RESOLU_FAIBLE quoi qu'il arrive
- Regex NE_PAS simplifié, normalisation T1_base appliquée avant test

PATCH v10.2 :
-------------
- Dédoublonnage intelligent par ID_PROGRAMME :
  PRODUIT concaténé / DATE plus ancienne / APPARITION plus récente
  Priorité PRGM > ACHETEUR pour libellé/RS/ETAT/IBAN
  ID_RC : priorité PRGM, sinon le plus long
  Résolution : meilleure phase + meilleur score

PATCH v10.1 :
-------------
- Fix regex méthode 0 (lookbehind/lookahead)
- Seuils durcis (méthode 5 corruption RC, fuzzy 0.95)
- Substrings refactorés : méthode 14 (≥25 chars), méthode 15 (≥18 chars)

NOUVEAUTÉS vs J8WQT :
---------------------
1. Méthode 0 EXCLU_NE_PAS_TOUCHER : détection regex robuste sur libellé
2. Pipeline RS atomique T1→T6 (transformations combinables)
3. 16 méthodes ordonnées au lieu de 6 blocs parallèles
4. 22 colonnes au lieu de ~70
5. AUDIT_RESOLUTION narratif lisible étape par étape
6. Suffixe juridique "LDT" (coquille fréquente de "LTD") ajouté
7. Normalisation pays expand-to-ISO (20+ pays courants)
8. WORLDLINE_LIVRABLE format final : "Inconnu1234" (sans underscore/tiret/espace)
9. ANALYSE avec descriptions claires de chaque méthode
10. Filtre juridique cross-country (v10.3)

LISTE DES MÉTHODES :
0.  EXCLU_NE_PAS_TOUCHER
1.  IBAN_ACCOUNT
2.  IBAN_SINGLETON
3.  RC_PARC_DIRECT
4.  RC_PARC_VARIANTE       (préfixe "00" ou zfill 17)
5.  RC_PARC_CORRUPTION     (len 14-16 reconstruit + validé par sim RS ≥ 0.75)
6.  RS_EXACT               (T1)
7.  RS_CLEAN_PUNCT         (T1+T2+T3)
8.  RS_CLEAN_DEVISE        (T1+T2+T3+T4)
9.  RS_CLEAN_JURIDIQUE     (T1+T2+T3+T5) + filtre cross-country
10. RS_CLEAN_PAYS          (T1+T2+T3+T6)
11. RS_CLEAN_FULL          (T1+T2+T3+T4+T5+T6) + filtre cross-country
12. RS_TOKEN_JACCARD       (= 1.00 STRICT) + filtre cross-country — RESOLU_FAIBLE
13. RS_FUZZY_LEVENSHTEIN   (≥ 0.95) + filtre cross-country — RESOLU_FAIBLE
14. RS_SUBSTRING_25        (≥ 25 chars) + filtre cross-country — RESOLU_FAIBLE
15. RS_SUBSTRING_18        (≥ 18 chars) + filtre cross-country — RESOLU_FAIBLE
99. NON_RESOLU

BNP Paribas Cash Management - Direction Monétique
Juin 2026
"""

import argparse
import os
import re
import sys
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher
from collections import defaultdict
from pathlib import Path

import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Résolution auto des sources + lecture DuckDB (inliné depuis 06.fonctions) ──
def _find_sources_root():
    here = Path(__file__).resolve()
    for d in [here] + list(here.parents):
        c = d / "03.sources"
        if c.is_dir():
            return c
    return None


def resolve_source(name, required=False):
    root = _find_sources_root()
    if not root:
        if required:
            raise FileNotFoundError(f"Dossier 03.sources introuvable depuis {__file__}")
        return None
    folder = next((m for m in root.glob("*/" + name) if m.is_dir()), None)
    if not folder:
        if required:
            raise FileNotFoundError(f"Source introuvable : 03.sources/*/{name}")
        return None
    exts = (".csv", ".xlsx", ".xls", ".txt")
    files = [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() in exts]
    if not files:
        if required:
            raise FileNotFoundError(f"Aucun fichier source dans {folder}")
        return None

    def _np(f):
        m = re.match(r"\s*(\d+)", f.name)
        return (int(m.group(1)) if m else -1, f.name)
    return max(files, key=_np)


def _read_duck(path, nrows=None):
    p = Path(path)
    if nrows is not None or p.suffix.lower() not in (".csv", ".txt", ""):
        return None
    try:
        import duckdb
        return duckdb.read_csv(str(p), all_varchar=True).df()
    except Exception:
        return None


# =============================================================================
# CONSTANTES
# =============================================================================

VERSION_ID = "Q3JKL"

# Couleurs
GRN   = "00915A"; GRN2  = "E8F5E9"; GRN3  = "005A3C"
DARK  = "1C3A2D"; GREY  = "F5F5F5"; GREY2 = "EEEEEE"
WHT   = "FFFFFF"; RED   = "B71C1C"; RED2  = "FFEBEE"
BLU   = "1A4D7A"; BLU2  = "E3F2FD"
ORG   = "E65100"; ORG2  = "FFF3E0"
PUR   = "7B1FA2"; PUR2  = "F3E5F5"
YLO   = "F57F17"; YLO2  = "FFF9C4"

RC_LENGTH = 17

# --- Méthodes et scoring ----------------------------------------------------

METHODE_LABELS = {
    0:  "EXCLU_NE_PAS_TOUCHER",
    1:  "IBAN_ACCOUNT",
    2:  "IBAN_SINGLETON",
    3:  "RC_PARC_DIRECT",
    4:  "RC_PARC_VARIANTE",
    5:  "RC_PARC_CORRUPTION",
    6:  "RS_EXACT",
    7:  "RS_CLEAN_PUNCT",
    8:  "RS_CLEAN_DEVISE",
    9:  "RS_CLEAN_JURIDIQUE",
    10: "RS_CLEAN_PAYS",
    11: "RS_CLEAN_FULL",
    12: "RS_TOKEN_JACCARD",
    13: "RS_FUZZY_LEVENSHTEIN",
    14: "RS_SUBSTRING_25",
    15: "RS_SUBSTRING_18",
    99: "NON_RESOLU",
}

# Score de base par phase (modulé par similarité RS dans certains cas)
SCORE_BASE = {
    0:  0.00,  # EXCLU = pas de RMPM
    1:  1.00,  # IBAN_ACCOUNT
    2:  0.90,  # IBAN_SINGLETON
    3:  0.95,  # RC_PARC_DIRECT
    4:  0.85,  # RC_PARC_VARIANTE
    5:  0.70,  # RC_PARC_CORRUPTION
    6:  0.95,  # RS_EXACT
    7:  0.90,  # RS_CLEAN_PUNCT
    8:  0.88,  # RS_CLEAN_DEVISE
    9:  0.88,  # RS_CLEAN_JURIDIQUE
    10: 0.85,  # RS_CLEAN_PAYS
    11: 0.82,  # RS_CLEAN_FULL
    12: 0.78,  # RS_TOKEN_JACCARD (modulé par jaccard)
    13: 0.75,  # RS_FUZZY_LEVENSHTEIN (modulé par ratio)
    14: 0.70,  # RS_SUBSTRING_25 → FAIBLE mais score relevé (substr 25 chars très restrictif)
    15: 0.55,  # RS_SUBSTRING_18 → FAIBLE (substr 18 chars, fiabilité moyenne)
    99: 0.00,
}

# Phases 12-15 → RESOLU_FAIBLE quoi qu'il arrive (v10.3)
PHASES_FAIBLES = {12, 13, 14, 15}

# Seuils (v10.3 — Jaccard durci à 1.00 par décision métier)
MIN_SIM_CORRUPTION = 0.75   # méthode 5
MIN_JACCARD        = 1.00   # méthode 12 (était 0.80) — exact token match requis
MIN_FUZZY          = 0.95   # méthode 13
NGRAM_LEN_LONG     = 25     # méthode 14
NGRAM_LEN_COURT    = 18     # méthode 15
FUZZY_BUCKET_SIZE  = 5
FUZZY_MAX_CANDIDATES = 200

# --- Devises ---------------------------------------------------------------

CURRENCIES = [
    "EUR", "USD", "GBP", "CHF", "JPY", "CAD", "AUD", "CNY", "HKD", "SGD",
    "NOK", "SEK", "DKK", "PLN", "CZK", "HUF", "RON", "BGN", "HRK", "RUB",
    "TRY", "MAD", "TND", "DZD", "ZAR", "INR", "THB", "BRL", "MXN", "AED",
    "SAR", "QAR", "KWD", "OMR", "JOD", "EGP", "NGN", "GHS", "KES", "TWD",
    "KRW", "VND", "IDR", "MYR", "PHP", "NZD", "CLP", "COP", "PEN", "ARS",
    "ILS", "ISK", "BHD",
]

# --- Suffixes juridiques (avec LDT = coquille de LTD) ---------------------

LEGAL_SUFFIXES = [
    "SASU", "SARL", "EURL", "SNC", "SCS", "SCA", "SAS", "SA",
    "LIMITED", "LTD", "LDT",   # LDT = coquille fréquente de LTD
    "LLC", "INCORPORATED", "INC", "CORPORATION", "CORP", "COMPANY", "CO",
    "GMBH", "AG", "KG", "OHG", "MBH",
    "BV", "NV", "OY", "AB", "AS", "PLC", "PTE", "SPA", "SRL", "SL", "SLU",
    "LP", "LLP", "LTDA", "SDN", "BHD", "SE", "PT", "TBK",
]
LEGAL_SUFFIXES_SORTED = sorted(LEGAL_SUFFIXES, key=len, reverse=True)

# --- Pays : expand-to-ISO bidirectionnel (nom complet → code ISO) --------
# On expand TOUJOURS vers ISO. Pas l'inverse (trop risqué : "FR" peut être
# un fragment d'autre chose, "AT" idem, etc.).

COUNTRY_EXPAND_TO_ISO = {
    "FRANCE": "FR",
    "GERMANY": "DE", "ALLEMAGNE": "DE", "DEUTSCHLAND": "DE",
    "ITALY": "IT", "ITALIE": "IT", "ITALIA": "IT",
    "SPAIN": "ES", "ESPAGNE": "ES", "ESPANA": "ES",
    "BELGIUM": "BE", "BELGIQUE": "BE", "BELGIE": "BE",
    "NETHERLANDS": "NL", "PAYS BAS": "NL", "HOLLAND": "NL", "HOLLANDE": "NL",
    "LUXEMBOURG": "LU",
    "PORTUGAL": "PT",
    "UNITED KINGDOM": "GB", "GREAT BRITAIN": "GB", "ROYAUME UNI": "GB", "ENGLAND": "GB",
    "IRELAND": "IE", "IRLANDE": "IE",
    "SWITZERLAND": "CH", "SUISSE": "CH", "SCHWEIZ": "CH",
    "AUSTRIA": "AT", "AUTRICHE": "AT", "OSTERREICH": "AT",
    "UNITED STATES": "US", "ETATS UNIS": "US", "USA": "US",
    "CANADA": "CA",
    "POLAND": "PL", "POLOGNE": "PL", "POLSKA": "PL",
    "SWEDEN": "SE", "SUEDE": "SE", "SVERIGE": "SE",
    "NORWAY": "NO", "NORVEGE": "NO", "NORGE": "NO",
    "DENMARK": "DK", "DANEMARK": "DK", "DANMARK": "DK",
    "FINLAND": "FI", "FINLANDE": "FI", "SUOMI": "FI",
    "CZECH REPUBLIC": "CZ", "TCHEQUIE": "CZ", "REPUBLIQUE TCHEQUE": "CZ",
}
# Pré-compilation : on trie par longueur décroissante pour éviter "UNITED"
# matché avant "UNITED KINGDOM"
COUNTRY_EXPAND_SORTED = sorted(COUNTRY_EXPAND_TO_ISO.items(),
                                key=lambda kv: -len(kv[0]))

# --- Stop tokens (Jaccard) ------------------------------------------------

STOP_TOKENS = {
    "DE", "DU", "DES", "LA", "LE", "LES", "ET", "EN", "AU", "AUX", "UN", "UNE",
    "OF", "THE", "AND", "FOR", "GROUP", "GROUPE", "INTERNATIONAL", "INTL",
    "FRANCE", "EUROPE", "WORLD", "GLOBAL",
}

# --- Pattern méthode 0 (EXCLU "NE PAS") -----------------------------------
# Matche (sur libellé déjà uppercase) :
#   - "NE PAS TOUCHER" / "NE PAS UTILISER"
#   - "NE_PAS_UTILISER" / "ne_pas_util" / "NE_PAS_FUTUR ECO HABITAT"
#   - "NE-PAS" / "NE.PAS"
#   - "NEPAS" (collé, validé par Ali — aucune entreprise ne s'appelle NEPAS)
# Ne matche PAS : "BENEPAS", "DUNEPAS" (lettre alpha collée avant ou après)
#
# v10.3 : on travaille sur libellé pré-normalisé (T1_base = upper + accents),
# donc plus de re.IGNORECASE. Lookaround [A-Z] sans risque de match A-Za-z.
PATTERN_NE_PAS = re.compile(
    r"(?<![A-Z])(?:NE[ _\-\.]?PAS|NEPAS)(?![A-Z])"
)

# --- Regex de nettoyage ---------------------------------------------------

_CURR_PATTERN = r"(?:^|[\s\-_.]+)(?:" + "|".join(re.escape(c) for c in CURRENCIES) + r")\s*$"
CURR_REGEX = re.compile(_CURR_PATTERN, re.IGNORECASE)

_LEGAL_PATTERN = r"(?:^|[\s\-_.]+)(?:" + "|".join(re.escape(s) for s in LEGAL_SUFFIXES_SORTED) + r")\s*$"
LEGAL_REGEX = re.compile(_LEGAL_PATTERN, re.IGNORECASE)

PUNCT_REGEX = re.compile(r"[\.\,\-\_\/\\\&\'\"\(\)\;\:]+")
SPACES_REGEX = re.compile(r"\s+")
TOKEN_SPLIT_REGEX = re.compile(r"[\s\-_.,;:'\"/\\&()]+")

# --- Positions par défaut (mêmes que J8WQT) -------------------------------

DEFAULT_POS = {
    "prgm_id": 4, "prgm_rc": 40, "prgm_iban": 9, "prgm_rs": 8,
    "prgm_libelle": 3, "prgm_produit": 5, "prgm_mois": 2,
    "ach_id": 1, "ach_libelle": 2, "ach_rc": 71, "ach_rs": 3,
    "ach_iban": 17, "ach_etat": 19, "ach_type": 15, "ach_periode": 13, "ach_date_crea": 72,
    "parc_rc": 13, "parc_rmpm": 6, "parc_rs": 8, "parc_code_ga": 11, "parc_nom_ga": 12,
    "ibg_iso_ga": 1, "ibg_code_ga": 2, "ibg_nom_ga": 3, "ibg_iso_ej": 4,
    "ibg_rmpm": 5, "ibg_rs": 6, "ibg_iban": 7,
    "ibs_iso": 1, "ibs_rmpm": 2, "ibs_rs": 3, "ibs_iban": 4,
    "abs_rmpm_src": 1, "abs_rmpm_new": 2, "abs_nom": 3,
}


# =============================================================================
# UTILITAIRES BAS NIVEAU
# =============================================================================

def load_csv_smart(path, nrows=None):
    _d = _read_duck(path, nrows)
    if _d is not None:
        return _d
    for sep in [";", ",", "\t"]:
        for enc in ["utf-8-sig", "utf-8", "latin1", "cp1252"]:
            try:
                df = pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                 keep_default_na=False, na_values=[],
                                 on_bad_lines="skip", nrows=5)
                if df.shape[1] > 1:
                    return pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                       keep_default_na=False, na_values=[],
                                       on_bad_lines="skip", nrows=nrows)
            except Exception:
                continue
    return pd.read_csv(path, sep=None, engine="python", dtype=str,
                       keep_default_na=False, na_values=[],
                       on_bad_lines="skip", nrows=nrows)


def clean_id_safe(series):
    s = series.astype(str).str.strip()
    s = s.replace(["", "nan", "NaN", "None", "NULL", "NA", "N/A"], "")
    mask = s.str.startswith('="') & s.str.endswith('"')
    s = s.where(~mask, s.str[2:-1])
    s = s.str.lstrip("'")
    mask2 = s.str.endswith(".0") & s.str[:-2].str.isdigit()
    s = s.where(~mask2, s.str[:-2])
    return s.str.strip()


def clean_iban(series):
    return clean_id_safe(series).str.upper().str.replace(" ", "", regex=False)


def excel_id(value):
    if not value or str(value) in ("nan", "None"):
        return ""
    return f'="{value}"'


def sanitize_id_for_inconnu(id_value):
    """Pour 'Inconnu1234' : supprime espace, underscore, tiret de l'ID."""
    if not id_value:
        return ""
    return re.sub(r"[\s_\-]", "", str(id_value))


def is_placeholder_rmpm(value):
    if value is None:
        return True
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "null", "n/a"):
        return True
    s_compact = s.replace(" ", "")
    if not s_compact:
        return True
    if all(c == "0" for c in s_compact):
        return True
    if all(c.upper() == "X" for c in s_compact):
        return True
    return False


# =============================================================================
# PIPELINE DE NETTOYAGE ATOMIQUE — T1 à T6
# =============================================================================

def T1_base(s):
    """T1_BASE : uppercase + strip accents."""
    if pd.isna(s) or str(s).strip() == "":
        return ""
    n = unicodedata.normalize("NFD", str(s).strip().upper())
    return "".join(c for c in n if unicodedata.category(c) != "Mn")


def T2_punct(s):
    """T2_PUNCT : supprime ponctuation."""
    return PUNCT_REGEX.sub(" ", s)


def T3_spaces(s):
    """T3_SPACES : réduit espaces multiples et trim."""
    return SPACES_REGEX.sub(" ", s).strip()


def T4_devise(s):
    """T4_DEVISE : strip suffixe devise (loop jusqu'à stabilité)."""
    prev = None
    while prev != s:
        prev = s
        s = CURR_REGEX.sub("", s).strip()
    return s


def T5_juridique(s):
    """T5_JURIDIQUE : strip suffixe juridique (loop jusqu'à stabilité)."""
    prev = None
    while prev != s:
        prev = s
        s = LEGAL_REGEX.sub("", s).strip()
    return s


def T6_pays(s):
    """T6_PAYS : expand noms de pays → code ISO (FRANCE→FR, GERMANY→DE…)."""
    if not s:
        return s
    for full_name, iso in COUNTRY_EXPAND_SORTED:
        # \b assure qu'on remplace que des mots entiers
        s = re.sub(r"\b" + re.escape(full_name) + r"\b", iso, s)
    return s


# Niveaux de normalisation combinés (pour les méthodes 6 à 11)

def norm_L1(s):
    """Méthode 6 — RS_EXACT : T1 seul."""
    return T1_base(s)


def norm_L2(s):
    """Méthode 7 — RS_CLEAN_PUNCT : T1+T2+T3."""
    return T3_spaces(T2_punct(T1_base(s)))


def norm_L3(s):
    """Méthode 8 — RS_CLEAN_DEVISE : T1+T2+T3+T4."""
    return T4_devise(T3_spaces(T2_punct(T1_base(s))))


def norm_L4(s):
    """Méthode 9 — RS_CLEAN_JURIDIQUE : T1+T2+T3+T5."""
    return T5_juridique(T3_spaces(T2_punct(T1_base(s))))


def norm_L5(s):
    """Méthode 10 — RS_CLEAN_PAYS : T1+T2+T3+T6."""
    return T3_spaces(T6_pays(T3_spaces(T2_punct(T1_base(s)))))


def norm_L6(s):
    """Méthode 11 — RS_CLEAN_FULL : T1+T2+T3+T4+T5+T6 (loop devise/juridique)."""
    out = T3_spaces(T2_punct(T1_base(s)))
    prev = None
    while prev != out:
        prev = out
        out = T4_devise(out)
        out = T5_juridique(out)
    out = T3_spaces(T6_pays(out))
    return out


NORM_FUNCS = {
    6:  ("RS_EXACT",            norm_L1),
    7:  ("RS_CLEAN_PUNCT",      norm_L2),
    8:  ("RS_CLEAN_DEVISE",     norm_L3),
    9:  ("RS_CLEAN_JURIDIQUE",  norm_L4),
    10: ("RS_CLEAN_PAYS",       norm_L5),
    11: ("RS_CLEAN_FULL",       norm_L6),
}


def tokenize_rs(s, min_len=3):
    """Tokens significatifs (longueur ≥3, hors stop tokens). Pipeline L6."""
    norm = norm_L6(s)
    tokens = TOKEN_SPLIT_REGEX.split(norm)
    return frozenset(t for t in tokens if len(t) >= min_len and t not in STOP_TOKENS)


def jaccard(set_a, set_b):
    if not set_a or not set_b:
        return 0.0
    inter = len(set_a & set_b)
    union = len(set_a | set_b)
    return inter / union if union else 0.0


def rs_similarity_L6(a, b):
    if not a or not b:
        return 0.0
    na = norm_L6(a); nb = norm_L6(b)
    if not na or not nb:
        return 0.0
    return SequenceMatcher(None, na, nb).ratio()


# =============================================================================
# MÉTHODE 0 — EXCLUSION "NE PAS TOUCHER"
# =============================================================================

def is_excluded(libelle):
    """Retourne True si le libellé contient un motif d'exclusion 'NE PAS'.
    Normalise d'abord le libellé : upper + strip accents + retire caractères
    invisibles (BOM, ZWJ, etc.) pour éviter les ratés à cause d'encodage."""
    if not libelle or pd.isna(libelle):
        return False
    # T1_base = upper + strip accents
    norm = T1_base(str(libelle))
    # Retire caractères de contrôle / invisibles potentiels
    norm = re.sub(r'[\x00-\x1f\x7f-\x9f\u200b-\u200d\ufeff]', '', norm)
    return bool(PATTERN_NE_PAS.search(norm))


# =============================================================================
# FILTRE JURIDIQUE CROSS-COUNTRY (v10.3)
# =============================================================================

def extract_legal_suffix(rs):
    """Extrait le suffixe juridique en fin de RS, après avoir nettoyé
    ponctuation + devise + pays. Retourne le suffixe normalisé (SAS, NV, SRL...)
    ou "" si aucun.

    Exemples :
        "BOPRO SAS - EUR"  → "SAS"
        "BOPRO NV"         → "NV"
        "ACME SRL"         → "SRL"
        "FACILITY"         → ""
        "ALTO - EUR"       → ""
    """
    if not rs:
        return ""
    # Pipeline jusqu'à T6 mais SANS strip juridique (T5)
    base = T3_spaces(T2_punct(T1_base(rs)))
    after_currency = T4_devise(base)
    after_country = T3_spaces(T6_pays(after_currency))
    m = LEGAL_REGEX.search(after_country)
    if not m:
        return ""
    matched = m.group(0).strip()
    # Trim séparateurs au début (le pattern inclut le séparateur)
    for sep in (' ', '_', '-', '.'):
        matched = matched.lstrip(sep)
    return matched.strip()


def is_juridique_compatible(source_suffix, target_suffix):
    """Filtre cross-country : rejette les matches dont les suffixes juridiques
    sont sémantiquement incompatibles.

    Règles :
    - source vide, target vide          → OK (aucune info juridique des deux côtés)
    - source vide, target plein         → REJECT (source ambiguë, ne peut confirmer le pays)
    - source plein, target vide         → OK (target abbréviation, source plus précise)
    - source plein = target plein       → OK (même juridiction)
    - source plein ≠ target plein       → REJECT (jurisdictions différentes)

    Exemples :
        ("SAS", "NV")   → False  (FR vs NL — REJECT)
        ("",    "SRL")  → False  (source vide, target italien — REJECT)
        ("CORP","")     → True   (source US, target abbrégé — ACCEPT)
        ("SAS", "SAS")  → True   (même juridiction — ACCEPT)
        ("",    "")     → True   (rien à comparer — ACCEPT)
    """
    if source_suffix == target_suffix:
        return True
    if source_suffix and not target_suffix:
        return True
    return False


# =============================================================================
# VALIDATION & VARIANTES RC
# =============================================================================

def validate_rc(rc_str):
    if not rc_str or rc_str.strip() == "":
        return False, "VIDE"
    s = rc_str.strip()
    if not s.isdigit():
        if re.fullmatch(r"[Xx]{2,}", s) or re.fullmatch(r"[Nn]/[Aa]", s) or not any(c.isdigit() for c in s):
            return False, "PLACEHOLDER"
        return False, "NON_NUMERIQUE"
    if all(c == "0" for c in s):
        return False, "ZERO_ONLY"
    if len(s) < RC_LENGTH:
        return False, "TROP_COURT"
    if len(s) > RC_LENGTH:
        return False, "TROP_LONG"
    return True, "OK"


def rc_variants(rc_str):
    """Variantes pour méthode 4 : original + préfixe 00 + zfill(17)."""
    if not rc_str:
        return []
    s = rc_str.strip()
    if not s:
        return []
    variants = [s]
    if s.isdigit():
        v_00 = "00" + s
        if v_00 != s and v_00 not in variants:
            variants.append(v_00)
        v_zf = s.zfill(RC_LENGTH)
        if v_zf not in variants:
            variants.append(v_zf)
    return variants


def corruption_candidates(rc_str):
    """Méthode 5 : reconstruction d'un RC tronqué (len 14-16).
    - len 16 → on présume qu'il manque un '0' au début → "0" + s
    - len 15 → on présume qu'il manque "00" au début → "00" + s
    - len 14 → on tente "00" + s + chaque digit 0-9 (brute force)
    """
    if not rc_str or not rc_str.strip().isdigit():
        return []
    s = rc_str.strip()
    L = len(s)
    candidates = []
    if L == 16:
        candidates.append(("0" + s, "len_16_prefix_0"))
    elif L == 15:
        candidates.append(("00" + s, "len_15_prefix_00"))
    elif L == 14:
        for d in "0123456789":
            candidates.append(("00" + s + d, "len_14_brute_suffix"))
    return candidates


# =============================================================================
# INDEX BUILDERS
# =============================================================================

def build_rs_indexes(sources):
    """
    Construit pour chaque source RS :
    - 6 index exact (L1..L6) : {norm_value: payload}
    - tokens : {token: [(tokens_set, payload), ...]}
    - fuzzy_buckets : {bucket_id: [(L6_value, payload), ...]}
    - ngram_long (25) et ngram_court (18) : {ngram: payload}

    sources = liste de tuples (source_name, df, col_rs, col_rmpm, col_cga, col_nga)

    v10.3 : payload = (rmpm, rs, cga, nga, src_name, legal_suffix)
    Le suffixe juridique de la RS cible est stocké pour permettre le filtre
    cross-country dans run_cascade.
    """
    idx = {k: {} for k in range(6, 12)}  # L1..L6 → méthodes 6..11
    token_idx = defaultdict(list)
    fuzzy_buckets = defaultdict(list)
    ngram_long = {}
    ngram_court = {}

    for src_name, df, col_rs, col_rmpm, col_cga, col_nga in sources:
        rs_arr   = df[col_rs].astype(str).values
        rmpm_arr = df[col_rmpm].astype(str).values
        cga_arr  = df[col_cga].astype(str).values if col_cga else [""] * len(df)
        nga_arr  = df[col_nga].astype(str).values if col_nga else [""] * len(df)

        for rs, rmpm, cga, nga in zip(rs_arr, rmpm_arr, cga_arr, nga_arr):
            rs = rs.strip(); rmpm = rmpm.strip()
            if not rs or not rmpm:
                continue
            target_suffix = extract_legal_suffix(rs)
            payload = (rmpm, rs, cga.strip(), nga.strip(), src_name, target_suffix)

            # Indexes exacts L1..L6
            for phase, (_, fn) in NORM_FUNCS.items():
                key = fn(rs)
                if key and key not in idx[phase]:
                    idx[phase][key] = payload

            rs_L6 = norm_L6(rs)

            # Tokens (méthode 12)
            tokens = tokenize_rs(rs)
            for tok in tokens:
                token_idx[tok].append((tokens, payload))

            # Fuzzy buckets (méthode 13)
            if len(rs_L6) >= 5:
                bucket = len(rs_L6) // FUZZY_BUCKET_SIZE
                fuzzy_buckets[bucket].append((rs_L6, payload))

            # N-grams long (méthode 14) et court (méthode 15)
            for L_ngram, ng_idx in [(NGRAM_LEN_LONG, ngram_long), (NGRAM_LEN_COURT, ngram_court)]:
                if len(rs_L6) >= L_ngram:
                    for i in range(len(rs_L6) - L_ngram + 1):
                        ng = rs_L6[i:i + L_ngram]
                        if ng not in ng_idx:
                            ng_idx[ng] = payload

    return idx, token_idx, fuzzy_buckets, ngram_long, ngram_court


# =============================================================================
# CASCADE ORCHESTRATOR + AUDIT BUILDER
# =============================================================================

class CascadeResult:
    __slots__ = ("phase", "methode", "rmpm", "rs_target", "cga", "nga",
                 "source", "val_source", "val_target", "score", "sim")

    def __init__(self, phase, methode, rmpm, rs_target, cga, nga,
                 source, val_source, val_target, score, sim=None):
        self.phase = phase
        self.methode = methode
        self.rmpm = rmpm
        self.rs_target = rs_target
        self.cga = cga
        self.nga = nga
        self.source = source            # nom du fichier source (PARC/IBAN_GROUPE/IBAN_SINGLETON)
        self.val_source = val_source    # valeur testée (RS nettoyée, IBAN, RC...)
        self.val_target = val_target    # valeur trouvée côté cible
        self.score = score
        self.sim = sim


def compute_score(phase, sim=None):
    base = SCORE_BASE.get(phase, 0.0)
    if phase in (1, 2, 3):  # IBAN_*, RC direct → score fixe
        return round(base, 2)
    if phase in (4,):  # variante : modulé légèrement
        return round(base, 2)
    if phase in (5,):  # corruption : modulé par sim
        if sim is None or sim <= 0:
            return round(base * 0.5, 2)
        return round(base * (0.5 + 0.5 * sim), 2)
    if phase in (6, 7, 8, 9, 10, 11):
        if sim is None:
            return round(base, 2)
        return round(base * (0.7 + 0.3 * sim), 2)
    if phase == 12:  # jaccard
        if sim is None:
            sim = 0.7
        return round(base * sim, 2)
    if phase == 13:  # fuzzy
        if sim is None:
            sim = 0.85
        return round(base * sim, 2)
    if phase == 14:  # substring long → faible
        return round(base, 2)
    if phase == 15:  # best guess → très faible
        return round(base, 2)
    return 0.0


def status_from_phase_and_score(phase, score):
    if phase == 0:
        return "EXCLU"
    if phase == 99 or score == 0.0:
        return "NON_RESOLU"
    if phase in PHASES_FAIBLES:
        return "RESOLU_FAIBLE"
    if score >= 0.75:
        return "RESOLU"
    if score >= 0.50:
        return "RESOLU_MOYEN"
    return "RESOLU_FAIBLE"


def run_cascade(libelle, rc_raw, iban, rs, audit_steps,
                ibg_iban_dict, ibs_iban_dict, parc_rc_dict,
                rs_indexes, token_idx, fuzzy_buckets,
                ngram_long, ngram_court,
                corruption_active=True):
    """
    Exécute la cascade des 16 méthodes. Première qui match gagne.
    Construit l'audit narratif au fil de l'eau.
    Retourne CascadeResult ou None (NON_RESOLU).

    v10.3 — Filtre juridique cross-country appliqué aux phases 9, 11, 12, 13,
    14, 15 (toutes celles qui strippent ou ignorent la juridique).
    """

    # Calcul du suffixe juridique de la source — utilisé par le filtre
    query_suffix = extract_legal_suffix(rs) if rs else ""

    # ── Méthode 0 : EXCLU_NE_PAS_TOUCHER ─────────────────────────────────
    if is_excluded(libelle):
        # Récupère le motif normalisé pour afficher (la regex tape sur T1_base)
        norm = T1_base(str(libelle))
        norm = re.sub(r'[\x00-\x1f\x7f-\x9f\u200b-\u200d\ufeff]', '', norm)
        match = PATTERN_NE_PAS.search(norm)
        match_text = match.group(0) if match else "NE PAS"
        audit_steps.append(
            f"[0] EXCLU_NE_PAS_TOUCHER ✓ : motif '{match_text}' détecté "
            f"dans libellé '{libelle}' → programme exclu, pas de résolution"
        )
        return CascadeResult(
            phase=0, methode=METHODE_LABELS[0],
            rmpm="", rs_target="", cga="", nga="",
            source="LIBELLE", val_source=str(libelle), val_target=match_text,
            score=0.0, sim=None,
        )

    # ── Méthode 1 : IBAN_ACCOUNT (dans IBAN_GROUPE) ──────────────────────
    if iban and iban in ibg_iban_dict:
        rmpm, rs_t, cga, nga = ibg_iban_dict[iban]
        audit_steps.append(
            f"[1] IBAN_ACCOUNT ✓ : IBAN '{iban}' trouvé dans IBAN_GROUPE → "
            f"RMPM {rmpm} (RS cible '{rs_t}', GA {cga}/{nga})"
        )
        return CascadeResult(
            phase=1, methode=METHODE_LABELS[1],
            rmpm=rmpm, rs_target=rs_t, cga=cga, nga=nga,
            source="IBAN_GROUPE", val_source=iban, val_target=iban,
            score=compute_score(1),
        )
    elif iban:
        audit_steps.append(f"[1] IBAN_ACCOUNT : IBAN '{iban}' cherché dans IBAN_GROUPE → non trouvé")
    else:
        audit_steps.append("[1] IBAN_ACCOUNT : pas d'IBAN à tester")

    # ── Méthode 2 : IBAN_SINGLETON ───────────────────────────────────────
    if iban and iban in ibs_iban_dict:
        rmpm, rs_t = ibs_iban_dict[iban]
        audit_steps.append(
            f"[2] IBAN_SINGLETON ✓ : IBAN '{iban}' trouvé dans IBAN_SINGLETON → "
            f"RMPM {rmpm} (RS cible '{rs_t}')"
        )
        return CascadeResult(
            phase=2, methode=METHODE_LABELS[2],
            rmpm=rmpm, rs_target=rs_t, cga="", nga="",
            source="IBAN_SINGLETON", val_source=iban, val_target=iban,
            score=compute_score(2),
        )
    elif iban:
        audit_steps.append(f"[2] IBAN_SINGLETON : IBAN '{iban}' cherché dans IBAN_SINGLETON → non trouvé")
    else:
        audit_steps.append("[2] IBAN_SINGLETON : pas d'IBAN à tester")

    # ── Méthode 3 : RC_PARC_DIRECT ───────────────────────────────────────
    rc = rc_raw.strip() if rc_raw else ""
    if rc and rc in parc_rc_dict:
        rmpm, rs_t, cga, nga = parc_rc_dict[rc]
        audit_steps.append(
            f"[3] RC_PARC_DIRECT ✓ : RC '{rc}' trouvé dans PARC → "
            f"RMPM {rmpm} (RS cible '{rs_t}', GA {cga}/{nga})"
        )
        return CascadeResult(
            phase=3, methode=METHODE_LABELS[3],
            rmpm=rmpm, rs_target=rs_t, cga=cga, nga=nga,
            source="PARC", val_source=rc, val_target=rc,
            score=compute_score(3),
        )
    else:
        audit_steps.append(f"[3] RC_PARC_DIRECT : RC '{rc}' cherché dans PARC → non trouvé")

    # ── Méthode 4 : RC_PARC_VARIANTE (préfixe 00 / zfill 17) ─────────────
    variants = rc_variants(rc) if rc else []
    variants_other = [v for v in variants if v != rc]
    for v in variants_other:
        if v in parc_rc_dict:
            rmpm, rs_t, cga, nga = parc_rc_dict[v]
            audit_steps.append(
                f"[4] RC_PARC_VARIANTE ✓ : RC '{rc}' testé sous forme '{v}' "
                f"trouvé dans PARC → RMPM {rmpm} (GA {cga}/{nga})"
            )
            return CascadeResult(
                phase=4, methode=METHODE_LABELS[4],
                rmpm=rmpm, rs_target=rs_t, cga=cga, nga=nga,
                source="PARC", val_source=rc, val_target=v,
                score=compute_score(4),
            )
    if variants_other:
        audit_steps.append(
            f"[4] RC_PARC_VARIANTE : variantes {variants_other} testées dans PARC → non trouvé"
        )
    else:
        audit_steps.append("[4] RC_PARC_VARIANTE : pas de variante générable")

    # ── Méthode 5 : RC_PARC_CORRUPTION (len 14-16) ───────────────────────
    if corruption_active and rc:
        cands = corruption_candidates(rc)
        if cands:
            best = None; best_sim = 0.0
            for cand_rc, tag in cands:
                if cand_rc in parc_rc_dict:
                    rmpm, rs_t, cga, nga = parc_rc_dict[cand_rc]
                    sim = rs_similarity_L6(rs, rs_t)
                    if sim >= MIN_SIM_CORRUPTION and sim > best_sim:
                        best = (cand_rc, tag, rmpm, rs_t, cga, nga, sim)
                        best_sim = sim
            if best:
                cand_rc, tag, rmpm, rs_t, cga, nga, sim = best
                audit_steps.append(
                    f"[5] RC_PARC_CORRUPTION ✓ : RC '{rc}' (len {len(rc)}) "
                    f"reconstruit en '{cand_rc}' ({tag}) trouvé dans PARC, "
                    f"validé par sim RS {sim:.2f} ≥ {MIN_SIM_CORRUPTION} → "
                    f"RMPM {rmpm} (RS cible '{rs_t}')"
                )
                return CascadeResult(
                    phase=5, methode=METHODE_LABELS[5],
                    rmpm=rmpm, rs_target=rs_t, cga=cga, nga=nga,
                    source="PARC", val_source=rc, val_target=cand_rc,
                    score=compute_score(5, sim), sim=sim,
                )
            audit_steps.append(
                f"[5] RC_PARC_CORRUPTION : {len(cands)} candidats testés (len {len(rc)}), "
                f"aucun avec sim RS ≥ {MIN_SIM_CORRUPTION}"
            )
        else:
            audit_steps.append(
                f"[5] RC_PARC_CORRUPTION : RC '{rc}' (len {len(rc)}) hors plage 14-16, ignoré"
            )
    elif not corruption_active:
        audit_steps.append("[5] RC_PARC_CORRUPTION : désactivé par l'utilisateur")
    else:
        audit_steps.append("[5] RC_PARC_CORRUPTION : pas de RC à tester")

    # ── Méthodes 6 à 11 : cascade nettoyage RS ───────────────────────────
    # Phases 9 et 11 : filtre juridique cross-country appliqué
    if rs:
        for phase in range(6, 12):
            label, fn = NORM_FUNCS[phase]
            rs_norm = fn(rs)
            if not rs_norm:
                audit_steps.append(f"[{phase}] {label} : RS vide après nettoyage, ignoré")
                continue
            if rs_norm in rs_indexes[phase]:
                rmpm, rs_t, cga, nga, src, target_suffix = rs_indexes[phase][rs_norm]

                # Filtre juridique cross-country (phases 9 et 11)
                if phase in (9, 11):
                    if not is_juridique_compatible(query_suffix, target_suffix):
                        audit_steps.append(
                            f"[{phase}] {label} : candidat '{rs_t}' (norm '{rs_norm}') "
                            f"REJETÉ — juridique incompatible "
                            f"(source='{query_suffix or '∅'}' vs target='{target_suffix or '∅'}')"
                        )
                        continue

                audit_steps.append(
                    f"[{phase}] {label} ✓ : RS '{rs}' nettoyée en '{rs_norm}' → "
                    f"matchée dans {src} contre cible '{rs_t}' → "
                    f"RMPM {rmpm}" + (f" (GA {cga}/{nga})" if cga else "")
                )
                sim = rs_similarity_L6(rs, rs_t)
                return CascadeResult(
                    phase=phase, methode=label,
                    rmpm=rmpm, rs_target=rs_t, cga=cga, nga=nga,
                    source=src, val_source=rs_norm, val_target=rs_t,
                    score=compute_score(phase, sim), sim=sim,
                )
            audit_steps.append(f"[{phase}] {label} : '{rs}' → '{rs_norm}' → non trouvé")
    else:
        audit_steps.append("[6-11] RS_CLEAN_* : pas de RS à tester")

    # ── Méthode 12 : RS_TOKEN_JACCARD (≥ 1.00 + filtre juridique) ───────
    if rs:
        tokens_in = tokenize_rs(rs)
        if tokens_in:
            candidates = {}
            for tok in tokens_in:
                for tokens_ref, payload in token_idx.get(tok, []):
                    candidates[id(tokens_ref)] = (tokens_ref, payload)
            # Collecte des candidats qui passent le seuil, triés par jaccard desc
            qualifying = []
            for tokens_ref, payload in candidates.values():
                j = jaccard(tokens_in, tokens_ref)
                if j >= MIN_JACCARD:
                    qualifying.append((j, payload))
            qualifying.sort(key=lambda x: -x[0])

            best_jac_seen = 0.0
            n_rejected_juridique = 0
            for j, payload in qualifying:
                rmpm, rs_t, cga, nga, src, target_suffix = payload
                best_jac_seen = max(best_jac_seen, j)
                if not is_juridique_compatible(query_suffix, target_suffix):
                    n_rejected_juridique += 1
                    audit_steps.append(
                        f"[12] RS_TOKEN_JACCARD : candidat '{rs_t}' (jaccard {j:.2f}) "
                        f"REJETÉ — juridique incompatible "
                        f"(source='{query_suffix or '∅'}' vs target='{target_suffix or '∅'}')"
                    )
                    continue
                audit_steps.append(
                    f"[12] RS_TOKEN_JACCARD ✓ : tokens {sorted(tokens_in)} matchés "
                    f"dans {src} contre cible '{rs_t}' (jaccard {j:.2f} ≥ {MIN_JACCARD}) → "
                    f"RMPM {rmpm}"
                )
                return CascadeResult(
                    phase=12, methode=METHODE_LABELS[12],
                    rmpm=rmpm, rs_target=rs_t, cga=cga, nga=nga,
                    source=src, val_source=" ".join(sorted(tokens_in)),
                    val_target=rs_t,
                    score=compute_score(12, j), sim=j,
                )
            if qualifying:
                audit_steps.append(
                    f"[12] RS_TOKEN_JACCARD : {len(qualifying)} candidat(s) au seuil, "
                    f"{n_rejected_juridique} rejeté(s) par filtre juridique"
                )
            else:
                # Aucun candidat n'atteint le seuil
                best_jac_all = 0.0
                for tokens_ref, payload in candidates.values():
                    j = jaccard(tokens_in, tokens_ref)
                    best_jac_all = max(best_jac_all, j)
                audit_steps.append(
                    f"[12] RS_TOKEN_JACCARD : meilleur jaccard {best_jac_all:.2f} < {MIN_JACCARD}"
                )
        else:
            audit_steps.append("[12] RS_TOKEN_JACCARD : aucun token significatif")
    else:
        audit_steps.append("[12] RS_TOKEN_JACCARD : pas de RS")

    # ── Méthode 13 : RS_FUZZY_LEVENSHTEIN (≥ 0.95 + filtre juridique) ────
    if rs:
        rs_L6 = norm_L6(rs)
        if len(rs_L6) >= 5:
            bucket = len(rs_L6) // FUZZY_BUCKET_SIZE
            candidates = []
            for b in (bucket - 1, bucket, bucket + 1):
                candidates.extend(fuzzy_buckets.get(b, []))
            if len(candidates) > FUZZY_MAX_CANDIDATES:
                candidates = candidates[:FUZZY_MAX_CANDIDATES]
            # Collecte tous les candidats qui passent le seuil, triés par ratio desc
            qualifying = []
            best_ratio_all = 0.0
            for rs_ref_L6, payload in candidates:
                r = SequenceMatcher(None, rs_L6, rs_ref_L6).ratio()
                best_ratio_all = max(best_ratio_all, r)
                if r >= MIN_FUZZY:
                    qualifying.append((r, rs_ref_L6, payload))
            qualifying.sort(key=lambda x: -x[0])

            n_rejected_juridique = 0
            for r, rs_ref_L6, payload in qualifying:
                rmpm, rs_t, cga, nga, src, target_suffix = payload
                if not is_juridique_compatible(query_suffix, target_suffix):
                    n_rejected_juridique += 1
                    audit_steps.append(
                        f"[13] RS_FUZZY_LEVENSHTEIN : candidat '{rs_t}' (ratio {r:.2f}) "
                        f"REJETÉ — juridique incompatible "
                        f"(source='{query_suffix or '∅'}' vs target='{target_suffix or '∅'}')"
                    )
                    continue
                audit_steps.append(
                    f"[13] RS_FUZZY_LEVENSHTEIN ✓ : '{rs_L6}' fuzzy-matché "
                    f"contre '{rs_ref_L6}' dans {src} (ratio {r:.2f} ≥ {MIN_FUZZY}) → "
                    f"RMPM {rmpm} (RS cible '{rs_t}')"
                )
                return CascadeResult(
                    phase=13, methode=METHODE_LABELS[13],
                    rmpm=rmpm, rs_target=rs_t, cga=cga, nga=nga,
                    source=src, val_source=rs_L6, val_target=rs_t,
                    score=compute_score(13, r), sim=r,
                )
            if qualifying:
                audit_steps.append(
                    f"[13] RS_FUZZY_LEVENSHTEIN : {len(qualifying)} candidat(s) au seuil, "
                    f"{n_rejected_juridique} rejeté(s) par filtre juridique"
                )
            else:
                audit_steps.append(
                    f"[13] RS_FUZZY_LEVENSHTEIN : meilleur ratio {best_ratio_all:.2f} < {MIN_FUZZY}"
                )
        else:
            audit_steps.append("[13] RS_FUZZY_LEVENSHTEIN : RS trop courte (<5 chars)")
    else:
        audit_steps.append("[13] RS_FUZZY_LEVENSHTEIN : pas de RS")

    # ── Méthode 14 : RS_SUBSTRING_25 (≥25 chars) — FAIBLE + filtre ───────
    if rs:
        rs_L6 = norm_L6(rs)
        if len(rs_L6) >= NGRAM_LEN_LONG:
            n_rejected = 0
            for i in range(len(rs_L6) - NGRAM_LEN_LONG + 1):
                ng = rs_L6[i:i + NGRAM_LEN_LONG]
                if ng in ngram_long:
                    rmpm, rs_t, cga, nga, src, target_suffix = ngram_long[ng]
                    if not is_juridique_compatible(query_suffix, target_suffix):
                        n_rejected += 1
                        continue
                    audit_steps.append(
                        f"[14] {METHODE_LABELS[14]} ✓ (FAIBLE) : substring '{ng}' "
                        f"(≥{NGRAM_LEN_LONG} chars) trouvé dans {src} contre cible '{rs_t}' → "
                        f"RMPM {rmpm}"
                    )
                    return CascadeResult(
                        phase=14, methode=METHODE_LABELS[14],
                        rmpm=rmpm, rs_target=rs_t, cga=cga, nga=nga,
                        source=src, val_source=ng, val_target=rs_t,
                        score=compute_score(14),
                    )
            if n_rejected:
                audit_steps.append(
                    f"[14] {METHODE_LABELS[14]} : {n_rejected} match(es) rejeté(s) par filtre juridique"
                )
            else:
                audit_steps.append(
                    f"[14] {METHODE_LABELS[14]} : aucun substring ≥{NGRAM_LEN_LONG} chars matché"
                )
        else:
            audit_steps.append(f"[14] {METHODE_LABELS[14]} : RS trop courte (<{NGRAM_LEN_LONG})")
    else:
        audit_steps.append(f"[14] {METHODE_LABELS[14]} : pas de RS")

    # ── Méthode 15 : RS_SUBSTRING_18 (≥18 chars) — FAIBLE + filtre ───────
    if rs:
        rs_L6 = norm_L6(rs)
        if len(rs_L6) >= NGRAM_LEN_COURT:
            n_rejected = 0
            for i in range(len(rs_L6) - NGRAM_LEN_COURT + 1):
                ng = rs_L6[i:i + NGRAM_LEN_COURT]
                if ng in ngram_court:
                    rmpm, rs_t, cga, nga, src, target_suffix = ngram_court[ng]
                    if not is_juridique_compatible(query_suffix, target_suffix):
                        n_rejected += 1
                        continue
                    audit_steps.append(
                        f"[15] {METHODE_LABELS[15]} ✓ (FAIBLE) : substring '{ng}' "
                        f"(≥{NGRAM_LEN_COURT} chars) trouvé dans {src} contre cible '{rs_t}' → "
                        f"RMPM {rmpm} — À VALIDER MANUELLEMENT"
                    )
                    return CascadeResult(
                        phase=15, methode=METHODE_LABELS[15],
                        rmpm=rmpm, rs_target=rs_t, cga=cga, nga=nga,
                        source=src, val_source=ng, val_target=rs_t,
                        score=compute_score(15),
                    )
            if n_rejected:
                audit_steps.append(
                    f"[15] {METHODE_LABELS[15]} : {n_rejected} match(es) rejeté(s) par filtre juridique"
                )
            else:
                audit_steps.append(
                    f"[15] {METHODE_LABELS[15]} : aucun substring ≥{NGRAM_LEN_COURT} chars matché"
                )
        else:
            audit_steps.append(f"[15] {METHODE_LABELS[15]} : RS trop courte (<{NGRAM_LEN_COURT})")
    else:
        audit_steps.append(f"[15] {METHODE_LABELS[15]} : pas de RS")

    # ── Aucune méthode n'a matché ────────────────────────────────────────
    audit_steps.append("[99] NON_RESOLU : aucune méthode n'a retourné de résultat")
    return None


# =============================================================================
# DÉDOUBLONNAGE INTELLIGENT PAR ID_PROGRAMME
# =============================================================================

def _parse_date_safe(s):
    """Parse une date string sous formes courantes. Retourne datetime ou None."""
    if not s or not str(s).strip():
        return None
    s = str(s).strip()
    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y", "%Y%m%d", "%d.%m.%Y"]:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, dayfirst=True, errors="raise").to_pydatetime()
    except Exception:
        return None


def smart_dedup_by_id_programme(df_out):
    """
    Déduplique par ID_PROGRAMME avec règles métier précises (v10.2).

    Règles :
    - PRODUIT : concaténation alphabétique de tous les produits distincts (séparateur " / ")
    - DATE_CREATION : la plus ANCIENNE (parsée pour ordre chronologique fiable)
    - DERNIERE_APPARITION : la plus RÉCENTE (tri numérique YYYYMM)
    - LIBELLE_PROGRAMME, RS_WORLDLINE, ETAT_ACTIF :
        priorité 1 = source PRGM (ou LES_DEUX) ; priorité 2 = ACHETEUR.
        Au sein du même tier : la valeur correspondant à la DERNIERE_APPARITION
        la plus récente. C'est le principe "PRGM est l'original, ACHETEUR enrichit".
    - ID_RC : priorité PRGM, sinon le plus long (conserve les zéros initiaux)
    - SOURCE : recalculée (LES_DEUX si présent dans plusieurs sources)
    - Résolution (METHODE/RMPM/AUDIT/...) : on retient la ligne avec la meilleure
        phase (1-15 prioritaires, 0 EXCLU et 99 NON_RESOLU repoussés en queue),
        puis le meilleur SCORE_CONFIANCE, puis priorité PRGM.

    Retourne (df_dedup, n_removed).
    """
    if df_out.empty:
        return df_out.copy(), 0

    mask_with_id = (df_out["ID_PROGRAMME"].astype(str).str.strip() != "")
    df_with_id = df_out[mask_with_id].copy()
    df_no_id = df_out[~mask_with_id].copy()

    if df_with_id.empty:
        return df_out.copy(), 0

    # Conversions auxiliaires pour tri
    df_with_id["_DERN_NUM"] = pd.to_numeric(
        df_with_id["DERNIERE_APPARITION"], errors="coerce").fillna(0).astype(int)
    df_with_id["_PHASE_NUM"] = pd.to_numeric(
        df_with_id["PHASE_RESOLUTION"], errors="coerce").fillna(99).astype(int)
    df_with_id["_SCORE_NUM"] = pd.to_numeric(
        df_with_id["SCORE_CONFIANCE"], errors="coerce").fillna(0.0)

    # Priorité de source : PRGM/LES_DEUX en premier, ACHETEUR ensuite
    source_order = {"PRGM": 0, "LES_DEUX": 0, "ACHETEUR": 1}
    df_with_id["_SRC_PRIO"] = df_with_id["SOURCE"].map(source_order).fillna(2).astype(int)

    # Priorité phase : phases 1-15 prioritaires, EXCLU et NON_RESOLU en dernier
    def _phase_priority(p):
        if p == 0:  return 100   # EXCLU repoussé
        if p == 99: return 101   # NON_RESOLU encore plus loin
        return int(p)
    df_with_id["_PHASE_PRIO"] = df_with_id["_PHASE_NUM"].apply(_phase_priority)

    def _pick_priority(group, col):
        """Priorité PRGM > ACHETEUR, puis DERNIERE_APPARITION la plus récente
        dans chaque tier. Retourne la première valeur non vide."""
        g = group.sort_values(["_SRC_PRIO", "_DERN_NUM"], ascending=[True, False])
        for v in g[col]:
            vs = str(v).strip()
            if vs and vs.lower() not in ("nan", "none", "null"):
                return vs
        return ""

    def _is_valid(v):
        vs = str(v).strip()
        return vs and vs.lower() not in ("nan", "none", "null")

    aggregated = []
    for id_prog, group in df_with_id.groupby("ID_PROGRAMME", sort=False):
        row = {}
        row["ID_PROGRAMME"] = id_prog

        # SOURCE : recalculée
        unique_sources = set(group["SOURCE"].unique())
        if "LES_DEUX" in unique_sources or ({"PRGM", "ACHETEUR"} <= unique_sources):
            row["SOURCE"] = "LES_DEUX"
        elif "PRGM" in unique_sources:
            row["SOURCE"] = "PRGM"
        elif "ACHETEUR" in unique_sources:
            row["SOURCE"] = "ACHETEUR"
        else:
            row["SOURCE"] = next(iter(unique_sources)) if unique_sources else ""

        # Champs avec règle "priorité PRGM, sinon plus récent"
        row["LIBELLE_PROGRAMME"] = _pick_priority(group, "LIBELLE_PROGRAMME")
        row["RS_WORLDLINE"]      = _pick_priority(group, "RS_WORLDLINE")
        row["ETAT_ACTIF"]        = _pick_priority(group, "ETAT_ACTIF")

        # ID_RC : priorité PRGM, sinon le plus long
        prgm_group = group[group["_SRC_PRIO"] == 0]
        prgm_rcs = [str(v).strip() for v in prgm_group["ID_RC"] if _is_valid(v)]
        if prgm_rcs:
            row["ID_RC"] = max(prgm_rcs, key=len)
        else:
            all_rcs = [str(v).strip() for v in group["ID_RC"] if _is_valid(v)]
            row["ID_RC"] = max(all_rcs, key=len) if all_rcs else ""

        # PRODUIT : concaténation avec " / "
        products = sorted(set(str(p).strip() for p in group["PRODUIT"] if _is_valid(p)))
        row["PRODUIT"] = " / ".join(products)

        # DATE_CREATION : la plus ANCIENNE
        dates = [str(d).strip() for d in group["DATE_CREATION"] if _is_valid(d)]
        if dates:
            parsed = [(_parse_date_safe(d), d) for d in dates]
            parsed_ok = [(p, d) for p, d in parsed if p is not None]
            if parsed_ok:
                row["DATE_CREATION"] = min(parsed_ok, key=lambda x: x[0])[1]
            else:
                row["DATE_CREATION"] = min(dates)
        else:
            row["DATE_CREATION"] = ""

        # DERNIERE_APPARITION : la plus RÉCENTE (YYYYMM numérique)
        derns = [str(d).strip() for d in group["DERNIERE_APPARITION"]
                 if _is_valid(d) and str(d).strip() != "0"]
        if derns:
            try:
                row["DERNIERE_APPARITION"] = max(derns, key=lambda x: int(x))
            except ValueError:
                row["DERNIERE_APPARITION"] = max(derns)
        else:
            row["DERNIERE_APPARITION"] = ""

        # Résolution : EXCLU sticky (v10.3) — si au moins une ligne du groupe
        # est EXCLU (phase 0), le programme entier est EXCLU. Cela évite que
        # la dedup masque un libellé "NE_PAS_..." venant d'une source secondaire.
        excluded_rows = group[group["_PHASE_NUM"] == 0]
        if not excluded_rows.empty:
            # On prend la ligne EXCLU la plus récente
            excluded_sorted = excluded_rows.sort_values("_DERN_NUM", ascending=False)
            best = excluded_sorted.iloc[0]
            # On surcharge également LIBELLE_PROGRAMME avec celui de la ligne EXCLU
            # (c'est celui qui contient "NE_PAS_..." — l'info utile)
            lib_excluded = str(best.get("LIBELLE_PROGRAMME", "")).strip()
            if lib_excluded and lib_excluded.lower() not in ("nan", "none"):
                row["LIBELLE_PROGRAMME"] = lib_excluded
        else:
            # Tri normal : meilleure phase, puis meilleur score, puis PRGM
            g_sorted = group.sort_values(
                ["_PHASE_PRIO", "_SCORE_NUM", "_SRC_PRIO"],
                ascending=[True, False, True]
            )
            best = g_sorted.iloc[0]

        for col in ["METHODE_RESOLUTION", "PHASE_RESOLUTION",
                    "VALEUR_RESOLUTION_SOURCE", "VALEUR_RESOLUTION_CIBLE",
                    "AUDIT_RESOLUTION", "SCORE_CONFIANCE", "STATUS",
                    "RMPM_FINAL", "NOM_ENTITE_JURIDIQUE", "CODE_GA", "NOM_GA",
                    "RMPM_ABSORBE", "RMPM_SUCCESSEUR"]:
            row[col] = best.get(col, "")

        aggregated.append(row)

    df_with_id_dedup = pd.DataFrame(aggregated)
    # Aligner sur l'ordre des colonnes de df_out
    df_with_id_dedup = df_with_id_dedup[df_out.columns.tolist()]

    if not df_no_id.empty:
        df_dedup = pd.concat([df_with_id_dedup, df_no_id[df_out.columns.tolist()]],
                             ignore_index=True)
    else:
        df_dedup = df_with_id_dedup

    df_dedup = df_dedup.sort_values(["ID_PROGRAMME", "ID_RC"]).reset_index(drop=True)
    n_removed = len(df_out) - len(df_dedup)
    return df_dedup, n_removed


# =============================================================================
# APPLICATION GUI
# =============================================================================

class _Const:
    """Petit conteneur immuable pour exposer une valeur via `.get()`,
    afin de preserver a l'identique les appels `self.mode_var.get()` et
    `self.corruption_var.get()` du worker GUI d'origine."""

    __slots__ = ("_value",)

    def __init__(self, value):
        self._value = value

    def get(self):
        return self._value


class RCIdentifierAnalyzer_Q3JKL:
    """Analyseur RC Worldline — version CLI autonome (sans GUI).

    L'__init__ affecte les memes attributs que la version GUI d'origine,
    mais depuis les arguments argparse au lieu des widgets tkinter.
    """

    def __init__(self, args: argparse.Namespace):
        # Mode et options — memes attributs que la version GUI (objets
        # exposant .get() pour preserver les appels du worker).
        self.mode_var = _Const(args.mode)
        self.corruption_var = _Const(not args.no_corruption)

        # Chemins des fichiers d'entree (memes cles que la version GUI).
        self.files = {
            "PRGM": str(args.prgm) if args.prgm else "",
            "ACHETEUR": str(args.acheteur) if args.acheteur else "",
            "PARC": str(args.parc) if args.parc else "",
            "IBAN_GROUPE": str(args.iban_groupe) if args.iban_groupe else "",
            "IBAN_SINGLETON": str(args.iban_singleton) if args.iban_singleton else "",
            "ABSORBES": str(args.absorbes) if args.absorbes else "",
            "LOOKUP_INPUT": str(args.lookup_input) if args.lookup_input else "",
        }

        # Sortie : remplace asksaveasfilename (filedialog) de la GUI.
        self.output_dir = Path(args.output_dir) if args.output_dir else Path.cwd()
        self.output_filename = args.output_filename  # None => nom auto-genere

        # Mapping/selection manuelle des colonnes : etape visuelle supprimee.
        # On retombe sur les positions par defaut de l'UI (DEFAULT_POS),
        # exactement comme les CTkComboBox preselectionnes de la GUI.
        self.dfs_preview = {}
        self.original_cols = {}
        # col_mapping_vars : { cle_logique -> nom_de_colonne } resolu apres
        # lecture des entetes CSV (voir _resolve_default_columns).
        self.col_mapping_vars = {}

        # Chemin du XLSX produit (renseigne en fin de worker).
        self.last_output_path = ""

    # ------------------------------------------------------------------
    # ENTREE PRINCIPALE
    # ------------------------------------------------------------------

    def run(self) -> None:
        """Point d'entree CLI : remplace le mainloop/threading GUI.

        Resout le mapping des colonnes par defaut puis execute le worker
        (corps preserve a l'identique depuis la version GUI)."""
        self._resolve_default_columns()
        self._worker()

    # ------------------------------------------------------------------
    # MAPPING COLONNES PAR DEFAUT (ex-etape visuelle ETAPE 2 supprimee)
    # ------------------------------------------------------------------

    def _resolve_default_columns(self) -> None:
        """Reconstruit self.col_mapping_vars depuis DEFAULT_POS.

        Dans la GUI, l'utilisateur selectionnait chaque colonne via un
        CTkComboBox preselectionne sur la position DEFAULT_POS. Ici, on
        applique directement ces positions par defaut (1-indexees) sur
        les entetes lus, sans interaction. Une colonne par defaut absente
        (position > nombre de colonnes) reste non mappee (None), comme un
        combobox vide cote GUI.
        """
        # Mapping cle_logique -> cle DEFAULT_POS (identique aux _map_block GUI).
        default_keys = [
            "prgm_id", "prgm_rc", "prgm_iban", "prgm_rs",
            "prgm_libelle", "prgm_produit", "prgm_mois",
            "ach_id", "ach_libelle", "ach_rc", "ach_rs", "ach_iban",
            "ach_etat", "ach_type", "ach_periode", "ach_date_crea",
            "parc_rc", "parc_rmpm", "parc_rs", "parc_code_ga", "parc_nom_ga",
            "ibg_iso_ga", "ibg_code_ga", "ibg_nom_ga", "ibg_iso_ej",
            "ibg_rmpm", "ibg_rs", "ibg_iban",
            "ibs_iso", "ibs_rmpm", "ibs_rs", "ibs_iban",
            "abs_rmpm_src", "abs_rmpm_new", "abs_nom",
        ]
        # Fichier source de chaque groupe de cles (prefixe -> cle self.files).
        prefix_to_file = {
            "prgm": "PRGM", "ach": "ACHETEUR", "parc": "PARC",
            "ibg": "IBAN_GROUPE", "ibs": "IBAN_SINGLETON", "abs": "ABSORBES",
        }
        # Lecture des entetes (5 lignes) des fichiers requis.
        required = ["PRGM", "ACHETEUR", "PARC", "IBAN_GROUPE",
                    "IBAN_SINGLETON", "ABSORBES"]
        for fkey in required:
            path = self.files.get(fkey, "")
            if not path:
                raise ValueError(f"Fichier {fkey} obligatoire manquant.")
            self.dfs_preview[fkey] = load_csv_smart(path, nrows=5)
            self.original_cols[fkey] = list(self.dfs_preview[fkey].columns)

        for key in default_keys:
            prefix = key.split("_", 1)[0]
            fkey = prefix_to_file[prefix]
            cols = self.original_cols.get(fkey, [])
            pos = DEFAULT_POS.get(key)  # 1-indexe
            if pos and pos <= len(cols):
                self.col_mapping_vars[key] = cols[pos - 1]
            else:
                self.col_mapping_vars[key] = None

    def _get_col(self, key):
        # Renvoie le nom de colonne mappe (ex-_get_col sur CTkComboBox).
        return self.col_mapping_vars.get(key)

    def _prog(self, msg: str) -> None:
        # Remplace self.lbl_status / progressbar de la GUI par un print.
        print(msg)

    def _upd(self, value, text):
        # Remplace progressbar + label de status de la GUI : on ne
        # garde que le message textuel (la valeur de progression est
        # ignoree). Corps du worker preserve a l'identique.
        self._prog(text)

    # ------------------------------------------------------------------
    # WORKER
    # ------------------------------------------------------------------

    def _worker(self):
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            mode = self.mode_var.get()
            corruption_active = self.corruption_var.get()

            # ── 1. CHARGEMENT ──────────────────────────────────────────
            self._upd(0.02, "Chargement PRGM...")
            df_prgm = load_csv_smart(self.files["PRGM"])
            self._upd(0.04, "Chargement ACHETEUR...")
            df_ach = load_csv_smart(self.files["ACHETEUR"])
            self._upd(0.06, "Chargement PARC...")
            df_parc = load_csv_smart(self.files["PARC"])
            self._upd(0.08, "Chargement IBAN_GROUPE...")
            df_ibg = load_csv_smart(self.files["IBAN_GROUPE"])
            self._upd(0.10, "Chargement IBAN_SINGLETON...")
            df_ibs = load_csv_smart(self.files["IBAN_SINGLETON"])
            self._upd(0.12, "Chargement ABSORBÉS...")
            df_abs = load_csv_smart(self.files["ABSORBES"])

            df_lookup_input = None
            if mode == "MODE_2":
                self._upd(0.13, "Chargement LOOKUP_INPUT...")
                df_lookup_input = load_csv_smart(self.files["LOOKUP_INPUT"])

            # ── 2. PRGM ────────────────────────────────────────────────
            self._upd(0.14, "Préparation PRGM...")
            col_id_p   = self._get_col("prgm_id")
            col_rc_p   = self._get_col("prgm_rc")
            col_iban_p = self._get_col("prgm_iban")
            col_rs_p   = self._get_col("prgm_rs")
            col_lib_p  = self._get_col("prgm_libelle")
            col_prod_p = self._get_col("prgm_produit")
            col_mois_p = self._get_col("prgm_mois")

            df_prgm["_ID"]       = clean_id_safe(df_prgm[col_id_p])
            df_prgm["_RC_RAW"]   = clean_id_safe(df_prgm[col_rc_p])
            df_prgm["_PROD"]     = df_prgm[col_prod_p].astype(str).str.strip()
            df_prgm["_MOIS"]     = df_prgm[col_mois_p].astype(str).str.strip()
            df_prgm["_IBAN_P"]   = clean_iban(df_prgm[col_iban_p])
            df_prgm["_RS_ORIG"]  = df_prgm[col_rs_p].astype(str).str.strip()
            df_prgm["_LIB"]      = df_prgm[col_lib_p].astype(str).str.strip() if col_lib_p else ""
            df_prgm["_MOIS_NUM"] = pd.to_numeric(df_prgm["_MOIS"], errors="coerce").fillna(0).astype(int)

            df_prgm_sorted = df_prgm.sort_values("_MOIS_NUM", ascending=False)
            df_prgm_dd = df_prgm_sorted.drop_duplicates(subset=["_RC_RAW", "_PROD"], keep="first").copy()
            last_seen = (df_prgm.groupby("_RC_RAW")["_MOIS_NUM"]
                         .max().reset_index()
                         .rename(columns={"_MOIS_NUM": "_DERNIERE_APPARITION"}))
            df_prgm_dd = df_prgm_dd.merge(last_seen, on="_RC_RAW", how="left")
            df_prgm_dd["DERNIERE_APPARITION"] = df_prgm_dd["_DERNIERE_APPARITION"].apply(
                lambda v: str(v) if v and v > 0 else "")
            df_prgm_dd["_SOURCE"] = "PRGM"
            df_prgm_dd["_ETAT_ACTIF"] = "OUI"
            df_prgm_dd["_DATE_CREATION"] = ""

            # ── 3. ACHETEUR ────────────────────────────────────────────
            self._upd(0.16, "Préparation ACHETEUR...")
            col_id_a      = self._get_col("ach_id")
            col_lib_a     = self._get_col("ach_libelle")
            col_rc_a      = self._get_col("ach_rc")
            col_rs_a      = self._get_col("ach_rs")
            col_iban_a    = self._get_col("ach_iban")
            col_etat      = self._get_col("ach_etat")
            col_type_a    = self._get_col("ach_type")
            col_periode   = self._get_col("ach_periode")
            col_date_crea = self._get_col("ach_date_crea")

            df_ach["_ID"]       = clean_id_safe(df_ach[col_id_a])
            df_ach["_LIB"]      = df_ach[col_lib_a].astype(str).str.strip() if col_lib_a else ""
            df_ach["_RC_RAW"]   = clean_id_safe(df_ach[col_rc_a])
            df_ach["_PROD"]     = df_ach[col_type_a].astype(str).str.strip()
            df_ach["_MOIS"]     = df_ach[col_periode].astype(str).str.strip()
            df_ach["_IBAN_P"]   = clean_iban(df_ach[col_iban_a])
            df_ach["_RS_ORIG"]  = df_ach[col_rs_a].astype(str).str.strip()
            df_ach["_MOIS_NUM"] = pd.to_numeric(df_ach["_MOIS"], errors="coerce").fillna(0).astype(int)

            df_ach_sorted = df_ach.sort_values("_MOIS_NUM", ascending=False)
            df_ach_dd = df_ach_sorted.drop_duplicates(subset=["_RC_RAW", "_PROD"], keep="first").copy()
            df_ach_dd["DERNIERE_APPARITION"] = df_ach_dd["_MOIS_NUM"].apply(
                lambda v: str(v) if v and v > 0 else "")
            df_ach_dd["_SOURCE"] = "ACHETEUR"
            etat_raw = df_ach_dd[col_etat].astype(str).str.strip().str.upper()
            df_ach_dd["_ETAT_ACTIF"] = np.where(
                etat_raw.isin(["OUI", "ACTIF", "TRUE", "1", "O", "Y", "YES"]), "OUI", "NON")
            df_ach_dd["_DATE_CREATION"] = df_ach_dd[col_date_crea].astype(str).str.strip()

            # Dictionnaires d'enrichissement ACHETEUR
            ach_lib_dict, ach_date_dict, ach_etat_dict, ach_id_to_rc = {}, {}, {}, {}
            for idp, lib, dc, rc in zip(df_ach["_ID"].values,
                                         df_ach["_LIB"].values,
                                         df_ach[col_date_crea].astype(str).str.strip().values,
                                         df_ach["_RC_RAW"].values):
                if idp and idp not in ach_lib_dict:
                    ach_lib_dict[idp]  = lib
                    ach_date_dict[idp] = dc
                    ach_id_to_rc[idp]  = rc
            for idp, etat in zip(df_ach_dd["_ID"].values, df_ach_dd["_ETAT_ACTIF"].values):
                if idp and idp not in ach_etat_dict:
                    ach_etat_dict[idp] = etat

            prgm_id_to_rc, prgm_id_to_iban = {}, {}
            for idp, rc, iban in zip(df_prgm["_ID"].values,
                                      df_prgm["_RC_RAW"].values,
                                      df_prgm["_IBAN_P"].values):
                if idp and idp not in prgm_id_to_rc:
                    prgm_id_to_rc[idp] = rc
                    prgm_id_to_iban[idp] = iban
            ach_id_to_iban = {}
            for idp, iban in zip(df_ach["_ID"].values, df_ach["_IBAN_P"].values):
                if idp and idp not in ach_id_to_iban:
                    ach_id_to_iban[idp] = iban

            # ── 4. df_union ────────────────────────────────────────────
            if mode == "MODE_1":
                self._upd(0.18, "Fusion PRGM + ACHETEUR...")
                cols_common = ["_ID", "_LIB", "_RC_RAW", "_PROD", "_IBAN_P", "_RS_ORIG",
                               "_MOIS_NUM", "DERNIERE_APPARITION", "_SOURCE",
                               "_ETAT_ACTIF", "_DATE_CREATION"]
                df_p = df_prgm_dd[cols_common].copy()
                df_a = df_ach_dd[cols_common].copy()
                df_p["_FUSION_KEY"] = df_p["_RC_RAW"] + "|" + df_p["_PROD"]
                df_a["_FUSION_KEY"] = df_a["_RC_RAW"] + "|" + df_a["_PROD"]
                keys_prgm = set(df_p["_FUSION_KEY"].values)
                keys_ach  = set(df_a["_FUSION_KEY"].values)
                keys_both = keys_prgm & keys_ach
                df_p["_SOURCE"] = np.where(df_p["_FUSION_KEY"].isin(keys_both), "LES_DEUX", "PRGM")

                for idx in df_p.index:
                    idp = df_p.at[idx, "_ID"]
                    if idp and idp in ach_date_dict:
                        if not df_p.at[idx, "_DATE_CREATION"]:
                            df_p.at[idx, "_DATE_CREATION"] = ach_date_dict[idp]
                        if idp in ach_etat_dict:
                            df_p.at[idx, "_ETAT_ACTIF"] = ach_etat_dict[idp]
                        if idp in ach_lib_dict and not df_p.at[idx, "_LIB"]:
                            df_p.at[idx, "_LIB"] = ach_lib_dict[idp]

                df_a_only = df_a[~df_a["_FUSION_KEY"].isin(keys_prgm)].copy()
                df_union = pd.concat([df_p, df_a_only], ignore_index=True)
                df_union.drop(columns=["_FUSION_KEY"], inplace=True)
            else:
                self._upd(0.18, "Construction df_union depuis LOOKUP_INPUT...")
                if df_lookup_input is None or df_lookup_input.shape[1] < 4:
                    raise ValueError("LOOKUP_INPUT doit avoir au moins 4 colonnes.")
                cols_lookup = list(df_lookup_input.columns)
                lk_id  = clean_id_safe(df_lookup_input[cols_lookup[0]]).values
                lk_lib = df_lookup_input[cols_lookup[1]].astype(str).str.strip().values
                lk_rs  = df_lookup_input[cols_lookup[2]].astype(str).str.strip().values

                rc_resolved, iban_resolved, source_resolved = [], [], []
                for idp in lk_id:
                    rc = iban = src = ""
                    if idp in ach_id_to_rc and ach_id_to_rc[idp]:
                        rc = ach_id_to_rc[idp]
                        iban = ach_id_to_iban.get(idp, "")
                        src = "ACHETEUR"
                    elif idp in prgm_id_to_rc and prgm_id_to_rc[idp]:
                        rc = prgm_id_to_rc[idp]
                        iban = prgm_id_to_iban.get(idp, "")
                        src = "PRGM"
                    rc_resolved.append(rc)
                    iban_resolved.append(iban)
                    source_resolved.append(src)

                df_union = pd.DataFrame({
                    "_ID":       lk_id, "_LIB": lk_lib,
                    "_RC_RAW":   rc_resolved, "_PROD": [""] * len(lk_id),
                    "_IBAN_P":   iban_resolved,
                    "_RS_ORIG":  lk_rs, "_MOIS_NUM": [0] * len(lk_id),
                    "DERNIERE_APPARITION": [""] * len(lk_id),
                    "_SOURCE":   source_resolved,
                    "_ETAT_ACTIF": ["N/A"] * len(lk_id),
                    "_DATE_CREATION": [""] * len(lk_id),
                })

            n = len(df_union)
            n_prgm_only = int((df_union["_SOURCE"] == "PRGM").sum())
            n_ach_only  = int((df_union["_SOURCE"] == "ACHETEUR").sum())
            n_both      = int((df_union["_SOURCE"] == "LES_DEUX").sum())
            self._upd(0.20, f"{n:,} lignes")

            # ── 5. Dictionnaires IBAN ──────────────────────────────────
            self._upd(0.22, "Dictionnaires IBAN...")
            col_ibg_code_ga = self._get_col("ibg_code_ga")
            col_ibg_nom_ga  = self._get_col("ibg_nom_ga")
            col_ibg_rmpm    = self._get_col("ibg_rmpm")
            col_ibg_rs      = self._get_col("ibg_rs")
            col_ibg_iban    = self._get_col("ibg_iban")

            df_ibg["_IBAN_G"]    = clean_iban(df_ibg[col_ibg_iban])
            df_ibg["_RMPM_G"]    = clean_id_safe(df_ibg[col_ibg_rmpm])
            df_ibg["_RS_ORIG_G"] = df_ibg[col_ibg_rs].astype(str).str.strip()
            df_ibg["_CODE_GA"]   = clean_id_safe(df_ibg[col_ibg_code_ga])
            df_ibg["_NOM_GA"]    = df_ibg[col_ibg_nom_ga].astype(str).str.strip()

            df_ibg_dd = df_ibg[df_ibg["_IBAN_G"] != ""].drop_duplicates("_IBAN_G").copy()
            ibg_iban_dict = {}
            for iban_g, rmpm_g, rs_g, code_ga, nom_ga in zip(
                df_ibg_dd["_IBAN_G"].values, df_ibg_dd["_RMPM_G"].values,
                df_ibg_dd["_RS_ORIG_G"].values, df_ibg_dd["_CODE_GA"].values,
                df_ibg_dd["_NOM_GA"].values,
            ):
                if iban_g:
                    ibg_iban_dict[iban_g] = (str(rmpm_g), str(rs_g), str(code_ga), str(nom_ga))

            df_ibg_rev_dd = df_ibg[df_ibg["_RMPM_G"] != ""].drop_duplicates("_RMPM_G").copy()
            rmpm_to_ga_dict = {}
            for rmpm_g, code_ga, nom_ga in zip(
                df_ibg_rev_dd["_RMPM_G"].values, df_ibg_rev_dd["_CODE_GA"].values,
                df_ibg_rev_dd["_NOM_GA"].values,
            ):
                if rmpm_g:
                    rmpm_to_ga_dict[rmpm_g] = (str(code_ga), str(nom_ga))

            col_ibs_rmpm = self._get_col("ibs_rmpm")
            col_ibs_rs   = self._get_col("ibs_rs")
            col_ibs_iban = self._get_col("ibs_iban")
            df_ibs["_IBAN_S"]    = clean_iban(df_ibs[col_ibs_iban])
            df_ibs["_RMPM_S"]    = clean_id_safe(df_ibs[col_ibs_rmpm])
            df_ibs["_RS_ORIG_S"] = df_ibs[col_ibs_rs].astype(str).str.strip()
            df_ibs_dd = df_ibs[df_ibs["_IBAN_S"] != ""].drop_duplicates("_IBAN_S").copy()
            ibs_iban_dict = {}
            for iban_s, rmpm_s, rs_s in zip(
                df_ibs_dd["_IBAN_S"].values, df_ibs_dd["_RMPM_S"].values,
                df_ibs_dd["_RS_ORIG_S"].values,
            ):
                if iban_s:
                    ibs_iban_dict[iban_s] = (str(rmpm_s), str(rs_s))

            rmpm_to_rs_dict = {}
            for r, s in zip(df_ibs_dd["_RMPM_S"].values, df_ibs_dd["_RS_ORIG_S"].values):
                if r and r not in rmpm_to_rs_dict:
                    rmpm_to_rs_dict[r] = s
            for r, s in zip(df_ibg_rev_dd["_RMPM_G"].values, df_ibg_rev_dd["_RS_ORIG_G"].values):
                if r:
                    rmpm_to_rs_dict[r] = s

            # ── 6. PARC ────────────────────────────────────────────────
            self._upd(0.26, "Dictionnaire PARC...")
            col_rc_parc      = self._get_col("parc_rc")
            col_rmpm_parc    = self._get_col("parc_rmpm")
            col_rs_parc      = self._get_col("parc_rs")
            col_code_ga_parc = self._get_col("parc_code_ga")
            col_nom_ga_parc  = self._get_col("parc_nom_ga")

            df_parc["_RC_P"]      = clean_id_safe(df_parc[col_rc_parc])
            df_parc["_RMPM_P"]    = clean_id_safe(df_parc[col_rmpm_parc])
            df_parc["_RS_ORIG_P"] = df_parc[col_rs_parc].astype(str).str.strip()
            if col_code_ga_parc and col_code_ga_parc in df_parc.columns:
                df_parc["_CODE_GA_P"] = clean_id_safe(df_parc[col_code_ga_parc])
            else:
                df_parc["_CODE_GA_P"] = ""
            if col_nom_ga_parc and col_nom_ga_parc in df_parc.columns:
                df_parc["_NOM_GA_P"]  = df_parc[col_nom_ga_parc].astype(str).str.strip()
            else:
                df_parc["_NOM_GA_P"]  = ""

            df_parc_rc = df_parc[df_parc["_RC_P"] != ""].drop_duplicates("_RC_P").copy()
            parc_rc_dict = {}
            for rc_p, rmpm_p, rs_p, code_ga_p, nom_ga_p in zip(
                df_parc_rc["_RC_P"].values, df_parc_rc["_RMPM_P"].values,
                df_parc_rc["_RS_ORIG_P"].values, df_parc_rc["_CODE_GA_P"].values,
                df_parc_rc["_NOM_GA_P"].values,
            ):
                if rc_p:
                    parc_rc_dict[rc_p] = (str(rmpm_p), str(rs_p),
                                          str(code_ga_p), str(nom_ga_p))

            # ── 7. ABSORBÉS ────────────────────────────────────────────
            self._upd(0.28, "Dictionnaire absorbés...")
            col_abs_src = self._get_col("abs_rmpm_src")
            col_abs_new = self._get_col("abs_rmpm_new")
            df_abs["_SRC_A"] = clean_id_safe(df_abs[col_abs_src])
            df_abs["_NEW_A"] = clean_id_safe(df_abs[col_abs_new])
            df_abs_dd = df_abs[df_abs["_SRC_A"] != ""].drop_duplicates("_SRC_A").copy()
            abs_dict = df_abs_dd.set_index("_SRC_A")["_NEW_A"].to_dict()

            # ── 8. INDEX RS (T1..T6 + tokens + fuzzy + ngrams) ─────────
            self._upd(0.30, "Construction indexes RS (T1→T6, tokens, fuzzy, ngrams)...")
            sources_rs = [
                ("PARC",           df_parc, "_RS_ORIG_P", "_RMPM_P", "_CODE_GA_P", "_NOM_GA_P"),
                ("IBAN_GROUPE",    df_ibg,  "_RS_ORIG_G", "_RMPM_G", "_CODE_GA",   "_NOM_GA"),
                ("IBAN_SINGLETON", df_ibs,  "_RS_ORIG_S", "_RMPM_S", None,         None),
            ]
            rs_indexes, token_idx, fuzzy_buckets, ngram_long, ngram_court = build_rs_indexes(sources_rs)
            self._upd(0.34, "Indexes construits.")

            # ── 9. PIPELINE ────────────────────────────────────────────
            self._upd(0.36, "Pipeline cascade (16 méthodes)...")

            id_arr        = df_union["_ID"].values
            lib_arr       = df_union["_LIB"].values
            rc_arr        = df_union["_RC_RAW"].values
            iban_arr      = df_union["_IBAN_P"].values
            rs_orig_arr   = df_union["_RS_ORIG"].values
            prod_arr      = df_union["_PROD"].values
            source_arr    = df_union["_SOURCE"].values
            etat_arr      = df_union["_ETAT_ACTIF"].values
            date_crea_arr = df_union["_DATE_CREATION"].values
            dern_app_arr  = df_union["DERNIERE_APPARITION"].values

            col_methode = np.empty(n, dtype=object)
            col_phase   = np.empty(n, dtype=object)
            col_val_src = np.empty(n, dtype=object)
            col_val_tgt = np.empty(n, dtype=object)
            col_audit   = np.empty(n, dtype=object)
            col_score   = np.empty(n, dtype=object)
            col_status  = np.empty(n, dtype=object)
            col_rmpm_provisoire = np.empty(n, dtype=object)
            col_nom_ej_raw = np.empty(n, dtype=object)
            col_cga_raw = np.empty(n, dtype=object)
            col_nga_raw = np.empty(n, dtype=object)
            col_source_match = np.empty(n, dtype=object)  # nom de la base où on a trouvé

            for i in range(n):
                if i % 1500 == 0:
                    self._upd(0.36 + 0.45 * i / n, f"Analyse {i:,}/{n:,}...")

                audit_steps = []
                result = run_cascade(
                    libelle=lib_arr[i],
                    rc_raw=rc_arr[i],
                    iban=iban_arr[i],
                    rs=rs_orig_arr[i],
                    audit_steps=audit_steps,
                    ibg_iban_dict=ibg_iban_dict,
                    ibs_iban_dict=ibs_iban_dict,
                    parc_rc_dict=parc_rc_dict,
                    rs_indexes=rs_indexes,
                    token_idx=token_idx,
                    fuzzy_buckets=fuzzy_buckets,
                    ngram_long=ngram_long,
                    ngram_court=ngram_court,
                    corruption_active=corruption_active,
                )

                if result is None:
                    col_methode[i] = f"99.{METHODE_LABELS[99]}"
                    col_phase[i] = 99
                    col_val_src[i] = ""
                    col_val_tgt[i] = ""
                    col_score[i] = "0.00"
                    col_status[i] = "NON_RESOLU"
                    col_rmpm_provisoire[i] = ""
                    col_nom_ej_raw[i] = ""
                    col_cga_raw[i] = ""
                    col_nga_raw[i] = ""
                    col_source_match[i] = ""
                else:
                    col_methode[i] = f"{result.phase}.{result.methode}"
                    col_phase[i] = result.phase
                    col_val_src[i] = result.val_source
                    col_val_tgt[i] = result.val_target
                    col_score[i] = f"{result.score:.2f}"
                    col_status[i] = status_from_phase_and_score(result.phase, result.score)
                    col_rmpm_provisoire[i] = result.rmpm
                    col_nom_ej_raw[i] = result.rs_target
                    col_cga_raw[i] = result.cga
                    col_nga_raw[i] = result.nga
                    col_source_match[i] = result.source

                col_audit[i] = "\n".join(audit_steps)

            # ── 10. POST-TRAITEMENT (absorption + enrichissement GA) ──
            self._upd(0.82, "Post-traitement (absorption, enrichissement)...")
            col_rmpm_final = np.empty(n, dtype=object)
            col_nom_ej     = np.empty(n, dtype=object)
            col_cga        = np.empty(n, dtype=object)
            col_nga        = np.empty(n, dtype=object)
            col_absorbe    = np.empty(n, dtype=object)
            col_successeur = np.empty(n, dtype=object)

            for i in range(n):
                phase = col_phase[i]
                rmpm_p = col_rmpm_provisoire[i]

                # Cas EXCLU ou NON_RESOLU
                if phase in (0, 99) or not rmpm_p:
                    col_rmpm_final[i] = ""
                    col_nom_ej[i]     = ""
                    col_cga[i]        = ""
                    col_nga[i]        = ""
                    col_absorbe[i]    = "N/A"
                    col_successeur[i] = ""
                    continue

                # Absorption
                if rmpm_p in abs_dict:
                    successeur = abs_dict[rmpm_p]
                    col_absorbe[i]    = "OUI"
                    col_successeur[i] = successeur
                    rmpm_definitif = successeur
                else:
                    col_absorbe[i]    = "NON"
                    col_successeur[i] = ""
                    rmpm_definitif = rmpm_p

                col_rmpm_final[i] = rmpm_definitif

                # NOM_ENTITE_JURIDIQUE : priorité au RS cible directement trouvé
                if col_nom_ej_raw[i]:
                    col_nom_ej[i] = col_nom_ej_raw[i]
                else:
                    col_nom_ej[i] = rmpm_to_rs_dict.get(rmpm_definitif, "")

                # CODE_GA / NOM_GA : si on a déjà cga/nga depuis la résolution, on garde
                # Sinon reverse lookup
                if col_cga_raw[i]:
                    col_cga[i] = col_cga_raw[i]
                    col_nga[i] = col_nga_raw[i]
                elif rmpm_definitif in rmpm_to_ga_dict:
                    cga_r, nga_r = rmpm_to_ga_dict[rmpm_definitif]
                    col_cga[i] = cga_r
                    col_nga[i] = nga_r
                else:
                    col_cga[i] = ""
                    col_nga[i] = ""

            # ── 11. df_out (22 colonnes) ───────────────────────────────
            self._upd(0.85, "Construction df_out...")
            df_out = pd.DataFrame({
                # IDENTIFICATION (9)
                "SOURCE":              source_arr,
                "ID_PROGRAMME":        id_arr,
                "LIBELLE_PROGRAMME":   lib_arr,
                "ID_RC":               rc_arr,
                "PRODUIT":             prod_arr,
                "RS_WORLDLINE":        rs_orig_arr,
                "ETAT_ACTIF":          etat_arr,
                "DATE_CREATION":       date_crea_arr,
                "DERNIERE_APPARITION": dern_app_arr,
                # RÉSOLUTION (7)
                "METHODE_RESOLUTION":      col_methode,
                "PHASE_RESOLUTION":        col_phase,
                "VALEUR_RESOLUTION_SOURCE": col_val_src,
                "VALEUR_RESOLUTION_CIBLE":  col_val_tgt,
                "AUDIT_RESOLUTION":        col_audit,
                "SCORE_CONFIANCE":         col_score,
                "STATUS":                  col_status,
                # ENRICHISSEMENT (6)
                "RMPM_FINAL":           col_rmpm_final,
                "NOM_ENTITE_JURIDIQUE": col_nom_ej,
                "CODE_GA":              col_cga,
                "NOM_GA":               col_nga,
                "RMPM_ABSORBE":         col_absorbe,
                "RMPM_SUCCESSEUR":      col_successeur,
            })

            # ── 12. DÉDOUBLONNAGE INTELLIGENT v10.2 ───────────────────
            # Règles métier :
            # - 1 ligne par ID_PROGRAMME
            # - PRODUIT : concat de tous les produits (" / ")
            # - DATE_CREATION : plus ancienne | DERNIERE_APPARITION : plus récente
            # - LIBELLE/RS/ETAT/IBAN : priorité PRGM > ACHETEUR, puis plus récent
            # - Résolution : ligne avec meilleure phase + meilleur score
            self._upd(0.87, "Dédoublonnage intelligent par ID_PROGRAMME...")
            df_dedup, n_dedup_removed = smart_dedup_by_id_programme(df_out)
            n_dedup = len(df_dedup)

            # ── 13. STATS PAR PHASE ───────────────────────────────────
            self._upd(0.88, "Stats par phase...")
            stats_phases = {p: int((col_phase == p).sum()) for p in range(0, 16)}
            stats_phases[99] = int((col_phase == 99).sum())
            n_excluded = stats_phases.get(0, 0)
            n_resolved = sum(stats_phases[p] for p in range(1, 16))
            n_non_resolu = stats_phases[99]
            n_resolved_faible = sum(stats_phases[p] for p in PHASES_FAIBLES)
            n_resolved_ferme = n_resolved - n_resolved_faible

            stats_status = {s: int((col_status == s).sum())
                            for s in ("RESOLU", "RESOLU_MOYEN", "RESOLU_FAIBLE",
                                       "NON_RESOLU", "EXCLU")}

            # ── 14. EXPORT XLSX ────────────────────────────────────────
            self._upd(0.89, "Sélection fichier...")
            default_name = f"RC_ANALYSIS_{ts}_{VERSION_ID}{'_M2' if mode == 'MODE_2' else ''}.xlsx"
            # Remplace filedialog.asksaveasfilename : on construit le chemin
            # de sortie depuis self.output_dir / self.output_filename.
            out_name = self.output_filename if self.output_filename else default_name
            self.output_dir.mkdir(parents=True, exist_ok=True)
            save_path = str(self.output_dir / out_name)

            self._upd(0.91, "Écriture XLSX...")
            wb = Workbook()
            self._write_data_sheet(wb, df_out, n)
            self._write_synthese_sheet(wb, df_dedup, n, n_dedup)
            self._write_analyse_sheet(wb, mode, n, n_dedup,
                                       n_prgm_only, n_ach_only, n_both,
                                       stats_phases, stats_status,
                                       n_excluded, n_resolved, n_resolved_ferme,
                                       n_resolved_faible, n_non_resolu)
            self._write_livrable_sheet(wb, df_dedup, n_dedup)
            if mode == "MODE_2":
                self._write_mode2_sheet(wb, df_out, n)

            self._upd(0.97, "Sauvegarde...")
            wb.save(save_path)
            self._upd(1.0, "Terminé !")

            # Recapitulatif final (remplace messagebox.showinfo de la GUI).
            self._prog(
                "Analyse terminee.\n"
                f"Fichier cree : {save_path}\n\n"
                f"Mode : {mode}  |  Corruption RC : {'ON' if corruption_active else 'OFF'}\n\n"
                f"BRUT : {n:,} lignes  ->  DEDUPLIQUE : {n_dedup:,}\n\n"
                "REPARTITION PAR PHASE (brut) :\n"
                + "\n".join(f"  [{p:2d}] {METHODE_LABELS[p]:<28} : {stats_phases.get(p, 0):,}"
                             for p in [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,99])
                + "\n\nSTATUS :\n"
                f"  RESOLU         : {stats_status['RESOLU']:,}\n"
                f"  RESOLU_MOYEN   : {stats_status['RESOLU_MOYEN']:,}\n"
                f"  RESOLU_FAIBLE  : {stats_status['RESOLU_FAIBLE']:,}\n"
                f"  EXCLU          : {stats_status['EXCLU']:,}\n"
                f"  NON_RESOLU     : {stats_status['NON_RESOLU']:,}\n"
            )
            # Memorise le chemin produit pour le message [OK] de main().
            self.last_output_path = save_path

        except Exception:
            # Propage vers main() qui gere les codes de sortie (1/2).
            self._upd(0, "Erreur")
            raise

    # ------------------------------------------------------------------
    # ÉCRITURE XLSX — helpers
    # ------------------------------------------------------------------

    def _fill(self, hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def _font(self, color="FFFFFF", bold=True, size=10):
        return Font(name="Segoe UI", size=size, bold=bold, color=color)

    def _hdr_align(self):
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _row_fill_status(self, status):
        if status == "EXCLU":          return self._fill(GREY2)
        if status == "RESOLU":         return self._fill(GRN2)
        if status == "RESOLU_MOYEN":   return self._fill(YLO2)
        if status == "RESOLU_FAIBLE":  return self._fill(ORG2)
        if status == "NON_RESOLU":     return self._fill(RED2)
        return None

    def _write_columns_with_groups(self, ws, df, groups, id_cols, header_row=4,
                                    status_col_name="STATUS"):
        """Écrit un DataFrame avec en-têtes groupés et coloration par status.
        groups = [(group_label, [col_names], header_color), ...]
        """
        start_col = 2
        col_layout = []
        for group_label, cols, color in groups:
            for j, cname in enumerate(cols):
                xl_col = start_col + j
                c_grp = ws.cell(row=header_row, column=xl_col,
                                value=group_label if j == 0 else "")
                c_grp.fill = self._fill(color); c_grp.font = self._font(size=9)
                c_grp.alignment = Alignment(horizontal="left")
                c_h = ws.cell(row=header_row + 1, column=xl_col, value=cname)
                c_h.fill = self._fill(GRN); c_h.font = self._font(size=9)
                c_h.alignment = self._hdr_align()
                # Largeur
                if cname == "AUDIT_RESOLUTION":     w = 90
                elif cname in ("LIBELLE_PROGRAMME", "RS_WORLDLINE",
                                "NOM_ENTITE_JURIDIQUE", "NOM_GA"): w = 30
                elif cname in ("VALEUR_RESOLUTION_SOURCE", "VALEUR_RESOLUTION_CIBLE",
                                "METHODE_RESOLUTION"): w = 28
                else: w = 18
                ws.column_dimensions[get_column_letter(xl_col)].width = w
                col_layout.append((cname, xl_col))
            start_col += len(cols)

        # Body
        for ri, (_, row) in enumerate(df.iterrows()):
            xl_row = header_row + 2 + ri
            status = str(row.get(status_col_name, ""))
            fill = self._row_fill_status(status)
            for cname, xl_col in col_layout:
                v = row.get(cname, "")
                vs = str(v) if v is not None and v != "" else ""
                if cname in id_cols and vs:
                    vs = excel_id(vs)
                cell = ws.cell(row=xl_row, column=xl_col, value=vs)
                # AUDIT_RESOLUTION en wrap
                if cname == "AUDIT_RESOLUTION":
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                if fill and cname == status_col_name:
                    cell.fill = fill
        ws.freeze_panes = f"B{header_row + 2}"

    def _write_data_sheet(self, wb, df_out, n_brut):
        ws = wb.active
        ws.title = "DATA"
        ws.column_dimensions["A"].width = 2.5
        ws["B2"] = (f"RC ANALYSIS [{VERSION_ID}] — {datetime.now().strftime('%d/%m/%Y %H:%M')} "
                    f"— Brut {n_brut:,} lignes")
        ws["B2"].font = Font(name="Segoe UI", bold=True, size=13, color=DARK)
        ID_COLS = {"ID_PROGRAMME", "ID_RC", "RMPM_FINAL", "RMPM_SUCCESSEUR"}
        groups = [
            ("IDENTIFICATION",
             ["SOURCE", "ID_PROGRAMME", "LIBELLE_PROGRAMME", "ID_RC", "PRODUIT",
              "RS_WORLDLINE", "ETAT_ACTIF", "DATE_CREATION", "DERNIERE_APPARITION"], GRN3),
            ("RÉSOLUTION",
             ["METHODE_RESOLUTION", "PHASE_RESOLUTION",
              "VALEUR_RESOLUTION_SOURCE", "VALEUR_RESOLUTION_CIBLE",
              "AUDIT_RESOLUTION", "SCORE_CONFIANCE", "STATUS"], PUR),
            ("ENRICHISSEMENT",
             ["RMPM_FINAL", "NOM_ENTITE_JURIDIQUE", "CODE_GA", "NOM_GA",
              "RMPM_ABSORBE", "RMPM_SUCCESSEUR"], DARK),
        ]
        self._write_columns_with_groups(ws, df_out, groups, ID_COLS, header_row=4)

    def _write_synthese_sheet(self, wb, df_dedup, n_brut, n_dedup):
        ws = wb.create_sheet(title="SYNTHESE_WORLDLINE")
        ws.column_dimensions["A"].width = 2.5
        ws["B2"] = (f"SYNTHÈSE WORLDLINE [{VERSION_ID}] — "
                    f"{n_brut:,} brutes → {n_dedup:,} lignes uniques")
        ws["B2"].font = Font(name="Segoe UI", bold=True, size=13, color=DARK)
        ws["B3"] = ("Vue dédupliquée par ID_PROGRAMME (RC le plus long retenu). "
                    "Coloration par STATUS.")
        ws["B3"].font = Font(name="Segoe UI", italic=True, size=10, color="666666")
        ID_COLS = {"ID_PROGRAMME", "ID_RC", "RMPM_FINAL", "RMPM_SUCCESSEUR"}
        groups = [
            ("IDENTIFICATION",
             ["SOURCE", "ID_PROGRAMME", "LIBELLE_PROGRAMME", "ID_RC", "PRODUIT",
              "RS_WORLDLINE", "ETAT_ACTIF"], GRN3),
            ("RÉSOLUTION",
             ["METHODE_RESOLUTION", "PHASE_RESOLUTION",
              "VALEUR_RESOLUTION_SOURCE", "VALEUR_RESOLUTION_CIBLE",
              "SCORE_CONFIANCE", "STATUS"], PUR),
            ("ENRICHISSEMENT",
             ["RMPM_FINAL", "NOM_ENTITE_JURIDIQUE", "CODE_GA", "NOM_GA",
              "RMPM_ABSORBE", "RMPM_SUCCESSEUR"], DARK),
            ("AUDIT",
             ["AUDIT_RESOLUTION"], BLU),
        ]
        self._write_columns_with_groups(ws, df_dedup, groups, ID_COLS, header_row=5)

    def _write_analyse_sheet(self, wb, mode, n, n_dedup,
                              n_prgm_only, n_ach_only, n_both,
                              stats_phases, stats_status,
                              n_excluded, n_resolved, n_resolved_ferme,
                              n_resolved_faible, n_non_resolu):
        ws = wb.create_sheet(title="ANALYSE")
        ws.column_dimensions["A"].width = 2.5
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 32
        ws.column_dimensions["D"].width = 80
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 10

        ws["B2"] = f"ANALYSE [{VERSION_ID}] — Mode {mode[-1]}"
        ws["B2"].font = Font(name="Segoe UI", bold=True, size=14, color=DARK)
        ws["B3"] = f"Brut : {n:,} lignes  |  Dédupliqué : {n_dedup:,}"
        ws["B3"].font = Font(name="Segoe UI", italic=True, size=10, color="666666")

        def pct(v):
            if n == 0 or not isinstance(v, (int, float)) or v == "": return ""
            return f"{100 * v / n:.1f}%"

        # --- Section 1 : périmètre ---
        ws["B5"] = "PÉRIMÈTRE"
        ws["B5"].font = self._font(color="FFFFFF", bold=True, size=11)
        ws["B5"].fill = self._fill(DARK)
        for col in "BCDEF":
            ws[f"{col}5"].fill = self._fill(DARK)

        ws["B6"] = "Total brut";       ws["E6"] = n;            ws["F6"] = pct(n)
        ws["B7"] = "Total dédupliqué"; ws["E7"] = n_dedup;      ws["F7"] = ""
        ws["B8"] = "  · PRGM only";    ws["E8"] = n_prgm_only;  ws["F8"] = pct(n_prgm_only)
        ws["B9"] = "  · ACHETEUR only";ws["E9"] = n_ach_only;   ws["F9"] = pct(n_ach_only)
        ws["B10"]= "  · LES_DEUX";     ws["E10"]= n_both;       ws["F10"]= pct(n_both)

        # --- Section 2 : phases détaillées avec descriptions ---
        ws["B12"] = "MÉTHODE"
        ws["C12"] = "NOM"
        ws["D12"] = "DESCRIPTION"
        ws["E12"] = "VOLUME"
        ws["F12"] = "%"
        for col in "BCDEF":
            ws[f"{col}12"].font = self._font(color="FFFFFF", bold=True, size=10)
            ws[f"{col}12"].fill = self._fill(GRN)
            ws[f"{col}12"].alignment = self._hdr_align()

        DESCRIPTIONS = {
            0:  "Le libellé programme contient le motif 'NE PAS' (variantes : NE PAS, NE_PAS, ne pas, NEPAS). Le programme est explicitement marqué à exclure côté métier — aucune tentative de résolution.",
            1:  "L'IBAN du programme est cherché directement dans la base IBAN_GROUPE (table groupe d'affaires). Match exact = résolution la plus fiable.",
            2:  "L'IBAN du programme est cherché dans la base IBAN_SINGLETON (entités isolées sans groupe). Match exact = résolution fiable.",
            3:  "Le RC tel quel (longueur 17) est cherché directement dans le Parc Client. Match exact = résolution fiable.",
            4:  "Si le RC direct n'a pas matché, on tente avec deux variantes : préfixe '00' ajouté, ou zfill à 17 caractères. Couvre les RC stockés sans zéros initiaux.",
            5:  "Reconstruction d'un RC tronqué. Len 16 → ajout '0' devant. Len 15 → ajout '00' devant. Len 14 → bruteforce '00' + RC + un digit 0-9. Chaque candidat est validé par sim RS ≥ 0.75 (durci v10.1) pour limiter les faux positifs.",
            6:  "Match exact sur la Raison Sociale après simple uppercase + suppression d'accents (T1). C'est le niveau de nettoyage minimum.",
            7:  "Match exact sur la RS après T1 + suppression ponctuation (., - _ / & ' \" ( ) ; :) + normalisation espaces. Couvre 'S.A.R.L.' = 'SARL'.",
            8:  "Comme RS_CLEAN_PUNCT + strip suffixe devise en fin de RS (EUR, USD, GBP, CHF, et 49 autres). Couvre 'ACME EUR' = 'ACME'.",
            9:  "Comme RS_CLEAN_PUNCT + strip suffixe juridique en fin de RS (SARL, SA, GMBH, LTD, LDT, LIMITED, INC, SRL, SPA, NV, BV, AG, et ~35 autres). Couvre 'ACME SARL' = 'ACME'. ⚠️ Filtre juridique cross-country actif : un match 'BOPRO SAS' vs 'BOPRO NV' est REJETÉ (jurisdictions FR vs NL).",
            10: "Comme RS_CLEAN_PUNCT + normalisation pays vers code ISO (FRANCE→FR, GERMANY→DE, ITALIE→IT, ROYAUME UNI→GB, etc., ~20 pays). Couvre 'ACME FRANCE' ↔ 'ACME FR'.",
            11: "Pipeline de nettoyage complet : T1 + T2 + T3 + T4 + T5 + T6. Strip répété en boucle jusqu'à stabilité. Méthode 'catch-all' du nettoyage RS. ⚠️ Filtre juridique cross-country actif (cf. méthode 9).",
            12: "Tokenisation de la RS (mots de ≥3 chars, hors stop-words : DE, DU, LA, LE, OF, THE, GROUP, FRANCE, EUROPE, etc.). Match si Jaccard = 1.00 STRICT (durci v10.3, ex-0.80) — tous les tokens significatifs doivent être identiques entre source et cible. ⚠️ Filtre juridique cross-country actif. ⚠️ Classée RESOLU_FAIBLE (v10.3).",
            13: "Distance de Levenshtein (SequenceMatcher) sur la RS nettoyée niveau L6. Match si ratio ≥ 0.95. Capture quasi-uniquement les typos évidentes (1-2 lettres). Bucketé par longueur (±5 chars) pour performance. ⚠️ Filtre juridique cross-country actif. ⚠️ Classée RESOLU_FAIBLE (v10.3).",
            14: "Substring d'au moins 25 caractères de la RS source trouvé dans une RS cible. Seuil très restrictif, adapté aux noms longs (ministères, sociétés à raison sociale étendue, joint ventures). ⚠️ Filtre juridique cross-country actif. ⚠️ Classée RESOLU_FAIBLE.",
            15: "Substring d'au moins 18 caractères. Plus permissif que la méthode 14 mais reste assez sélectif. Conçu pour rattraper les noms longs partiellement tronqués. ⚠️ Filtre juridique cross-country actif. ⚠️ Classée RESOLU_FAIBLE — À VALIDER.",
            99: "Aucune des 15 méthodes précédentes n'a retourné de résultat. Le programme reste non résolu.",
        }

        row_i = 13
        for phase in [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,99]:
            volume = stats_phases.get(phase, 0)
            label = METHODE_LABELS[phase]
            desc = DESCRIPTIONS[phase]
            ws.cell(row=row_i, column=2, value=phase)
            ws.cell(row=row_i, column=3, value=label)
            ws.cell(row=row_i, column=4, value=desc)
            ws.cell(row=row_i, column=5, value=volume)
            ws.cell(row=row_i, column=6, value=pct(volume))
            ws.cell(row=row_i, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row_i, column=2).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row=row_i, column=5).alignment = Alignment(horizontal="right", vertical="top")
            ws.cell(row=row_i, column=6).alignment = Alignment(horizontal="right", vertical="top")
            # Coloration : exclu = gris, faible = orange, non résolu = rouge
            if phase == 0:
                bg = GREY2
            elif phase in PHASES_FAIBLES:
                bg = ORG2
            elif phase == 99:
                bg = RED2
            else:
                bg = GRN2 if volume > 0 else "FFFFFF"
            for c in "BCDEF":
                ws[f"{c}{row_i}"].fill = self._fill(bg)
            ws.row_dimensions[row_i].height = 50
            row_i += 1

        # --- Section 3 : status récapitulatif ---
        row_i += 1
        ws.cell(row=row_i, column=2, value="STATUS DE CONFIANCE")
        for c in "BCDEF":
            ws[f"{c}{row_i}"].font = self._font(color="FFFFFF", bold=True, size=11)
            ws[f"{c}{row_i}"].fill = self._fill(DARK)
        row_i += 1
        for status, vol, descr, bg in [
            ("RESOLU",         stats_status["RESOLU"],
             "Score ≥ 0.75. Résolution forte.", GRN2),
            ("RESOLU_MOYEN",   stats_status["RESOLU_MOYEN"],
             "Score 0.50-0.75. Résolution acceptable.", YLO2),
            ("RESOLU_FAIBLE",  stats_status["RESOLU_FAIBLE"],
             "Score < 0.50 OU méthode 14/15 (substring). À vérifier.", ORG2),
            ("EXCLU",          stats_status["EXCLU"],
             "Programme exclu (méthode 0 : NE PAS TOUCHER).", GREY2),
            ("NON_RESOLU",     stats_status["NON_RESOLU"],
             "Aucune méthode n'a matché.", RED2),
        ]:
            ws.cell(row=row_i, column=3, value=status)
            ws.cell(row=row_i, column=4, value=descr)
            ws.cell(row=row_i, column=5, value=vol)
            ws.cell(row=row_i, column=6, value=pct(vol))
            ws.cell(row=row_i, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row_i, column=5).alignment = Alignment(horizontal="right")
            ws.cell(row=row_i, column=6).alignment = Alignment(horizontal="right")
            for c in "BCDEF":
                ws[f"{c}{row_i}"].fill = self._fill(bg)
            row_i += 1

    def _write_livrable_sheet(self, wb, df_dedup, n_dedup):
        """WORLDLINE_LIVRABLE : 6 colonnes finales avec format Inconnu1234."""
        ws = wb.create_sheet(title="WORLDLINE_LIVRABLE")
        ws.column_dimensions["A"].width = 2.5
        ws["B2"] = f"WORLDLINE LIVRABLE — Format final — {datetime.now().strftime('%d/%m/%Y')}"
        ws["B2"].font = Font(name="Segoe UI", bold=True, size=13, color=DARK)
        ws["B3"] = ("Format final pour Worldline. Si RMPM non trouvé : 'Inconnu<ID_PROGRAMME>' "
                    "(sans underscore/tiret/espace).")
        ws["B3"].font = Font(name="Segoe UI", italic=True, size=10, color="666666")

        headers = ["ID_PROGRAMME", "LIBELLE", "ID_RC", "RMPM_TROUVE", "METHODE", "RMPM"]
        widths  = [22, 40, 22, 14, 28, 24]
        for j, (h, w) in enumerate(zip(headers, widths)):
            c = ws.cell(row=5, column=2 + j, value=h)
            c.fill = self._fill(GRN); c.font = self._font(size=10)
            c.alignment = self._hdr_align()
            ws.column_dimensions[get_column_letter(2 + j)].width = w

        n_trouve = 0; n_non_trouve = 0
        for ri, (_, row) in enumerate(df_dedup.iterrows()):
            row_idx = ri + 6
            id_v = str(row["ID_PROGRAMME"]).strip() if row["ID_PROGRAMME"] else ""
            rc_v = str(row["ID_RC"]).strip() if row["ID_RC"] else ""
            rmpm_v = str(row["RMPM_FINAL"]).strip() if row["RMPM_FINAL"] else ""
            methode = str(row["METHODE_RESOLUTION"]) if row["METHODE_RESOLUTION"] else ""

            if is_placeholder_rmpm(rmpm_v):
                id_clean = sanitize_id_for_inconnu(id_v)
                rmpm_display = f"Inconnu{id_clean}" if id_clean else "Inconnu"
                rmpm_trouve = "NON"
                is_real_id = False
                n_non_trouve += 1
            else:
                rmpm_display = rmpm_v
                rmpm_trouve = "OUI"
                is_real_id = True
                n_trouve += 1

            ws.cell(row=row_idx, column=2, value=excel_id(id_v) if id_v else "")
            ws.cell(row=row_idx, column=3, value=str(row["LIBELLE_PROGRAMME"]))
            ws.cell(row=row_idx, column=4, value=excel_id(rc_v) if rc_v else "")
            ws.cell(row=row_idx, column=5, value=rmpm_trouve)
            ws.cell(row=row_idx, column=6, value=methode)
            if is_real_id:
                ws.cell(row=row_idx, column=7, value=excel_id(rmpm_display))
            else:
                ws.cell(row=row_idx, column=7, value=rmpm_display)

            # Coloration ligne
            status = str(row["STATUS"])
            fill = self._row_fill_status(status)
            if fill:
                for c in range(2, 8):
                    ws.cell(row=row_idx, column=c).fill = fill

        ws.freeze_panes = "B6"
        ws.auto_filter.ref = f"B5:G{5 + n_dedup}"

    def _write_mode2_sheet(self, wb, df_out, n):
        ws = wb.create_sheet(title="MODE2_LOOKUP_RESULT")
        ws.column_dimensions["A"].width = 2.5
        ws["B2"] = "MODE 2 — Lookup result (ordre d'origine)"
        ws["B2"].font = Font(name="Segoe UI", bold=True, size=13, color=DARK)
        headers = ["ID_PROGRAMME", "LIBELLE", "RAISON_SOCIALE", "RMPM_TROUVE", "METHODE", "RMPM"]
        widths  = [22, 40, 36, 14, 28, 24]
        for j, (h, w) in enumerate(zip(headers, widths)):
            c = ws.cell(row=5, column=2 + j, value=h)
            c.fill = self._fill(GRN); c.font = self._font(size=10)
            c.alignment = self._hdr_align()
            ws.column_dimensions[get_column_letter(2 + j)].width = w

        for ri in range(n):
            row = df_out.iloc[ri]
            row_idx = ri + 6
            id_v = str(row["ID_PROGRAMME"]).strip() if row["ID_PROGRAMME"] else ""
            rmpm_v = str(row["RMPM_FINAL"]).strip() if row["RMPM_FINAL"] else ""
            methode = str(row["METHODE_RESOLUTION"]) if row["METHODE_RESOLUTION"] else ""

            if is_placeholder_rmpm(rmpm_v):
                id_clean = sanitize_id_for_inconnu(id_v)
                rmpm_display = f"Inconnu{id_clean}" if id_clean else "Inconnu"
                rmpm_trouve = "NON"; is_real_id = False
            else:
                rmpm_display = rmpm_v
                rmpm_trouve = "OUI"; is_real_id = True

            ws.cell(row=row_idx, column=2, value=excel_id(id_v) if id_v else "")
            ws.cell(row=row_idx, column=3, value=str(row["LIBELLE_PROGRAMME"]))
            ws.cell(row=row_idx, column=4, value=str(row["RS_WORLDLINE"]))
            ws.cell(row=row_idx, column=5, value=rmpm_trouve)
            ws.cell(row=row_idx, column=6, value=methode)
            if is_real_id:
                ws.cell(row=row_idx, column=7, value=excel_id(rmpm_display))
            else:
                ws.cell(row=row_idx, column=7, value=rmpm_display)

            fill = self._row_fill_status(str(row["STATUS"]))
            if fill:
                for c in range(2, 8):
                    ws.cell(row=row_idx, column=c).fill = fill

        ws.freeze_panes = "B6"
        ws.auto_filter.ref = f"B5:G{5 + n}"


# =============================================================================
# CLI — argparse / main
# =============================================================================

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="01.Q3JKL.py",
        description=("RC Identifier Analyzer [Q3JKL] — analyseur d'identifiants "
                     "RC Worldline (BNP Paribas, Direction Monetique). "
                     "Pipeline 16 methodes + synthese de resolution."),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--mode", choices=["MODE_1", "MODE_2"], default="MODE_1",
                        help="Mode de fonctionnement.")

    # Entrees obligatoires (toujours requises).
    parser.add_argument("--prgm", type=Path, required=False, default=None,
                        help="CSV PRGM (Worldline consolide).")
    parser.add_argument("--acheteur", type=Path, required=False, default=None,
                        help="CSV ACHETEUR.")
    parser.add_argument("--parc", type=Path, required=False, default=None,
                        help="CSV PARC CLIENT (RC -> RMPM).")
    parser.add_argument("--iban-groupe", type=Path, required=False, default=None,
                        help="CSV IBAN_GROUPE (FORTIS groupe).")
    parser.add_argument("--iban-singleton", type=Path, required=False, default=None,
                        help="CSV IBAN_SINGLETON (FORTIS singleton).")
    parser.add_argument("--absorbes", type=Path, required=False, default=None,
                        help="CSV RMPM ABSORBES (source -> nouveau).")

    # Entree conditionnelle (obligatoire en MODE_2 seulement).
    parser.add_argument("--lookup-input", type=Path, default=None,
                        help="CSV 4 colonnes (positions absolues). "
                             "Obligatoire avec --mode MODE_2.")

    # Option metier (methode 5).
    parser.add_argument("--no-corruption", action="store_true",
                        help="Desactive la methode 5 (RC_PARC_CORRUPTION).")

    # Sortie.
    parser.add_argument("--output-dir", type=Path, default=None,
                        help="Repertoire de sortie (defaut : repertoire courant).")
    parser.add_argument("--output-filename", type=str, default=None,
                        help="Nom du fichier XLSX (defaut : auto-genere).")
    return parser


def main(argv=None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    # Validation des arguments (code de sortie 2 = mauvais usage).
    if args.mode == "MODE_2" and args.lookup_input is None:
        parser.error("--lookup-input est obligatoire avec --mode MODE_2.")

    if args.mode == "MODE_2" and not Path(args.lookup_input).is_file():
        parser.error(f"--lookup-input : fichier introuvable : {args.lookup_input}")

    # Execution (code de sortie 1 = erreur d'execution).
    try:
        # Auto-résolution des sources OBLIGATOIRES depuis 03.sources/ si absentes.
        if not args.prgm:
            args.prgm = resolve_source("PRGM_AGREGE", required=True)
        if not args.acheteur:
            args.acheteur = resolve_source("FICHIER_ACHETEUR", required=True)
        if not args.parc:
            args.parc = resolve_source("PARC_CLIENT", required=True)
        if not args.iban_groupe:
            args.iban_groupe = resolve_source("IBAN_ACCOUNT", required=True)
        if not args.iban_singleton:
            args.iban_singleton = resolve_source("IBAN_SINGLETON", required=True)
        if not args.absorbes:
            args.absorbes = resolve_source("ABSORBES", required=True)

        app = RCIdentifierAnalyzer_Q3JKL(args)
        app.run()
    except FileNotFoundError as exc:
        print(f"[ERREUR] {exc}", file=sys.stderr)
        return 1
    except Exception as exc:  # noqa: BLE001 — frontiere CLI
        import traceback
        traceback.print_exc()
        print(f"[ERREUR] {exc}", file=sys.stderr)
        return 1

    out = getattr(app, "last_output_path", "")
    print(f"[OK] Analyse terminee — fichier produit : {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())