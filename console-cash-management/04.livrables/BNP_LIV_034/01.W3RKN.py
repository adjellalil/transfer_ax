
"""
CIB REVENUS ANALYZER v34 [W3RKN] — CLI

DESCRIPTION
    Analyseur des revenus clients CIB de BNP Paribas Cash Management (Direction
    Monetique). Livrable BNP_LIV_034. Croise les flux CCO (Monext) et CPC
    (Worldline) avec les pivots d'identite (IBAN_ACCOUNT, REFERENTIEL_CLIENT,
    IDENTIFIANT_SEGMENT) pour reconstituer, par client, le nombre de cartes, les
    flux, le PNB, l'exposition reelle (ACTUAL_EXPOSURE) et l'utilisation, avec
    une cascade de matching et une recuperation etendue (PARC_CLIENT, SINGLETON,
    nom exact, RP/RC toutes combinaisons). Conversion des devises Worldline en
    EUR, gestion explicite du "NON CALCULABLE", agregats trimestriels/annuels,
    snapshots EXPOSURE/CUSTOMERS et sortie XLSX en nombres natifs.

    Cette version est une transposition CLI (argparse) de l'application GUI
    customtkinter d'origine. La LOGIQUE METIER (helpers, cascade de matching,
    calculs, colonnes de sortie, XLSX) est PRESERVEE A L'IDENTIQUE. Seules les
    couches d'interface (customtkinter / tkinter, dialogues, threading GUI) ont
    ete retirees et remplacees par des arguments de ligne de commande.

    Le mapping/selection manuelle des colonnes (etape 2 de l'ancienne UI) et le
    cleaning interactif des valeurs non mappees (etape 3) ne sont plus visuels :
    on applique les positions par defaut (DEFAULT_POSITIONS) qui etaient
    preselectionnees dans l'UI, et aucun mapping manuel additionnel n'est
    fourni (self.manual_map reste vide). Voir commentaires "[CLI]" dans le code.

SOURCES REQUISES (obligatoires)
    --source         00. MONITORING_CIB / ONBORDING (base a enrichir)
    --worldline      05. PRGM Worldline (flux CPC)
    --monext         04. MONEXT (flux CCO)
    --account        01. IBAN_ACCOUNT (pivot prioritaire 1)
    --ref-client     02. REFERENTIEL_CLIENT (pivot priorite 2)
    --idseg          03. IDENTIFIANT_SEGMENT (pivot priorite 3)
    --devises        09. DEVISES (Date YYYYMM | Code | Taux vers EUR)
    --type-of-card   08. CLEAN_TYPE_OF_CARD (Type Original -> New)

SOURCES OPTIONNELLES
    --bejo-cartes    12. BEJO_CARTES (entite | nb cartes)        [requiert --use-bejo]
    --bejo-flux      13. BEJO_FLUX   (entite | IBAN)             [requiert --use-bejo]
    --country        06. CLEAN_COUNTRY (Pays Original -> New)
    --sales          07. CLEAN_SALES   (Sales Original -> New)
    --parc-client    14. PARC_CLIENT (RMPM -> ID_RC + ID_RP)
    --singleton      15. SINGLETON   (Nom entite -> IBAN)

OUTPUTS PRODUITS
    <output-dir>/<output-filename>.csv    Tableau complet (sep ';', utf-8-sig)
    <output-dir>/<output-filename>.xlsx   Classeur DICTIONNAIRE / SYNTHESE /
                                          REVENU GLOBAL / REVENU CCO (MONEXT) /
                                          REVENU CPC (WORLDLINE), nombres natifs.
    (XLSX uniquement si openpyxl est disponible.)

ARGUMENTS CLI
    Fichiers       : voir SOURCES REQUISES / OPTIONNELLES ci-dessus.
    Sortie         : --output-dir (dossier), --output-filename (base sans extension).
    Options metier :
      --use-bejo                 Active l'integration BEJO (requiert bejo-cartes/flux).
      --ref-mois MOIS            Mois de reference (nom FR, ex. "Decembre"). Def. Decembre.
      --ref-annee ANNEE          Annee de reference (ex. 2025). Def. 2025.

DECOMPOSITION (arborescent)
    main()
    |-- argparse : lecture des 8 flags obligatoires + 6 optionnels + sortie + options
    |-- validation des chemins / coherence BEJO
    |-- CIBRevenusAnalyzerCLI(args)
    |   |-- __init__       : remplit self.files{...}, options, self.manual_map (vide)
    |   |-- _build_mapping : positions par defaut -> noms de colonnes (ex-UI etape 2)
    |   |-- run()          : prepare les arguments du worker puis appelle worker(...)
    |       |-- worker(...) : [VERBATIM] coeur du traitement
    |           |-- chargement des sources (load_csv_smart)
    |           |-- cleaning SOURCE (mapping pays/sales/type of card)
    |           |-- indexation des pivots (ACCOUNT / REF_CLIENT / IDSEG)
    |           |-- recuperation etendue (PARC -> RC/RP, nom -> IBAN)
    |           |-- indexation des flux MX / WL + conversion devises -> EUR
    |           |-- resolution d'identite (cascade) + matching des flux
    |           |-- exposition / utilisation "non calculable"
    |           |-- agregats trimestriels / annuels
    |           |-- ecriture CSV (output-dir / output-filename)
    |           |-- generate_xlsx(...) : DICTIONNAIRE / SYNTHESE / detail
    |-- sorties : 0 = OK, 1 = erreur de traitement, 2 = erreur d'arguments
"""

import argparse
import math
import os
import re
import re as _re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ─── Résolution auto des sources + lecture DuckDB (inliné depuis 06.fonctions) ──
def _find_sources_root():
    here = Path(__file__).resolve()
    for d in [here] + list(here.parents):
        c = d / "03.sources"
        if c.is_dir():
            return c
    return None


def resolve_source(name, required=False):
    root = _find_sources_root()
    if not root:
        if required:
            raise FileNotFoundError(f"Dossier 03.sources introuvable depuis {__file__}")
        return None
    folder = next((m for m in root.glob("*/" + name) if m.is_dir()), None)
    if not folder:
        if required:
            raise FileNotFoundError(f"Source introuvable : 03.sources/*/{name}")
        return None
    exts = (".csv", ".xlsx", ".xls", ".txt")
    files = [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() in exts]
    if not files:
        if required:
            raise FileNotFoundError(f"Aucun fichier source dans {folder}")
        return None

    def _np(f):
        m = re.match(r"\s*(\d+)", f.name)
        return (int(m.group(1)) if m else -1, f.name)
    return max(files, key=_np)


def _read_duck(path, nrows=None):
    p = Path(path)
    if nrows is not None or p.suffix.lower() not in (".csv", ".txt", ""):
        return None
    try:
        import duckdb
        return duckdb.read_csv(str(p), all_varchar=True).df()
    except Exception:
        return None


# =============================================================================
# CONFIGURATION
# =============================================================================
BNP_GREEN = "#00915A"
VERSION_ID = "W3RKN"

MOIS_NOMS = [
    'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
    'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'
]
MOIS_LABELS = {
    '01': 'JANVIER',  '02': 'FEVRIER',   '03': 'MARS',
    '04': 'AVRIL',    '05': 'MAI',       '06': 'JUIN',
    '07': 'JUILLET',  '08': 'AOUT',      '09': 'SEPTEMBRE',
    '10': 'OCTOBRE',  '11': 'NOVEMBRE',  '12': 'DECEMBRE'
}

# Positions par défaut des colonnes dans chaque fichier source (présélection UI).
DEFAULT_POSITIONS = {
    # 00 — MONITORING CIB (SOURCE / ONBORDING à enrichir)
    'source_statut':                6,   # col 6 = statut ("Confirmé") — snapshots
    'source_pays_1':                3,
    'source_pays_2':                4,
    'source_sales':                10,
    'source_type_of_card':         16,
    'source_rs':                    7,   # col 7 = nom entité (match exact possible vs IBAN col 6 / SINGLETON)
    'source_rc':                    7,
    'source_rp':                   12,
    'source_rmpm':                 15,
    'source_iban':                  0,   # SOURCE peut ne pas avoir d'IBAN
    'source_differe':              18,   # col S Bénédicte = source 18
    'source_plafond':              20,
    'source_exposition_theorique': 32,   # col AG Bénédicte = source 32

    # 01 — IBAN_ACCOUNT (pivot maître)
    'acc_pays_ga':  1, 'acc_code_ga':  2, 'acc_nom_ga':   3,
    'acc_pays_le':  4, 'acc_rmpm':     5, 'acc_nom_le':   6, 'acc_iban':     7,

    # 02 — REFERENTIEL_CLIENT (Olivier)
    'ref_rp':      1, 'ref_rib':     2, 'ref_devise':  3, 'ref_rmpm':    4,
    'ref_siren':   5, 'ref_rs':      6, 'ref_rc':      7, 'ref_segment': 8,
    'ref_marche':  9, 'ref_code_ga': 10, 'ref_nom_ga':  11,

    # 03 — IDSEG (segment uniquement)
    'idseg_type':    1, 'idseg_id':      2, 'idseg_segment': 3,

    # 04 — MONEXT
    'monext_mois':          1, 'monext_rs':            4, 'monext_id_rp':         9,
    'monext_id_rc':        10, 'monext_iban':         11, 'monext_nb_cartes':    12,
    'monext_depenses':     15, 'monext_retraits':     17, 'monext_pnb_first':    19,
    'monext_pnb_last':     55, 'monext_pnb_exclude':  33, 'monext_interchange':  21,

    # 05 — PRGM Worldline
    'worldline_mois':        2, 'worldline_rs':          8, 'worldline_iban':        9,
    'worldline_id_rc':      40, 'worldline_devise':     12, 'worldline_nb_cartes':  17,
    'worldline_depenses_1': 28, 'worldline_depenses_2': 29, 'worldline_plafond':    13,
    'worldline_periodicite': 14,

    # 09 — DEVISES (taux de change vers EUR)
    'devise_date':  1, 'devise_code':  2, 'devise_taux':  3,

    # 06/07/08 — Cleaning
    'country_original': 2, 'country_new': 3,
    'sales_original':   1, 'sales_new':   2,
    'toc_original':     1, 'toc_new':     2,

    # 12/13 — BEJO
    'bejo_cartes_entite': 1, 'bejo_cartes_nb':  3,
    'bejo_flux_entite':   1, 'bejo_flux_iban':  6,

    # 14 — PARC_CLIENT (NOUVEAU, optionnel) : RMPM → ID_RC + ID_RP (même ligne)
    #   positions DEVINÉES — à vérifier/ajuster.
    'parc_rmpm': 1, 'parc_rc': 2, 'parc_rp': 3,

    # 15 — IBAN_SINGLETON (NOUVEAU, optionnel) : nom entité → IBAN (même esprit que IBAN_ACCOUNT)
    #   positions DEVINÉES — à vérifier/ajuster.
    'sing_nom_le': 6, 'sing_iban': 7,
}

DEFAULT_PNB_WORLDLINE   = [30, 31, 32, 33, 34, 35, 36]
RS_INCLUSION_MIN        = 14
COL_GAP = 2

# =============================================================================
# CODE COULEUR DES BLOCS (palette sobre BNP — reprise VX7M4)
# =============================================================================
BLOC_COLORS = {
    'A_SOURCE':    {'name': 'Colonnes SOURCE originales', 'header': '4A5560', 'banner': '8A95A0', 'cell': 'F5F6F7', 'font': 'FFFFFF'},
    'B_IDENTITY':  {'name': 'Identité client',            'header': '1F4E79', 'banner': '4A78A8', 'cell': 'EAF1F8', 'font': 'FFFFFF'},
    'C_MATCHING':  {'name': 'Méthode de matching',        'header': '00665E', 'banner': '4E8E88', 'cell': 'E6EFEE', 'font': 'FFFFFF'},
    'D1_NB_MX':    {'name': 'Nombre de cartes MONEXT (CCO)', 'header': '4A4458', 'banner': '78708E', 'cell': 'E8E5ED', 'font': 'FFFFFF'},
    'D2_FLUX_MX':  {'name': 'Flux MONEXT (CCO)',          'header': '5C5470', 'banner': '8A82A0', 'cell': 'EFEDF2', 'font': 'FFFFFF'},
    'D3_PNB_MX':   {'name': 'PNB MONEXT (CCO)',           'header': '6F6488', 'banner': '9A91AE', 'cell': 'F2EFF6', 'font': 'FFFFFF'},
    'E1_NB_WL':    {'name': 'Nombre de cartes WORLDLINE (CPC)', 'header': '583618', 'banner': '85603D', 'cell': 'EEE6DC', 'font': 'FFFFFF'},
    'E2_FLUX_WL':  {'name': 'Flux WORLDLINE (CPC)',       'header': '6B4423', 'banner': '9C7A55', 'cell': 'F2EDE6', 'font': 'FFFFFF'},
    'E3_PNB_WL':   {'name': 'PNB WORLDLINE (CPC)',        'header': '7E5530', 'banner': 'AE8A66', 'cell': 'F5F0EA', 'font': 'FFFFFF'},
    'F_EXPOSITION':{'name': 'Exposition + utilisation + composantes', 'header': '0D3B66', 'banner': '4A6A8E', 'cell': 'E6ECF2', 'font': 'FFFFFF'},
    'G_DIAG':      {'name': 'Diagnostic', 'header': '595959', 'banner': '8C8C8C', 'cell': 'F7F7F7', 'font': 'FFFFFF'},
}

# Couleurs onglets ANALYSE
WHT = "FFFFFF"; DARK = "1C3A2D"
GRN  = BLOC_COLORS['C_MATCHING']['header']; GRN2 = BLOC_COLORS['C_MATCHING']['cell']
BLU  = BLOC_COLORS['B_IDENTITY']['header']; BLU2 = BLOC_COLORS['B_IDENTITY']['cell']
PUR  = BLOC_COLORS['D2_FLUX_MX']['header']; PUR2 = BLOC_COLORS['D2_FLUX_MX']['cell']
ORG  = BLOC_COLORS['E2_FLUX_WL']['header']; ORG2 = BLOC_COLORS['E2_FLUX_WL']['cell']


# =============================================================================
# ANALYSEUR (CLI) — sans GUI, logique métier préservée à l'identique
# =============================================================================
class CIBRevenusAnalyzerCLI:

    def __init__(self, args: argparse.Namespace) -> None:
        # [CLI] Remplace l'ancien __init__(self) qui montait l'UI customtkinter.
        # Les attributs portent les mêmes noms / structures que la version GUI.
        self.files: Dict[str, str] = {
            "SOURCE":      args.source      or "",
            "WORLDLINE":   args.worldline   or "",
            "MONEXT":      args.monext      or "",
            "ACCOUNT":     args.account     or "",
            "REF_CLIENT":  args.ref_client  or "",
            "IDSEG":       args.idseg       or "",
            "DEVISES":     args.devises     or "",
            "BEJO_CARTES": args.bejo_cartes or "",
            "BEJO_FLUX":   args.bejo_flux   or "",
            "TYPE_OF_CARD": args.type_of_card or "",
            "COUNTRY":     args.country     or "",
            "SALES":       args.sales       or "",
            # NOUVEAU (V2) — récupération étendue, optionnels
            "PARC_CLIENT": args.parc_client or "",
            "SINGLETON":   args.singleton   or "",
        }

        self.use_bejo: bool = bool(args.use_bejo)
        self.output_dir: Path = Path(args.output_dir)
        self.output_filename: str = args.output_filename

        self.ref_mois: str = args.ref_mois
        self.ref_annee: str = str(args.ref_annee)

        # [CLI] Mapping manuel des valeurs non mappées (ex-étape 3, interactif).
        # En CLI on ne propose pas de cleaning interactif : table vide -> seules
        # les correspondances présentes dans les fichiers de mapping s'appliquent.
        self.manual_map: Dict[str, Dict[str, str]] = {'COUNTRY': {}, 'SALES': {}, 'TYPE_OF_CARD': {}}

        # Colonnes détectées (équivalent self.original_cols de l'UI) + dictionnaire m{}.
        self.original_cols: Dict[str, List[str]] = {}
        self.dfs_preview: Dict[str, Any] = {}

    # -------------------------------------------------------------------------
    # CHARGEMENT CSV (VERBATIM — identique à la version GUI)
    # -------------------------------------------------------------------------
    def load_csv_smart(self, path, nrows=None):
        _d = _read_duck(path, nrows)
        if _d is not None:
            return _d
        for sep in [';', ',', '\t']:
            for enc in ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']:
                try:
                    df_t = pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                       keep_default_na=False, na_values=[],
                                       on_bad_lines='skip', nrows=5)
                    if df_t.shape[1] > 1:
                        return pd.read_csv(path, sep=sep, encoding=enc, dtype=str,
                                           keep_default_na=False, na_values=[],
                                           on_bad_lines='skip', nrows=nrows)
                except Exception:
                    continue
        return pd.read_csv(path, sep=None, engine='python', dtype=str,
                           keep_default_na=False, na_values=[],
                           on_bad_lines='skip', nrows=nrows)

    # -------------------------------------------------------------------------
    # MAPPING DES COLONNES (ex-étape 2 UI) -> positions par défaut
    # -------------------------------------------------------------------------
    def _build_mapping(self):
        """[CLI] Reconstruit le dictionnaire `m` (mapping_key -> nom de colonne)
        à partir des positions par défaut DEFAULT_POSITIONS, exactement comme la
        présélection des CTkComboBox dans l'ancienne étape 2 (étape visuelle
        supprimée). Renvoie aussi mx_pnb_cfg, wl_pnb_cols, ref_yyyymm, differe_col
        et expo_cfg construits sur les mêmes valeurs par défaut."""
        # Lecture des colonnes de chaque fichier fourni (équivalent load_previews).
        keys = ["SOURCE", "ACCOUNT", "REF_CLIENT", "IDSEG", "MONEXT", "WORLDLINE",
                "DEVISES", "TYPE_OF_CARD"]
        for opt in ["COUNTRY", "SALES"]:
            if self.files.get(opt):
                keys.append(opt)
        if self.use_bejo:
            keys += ["BEJO_CARTES", "BEJO_FLUX"]
        for opt in ["PARC_CLIENT", "SINGLETON"]:
            if self.files.get(opt):
                keys.append(opt)
        for k in keys:
            df_prev = self.load_csv_smart(self.files[k], nrows=200)
            self.dfs_preview[k]   = df_prev.head(5)
            self.original_cols[k] = list(df_prev.columns)
            print(f"  {k:<14} : {len(df_prev.columns)} colonnes detectees")

        def col_at(fkey, pos):
            # pos = position 1-based comme dans l'UI ; 0 ou hors borne -> "".
            cols = self.original_cols.get(fkey, [])
            if pos and 0 < pos <= len(cols):
                return cols[pos - 1]
            return ""

        # Association mapping_key -> (fichier, position par défaut).
        FIELD_FILE = {
            'source_statut': 'SOURCE', 'source_pays_1': 'SOURCE', 'source_pays_2': 'SOURCE',
            'source_sales': 'SOURCE', 'source_type_of_card': 'SOURCE', 'source_rs': 'SOURCE',
            'source_rc': 'SOURCE', 'source_rp': 'SOURCE', 'source_rmpm': 'SOURCE',
            'source_iban': 'SOURCE',
            'country_original': 'COUNTRY', 'country_new': 'COUNTRY',
            'sales_original': 'SALES', 'sales_new': 'SALES',
            'toc_original': 'TYPE_OF_CARD', 'toc_new': 'TYPE_OF_CARD',
            'acc_pays_ga': 'ACCOUNT', 'acc_code_ga': 'ACCOUNT', 'acc_nom_ga': 'ACCOUNT',
            'acc_pays_le': 'ACCOUNT', 'acc_rmpm': 'ACCOUNT', 'acc_nom_le': 'ACCOUNT',
            'acc_iban': 'ACCOUNT',
            'ref_rp': 'REF_CLIENT', 'ref_rib': 'REF_CLIENT', 'ref_devise': 'REF_CLIENT',
            'ref_rmpm': 'REF_CLIENT', 'ref_siren': 'REF_CLIENT', 'ref_rs': 'REF_CLIENT',
            'ref_rc': 'REF_CLIENT', 'ref_segment': 'REF_CLIENT', 'ref_marche': 'REF_CLIENT',
            'ref_code_ga': 'REF_CLIENT', 'ref_nom_ga': 'REF_CLIENT',
            'idseg_type': 'IDSEG', 'idseg_id': 'IDSEG', 'idseg_segment': 'IDSEG',
            'monext_mois': 'MONEXT', 'monext_rs': 'MONEXT', 'monext_id_rp': 'MONEXT',
            'monext_id_rc': 'MONEXT', 'monext_iban': 'MONEXT', 'monext_nb_cartes': 'MONEXT',
            'monext_depenses': 'MONEXT', 'monext_retraits': 'MONEXT',
            'worldline_mois': 'WORLDLINE', 'worldline_rs': 'WORLDLINE', 'worldline_iban': 'WORLDLINE',
            'worldline_id_rc': 'WORLDLINE', 'worldline_devise': 'WORLDLINE',
            'worldline_nb_cartes': 'WORLDLINE', 'worldline_depenses_1': 'WORLDLINE',
            'worldline_depenses_2': 'WORLDLINE', 'worldline_plafond': 'WORLDLINE',
            'worldline_periodicite': 'WORLDLINE',
            'devise_date': 'DEVISES', 'devise_code': 'DEVISES', 'devise_taux': 'DEVISES',
            'bejo_cartes_entite': 'BEJO_CARTES', 'bejo_cartes_nb': 'BEJO_CARTES',
            'bejo_flux_entite': 'BEJO_FLUX', 'bejo_flux_iban': 'BEJO_FLUX',
            'parc_rmpm': 'PARC_CLIENT', 'parc_rc': 'PARC_CLIENT', 'parc_rp': 'PARC_CLIENT',
            'sing_nom_le': 'SINGLETON', 'sing_iban': 'SINGLETON',
        }

        m: Dict[str, str] = {}
        for mk, default_pos in DEFAULT_POSITIONS.items():
            fkey = FIELD_FILE.get(mk)
            if fkey is None:
                continue
            # Champs des fichiers optionnels non fournis -> "" (ignorés en aval).
            if fkey not in self.original_cols:
                m[mk] = ""
                continue
            m[mk] = col_at(fkey, default_pos)

        # PNB MONEXT : plage first/last + exclusion + interchange (ex-_pnb_monext_block).
        mx_cols = self.original_cols.get("MONEXT", [])
        mx_pnb_cfg = {
            'first':       DEFAULT_POSITIONS['monext_pnb_first'] if DEFAULT_POSITIONS['monext_pnb_first'] <= len(mx_cols) else 0,
            'last':        DEFAULT_POSITIONS['monext_pnb_last']  if DEFAULT_POSITIONS['monext_pnb_last']  <= len(mx_cols) else 0,
            'excl':        DEFAULT_POSITIONS['monext_pnb_exclude'] if DEFAULT_POSITIONS['monext_pnb_exclude'] <= len(mx_cols) else 0,
            'interchange': DEFAULT_POSITIONS['monext_interchange'] if DEFAULT_POSITIONS['monext_interchange'] <= len(mx_cols) else 0,
        }

        # PNB WORLDLINE : colonnes (max 7) (ex-_pnb_worldline_block).
        wl_cols = self.original_cols.get("WORLDLINE", [])
        wl_pnb_cols = []
        for i in range(7):
            if i < len(DEFAULT_PNB_WORLDLINE) and DEFAULT_PNB_WORLDLINE[i] <= len(wl_cols):
                wl_pnb_cols.append(wl_cols[DEFAULT_PNB_WORLDLINE[i] - 1])

        # Date de référence (ex-_utilisation_block).
        mois_idx = MOIS_NOMS.index(self.ref_mois) + 1 if self.ref_mois in MOIS_NOMS else 12
        ref_yyyymm = f"{self.ref_annee}{str(mois_idx).zfill(2)}"

        # Colonne différé (ex-_utilisation_block).
        differe_col = col_at("SOURCE", DEFAULT_POSITIONS['source_differe'])

        # Exposition (ex-_exposition_block).
        expo_cfg = {
            'use_bejo':           self.use_bejo,
            'col_expo_theorique': col_at("SOURCE", DEFAULT_POSITIONS['source_exposition_theorique']),
            'col_plafond_source': col_at("SOURCE", DEFAULT_POSITIONS['source_plafond']),
        }

        return m, mx_pnb_cfg, wl_pnb_cols, ref_yyyymm, differe_col, expo_cfg

    # -------------------------------------------------------------------------
    # UTILITAIRES (VERBATIM)
    # -------------------------------------------------------------------------
    @staticmethod
    def norm_map(val):
        """Normalisation pour comparaison de mapping : upper, sans accent, espaces simples."""
        if pd.isna(val) or str(val).strip() == '':
            return ''
        s = str(val).strip().upper()
        s = unicodedata.normalize('NFD', s)
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return ' '.join(s.split())

    def _col_name(self, s):
        if not s or s.startswith("(Aucune"):
            return ""
        return s.split(". ", 1)[1] if ". " in s else s

    def _col_idx(self, s):
        if not s or s.startswith("(Aucune"):
            return 0
        try:
            return int(s.split(". ")[0])
        except Exception:
            return 0

    @staticmethod
    def round_half_up(x):
        """Arrondi à l'entier le plus proche, demi vers le haut (1.5→2, 0.5→1)."""
        try:
            return int(math.floor(float(x) + 0.5))
        except Exception:
            return 0

    @staticmethod
    def quarter_of(yyyymm):
        if not yyyymm or len(yyyymm) != 6:
            return ('', 0, 0)
        an = int(yyyymm[:4]); mo = int(yyyymm[4:])
        q = (mo - 1) // 3 + 1
        return (f"T{q}_{an}", an, q)

    @staticmethod
    def quarter_label_long(an, q):
        return f"Q{q} {an}"

    @staticmethod
    def clean_id_safe(series):
        s = series.astype(str).str.strip()
        s = s.replace(['', 'nan', 'NaN', 'None', 'NULL', 'NA', 'N/A', 'NAN', 'NONE'], '')
        mask = s.str.startswith('="') & s.str.endswith('"')
        s = s.where(~mask, s.str[2:-1])
        s = s.str.lstrip("'")
        mask2 = s.str.endswith('.0') & s.str[:-2].str.isdigit()
        s = s.where(~mask2, s.str[:-2])
        return s.str.strip()

    @staticmethod
    def clean_id_strip0(series):
        s = CIBRevenusAnalyzerCLI.clean_id_safe(series)
        stripped = s.str.lstrip('0')
        return stripped.where(stripped != '', s)

    @staticmethod
    def clean_iban(series):
        s = CIBRevenusAnalyzerCLI.clean_id_safe(series)
        return s.str.upper().str.replace(' ', '', regex=False)

    @staticmethod
    def normalize_rs(series):
        def _n(val):
            if pd.isna(val) or str(val).strip() == '':
                return ''
            s = str(val).strip().upper()
            s = unicodedata.normalize('NFD', s)
            s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
            return ' '.join(s.split())
        return series.apply(_n)

    @staticmethod
    def normalize_label(series):
        return CIBRevenusAnalyzerCLI.normalize_rs(series)

    @staticmethod
    def _kv(v):
        """Variantes d'une clé identifiant : valeur nettoyée + sans zéros initiaux.
        Sert à tester toutes les combinaisons (avec/sans zéros) des deux côtés."""
        if v is None:
            return set()
        s = str(v).strip().replace(' ', '')
        if s in ('', 'nan', 'NaN', 'None', 'NULL', 'NA', 'N/A'):
            return set()
        out = {s}
        stripped = s.lstrip('0')
        if stripped:
            out.add(stripped)
        return out

    @staticmethod
    def to_float(series):
        s = series.astype(str)
        for rep in ['"', "'", ' ', '\xa0', ' ', '€']:
            s = s.str.replace(rep, '', regex=False)
        s = s.str.replace('EUR', '', regex=False)
        mask = s.str.endswith('-')
        s = s.where(~mask, '-' + s.str[:-1])
        s = s.str.replace(',', '.', regex=False)
        return pd.to_numeric(s, errors='coerce').fillna(0.0)

    @staticmethod
    def parse_differe(val):
        if pd.isna(val):
            return None
        s = str(val).strip().replace(',', '.').replace(' ', '')
        if s == '' or s.lower() in ['nan', 'none', 'null', 'n/a', 'na']:
            return None
        try:
            return float(s)
        except Exception:
            return None

    @staticmethod
    def parse_mois(series):
        def _p(val):
            if pd.isna(val):
                return ''
            s = str(val).strip()
            if s.startswith('="') and s.endswith('"'):
                s = s[2:-1].strip()
            s = s.lstrip("'").strip()
            if s.endswith('.0') and s[:-2].isdigit():
                s = s[:-2]
            if _re.fullmatch(r'\d{6}', s):
                return s
            m = _re.match(r'^(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})$', s)
            if m:
                mois, annee = int(m.group(2)), m.group(3)
                if 1 <= mois <= 12:
                    return f"{annee}{str(mois).zfill(2)}"
            m = _re.match(r'^(\d{1,2})[/\-\.](\d{4})$', s)
            if m:
                mois, annee = int(m.group(1)), m.group(2)
                if 1 <= mois <= 12:
                    return f"{annee}{str(mois).zfill(2)}"
            m = _re.match(r'^(\d{4})[/\-\.](\d{1,2})$', s)
            if m:
                annee, mois = m.group(1), int(m.group(2))
                if 1 <= mois <= 12:
                    return f"{annee}{str(mois).zfill(2)}"
            return ''
        return series.apply(_p)

    @staticmethod
    def normalize_segment(raw):
        if pd.isna(raw):
            return ''
        s = str(raw).strip().upper()
        if 'BPE' in s:
            return 'BPE'
        if 'ENTREPRISE' in s or 'CORPORATE' in s or 'BCEF' in s:
            return 'ENTREPRISE'
        return ''

    @staticmethod
    def fmt(val):
        try:
            return str(round(float(val), 2)).replace('.', ',')
        except Exception:
            return str(val)

    @staticmethod
    def fmt_diag(val):
        try:
            return f"{round(float(val), 2):g}"
        except Exception:
            return str(val)

    @staticmethod
    def mois_label(yyyymm):
        if len(yyyymm) == 6:
            return f"{yyyymm[:4]}_{MOIS_LABELS.get(yyyymm[4:], yyyymm[4:])}"
        return yyyymm

    # -------------------------------------------------------------------------
    # PROGRESSION (ex-self._prog GUI -> print)
    # -------------------------------------------------------------------------
    def _prog(self, val, txt):
        # [CLI] Remplace l'avancement de la barre de progression + label statut.
        print(f"[{val * 100:5.1f}%] {txt}")

    # -------------------------------------------------------------------------
    # LANCEMENT (ex-start_thread/threading) -> run()
    # -------------------------------------------------------------------------
    def run(self) -> Dict[str, Any]:
        # [CLI] Remplace start_thread() + threading.Thread(target=self.worker...).
        print("Lecture des colonnes...")
        m, mx_pnb_cfg, wl_pnb_cols, ref_yyyymm, differe_col, expo_cfg = self._build_mapping()

        # Contrôles repris de start_thread() (plage PNB MONEXT).
        if not mx_pnb_cfg['first'] or not mx_pnb_cfg['last']:
            raise ValueError("Plage PNB MONEXT non configurée (colonnes par défaut absentes du fichier MONEXT).")
        if mx_pnb_cfg['first'] > mx_pnb_cfg['last']:
            raise ValueError("Première colonne PNB doit être avant la dernière.")

        return self.worker(m, mx_pnb_cfg, wl_pnb_cols, ref_yyyymm, differe_col, expo_cfg)

    # =========================================================================
    # WORKER — cœur du traitement (VERBATIM, hors lignes UI commentées)
    # =========================================================================
    def worker(self, m, mx_pnb_cfg, wl_pnb_cols, ref_yyyymm, differe_col, expo_cfg):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fn = {k: (os.path.basename(p) if p else f"<{k}>") for k, p in self.files.items()}

        # ── ÉTAPE 1 — CHARGEMENT ─────────────────────────────────────────
        self._prog(0.02, "Chargement SOURCE (Monitoring/Onbording CIB)...")
        df_source  = self.load_csv_smart(self.files["SOURCE"]); total_rows = len(df_source)
        self._prog(0.03, "Chargement IBAN_ACCOUNT...");        df_account   = self.load_csv_smart(self.files["ACCOUNT"])
        self._prog(0.04, "Chargement REFERENTIEL_CLIENT..."); df_ref       = self.load_csv_smart(self.files["REF_CLIENT"])
        self._prog(0.05, "Chargement IDENTIFIANT_SEGMENT..."); df_idseg    = self.load_csv_smart(self.files["IDSEG"])
        self._prog(0.06, "Chargement MONEXT...");             df_monext    = self.load_csv_smart(self.files["MONEXT"])
        self._prog(0.08, "Chargement WORLDLINE...");          df_worldline = self.load_csv_smart(self.files["WORLDLINE"])
        self._prog(0.085, "Chargement DEVISES...");           df_devises   = self.load_csv_smart(self.files["DEVISES"])
        self._prog(0.09, "Chargement tables de mapping...")
        df_country = self.load_csv_smart(self.files["COUNTRY"]) if self.files.get("COUNTRY") else pd.DataFrame()
        df_sales   = self.load_csv_smart(self.files["SALES"]) if self.files.get("SALES") else pd.DataFrame()
        df_toc     = self.load_csv_smart(self.files["TYPE_OF_CARD"])

        # NOUVEAU (V2) — fichiers de récupération optionnels
        df_parc = None; df_singleton = None
        if self.files.get("PARC_CLIENT"):
            self._prog(0.092, "Chargement PARC_CLIENT...")
            df_parc = self.load_csv_smart(self.files["PARC_CLIENT"])
        if self.files.get("SINGLETON"):
            self._prog(0.094, "Chargement SINGLETON...")
            df_singleton = self.load_csv_smart(self.files["SINGLETON"])

        # ── ÉTAPE 2 — CLEANING SOURCE (normalisation + mapping + manuel) ──
        self._prog(0.10, "Cleaning SOURCE (normalisation avant mapping)...")

        def build_mapping_dict(df_map, col_orig, col_new, manual):
            d = {}
            if col_orig in df_map.columns and col_new in df_map.columns:
                ko = df_map[col_orig].astype(str)
                vn = df_map[col_new].astype(str).str.strip()
                for o, n in zip(ko.values, vn.values):
                    nk = self.norm_map(o)
                    if nk and nk not in d and n.strip():
                        d[nk] = n.strip()
            for nk, target in (manual or {}).items():
                d[nk] = target
            return d

        def apply_mapping(series, mapping_dict):
            def _ap(v):
                nk = self.norm_map(v)
                if nk in mapping_dict:
                    return mapping_dict[nk]
                return str(v).strip()
            return series.apply(_ap)

        country_map = build_mapping_dict(df_country, m['country_original'], m['country_new'], self.manual_map.get('COUNTRY'))
        sales_map   = build_mapping_dict(df_sales,   m['sales_original'],   m['sales_new'],   self.manual_map.get('SALES'))
        toc_map     = build_mapping_dict(df_toc,     m['toc_original'],     m['toc_new'],     self.manual_map.get('TYPE_OF_CARD'))

        if m.get('source_pays_1') and m['source_pays_1'] in df_source.columns:
            df_source[m['source_pays_1']] = apply_mapping(df_source[m['source_pays_1']], country_map)
        if m.get('source_pays_2') and m['source_pays_2'] in df_source.columns:
            df_source[m['source_pays_2']] = apply_mapping(df_source[m['source_pays_2']], country_map)
        if m.get('source_sales') and m['source_sales'] in df_source.columns:
            df_source[m['source_sales']] = apply_mapping(df_source[m['source_sales']], sales_map)
        if m.get('source_type_of_card') and m['source_type_of_card'] in df_source.columns:
            df_source[m['source_type_of_card']] = apply_mapping(df_source[m['source_type_of_card']], toc_map)

        if m.get('source_statut') and m['source_statut'] in df_source.columns:
            arr_statut = df_source[m['source_statut']].astype(str).str.strip().values
        else:
            arr_statut = np.full(total_rows, '', dtype=object)

        # ── ÉTAPE 3 — INDEXATION DES PIVOTS ──────────────────────────────
        self._prog(0.12, "Indexation IBAN_ACCOUNT...")
        df_account['_K_RMPM']       = self.clean_id_safe(df_account[m['acc_rmpm']])
        df_account['_K_RMPM_STRIP'] = df_account['_K_RMPM'].str.lstrip('0').where(
                                          df_account['_K_RMPM'].str.lstrip('0') != '', df_account['_K_RMPM'])
        df_account['_K_IBAN']  = self.clean_iban(df_account[m['acc_iban']])
        df_account['_CODE_GA'] = self.clean_id_strip0(df_account[m['acc_code_ga']])
        df_account['_NOM_GA']  = df_account[m['acc_nom_ga']].astype(str).str.strip()
        df_account['_PAYS_GA'] = df_account[m['acc_pays_ga']].astype(str).str.strip()
        df_account['_NOM_LE']  = df_account[m['acc_nom_le']].astype(str).str.strip()
        df_account['_PAYS_LE'] = df_account[m['acc_pays_le']].astype(str).str.strip()
        df_account['_K_NOM_LE'] = self.normalize_rs(df_account[m['acc_nom_le']])

        acc_by_iban = {}
        for _, r in df_account[df_account['_K_IBAN'] != ''].iterrows():
            ib = r['_K_IBAN']
            if ib not in acc_by_iban:
                acc_by_iban[ib] = {'rmpm': r['_K_RMPM'], 'code_ga': r['_CODE_GA'], 'nom_ga': r['_NOM_GA'],
                                   'pays_ga': r['_PAYS_GA'], 'nom_le': r['_NOM_LE'], 'pays_le': r['_PAYS_LE']}
        acc_by_rmpm = {}
        df_acc_ok = df_account[(df_account['_K_RMPM'] != '') & (df_account['_K_IBAN'] != '')]
        for _, r in df_acc_ok.iterrows():
            for k in {r['_K_RMPM'], r['_K_RMPM_STRIP']}:
                acc_by_rmpm.setdefault(k, [])
                entry = (r['_K_IBAN'], r['_NOM_LE'], r['_CODE_GA'], r['_NOM_GA'], r['_PAYS_GA'], r['_PAYS_LE'])
                if entry not in acc_by_rmpm[k]:
                    acc_by_rmpm[k].append(entry)

        self._prog(0.14, "Indexation REFERENTIEL_CLIENT...")
        df_ref['_K_RC']      = self.clean_id_strip0(df_ref[m['ref_rc']])
        df_ref['_K_RC_SAFE'] = self.clean_id_safe(df_ref[m['ref_rc']])
        df_ref['_K_RP']      = self.clean_id_strip0(df_ref[m['ref_rp']])
        df_ref['_K_RP_SAFE'] = self.clean_id_safe(df_ref[m['ref_rp']])
        df_ref['_K_RIB']     = self.clean_iban(df_ref[m['ref_rib']])
        df_ref['_K_RMPM']    = self.clean_id_safe(df_ref[m['ref_rmpm']])
        df_ref['_RS']        = self.normalize_rs(df_ref[m['ref_rs']])
        df_ref['_CODE_GA']   = self.clean_id_strip0(df_ref[m['ref_code_ga']])
        df_ref['_NOM_GA']    = df_ref[m['ref_nom_ga']].astype(str).str.strip()
        df_ref['_SEGMENT']   = df_ref[m['ref_segment']].astype(str).apply(self.normalize_segment)

        def _ref_entry(row):
            return {'rmpm': row['_K_RMPM'], 'code_ga': row['_CODE_GA'], 'nom_ga': row['_NOM_GA'],
                    'rs': row['_RS'], 'segment': row['_SEGMENT']}
        ref_by_rc = {}
        for _, r in df_ref[df_ref['_K_RC'] != ''].iterrows():
            for k in {r['_K_RC'], r['_K_RC_SAFE']}:
                if k and k not in ref_by_rc:
                    ref_by_rc[k] = _ref_entry(r)
        ref_by_rib = {}
        for _, r in df_ref[df_ref['_K_RIB'] != ''].iterrows():
            if r['_K_RIB'] not in ref_by_rib:
                ref_by_rib[r['_K_RIB']] = _ref_entry(r)
        ref_by_rp = {}
        for _, r in df_ref[df_ref['_K_RP'] != ''].iterrows():
            for k in {r['_K_RP'], r['_K_RP_SAFE']}:
                if k and k not in ref_by_rp:
                    ref_by_rp[k] = _ref_entry(r)

        self._prog(0.16, "Indexation IDENTIFIANT_SEGMENT...")
        seg_by_rc = {}; seg_by_rmpm = {}; seg_by_rp = {}; seg_by_ca = {}
        idseg_types = df_idseg[m['idseg_type']].astype(str).str.strip().str.upper()
        idseg_vals  = self.clean_id_safe(df_idseg[m['idseg_id']])
        idseg_segs  = df_idseg[m['idseg_segment']].astype(str).apply(self.normalize_segment)
        TYPE_MAP = {'RC': seg_by_rc, 'RMPM': seg_by_rmpm, 'RP': seg_by_rp,
                    'CODE_AGENCE': seg_by_ca, 'CODE GA': seg_by_ca}
        for t, v, seg in zip(idseg_types, idseg_vals, idseg_segs):
            if not seg or not v:
                continue
            td = TYPE_MAP.get(t)
            if td is None:
                continue
            if v not in td:
                td[v] = seg
            stripped = v.lstrip('0')
            if stripped and stripped != v and stripped not in td:
                td[stripped] = seg

        # ── ÉTAPE 3bis (V2) — PARC_CLIENT (RMPM→RC/RP) + NOM→IBAN ────────
        self._prog(0.17, "Indexation récupération étendue (PARC + nom→IBAN)...")
        # parc_by_rmpm : clé RMPM (brut + sans zéros) → liste de (rc_variants, rp_variants)
        parc_by_rmpm = {}
        if df_parc is not None and m.get('parc_rmpm') in df_parc.columns:
            p_rmpm = self.clean_id_safe(df_parc[m['parc_rmpm']])
            p_rc   = self.clean_id_safe(df_parc[m['parc_rc']]) if m.get('parc_rc') in df_parc.columns else pd.Series([''] * len(df_parc))
            p_rp   = self.clean_id_safe(df_parc[m['parc_rp']]) if m.get('parc_rp') in df_parc.columns else pd.Series([''] * len(df_parc))
            for rmpm_v, rc_v, rp_v in zip(p_rmpm.values, p_rc.values, p_rp.values):
                if not rmpm_v:
                    continue
                rc_set = self._kv(rc_v); rp_set = self._kv(rp_v)
                if not rc_set and not rp_set:
                    continue
                for kk in self._kv(rmpm_v):
                    slot = parc_by_rmpm.setdefault(kk, {'rc': set(), 'rp': set()})
                    slot['rc'] |= rc_set
                    slot['rp'] |= rp_set

        # name_to_ibans : nom entité normalisé → set d'IBAN (IBAN_ACCOUNT col nom_le + SINGLETON)
        name_to_ibans = {}
        for nom_norm, ib in zip(df_account['_K_NOM_LE'].values, df_account['_K_IBAN'].values):
            if nom_norm and ib:
                name_to_ibans.setdefault(nom_norm, set()).add(ib)
        if df_singleton is not None and m.get('sing_nom_le') in df_singleton.columns \
           and m.get('sing_iban') in df_singleton.columns:
            s_nom = self.normalize_rs(df_singleton[m['sing_nom_le']])
            s_ib  = self.clean_iban(df_singleton[m['sing_iban']])
            for nom_norm, ib in zip(s_nom.values, s_ib.values):
                if nom_norm and ib:
                    name_to_ibans.setdefault(nom_norm, set()).add(ib)

        # ── ÉTAPE 4 — INDEXATION DES FLUX MX/WL ──────────────────────────
        self._prog(0.20, "Préparation MONEXT...")
        df_monext['_K_RP']      = self.clean_id_strip0(df_monext[m['monext_id_rp']])
        df_monext['_K_RP_SAFE'] = self.clean_id_safe(df_monext[m['monext_id_rp']])
        df_monext['_K_RC']      = self.clean_id_strip0(df_monext[m['monext_id_rc']])
        df_monext['_K_RC_SAFE'] = self.clean_id_safe(df_monext[m['monext_id_rc']])
        df_monext['_K_RS']      = self.normalize_rs(df_monext[m['monext_rs']])
        df_monext['_K_IBAN']    = self.clean_iban(df_monext[m['monext_iban']])
        df_monext['_DEPENSES']  = self.to_float(df_monext[m['monext_depenses']])
        df_monext['_RETRAITS']  = self.to_float(df_monext[m['monext_retraits']])
        df_monext['_NB_CARTES'] = self.to_float(df_monext[m['monext_nb_cartes']])
        df_monext['_FLUX']      = df_monext['_DEPENSES'] + df_monext['_RETRAITS']
        df_monext['_MOIS']      = self.parse_mois(df_monext[m['monext_mois']])

        mx_cols = list(df_monext.columns)
        f0  = mx_pnb_cfg['first'] - 1; l0 = mx_pnb_cfg['last'] - 1
        e0  = mx_pnb_cfg['excl'] - 1 if mx_pnb_cfg['excl'] > 0 else -1
        ic0 = mx_pnb_cfg['interchange'] - 1 if mx_pnb_cfg['interchange'] > 0 else -1
        pnb_range     = mx_cols[f0:l0+1]
        excl_col_name = mx_cols[e0]  if 0 <= e0  < len(mx_cols) else None
        ic_col_name   = mx_cols[ic0] if 0 <= ic0 < len(mx_cols) else None
        for c in pnb_range:
            if c in df_monext.columns:
                df_monext[c] = self.to_float(df_monext[c])
        if ic_col_name and ic_col_name in df_monext.columns:
            df_monext[ic_col_name] = self.to_float(df_monext[ic_col_name]) * -1
        pnb_filtered = [c for c in pnb_range if c != excl_col_name]
        df_monext['_PNB'] = df_monext[pnb_filtered].sum(axis=1) if pnb_filtered else 0.0

        agg_dict = {'_FLUX': 'sum', '_PNB': 'sum', '_NB_CARTES': 'sum'}
        def agg_groupby(df, key):
            return df[df[key] != ''].groupby(key).agg(agg_dict).T.to_dict('list')
        agg_mx_iban = agg_groupby(df_monext, '_K_IBAN')
        agg_mx_rp = dict(agg_groupby(df_monext, '_K_RP'))
        for k, v in agg_groupby(df_monext, '_K_RP_SAFE').items():
            agg_mx_rp.setdefault(k, v)
        agg_mx_rc = dict(agg_groupby(df_monext, '_K_RC'))
        for k, v in agg_groupby(df_monext, '_K_RC_SAFE').items():
            agg_mx_rc.setdefault(k, v)
        agg_mx_rs = agg_groupby(df_monext, '_K_RS')
        mx_rs_incl = [(k, v) for k, v in agg_mx_rs.items() if len(k) >= RS_INCLUSION_MIN]

        self._prog(0.23, "Agrégats mensuels MONEXT...")
        mois_mx = sorted([x for x in df_monext['_MOIS'].unique() if x and len(x) == 6])
        mx_mo_iban = {}; mx_mo_rp = {}; mx_mo_rc = {}; mx_global_by_mois = {}
        for mois in mois_mx:
            df_m = df_monext[df_monext['_MOIS'] == mois]
            mx_mo_iban[mois] = df_m[df_m['_K_IBAN'] != ''].groupby('_K_IBAN').agg(agg_dict).T.to_dict('list')
            d_rp = dict(df_m[df_m['_K_RP'] != ''].groupby('_K_RP').agg(agg_dict).T.to_dict('list'))
            for k, v in df_m[df_m['_K_RP_SAFE'] != ''].groupby('_K_RP_SAFE').agg(agg_dict).T.to_dict('list').items():
                d_rp.setdefault(k, v)
            mx_mo_rp[mois] = d_rp
            d_rc = dict(df_m[df_m['_K_RC'] != ''].groupby('_K_RC').agg(agg_dict).T.to_dict('list'))
            for k, v in df_m[df_m['_K_RC_SAFE'] != ''].groupby('_K_RC_SAFE').agg(agg_dict).T.to_dict('list').items():
                d_rc.setdefault(k, v)
            mx_mo_rc[mois] = d_rc
            mx_global_by_mois[mois] = {'flux': df_m['_FLUX'].sum(), 'pnb': df_m['_PNB'].sum(), 'nb': df_m['_NB_CARTES'].sum()}

        self._prog(0.28, "Préparation WORLDLINE...")
        df_worldline['_K_RC']      = self.clean_id_strip0(df_worldline[m['worldline_id_rc']])
        df_worldline['_K_RC_SAFE'] = self.clean_id_safe(df_worldline[m['worldline_id_rc']])
        df_worldline['_K_RS']      = self.normalize_rs(df_worldline[m['worldline_rs']])
        df_worldline['_K_IBAN']    = self.clean_iban(df_worldline[m['worldline_iban']])
        df_worldline['_NB_CARTES'] = self.to_float(df_worldline[m['worldline_nb_cartes']])
        df_worldline['_MOIS']      = self.parse_mois(df_worldline[m['worldline_mois']])

        self._prog(0.285, "Conversion devises Worldline → EUR...")
        col_devise_wl = m.get('worldline_devise', '')
        col_dd = m.get('devise_date', ''); col_dc = m.get('devise_code', ''); col_dt = m.get('devise_taux', '')
        taux_dict = {}; mois_par_devise = {}
        if col_dd and col_dc and col_dt:
            dev_dates = df_devises[col_dd].astype(str).str.strip()
            dev_codes = df_devises[col_dc].astype(str).str.strip().str.upper()
            dev_taux  = pd.to_numeric(df_devises[col_dt].astype(str).str.strip().str.replace(',', '.', regex=False), errors='coerce')
            for d, c, t in zip(dev_dates.values, dev_codes.values, dev_taux.values):
                if not d or not c or pd.isna(t) or t <= 0:
                    continue
                d_clean = ''.join(ch for ch in d if ch.isdigit())
                if len(d_clean) != 6:
                    continue
                taux_dict[(d_clean, c)] = float(t)
                mois_par_devise.setdefault(c, set()).add(d_clean)
            mois_par_devise = {dev: sorted(list(s)) for dev, s in mois_par_devise.items()}
        def resolve_taux(mois_val, devise_val):
            if not devise_val or devise_val == 'EUR':
                return 1.0
            if devise_val not in mois_par_devise:
                return 1.0
            m_clean = ''.join(ch for ch in str(mois_val) if ch.isdigit())
            if len(m_clean) != 6:
                return 1.0
            key = (m_clean, devise_val)
            if key in taux_dict:
                return taux_dict[key]
            candidat = None
            for md in mois_par_devise[devise_val]:
                if md <= m_clean:
                    candidat = md
                else:
                    break
            return taux_dict[(candidat, devise_val)] if candidat else 1.0
        n_wl = len(df_worldline)
        if col_devise_wl and col_devise_wl in df_worldline.columns:
            arr_devise = df_worldline[col_devise_wl].astype(str).str.strip().str.upper().values
        else:
            arr_devise = np.full(n_wl, 'EUR', dtype=object)
        arr_mois_wl = df_worldline['_MOIS'].values
        cache_taux = {}; taux_array = np.ones(n_wl, dtype=float)
        for i in range(n_wl):
            key = (arr_mois_wl[i], arr_devise[i])
            if key not in cache_taux:
                cache_taux[key] = resolve_taux(arr_mois_wl[i], arr_devise[i])
            taux_array[i] = cache_taux[key]
        df_worldline['_DEP1'] = self.to_float(df_worldline[m['worldline_depenses_1']]) * taux_array
        df_worldline['_DEP2'] = self.to_float(df_worldline[m['worldline_depenses_2']]) * taux_array
        if wl_pnb_cols:
            for c in wl_pnb_cols:
                df_worldline[c] = self.to_float(df_worldline[c]) * taux_array
        nb_lignes_non_eur = int((arr_devise != 'EUR').sum())
        df_worldline['_FLUX'] = df_worldline['_DEP1'] + df_worldline['_DEP2']
        df_worldline['_PNB'] = df_worldline[wl_pnb_cols].sum(axis=1) if wl_pnb_cols else 0.0

        agg_wl_iban = agg_groupby(df_worldline, '_K_IBAN')
        agg_wl_rc = dict(agg_groupby(df_worldline, '_K_RC'))
        for k, v in agg_groupby(df_worldline, '_K_RC_SAFE').items():
            agg_wl_rc.setdefault(k, v)
        agg_wl_rs = agg_groupby(df_worldline, '_K_RS')
        wl_rs_incl = [(k, v) for k, v in agg_wl_rs.items() if len(k) >= RS_INCLUSION_MIN]

        self._prog(0.31, "Agrégats mensuels WORLDLINE...")
        mois_wl = sorted([x for x in df_worldline['_MOIS'].unique() if x and len(x) == 6])
        wl_mo_iban = {}; wl_mo_rc = {}; wl_global_by_mois = {}
        for mois in mois_wl:
            df_w = df_worldline[df_worldline['_MOIS'] == mois]
            wl_mo_iban[mois] = df_w[df_w['_K_IBAN'] != ''].groupby('_K_IBAN').agg(agg_dict).T.to_dict('list')
            d_rc = dict(df_w[df_w['_K_RC'] != ''].groupby('_K_RC').agg(agg_dict).T.to_dict('list'))
            for k, v in df_w[df_w['_K_RC_SAFE'] != ''].groupby('_K_RC_SAFE').agg(agg_dict).T.to_dict('list').items():
                d_rc.setdefault(k, v)
            wl_mo_rc[mois] = d_rc
            wl_global_by_mois[mois] = {'flux': df_w['_FLUX'].sum(), 'pnb': df_w['_PNB'].sum(), 'nb': df_w['_NB_CARTES'].sum()}

        all_mois = sorted(set(mois_mx) | set(mois_wl))

        # ── ÉTAPE 5 — PRÉPARATION SOURCE ─────────────────────────────────
        self._prog(0.34, "Préparation SOURCE (clés de matching)...")
        df_source['_K_RP']      = self.clean_id_strip0(df_source[m['source_rp']])
        df_source['_K_RP_SAFE'] = self.clean_id_safe(df_source[m['source_rp']])
        df_source['_K_RC']      = self.clean_id_strip0(df_source[m['source_rc']])
        df_source['_K_RC_SAFE'] = self.clean_id_safe(df_source[m['source_rc']])
        df_source['_K_RMPM']    = self.clean_id_safe(df_source[m['source_rmpm']])
        df_source['_K_RS']      = self.normalize_rs(df_source[m['source_rs']])
        if m.get('source_iban'):
            df_source['_K_IBAN'] = self.clean_iban(df_source[m['source_iban']])
        else:
            df_source['_K_IBAN'] = ''

        arr_rp = df_source['_K_RP'].values;      arr_rp_safe = df_source['_K_RP_SAFE'].values
        arr_rc = df_source['_K_RC'].values;      arr_rc_safe = df_source['_K_RC_SAFE'].values
        arr_rmpm = df_source['_K_RMPM'].values;  arr_rs = df_source['_K_RS'].values
        arr_iban = df_source['_K_IBAN'].values

        # ── ÉTAPE 6 — RÉSOLUTION D'IDENTITÉ (cascade + récupération V2) ───
        self._prog(0.38, "PHASE 1 : Résolution d'identité...")
        id_rmpm    = np.full(total_rows, '', dtype=object); id_code_ga = np.full(total_rows, '', dtype=object)
        id_nom_ga  = np.full(total_rows, '', dtype=object); id_pays_ga = np.full(total_rows, '', dtype=object)
        id_nom_le  = np.full(total_rows, '', dtype=object); id_pays_le = np.full(total_rows, '', dtype=object)
        id_rs_ref  = np.full(total_rows, '', dtype=object); id_segment = np.full(total_rows, '', dtype=object)
        id_ibans   = [[] for _ in range(total_rows)]
        id_source  = np.full(total_rows, 'NO_IDENTITY', dtype=object)
        id_seg_src = np.full(total_rows, 'NO_SEGMENT', dtype=object)
        diag_identity = np.full(total_rows, '', dtype=object)
        # NOUVEAU (V2) — clés candidates RC/RP par ligne (toutes combinaisons)
        cand_rc = [set() for _ in range(total_rows)]
        cand_rp = [set() for _ in range(total_rows)]
        # NOUVEAU (V2) — flags de récupération étendue (pour stats)
        parc_flag = np.zeros(total_rows, dtype=bool)   # PARC a apporté des clés candidates
        nom_flag  = np.zeros(total_rows, dtype=bool)   # NOM exact → IBAN(s) ajouté(s)

        def _fill(arr, i, value):
            if value and not arr[i]:
                arr[i] = value

        for i in range(total_rows):
            if i % 2000 == 0:
                self._prog(0.38 + 0.10 * i / max(total_rows, 1), f"Identité : {i:,}/{total_rows:,}...")
            rmpm = arr_rmpm[i]; rmpm_strip = (rmpm.lstrip('0') or rmpm) if rmpm else ''
            rc = arr_rc[i]; rc_safe = arr_rc_safe[i]; rp = arr_rp[i]; rp_safe = arr_rp_safe[i]
            rs = arr_rs[i]; iban = arr_iban[i]
            steps = []

            if iban and iban in acc_by_iban:
                info = acc_by_iban[iban]
                _fill(id_rmpm, i, info['rmpm']); _fill(id_code_ga, i, info['code_ga']); _fill(id_nom_ga, i, info['nom_ga'])
                _fill(id_pays_ga, i, info['pays_ga']); _fill(id_nom_le, i, info['nom_le']); _fill(id_pays_le, i, info['pays_le'])
                if iban not in id_ibans[i]:
                    id_ibans[i].append(iban)
                if id_source[i] == 'NO_IDENTITY':
                    id_source[i] = 'IBAN_ACCOUNT_IBAN'
                steps.append(f"Identité trouvée dans {fn['ACCOUNT']} via IBAN={iban} → RMPM={info['rmpm']}, GA={info['code_ga']}")

            acc_rmpm_hit = acc_by_rmpm.get(rmpm) or acc_by_rmpm.get(rmpm_strip)
            if acc_rmpm_hit:
                for (ib, nom_le, code_ga, nom_ga, pays_ga, pays_le) in acc_rmpm_hit:
                    if ib and ib not in id_ibans[i]:
                        id_ibans[i].append(ib)
                    _fill(id_nom_le, i, nom_le); _fill(id_code_ga, i, code_ga); _fill(id_nom_ga, i, nom_ga)
                    _fill(id_pays_ga, i, pays_ga); _fill(id_pays_le, i, pays_le)
                _fill(id_rmpm, i, rmpm)
                if id_source[i] == 'NO_IDENTITY':
                    id_source[i] = 'IBAN_ACCOUNT_RMPM'
                steps.append(f"IBAN(s) du client retrouvé(s) dans {fn['ACCOUNT']} via RMPM={rmpm} → {len(acc_rmpm_hit)} IBAN rattaché(s)")

            ref_hit = None
            for k in [rc, rc_safe]:
                if k and k in ref_by_rc:
                    ref_hit = ref_by_rc[k]; break
            if ref_hit:
                _fill(id_rmpm, i, ref_hit['rmpm']); _fill(id_code_ga, i, ref_hit['code_ga'])
                _fill(id_nom_ga, i, ref_hit['nom_ga']); _fill(id_rs_ref, i, ref_hit['rs'])
                if ref_hit['segment'] and not id_segment[i]:
                    id_segment[i] = ref_hit['segment']; id_seg_src[i] = 'REF_CLIENT_RC'
                if id_source[i] == 'NO_IDENTITY':
                    id_source[i] = 'REF_CLIENT_RC'
                steps.append(f"Identité complétée dans {fn['REF_CLIENT']} via RC={rc} → segment={ref_hit['segment'] or 'n/a'}")

            if iban and iban in ref_by_rib:
                ref_hit = ref_by_rib[iban]
                _fill(id_rmpm, i, ref_hit['rmpm']); _fill(id_code_ga, i, ref_hit['code_ga'])
                _fill(id_nom_ga, i, ref_hit['nom_ga']); _fill(id_rs_ref, i, ref_hit['rs'])
                if ref_hit['segment'] and not id_segment[i]:
                    id_segment[i] = ref_hit['segment']; id_seg_src[i] = 'REF_CLIENT_RIB'
                if id_source[i] == 'NO_IDENTITY':
                    id_source[i] = 'REF_CLIENT_RIB'
                steps.append(f"Identité complétée dans {fn['REF_CLIENT']} via RIB={iban}")

            if id_source[i] == 'NO_IDENTITY':
                for k in [rp, rp_safe]:
                    if k and k in ref_by_rp:
                        ref_hit = ref_by_rp[k]
                        _fill(id_rmpm, i, ref_hit['rmpm']); _fill(id_code_ga, i, ref_hit['code_ga'])
                        _fill(id_nom_ga, i, ref_hit['nom_ga']); _fill(id_rs_ref, i, ref_hit['rs'])
                        if ref_hit['segment'] and not id_segment[i]:
                            id_segment[i] = ref_hit['segment']; id_seg_src[i] = 'REF_CLIENT_RP'
                        id_source[i] = 'REF_CLIENT_RP'
                        steps.append(f"Identité trouvée dans {fn['REF_CLIENT']} via RP={k}")
                        break

            # ── V2 : clés candidates RC/RP (toutes combinaisons espaces/zéros) ──
            cset = self._kv(rc) | self._kv(rc_safe)
            pset = self._kv(rp) | self._kv(rp_safe)

            # ── V2 : PARC_CLIENT — RMPM → RC/RP (clés candidates supplémentaires) ──
            parc_hit = None
            if parc_by_rmpm:
                parc_hit = parc_by_rmpm.get(rmpm) or parc_by_rmpm.get(rmpm_strip)
            if parc_hit:
                before_n = len(cset) + len(pset)
                cset |= parc_hit['rc']; pset |= parc_hit['rp']
                if len(cset) + len(pset) > before_n:
                    parc_flag[i] = True
                steps.append(f"PARC_CLIENT ({fn['PARC_CLIENT']}) via RMPM={rmpm or rmpm_strip} → "
                             f"RC={sorted(parc_hit['rc']) if parc_hit['rc'] else '∅'}, "
                             f"RP={sorted(parc_hit['rp']) if parc_hit['rp'] else '∅'} (clés candidates flux)")
                if id_source[i] == 'NO_IDENTITY':
                    done = False
                    for k in sorted(parc_hit['rc']):
                        if k in ref_by_rc:
                            rh = ref_by_rc[k]
                            _fill(id_rmpm, i, rh['rmpm']); _fill(id_code_ga, i, rh['code_ga'])
                            _fill(id_nom_ga, i, rh['nom_ga']); _fill(id_rs_ref, i, rh['rs'])
                            if rh['segment'] and not id_segment[i]:
                                id_segment[i] = rh['segment']; id_seg_src[i] = 'PARC_REF_RC'
                            id_source[i] = 'PARC_REF_RC'
                            steps.append(f"Identité complétée via PARC→RC={k} dans {fn['REF_CLIENT']}")
                            done = True; break
                    if not done:
                        for k in sorted(parc_hit['rp']):
                            if k in ref_by_rp:
                                rh = ref_by_rp[k]
                                _fill(id_rmpm, i, rh['rmpm']); _fill(id_code_ga, i, rh['code_ga'])
                                _fill(id_nom_ga, i, rh['nom_ga']); _fill(id_rs_ref, i, rh['rs'])
                                if rh['segment'] and not id_segment[i]:
                                    id_segment[i] = rh['segment']; id_seg_src[i] = 'PARC_REF_RP'
                                id_source[i] = 'PARC_REF_RP'
                                steps.append(f"Identité complétée via PARC→RP={k} dans {fn['REF_CLIENT']}")
                                break
            cand_rc[i] = cset; cand_rp[i] = pset

            # ── V2 : match exact du NOM d'entité (col 7 SOURCE) → IBAN(s) ──
            if rs and rs in name_to_ibans:
                new_ibans = [ib for ib in sorted(name_to_ibans[rs]) if ib not in id_ibans[i]]
                for ib in new_ibans:
                    id_ibans[i].append(ib)
                if new_ibans:
                    nom_flag[i] = True
                    steps.append(f"Nom entité « {rs} » trouvé en exact dans {fn['ACCOUNT']}/{fn['SINGLETON']} "
                                 f"→ {len(new_ibans)} IBAN rattaché(s) : "
                                 f"{', '.join(new_ibans[:5])}{'…' if len(new_ibans) > 5 else ''}")
                    if id_source[i] == 'NO_IDENTITY':
                        id_source[i] = 'NOM_EXACT_IBAN'

            # ── Segment fallback IDSEG ──
            if not id_segment[i]:
                for k in [rc, rc_safe]:
                    if k and k in seg_by_rc:
                        id_segment[i] = seg_by_rc[k]; id_seg_src[i] = 'IDSEG_RC'
                        steps.append(f"Segment trouvé dans {fn['IDSEG']} via RC={k} → {id_segment[i]}"); break
            if not id_segment[i]:
                for k in [rmpm, rmpm_strip]:
                    if k and k in seg_by_rmpm:
                        id_segment[i] = seg_by_rmpm[k]; id_seg_src[i] = 'IDSEG_RMPM'
                        steps.append(f"Segment trouvé dans {fn['IDSEG']} via RMPM={k} → {id_segment[i]}"); break
            if not id_segment[i]:
                for k in [rp, rp_safe]:
                    if k and k in seg_by_rp:
                        id_segment[i] = seg_by_rp[k]; id_seg_src[i] = 'IDSEG_RP'
                        steps.append(f"Segment trouvé dans {fn['IDSEG']} via RP={k} → {id_segment[i]}"); break

            if not steps:
                steps.append(f"Aucune identité : RMPM/RC/RP/IBAN/nom du client absents de "
                             f"{fn['ACCOUNT']}, {fn['REF_CLIENT']}, {fn['IDSEG']} (et PARC/SINGLETON si fournis)")
            diag_identity[i] = " | ".join(steps)

        # ── ÉTAPE 7 — MATCHING DES FLUX (cascade + diag complet V2) ──────
        self._prog(0.50, "PHASE 2 : Matching des flux MX/WL...")
        flux_mx_total = np.zeros(total_rows); pnb_mx_total = np.zeros(total_rows); nb_mx_total = np.zeros(total_rows)
        match_mx = np.full(total_rows, 'NO_MATCH', dtype=object); matched_keys_mx = [None]*total_rows
        flux_wl_total = np.zeros(total_rows); pnb_wl_total = np.zeros(total_rows); nb_wl_total = np.zeros(total_rows)
        match_wl = np.full(total_rows, 'NO_MATCH', dtype=object); matched_keys_wl = [None]*total_rows
        nb_iban_trouves = np.zeros(total_rows, dtype=int)
        nb_iban_dans_mx = np.zeros(total_rows, dtype=int); nb_iban_dans_wl = np.zeros(total_rows, dtype=int)
        diag_flux_mx = np.full(total_rows, '', dtype=object); diag_flux_wl = np.full(total_rows, '', dtype=object)

        def _lst(s):
            return sorted(s) if s else '∅'

        for i in range(total_rows):
            if i % 2000 == 0:
                self._prog(0.50 + 0.15 * i / max(total_rows, 1), f"Matching : {i:,}/{total_rows:,}...")
            ibans_client = id_ibans[i]; nb_iban_trouves[i] = len(ibans_client)
            rs = arr_rs[i]
            rp_keys = sorted(cand_rp[i]); rc_keys = sorted(cand_rc[i])
            steps_mx = []; steps_wl = []

            # ===== MONEXT =====
            if ibans_client:
                ibans_in_mx = [ib for ib in ibans_client if ib in agg_mx_iban]
                if ibans_in_mx:
                    flux_mx_total[i] = sum(agg_mx_iban[ib][0] for ib in ibans_in_mx)
                    pnb_mx_total[i]  = sum(agg_mx_iban[ib][1] for ib in ibans_in_mx)
                    nb_mx_total[i]   = sum(agg_mx_iban[ib][2] for ib in ibans_in_mx)
                    match_mx[i] = 'IBAN_DIRECT'; matched_keys_mx[i] = ('IBAN', ibans_in_mx); nb_iban_dans_mx[i] = len(ibans_in_mx)
                    steps_mx.append(f"Flux MONEXT trouvés dans {fn['MONEXT']} via IBAN "
                                    f"({len(ibans_in_mx)}/{len(ibans_client)} IBAN rattachés) : {self.fmt(flux_mx_total[i])} €")
            if match_mx[i] == 'NO_MATCH':
                for key in rp_keys:
                    if key in agg_mx_rp:
                        flux_mx_total[i], pnb_mx_total[i], nb_mx_total[i] = agg_mx_rp[key]
                        match_mx[i] = 'DIRECT_RP'; matched_keys_mx[i] = ('RP', [key])
                        steps_mx.append(f"Flux MONEXT trouvés dans {fn['MONEXT']} via RP={key} "
                                        f"(testé : {rp_keys}) : {self.fmt(flux_mx_total[i])} €"); break
            if match_mx[i] == 'NO_MATCH':
                for key in rc_keys:
                    if key in agg_mx_rc:
                        flux_mx_total[i], pnb_mx_total[i], nb_mx_total[i] = agg_mx_rc[key]
                        match_mx[i] = 'DIRECT_RC'; matched_keys_mx[i] = ('RC', [key])
                        steps_mx.append(f"Flux MONEXT trouvés dans {fn['MONEXT']} via RC={key} "
                                        f"(testé : {rc_keys}) : {self.fmt(flux_mx_total[i])} €"); break
            if match_mx[i] == 'NO_MATCH' and rs and rs in agg_mx_rs:
                flux_mx_total[i], pnb_mx_total[i], nb_mx_total[i] = agg_mx_rs[rs]
                match_mx[i] = 'RS_EXACTE'; matched_keys_mx[i] = ('RS', [rs])
                steps_mx.append(f"Flux MONEXT trouvés dans {fn['MONEXT']} via Raison Sociale exacte « {rs} » : {self.fmt(flux_mx_total[i])} €")
            if match_mx[i] == 'NO_MATCH' and rs and len(rs) >= RS_INCLUSION_MIN:
                for k, vals in mx_rs_incl:
                    if rs in k or k in rs:
                        flux_mx_total[i], pnb_mx_total[i], nb_mx_total[i] = vals
                        match_mx[i] = 'RS_INCLUSION'; matched_keys_mx[i] = ('RS', [k])
                        steps_mx.append(f"Flux MONEXT trouvés dans {fn['MONEXT']} via Raison Sociale approchante « {rs} » ≈ « {k} » : {self.fmt(flux_mx_total[i])} €"); break
            if match_mx[i] == 'NO_MATCH':
                parts = [f"Client NON trouvé dans {fn['MONEXT']} après toutes les tentatives"]
                parts.append(f"IBAN testés ({len(ibans_client)}) : aucun présent dans MONEXT"
                             if ibans_client else "IBAN : aucun IBAN rattaché au client")
                parts.append(f"RP testés {_lst(cand_rp[i])} : aucun match")
                parts.append(f"RC testés {_lst(cand_rc[i])} : aucun match")
                parts.append(f"RS exacte « {rs} » : absente" if rs else "RS : nom entité vide")
                if rs and len(rs) >= RS_INCLUSION_MIN:
                    parts.append("RS inclusion : aucun rapprochement")
                parts.append("→ flux/utilisation CCO non renseignés (pas 0)")
                steps_mx.append(" | ".join(parts))

            # ===== WORLDLINE (pas d'index RP propre → RC+RP testés sur l'index RC) =====
            wl_id_keys = sorted(cand_rc[i] | cand_rp[i])
            if ibans_client:
                ibans_in_wl = [ib for ib in ibans_client if ib in agg_wl_iban]
                if ibans_in_wl:
                    flux_wl_total[i] = sum(agg_wl_iban[ib][0] for ib in ibans_in_wl)
                    pnb_wl_total[i]  = sum(agg_wl_iban[ib][1] for ib in ibans_in_wl)
                    nb_wl_total[i]   = sum(agg_wl_iban[ib][2] for ib in ibans_in_wl)
                    match_wl[i] = 'IBAN_DIRECT'; matched_keys_wl[i] = ('IBAN', ibans_in_wl); nb_iban_dans_wl[i] = len(ibans_in_wl)
                    steps_wl.append(f"Flux WORLDLINE trouvés dans {fn['WORLDLINE']} via IBAN "
                                    f"({len(ibans_in_wl)}/{len(ibans_client)} IBAN rattachés) : {self.fmt(flux_wl_total[i])} €")
            if match_wl[i] == 'NO_MATCH':
                for key in wl_id_keys:
                    if key in agg_wl_rc:
                        flux_wl_total[i], pnb_wl_total[i], nb_wl_total[i] = agg_wl_rc[key]
                        match_wl[i] = 'DIRECT_RC'; matched_keys_wl[i] = ('RC', [key])
                        steps_wl.append(f"Flux WORLDLINE trouvés dans {fn['WORLDLINE']} via RC/RP={key} "
                                        f"(testé : {wl_id_keys}) : {self.fmt(flux_wl_total[i])} €"); break
            if match_wl[i] == 'NO_MATCH' and rs and rs in agg_wl_rs:
                flux_wl_total[i], pnb_wl_total[i], nb_wl_total[i] = agg_wl_rs[rs]
                match_wl[i] = 'RS_EXACTE'; matched_keys_wl[i] = ('RS', [rs])
                steps_wl.append(f"Flux WORLDLINE trouvés dans {fn['WORLDLINE']} via Raison Sociale exacte « {rs} » : {self.fmt(flux_wl_total[i])} €")
            if match_wl[i] == 'NO_MATCH' and rs and len(rs) >= RS_INCLUSION_MIN:
                for k, vals in wl_rs_incl:
                    if rs in k or k in rs:
                        flux_wl_total[i], pnb_wl_total[i], nb_wl_total[i] = vals
                        match_wl[i] = 'RS_INCLUSION'; matched_keys_wl[i] = ('RS', [k])
                        steps_wl.append(f"Flux WORLDLINE trouvés dans {fn['WORLDLINE']} via Raison Sociale approchante « {rs} » ≈ « {k} » : {self.fmt(flux_wl_total[i])} €"); break
            if match_wl[i] == 'NO_MATCH':
                parts = [f"Client NON trouvé dans {fn['WORLDLINE']} après toutes les tentatives"]
                parts.append(f"IBAN testés ({len(ibans_client)}) : aucun présent dans Worldline"
                             if ibans_client else "IBAN : aucun IBAN rattaché au client")
                parts.append(f"RC/RP testés {_lst(cand_rc[i] | cand_rp[i])} : aucun match")
                parts.append(f"RS exacte « {rs} » : absente" if rs else "RS : nom entité vide")
                if rs and len(rs) >= RS_INCLUSION_MIN:
                    parts.append("RS inclusion : aucun rapprochement")
                parts.append("→ flux/utilisation CPC non renseignés (pas 0)")
                steps_wl.append(" | ".join(parts))

            diag_flux_mx[i] = " | ".join(steps_mx); diag_flux_wl[i] = " | ".join(steps_wl)

        self._prog(0.66, "Matching terminé.")

        # ── Comptage des récupérations dues à la cascade étendue V2 ──────
        # PARC : ligne où PARC a apporté des clés ET match obtenu via RP/RC.
        n_parc_recovered = int(sum(
            1 for i in range(total_rows)
            if parc_flag[i] and (
                match_mx[i] in ('DIRECT_RP', 'DIRECT_RC') or match_wl[i] == 'DIRECT_RC')
        ))
        # NOM→IBAN : ligne où le nom exact a ajouté des IBAN ET match obtenu via IBAN.
        n_nom_recovered = int(sum(
            1 for i in range(total_rows)
            if nom_flag[i] and (
                match_mx[i] == 'IBAN_DIRECT' or match_wl[i] == 'IBAN_DIRECT')
        ))

        # ── ÉTAPE 8 — BEJO override (MX) ─────────────────────────────────
        _mx_monthly_bejo = {}; n_bejo_ok = 0
        if expo_cfg['use_bejo'] and self.files.get("BEJO_CARTES") and self.files.get("BEJO_FLUX"):
            self._prog(0.68, "Traitement BEJO...")
            df_bc = self.load_csv_smart(self.files["BEJO_CARTES"]); df_bf = self.load_csv_smart(self.files["BEJO_FLUX"])
            col_bc_ent = m.get('bejo_cartes_entite', df_bc.columns[0])
            col_bc_nb  = m.get('bejo_cartes_nb', df_bc.columns[2] if len(df_bc.columns) > 2 else df_bc.columns[-1])
            col_bf_ent = m.get('bejo_flux_entite', df_bf.columns[0])
            col_bf_iban= m.get('bejo_flux_iban', df_bf.columns[5] if len(df_bf.columns) > 5 else df_bf.columns[-1])
            df_bc['_K_ENT'] = self.normalize_label(df_bc[col_bc_ent])
            bejo_cartes = {}
            for _, r in df_bc[df_bc['_K_ENT'] != ''].drop_duplicates('_K_ENT').iterrows():
                bejo_cartes[r['_K_ENT']] = r[col_bc_nb]
            df_bf['_K_ENT'] = self.normalize_label(df_bf[col_bf_ent]); df_bf['_K_IBAN'] = self.clean_iban(df_bf[col_bf_iban])
            bejo_flux = {}
            for _, r in df_bf[(df_bf['_K_ENT'] != '') & (df_bf['_K_IBAN'] != '')].iterrows():
                bejo_flux.setdefault(r['_K_ENT'], [])
                if r['_K_IBAN'] not in bejo_flux[r['_K_ENT']]:
                    bejo_flux[r['_K_ENT']].append(r['_K_IBAN'])
            arr_ent_src = self.normalize_label(df_source[m['source_rs']]).values
            for i in range(total_rows):
                ent_norm = arr_ent_src[i]
                if not ent_norm or ent_norm not in bejo_cartes:
                    continue
                ibans = bejo_flux.get(ent_norm, [])
                ibans_ok = [ib for ib in ibans if ib in agg_mx_iban]
                if not ibans_ok:
                    continue
                n_bejo_ok += 1
                nb_ref = self.to_float(pd.Series([str(bejo_cartes[ent_norm])])).iloc[0]
                flux_mx_total[i] = sum(agg_mx_iban[ib][0] for ib in ibans_ok)
                pnb_mx_total[i]  = sum(agg_mx_iban[ib][1] for ib in ibans_ok)
                nb_mx_total[i]   = nb_ref
                match_mx[i] = 'BEJO'; matched_keys_mx[i] = ('IBAN_BEJO', ibans_ok)
                for mois in all_mois:
                    d_iban = mx_mo_iban.get(mois, {})
                    fl_s = sum(d_iban[ib][0] for ib in ibans_ok if ib in d_iban)
                    p_s  = sum(d_iban[ib][1] for ib in ibans_ok if ib in d_iban)
                    nb_s = sum(d_iban[ib][2] for ib in ibans_ok if ib in d_iban)
                    nb_final = nb_ref if mois == ref_yyyymm else nb_s
                    _mx_monthly_bejo.setdefault(mois, {})[i] = (fl_s, p_s, nb_final)
                diag_flux_mx[i] += f" | Override BEJO appliqué ({len(ibans_ok)} IBAN via {fn['BEJO_FLUX']})"
            self._prog(0.70, f"BEJO : {n_bejo_ok} entités matchées")

        # ── ÉTAPE 9 — SÉRIES MENSUELLES ──────────────────────────────────
        self._prog(0.72, f"Séries mensuelles ({len(all_mois)} mois)...")
        monthly_mx = {}; monthly_wl = {}
        for mois in all_mois:
            fc = np.zeros(total_rows); pc = np.zeros(total_rows); nc = np.zeros(total_rows)
            d_iban = mx_mo_iban.get(mois, {}); d_rp = mx_mo_rp.get(mois, {}); d_rc = mx_mo_rc.get(mois, {})
            bejo_mo = _mx_monthly_bejo.get(mois, {})
            for i in range(total_rows):
                if i in bejo_mo:
                    fc[i], pc[i], nc[i] = bejo_mo[i]; continue
                mk = matched_keys_mx[i]
                if mk is None:
                    continue
                mtype, keys = mk
                if mtype == 'IBAN':
                    for k in keys:
                        if k in d_iban:
                            fl, pn, n = d_iban[k]; fc[i] += fl; pc[i] += pn; nc[i] += n
                elif mtype == 'RP':
                    for k in keys:
                        if k in d_rp:
                            fc[i], pc[i], nc[i] = d_rp[k]; break
                elif mtype == 'RC':
                    for k in keys:
                        if k in d_rc:
                            fc[i], pc[i], nc[i] = d_rc[k]; break
                elif mtype == 'RS':
                    df_m = df_monext[(df_monext['_MOIS'] == mois) & (df_monext['_K_RS'].isin(keys))]
                    if len(df_m) > 0:
                        fc[i] = df_m['_FLUX'].sum(); pc[i] = df_m['_PNB'].sum(); nc[i] = df_m['_NB_CARTES'].sum()
            monthly_mx[mois] = (fc, pc, nc)
        for mois in all_mois:
            fc = np.zeros(total_rows); pc = np.zeros(total_rows); nc = np.zeros(total_rows)
            d_iban = wl_mo_iban.get(mois, {}); d_rc = wl_mo_rc.get(mois, {})
            for i in range(total_rows):
                mk = matched_keys_wl[i]
                if mk is None:
                    continue
                mtype, keys = mk
                if mtype == 'IBAN':
                    for k in keys:
                        if k in d_iban:
                            fl, pn, n = d_iban[k]; fc[i] += fl; pc[i] += pn; nc[i] += n
                elif mtype == 'RC':
                    for k in keys:
                        if k in d_rc:
                            fc[i], pc[i], nc[i] = d_rc[k]; break
                elif mtype == 'RS':
                    df_w = df_worldline[(df_worldline['_MOIS'] == mois) & (df_worldline['_K_RS'].isin(keys))]
                    if len(df_w) > 0:
                        fc[i] = df_w['_FLUX'].sum(); pc[i] = df_w['_PNB'].sum(); nc[i] = df_w['_NB_CARTES'].sum()
            monthly_wl[mois] = (fc, pc, nc)

        # ── ÉTAPE 10 — ATTRIBUTS WORLDLINE (plafond + périodicité) ───────
        self._prog(0.74, "Plafonds & périodicité Worldline...")
        df_worldline['_PLAFOND'] = self.to_float(df_worldline[m['worldline_plafond']]) * taux_array
        df_worldline['_PERIO']   = self.to_float(df_worldline[m['worldline_periodicite']])

        def build_wl_attr(keycol):
            d = {}
            sub = df_worldline[(df_worldline[keycol] != '') & (df_worldline['_MOIS'] != '')]
            g = sub.groupby([keycol, '_MOIS']).agg(_PLAF=('_PLAFOND', 'sum'), _PERIO=('_PERIO', 'max'))
            for (k, mo), row in g.iterrows():
                d.setdefault(k, {})[mo] = (float(row['_PLAF']), float(row['_PERIO']))
            return d
        wl_attr_iban = build_wl_attr('_K_IBAN')
        wl_attr_rc   = dict(build_wl_attr('_K_RC'))
        for k, v in build_wl_attr('_K_RC_SAFE').items():
            wl_attr_rc.setdefault(k, v)

        def pick_attr(d, key, ref):
            if key not in d:
                return None
            months_le = sorted([mo for mo in d[key] if mo <= ref])
            if months_le:
                return d[key][months_le[-1]]
            allm = sorted(d[key].keys())
            return d[key][allm[0]] if allm else None

        def wl_plafond_perio(i):
            """Renvoie (plafond_total, perio_max, trouve:bool) pour la ligne i côté WL."""
            mk = matched_keys_wl[i]
            if mk is None:
                return (0.0, 1.0, False)
            mtype, keys = mk
            plaf = 0.0; perio = 1.0; found = False
            if mtype == 'IBAN':
                for k in keys:
                    a = pick_attr(wl_attr_iban, k, ref_yyyymm)
                    if a is not None:
                        plaf += a[0]; perio = max(perio, a[1]); found = True
            elif mtype == 'RC':
                for k in keys:
                    a = pick_attr(wl_attr_rc, k, ref_yyyymm)
                    if a is not None:
                        plaf += a[0]; perio = max(perio, a[1]); found = True; break
            elif mtype == 'RS':
                sub = df_worldline[(df_worldline['_K_RS'].isin(keys)) & (df_worldline['_MOIS'] != '') &
                                   (df_worldline['_MOIS'] <= ref_yyyymm)]
                if len(sub) > 0:
                    last_mo = sub['_MOIS'].max()
                    sub2 = sub[sub['_MOIS'] == last_mo]
                    plaf = float(sub2['_PLAFOND'].sum()); perio = float(sub2['_PERIO'].max()); found = True
            return (plaf, perio, found)

        # ── ÉTAPE 11 — DIFFÉRÉ, PLAFOND SOURCE, EXPO THÉORIQUE ───────────
        self._prog(0.76, "Lecture différé / plafonds / exposition théorique...")
        if differe_col and differe_col in df_source.columns:
            arr_differe = df_source[differe_col].apply(self.parse_differe).values
        else:
            arr_differe = np.full(total_rows, None, dtype=object)

        col_plaf_src = expo_cfg['col_plafond_source']
        if col_plaf_src and col_plaf_src in df_source.columns:
            raw_plaf = df_source[col_plaf_src].astype(str).str.strip().values
            val_plaf = self.to_float(df_source[col_plaf_src]).values
        else:
            raw_plaf = np.full(total_rows, '', dtype=object); val_plaf = np.zeros(total_rows)

        col_expo_th = expo_cfg['col_expo_theorique']
        if col_expo_th and col_expo_th in df_source.columns:
            raw_expo_th = df_source[col_expo_th].astype(str).str.strip().values
            val_expo_th = self.to_float(df_source[col_expo_th]).values
        else:
            raw_expo_th = np.full(total_rows, '', dtype=object); val_expo_th = np.zeros(total_rows)

        def is_empty(s):
            return (s is None) or (str(s).strip().lower() in ('', 'nan', 'none', 'null', 'n/a', 'na'))

        # ── ÉTAPE 12 — NB CARTES RÉF (MX) : moyenne 3 mois, anti-zéro ────
        self._prog(0.78, "Nb cartes de référence (moyenne 3 mois)...")
        mois_le_ref = sorted([mo for mo in all_mois if mo <= ref_yyyymm])
        window_ref  = mois_le_ref[-3:]  # mois réf + 2 précédents (présents)
        nb_cartes_ref = np.full(total_rows, None, dtype=object)
        for i in range(total_rows):
            if match_mx[i] == 'NO_MATCH':
                continue  # client non trouvé en MX → non calculable
            if not window_ref:
                continue
            vals = [float(monthly_mx[mo][2][i]) for mo in window_ref]
            if all(v == 0 for v in vals):
                nb_cartes_ref[i] = 0.0
                continue
            avg = sum(vals) / len(vals)
            r = self.round_half_up(avg)
            if r == 0 and any(v > 0 for v in vals):
                r = 1
            nb_cartes_ref[i] = float(r)

        # ── ÉTAPE 13 — EXPOSITION (gérée "non calculable") ───────────────
        self._prog(0.80, "Calcul de l'exposition réelle...")
        expo_mx = np.full(total_rows, None, dtype=object); expo_wl = np.full(total_rows, None, dtype=object)
        expo_reelle = np.full(total_rows, None, dtype=object)
        mult_arr = np.full(total_rows, None, dtype=object); fact_perio_arr = np.full(total_rows, None, dtype=object)
        plaf_wl_arr = np.full(total_rows, None, dtype=object)
        expo_mx_statut = np.full(total_rows, '', dtype=object); expo_wl_statut = np.full(total_rows, '', dtype=object)
        expo_statut = np.full(total_rows, '', dtype=object)
        expo_ecart = np.full(total_rows, None, dtype=object); expo_comp = np.full(total_rows, '', dtype=object)

        for i in range(total_rows):
            d = arr_differe[i]
            found_mx = match_mx[i] != 'NO_MATCH'; found_wl = match_wl[i] != 'NO_MATCH'

            if d is None:
                expo_statut[i] = "NON CALCULABLE – différé non renseigné"
                expo_mx_statut[i] = expo_statut[i]; expo_wl_statut[i] = expo_statut[i]
                continue
            mult = 0 if d < 1 else (2 if d < 28 else 3)
            mult_arr[i] = mult

            # ----- côté MX -----
            plaf_src_present = not is_empty(raw_plaf[i])
            mx_calc = found_mx and plaf_src_present and (nb_cartes_ref[i] is not None)
            if not found_mx:
                expo_mx_statut[i] = "N/A (client absent de MONEXT)"
            elif not plaf_src_present:
                expo_mx_statut[i] = "NON CALCULABLE – plafond source (MONEXT) introuvable"
            else:
                expo_mx_statut[i] = "CALCULÉE"
                expo_mx[i] = mult * float(val_plaf[i]) * float(nb_cartes_ref[i])

            # ----- côté WL -----
            plaf_wl, perio, wl_found_attr = wl_plafond_perio(i)
            fact_perio = (1.0 / 12.0) if perio == 12 else 1.0
            fact_perio_arr[i] = fact_perio
            plaf_wl_present = found_wl and wl_found_attr
            if found_wl:
                plaf_wl_arr[i] = plaf_wl
            wl_calc = found_wl and plaf_wl_present
            if not found_wl:
                expo_wl_statut[i] = "N/A (client absent de Worldline)"
            elif not plaf_wl_present:
                expo_wl_statut[i] = "NON CALCULABLE – plafond Worldline introuvable"
            else:
                expo_wl_statut[i] = "CALCULÉE"
                expo_wl[i] = mult * fact_perio * float(plaf_wl)

            # ----- global -----
            if not found_mx and not found_wl:
                expo_statut[i] = "NON CALCULABLE – client non trouvé (ni MONEXT ni Worldline)"
                continue
            problems = []
            if found_mx and not mx_calc:
                problems.append("plafond MONEXT (source) introuvable")
            if found_wl and not wl_calc:
                problems.append("plafond Worldline introuvable")
            if problems:
                expo_statut[i] = "NON CALCULABLE – " + " ; ".join(problems)
                continue
            total = (float(expo_mx[i]) if mx_calc else 0.0) + (float(expo_wl[i]) if wl_calc else 0.0)
            expo_reelle[i] = total
            expo_statut[i] = "CALCULÉE"
            ec = total - float(val_expo_th[i])
            expo_ecart[i] = ec
            if abs(ec) < 0.005:
                expo_comp[i] = "ÉGALE"
            elif ec > 0:
                expo_comp[i] = "SUPÉRIEURE (dépassement vs théorique)"
            else:
                expo_comp[i] = "INFÉRIEURE"

        # ── ÉTAPE 14 — UTILISATION (fenêtre différé, vide si non trouvé) ─
        self._prog(0.82, "Calcul de l'utilisation...")
        util_mx = np.full(total_rows, None, dtype=object); util_wl = np.full(total_rows, None, dtype=object)
        util_reelle = np.full(total_rows, None, dtype=object)
        for i in range(total_rows):
            d = arr_differe[i]
            found_mx = match_mx[i] != 'NO_MATCH'; found_wl = match_wl[i] != 'NO_MATCH'
            if d is None:
                continue
            if not found_mx and not found_wl:
                continue  # vide + raison (cf. diag)
            N = 0 if d < 1 else (2 if d < 28 else 3)
            win = mois_le_ref[-N:] if N > 0 else []
            umx = sum(float(monthly_mx[mo][0][i]) for mo in win)
            uwl = sum(float(monthly_wl[mo][0][i]) for mo in win)
            util_mx[i] = umx; util_wl[i] = uwl; util_reelle[i] = umx + uwl

        # ── ÉTAPE 15 — TRIMESTRES & TOTAUX ANNUELS ───────────────────────
        self._prog(0.84, "Agrégats trimestriels & annuels...")

        cur_an, cur_q = (0, 0)
        if all_mois:
            _, cur_an, cur_q = self.quarter_of(max(all_mois))
        quarters = []
        if cur_an >= 2025:
            an = 2025; q = 1
            while (an < cur_an) or (an == cur_an and q <= cur_q):
                quarters.append((an, q))
                q += 1
                if q > 4:
                    q = 1; an += 1
        else:
            seen = []
            for mo in all_mois:
                _, a, qq = self.quarter_of(mo)
                if (a, qq) not in seen:
                    seen.append((a, qq))
            quarters = sorted(seen)

        def months_of_quarter(an, q):
            return [mo for mo in all_mois if self.quarter_of(mo)[1] == an and self.quarter_of(mo)[2] == q]

        q_flux_mx = {}; q_flux_wl = {}; q_cartes_mx = {}; q_cartes_wl = {}; q_pnb_mx = {}; q_pnb_wl = {}
        for (an, q) in quarters:
            mos = months_of_quarter(an, q)
            fmx = np.zeros(total_rows); fwl = np.zeros(total_rows)
            cmx = np.zeros(total_rows); cwl = np.zeros(total_rows)
            pmx = np.zeros(total_rows); pwl = np.zeros(total_rows)
            for mo in mos:
                fmx += monthly_mx[mo][0]; pmx += monthly_mx[mo][1]; cmx += monthly_mx[mo][2]
                fwl += monthly_wl[mo][0]; pwl += monthly_wl[mo][1]; cwl += monthly_wl[mo][2]
            key = (an, q)
            q_flux_mx[key] = fmx; q_flux_wl[key] = fwl
            q_cartes_mx[key] = cmx; q_cartes_wl[key] = cwl
            q_pnb_mx[key] = pmx; q_pnb_wl[key] = pwl

        annees = sorted({mo[:4] for mo in all_mois})
        an_flux_mx = {}; an_flux_wl = {}; an_cartes_mx = {}; an_cartes_wl = {}; an_pnb_mx = {}; an_pnb_wl = {}
        for an in annees:
            mos = [mo for mo in all_mois if mo[:4] == an]
            fmx = np.zeros(total_rows); fwl = np.zeros(total_rows)
            cmx = np.zeros(total_rows); cwl = np.zeros(total_rows)
            pmx = np.zeros(total_rows); pwl = np.zeros(total_rows)
            for mo in mos:
                fmx += monthly_mx[mo][0]; pmx += monthly_mx[mo][1]; cmx += monthly_mx[mo][2]
                fwl += monthly_wl[mo][0]; pwl += monthly_wl[mo][1]; cwl += monthly_wl[mo][2]
            an_flux_mx[an] = fmx; an_flux_wl[an] = fwl
            an_cartes_mx[an] = cmx; an_cartes_wl[an] = cwl
            an_pnb_mx[an] = pmx; an_pnb_wl[an] = pwl

        flux_mx_tot = np.zeros(total_rows); flux_wl_tot = np.zeros(total_rows)
        pnb_mx_tot  = np.zeros(total_rows); pnb_wl_tot  = np.zeros(total_rows)
        nb_mx_tot   = np.zeros(total_rows); nb_wl_tot   = np.zeros(total_rows)
        for mo in all_mois:
            flux_mx_tot += monthly_mx[mo][0]; pnb_mx_tot += monthly_mx[mo][1]; nb_mx_tot += monthly_mx[mo][2]
            flux_wl_tot += monthly_wl[mo][0]; pnb_wl_tot += monthly_wl[mo][1]; nb_wl_tot += monthly_wl[mo][2]

        depassement = np.full(total_rows, None, dtype=object)
        cur_key = (cur_an, cur_q)
        if cur_key in q_flux_mx:
            for i in range(total_rows):
                if is_empty(raw_expo_th[i]):
                    continue
                depassement[i] = float(q_flux_mx[cur_key][i]) + float(q_flux_wl[cur_key][i]) - float(val_expo_th[i])

        # ── ÉTAPE 16 — CONSTRUCTION DE LA SORTIE ─────────────────────────
        self._prog(0.87, "Construction du tableau de sortie...")
        out = {}; col_block = []

        def push(name, vals, block, kind='num'):
            if kind == 'num':
                out[name] = ['' if (v is None or (isinstance(v, float) and math.isnan(v))) else self.fmt(v) for v in vals]
            elif kind == 'int':
                out[name] = ['' if v is None else str(int(round(float(v)))) for v in vals]
            else:
                out[name] = ['' if v is None else str(v) for v in vals]
            col_block.append((name, block))

        # A — colonnes SOURCE d'origine
        src_orig_cols = [c for c in df_source.columns if not c.startswith('_K_')]
        # Renommage à l'affichage des colonnes SOURCE (demande Bénédicte) :
        # on ne touche NI la donnée NI les clés de calcul, uniquement le libellé en sortie.
        def _norm_src_name(s):
            return " ".join(str(s).split()).lower()
        SOURCE_COL_RENAME = {
            "exposition - total exposition (eur)": "Projected EXPOSURE (EUR)",
        }
        for c in src_orig_cols:
            disp = SOURCE_COL_RENAME.get(_norm_src_name(c), c)
            push(disp, df_source[c].astype(str).values, 'A_SOURCE', kind='text')

        # B — identité
        push("RMPM",            id_rmpm,   'B_IDENTITY', kind='text')
        push("CODE_GA",         id_code_ga,'B_IDENTITY', kind='text')
        push("NOM_GA",          id_nom_ga, 'B_IDENTITY', kind='text')
        push("PAYS_GA",         id_pays_ga,'B_IDENTITY', kind='text')
        push("NOM_LE",          id_nom_le, 'B_IDENTITY', kind='text')
        push("PAYS_LE",         id_pays_le,'B_IDENTITY', kind='text')
        push("RS_REFERENTIEL",  id_rs_ref, 'B_IDENTITY', kind='text')
        push("SEGMENT",         id_segment,'B_IDENTITY', kind='text')
        push("SOURCE_IDENTITE", id_source, 'B_IDENTITY', kind='text')
        push("SOURCE_SEGMENT",  id_seg_src,'B_IDENTITY', kind='text')
        push("IBANS_RATTACHES", [" ; ".join(x) for x in id_ibans], 'B_IDENTITY', kind='text')

        # C — matching
        push("MATCH_MONEXT",   match_mx, 'C_MATCHING', kind='text')
        push("MATCH_WORLDLINE",match_wl, 'C_MATCHING', kind='text')
        push("NB_IBAN_TROUVES",nb_iban_trouves,'C_MATCHING', kind='int')
        push("NB_IBAN_DANS_MX",nb_iban_dans_mx,'C_MATCHING', kind='int')
        push("NB_IBAN_DANS_WL",nb_iban_dans_wl,'C_MATCHING', kind='int')

        # D1 — nb cartes MX (mensuel + total + annuel)
        for mo in all_mois:
            push(f"NB CARTES MX {self.mois_label(mo)}", monthly_mx[mo][2], 'D1_NB_MX', kind='int')
        push("NB CARTES MX TOTAL", nb_mx_tot, 'D1_NB_MX', kind='int')
        for an in annees:
            push(f"NB CARTES MX {an}", an_cartes_mx[an], 'D1_NB_MX', kind='int')

        # D2 — flux MX (mensuel + trimestriel + total + annuel "Mt MX")
        for mo in all_mois:
            push(f"FLUX MX {self.mois_label(mo)}", monthly_mx[mo][0], 'D2_FLUX_MX')
        for (an, q) in quarters:
            push(f"FLUX TRIM MX {self.quarter_label_long(an, q)}", q_flux_mx[(an, q)], 'D2_FLUX_MX')
        push("FLUX MX TOTAL", flux_mx_tot, 'D2_FLUX_MX')
        for an in annees:
            push(f"Mt MX {an}", an_flux_mx[an], 'D2_FLUX_MX')

        # D3 — PNB MX
        for mo in all_mois:
            push(f"PNB MX {self.mois_label(mo)}", monthly_mx[mo][1], 'D3_PNB_MX')
        push("PNB MX TOTAL", pnb_mx_tot, 'D3_PNB_MX')
        for an in annees:
            push(f"PNB MX {an}", an_pnb_mx[an], 'D3_PNB_MX')

        # E1 — nb cartes WL
        for mo in all_mois:
            push(f"NB CARTES WL {self.mois_label(mo)}", monthly_wl[mo][2], 'E1_NB_WL', kind='int')
        push("NB CARTES WL TOTAL", nb_wl_tot, 'E1_NB_WL', kind='int')
        for an in annees:
            push(f"NB CARTES WL {an}", an_cartes_wl[an], 'E1_NB_WL', kind='int')

        # E2 — flux WL
        for mo in all_mois:
            push(f"FLUX WL {self.mois_label(mo)}", monthly_wl[mo][0], 'E2_FLUX_WL')
        for (an, q) in quarters:
            push(f"FLUX TRIM WL {self.quarter_label_long(an, q)}", q_flux_wl[(an, q)], 'E2_FLUX_WL')
        push("FLUX WL TOTAL", flux_wl_tot, 'E2_FLUX_WL')
        for an in annees:
            push(f"Mt WL {an}", an_flux_wl[an], 'E2_FLUX_WL')

        # E3 — PNB WL
        for mo in all_mois:
            push(f"PNB WL {self.mois_label(mo)}", monthly_wl[mo][1], 'E3_PNB_WL')
        push("PNB WL TOTAL", pnb_wl_tot, 'E3_PNB_WL')
        for an in annees:
            push(f"PNB WL {an}", an_pnb_wl[an], 'E3_PNB_WL')

        # F — exposition / utilisation / composantes
        push("DIFFERE_JOURS",        arr_differe,     'F_EXPOSITION')
        push("MULTIPLICATEUR_N",     mult_arr,        'F_EXPOSITION', kind='int')
        push("PLAFOND_SOURCE",       [val_plaf[i] if not is_empty(raw_plaf[i]) else None for i in range(total_rows)], 'F_EXPOSITION')
        push("NB_CARTES_REF_MX",     nb_cartes_ref,   'F_EXPOSITION', kind='int')
        push("PLAFOND_WL",           plaf_wl_arr,     'F_EXPOSITION')
        push("FACTEUR_PERIODICITE",  fact_perio_arr,  'F_EXPOSITION')
        push("EXPOSITION_MX",        expo_mx,         'F_EXPOSITION')
        push("EXPOSITION_WL",        expo_wl,         'F_EXPOSITION')
        push("ACTUAL_EXPOSURE",    expo_reelle,     'F_EXPOSITION')
        push("EXPOSITION_THEORIQUE", [val_expo_th[i] if not is_empty(raw_expo_th[i]) else None for i in range(total_rows)], 'F_EXPOSITION')
        push("EXPO_ECART",           expo_ecart,      'F_EXPOSITION')
        push("EXPO_COMPARAISON",     expo_comp,       'F_EXPOSITION', kind='text')
        push("EXPO_MX_STATUT",       expo_mx_statut,  'F_EXPOSITION', kind='text')
        push("EXPO_WL_STATUT",       expo_wl_statut,  'F_EXPOSITION', kind='text')
        push("EXPO_STATUT",          expo_statut,     'F_EXPOSITION', kind='text')
        push("UTILISATION_MX",       util_mx,         'F_EXPOSITION')
        push("UTILISATION_WL",       util_wl,         'F_EXPOSITION')
        push("UTILISATION_REELLE",   util_reelle,     'F_EXPOSITION')
        push("DEPASSEMENT_UTILISATION", depassement,  'F_EXPOSITION')

        # G — diagnostic
        push("DIAG_IDENTITE", diag_identity, 'G_DIAG', kind='text')
        push("DIAG_FLUX_MX",  diag_flux_mx,  'G_DIAG', kind='text')
        push("DIAG_FLUX_WL",  diag_flux_wl,  'G_DIAG', kind='text')

        df_out = pd.DataFrame(out)

        # ── ÉTAPE 17 — EMPLACEMENT DE SORTIE + SAUVEGARDE CSV ────────────
        # [CLI] Remplace l'ancien dialogue asksaveasfilename (boîte d'emplacement) :
        # le dossier et le nom de base proviennent de --output-dir / --output-filename.
        self._prog(0.89, "Préparation de l'emplacement de sortie...")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        base_noext = str(self.output_dir / self.output_filename)
        csv_path  = base_noext + ".csv"
        xlsx_path = base_noext + ".xlsx"
        save_note = ""

        self._prog(0.90, "Sauvegarde CSV...")
        df_out.to_csv(csv_path, sep=';', index=False, encoding='utf-8-sig')

        # ── Données pour les snapshots de synthèse ───────────────────────
        snap = {
            'ref_yyyymm': ref_yyyymm,
            'rmpm': arr_rmpm, 'statut': arr_statut, 'differe': arr_differe,
            'found_mx': np.array([match_mx[i] != 'NO_MATCH' for i in range(total_rows)]),
            'found_wl': np.array([match_wl[i] != 'NO_MATCH' for i in range(total_rows)]),
            'expo_mx': expo_mx, 'expo_wl': expo_wl, 'expo_reelle': expo_reelle,
            'expo_statut': expo_statut, 'expo_comp': expo_comp,
            'val_expo_th': val_expo_th, 'raw_expo_th': raw_expo_th,
            'quarters': quarters,
            'q_flux_mx': q_flux_mx, 'q_flux_wl': q_flux_wl,
            'q_cartes_mx': q_cartes_mx, 'q_cartes_wl': q_cartes_wl,
            'q_pnb_mx': q_pnb_mx, 'q_pnb_wl': q_pnb_wl,
            'cur_key': cur_key,
        }

        stats = {
            'total_rows': total_rows,
            'matched_mx': int((np.array(match_mx) != 'NO_MATCH').sum()),
            'matched_wl': int((np.array(match_wl) != 'NO_MATCH').sum()),
            'expo_calc': int((np.array(expo_statut) == 'CALCULÉE').sum()),
            'expo_noncalc': int(sum(1 for s in expo_statut if str(s).startswith('NON CALCULABLE'))),
            'nb_lignes_non_eur': nb_lignes_non_eur,
            'n_bejo': n_bejo_ok,
            'n_parc': n_parc_recovered,
            'n_nom_iban': n_nom_recovered,
            'all_mois': all_mois,
        }

        # ── ÉTAPE 18 — XLSX ──────────────────────────────────────────────
        self._prog(0.93, "Génération du fichier Excel...")
        if OPENPYXL_OK:
            self.generate_xlsx(xlsx_path, df_out, col_block, all_mois, snap, stats,
                               mx_global_by_mois, wl_global_by_mois)
        else:
            xlsx_path = None

        self._prog(1.0, "Terminé.")
        msg = (f"ANALYSE TERMINÉE [{VERSION_ID}]\n\n"
               f"Lignes traitées        : {stats['total_rows']:,}\n"
               f"Matchées MONEXT        : {stats['matched_mx']:,}\n"
               f"Matchées Worldline     : {stats['matched_wl']:,}\n"
               f"Récup. via PARC_CLIENT : {stats['n_parc']:,}\n"
               f"Récup. via NOM→IBAN    : {stats['n_nom_iban']:,}\n"
               f"Exposition calculée    : {stats['expo_calc']:,}\n"
               f"Exposition non calc.   : {stats['expo_noncalc']:,}\n"
               f"Lignes WL converties   : {stats['nb_lignes_non_eur']:,}\n"
               f"Override BEJO          : {stats['n_bejo']:,}\n\n"
               f"{(save_note + chr(10) + chr(10)) if save_note else ''}"
               f"CSV  : {csv_path}\n"
               f"XLSX : {xlsx_path if xlsx_path else 'openpyxl indisponible'}")
        print(msg)
        return {'csv_path': csv_path, 'xlsx_path': xlsx_path, 'stats': stats}

    # =========================================================================
    # GÉNÉRATION XLSX (VERBATIM)
    # =========================================================================
    NUMERIC_BLOCKS = {'D1_NB_MX', 'D2_FLUX_MX', 'D3_PNB_MX', 'E1_NB_WL', 'E2_FLUX_WL', 'E3_PNB_WL'}
    F_TEXT_COLS = {'EXPO_COMPARAISON', 'EXPO_MX_STATUT', 'EXPO_WL_STATUT', 'EXPO_STATUT'}
    INT_HINTS = ('NB CARTES', 'NB_CARTES', 'NB_IBAN', 'MULTIPLICATEUR')

    @staticmethod
    def _parse_fr(s):
        if s is None:
            return None
        t = str(s).strip()
        if t == '':
            return None
        try:
            return float(t.replace('\xa0', '').replace(' ', '').replace(',', '.'))
        except Exception:
            return None

    def _is_numeric_col(self, name, block):
        if block in self.NUMERIC_BLOCKS:
            return True
        if block == 'F_EXPOSITION' and name not in self.F_TEXT_COLS:
            return True
        return False

    # Affectation de chaque colonne à un côté pour le périmètre par feuille :
    #   SHARED  → présent partout (GLOBAL + CCO + CPC)
    #   MX      → présent dans GLOBAL + CCO (MONEXT)
    #   WL      → présent dans GLOBAL + CPC (WORLDLINE)
    #   GLOBAL  → présent uniquement dans REVENU GLOBAL
    _F_MX_COLS     = {'PLAFOND_SOURCE', 'NB_CARTES_REF_MX', 'EXPOSITION_MX', 'EXPO_MX_STATUT', 'UTILISATION_MX'}
    _F_WL_COLS     = {'PLAFOND_WL', 'FACTEUR_PERIODICITE', 'EXPOSITION_WL', 'EXPO_WL_STATUT', 'UTILISATION_WL'}
    _F_SHARED_COLS = {'DIFFERE_JOURS', 'MULTIPLICATEUR_N'}

    def _col_side(self, name, block):
        if block in ('A_SOURCE', 'B_IDENTITY'):
            return 'SHARED'
        if block in ('D1_NB_MX', 'D2_FLUX_MX', 'D3_PNB_MX'):
            return 'MX'
        if block in ('E1_NB_WL', 'E2_FLUX_WL', 'E3_PNB_WL'):
            return 'WL'
        if block == 'C_MATCHING':
            if name in ('MATCH_MONEXT', 'NB_IBAN_DANS_MX'):
                return 'MX'
            if name in ('MATCH_WORLDLINE', 'NB_IBAN_DANS_WL'):
                return 'WL'
            return 'SHARED'  # NB_IBAN_TROUVES
        if block == 'G_DIAG':
            if name == 'DIAG_FLUX_MX':
                return 'MX'
            if name == 'DIAG_FLUX_WL':
                return 'WL'
            return 'SHARED'  # DIAG_IDENTITE
        if block == 'F_EXPOSITION':
            if name in self._F_SHARED_COLS:
                return 'SHARED'
            if name in self._F_MX_COLS:
                return 'MX'
            if name in self._F_WL_COLS:
                return 'WL'
            return 'GLOBAL'  # ACTUAL_EXPOSURE, EXPOSITION_THEORIQUE, EXPO_ECART, EXPO_COMPARAISON, EXPO_STATUT, UTILISATION_REELLE, DEPASSEMENT_UTILISATION
        return 'SHARED'

    def _cols_for_sheet(self, col_block, kind):
        keep = []
        for (name, block) in col_block:
            side = self._col_side(name, block)
            if kind == 'GLOBAL' or side == 'SHARED' \
               or (kind == 'CCO' and side == 'MX') \
               or (kind == 'CPC' and side == 'WL'):
                keep.append(name)
        return keep

    def generate_xlsx(self, path, df_out, col_block, all_mois, snap, stats,
                      mx_global_by_mois, wl_global_by_mois):
        wb = Workbook()
        numeric_cols = {n for (n, b) in col_block if self._is_numeric_col(n, b)}

        # 1) DICTIONNAIRE (feuille 1)
        ws_dic = wb.active
        ws_dic.title = "DICTIONNAIRE"
        self._build_dictionnaire_sheet(ws_dic, col_block, all_mois, snap)

        # 2) SYNTHÈSE
        ws_syn = wb.create_sheet("SYNTHESE")
        self._build_synthese(ws_syn, df_out, snap, stats, mx_global_by_mois, wl_global_by_mois)

        # 3) Feuilles de détail GLOBAL / CCO / CPC
        found_mx = snap['found_mx']; found_wl = snap['found_wl']
        idx_global = list(range(len(df_out)))
        idx_cco = [i for i in range(len(df_out)) if found_mx[i]]
        idx_cpc = [i for i in range(len(df_out)) if found_wl[i]]
        for sheet_name, idx, kind in [("REVENU GLOBAL", idx_global, 'GLOBAL'),
                                      ("REVENU CCO (MONEXT)", idx_cco, 'CCO'),
                                      ("REVENU CPC (WORLDLINE)", idx_cpc, 'CPC')]:
            ws = wb.create_sheet(sheet_name)
            keep = self._cols_for_sheet(col_block, kind)
            sub = df_out.iloc[idx][keep].reset_index(drop=True)
            self._build_detail_sheet(ws, sub, col_block, numeric_cols)

        wb.save(path)

    def _build_dictionnaire_sheet(self, ws, col_block, all_mois, snap):
        thin = Side(style='thin', color='D0D0D0')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.merge_cells('A1:D1')
        c = ws['A1']; c.value = f"DICTIONNAIRE DES COLONNES CRÉÉES — CIB REVENUS ANALYZER v32 [{VERSION_ID}]"
        c.font = Font(bold=True, size=13, color='FFFFFF'); c.fill = PatternFill('solid', fgColor=BNP_GREEN.lstrip('#'))
        c.alignment = Alignment(horizontal='center', vertical='center'); ws.row_dimensions[1].height = 26
        hdr = ['Bloc', 'Colonne', 'Définition', 'Logique de calcul']
        for j, h in enumerate(hdr, start=1):
            cell = ws.cell(row=2, column=j, value=h)
            cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor='1E293B')
            cell.alignment = Alignment(horizontal='center'); cell.border = border

        DEFS = {
            "RMPM": ("Identifiant client RMPM résolu", "Cascade IBAN_ACCOUNT > REF_CLIENT (RC/RIB/RP) > PARC_CLIENT > héritage source."),
            "CODE_GA": ("Code Groupe d'Affaires", "Repris de IBAN_ACCOUNT ou REFERENTIEL_CLIENT selon le pivot ayant matché."),
            "SEGMENT": ("Segment marché (BPE / ENTREPRISE)", "REF_CLIENT puis fallback IDENTIFIANT_SEGMENT (RC > RMPM > RP)."),
            "SOURCE_IDENTITE": ("Pivot ayant fourni l'identité", "Trace la règle gagnante : IBAN/RMPM/RC/RIB/RP, PARC_REF_*, ou NOM_EXACT_IBAN."),
            "IBANS_RATTACHES": ("IBAN(s) rattaché(s) au client", "Union des IBAN trouvés via IBAN direct + RMPM (IBAN_ACCOUNT) + NOM exact (ACCOUNT/SINGLETON)."),
            "MATCH_MONEXT": ("Méthode de matching des flux MONEXT", "IBAN > DIRECT(RP/RC, toutes combinaisons) > RS exacte > RS inclusion > BEJO. NO_MATCH si rien."),
            "MATCH_WORLDLINE": ("Méthode de matching des flux Worldline", "IBAN > DIRECT(RC+RP) > RS exacte > RS inclusion. NO_MATCH si rien."),
            "FLUX MX <mois>": ("Flux MONEXT (dépenses+retraits) du mois", "Somme des lignes MONEXT du mois rattachées au client (déjà en EUR)."),
            "FLUX WL <mois>": ("Flux Worldline du mois", "Somme dépenses1+dépenses2 du mois, converties en EUR via la table devises."),
            "FLUX TRIM MX <Qx>": ("Flux MONEXT du trimestre", "Somme des 3 mois du trimestre. Trimestres de T1 2025 au trimestre courant."),
            "FLUX TRIM WL <Qx>": ("Flux Worldline du trimestre", "Somme des 3 mois du trimestre."),
            "Mt MX <année>": ("Flux MONEXT annuel (nom Bénédicte)", "Somme de tous les mois de l'année côté MONEXT."),
            "Mt WL <année>": ("Flux Worldline annuel (nom Bénédicte)", "Somme de tous les mois de l'année côté Worldline."),
            "NB CARTES MX <mois>": ("Nb cartes MONEXT du mois", "Somme du nb de cartes MONEXT du mois pour le client."),
            "PNB MX <mois>": ("PNB MONEXT du mois", "Somme de la plage PNB (col first→last, hors exclusion, interchange ×-1)."),
            "DIFFERE_JOURS": ("Différé de paiement (jours)", "Lu dans la colonne différé SOURCE. Vide si non renseigné."),
            "MULTIPLICATEUR_N": ("Multiplicateur N d'exposition", "N=0 si différé<1 ; N=2 si 1≤différé<28 ; N=3 si différé≥28."),
            "PLAFOND_SOURCE": ("Plafond mensuel carte (fichier source)", "Col 20 SOURCE (Onbording/Manel). Vide si absent → MX non calculable."),
            "NB_CARTES_REF_MX": ("Nb cartes de référence MONEXT", "Moyenne du nb cartes sur 3 mois (réf + 2 préc.), arrondi demi-haut. "
                                 "Anti-zéro : ≥1 carte → résultat ≥1. Vide si client absent de MONEXT."),
            "PLAFOND_WL": ("Plafond mensuel Worldline (EUR)", "Somme des plafonds des cartes WL au mois de réf (ou dernier mois ≤ réf), converti EUR."),
            "FACTEUR_PERIODICITE": ("Facteur de périodicité WL", "1/12 si périodicité=12 (annuel) ; 1 sinon."),
            "EXPOSITION_MX": ("Exposition CCO (MONEXT)", "N × plafond_source × nb_cartes_réf. Vide si un membre manque (cf. EXPO_MX_STATUT)."),
            "EXPOSITION_WL": ("Exposition CPC (Worldline)", "N × facteur_périodicité × plafond_WL. Vide si un membre manque."),
            "ACTUAL_EXPOSURE": ("Actual Exposure (exposition réelle totale)", "EXPO_MX + EXPO_WL. Vide si différé manquant, client introuvable, ou plafond introuvable d'un côté présent."),
            "EXPOSITION_THEORIQUE": ("Exposition théorique (col AG)", "Reprise de la colonne AG (source 32). Vide si non renseignée."),
            "EXPO_ECART": ("Écart exposition réelle − théorique", "Rempli uniquement si EXPO_STATUT = CALCULÉE. Positif = dépassement."),
            "EXPO_COMPARAISON": ("Position réelle vs théorique", "SUPÉRIEURE / ÉGALE / INFÉRIEURE. Rempli si CALCULÉE."),
            "EXPO_MX_STATUT": ("Statut calcul exposition MONEXT", "CALCULÉE / N/A (absent MONEXT) / NON CALCULABLE – plafond source introuvable."),
            "EXPO_WL_STATUT": ("Statut calcul exposition Worldline", "CALCULÉE / N/A (absent Worldline) / NON CALCULABLE – plafond Worldline introuvable."),
            "EXPO_STATUT": ("Statut global de l'exposition", "CALCULÉE ou NON CALCULABLE avec la raison explicite (jamais un 0 trompeur)."),
            "UTILISATION_MX": ("Utilisation CCO sur la fenêtre différé", "Somme des flux MONEXT sur les N derniers mois (N = multiplicateur)."),
            "UTILISATION_WL": ("Utilisation CPC sur la fenêtre différé", "Somme des flux Worldline sur les N derniers mois."),
            "UTILISATION_REELLE": ("Utilisation réelle totale", "MX + WL. Vide si client introuvable des deux côtés ou différé manquant."),
            "DEPASSEMENT_UTILISATION": ("Dépassement d'utilisation (trim. courant)", "Flux MX + WL du trimestre courant − exposition théorique (AG)."),
            "DIAG_IDENTITE": ("Diagnostic d'identité (lisible)", "Raconte la résolution : fichier + référence utilisés (IBAN/RMPM/RC/RP/PARC/NOM exact)."),
            "DIAG_FLUX_MX": ("Diagnostic flux MONEXT (lisible)", "Toutes les tentatives : IBAN testés, RP testés, RC testés, RS exacte, RS inclusion."),
            "DIAG_FLUX_WL": ("Diagnostic flux Worldline (lisible)", "Toutes les tentatives : IBAN testés, RC/RP testés, RS exacte, RS inclusion."),
        }
        BLOCK_LABEL = {k: v['name'] for k, v in BLOC_COLORS.items()}

        def generic_def(name):
            if name.startswith("FLUX MX "): return DEFS["FLUX MX <mois>"]
            if name.startswith("FLUX WL "): return DEFS["FLUX WL <mois>"]
            if name.startswith("FLUX TRIM MX "): return DEFS["FLUX TRIM MX <Qx>"]
            if name.startswith("FLUX TRIM WL "): return DEFS["FLUX TRIM WL <Qx>"]
            if name.startswith("Mt MX "): return DEFS["Mt MX <année>"]
            if name.startswith("Mt WL "): return DEFS["Mt WL <année>"]
            if name.startswith("NB CARTES MX "): return DEFS["NB CARTES MX <mois>"]
            if name.startswith("NB CARTES WL "): return ("Nb cartes Worldline", "Somme du nb de cartes Worldline pour la période.")
            if name.startswith("PNB MX "): return DEFS["PNB MX <mois>"]
            if name.startswith("PNB WL "): return ("PNB Worldline", "Somme de la plage PNB Worldline convertie en EUR.")
            return DEFS.get(name)

        r = 3
        seen = set()
        for (name, block) in col_block:
            if block == 'A_SOURCE':
                continue
            d = generic_def(name)
            if d is None:
                continue
            key = (block, d[0])
            if key in seen and not name.startswith(("FLUX", "PNB", "NB CARTES", "Mt ")):
                continue
            fam = None
            for pref in ["FLUX TRIM MX ", "FLUX TRIM WL ", "FLUX MX ", "FLUX WL ", "Mt MX ", "Mt WL ",
                         "NB CARTES MX ", "NB CARTES WL ", "PNB MX ", "PNB WL "]:
                if name.startswith(pref):
                    fam = pref; break
            if fam:
                if fam in seen:
                    continue
                seen.add(fam)
                disp = fam.strip() + " …"
            else:
                if name in seen:
                    continue
                seen.add(name); disp = name
            ws.cell(row=r, column=1, value=BLOCK_LABEL.get(block, block)).border = border
            ws.cell(row=r, column=2, value=disp).border = border
            ws.cell(row=r, column=3, value=d[0]).border = border
            ws.cell(row=r, column=4, value=d[1]).border = border
            ws.cell(row=r, column=2).font = Font(bold=True)
            for col in range(1, 5):
                ws.cell(row=r, column=col).alignment = Alignment(vertical='top', wrap_text=True)
            r += 1
        ws.column_dimensions['A'].width = 28; ws.column_dimensions['B'].width = 26
        ws.column_dimensions['C'].width = 42; ws.column_dimensions['D'].width = 80
        ws.freeze_panes = "A3"

    def _build_detail_sheet(self, ws, df, col_block, numeric_cols):
        thin = Side(style='thin', color='E0E0E0')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cols = list(df.columns)
        name2block = {n: b for (n, b) in col_block}

        col_idx = 1; start = 1; prev_block = None
        def write_banner(c0, c1, block):
            if block is None:
                return
            info = BLOC_COLORS.get(block, {'banner': '888888', 'name': block, 'font': 'FFFFFF'})
            ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
            cell = ws.cell(row=1, column=c0, value=info['name'])
            cell.font = Font(bold=True, size=10, color=info.get('font', 'FFFFFF'))
            cell.fill = PatternFill('solid', fgColor=info['banner'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for j, name in enumerate(cols, start=1):
            b = name2block.get(name)
            if prev_block is None:
                prev_block = b; start = j
            elif b != prev_block:
                write_banner(start, j - 1, prev_block); prev_block = b; start = j
        write_banner(start, len(cols), prev_block)

        for j, name in enumerate(cols, start=1):
            b = name2block.get(name, '')
            info = BLOC_COLORS.get(b, {'header': '444444', 'font': 'FFFFFF'})
            cell = ws.cell(row=2, column=j, value=name)
            cell.font = Font(bold=True, size=9, color=info.get('font', 'FFFFFF'))
            cell.fill = PatternFill('solid', fgColor=info['header'])
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        ws.row_dimensions[2].height = 40

        for i in range(len(df)):
            row_xl = i + 3
            for j, name in enumerate(cols, start=1):
                raw = df.iloc[i, j - 1]
                cell = ws.cell(row=row_xl, column=j)
                if name in numeric_cols:
                    v = self._parse_fr(raw)
                    if v is None:
                        cell.value = None
                    else:
                        if any(h in name for h in self.INT_HINTS):
                            cell.value = int(round(v)); cell.number_format = '#,##0'
                        else:
                            cell.value = round(v, 2); cell.number_format = '#,##0.00'
                else:
                    cell.value = ('' if raw is None else str(raw))
                b = name2block.get(name, '')
                info = BLOC_COLORS.get(b)
                if info and (i % 2 == 1):
                    cell.fill = PatternFill('solid', fgColor=info['cell'])
                cell.border = border
                cell.alignment = Alignment(vertical='center')

        for j, name in enumerate(cols, start=1):
            w = 14
            if name.startswith("DIAG"): w = 70
            elif name in ('IBANS_RATTACHES', 'EXPO_STATUT', 'NOM_LE', 'NOM_GA'): w = 30
            elif name in self.F_TEXT_COLS: w = 22
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{len(df) + 2}"

    # =========================================================================
    # SYNTHÈSE (VERBATIM)
    # =========================================================================
    def _build_synthese(self, ws, df_out, snap, stats, mx_global_by_mois, wl_global_by_mois):
        thin = Side(style='thin', color='D8D8D8')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def title(r, text, fill):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            c = ws.cell(row=r, column=1, value=text)
            c.font = Font(bold=True, size=12, color='FFFFFF'); c.fill = PatternFill('solid', fgColor=fill)
            c.alignment = Alignment(horizontal='left', vertical='center'); ws.row_dimensions[r].height = 22
            return r + 1

        def table(r, headers, rows, hdr_fill):
            for j, h in enumerate(headers, start=1):
                c = ws.cell(row=r, column=j, value=h)
                c.font = Font(bold=True, color='FFFFFF', size=10); c.fill = PatternFill('solid', fgColor=hdr_fill)
                c.alignment = Alignment(horizontal='center'); c.border = border
            r += 1
            for row in rows:
                for j, val in enumerate(row, start=1):
                    c = ws.cell(row=r, column=j, value=val)
                    c.border = border
                    if j == 1:
                        c.alignment = Alignment(horizontal='left')
                    else:
                        c.alignment = Alignment(horizontal='right')
                        if isinstance(val, (int, float)):
                            c.number_format = '#,##0' if isinstance(val, int) else '#,##0.00'
                r += 1
            return r + 1

        ws.merge_cells('A1:D1')
        c = ws['A1']; c.value = f"SYNTHÈSE — CIB REVENUS ANALYZER v32 [{VERSION_ID}]"
        c.font = Font(bold=True, size=14, color='FFFFFF'); c.fill = PatternFill('solid', fgColor=BNP_GREEN.lstrip('#'))
        c.alignment = Alignment(horizontal='center', vertical='center'); ws.row_dimensions[1].height = 28
        r = 3

        all_mois = stats['all_mois']
        for metric, idx, fmt_int in [("FLUX (EUR)", 'flux', False), ("PNB (EUR)", 'pnb', False), ("NB CARTES", 'nb', True)]:
            r = title(r, f"COMPARATIF MENSUEL — {metric}", '1F4E79')
            rows = []
            tot_g = tot_c = tot_p = 0.0
            for mo in all_mois:
                cco = float(mx_global_by_mois.get(mo, {}).get(idx, 0.0))
                cpc = float(wl_global_by_mois.get(mo, {}).get(idx, 0.0))
                g = cco + cpc
                tot_g += g; tot_c += cco; tot_p += cpc
                if fmt_int:
                    rows.append([self.mois_label(mo), int(round(g)), int(round(cco)), int(round(cpc))])
                else:
                    rows.append([self.mois_label(mo), round(g, 2), round(cco, 2), round(cpc, 2)])
            if fmt_int:
                rows.append(["TOTAL", int(round(tot_g)), int(round(tot_c)), int(round(tot_p))])
            else:
                rows.append(["TOTAL", round(tot_g, 2), round(tot_c, 2), round(tot_p, 2)])
            r = table(r, ["Mois", "GLOBAL", "CCO (MX)", "CPC (WL)"], rows, '4A78A8')

        quarters = snap['quarters']; rmpm = snap['rmpm']
        def q_sum(dmx, dwl, key):
            cco = float(np.sum(dmx[key])); cpc = float(np.sum(dwl[key])); return (cco + cpc, cco, cpc)
        def q_customers(key):
            fmx = snap['q_flux_mx'][key]; fwl = snap['q_flux_wl'][key]
            cco = {rmpm[i] for i in range(len(rmpm)) if rmpm[i] and fmx[i] > 0}
            cpc = {rmpm[i] for i in range(len(rmpm)) if rmpm[i] and fwl[i] > 0}
            tot = {rmpm[i] for i in range(len(rmpm)) if rmpm[i] and (fmx[i] > 0 or fwl[i] > 0)}
            return (len(tot), len(cco), len(cpc))

        for label, src, is_int, is_cust in [
            ("CARTES", (snap['q_cartes_mx'], snap['q_cartes_wl']), True, False),
            ("FLUX (EUR)", (snap['q_flux_mx'], snap['q_flux_wl']), False, False),
            ("REVENUS / PNB (EUR)", (snap['q_pnb_mx'], snap['q_pnb_wl']), False, False),
            ("CUSTOMERS (RMPM distincts)", None, True, True),
        ]:
            r = title(r, f"TRIMESTRIEL — {label}", '00665E')
            rows = []
            for (an, q) in quarters:
                key = (an, q)
                if is_cust:
                    tot, cco, cpc = q_customers(key)
                else:
                    tot, cco, cpc = q_sum(src[0], src[1], key)
                if is_int:
                    rows.append([self.quarter_label_long(an, q), int(round(tot)), int(round(cco)), int(round(cpc))])
                else:
                    rows.append([self.quarter_label_long(an, q), round(tot, 2), round(cco, 2), round(cpc, 2)])
            r = table(r, ["Trimestre", "TOTAL", "CCO (MX)", "CPC (WL)"], rows, '4E8E88')

        ref = snap['ref_yyyymm']; ref_lbl = self.mois_label(ref)
        found_mx = snap['found_mx']; found_wl = snap['found_wl']
        expo_mx = snap['expo_mx']; expo_wl = snap['expo_wl']; expo_reelle = snap['expo_reelle']
        expo_statut = snap['expo_statut']; expo_comp = snap['expo_comp']
        val_expo_th = snap['val_expo_th']; raw_expo_th = snap['raw_expo_th']
        statut = snap['statut']; differe = snap['differe']
        n = len(rmpm)

        def is_confirmed(i):
            return 'CONFIRM' in str(statut[i]).upper()
        def th_present(i):
            return str(raw_expo_th[i]).strip().lower() not in ('', 'nan', 'none', 'null', 'n/a', 'na')

        proj_total = sum(float(val_expo_th[i]) for i in range(n) if th_present(i))
        proj_cco = proj_cpc = 0.0
        for i in range(n):
            if not th_present(i):
                continue
            ag = float(val_expo_th[i])
            if found_mx[i] and not found_wl[i]:
                proj_cco += ag
            elif found_wl[i] and not found_mx[i]:
                proj_cpc += ag
            elif found_mx[i] and found_wl[i]:
                emx = float(expo_mx[i]) if expo_mx[i] is not None else 0.0
                ewl = float(expo_wl[i]) if expo_wl[i] is not None else 0.0
                base = emx + ewl
                if base > 0:
                    proj_cco += ag * emx / base; proj_cpc += ag * ewl / base
                else:
                    proj_cco += ag / 2.0; proj_cpc += ag / 2.0

        act_total = act_cco = act_cpc = 0.0
        for i in range(n):
            if not (is_confirmed(i) and differe[i] not in (None,) and (differe[i] or 0) != 0 and expo_statut[i] == 'CALCULÉE'):
                continue
            if expo_reelle[i] is not None:
                act_total += float(expo_reelle[i])
            if expo_mx[i] is not None:
                act_cco += float(expo_mx[i])
            if expo_wl[i] is not None:
                act_cpc += float(expo_wl[i])

        r = title(r, f"EXPOSURE — SNAPSHOT {ref_lbl}", '0D3B66')
        r = table(r, ["", "TOTAL", "CCO (MX)", "CPC (WL)"], [
            ["Projected Exposure (théorique, AG)", round(proj_total, 2), round(proj_cco, 2), round(proj_cpc, 2)],
            ["Actual Exposure (réelle, Confirmé+différé≠0)", round(act_total, 2), round(act_cco, 2), round(act_cpc, 2)],
        ], '4A6A8E')
        r += 1

        pop = [i for i in range(n) if is_confirmed(i)]
        def rmset(cond):
            return len({rmpm[i] for i in pop if rmpm[i] and cond(i)})
        exceeding = rmset(lambda i: str(expo_comp[i]).startswith('SUP'))
        equal     = rmset(lambda i: expo_comp[i] == 'ÉGALE')
        lower     = rmset(lambda i: str(expo_comp[i]).startswith('INF'))
        no_expo   = rmset(lambda i: expo_statut[i] != 'CALCULÉE' or (expo_reelle[i] in (None, 0) or float(expo_reelle[i] or 0) == 0))
        immediat  = rmset(lambda i: (differe[i] is not None) and (differe[i] < 1))
        total_pop = len({rmpm[i] for i in pop if rmpm[i]})
        r = title(r, f"CUSTOMERS — SNAPSHOT {ref_lbl}", '0D3B66')
        r = table(r, ["Catégorie", "Nb sociétés (RMPM)", "", ""], [
            ["Exposition réelle SUPÉRIEURE à la théorique", exceeding, "", ""],
            ["Exposition réelle ÉGALE à la théorique", equal, "", ""],
            ["Exposition réelle INFÉRIEURE à la théorique", lower, "", ""],
            ["Sans exposition exploitable (non calculable / nulle)", no_expo, "", ""],
            ["Paiement immédiat (différé < 1 j)", immediat, "", ""],
            ["TOTAL sociétés (statut Confirmé)", total_pop, "", ""],
        ], '4A6A8E')
        r += 1

        # ----- AJOUT V2 : récupération étendue (PARC / NOM→IBAN) -----
        r = title(r, "RÉCUPÉRATION ÉTENDUE (V2) — clients repêchés", '583618')
        r = table(r, ["Voie de récupération", "Nb lignes", "", ""], [
            ["Via PARC_CLIENT (RMPM → RC/RP → flux)", int(stats.get('n_parc', 0)), "", ""],
            ["Via NOM exact (SOURCE → ACCOUNT/SINGLETON → IBAN → flux)", int(stats.get('n_nom_iban', 0)), "", ""],
            ["Override BEJO (MX)", int(stats.get('n_bejo', 0)), "", ""],
        ], '85603D')

        ws.column_dimensions['A'].width = 52
        for col in ('B', 'C', 'D'):
            ws.column_dimensions[col].width = 20
        ws.freeze_panes = "A2"


# =============================================================================
# POINT D'ENTRÉE — CLI argparse
# =============================================================================
def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="01.W3RKN.py",
        description=f"CIB REVENUS ANALYZER v34 [{VERSION_ID}] — version CLI (sans GUI).",
        formatter_class=argparse.RawTextHelpFormatter,
    )

    # --- 8 entrées obligatoires ---
    req = p.add_argument_group("Sources requises (obligatoires)")
    req.add_argument("--source",       required=False, default=None, help="00. MONITORING_CIB / ONBORDING (base a enrichir).")
    req.add_argument("--worldline",    required=False, default=None, help="05. PRGM Worldline (flux CPC).")
    req.add_argument("--monext",       required=False, default=None, help="04. MONEXT (flux CCO).")
    req.add_argument("--account",      required=False, default=None, help="01. IBAN_ACCOUNT (pivot priorite 1).")
    req.add_argument("--ref-client",   required=False, default=None, help="02. REFERENTIEL_CLIENT (pivot priorite 2).")
    req.add_argument("--idseg",        required=False, default=None, help="03. IDENTIFIANT_SEGMENT (pivot priorite 3).")
    req.add_argument("--devises",      required=False, default=None, help="09. DEVISES (Date YYYYMM | Code | Taux vers EUR).")
    req.add_argument("--type-of-card", required=False, default=None, help="08. CLEAN_TYPE_OF_CARD (Type Original -> New).")

    # --- entrées optionnelles ---
    opt = p.add_argument_group("Sources optionnelles")
    opt.add_argument("--bejo-cartes", default="", help="12. BEJO_CARTES (entite | nb cartes). Requiert --use-bejo.")
    opt.add_argument("--bejo-flux",   default="", help="13. BEJO_FLUX (entite | IBAN). Requiert --use-bejo.")
    opt.add_argument("--country",     default="", help="06. CLEAN_COUNTRY (Pays Original -> New).")
    opt.add_argument("--sales",       default="", help="07. CLEAN_SALES (Sales Original -> New).")
    opt.add_argument("--parc-client", default="", help="14. PARC_CLIENT (RMPM -> ID_RC + ID_RP).")
    opt.add_argument("--singleton",   default="", help="15. SINGLETON (Nom entite -> IBAN).")

    # --- sortie ---
    outg = p.add_argument_group("Sortie")
    outg.add_argument("--output-dir", required=True, help="Dossier de destination du CSV et du XLSX.")
    outg.add_argument("--output-filename", default=f"REVENUS_CIB_{VERSION_ID}",
                      help="Nom de base (sans extension) des fichiers produits.")

    # --- options metier ---
    optm = p.add_argument_group("Options metier")
    optm.add_argument("--use-bejo", action="store_true",
                      help="Active l'integration BEJO (requiert --bejo-cartes et --bejo-flux).")
    optm.add_argument("--ref-mois", default="Décembre",
                      help="Mois de reference (nom FR, ex. \"Décembre\"). Defaut : Décembre.")
    optm.add_argument("--ref-annee", default="2025",
                      help="Annee de reference (ex. 2025). Defaut : 2025.")

    return p


def main(argv: Optional[List[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if not OPENPYXL_OK:
        print("ATTENTION : openpyxl absent — installez-le pour la sortie XLSX (pip install openpyxl).")

    # Auto-résolution des sources OBLIGATOIRES depuis 03.sources/ si non fournies.
    try:
        if not args.source:
            args.source = resolve_source("MONITORING_ONBOARDING_CIB", required=True)
        if not args.worldline:
            args.worldline = resolve_source("PRGM_AGREGE", required=True)
        if not args.monext:
            args.monext = resolve_source("MONEXT_AGREGE", required=True)
        if not args.account:
            args.account = resolve_source("IBAN_ACCOUNT", required=True)
        if not args.ref_client:
            args.ref_client = resolve_source("REFERENTIEL_CLIENT", required=True)
        if not args.idseg:
            args.idseg = resolve_source("IDENTIFIANT_SEGMENT", required=True)
        if not args.devises:
            args.devises = resolve_source("DEVISES", required=True)
        if not args.type_of_card:
            args.type_of_card = resolve_source("CLEAN_TYPE_OF_CARD", required=True)
    except FileNotFoundError as e:
        print(f"ERREUR : {e}", file=sys.stderr)
        return 1

    # Validation des chemins d'entrée (obligatoires + optionnels fournis).
    to_check = {
        "--source": args.source, "--worldline": args.worldline, "--monext": args.monext,
        "--account": args.account, "--ref-client": args.ref_client, "--idseg": args.idseg,
        "--devises": args.devises, "--type-of-card": args.type_of_card,
    }
    for flag, opt_val in {
        "--bejo-cartes": args.bejo_cartes, "--bejo-flux": args.bejo_flux,
        "--country": args.country, "--sales": args.sales,
        "--parc-client": args.parc_client, "--singleton": args.singleton,
    }.items():
        if opt_val:
            to_check[flag] = opt_val

    missing = [f"{flag} ({path})" for flag, path in to_check.items() if not Path(path).is_file()]
    if missing:
        print("ERREUR : fichier(s) introuvable(s) :", file=sys.stderr)
        for mm in missing:
            print(f"  {mm}", file=sys.stderr)
        return 2

    # Cohérence BEJO (reprise de load_previews/start_thread).
    if args.use_bejo and (not args.bejo_cartes or not args.bejo_flux):
        print("ERREUR : --use-bejo requiert --bejo-cartes ET --bejo-flux.", file=sys.stderr)
        return 2
    if args.ref_mois not in MOIS_NOMS:
        print(f"ERREUR : --ref-mois doit etre l'un de : {', '.join(MOIS_NOMS)}.", file=sys.stderr)
        return 2

    try:
        analyzer = CIBRevenusAnalyzerCLI(args)
        result = analyzer.run()
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"ERREUR de traitement : {e}", file=sys.stderr)
        return 1

    print(f"[OK] Analyse terminee. CSV : {result['csv_path']}"
          + (f" | XLSX : {result['xlsx_path']}" if result.get('xlsx_path') else " | XLSX : (openpyxl indisponible)"))
    return 0


if __name__ == "__main__":
    sys.exit(main())
